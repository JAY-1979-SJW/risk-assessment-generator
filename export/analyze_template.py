# -*- coding: utf-8 -*-
"""
수정된 엑셀 템플릿 분석 스크립트
셀 병합, 열 너비, 행 높이, 고정 틀 등 추출
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def analyze_workbook(file_path):
    """워크북 분석"""
    wb = load_workbook(file_path)

    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        sheet_info = {
            "merged_cells": [],
            "column_widths": {},
            "row_heights": {},
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        }

        # 병합 셀 정보
        for merged_range in ws.merged_cells.ranges:
            sheet_info["merged_cells"].append(str(merged_range))

        # 열 너비
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            width = ws.column_dimensions[col_letter].width
            if width:
                sheet_info["column_widths"][col_letter] = round(width, 2)

        # 행 높이
        for row in range(1, min(ws.max_row + 1, 30)):  # 처음 30행만
            height = ws.row_dimensions[row].height
            if height:
                sheet_info["row_heights"][row] = round(height, 2)

        result[sheet_name] = sheet_info

    return result

def main():
    file_path = r"C:\Users\skyjw\OneDrive\03. PYTHON\15. 위험성평가표 자동생성기\export\위험성평가표_표준양식_테스트.xlsx"

    print("=" * 70)
    print("엑셀 템플릿 분석 결과")
    print("=" * 70)

    result = analyze_workbook(file_path)

    for sheet_name, info in result.items():
        print(f"\n### {sheet_name} ###")
        print(f"\n[고정 틀] {info['freeze_panes']}")

        print(f"\n[병합 셀] ({len(info['merged_cells'])}개)")
        for mc in info['merged_cells']:
            print(f"  - {mc}")

        print(f"\n[열 너비]")
        for col, width in info['column_widths'].items():
            print(f"  {col}: {width}")

        print(f"\n[행 높이]")
        for row, height in info['row_heights'].items():
            print(f"  {row}: {height}")

        print("-" * 50)

    # JSON으로도 저장
    output_json = file_path.replace('.xlsx', '_structure.json')
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\n구조 정보 저장: {output_json}")

if __name__ == "__main__":
    main()
