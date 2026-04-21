# -*- coding: utf-8 -*-
"""
엑셀 내보내기 모듈
KRAS 표준 양식에 맞게 엑셀 파일 생성

KRAS 표준 위험성평가표 구성:
1. 안전보건방침 및 추진목표
2. 위험성평가 조직구성
3. 위험성 추정 및 결정
4. 위험성평가 회의 결과
5. 위험성평가 실시표 (핵심)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.properties import PageSetupProperties
from datetime import datetime


class ExcelExporter:
    """KRAS 표준 양식 엑셀 내보내기 클래스"""

    def __init__(self, data_manager):
        self.data_manager = data_manager
        self._init_styles()

    def _init_styles(self):
        # 색상 정의
        self.HEADER_BG = "4472C4"
        self.HEADER_FG = "FFFFFF"
        self.SUBHEADER_BG = "D6DCE5"

        # 위험성 등급별 색상
        self.RISK_COLORS = {
            "높음": "FF6B6B",
            "보통": "FFE066",
            "낮음": "69DB7C",
        }

        # 폰트 정의
        self.title_font = Font(name="맑은 고딕", size=16, bold=True)
        self.subtitle_font = Font(name="맑은 고딕", size=12, bold=True)
        self.header_font = Font(name="맑은 고딕", size=10, bold=True, color=self.HEADER_FG)
        self.normal_font = Font(name="맑은 고딕", size=10)
        self.small_font = Font(name="맑은 고딕", size=9)

        # 채우기 정의
        self.header_fill = PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type="solid")
        self.subheader_fill = PatternFill(start_color=self.SUBHEADER_BG, end_color=self.SUBHEADER_BG, fill_type="solid")

        # 테두리 정의
        thin = Side(style='thin', color='000000')
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 정렬 정의
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def _get_risk_fill(self, level: str) -> PatternFill:
        """위험성 등급에 따른 배경색 반환"""
        color = self.RISK_COLORS.get(level, "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _apply_cell_style(self, cell, font=None, fill=None, border=None, alignment=None):
        """셀 스타일 일괄 적용"""
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment

    def _apply_border_to_range(self, ws, range_str):
        """범위에 테두리 적용"""
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = self.thin_border

    def _apply_print_settings(self, ws, print_area=None, fit_to_height=1):
        """가로 출력 설정 적용

        Args:
            ws: 워크시트
            print_area: 인쇄 영역 (예: 'A1:H22')
            fit_to_height: 높이 맞춤 페이지 수 (0=자동, 1=1페이지)
        """
        # 용지 방향: 가로
        ws.page_setup.orientation = 'landscape'

        # 용지 크기: A4
        ws.page_setup.paperSize = 9

        # 1페이지 맞춤
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = fit_to_height

        # fitToPage 활성화
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 여백 설정 (인치) - 상단 편철 여백 확보
        ws.page_margins = PageMargins(
            left=0.15,     # 약 0.4cm
            right=0.15,
            top=0.6,       # 약 1.5cm (편철용)
            bottom=0.2,
            header=0.3,
            footer=0.1
        )

        # 용지 정렬: 수평 가운데, 수직 상단
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False  # 상단 배치

        # 인쇄 영역 설정
        if print_area:
            ws.print_area = print_area

        # 눈금선 인쇄 안함
        ws.print_options.gridLines = False

    def export(self, file_path: str):
        """엑셀 파일로 내보내기"""
        wb = Workbook()
        wb.remove(wb.active)

        # 시트 생성 (KRAS 표준 순서)
        self._create_safety_policy_sheet(wb)
        self._create_organization_sheet(wb)
        self._create_risk_criteria_sheet(wb)
        self._create_meeting_sheet(wb)
        self._create_assessment_sheet(wb)

        wb.save(file_path)
        print(f"엑셀 파일 생성 완료: {file_path}")

    def _create_safety_policy_sheet(self, wb):
        """1. 안전보건방침 및 추진목표 시트"""
        ws = wb.create_sheet("1.안전보건방침")
        info = self.data_manager.company_info

        # 열 너비 설정 (A4 가로 인쇄 영역에 맞춤, 총 약 145자)
        col_widths = {'A': 15, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 22}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 제목 (A1:H1)
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = "안전보건방침 및 추진목표"
        self._apply_cell_style(cell, self.title_font, alignment=self.center_align)
        ws.row_dimensions[1].height = 35

        # 회사정보 테이블 (3~6행)
        info_data = [
            # row 3
            [("A3:B3", "회사명", True), ("C3:D3", info.get("company_name", ""), False),
             ("E3:F3", "대표자", True), ("G3:H3", info.get("ceo_name", ""), False)],
            # row 4
            [("A4:B4", "업종", True), ("C4:D4", info.get("business_type", ""), False),
             ("E4:F4", "평가유형", True), ("G4:H4", info.get("eval_type", ""), False)],
            # row 5
            [("A5:B5", "현장명", True), ("C5:D5", info.get("site_name", ""), False),
             ("E5:F5", "평가일자", True), ("G5:H5", info.get("eval_date", ""), False)],
            # row 6
            [("A6:B6", "주소", True), ("C6:H6", info.get("address", ""), False)],
        ]

        for row_data in info_data:
            for merge_range, value, is_header in row_data:
                ws.merge_cells(merge_range)
                cell = ws[merge_range.split(':')[0]]
                cell.value = value
                if is_header:
                    self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
                else:
                    self._apply_cell_style(cell, self.normal_font, border=self.thin_border, alignment=self.left_align)
                self._apply_border_to_range(ws, merge_range)

        # 안전보건방침 헤더 (A8:H8)
        ws.merge_cells('A8:H8')
        cell = ws['A8']
        cell.value = "안전보건방침"
        self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'A8:H8')

        # 안전보건방침 내용 (A9:H15)
        ws.merge_cells('A9:H15')
        cell = ws['A9']
        cell.value = info.get("safety_policy", "")
        self._apply_cell_style(cell, self.normal_font, border=self.thin_border, alignment=self.top_left_align)
        self._apply_border_to_range(ws, 'A9:H15')

        # 추진목표 헤더 (A17:H17)
        ws.merge_cells('A17:H17')
        cell = ws['A17']
        cell.value = "추진목표"
        self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'A17:H17')

        # 추진목표 내용 (A18:H22)
        ws.merge_cells('A18:H22')
        cell = ws['A18']
        cell.value = info.get("safety_goal", "")
        self._apply_cell_style(cell, self.normal_font, border=self.thin_border, alignment=self.top_left_align)
        self._apply_border_to_range(ws, 'A18:H22')

        # 틀 고정 (C7)
        ws.freeze_panes = 'C7'

        # 페이지 설정: 가로 방향, A4
        self._apply_print_settings(ws, 'A1:H22')

    def _create_organization_sheet(self, wb):
        """2. 위험성평가 조직구성 시트"""
        ws = wb.create_sheet("2.조직구성")
        info = self.data_manager.company_info
        org = self.data_manager.organization

        # 열 너비 설정 (A4 가로 인쇄 영역에 맞춤, 총 약 145자)
        col_widths = {'A': 18, 'B': 15, 'C': 15, 'D': 55, 'E': 20, 'F': 22}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 제목 (A1:F1)
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "위험성평가 실시 담당 조직 구성"
        self._apply_cell_style(cell, self.title_font, alignment=self.center_align)
        ws.row_dimensions[1].height = 35

        # 부제목 (A2:F2)
        ws.merge_cells('A2:F2')
        cell = ws['A2']
        cell.value = "(표준실시규정 제3조, 제4조 참조)"
        self._apply_cell_style(cell, Font(name="맑은 고딕", size=10, color="666666"), alignment=self.center_align)

        # 빈 행 (A3:F3)
        ws.merge_cells('A3:F3')

        # 회사명 (row 4)
        ws.cell(row=4, column=1, value="회사명")
        self._apply_cell_style(ws.cell(row=4, column=1), self.header_font, self.header_fill, self.thin_border, self.center_align)
        ws.merge_cells('B4:F4')
        ws['B4'] = info.get("company_name", "")
        self._apply_cell_style(ws['B4'], self.normal_font, border=self.thin_border, alignment=self.left_align)
        self._apply_border_to_range(ws, 'B4:F4')

        # 현장명 (row 5)
        ws.cell(row=5, column=1, value="현장명")
        self._apply_cell_style(ws.cell(row=5, column=1), self.header_font, self.header_fill, self.thin_border, self.center_align)
        ws.merge_cells('B5:F5')
        ws['B5'] = info.get("site_name", "")
        self._apply_cell_style(ws['B5'], self.normal_font, border=self.thin_border, alignment=self.left_align)
        self._apply_border_to_range(ws, 'B5:F5')

        # 빈 행 (A6:F6)
        ws.merge_cells('A6:F6')

        # 조직 테이블 헤더 (row 7)
        headers = ["직위/직책", "성명", "역할", "책임 및 권한"]
        header_cols = [1, 2, 3, 4]  # A, B, C, D(D:F 병합)

        ws.cell(row=7, column=1, value=headers[0])
        ws.cell(row=7, column=2, value=headers[1])
        ws.cell(row=7, column=3, value=headers[2])
        ws.merge_cells('D7:F7')
        ws.cell(row=7, column=4, value=headers[3])

        for col in [1, 2, 3, 4]:
            self._apply_cell_style(ws.cell(row=7, column=col), self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'D7:F7')
        ws.row_dimensions[7].height = 28

        # 조직 데이터
        members = org.get("members", [])
        if not members:
            members = [
                {"position": "대표이사", "name": "", "role": "총괄관리", "responsibility": "위험성평가 총괄책임, 안전보건방침 결정"},
                {"position": "안전관리자", "name": "", "role": "실무책임", "responsibility": "위험성평가 실시 총괄, 개선대책 수립 및 이행"},
                {"position": "관리감독자", "name": "", "role": "현장감독", "responsibility": "작업자 지휘·감독, 위험요인 발굴 및 보고"},
                {"position": "근로자대표", "name": "", "role": "근로자대표", "responsibility": "위험성평가 참여, 근로자 의견 수렴"},
            ]

        for i, member in enumerate(members):
            row = 8 + i
            ws.cell(row=row, column=1, value=member.get("position", ""))
            ws.cell(row=row, column=2, value=member.get("name", ""))
            ws.cell(row=row, column=3, value=member.get("role", ""))
            ws.merge_cells(f'D{row}:F{row}')
            ws.cell(row=row, column=4, value=member.get("responsibility", ""))

            for col in [1, 2, 3]:
                self._apply_cell_style(ws.cell(row=row, column=col), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_cell_style(ws.cell(row=row, column=4), self.normal_font, border=self.thin_border, alignment=self.left_align)
            self._apply_border_to_range(ws, f'D{row}:F{row}')
            ws.row_dimensions[row].height = 28

        # 틀 고정 (B8)
        ws.freeze_panes = 'B8'

        # 마지막 행 계산 (조직 멤버 수에 따라)
        members = org.get("members", [])
        last_row = 7 + max(len(members), 4)

        # 페이지 설정: 가로 방향, A4
        self._apply_print_settings(ws, f'A1:F{last_row}')

    def _create_risk_criteria_sheet(self, wb):
        """3. 위험성 추정 및 결정 시트"""
        ws = wb.create_sheet("3.위험성기준")

        # 열 너비 설정 (A4 가로 인쇄 영역에 맞춤, 총 약 145자)
        col_widths = {'A': 15, 'B': 12, 'C': 18, 'D': 25, 'E': 25, 'F': 25, 'G': 25}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 제목 (A1:G1)
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = "위험성 추정 및 결정 방법"
        self._apply_cell_style(cell, self.title_font, alignment=self.center_align)
        ws.row_dimensions[1].height = 35

        # 부제목 (A3:G3)
        ws.merge_cells('A3:G3')
        cell = ws['A3']
        cell.value = "빈도·강도법(곱셈식)에 의한 위험성 추정 및 결정표"
        self._apply_cell_style(cell, self.subtitle_font, alignment=self.center_align)

        # 설명 (A4:G4)
        ws.merge_cells('A4:G4')
        cell = ws['A4']
        cell.value = "※ 실시방법: 가능성(빈도)과 중대성(강도)을 추정한 수치를 곱셈에 의해 위험성을 구하고 위험성 수준을 결정함"
        self._apply_cell_style(cell, self.normal_font, alignment=self.left_align)

        # 가능성(빈도) 섹션
        ws.merge_cells('A6:G6')
        ws['A6'] = "▶ 가능성(빈도): 사고나 질병으로 이어질 가능성(확률)을 파악"
        self._apply_cell_style(ws['A6'], Font(name="맑은 고딕", size=10, bold=True), alignment=self.left_align)

        # 가능성 테이블 헤더 (row 7)
        ws.merge_cells('C7:G7')
        for col, header in [(1, "구분"), (2, "점수"), (3, "기준")]:
            cell = ws.cell(row=7, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'C7:G7')

        possibility_data = [
            ("상", 3, "발생가능성이 높음. 일상적으로 장시간 이루어지는 작업에 수반하는 것으로 피하기 어려운 것"),
            ("중", 2, "발생가능성이 있음. 일상적인 작업에 수반하는 것으로 피할 수 있는 것"),
            ("하", 1, "발생가능성이 낮음. 비정상적인 작업에 수반하는 것으로 피할 수 있는 것"),
        ]

        for i, (grade, score, desc) in enumerate(possibility_data):
            row = 8 + i
            ws.cell(row=row, column=1, value=grade)
            ws.cell(row=row, column=2, value=score)
            ws.merge_cells(f'C{row}:G{row}')
            ws.cell(row=row, column=3, value=desc)

            self._apply_cell_style(ws.cell(row=row, column=1), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_cell_style(ws.cell(row=row, column=2), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_cell_style(ws.cell(row=row, column=3), self.normal_font, border=self.thin_border, alignment=self.left_align)
            self._apply_border_to_range(ws, f'C{row}:G{row}')

        # 중대성(강도) 섹션
        ws.merge_cells('A12:G12')
        ws['A12'] = "▶ 중대성(강도): 사고나 질병으로 이어졌을 때 그 중대성(강도)을 파악"
        self._apply_cell_style(ws['A12'], Font(name="맑은 고딕", size=10, bold=True), alignment=self.left_align)

        # 중대성 테이블 헤더 (row 13)
        ws.merge_cells('C13:G13')
        for col, header in [(1, "구분"), (2, "점수"), (3, "기준")]:
            cell = ws.cell(row=13, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'C13:G13')

        severity_data = [
            ("대", 3, "사망을 초래할 수 있는 사고. 신체 일부에 영구손상을 수반하는 것"),
            ("중", 2, "휴업재해, 한번에 다수의 피해자가 수반하는 것. 실명, 절단 등 상해를 초래할 수 있는 사고"),
            ("소", 1, "아차 사고. 처치 후 바로 원래의 작업을 수행할 수 있는 경미한 부상 또는 질병"),
        ]

        for i, (grade, score, desc) in enumerate(severity_data):
            row = 14 + i
            ws.cell(row=row, column=1, value=grade)
            ws.cell(row=row, column=2, value=score)
            ws.merge_cells(f'C{row}:G{row}')
            ws.cell(row=row, column=3, value=desc)

            self._apply_cell_style(ws.cell(row=row, column=1), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_cell_style(ws.cell(row=row, column=2), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_cell_style(ws.cell(row=row, column=3), self.normal_font, border=self.thin_border, alignment=self.left_align)
            self._apply_border_to_range(ws, f'C{row}:G{row}')

        # 공식 (A18:G18)
        ws.merge_cells('A18:G18')
        ws['A18'] = "★ 위험성 = 가능성(빈도) × 중대성(강도)"
        self._apply_cell_style(ws['A18'], Font(name="맑은 고딕", size=12, bold=True, color="FF0000"), alignment=self.center_align)

        # 위험성 결정 섹션 (A20:G20)
        ws.merge_cells('A20:G20')
        ws['A20'] = "▶ 위험성 결정: 평가 3단계로 구분하고 평가 점수가 높은 순서대로 우선순위 결정"
        self._apply_cell_style(ws['A20'], Font(name="맑은 고딕", size=10, bold=True), alignment=self.left_align)

        # 위험성 결정 테이블 헤더 (row 21)
        ws.merge_cells('D21:G21')
        for col, header in [(1, "위험성 점수"), (2, "등급"), (3, "허용 여부"), (4, "조치 사항")]:
            cell = ws.cell(row=21, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'D21:G21')

        decision_data = [
            ("6 ~ 9", "높음", "허용 불가", "작업을 지속하려면 즉시 개선을 실행"),
            ("3 ~ 4", "보통", "허용 불가", "안전보건대책을 수립하고 개선"),
            ("1 ~ 2", "낮음", "허용 가능", "근로자에게 유해위험성 정보를 제공 및 교육"),
        ]

        for i, (score_range, level, acceptable, action) in enumerate(decision_data):
            row = 22 + i
            ws.cell(row=row, column=1, value=score_range)
            ws.cell(row=row, column=2, value=level)
            ws.cell(row=row, column=2).fill = self._get_risk_fill(level)
            ws.cell(row=row, column=3, value=acceptable)
            ws.merge_cells(f'D{row}:G{row}')
            ws.cell(row=row, column=4, value=action)

            for col in [1, 2, 3, 4]:
                self._apply_cell_style(ws.cell(row=row, column=col), self.normal_font, border=self.thin_border, alignment=self.center_align)
            self._apply_border_to_range(ws, f'D{row}:G{row}')

        # 틀 고정 (B8)
        ws.freeze_panes = 'B8'

        # 페이지 설정: 가로 방향, A4
        self._apply_print_settings(ws, 'A1:G24')

    def _create_meeting_sheet(self, wb):
        """4. 위험성평가 회의 결과 시트"""
        ws = wb.create_sheet("4.회의결과")
        meeting = self.data_manager.meeting

        # 열 너비 설정 (A4 가로 인쇄 영역에 맞춤, 총 약 145자)
        col_widths = {'A': 15, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 22}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 제목 (A1:H1)
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = "위험성평가 회의 결과"
        self._apply_cell_style(cell, self.title_font, alignment=self.center_align)
        ws.row_dimensions[1].height = 35

        # 회의정보 (row 3)
        ws.cell(row=3, column=1, value="회의일시")
        self._apply_cell_style(ws.cell(row=3, column=1), self.header_font, self.header_fill, self.thin_border, self.center_align)

        ws.merge_cells('B3:D3')
        date_str = meeting.get("date", "")
        time_start = meeting.get("time_start", "")
        time_end = meeting.get("time_end", "")
        ws['B3'] = f"{date_str}  {time_start} ~ {time_end}"
        self._apply_cell_style(ws['B3'], self.normal_font, border=self.thin_border, alignment=self.left_align)
        self._apply_border_to_range(ws, 'B3:D3')

        ws.cell(row=3, column=5, value="장소")
        self._apply_cell_style(ws.cell(row=3, column=5), self.header_font, self.header_fill, self.thin_border, self.center_align)

        ws.merge_cells('F3:H3')
        ws['F3'] = meeting.get("location", "")
        self._apply_cell_style(ws['F3'], self.normal_font, border=self.thin_border, alignment=self.left_align)
        self._apply_border_to_range(ws, 'F3:H3')

        # 회의내용 헤더 (A5:H5)
        ws.merge_cells('A5:H5')
        ws['A5'] = "회의내용"
        self._apply_cell_style(ws['A5'], self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'A5:H5')

        # 회의내용 본문 (A6:H13)
        ws.merge_cells('A6:H13')
        ws['A6'] = meeting.get("content", "")
        self._apply_cell_style(ws['A6'], self.normal_font, border=self.thin_border, alignment=self.top_left_align)
        self._apply_border_to_range(ws, 'A6:H13')

        # 참석자 명단 헤더 (A15:H15)
        ws.merge_cells('A15:H15')
        ws['A15'] = "참석자 명단"
        self._apply_cell_style(ws['A15'], self.header_font, self.header_fill, self.thin_border, self.center_align)
        self._apply_border_to_range(ws, 'A15:H15')

        # 참석자 테이블 헤더 (row 16)
        headers = ["소속", "직책", "성명", "서명", "소속", "직책", "성명", "서명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.subheader_fill, self.thin_border, self.center_align)

        # 참석자 데이터 (2명씩 한 행)
        attendees = meeting.get("attendees", [])
        row = 17
        for i in range(0, max(len(attendees), 6), 2):
            for j in range(2):
                col_offset = j * 4
                if i + j < len(attendees):
                    att = attendees[i + j]
                    ws.cell(row=row, column=1 + col_offset, value=att.get("department", ""))
                    ws.cell(row=row, column=2 + col_offset, value=att.get("position", ""))
                    ws.cell(row=row, column=3 + col_offset, value=att.get("name", ""))
                    ws.cell(row=row, column=4 + col_offset, value="")

                for col in range(1 + col_offset, 5 + col_offset):
                    self._apply_cell_style(ws.cell(row=row, column=col), self.normal_font, border=self.thin_border, alignment=self.center_align)
            row += 1

        # 틀 고정 (B4)
        ws.freeze_panes = 'B4'

        # 마지막 행 계산 (참석자 수에 따라)
        attendees = meeting.get("attendees", [])
        attendee_rows = max((len(attendees) + 1) // 2, 3)  # 2명씩 한 행, 최소 3행
        last_row = 16 + attendee_rows

        # 페이지 설정: 가로 방향, A4
        self._apply_print_settings(ws, f'A1:H{last_row}')

    def _create_assessment_sheet(self, wb):
        """5. 위험성평가 실시표 시트 (핵심)"""
        ws = wb.create_sheet("5.위험성평가실시")
        info = self.data_manager.company_info

        # 열 너비 설정 (A4 가로 인쇄 영역에 맞춤, 총 약 145자)
        # 번호, 공정/작업명, 유해위험요인, 관련근거, 현재조치, 가능성, 중대성, 위험성, 감소대책, 개선후, 담당자, 완료일
        col_widths = {'A': 5, 'B': 14, 'C': 28, 'D': 12, 'E': 18, 'F': 8, 'G': 8, 'H': 8, 'I': 22, 'J': 8, 'K': 8, 'L': 10}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 제목 (A1:L1)
        ws.merge_cells('A1:L1')
        ws['A1'] = "위험성평가 실시표"
        self._apply_cell_style(ws['A1'], self.title_font, alignment=self.center_align)
        ws.row_dimensions[1].height = 30

        # KRAS 표시 (A2:L2)
        ws.merge_cells('A2:L2')
        ws['A2'] = "KRAS 표준 위험성평가 (https://kras.kosha.or.kr)"
        self._apply_cell_style(ws['A2'], self.small_font, alignment=self.center_align)

        # 기본정보 (row 4)
        ws.merge_cells('A4:C4')
        ws['A4'] = f"현장명: {info.get('site_name', '')}"
        self._apply_cell_style(ws['A4'], Font(name="맑은 고딕", size=10, bold=True), alignment=self.left_align)

        ws.merge_cells('J4:L4')
        ws['J4'] = f"평가일자: {info.get('eval_date', '')}"
        self._apply_cell_style(ws['J4'], self.normal_font, alignment=Alignment(horizontal='right', vertical='center'))

        # ===== 2행 헤더 구조 (row 6-7) =====
        # 병합 셀들
        header_merges = [
            'A6:A7', 'B6:B7', 'C6:C7', 'D6:D7', 'E6:E7',  # 개별 헤더 (세로 병합)
            'F6:H6',  # 위험성 평가 (가로 병합)
            'I6:I7', 'J6:J7', 'K6:K7', 'L6:L7'  # 개별 헤더 (세로 병합)
        ]
        for merge_range in header_merges:
            ws.merge_cells(merge_range)

        # 1행 헤더 내용 (row 6)
        headers_row6 = [
            (1, "번호"), (2, "공정/작업명"), (3, "유해위험요인"), (4, "관련근거"),
            (5, "현재\n안전보건조치"), (6, "위험성 평가"),
            (9, "위험성\n감소대책"), (10, "개선후\n위험성"), (11, "담당자"), (12, "완료일")
        ]

        for col, text in headers_row6:
            cell = ws.cell(row=6, column=col, value=text)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)

        # 2행 헤더 내용 (row 7) - 위험성 평가 하위
        sub_headers = [(6, "가능성\n(빈도)"), (7, "중대성\n(강도)"), (8, "위험성")]
        for col, text in sub_headers:
            cell = ws.cell(row=7, column=col, value=text)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.thin_border, self.center_align)

        # 모든 헤더 셀에 테두리 적용
        for row in [6, 7]:
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = self.thin_border

        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 30

        # ===== 데이터 행 =====
        data_row = 8
        assessments = self.data_manager.assessments

        for idx, assessment in enumerate(assessments, 1):
            current_level = assessment.get("current_risk_level", "낮음")
            after_level = assessment.get("after_risk_level", "낮음")

            # 유해위험요인 조합
            risk_factor_parts = []
            if assessment.get("risk_category"):
                risk_factor_parts.append(f"[{assessment.get('risk_category')}]")
            if assessment.get("risk_detail"):
                risk_factor_parts.append(assessment.get("risk_detail"))
            if assessment.get("risk_situation"):
                risk_factor_parts.append(assessment.get("risk_situation"))
            risk_factor = " ".join(risk_factor_parts)

            # 공정/작업명 조합
            process_work = assessment.get("process", "")
            if assessment.get("sub_work"):
                process_work += f"\n({assessment.get('sub_work')})"

            # 위험성 점수
            possibility = assessment.get("possibility", 1)
            severity = assessment.get("severity", 1)
            current_risk = assessment.get("current_risk", possibility * severity)
            after_risk = assessment.get("after_risk", 1)

            data = [
                idx,
                process_work,
                risk_factor,
                assessment.get("legal_basis", ""),
                assessment.get("current_measures", ""),
                self.data_manager.format_possibility(possibility),
                self.data_manager.format_severity(severity),
                f"{current_risk}\n({current_level})",
                assessment.get("reduction_measures", ""),
                f"{after_risk}\n({after_level})",
                assessment.get("manager", ""),
                assessment.get("complete_date", assessment.get("due_date", "")),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                self._apply_cell_style(cell, self.normal_font, border=self.thin_border)

                if col in [1, 6, 7, 8, 10, 11, 12]:
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.left_align

                if col == 8:
                    cell.fill = self._get_risk_fill(current_level)
                elif col == 10:
                    cell.fill = self._get_risk_fill(after_level)

            ws.row_dimensions[data_row].height = 50
            data_row += 1

        # 데이터가 없을 경우 빈 행 추가
        if not assessments:
            for _ in range(5):
                for col in range(1, 13):
                    cell = ws.cell(row=data_row, column=col, value="")
                    self._apply_cell_style(cell, self.normal_font, border=self.thin_border, alignment=self.center_align)
                ws.row_dimensions[data_row].height = 50
                data_row += 1

        # 틀 고정 (C8)
        ws.freeze_panes = 'C8'

        # 마지막 행 계산 (헤더 7행 + 데이터 행)
        assessment_count = len(assessments) if assessments else 5
        last_row = 7 + assessment_count

        # 페이지 설정: 가로 방향, A4
        # 데이터가 10행 이상이면 세로는 자동 (여러 페이지 허용)
        fit_height = 0 if assessment_count > 10 else 1
        self._apply_print_settings(ws, f'A1:L{last_row}', fit_to_height=fit_height)
