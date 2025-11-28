# -*- coding: utf-8 -*-
"""
위험성평가표 엑셀 생성 모듈
KRAS 표준 양식 기준
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from risk_data import find_risks_for_work, calculate_risk_level, RISK_DATA


class RiskAssessmentGenerator:
    """위험성평가표 생성기"""

    def __init__(self):
        self.headers = [
            "공정명",
            "세부작업명",
            "위험분류",
            "위험세부분류",
            "위험발생 상황 및 결과",
            "관련근거(법적기준)",
            "현재의 안전보건조치",
            "평가척도",
            "가능성(빈도)",
            "중대성(강도)",
            "현재 위험성",
            "위험성 감소대책",
            "개선후 위험성",
            "개선예정일",
            "완료일",
            "담당자"
        ]

        # 스타일 설정
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 위험성 등급별 색상
        self.risk_colors = {
            "낮음": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # 초록
            "보통": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # 노랑
            "높음": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # 주황
            "매우높음": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # 빨강
        }

    def _format_possibility(self, value: int) -> str:
        """가능성 값 포맷팅"""
        labels = {1: "1(하)", 2: "2(중)", 3: "3(상)"}
        return labels.get(value, str(value))

    def _format_severity(self, value: int) -> str:
        """중대성 값 포맷팅"""
        labels = {1: "1(소)", 2: "2(중)", 3: "3(대)"}
        return labels.get(value, str(value))

    def _format_risk_level(self, score: int, level: str) -> str:
        """위험성 등급 포맷팅"""
        return f"{score}({level})"

    def generate_from_works(self, work_list: list, site_name: str = "", date: str = None) -> list:
        """
        작업 목록에서 위험성평가 데이터 생성

        Args:
            work_list: 작업 내용 리스트
            site_name: 현장명
            date: 평가일자

        Returns:
            위험성평가 데이터 리스트
        """
        if date is None:
            date = datetime.now().strftime("%Y-%m-%d")

        rows = []
        for work in work_list:
            risks = find_risks_for_work(work)
            for risk in risks:
                score, level = calculate_risk_level(risk["가능성"], risk["중대성"])
                row = {
                    "공정명": risk["공정명"],
                    "세부작업명": risk["세부작업명"],
                    "위험분류": risk["위험분류"],
                    "위험세부분류": risk["위험세부분류"],
                    "위험발생 상황 및 결과": risk["위험상황"],
                    "관련근거(법적기준)": risk["관련근거"],
                    "현재의 안전보건조치": "",
                    "평가척도": "3x3",
                    "가능성(빈도)": self._format_possibility(risk["가능성"]),
                    "중대성(강도)": self._format_severity(risk["중대성"]),
                    "현재 위험성": self._format_risk_level(score, level),
                    "위험성 감소대책": risk["감소대책"],
                    "개선후 위험성": self._format_risk_level(score, level),
                    "개선예정일": "",
                    "완료일": "",
                    "담당자": "",
                    "_score": score,
                    "_level": level,
                }
                rows.append(row)

        return rows

    def create_excel(self, data: list, output_path: str, site_name: str = "",
                     eval_date: str = None, company_name: str = ""):
        """
        위험성평가표 엑셀 파일 생성

        Args:
            data: generate_from_works()로 생성한 데이터
            output_path: 출력 파일 경로
            site_name: 현장명
            eval_date: 평가일자
            company_name: 업체명
        """
        if eval_date is None:
            eval_date = datetime.now().strftime("%Y-%m-%d")

        wb = Workbook()
        ws = wb.active
        ws.title = "위험성평가표"

        # 제목 영역
        ws.merge_cells('A1:P1')
        ws['A1'] = "위험성평가 실시표"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 정보 영역
        ws.merge_cells('A2:D2')
        ws['A2'] = f"현장명: {site_name}"
        ws['A2'].font = Font(size=11)

        ws.merge_cells('E2:H2')
        ws['E2'] = f"업체명: {company_name}"
        ws['E2'].font = Font(size=11)

        ws.merge_cells('I2:L2')
        ws['I2'] = f"평가일자: {eval_date}"
        ws['I2'].font = Font(size=11)

        # 헤더 행
        header_row = 4
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_align

        # 데이터 행
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, header in enumerate(self.headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border

                # 가운데 정렬 컬럼
                if header in ["평가척도", "가능성(빈도)", "중대성(강도)", "현재 위험성",
                              "개선후 위험성", "개선예정일", "완료일", "담당자"]:
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.left_align

                # 위험성 셀에 색상 적용
                if header in ["현재 위험성", "개선후 위험성"]:
                    level = row_data.get("_level", "")
                    if level in self.risk_colors:
                        cell.fill = self.risk_colors[level]

        # 열 너비 조정
        column_widths = {
            'A': 15,  # 공정명
            'B': 20,  # 세부작업명
            'C': 12,  # 위험분류
            'D': 15,  # 위험세부분류
            'E': 40,  # 위험발생 상황 및 결과
            'F': 25,  # 관련근거
            'G': 15,  # 현재의 안전보건조치
            'H': 8,   # 평가척도
            'I': 10,  # 가능성
            'J': 10,  # 중대성
            'K': 12,  # 현재 위험성
            'L': 35,  # 위험성 감소대책
            'M': 12,  # 개선후 위험성
            'N': 12,  # 개선예정일
            'O': 12,  # 완료일
            'P': 10,  # 담당자
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 행 높이 조정
        for row in range(header_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 30

        wb.save(output_path)
        print(f"위험성평가표 생성 완료: {output_path}")
        return output_path


def main():
    """테스트 실행"""
    generator = RiskAssessmentGenerator()

    # 테스트: 작업 목록
    work_list = [
        "도면검토및 서류작업",
        "지하1층 CD매립배관",
    ]

    # 위험성평가 데이터 생성
    data = generator.generate_from_works(work_list)

    print(f"\n작업 목록: {work_list}")
    print(f"생성된 위험요소 수: {len(data)}개")
    print("\n위험요소 목록:")
    for i, item in enumerate(data, 1):
        print(f"  {i}. [{item['위험분류']}] {item['위험세부분류']} - {item['현재 위험성']}")

    # 엑셀 파일 생성
    output_path = r"C:\Users\skyjw\OneDrive\03. PYTHON\10. 위험성평가표 자동생성기\test_output.xlsx"
    generator.create_excel(
        data=data,
        output_path=output_path,
        site_name="서부청소년시설 건립공사(소방)",
        company_name="㈜파워이에프씨",
        eval_date="2024-09-30"
    )


if __name__ == "__main__":
    main()
