import io, time, json, logging
from pathlib import Path
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from db import fetchone, fetchall, risk_level

router = APIRouter(prefix="/projects/{pid}/export", tags=["export"])

LOG_DIR = Path("/app/logs")
export_log = logging.getLogger("kras.export")


def _log_export(pid: int, fmt: str, filename: str, rows: int):
    record = {
        "ts": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "project_id": pid,
        "format": fmt,
        "filename": filename,
        "assessment_rows": rows,
    }
    export_log.info(json.dumps(record, ensure_ascii=False))
    try:
        with open(LOG_DIR / "export.jsonl", "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except Exception:
        pass


@router.get("/excel")
def export_excel(pid: int):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(503, "openpyxl not installed")

    # 데이터 조회
    company = fetchone("SELECT * FROM project_company_info WHERE project_id=%s", (pid,)) or {}
    members = fetchall("SELECT * FROM project_org_members WHERE project_id=%s ORDER BY sort_order", (pid,))
    assessments = fetchall("SELECT * FROM project_assessments WHERE project_id=%s ORDER BY sort_order, id", (pid,))
    meeting = fetchone("SELECT * FROM project_forms WHERE project_id=%s AND form_type='meeting'", (pid,))

    wb = Workbook()

    # ── 스타일 공통 ──────────────────────────────────────────────────────────
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    hdr_font  = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subhdr_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    title_font = Font(name="맑은 고딕", size=12, bold=True)
    norm_font  = Font(name="맑은 고딕", size=10)

    RISK_FILLS = {
        "높음": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "보통": PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid"),
        "낮음": PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
    }

    def style(cell, font=None, fill=None, al=None, bdr=True):
        if font: cell.font = font
        if fill: cell.fill = fill
        if al:   cell.alignment = al
        if bdr:  cell.border = border

    def merge_write(ws, rng, value, font=None, fill=None, al=None):
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = value
        style(cell, font=font or norm_font, fill=fill, al=al or center)

    # ── 시트1: 안전보건방침 ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1.안전보건방침"
    ws1.column_dimensions["A"].width = 15
    for col in "BCDEFGH": ws1.column_dimensions[col].width = 18

    ws1.merge_cells("A1:H1")
    ws1["A1"].value = "안전보건방침 및 추진목표"
    ws1["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    ws1["A1"].alignment = center

    info_rows = [
        ("회사명", company.get("company_name",""), "대표자", company.get("ceo_name","")),
        ("업종",   company.get("business_type",""), "평가유형", company.get("eval_type","")),
        ("현장명", company.get("site_name",""),    "평가일자", str(company.get("eval_date","") or "")),
        ("주소",   company.get("address",""),      None, None),
    ]
    r = 3
    for label1, val1, label2, val2 in info_rows:
        merge_write(ws1, f"A{r}:B{r}", label1, fill=subhdr_fill)
        merge_write(ws1, f"C{r}:D{r}", val1)
        if label2:
            merge_write(ws1, f"E{r}:F{r}", label2, fill=subhdr_fill)
            merge_write(ws1, f"G{r}:H{r}", val2)
        else:
            merge_write(ws1, f"E{r}:H{r}", val1)
        r += 1

    ws1.row_dimensions[r].height = 20
    merge_write(ws1, f"A{r}:H{r}", "안전보건방침", font=Font(name="맑은 고딕", size=10, bold=True), fill=subhdr_fill)
    r += 1
    ws1.row_dimensions[r].height = 80
    merge_write(ws1, f"A{r}:H{r}", company.get("safety_policy",""), al=left)
    r += 1
    merge_write(ws1, f"A{r}:H{r}", "추진목표", font=Font(name="맑은 고딕", size=10, bold=True), fill=subhdr_fill)
    r += 1
    ws1.row_dimensions[r].height = 80
    merge_write(ws1, f"A{r}:H{r}", company.get("safety_goal",""), al=left)

    # ── 시트2: 조직구성 ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("2.조직구성")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 55

    merge_write(ws2, "A1:D1", "위험성평가 실시 담당 조직 구성",
                font=title_font, fill=subhdr_fill)
    for col, hdr in zip("ABCD", ["직위/직책","성명","역할","책임 및 권한"]):
        c = ws2[f"{col}2"]
        c.value = hdr; style(c, font=hdr_font, fill=hdr_fill, al=center)

    for i, m in enumerate(members, 3):
        for col, key in zip("ABCD", ["position","name","role","responsibility"]):
            c = ws2[f"{col}{i}"]
            c.value = m.get(key, "")
            style(c, font=norm_font, al=center if col != "D" else left)

    # ── 시트5: 위험성평가실시 (핵심) ──────────────────────────────────────────
    ws5 = wb.create_sheet("5.위험성평가실시")
    hdrs = ["번호","공정/작업명","유해위험요인","관련근거","현재 안전보건조치",
            "가능성(빈도)","중대성(강도)","위험성","위험성 감소대책","개선후 위험성","담당자","완료일"]
    widths = [5,14,28,12,18,8,8,8,22,10,8,10]
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        col_letter = get_column_letter(i)
        ws5.column_dimensions[col_letter].width = w
        c = ws5[f"{col_letter}1"]
        c.value = h
        style(c, font=hdr_font, fill=hdr_fill, al=center)

    def fmt_risk(score, level):
        labels = {"낮음":"낮음","보통":"보통","높음":"높음"}
        return f"{score}({labels.get(level, level)})"

    for i, a in enumerate(assessments, 2):
        ws5.row_dimensions[i].height = 50
        prob, sev = a.get("possibility", 1), a.get("severity", 1)
        cur_score  = a.get("current_risk", prob * sev)
        cur_level  = a.get("current_risk_level", risk_level(cur_score))
        ap, as_    = a.get("after_possibility", 1), a.get("after_severity", 1)
        aft_score  = a.get("after_risk", ap * as_)
        aft_level  = a.get("after_risk_level", risk_level(aft_score))

        hazard = f"[{a.get('risk_category','')}] {a.get('risk_detail','')}\n{a.get('risk_situation','')}"
        row_data = [
            i - 1,
            f"{a.get('process','')}\n{a.get('sub_work','')}",
            hazard,
            a.get("legal_basis", ""),
            a.get("current_measures", ""),
            f"{prob}({'상' if prob==3 else '중' if prob==2 else '하'})",
            f"{sev}({'대' if sev==3 else '중' if sev==2 else '소'})",
            fmt_risk(cur_score, cur_level),
            a.get("reduction_measures", ""),
            fmt_risk(aft_score, aft_level),
            a.get("manager", ""),
            str(a.get("complete_date") or a.get("due_date") or ""),
        ]
        for j, val in enumerate(row_data, 1):
            cl = get_column_letter(j)
            c = ws5[f"{cl}{i}"]
            c.value = val
            c.font = norm_font
            c.alignment = left if j in (2,3,5,9) else center
            c.border = border
            if j == 8:  c.fill = RISK_FILLS.get(cur_level, PatternFill())
            if j == 10: c.fill = RISK_FILLS.get(aft_level, PatternFill())

    # 엑셀 파일 스트림 반환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    company_name = company.get("company_name", "위험성평가") or "위험성평가"
    eval_date    = str(company.get("eval_date", "") or "")
    filename = f"{company_name}_{eval_date}_위험성평가표.xlsx".replace(" ", "_")

    _log_export(pid, "excel", filename, len(assessments))

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
