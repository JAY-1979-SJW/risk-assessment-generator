"""Excel A4 레이아웃 품질 감사 스크립트.

핵심 안전서류 87종 + 부대서류 10종, 총 97개 출력물을 대상으로
A4 인쇄 레이아웃 품질을 자동 감사한다.

Usage:
    python scripts/audit_excel_layout_quality.py
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import engine.output.form_registry as regmod          # noqa: E402
from engine.output.supplementary_registry import (    # noqa: E402
    list_supplemental_types,
    build_supplemental_excel,
)

# ---------------------------------------------------------------------------
# 출력 경로
# ---------------------------------------------------------------------------
JSON_OUT_DIR  = ROOT / "runs/excel_layout_quality_audit"
JSON_OUT_PATH = JSON_OUT_DIR / "latest_layout_audit.json"
MD_OUT_PATH   = ROOT / "docs/reports/excel_layout_quality_audit.md"

# ---------------------------------------------------------------------------
# A4 상수 (단위: 인치, Excel 기본)
# ---------------------------------------------------------------------------
A4_WIDTH_PT  = 595.28   # A4 폭 포인트 (72pt = 1inch)
A4_HEIGHT_PT = 841.89
A4_WIDTH_IN  = 8.27     # 인치
A4_HEIGHT_IN = 11.69

# 인쇄 여백 기본값 (인치)
MARGIN_LEFT_DEFAULT  = 0.75
MARGIN_RIGHT_DEFAULT = 0.75
PRINTABLE_WIDTH_IN   = A4_WIDTH_IN - MARGIN_LEFT_DEFAULT - MARGIN_RIGHT_DEFAULT  # 6.77 in

# Excel 기본 열 너비 단위 → 인치 근사 (문자 너비 기준, 맑은 고딕 10pt)
# 1 column width unit ≈ 0.1098 inch (Excel 내부 환산)
COL_UNIT_TO_INCH = 0.1098

# 행 높이 단위: pt (1pt = 1/72 inch)
# 서명란 최소 권장 높이
SIGNATURE_ROW_HEIGHT_MIN = 18.0  # pt (18pt 미만이면 서명 공간 부족)

# 판정 임계값
WARN_WIDTH_UTILIZATION_LOW  = 0.70  # 70% 미만이면 여백 과다 WARN
WARN_WIDTH_UTILIZATION_HIGH = 1.25  # 125% 초과이면 폭 초과 WARN (fitToWidth 미설정 시)
WARN_WIDTH_UTILIZATION_HIGH_FTW = 1.60  # fitToWidth=1 적용 시 160% 초과만 WARN
FAIL_WIDTH_UTILIZATION_HIGH = 1.40  # 140% 초과이면 FAIL 후보 (fitToWidth 미설정 시)
WARN_USED_ROWS_REPEAT_HDR   = 45   # 45행 초과 반복 헤더 WARN
FAIL_USED_ROWS_REPEAT_HDR   = 60   # 60행 초과 반복 헤더 없으면 FAIL 후보
WARN_SCALE_LOW              = 80   # scale 80 미만 WARN
FAIL_SCALE_LOW              = 70   # scale 70 미만 FAIL 후보

# 표형 문서 keywords (반복 헤더 필요 가능성)
TABLE_KEYWORDS = [
    "명부", "목록", "일지", "대장", "등록부", "점검표", "기록표",
    "측정기록", "개선", "확인서", "서명", "교육일지",
]

# 서명 관련 키워드 (서명란 높이 확인)
SIGNATURE_KEYWORDS = ["서명", "확인자", "승인자", "작성자", "검토자"]


# ---------------------------------------------------------------------------
# 샘플 데이터 생성
# ---------------------------------------------------------------------------
def _make_form_sample(spec: regmod.FormSpec) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for field in spec.required_fields:
        if "date" in field or "year" in field:
            data[field] = "2026-01-01"
        elif "hours" in field or "qty" in field or "count" in field:
            data[field] = 1
        else:
            data[field] = f"[{field}]"
    return data


def _make_supplemental_sample(spec: Dict[str, Any]) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for field in spec.get("required_fields", []):
        if "date" in field or "year" in field:
            data[field] = "2026-01-01"
        elif "hours" in field or "qty" in field or "count" in field:
            data[field] = 1
        else:
            data[field] = f"[{field}]"
    return data


# ---------------------------------------------------------------------------
# openpyxl 분석 함수
# ---------------------------------------------------------------------------

def _get_col_widths_total(ws) -> float:
    """모든 열 너비 합계를 인치로 반환."""
    total_units = 0.0
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width:
            total_units += dim.width
        else:
            total_units += 8.43  # Excel 기본 열 너비
    return total_units * COL_UNIT_TO_INCH


def _get_print_setup(ws) -> Dict[str, Any]:
    """페이지 설정 정보 추출."""
    ps = ws.page_setup
    pm = ws.page_margins
    result = {
        "orientation":    getattr(ps, "orientation", None),
        "fitToWidth":     getattr(ps, "fitToWidth",  None),
        "fitToHeight":    getattr(ps, "fitToHeight", None),
        "scale":          getattr(ps, "scale",       None),
        "paperSize":      getattr(ps, "paperSize",   None),
        "margin_left":    getattr(pm, "left",        0.75),
        "margin_right":   getattr(pm, "right",       0.75),
        "margin_top":     getattr(pm, "top",         1.0),
        "margin_bottom":  getattr(pm, "bottom",      1.0),
        "print_area":     ws.print_area,
        "print_title_rows": getattr(ws, "print_title_rows", None),
    }
    return result


def _get_used_dims(ws) -> Tuple[int, int]:
    """실제 사용된 (행, 열) 수 반환."""
    max_r = ws.max_row   or 1
    max_c = ws.max_column or 1
    return max_r, max_c


def _find_signature_rows(ws) -> List[int]:
    """서명 관련 키워드가 있는 행 번호 목록 반환."""
    sig_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and any(kw in str(cell.value) for kw in SIGNATURE_KEYWORDS):
                sig_rows.append(cell.row)
                break
    return sig_rows


def _check_wrap_text_issues(ws) -> List[str]:
    """긴 문구 셀에 wrap_text 미설정 목록 반환."""
    issues = []
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            val = str(cell.value)
            if len(val) < 20:
                continue
            align = cell.alignment
            if align and align.wrap_text:
                continue
            # 병합된 셀인지 확인
            if cell.coordinate in [str(mr) for mr in ws.merged_cells.ranges]:
                continue
            issues.append(f"R{cell.row}C{cell.column} 긴문구({len(val)}자) wrap_text 없음")
    return issues[:5]  # 최대 5개만 반환


def _check_row_height_issues(ws, sig_rows: List[int]) -> List[str]:
    """서명란/비고란 행 높이 부족 목록."""
    issues = []
    for r in sig_rows:
        dim = ws.row_dimensions.get(r)
        h = dim.height if dim and dim.height else 15.0
        if h < SIGNATURE_ROW_HEIGHT_MIN:
            issues.append(f"R{r} 서명란 높이 {h:.0f}pt < {SIGNATURE_ROW_HEIGHT_MIN:.0f}pt")
    return issues[:3]


def _is_table_form(display_name: str) -> bool:
    return any(kw in display_name for kw in TABLE_KEYWORDS)


# ---------------------------------------------------------------------------
# 단일 xlsx 분석
# ---------------------------------------------------------------------------

def _analyze_wb(
    wb_bytes: bytes,
    display_name: str,
) -> Dict[str, Any]:
    """xlsx bytes를 분석하여 레이아웃 품질 결과 dict 반환."""
    fails: List[str] = []
    warns: List[str] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    except Exception as e:
        return {"error": str(e), "verdict": "ERROR", "fails": [], "warns": []}

    ws = wb.active

    # ── 페이지 설정 ────────────────────────────────────────────────────────
    ps = _get_print_setup(ws)
    used_rows, used_cols = _get_used_dims(ws)

    # A4 용지 설정 확인 (paperSize 9 = A4)
    paper_size = ps["paperSize"]
    is_a4 = (paper_size == 9 or paper_size is None)  # None이면 기본=Letter지만 묵인
    if paper_size is not None and paper_size != 9:
        warns.append(f"A4 용지 미설정 (paperSize={paper_size})")

    # orientation
    orientation = ps["orientation"]
    is_landscape = (orientation == "landscape")

    # fitToWidth
    ftw = ps["fitToWidth"]
    fth = ps["fitToHeight"]
    scale = ps["scale"]

    if ftw is None or ftw == 0:
        warns.append("fitToWidth 미설정 — A4 폭 자동 맞춤 없음")

    # fitToHeight 경고
    if fth == 1:
        if used_rows > FAIL_USED_ROWS_REPEAT_HDR:
            fails.append(f"fitToHeight=1 + used_rows={used_rows} — 다행 문서 강제 축소 FAIL")
        elif used_rows > WARN_USED_ROWS_REPEAT_HDR:
            warns.append(f"fitToHeight=1 + used_rows={used_rows} — 과도한 축소 가능")

    # scale 확인
    if scale is not None:
        if scale < FAIL_SCALE_LOW:
            fails.append(f"scale={scale} < {FAIL_SCALE_LOW} — 글자 너무 작음")
        elif scale < WARN_SCALE_LOW:
            warns.append(f"scale={scale} < {WARN_SCALE_LOW} — 가독성 저하 가능")

    # print_area — 미설정은 권고사항으로만 기록 (WARN 대상 아님)
    _has_print_area = bool(ps["print_area"])

    # ── 폭 활용도 ──────────────────────────────────────────────────────────
    margin_l = ps["margin_left"]  or MARGIN_LEFT_DEFAULT
    margin_r = ps["margin_right"] or MARGIN_RIGHT_DEFAULT
    if is_landscape:
        printable_w = A4_HEIGHT_IN - margin_l - margin_r  # 가로 방향
    else:
        printable_w = A4_WIDTH_IN  - margin_l - margin_r

    col_total_in = _get_col_widths_total(ws)
    util_ratio = col_total_in / printable_w if printable_w > 0 else 0.0

    if ftw == 1:
        # fitToWidth=1: Excel이 자동 축소 — 160% 초과 시만 WARN (글자 너무 작아짐)
        if util_ratio > WARN_WIDTH_UTILIZATION_HIGH_FTW:
            warns.append(
                f"열 너비 합계 {col_total_in:.2f}in > 출력폭 {printable_w:.2f}in "
                f"({util_ratio:.0%}) — fitToWidth=1이나 과도한 축소 가능"
            )
    else:
        if util_ratio > FAIL_WIDTH_UTILIZATION_HIGH:
            fails.append(
                f"열 너비 합계 {col_total_in:.2f}in > 출력폭 {printable_w:.2f}in "
                f"({util_ratio:.0%}) — fitToWidth 미설정, 인쇄 시 잘림"
            )
        elif util_ratio > WARN_WIDTH_UTILIZATION_HIGH:
            warns.append(
                f"열 너비 합계 {col_total_in:.2f}in가 출력폭 초과 가능 ({util_ratio:.0%})"
            )
    if util_ratio < WARN_WIDTH_UTILIZATION_LOW and used_cols >= 3:
        warns.append(
            f"폭 활용도 {util_ratio:.0%} < {WARN_WIDTH_UTILIZATION_LOW:.0%} — 여백 과다"
        )

    # ── 반복 헤더 ──────────────────────────────────────────────────────────
    ptr = ps["print_title_rows"]
    has_repeat_hdr = bool(ptr)

    if not has_repeat_hdr:
        if used_rows > FAIL_USED_ROWS_REPEAT_HDR:
            fails.append(
                f"used_rows={used_rows} > {FAIL_USED_ROWS_REPEAT_HDR} — "
                "반복 헤더(print_title_rows) 없음"
            )
        elif used_rows > WARN_USED_ROWS_REPEAT_HDR:
            warns.append(
                f"used_rows={used_rows} > {WARN_USED_ROWS_REPEAT_HDR} — "
                "반복 헤더 권장"
            )
        elif _is_table_form(display_name):
            warns.append("표형 문서 — 반복 헤더 설정 권장")

    # ── wrap_text ──────────────────────────────────────────────────────────
    wrap_issues = _check_wrap_text_issues(ws)
    if wrap_issues:
        warns.append(f"긴 문구 wrap_text 누락 {len(wrap_issues)}건")

    # ── 서명란 높이 ────────────────────────────────────────────────────────
    sig_rows = _find_signature_rows(ws)
    sig_height_issues = _check_row_height_issues(ws, sig_rows)
    for iss in sig_height_issues:
        warns.append(iss)

    # 서명란 페이지 끝 잘림 위험 — 서명란이 맨 마지막 행이고 문서가 50행 이상인 경우만
    if sig_rows and used_rows >= 50:
        last_sig = max(sig_rows)
        if last_sig == used_rows and fth != 1:
            warns.append(
                f"서명란(R{last_sig})이 최종행 — 페이지 분리 시 잘림 가능"
            )

    # ── 최종 판정 ──────────────────────────────────────────────────────────
    if fails:
        verdict = "FAIL"
    elif warns:
        verdict = "WARN"
    else:
        verdict = "PASS"

    return {
        "error":              None,
        "verdict":            verdict,
        "fails":              fails,
        "warns":              warns,
        "used_rows":          used_rows,
        "used_cols":          used_cols,
        "orientation":        orientation or "portrait",
        "fitToWidth":         ftw,
        "fitToHeight":        fth,
        "scale":              scale,
        "paper_size":         paper_size,
        "print_area":         ps["print_area"],
        "print_title_rows":   ptr,
        "col_total_in":       round(col_total_in, 3),
        "printable_width_in": round(printable_w, 3),
        "width_utilization":  round(util_ratio, 3),
        "sig_rows":           sig_rows,
        "is_table_form":      _is_table_form(display_name),
        "xlsx_bytes":         len(wb_bytes),
    }


# ---------------------------------------------------------------------------
# 메인 감사 루프
# ---------------------------------------------------------------------------

def run_audit() -> Dict[str, Any]:
    results_core: List[Dict[str, Any]] = []
    results_supp: List[Dict[str, Any]] = []

    # ── 핵심서류 87종 ─────────────────────────────────────────────────────
    print("[ 핵심서류 감사 중... ]")
    for form_type, spec in sorted(regmod._REGISTRY.items()):
        rec: Dict[str, Any] = {
            "kind":         "core",
            "form_type":    form_type,
            "display_name": spec.display_name,
            "builder_file": f"engine/output/{form_type}_builder.py",
            "verdict":      "ERROR",
            "fails": [], "warns": [],
            "used_rows": 0, "used_cols": 0,
            "orientation": "portrait",
            "fitToWidth": None, "fitToHeight": None, "scale": None,
            "paper_size": None, "print_area": None, "print_title_rows": None,
            "col_total_in": 0.0, "printable_width_in": 6.77, "width_utilization": 0.0,
            "sig_rows": [], "is_table_form": False,
            "xlsx_bytes": 0, "error": None,
        }
        try:
            sample = _make_form_sample(spec)
            raw = spec.builder(sample)
            analysis = _analyze_wb(raw, spec.display_name)
            rec.update(analysis)
        except Exception as e:
            rec["error"] = str(e)
            rec["verdict"] = "ERROR"

        results_core.append(rec)
        sym = {"PASS": "✓", "WARN": "!", "FAIL": "✗", "ERROR": "E"}.get(rec["verdict"], "?")
        print(f"  {sym} {form_type:<55} {rec['verdict']}")

    # ── 부대서류 10종 ─────────────────────────────────────────────────────
    print("\n[ 부대서류 감사 중... ]")
    for spec_meta in list_supplemental_types():
        st = spec_meta["supplemental_type"]
        rec: Dict[str, Any] = {
            "kind":            "supplemental",
            "form_type":       st,
            "display_name":    spec_meta["display_name"],
            "builder_file":    f"engine/output/{st}_builder.py",
            "verdict":         "ERROR",
            "fails": [], "warns": [],
            "used_rows": 0, "used_cols": 0,
            "orientation": "portrait",
            "fitToWidth": None, "fitToHeight": None, "scale": None,
            "paper_size": None, "print_area": None, "print_title_rows": None,
            "col_total_in": 0.0, "printable_width_in": 6.77, "width_utilization": 0.0,
            "sig_rows": [], "is_table_form": False,
            "xlsx_bytes": 0, "error": None,
        }
        if not spec_meta["has_builder"]:
            rec["error"] = "TODO — builder 미구현"
            results_supp.append(rec)
            continue
        try:
            sample = _make_supplemental_sample(spec_meta)
            raw = build_supplemental_excel(st, sample)
            analysis = _analyze_wb(raw, spec_meta["display_name"])
            rec.update(analysis)
        except Exception as e:
            rec["error"] = str(e)
            rec["verdict"] = "ERROR"

        results_supp.append(rec)
        sym = {"PASS": "✓", "WARN": "!", "FAIL": "✗", "ERROR": "E"}.get(rec["verdict"], "?")
        print(f"  {sym} {st:<55} {rec['verdict']}")

    all_results = results_core + results_supp
    return {
        "generated_at": datetime.now().isoformat(),
        "core_count":   len(results_core),
        "supp_count":   len(results_supp),
        "results":      all_results,
    }


# ---------------------------------------------------------------------------
# 보고서 생성
# ---------------------------------------------------------------------------

def _count_by_verdict(results: List[Dict], verdicts) -> int:
    if isinstance(verdicts, str):
        verdicts = [verdicts]
    return sum(1 for r in results if r.get("verdict") in verdicts)


def write_reports(data: Dict[str, Any]) -> None:
    JSON_OUT_DIR.mkdir(parents=True, exist_ok=True)
    MD_OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with open(JSON_OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    results  = data["results"]
    core     = [r for r in results if r["kind"] == "core"]
    supp     = [r for r in results if r["kind"] == "supplemental"]

    total        = len(results)
    pass_n       = _count_by_verdict(results, "PASS")
    warn_n       = _count_by_verdict(results, "WARN")
    fail_n       = _count_by_verdict(results, "FAIL")
    error_n      = _count_by_verdict(results, "ERROR")

    core_pass    = _count_by_verdict(core, "PASS")
    core_warn    = _count_by_verdict(core, "WARN")
    core_fail    = _count_by_verdict(core, "FAIL")

    supp_pass    = _count_by_verdict(supp, "PASS")
    supp_warn    = _count_by_verdict(supp, "WARN")
    supp_fail    = _count_by_verdict(supp, "FAIL")

    # A4 설정 통계
    ftw_set      = sum(1 for r in results if r.get("fitToWidth") == 1)
    fth_set      = sum(1 for r in results if r.get("fitToHeight") == 1)
    scale_low    = sum(1 for r in results if r.get("scale") and r["scale"] < WARN_SCALE_LOW)
    landscape_n  = sum(1 for r in results if r.get("orientation") == "landscape")
    no_print_area= sum(1 for r in results if not r.get("print_area"))
    no_fittowidth= sum(1 for r in results if r.get("fitToWidth") != 1 and r.get("verdict") != "ERROR")

    # 폭 활용도 이슈
    low_util  = [r for r in results if r.get("width_utilization", 1) < WARN_WIDTH_UTILIZATION_LOW
                 and r.get("used_cols", 0) >= 3 and r.get("verdict") != "ERROR"]
    high_util = [r for r in results if r.get("width_utilization", 0) > WARN_WIDTH_UTILIZATION_HIGH
                 and r.get("verdict") != "ERROR"]

    low_util.sort(key=lambda r: r.get("width_utilization", 1))
    high_util.sort(key=lambda r: -r.get("width_utilization", 0))

    # 반복 헤더 이슈
    heavy_rows   = [r for r in results if r.get("used_rows", 0) > WARN_USED_ROWS_REPEAT_HDR
                    and r.get("verdict") != "ERROR"]
    no_hdr       = [r for r in heavy_rows if not r.get("print_title_rows")]

    # 위험도 상위 정렬 (FAIL > WARN)
    risk_order   = {"FAIL": 0, "WARN": 1, "PASS": 2, "ERROR": 3}
    top30        = sorted(
        [r for r in results if r.get("verdict") in ("FAIL", "WARN")],
        key=lambda r: (risk_order.get(r.get("verdict"), 9),
                       -(len(r.get("fails", [])) * 10 + len(r.get("warns", []))))
    )[:30]

    lines = []
    a = lines.append

    a("# Excel A4 레이아웃 품질 감사 보고서")
    a("")
    a(f"**생성일**: {data['generated_at'][:10]}  ")
    a(f"**검사 대상**: 핵심서류 {data['core_count']}종 + 부대서류 {data['supp_count']}종 = **총 {total}종**  ")
    a(f"**분석 완료**: {total - error_n}종  ")
    a("")
    a("---")
    a("")

    # 1. 전체 요약
    a("## 1. 전체 요약")
    a("")
    a("| 구분 | 검사 수 | PASS | WARN | FAIL 후보 | ERROR |")
    a("|------|---------|------|------|-----------|-------|")
    a(f"| 핵심서류 | {len(core)} | {core_pass} | {core_warn} | {core_fail} | {_count_by_verdict(core, 'ERROR')} |")
    a(f"| 부대서류 | {len(supp)} | {supp_pass} | {supp_warn} | {supp_fail} | {_count_by_verdict(supp, 'ERROR')} |")
    a(f"| **합계** | **{total}** | **{pass_n}** | **{warn_n}** | **{fail_n}** | **{error_n}** |")
    a("")

    # 2. A4 페이지 설정 요약
    a("## 2. A4 페이지 설정 요약")
    a("")
    a("| 항목 | 수량 | 비율 |")
    a("|------|------|------|")
    a(f"| fitToWidth=1 적용 | {ftw_set} | {ftw_set/total:.0%} |")
    a(f"| fitToWidth 미설정 | {no_fittowidth} | {no_fittowidth/total:.0%} |")
    a(f"| fitToHeight=1 (강제 축소) | {fth_set} | {fth_set/total:.0%} |")
    a(f"| scale 80 미만 | {scale_low} | {scale_low/total:.0%} |")
    a(f"| 가로 (landscape) | {landscape_n} | {landscape_n/total:.0%} |")
    a(f"| print_area 미설정 | {no_print_area} | {no_print_area/total:.0%} |")
    a("")

    # 3. 폭 활용도
    a("## 3. 페이지 폭 활용도")
    a("")
    a(f"**여백 과다** (활용도 < {WARN_WIDTH_UTILIZATION_LOW:.0%}): **{len(low_util)}종**  ")
    a(f"**폭 초과** (활용도 > {WARN_WIDTH_UTILIZATION_HIGH:.0%}): **{len(high_util)}종**  ")
    a("")

    if low_util:
        a(f"### 3-1. 여백 과다 상위 (최대 20종)")
        a("")
        a("| form_type | 표시명 | 활용도 | 열너비합(in) | 출력폭(in) |")
        a("|-----------|--------|--------|-------------|-----------|")
        for r in low_util[:20]:
            a(f"| {r['form_type']} | {r['display_name']} | "
              f"{r['width_utilization']:.0%} | {r['col_total_in']:.2f} | {r['printable_width_in']:.2f} |")
        a("")

    if high_util:
        a(f"### 3-2. 폭 초과 상위 (최대 20종)")
        a("")
        a("| form_type | 표시명 | 활용도 | 열너비합(in) | 출력폭(in) |")
        a("|-----------|--------|--------|-------------|-----------|")
        for r in high_util[:20]:
            a(f"| {r['form_type']} | {r['display_name']} | "
              f"{r['width_utilization']:.0%} | {r['col_total_in']:.2f} | {r['printable_width_in']:.2f} |")
        a("")

    # 4. 반복 헤더
    a("## 4. 페이지 분할 / 반복 헤더")
    a("")
    a(f"**행 수 > {WARN_USED_ROWS_REPEAT_HDR}인 문서**: {len(heavy_rows)}종  ")
    a(f"**반복 헤더 미설정**: {len(no_hdr)}종  ")
    a(f"**반복 헤더 설정**: {sum(1 for r in results if r.get('print_title_rows'))}종  ")
    a("")

    if no_hdr:
        a("### 반복 헤더 필요 후보")
        a("")
        a("| form_type | 표시명 | used_rows | 표형 |")
        a("|-----------|--------|-----------|------|")
        for r in sorted(no_hdr, key=lambda x: -x.get("used_rows", 0)):
            a(f"| {r['form_type']} | {r['display_name']} | "
              f"{r['used_rows']} | {'Y' if r.get('is_table_form') else '-'} |")
        a("")

    # 5. 셀 겹침/잘림 위험 요약
    a("## 5. 셀 겹침 / 잘림 위험 요약")
    a("")
    wrap_issue_n = sum(1 for r in results
                       if any("wrap_text" in w for w in r.get("warns", [])))
    sig_issue_n  = sum(1 for r in results
                       if any("서명란" in w for w in r.get("warns", [])))
    sig_page_n   = sum(1 for r in results
                       if any("잘림 가능" in w for w in r.get("warns", [])))

    a(f"| 위험 유형 | 건수 |")
    a(f"|----------|------|")
    a(f"| 긴 문구 wrap_text 누락 | {wrap_issue_n} |")
    a(f"| 서명란 높이 부족 가능 | {sig_issue_n} |")
    a(f"| 서명란 페이지 끝 잘림 가능 | {sig_page_n} |")
    a("")

    # 6. 위험도 Top 30
    a("## 6. 문서별 위험도 Top 30")
    a("")
    a("| # | form_type | 표시명 | 판정 | FAIL수 | WARN수 | 주요 문제 | 권장 조치 |")
    a("|---|-----------|--------|------|--------|--------|----------|----------|")
    for i, r in enumerate(top30, 1):
        fails_str = "; ".join(r.get("fails", []))[:60]
        warns_str = "; ".join(r.get("warns", []))[:50]
        problem   = (fails_str or warns_str)[:60]
        # 권장 조치 자동 분류
        if any("fitToWidth" in w for w in r.get("warns", []) + r.get("fails", [])):
            action = "인쇄설정 추가"
        elif any("잘림" in w or "초과" in w for w in r.get("fails", [])):
            action = "열너비 조정"
        elif any("반복 헤더" in w for w in r.get("warns", []) + r.get("fails", [])):
            action = "print_title_rows 추가"
        elif any("여백" in w for w in r.get("warns", [])):
            action = "열너비 확대"
        elif any("서명" in w for w in r.get("warns", [])):
            action = "서명란 높이 조정"
        else:
            action = "개별 확인"
        a(f"| {i} | {r['form_type'][:40]} | {r['display_name'][:18]} | "
          f"**{r['verdict']}** | {len(r.get('fails',[]))} | {len(r.get('warns',[]))} | "
          f"{problem[:55]} | {action} |")
    a("")

    # 7. 공통 원인 분석
    a("## 7. 공통 원인 분석")
    a("")
    helper_fixable = sum(1 for r in results
                         if any("fitToWidth" in w or "print_area" in w
                                for w in r.get("warns", []) + r.get("fails", [])))
    individual_fix = sum(1 for r in results
                         if any("열너비" in w or "초과" in w or "반복 헤더" in w
                                for w in r.get("warns", []) + r.get("fails", [])))
    supp_fix       = sum(1 for r in supp
                         if r.get("verdict") in ("WARN", "FAIL"))

    a(f"| 원인 분류 | 해당 문서 수 | 설명 |")
    a(f"|----------|------------|------|")
    a(f"| excel_style_helpers 공통 보정 | {helper_fixable} | fitToWidth, print_area, 여백 미설정 |")
    a(f"| 개별 builder 보정 필요 | {individual_fix} | 열너비 초과/부족, 반복 헤더 누락 |")
    a(f"| 부대서류 보정 필요 | {supp_fix} | 부대서류 builder 레이아웃 개선 |")
    a("")
    a("### 7-1. excel_style_helpers에서 공통 보정 가능한 항목")
    a("")
    a("- `apply_col_widths()` 호출 시 `fitToWidth=1`, `print_area` 자동 설정 추가")
    a("- 공통 `set_print_setup(ws, landscape=False)` 헬퍼 함수 추가 권장")
    a("- 여백: `ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)`")
    a("")
    a("### 7-2. 개별 builder 보정이 필요한 항목")
    a("")
    a("- 열 너비 합계가 출력폭 초과하는 경우 → 각 builder의 `_COL_WIDTHS` 조정")
    a("- 45행 이상 표형 문서 → `ws.print_title_rows = '1:3'` 추가")
    a("- 서명란 row_height 부족 → 서명 행에 `ws.row_dimensions[row].height = 30` 적용")
    a("")
    a("### 7-3. 부대서류 전용 보정 필요")
    a("")
    a("- 부대서류 10종은 excel_style_helpers 기반으로 작성되었으나 인쇄 설정 미추가")
    a("- `build_*` 함수 말미에 공통 `set_print_setup(ws)` 호출 추가로 일괄 해결 가능")
    a("")

    # 8. 다음 단계
    a("## 8. 다음 단계")
    a("")
    a("| 단계 | 작업 | 예상 효과 |")
    a("|------|------|----------|")
    a("| 2단계 | excel_style_helpers에 `set_print_setup()` 공통 함수 추가 | fitToWidth/print_area 일괄 적용 |")
    a("| 3단계 | 핵심서류 Top 위험 문서 개별 열너비·반복헤더 보정 | FAIL 후보 0건 목표 |")
    a("| 4단계 | 부대서류 10종 인쇄 설정 일괄 추가 | 부대서류 WARN 해소 |")
    a("| 5단계 | 전체 97종 재감사 — Excel Layout QA v1.0 마감 보고서 작성 | PASS 90% 이상 목표 |")
    a("")

    # 최종 판정
    overall = "PASS" if fail_n == 0 else ("WARN" if fail_n < 5 else "FAIL")
    a("---")
    a("")
    a(f"```")
    a(f"전체: {total}종  |  PASS: {pass_n}  |  WARN: {warn_n}  |  FAIL 후보: {fail_n}  |  ERROR: {error_n}")
    a(f"핵심: {len(core)}종  PASS: {core_pass}  WARN: {core_warn}  FAIL: {core_fail}")
    a(f"부대: {len(supp)}종  PASS: {supp_pass}  WARN: {supp_warn}  FAIL: {supp_fail}")
    a(f"최종 판정: {overall}")
    a(f"```")

    with open(MD_OUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"\n  JSON  → {JSON_OUT_PATH}")
    print(f"  보고서 → {MD_OUT_PATH}")
    print("=" * 70)
    print(f"  전체: {total}종  |  PASS: {pass_n}  |  WARN: {warn_n}  |  "
          f"FAIL 후보: {fail_n}  |  ERROR: {error_n}")
    print("=" * 70)
    print(f"  최종 판정: {overall}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    data = run_audit()
    write_reports(data)
