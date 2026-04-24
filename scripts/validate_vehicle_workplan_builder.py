"""
차량계 건설기계 작업계획서 Excel builder 검증 스크립트.

검증 항목:
  1.  xlsx bytes 정상 생성 여부
  2.  openpyxl 재오픈 성공
  3.  시트명 == "차량계건설기계작업계획서"
  4.  제목 셀(A1) == "차량계 건설기계 작업계획서"
  5.  부제 셀(A2) == SHEET_SUBTITLE
  6.  법정 필수항목 라벨 존재 (LEGAL_LABELS 전체)
  7.  운행경로 텍스트 영역 라벨 존재 ("운행경로 기술")
  8.  운행경로 스케치 박스 존재 (헤더 행 + 빈 행 5행)
  9.  guide_worker_required 라벨 존재
  10. access_control 라벨 존재
  11. emergency_contact 라벨 존재
  12. 위험요소/안전조치 표 헤더 존재
  13. 작업 전 점검 사항 표 헤더 존재
  14. print_area 존재
  15. 빈 form_data에서도 파일 생성 성공
  16. 샘플 form_data에서 주요 값 정상 렌더링

실행:
  python scripts/validate_vehicle_workplan_builder.py
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from engine.output.vehicle_workplan_builder import (
    LEGAL_LABELS,
    MAX_HAZARD,
    MAX_CHECKS,
    SHEET_HEADING,
    SHEET_NAME,
    SHEET_SUBTITLE,
    build_vehicle_workplan_excel,
)

# ---------------------------------------------------------------------------
# 샘플 form_data
# ---------------------------------------------------------------------------
SAMPLE_FORM_DATA = {
    "site_name": "테스트 사업장 (주)",
    "project_name": "○○ 공장 신축공사 현장",
    "work_location": "현장 서측 진입로 구간 및 토공 작업구역",
    "work_date": "2026-04-25 ~ 2026-06-30",
    "supervisor": "김현장 (010-0000-1111)",
    "contractor": "㈜테스트건설",
    "prepared_by": "이안전",
    # 장비정보 [LAW]
    "machine_type": "굴착기 (백호우, 불도저)",
    "machine_capacity": "굴착기 0.8m³, 불도저 D6T (95kW)",
    "operator_name": "박기사",
    "operator_license": "건설기계조종사면허 제20-가-12345호",
    # 작업조건 [LAW+PRAC]
    "work_method": (
        "1. 굴착기로 표층 토사 제거 → 2. 불도저로 정지 작업 → "
        "3. 덤프트럭 반출 → 4. 구간별 단계 반복 (STA.순)"
    ),
    "guide_worker_required": "배치 필요 — 유도자 1명 상시 배치 (무전기 휴대), 기계 후진 시 신호 확인",
    "speed_limit": "현장 내 20km/h 이하",
    "work_radius": "굴착기 작업반경 7m, 출입 금지구역 표시",
    "ground_survey": "점토질 지반, 지하수위 G.L.-2.5m, 인근 구조물 없음",
    # 운행경로 [LAW]
    "travel_route_text": (
        "진입: 현장 정문(동측) → 현장 내 주도로(폭 6m) → 토공작업구역(서측)\n"
        "반출: 작업구역 → 주도로 → 출구(남측 임시 게이트) → 공사용 도로\n"
        "속도제한 구역: 정문~교차로 구간 (20km/h), 작업구역 인근 (10km/h)"
    ),
    "travel_route_sketch_note": "※ 운행경로 개략도를 아래 공간에 수기로 기재하시오 (진입로·구내동선·주요 위험구간 표시)",
    # 출입통제/유도자/비상연락망
    "access_control": "작업반경 7m 내 로프 바리케이드 설치, 출입금지 표지판 부착, 유도자 통제",
    "emergency_contact": "현장소장: 010-0000-1111 / 안전관리자: 010-2222-3333 / 119 소방서",
    "emergency_measure": "전도·충돌 발생 시: 즉시 작업 중단 → 안전관리자 연락 → 119 신고 → 현장 폐쇄",
    # 위험요소/안전조치
    "hazard_items": [
        {"hazard": "굴착기 후진 중 근로자 충돌",    "safety_measure": "유도자 배치 및 후방 경보 장치 작동 확인"},
        {"hazard": "불도저 경사지 주행 중 전도",     "safety_measure": "경사도 30° 이하 구간 작업, 안전벨트 착용"},
        {"hazard": "덤프트럭 과적 및 운행 중 낙석", "safety_measure": "적재함 덮개 사용, 과적 금지 (최대 적재량 준수)"},
        {"hazard": "작업반경 내 근로자 접근",        "safety_measure": "출입금지 구역 바리케이드 및 표지판 설치"},
    ],
    # 작업 전 점검
    "pre_check_items": [
        {"check_item": "제동장치·조향장치 이상 유무",    "result": "",  "note": ""},
        {"check_item": "유압장치 오일 누유 여부",         "result": "",  "note": ""},
        {"check_item": "버킷·블레이드 균열 여부",         "result": "",  "note": ""},
        {"check_item": "안전벨트 작동 상태",              "result": "",  "note": ""},
        {"check_item": "후방 카메라·경보 장치 작동 여부", "result": "",  "note": ""},
        {"check_item": "연료·냉각수·오일 레벨 확인",     "result": "",  "note": ""},
    ],
    "sign_date": "2026-04-25",
}

SAMPLE_EMPTY_DATA: dict = {}

# ---------------------------------------------------------------------------
# 검증 헬퍼
# ---------------------------------------------------------------------------

def _check(condition: bool, name: str, detail: str = "") -> bool:
    status = "PASS" if condition else "FAIL"
    msg = f"  [{status}] {name}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    return condition


def _all_cell_values(ws) -> set[str]:
    vals: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                vals.add(str(v))
    return vals


def _merged_ranges(ws) -> set[str]:
    return {str(m) for m in ws.merged_cells.ranges}


# ---------------------------------------------------------------------------
# 검증 함수
# ---------------------------------------------------------------------------

def validate_full(xlsx_bytes: bytes, label: str) -> list[bool]:
    results: list[bool] = []
    print(f"\n=== {label} ===")

    # 1. bytes 생성
    results.append(_check(
        isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
        "xlsx bytes 정상 생성",
        f"{len(xlsx_bytes):,} bytes",
    ))

    # 2. 재오픈
    try:
        wb = load_workbook(BytesIO(xlsx_bytes))
        results.append(_check(True, "openpyxl 재오픈 성공"))
    except Exception as e:
        results.append(_check(False, "openpyxl 재오픈", str(e)))
        return results

    # 3. 시트명
    results.append(_check(
        SHEET_NAME in wb.sheetnames,
        f"시트명 == '{SHEET_NAME}'",
        str(wb.sheetnames),
    ))
    ws = wb[SHEET_NAME]

    # 4. 제목 셀 A1
    title_val = ws.cell(row=1, column=1).value
    results.append(_check(
        title_val == SHEET_HEADING,
        f"제목(A1) == '{SHEET_HEADING}'",
        repr(title_val),
    ))

    # 5. 부제 셀 A2
    sub_val = ws.cell(row=2, column=1).value
    results.append(_check(
        sub_val == SHEET_SUBTITLE,
        "부제(A2) == SHEET_SUBTITLE",
        repr(sub_val),
    ))

    all_vals = _all_cell_values(ws)

    # 6. 법정 필수항목 라벨 전체 존재
    for lbl in LEGAL_LABELS:
        results.append(_check(
            lbl in all_vals,
            f"법정항목 라벨 존재: '{lbl}'",
        ))

    # 7. 운행경로 텍스트 영역 라벨 존재
    results.append(_check(
        "운행경로 기술" in all_vals,
        "운행경로 텍스트 영역 라벨 ('운행경로 기술') 존재",
    ))

    # 8. 운행경로 스케치 박스 존재 — 스케치 헤더 문구 포함 여부로 확인
    sketch_present = any("운행경로 개략도" in v or "수기로 기재" in v for v in all_vals)
    results.append(_check(
        sketch_present,
        "운행경로 스케치 박스 헤더 존재",
    ))

    # 9. guide_worker_required 라벨 존재
    results.append(_check(
        "유도자 배치 여부 및 방법" in all_vals,
        "유도자 배치 여부 및 방법 라벨 존재",
    ))

    # 10. access_control 라벨 존재
    results.append(_check(
        "출입통제 방법" in all_vals,
        "출입통제 방법 라벨 존재",
    ))

    # 11. emergency_contact 라벨 존재
    results.append(_check(
        "비상연락망" in all_vals,
        "비상연락망 라벨 존재",
    ))

    # 12. 위험요소/안전조치 표 헤더 존재
    hazard_table_ok = ("위험요소 및 안전조치" in all_vals
                       and "위험 요소" in all_vals
                       and "안전 조치" in all_vals)
    results.append(_check(
        hazard_table_ok,
        "위험요소/안전조치 표 헤더 존재",
    ))

    # 13. 작업 전 점검 사항 표 헤더 존재 (부분문자열 검색)
    precheck_section = any("작업 전 점검 사항" in v for v in all_vals)
    precheck_col_hdr = any("점검 항목" in v for v in all_vals) and any("이상유무" in v for v in all_vals)
    results.append(_check(
        precheck_section and precheck_col_hdr,
        "작업 전 점검 사항 표 헤더 존재",
    ))

    # 14. print_area 존재
    results.append(_check(
        bool(ws.print_area),
        "print_area 설정 존재",
        repr(ws.print_area),
    ))

    # 15. 주요 병합셀 확인
    merged = _merged_ranges(ws)
    key_merges = [
        "A1:H1",   # 제목
        "A2:H2",   # 부제
        "B3:D3",   # site_name
        "F3:H3",   # project_name
        "B4:H4",   # work_location 전폭
        "B8:H8",   # machine_type
        "B9:H9",   # machine_capacity
        "B10:H10", # work_method
        "B12:H12", # guide_worker_required
        "B16:H16", # travel_route_text
    ]
    for mr in key_merges:
        results.append(_check(mr in merged, f"병합셀 존재: {mr}"))

    return results


def validate_empty(xlsx_bytes: bytes) -> list[bool]:
    """빈 form_data로 생성: 헤더·라벨 외 값 필드가 공란인지 확인."""
    results: list[bool] = []
    print("\n=== 공란 유지 검증 (빈 form_data) ===")

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb[SHEET_NAME]

    # 헤더/라벨로 허용되는 접두사 집합 (부분문자열 포함 여부로 판단)
    header_prefixes = (
        SHEET_HEADING, SHEET_SUBTITLE,
        "사업장명", "현장명", "작업 위치", "작업 기간", "작업 책임자",
        "도급업체", "작성자",
        "법정 기재 사항", "기계의 종류", "기계의 성능", "작업방법",
        "운전자 성명", "운전자 자격", "유도자 배치",
        "안전속도 제한", "작업반경", "지형·지반",
        "운행경로", "출입통제", "비상연락망", "비상조치",
        "위험요소 및 안전조치", "순번", "위험 요소", "안전 조치",
        "작업 전 점검 사항", "점검 항목", "이상유무", "비고",
        "확인 및 서명", "작성자 서명", "작업책임자 서명", "작성일",
    )

    def _is_header(v: str) -> bool:
        return any(v.startswith(p) or p in v for p in header_prefixes)
    non_blank_non_header: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is None or v == "":
                continue
            sv = str(v)
            if _is_header(sv):
                continue
            if isinstance(v, int):  # 순번 정수 허용
                continue
            non_blank_non_header.append(sv)

    results.append(_check(
        len(non_blank_non_header) == 0,
        "모든 값 필드 공란 유지",
        f"비공란 비헤더 값: {non_blank_non_header[:5] if non_blank_non_header else '없음'}",
    ))
    return results


def validate_sample_values(xlsx_bytes: bytes) -> list[bool]:
    """샘플 form_data에서 주요 값이 정상 렌더링되는지 확인."""
    results: list[bool] = []
    print("\n=== 샘플 값 렌더링 검증 ===")

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb[SHEET_NAME]
    all_vals = _all_cell_values(ws)

    checks = [
        ("테스트 사업장 (주)",    "site_name"),
        ("○○ 공장 신축공사 현장", "project_name"),
        ("굴착기 (백호우, 불도저)", "machine_type"),
        ("굴착기 작업반경 7m, 출입 금지구역 표시", "work_radius 포함 텍스트"),
        ("배치 필요 — 유도자 1명 상시 배치", "guide_worker_required 포함 텍스트"),
        ("진입: 현장 정문(동측)", "travel_route_text 포함 텍스트"),
        ("로프 바리케이드 설치", "access_control 포함 텍스트"),
        ("안전관리자: 010-2222-3333", "emergency_contact 포함 텍스트"),
        ("2026-04-25",            "sign_date"),
    ]

    for expected_part, field_label in checks:
        found = any(expected_part in v for v in all_vals)
        results.append(_check(found, f"값 렌더링 확인: {field_label}", repr(expected_part)))

    # hazard_items 첫 번째 항목 확인
    results.append(_check(
        any("굴착기 후진 중 근로자 충돌" in v for v in all_vals),
        "hazard_items[0].hazard 렌더링",
    ))
    results.append(_check(
        any("유도자 배치 및 후방 경보" in v for v in all_vals),
        "hazard_items[0].safety_measure 렌더링",
    ))

    # pre_check_items 첫 번째 항목 확인
    results.append(_check(
        any("제동장치·조향장치 이상 유무" in v for v in all_vals),
        "pre_check_items[0].check_item 렌더링",
    ))

    return results


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def main() -> int:
    print("차량계 건설기계 작업계획서 Excel Builder 검증 시작")
    print("=" * 60)

    all_results: list[bool] = []

    # 샘플 데이터 검증
    xlsx_sample = build_vehicle_workplan_excel(SAMPLE_FORM_DATA)
    all_results += validate_full(xlsx_sample, "샘플 데이터 검증")
    all_results += validate_sample_values(xlsx_sample)

    # 빈 form_data 검증
    xlsx_empty = build_vehicle_workplan_excel(SAMPLE_EMPTY_DATA)
    all_results += validate_full(xlsx_empty, "빈 form_data 검증")
    all_results += validate_empty(xlsx_empty)

    total  = len(all_results)
    passed = sum(all_results)
    failed = total - passed

    print("\n" + "=" * 60)
    print(f"결과: {passed}/{total} PASS, {failed} FAIL")
    if failed == 0:
        print("최종 판정: PASS")
        return 0
    else:
        print("최종 판정: FAIL")
        return 1


if __name__ == "__main__":
    sys.exit(main())
