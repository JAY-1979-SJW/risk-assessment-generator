"""
Excel builder 구조 검증 스크립트.

registry에 등록된 모든 form_type에 대해 아래 항목을 검증한다.

검증 항목:
  F1  빈 dict {}로 xlsx 생성 — 크래시 없음
  F2  required_fields만 채운 minimal 입력으로 xlsx 생성
  F3  openpyxl.load_workbook으로 재오픈 가능
  F4  sheet title ≤ 31자
  F5  sheet title에 금지 문자 없음 (: ? * [ ] / \\)
  F6  A1 셀 값이 비어있지 않음 (제목 있음)
  F7  열 너비: 1열 이상에 너비가 명시적으로 설정됨 (기본값 8.0만이 아님)
  F8  행 높이: 1행 이상에 높이가 명시적으로 설정됨 (기본값만이 아님)
  F9  병합 셀 범위 유효성 (min_col ≤ max_col, min_row ≤ max_row)
  F10 병합 셀 내 비앵커 셀에 값 없음 (openpyxl 규칙)
  F11 전체 셀 중 None 값 비율 ≤ 95% (대부분 셀이 빈 문자열로 처리됨)
  F12 행 높이 이상값 없음 (0 < height ≤ 200)
  F13 열 너비 이상값 없음 (0 < width ≤ 100)
  F14 xlsx 파일 크기 > 3,000 bytes (내용 있는 파일)
  F15 xlsx 파일 크기 < 500,000 bytes (비정상 팽창 없음)

실행:
  python scripts/validate_excel_builders.py
  python scripts/validate_excel_builders.py --form_type piling_workplan
  python scripts/validate_excel_builders.py --verbose
  python scripts/validate_excel_builders.py --fail-only
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import tempfile
import os
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from engine.output.form_registry import build_form_excel, list_supported_forms

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------

SHEET_TITLE_MAX_LEN   = 31
SHEET_TITLE_FORBIDDEN = re.compile(r'[:\?*\[\]/\\]')
ROW_HEIGHT_MAX        = 200.0
COL_WIDTH_MAX         = 100.0
FILE_SIZE_MIN         = 3_000
FILE_SIZE_MAX         = 500_000
NONE_RATIO_MAX        = 0.95   # 전체 셀 중 None 비율 허용 상한

# ---------------------------------------------------------------------------
# 결과 구조체
# ---------------------------------------------------------------------------

@dataclass
class CheckResult:
    code: str
    ok: bool
    detail: str = ""

@dataclass
class FormResult:
    form_type: str
    display_name: str
    checks: list[CheckResult] = field(default_factory=list)

    @property
    def passed(self) -> int:
        return sum(1 for c in self.checks if c.ok)

    @property
    def failed(self) -> int:
        return sum(1 for c in self.checks if not c.ok)

    @property
    def verdict(self) -> str:
        if self.failed == 0:
            return "PASS"
        return "FAIL"


# ---------------------------------------------------------------------------
# 검증 함수
# ---------------------------------------------------------------------------

def _make_minimal_data(spec: dict[str, Any]) -> dict[str, Any]:
    """required_fields를 최소값으로 채운 dict."""
    data: dict[str, Any] = {}
    for f in spec.get("required_fields") or []:
        data[f] = "테스트"
    return data


def _check_worksheet(ws: Worksheet, xlsx_bytes: bytes) -> list[CheckResult]:
    results: list[CheckResult] = []

    def ok(code: str, detail: str = "") -> CheckResult:
        return CheckResult(code, True, detail)

    def fail(code: str, detail: str) -> CheckResult:
        return CheckResult(code, False, detail)

    # F4 sheet title 길이
    title = ws.title or ""
    if len(title) <= SHEET_TITLE_MAX_LEN:
        results.append(ok("F4", f"title={title!r} ({len(title)}자)"))
    else:
        results.append(fail("F4", f"title={title!r} — {len(title)}자 > {SHEET_TITLE_MAX_LEN}"))

    # F5 sheet title 금지 문자
    bad = SHEET_TITLE_FORBIDDEN.findall(title)
    if not bad:
        results.append(ok("F5"))
    else:
        results.append(fail("F5", f"금지 문자: {bad}"))

    # F6 A1 값 있음
    a1_val = ws.cell(row=1, column=1).value
    if a1_val and str(a1_val).strip():
        results.append(ok("F6", f"A1={str(a1_val)[:30]!r}"))
    else:
        results.append(fail("F6", f"A1 비어있음: {a1_val!r}"))

    # F7 열 너비 설정 여부
    non_default_cols = [
        col for col, cd in ws.column_dimensions.items()
        if cd.width is not None and cd.width != 8.0
    ]
    if non_default_cols:
        results.append(ok("F7", f"{len(non_default_cols)}개 열 너비 설정됨"))
    else:
        results.append(fail("F7", "모든 열 너비가 기본값(8.0) — apply_col_widths 미적용 의심"))

    # F8 행 높이 설정 여부
    non_default_rows = [
        r for r, rd in ws.row_dimensions.items()
        if rd.height is not None and rd.height > 0
    ]
    if non_default_rows:
        results.append(ok("F8", f"{len(non_default_rows)}개 행 높이 설정됨"))
    else:
        results.append(fail("F8", "모든 행 높이가 기본값 — write_cell height 미적용 의심"))

    # F9 병합 셀 범위 유효성
    invalid_merges = []
    for mc in ws.merged_cells.ranges:
        if mc.min_col > mc.max_col or mc.min_row > mc.max_row:
            invalid_merges.append(str(mc))
    if not invalid_merges:
        results.append(ok("F9", f"병합 셀 {len(ws.merged_cells.ranges)}개 모두 유효"))
    else:
        results.append(fail("F9", f"비정상 병합: {invalid_merges}"))

    # F10 병합 셀 비앵커에 값 없음
    bad_merge_cells = []
    for mc in ws.merged_cells.ranges:
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                if r == mc.min_row and c == mc.min_col:
                    continue  # 앵커 셀
                val = ws.cell(row=r, column=c).value
                if val not in (None, ""):
                    from openpyxl.utils import get_column_letter
                    bad_merge_cells.append(f"{get_column_letter(c)}{r}={val!r}")
    if not bad_merge_cells:
        results.append(ok("F10"))
    else:
        results.append(fail("F10", f"비앵커 셀에 값: {bad_merge_cells[:5]}"))

    # F11 None 값 비율
    total_cells = 0
    none_cells  = 0
    for row in ws.iter_rows():
        for cell in row:
            total_cells += 1
            if cell.value is None:
                none_cells += 1
    ratio = none_cells / total_cells if total_cells else 0
    if ratio <= NONE_RATIO_MAX:
        results.append(ok("F11", f"None 비율={ratio:.1%} ({none_cells}/{total_cells})"))
    else:
        results.append(fail("F11", f"None 비율={ratio:.1%} — 대부분 셀이 None (빈 문자열 처리 누락)"))

    # F12 행 높이 이상값
    bad_heights = [
        (r, rd.height) for r, rd in ws.row_dimensions.items()
        if rd.height is not None and not (0 < rd.height <= ROW_HEIGHT_MAX)
    ]
    if not bad_heights:
        results.append(ok("F12"))
    else:
        results.append(fail("F12", f"이상 행 높이: {bad_heights[:5]}"))

    # F13 열 너비 이상값
    bad_widths = [
        (col, cd.width) for col, cd in ws.column_dimensions.items()
        if cd.width is not None and not (0 < cd.width <= COL_WIDTH_MAX)
    ]
    if not bad_widths:
        results.append(ok("F13"))
    else:
        results.append(fail("F13", f"이상 열 너비: {bad_widths[:5]}"))

    # F14/F15 파일 크기
    sz = len(xlsx_bytes)
    if sz >= FILE_SIZE_MIN:
        results.append(ok("F14", f"{sz:,} bytes"))
    else:
        results.append(fail("F14", f"파일 크기 너무 작음: {sz:,} bytes < {FILE_SIZE_MIN:,}"))

    if sz <= FILE_SIZE_MAX:
        results.append(ok("F15", f"{sz:,} bytes"))
    else:
        results.append(fail("F15", f"파일 크기 비정상: {sz:,} bytes > {FILE_SIZE_MAX:,}"))

    return results


def validate_form(spec: dict[str, Any]) -> FormResult:
    form_type    = spec["form_type"]
    display_name = spec.get("display_name", form_type)
    result = FormResult(form_type=form_type, display_name=display_name)

    def ok(code: str, detail: str = "") -> None:
        result.checks.append(CheckResult(code, True, detail))

    def fail(code: str, detail: str) -> None:
        result.checks.append(CheckResult(code, False, detail))

    # F1 빈 dict 크래시 없음
    try:
        xlsx_empty = build_form_excel(form_type, {})
        ok("F1", f"{len(xlsx_empty):,} bytes")
    except Exception as e:
        fail("F1", f"{type(e).__name__}: {e}")
        return result  # 이후 검증 불가

    # F2 minimal 입력으로 생성
    minimal = _make_minimal_data(spec)
    try:
        xlsx_min = build_form_excel(form_type, minimal)
        ok("F2", f"{len(xlsx_min):,} bytes, {len(minimal)} 필드")
    except Exception as e:
        fail("F2", f"{type(e).__name__}: {e}")
        xlsx_min = xlsx_empty  # fallback

    # F3 load_workbook
    try:
        wb = load_workbook(io.BytesIO(xlsx_min))
        ws = wb.active
        ok("F3", f"sheet={ws.title!r}")
    except Exception as e:
        fail("F3", f"{type(e).__name__}: {e}")
        return result

    # F4~F15 worksheet 구조 검증
    result.checks.extend(_check_worksheet(ws, xlsx_min))

    return result


# ---------------------------------------------------------------------------
# 출력
# ---------------------------------------------------------------------------

def print_result(r: FormResult, verbose: bool) -> None:
    verdict_sym = "✅" if r.verdict == "PASS" else "❌"
    print(f"\n{verdict_sym} [{r.verdict}] {r.form_type}  ({r.display_name})")
    print(f"         PASS {r.passed}/{len(r.checks)}  FAIL {r.failed}/{len(r.checks)}")
    for c in r.checks:
        if not c.ok:
            print(f"    ❌ {c.code}  {c.detail}")
        elif verbose:
            print(f"    ✓  {c.code}  {c.detail}")


def print_summary(results: list[FormResult]) -> None:
    total   = len(results)
    passed  = sum(1 for r in results if r.verdict == "PASS")
    failed  = total - passed
    total_checks  = sum(len(r.checks) for r in results)
    failed_checks = sum(r.failed for r in results)

    print("\n" + "=" * 70)
    print(f"  서식 전체: {total}종  |  PASS: {passed}  |  FAIL: {failed}")
    print(f"  검증 항목: {total_checks}건  |  통과: {total_checks - failed_checks}  |  실패: {failed_checks}")
    print("=" * 70)
    if failed > 0:
        print("\n  실패 서식 목록:")
        for r in results:
            if r.verdict == "FAIL":
                fail_codes = [c.code for c in r.checks if not c.ok]
                print(f"    ❌ {r.form_type:45} {fail_codes}")
    verdict = "PASS" if failed == 0 else "FAIL"
    print(f"\n  최종 판정: {verdict}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# 진입점
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Excel builder 구조 검증")
    parser.add_argument("--form_type", help="특정 form_type만 검증")
    parser.add_argument("--verbose",   action="store_true", help="PASS 항목도 출력")
    parser.add_argument("--fail-only", action="store_true", dest="fail_only",
                        help="FAIL 서식만 출력")
    args = parser.parse_args()

    all_specs = list_supported_forms()
    if args.form_type:
        specs = [s for s in all_specs if s["form_type"] == args.form_type]
        if not specs:
            print(f"[ERROR] form_type 미등록: {args.form_type!r}")
            print(f"  등록된 form_type: {[s['form_type'] for s in all_specs]}")
            return 1
    else:
        specs = all_specs

    print(f"Excel builder 구조 검증 — {len(specs)}종")
    print("검증 항목: F1(empty) F2(minimal) F3(load) F4(title길이) F5(title문자)")
    print("           F6(A1값) F7(열너비) F8(행높이) F9(병합유효) F10(병합값)")
    print("           F11(None비율) F12(행높이범위) F13(열너비범위) F14(최소크기) F15(최대크기)")

    results: list[FormResult] = []
    for spec in specs:
        r = validate_form(spec)
        results.append(r)
        if not args.fail_only or r.verdict == "FAIL":
            print_result(r, verbose=args.verbose)

    print_summary(results)
    return 0 if all(r.verdict == "PASS" for r in results) else 1


if __name__ == "__main__":
    sys.exit(main())
