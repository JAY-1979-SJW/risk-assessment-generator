"""
교육일지 Excel builder 검증 스크립트.

검증 항목:
  1. xlsx bytes 정상 생성 여부
  2. 시트명 == "교육일지"
  3. 제목 셀(A1) == "안전보건교육일지"
  4. 필수 헤더 존재 (교육 내용 표, 수강자 명단 표)
  5. 수강자 30행 출력 확인
  6. 사용자 미입력 필드 → 공란 유지 확인
  7. 서명란(attendee_signature) 항상 공란 확인
  8. openpyxl 로 재오픈 성공 여부

실행:
  python scripts/validate_education_log_builder.py
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from engine.output.education_log_builder import (
    MAX_ATTENDEES,
    SHEET_NAME,
    SHEET_HEADING,
    _SHEET_SUBTITLE,
    build_education_log_excel,
)

# ---------------------------------------------------------------------------
# 샘플 form_data (검증용)
# ---------------------------------------------------------------------------
SAMPLE_FORM_DATA = {
    "site_name": "테스트 사업장",
    "site_address": "서울시 강남구 테헤란로 123",
    "education_type": "정기교육",
    "education_date": "2026-04-24 09:00~11:00",
    "education_location": "본사 2층 교육실",
    "education_duration_hours": "2",
    "education_target_job": "생산직 근로자 전원",
    "instructor_name": "홍길동",
    "instructor_qualification": "산업안전지도사",
    "subjects": [
        {
            "subject_name": "산업안전 및 사고 예방",
            "subject_content": "추락·낙하·협착 위험 및 예방 조치",
            "subject_hours": "1",
        },
        {
            "subject_name": "산업보건 및 직업병 예방",
            "subject_content": "근골격계 질환 예방, 유해인자 관리",
            "subject_hours": "1",
        },
    ],
    "attendees": [
        {"attendee_name": "김철수", "attendee_job_type": "용접공"},
        {"attendee_name": "이영희", "attendee_job_type": "생산직"},
        {"attendee_name": "박민준", "attendee_job_type": "설비보전"},
    ],
    "confirmer_name": "안전관리자 이담당",
    "confirmer_role": "안전보건관리책임자",
    "confirm_date": "2026-04-24",
}

SAMPLE_EMPTY_DATA: dict = {}  # 모든 필드 공란 테스트용


# ---------------------------------------------------------------------------
# 검증 함수
# ---------------------------------------------------------------------------

def _check(condition: bool, name: str, detail: str = "") -> bool:
    status = "PASS" if condition else "FAIL"
    msg = f"  [{status}] {name}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    return condition


def validate_full(xlsx_bytes: bytes, label: str) -> list[bool]:
    """xlsx bytes에 대한 전체 검증. 결과 bool 목록 반환."""
    results = []
    print(f"\n=== {label} ===")

    # 1. bytes 생성 여부
    results.append(_check(isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
                          "xlsx bytes 정상 생성",
                          f"{len(xlsx_bytes):,} bytes"))

    # 2. openpyxl 재오픈
    try:
        wb = load_workbook(BytesIO(xlsx_bytes))
        results.append(_check(True, "openpyxl 재오픈 성공"))
    except Exception as e:
        results.append(_check(False, "openpyxl 재오픈", str(e)))
        return results

    # 3. 시트명
    results.append(_check(SHEET_NAME in wb.sheetnames,
                          f"시트명 == '{SHEET_NAME}'",
                          str(wb.sheetnames)))
    ws = wb[SHEET_NAME]

    # 4. 제목 셀
    title_val = ws.cell(row=1, column=1).value
    results.append(_check(title_val == SHEET_HEADING,
                          f"제목 셀(A1) == '{SHEET_HEADING}'",
                          repr(title_val)))

    # 5. 필수 헤더 텍스트 존재 (셀 스캔)
    all_values = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                all_values.add(str(v))

    required_texts = ["교육 내용", "수강자 명단", "교육 과목명", "교육 내용 요약",
                      "성명", "직종(직위)", "서명 또는 날인"]
    for text in required_texts:
        results.append(_check(text in all_values,
                              f"필수 헤더 존재: '{text}'"))

    # 6. 수강자 30행 확인: "번호" 열(A)에서 숫자 값 30까지 존재 여부
    attendee_numbers = []
    for row in ws.iter_rows(values_only=True):
        val = row[0]  # 컬럼 A
        if isinstance(val, int) and 1 <= val <= MAX_ATTENDEES:
            attendee_numbers.append(val)
    has_30 = (MAX_ATTENDEES in attendee_numbers)
    results.append(_check(has_30,
                          f"수강자 {MAX_ATTENDEES}행 존재",
                          f"발견된 번호: {sorted(attendee_numbers)[:5]}..."))

    return results


def validate_empty(xlsx_bytes: bytes) -> list[bool]:
    """빈 form_data로 생성한 xlsx: 모든 값 필드가 공란인지 확인."""
    results = []
    print("\n=== 공란 유지 검증 (빈 form_data) ===")

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb[SHEET_NAME]

    # 제목·헤더 외 값 셀이 모두 빈 문자열 또는 정수(순번)인지 확인
    non_blank_non_header = []
    header_texts = {
        SHEET_HEADING, _SHEET_SUBTITLE, "교육 내용", "수강자 명단",
        "사업장명", "사업장 소재지", "교육 종류", "교육 장소",
        "교육 일시", "교육 시간", "교육 대상", "강사명", "강사 자격",
        "순번", "교육 과목명", "교육 내용 요약", "시간(h)",
        "번호", "성명", "직종(직위)", "서명 또는 날인",
        "확인자 성명", "확인자 직위", "서명", "확인 일자",
    }
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is None or v == "":
                continue
            sv = str(v)
            if sv in header_texts:
                continue
            # 순번(정수)는 허용
            if isinstance(v, int):
                continue
            non_blank_non_header.append(sv)

    results.append(_check(len(non_blank_non_header) == 0,
                          "모든 값 필드 공란 유지",
                          f"비공란 비헤더 값: {non_blank_non_header[:5] if non_blank_non_header else '없음'}"))
    return results


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def main() -> int:
    print("교육일지 Excel Builder 검증 시작")
    print("=" * 50)

    all_results: list[bool] = []

    # 샘플 데이터 검증
    xlsx_full = build_education_log_excel(SAMPLE_FORM_DATA)
    all_results += validate_full(xlsx_full, "샘플 데이터 검증")

    # 빈 데이터 검증
    xlsx_empty = build_education_log_excel(SAMPLE_EMPTY_DATA)
    all_results += validate_full(xlsx_empty, "빈 form_data 검증")
    all_results += validate_empty(xlsx_empty)

    # 최종 판정
    total = len(all_results)
    passed = sum(all_results)
    failed = total - passed

    print("\n" + "=" * 50)
    print(f"결과: {passed}/{total} PASS, {failed} FAIL")
    if failed == 0:
        print("최종 판정: PASS")
        return 0
    else:
        print("최종 판정: FAIL")
        return 1


if __name__ == "__main__":
    sys.exit(main())
