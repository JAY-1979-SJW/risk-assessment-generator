"""
P1 구현 + evidence 연결 완료 — ED-003 / WP-015 smoke test.

검증 항목:
  1. registry: special_education_log 등록됨
  2. get_form_spec() 검증 (display_name, required_fields, repeat_field, max_repeat_rows)
  3. 최소 샘플 입력으로 Excel bytes 생성 확인
  4. 공란 form_data로 bytes 생성 (오류 없음)
  5. 전체 샘플 bytes 생성
  6. 필수 섹션 제목 포함 확인:
     - 특별 안전보건교육 교육일지
     - 교육 기본정보
     - 교육시간 구분
     - 특별교육 대상 작업
     - 교육내용
     - 교육대상자 명단
     - 확인 서명
  7. 개인정보 과다 필드 없음: 주민등록번호, 건강정보, 질병명
  8. 별표 4/별표 5 기준 문구 포함 (VERIFIED 상태)
  9. ED-003 catalog DONE 확인
 10. ED-003 evidence_status == VERIFIED (PARTIAL_VERIFIED도 허용)
 11. evidence_id 4개 연결 확인 (ED-003-L1 ~ L4)
 12. evidence_file 4개 실제 파일 존재 확인
 13. 각 evidence_file verification_result 확인 (VERIFIED 또는 PARTIAL_VERIFIED)
 14. 별표 4 교육시간 구조화 정보 존재 확인
 15. 별표 5 대상 작업 번호 범위 (제1호~제39호) 확인

실행:
    cd <project_root>
    python scripts/smoke_test_p1_forms.py

판정:
    PASS — 모든 항목 통과
    WARN — bytes 생성됐으나 경고 발생
    FAIL — 예외 발생 또는 필수 항목 미충족
"""
from __future__ import annotations

import json
import sys
import traceback
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, ".")

from engine.output.form_registry import build_form_excel, get_form_spec, list_supported_forms

ROOT         = Path(".")
EVIDENCE_DIR = ROOT / "data" / "evidence" / "safety_law_refs"

# ---------------------------------------------------------------------------
# ED-003 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_ED003_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "site_address": "서울특별시 중구 세종대로 110",
    "education_name": "밀폐공간 작업 특별 안전보건교육",
    "education_date": "2026-04-25",
    "education_location": "현장 교육장",
    "education_target_work": "밀폐공간 작업",
    "related_education": "신규채용 시 교육 병행 실시",
    "duration_category": "16시간 이상 (최초 배치 전)",
    "actual_duration_hours": "16",
    "remaining_hours": "0",
    "subjects": [
        {
            "subject_name": "밀폐공간 작업의 위험성",
            "subject_content": "산소결핍·유해가스 발생 원인 및 사례",
            "subject_hours": "4",
        },
        {
            "subject_name": "가스 측정 및 환기",
            "subject_content": "가스 측정기 사용법, 강제 환기 설비 운영",
            "subject_hours": "4",
        },
        {
            "subject_name": "비상대피 및 구조",
            "subject_content": "비상신호, 대피 절차, 공기호흡기 착용법",
            "subject_hours": "4",
        },
        {
            "subject_name": "실습: 가스 측정·환기·비상대피",
            "subject_content": "현장 실습 (집수정 모의 훈련)",
            "subject_hours": "4",
        },
    ],
    "instructor_name": "홍길동",
    "instructor_org": "현장안전팀",
    "instructor_role": "안전관리자",
    "instructor_qualification": "산업안전지도사 (제2020-00001호)",
    "attendees": [
        {
            "attendee_name": "이순신",
            "attendee_org": "테스트건설(주)",
            "attendee_job_type": "배관공",
            "attendee_birth_year": "1985",
            "attendee_completed": "완료",
        },
        {
            "attendee_name": "강감찬",
            "attendee_org": "협력업체 A",
            "attendee_job_type": "감시인",
            "attendee_birth_year": "1990",
            "attendee_completed": "완료",
        },
    ],
    "comprehension_verbal": "구두 질의응답 완료 (전원 이해 확인)",
    "comprehension_checklist": "10문항 체크리스트 실시 — 전원 8점 이상",
    "comprehension_practice": "가스측정기 사용 실습 완료",
    "retraining_targets": "없음",
    "attachments": "교육자료(PPT) / 현장 실습 사진 3장 / 참석자 서명부",
    "confirmer_name": "김안전",
    "confirmer_role": "안전보건관리책임자",
    "supervisor_name": "박감독",
    "site_manager_name": "최소장",
    "confirm_date": "2026-04-25",
}

SAMPLE_ED003_MINIMAL: dict = {
    "education_date": "2026-04-25",
    "education_location": "현장 교육장",
    "education_target_work": "밀폐공간 작업",
    "instructor_name": "홍길동",
    "confirmer_name": "김안전",
    "confirmer_role": "안전보건관리책임자",
}

# ---------------------------------------------------------------------------
# 필수 섹션 제목
# ---------------------------------------------------------------------------

REQUIRED_HEADINGS = [
    "특별 안전보건교육 교육일지",
    "교육 기본정보",
    "교육시간 구분",
    "특별교육 대상 작업",
    "교육내용",
    "교육대상자 명단",
    "확인 서명",
]

# 워크북에 반드시 포함되어야 하는 별표 관련 문구
REQUIRED_BYEOLPYO_KEYWORDS = ["별표 4", "별표 5"]

# 기본 필드명으로 출력되면 안 되는 개인정보 키워드
FORBIDDEN_PII_KEYWORDS = [
    "주민등록번호",
    "건강정보",
    "질병명",
]

# ED-003 evidence ID 목록
EXPECTED_EVIDENCE_IDS = ["ED-003-L1", "ED-003-L2", "ED-003-L3", "ED-003-L4"]

# evidence 파일명
EXPECTED_EVIDENCE_FILES = [
    "ED-003-L1_industrial_safety_health_act_article_29.json",
    "ED-003-L2_industrial_safety_health_rule_article_26.json",
    "ED-003-L3_industrial_safety_health_rule_attached_table_4.json",
    "ED-003-L4_industrial_safety_health_rule_attached_table_5.json",
]

# catalog evidence_status 허용값
VALID_EVIDENCE_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _all_cell_values(wb) -> list[str]:
    return [
        str(cell.value or "")
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    ]


def _check(ok: bool, name: str, detail: str = "") -> tuple[str, str, str]:
    verdict = "PASS" if ok else "FAIL"
    return (verdict, name, detail)


# ---------------------------------------------------------------------------
# 검증 실행
# ---------------------------------------------------------------------------
# (run_smoke_test 정의는 WP-015 추가 후 파일 하단에 통합됨)
# ---------------------------------------------------------------------------

def _run_ed003_checks_legacy(results, supported) -> None:
    """ED-003 검증 (run_smoke_test에 통합됨, 직접 호출 금지)."""
    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "special_education_log" in supported,
        "registry: special_education_log 등록됨",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    try:
        spec = get_form_spec("special_education_log")
        results.append(_check(isinstance(spec, dict),
                               "get_form_spec() → dict"))
        results.append(_check(spec.get("display_name") == "특별 안전보건교육 교육일지",
                               "display_name 확인",
                               repr(spec.get("display_name"))))
        results.append(_check(isinstance(spec.get("required_fields"), list)
                               and len(spec["required_fields"]) > 0,
                               "required_fields 비어있지 않음"))
        results.append(_check(spec.get("repeat_field") == "attendees",
                               "repeat_field == 'attendees'"))
        results.append(_check(spec.get("max_repeat_rows") == 30,
                               "max_repeat_rows == 30"))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("special_education_log", SAMPLE_ED003_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("special_education_log", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~8. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("special_education_log", SAMPLE_ED003_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = _all_cell_values(wb_full)
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 개인정보 과다 필드 없음
        for keyword in FORBIDDEN_PII_KEYWORDS:
            found = keyword in all_text
            results.append(_check(
                not found,
                f"개인정보 과다 필드 없음: '{keyword}'",
                "FAIL — 기본 필드명에 해당 문구 발견" if found else "",
            ))

        # 별표 4/5 기준 문구 포함 (VERIFIED 상태)
        for kw in REQUIRED_BYEOLPYO_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"별표 기준 문구 포함: '{kw}'",
            ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 9~15. ED-003 catalog + evidence 검증 ────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        ed003 = next(
            (d for d in cat["documents"] if d["id"] == "ED-003"), None
        )

        # 9. catalog 존재
        results.append(_check(ed003 is not None, "catalog: ED-003 항목 존재"))

        if ed003:
            # 10. implementation_status == DONE
            results.append(_check(
                ed003.get("implementation_status") == "DONE",
                "catalog: ED-003 implementation_status == DONE",
                repr(ed003.get("implementation_status")),
            ))
            # form_type 확인
            results.append(_check(
                ed003.get("form_type") == "special_education_log",
                "catalog: ED-003 form_type == 'special_education_log'",
                repr(ed003.get("form_type")),
            ))
            # 11. evidence_status VERIFIED 또는 PARTIAL_VERIFIED
            ev_status = ed003.get("evidence_status", "")
            results.append(_check(
                ev_status in VALID_EVIDENCE_STATUSES,
                f"catalog: ED-003 evidence_status in {VALID_EVIDENCE_STATUSES}",
                repr(ev_status),
            ))

            # 12. evidence_id 4개 포함
            ev_ids = ed003.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # 13. evidence_file 실제 존재 확인
            ev_files = ed003.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            for efname in EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # 14~15. 각 evidence 파일 내용 검증
        for efname in EXPECTED_EVIDENCE_FILES:
            fpath = EVIDENCE_DIR / efname
            if not fpath.exists():
                continue
            try:
                ev_data = json.loads(fpath.read_text(encoding="utf-8"))
                vr = ev_data.get("verification_result", "")
                results.append(_check(
                    vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                    f"evidence 검증결과: {efname[:30]}... → {vr}",
                ))
                # 별표 4: special_education_hours 구조 확인
                if "table_4" in efname:
                    hours_data = ev_data.get("special_education_hours", {})
                    results.append(_check(
                        bool(hours_data.get("일반근로자")),
                        "별표 4: 일반근로자 특별교육 시간 구조화",
                        repr(hours_data.get("일반근로자", ""))[:60],
                    ))
                # 별표 5: range_phrase 확인
                if "table_5" in efname:
                    works_data = ev_data.get("special_education_target_works", {})
                    results.append(_check(
                        "제1호부터 제39호" in works_data.get("range_phrase", ""),
                        "별표 5: 대상 작업 범위 '제1호부터 제39호까지' 확인",
                        repr(works_data.get("range_phrase", ""))[:60],
                    ))
            except Exception as exc:
                results.append(_check(False, f"evidence 파일 로딩: {efname[:30]}...", str(exc)))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "catalog/evidence 검증 중 예외", tb))

    # ── 출력 ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  KRAS P1 Smoke Test — ED-003 특별 안전보건교육 교육일지")
    print("=" * 62)

    pass_cnt = warn_cnt = fail_cnt = 0
    for verdict, name, detail in results:
        icon = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌"}[verdict]
        line = f"  {icon} [{verdict}] {name}"
        if detail:
            line += f" — {detail}"
        print(line)
        if verdict == "PASS":   pass_cnt += 1
        elif verdict == "WARN": warn_cnt += 1
        else:                   fail_cnt += 1; overall = "FAIL"

    print("-" * 62)
    total = len(results)
    print(f"  합계: PASS {pass_cnt}/{total}  WARN {warn_cnt}  FAIL {fail_cnt}")
    print(f"\n  최종 판정: {overall}")
    print("=" * 62 + "\n")

    sys.exit(0 if overall != "FAIL" else 1)


# ---------------------------------------------------------------------------
# WP-011 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_WP011_MINIMAL: dict = {
    "work_location": "지하 1층 전기실 수배전반",
    "work_supervisor": "홍길동",
}

SAMPLE_WP011_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "work_date": "2026-04-25",
    "work_period": "2026-04-25 ~ 2026-04-26",
    "contractor": "협력업체 전기",
    "prepared_by": "홍길동",
    "reviewer": "김감독",
    "work_name": "수배전반 분기차단기 교체 작업",
    "work_location": "지하 1층 전기실 수배전반",
    "work_datetime": "2026-04-25 09:00 ~ 18:00",
    "voltage": "22.9kV / 저압 380V",
    "work_category": "정전작업 후 활선 근접",
    "work_supervisor": "홍길동",
    "type_outage": "○",
    "type_live": "",
    "type_near": "○",
    "type_temp_elec": "",
    "type_panel": "○",
    "type_cable": "",
    "type_power_tool": "",
    "type_test_measure": "○",
    "hazard_electric_shock": "○",
    "hazard_arc": "○",
    "hazard_short_circuit": "○",
    "hazard_leakage": "○",
    "hazard_fire": "○",
    "hazard_explosion": "",
    "hazard_fall": "",
    "hazard_pinch": "",
    "prereq_ra001": "확인 완료",
    "prereq_ra004": "확인 완료",
    "prereq_ptw004": "발행 완료",
    "prereq_cl004": "확인 완료",
    "prereq_ppe001": "확인 완료",
    "loto_scope": "MCC 판넬 3번 회로 전체",
    "loto_breaker_location": "지하1층 전기실 주차단기 CB-3",
    "loto_lock_installed": "완료",
    "loto_sign_attached": "완료",
    "loto_residual_voltage": "검전기 확인 — 0V 확인",
    "loto_re_energize": "LOTO 잠금 유지, 작업완료 후 작업책임자 해제",
    "live_approach_limit": "300mm 이상 이격",
    "live_insulation_ppe": "절연장갑·절연화·보안면 착용",
    "live_insulation_tools": "절연공구 전 수량 지참",
    "live_monitor": "감시인 1명 상주",
    "live_energized_protect": "절연 커버 설치",
    "temp_elcb": "해당 없음 (정전작업)",
    "temp_grounding": "접지선 연결 확인",
    "temp_wire_protect": "해당 없음",
    "temp_waterproof": "해당 없음",
    "temp_overload": "해당 없음",
    "temp_panel_lock": "분전반 잠금 완료",
    "tool_body_damage": "이상 없음",
    "tool_wire_insulation": "이상 없음",
    "tool_plug": "이상 없음",
    "tool_ground_wire": "연결 확인",
    "tool_elcb": "작동 확인",
    "ppe_insulated_gloves": "○",
    "ppe_insulated_shoes": "○",
    "ppe_face_shield": "○",
    "ppe_insulation_mat": "○",
    "equip_voltage_tester": "○",
    "equip_insulation_meter": "○",
    "mgmt_zone_control": "작업구역 접근금지 로프 및 표지판 설치",
    "mgmt_monitor": "감시인 1명 (이감시 — 010-0000-0000)",
    "mgmt_emergency_stop": "이상 발생 시 즉시 작업중지, 전원 차단",
    "mgmt_fire_response": "소화기 1대 비치, 화재 시 119 신고 후 대피",
    "mgmt_reenergize_proc": "작업완료 → LOTO 해제 → 책임자 확인 → 재투입",
    "nonconformance_items": [],
    "work_verdict": "적합",
    "verdict_condition": "",
    "sign_date": "2026-04-25",
}

# WP-011 필수 섹션 제목
WP011_REQUIRED_HEADINGS = [
    "전기 작업계획서",
    "작업 기본정보",
    "전기작업 유형",
    "전기 위험요인 확인",
    "작업 전 선행서류 확인",
    "정전 및 LOTO 계획",
    "활선·근접작업 안전조치",
    "임시전기 및 분전반 안전조치",
    "전동공구 및 이동식 전기기계기구 점검",
    "보호구 및 측정장비 확인",
    "작업 중 관리계획",
    "부적합 사항 및 시정조치",
    "작업 가능 여부 판정",
    "확인 서명",
]

# WP-011 필수 키워드
WP011_REQUIRED_KEYWORDS = [
    "정전 및 LOTO",
    "활선",
    "누전차단기",
    "접지",
    "작업중지",
]

# WP-011 evidence
WP011_EXPECTED_EVIDENCE_IDS = ["WP-011-L1"]
WP011_EXPECTED_EVIDENCE_FILES = [
    "WP-011-L1_safety_rule_electrical_work.json",
]
WP011_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_wp011_smoke_test() -> list[tuple[str, str, str]]:
    """WP-011 전기 작업계획서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "electrical_workplan" in supported,
        "registry: electrical_workplan 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("electrical_workplan")
        results.append(_check(isinstance(spec, dict),
                               "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "전기 작업계획서",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "nonconformance_items",
            "repeat_field == 'nonconformance_items'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 5,
            "max_repeat_rows == 5",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("electrical_workplan", SAMPLE_WP011_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("electrical_workplan", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~9. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("electrical_workplan", SAMPLE_WP011_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in WP011_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in WP011_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 10~13. WP-011 catalog + evidence 검증 ─────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        wp011 = next(
            (d for d in cat["documents"] if d["id"] == "WP-011"), None
        )

        results.append(_check(wp011 is not None, "catalog: WP-011 항목 존재"))

        if wp011:
            results.append(_check(
                wp011.get("implementation_status") == "DONE",
                "catalog: WP-011 implementation_status == DONE",
                repr(wp011.get("implementation_status")),
            ))
            results.append(_check(
                wp011.get("form_type") == "electrical_workplan",
                "catalog: WP-011 form_type == 'electrical_workplan'",
                repr(wp011.get("form_type")),
            ))
            ev_status = wp011.get("evidence_status", "")
            results.append(_check(
                ev_status in WP011_VALID_EV_STATUSES,
                f"catalog: WP-011 evidence_status in {WP011_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # related_documents 필수 포함 여부
            rel_docs = wp011.get("related_documents") or []
            for req_doc in ["RA-001", "RA-004", "PTW-004", "CL-004"]:
                results.append(_check(
                    req_doc in rel_docs,
                    f"catalog: WP-011 related_documents 포함 — {req_doc}",
                ))

            # evidence_id 포함
            ev_ids = wp011.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in WP011_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in WP011_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "WP-011 catalog/evidence 검증 중 예외", tb))

    # ── 14. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in WP011_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                f"evidence 검증결과: {efname[:40]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:40]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# WP-015 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_WP015_MINIMAL: dict = {
    "structural_review_done": "실시",
    "assembly_drawing_done": "작성 완료",
    "work_location": "3층 슬래브 구간",
}

SAMPLE_WP015_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "work_date": "2026-04-25",
    "work_period": "2026-04-25 ~ 2026-05-10",
    "contractor": "협력업체 A",
    "prepared_by": "홍길동",
    "reviewer": "이순신",
    "approver": "김안전",
    "work_location": "3층 슬래브 구간",
    "work_scope": "3층 슬래브 전면 (240㎡)",
    "formwork_type": "유로폼",
    "shoring_type": "파이프 서포트",
    "floor_section": "3층 (EL+9.0m)",
    "work_phases": "반입 → 운반 → 동바리 설치 → 거푸집 설치 → 타설 전 점검 → 콘크리트 타설 → 양생 → 해체",
    "survey_ground": "콘크리트 슬래브 (하층 완성, 이상 없음)",
    "survey_substructure": "2층 슬래브 및 보 구조물 확인 완료",
    "survey_opening": "계단 개구부 1개소 — 방호조치 필요",
    "survey_material_place": "외부 야적장 (건물 서측)",
    "survey_equipment_route": "타워크레인 양중 후 수레 운반",
    "survey_weather": "강풍(풍속 10m/s 이상), 강우 시 작업 중지",
    "survey_lighting": "조명 설치 완료, 통로 1.5m 이상 확보",
    "structural_review_done": "실시",
    "structural_reviewer": "박구조 (구조기술사)",
    "assembly_drawing_done": "작성 완료",
    "member_spec": "파이프 서포트 Ø60.5×2.3t, 멍에 H-100, 장선 각목 90×90",
    "install_interval": "서포트 간격 900×900mm",
    "joint_method": "커플러 체결, 핀 고정",
    "brace_plan": "수평연결재 3단 이상, 가새 X형 설치",
    "base_plate_plan": "깔판 두께 30mm 이상, 쐐기 고정",
    "structural_doc_attached": "첨부 (별도 보관)",
    "assembly_drawing_attached": "첨부 (별도 보관)",
    "work_sequence": "1.자재반입→ 2.먹매김/위치확인→ 3.동바리설치→ 4.멍에/장선설치→ 5.거푸집설치→ 6.수직·수평도확인→ 7.타설전점검→ 8.콘크리트타설→ 9.양생→ 10.해체전강도확인→ 11.거푸집해체→ 12.동바리해체",
    "hazard_items": [
        {"hazard": "동바리 침하/전도", "safety_measure": "깔판·베이스플레이트 설치, 수직도 확인"},
        {"hazard": "콘크리트 타설 중 붕괴", "safety_measure": "편심하중 방지, 타설 속도 관리"},
        {"hazard": "추락 (개구부)", "safety_measure": "개구부 덮개·안전망 설치"},
        {"hazard": "낙하물", "safety_measure": "낙하물 방지망, 출입통제"},
    ],
    "safety_measures_text": "1)작업구역 출입통제 2)추락방지 안전난간 설치 3)낙하물 방지망 4)동바리 침하방지 깔판 5)타설 순서·속도 관리 6)작업지휘자 배치 7)전원 보호구 착용",
    "work_commander_name": "강감찬",
    "work_commander_org": "테스트건설(주) 안전팀",
    "work_commander_contact": "010-0000-0000",
    "work_commander_duties": "거푸집·동바리 설치 및 해체 전 과정 지휘",
    "work_commander_educated": "완료 (2026-04-24)",
    "education_done": "완료",
    "tbm_done": "완료",
    "sign_date": "2026-04-25",
}

# WP-015 필수 섹션 제목
WP015_REQUIRED_HEADINGS = [
    "거푸집·동바리 작업계획서",
    "작업 개요",
    "사전조사",
    "구조검토 및 조립도",
    "작업 순서 및 방법",
    "주요 위험요인",
    "안전대책",
    "작업지휘자",
    "작업 전 점검표",
    "확인 서명",
]

# WP-015 필수 포함 문구
WP015_REQUIRED_PHRASES = [
    "구조검토서 및 조립도 원본 첨부·보관 필요",
]

# WP-015 evidence ID 목록
WP015_EXPECTED_EVIDENCE_IDS = ["WP-015-L1", "WP-015-L2", "WP-015-L3", "WP-015-L4"]

# WP-015 evidence 파일명 (2026-04-26 정규화 배치로 표준명 적용)
WP015_EXPECTED_EVIDENCE_FILES = [
    "WP-015-L1_safety_rule_article_38.json",
    "WP-015-L2_safety_rule_article_39.json",
    "WP-015-L3_safety_rule_article_331.json",
    "WP-015-L4_safety_rule_articles_328_337.json",
]

# 허용 evidence_status
WP015_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_wp015_smoke_test() -> list[tuple[str, str, str]]:
    """WP-015 거푸집·동바리 작업계획서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "formwork_shoring_workplan" in supported,
        "registry: formwork_shoring_workplan 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("formwork_shoring_workplan")
        results.append(_check(isinstance(spec, dict),
                               "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "거푸집·동바리 작업계획서",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "hazard_items",
            "repeat_field == 'hazard_items'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("formwork_shoring_workplan", SAMPLE_WP015_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("formwork_shoring_workplan", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("formwork_shoring_workplan", SAMPLE_WP015_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = _all_cell_values(wb_full)
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in WP015_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 문구 포함
        for phrase in WP015_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"필수 문구 포함: '{phrase}'",
            ))

        # 개인정보 과다 필드 없음
        for keyword in FORBIDDEN_PII_KEYWORDS:
            found = keyword in all_text
            results.append(_check(
                not found,
                f"개인정보 과다 필드 없음: '{keyword}'",
                "FAIL — 기본 필드명에 해당 문구 발견" if found else "",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8~11. WP-015 catalog + evidence 검증 ─────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        wp015 = next(
            (d for d in cat["documents"] if d["id"] == "WP-015"), None
        )

        results.append(_check(wp015 is not None, "catalog: WP-015 항목 존재"))

        if wp015:
            results.append(_check(
                wp015.get("implementation_status") == "DONE",
                "catalog: WP-015 implementation_status == DONE",
                repr(wp015.get("implementation_status")),
            ))
            results.append(_check(
                wp015.get("form_type") == "formwork_shoring_workplan",
                "catalog: WP-015 form_type == 'formwork_shoring_workplan'",
                repr(wp015.get("form_type")),
            ))
            ev_status = wp015.get("evidence_status", "")
            results.append(_check(
                ev_status in WP015_VALID_EV_STATUSES,
                f"catalog: WP-015 evidence_status in {WP015_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 4개 포함
            ev_ids = wp015.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in WP015_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            ev_files = wp015.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            for efname in WP015_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "WP-015 catalog/evidence 검증 중 예외", tb))

    # ── 12. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in WP015_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                f"evidence 검증결과: {efname[:35]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:35]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# ED-004 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_ED004_MINIMAL: dict = {
    "role_type": "안전관리자",
    "training_category": "신규교육",
}

SAMPLE_ED004_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "field_name": "테스트현장 신축공사",
    "employer_name": "테스트건설(주)",
    "write_date": "2026-04-25",
    "department": "안전보건팀",
    "person_name": "홍길동",
    "person_org": "테스트건설(주)",
    "person_title": "안전관리자",
    "role_type": "안전관리자",
    "appointment_date": "2026-01-01",
    "is_training_target": "대상",
    "training_category": "신규교육",
    "legal_basis_text": "산업안전보건법 제32조, 제17조, 시행규칙 직무교육 조항",
    "new_training_deadline": "선임 후 3개월 이내 (NEEDS_VERIFICATION)",
    "refresher_cycle": "매 2년마다 (NEEDS_VERIFICATION)",
    "doctor_special_case": "해당 없음 (안전관리자)",
    "training_exemption": "없음",
    "training_org": "한국산업안전보건교육원",
    "training_course": "안전관리자 신규 직무교육",
    "training_start_date": "2026-03-10",
    "training_end_date": "2026-03-12",
    "training_hours": "24",
    "completion_no": "2026-SM-00001",
    "certificate_date": "2026-03-12",
    "completion_status": "수료",
    "cert_attached": "첨부",
    "agency_confirm_attached": "첨부",
    "appointment_doc_attached": "첨부",
    "refresher_basis_attached": "해당 없음 (신규교육)",
    "not_completed": "없음",
    "not_completed_reason": "",
    "action_plan": "",
    "scheduled_training_date": "",
    "manager_name": "김안전",
    "writer_name": "홍길동",
    "safety_manager_sign": "홍길동",
    "supervisor_sign": "이감독",
    "site_manager_sign": "최소장",
    "sign_date": "2026-04-25",
}

# ED-004 필수 섹션 제목/문구
ED004_REQUIRED_HEADINGS = [
    "안전보건관리자 직무교육 이수 확인서",
    "직무교육 대상자 정보",
    "역할 구분",
    "신규교육",
    "보수교육",
    "교육 이수 내역",
    "수료증 첨부",
    "확인 및 서명",
]

ED004_REQUIRED_PHRASES = [
    "공식 수료증을 대체하지 않음",
]

ED004_FORBIDDEN_PII = [
    "주민등록번호",
    "건강정보",
    "질병명",
]

ED004_TRAINING_CODES = [
    "EDU_RESP_MANAGER_DUTY",
    "EDU_SAFETY_MANAGER_DUTY",
    "EDU_HEALTH_MANAGER_DUTY",
]

ED004_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}

EVIDENCE_DIR_ED004 = ROOT / "data" / "evidence" / "safety_law_refs"

ED004_EXPECTED_EVIDENCE_FILES = [
    "ED-004-L1_industrial_safety_health_act_article_32.json",
    "ED-004-L2_industrial_safety_health_act_articles_15_17_18.json",
    "ED-004-L3_industrial_safety_health_rule_job_training.json",
]


def run_ed004_smoke_test() -> list[tuple[str, str, str]]:
    """ED-004 안전보건관리자 직무교육 이수 확인서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "manager_job_training_record" in supported,
        "registry: manager_job_training_record 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("manager_job_training_record")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "안전보건관리자 직무교육 이수 확인서",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") is None,
            "repeat_field is None (단일 레코드 서식)",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("manager_job_training_record", SAMPLE_ED004_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("manager_job_training_record", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~8. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("manager_job_training_record", SAMPLE_ED004_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO
        from openpyxl import load_workbook
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 필수 섹션/제목/문구 포함
        for heading in ED004_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 포함: '{heading}'",
            ))

        # 필수 고지 문구
        for phrase in ED004_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"필수 문구 포함: '{phrase}'",
            ))

        # 개인정보 과다 필드 없음
        for keyword in ED004_FORBIDDEN_PII:
            found = keyword in all_text
            results.append(_check(
                not found,
                f"개인정보 과다 필드 없음: '{keyword}'",
                "FAIL — 해당 문구 발견" if found else "",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 9. ED-004 catalog 검증 ────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        ed004 = next((d for d in cat["documents"] if d["id"] == "ED-004"), None)

        results.append(_check(ed004 is not None, "catalog: ED-004 항목 존재"))

        if ed004:
            results.append(_check(
                ed004.get("implementation_status") == "DONE",
                "catalog: ED-004 implementation_status == DONE",
                repr(ed004.get("implementation_status")),
            ))
            results.append(_check(
                ed004.get("form_type") == "manager_job_training_record",
                "catalog: ED-004 form_type == 'manager_job_training_record'",
                repr(ed004.get("form_type")),
            ))
            ev_status = ed004.get("evidence_status", "")
            results.append(_check(
                ev_status in ED004_VALID_EV_STATUSES,
                f"catalog: ED-004 evidence_status in {ED004_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_file 존재 확인
            ev_files = ed004.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            for efname in ED004_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR_ED004 / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:50]}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "ED-004 catalog 검증 중 예외", tb))

    # ── 10. training_types.yml 3종 코드 검증 ─────────────────────────────
    try:
        import yaml
        tt_path = Path("data/masters/safety/training/training_types.yml")
        with open(tt_path, encoding="utf-8") as f:
            tt = yaml.safe_load(f)
        tt_codes = {t["training_code"] for t in tt["training_types"]}

        for code in ED004_TRAINING_CODES:
            results.append(_check(
                code in tt_codes,
                f"training_types: {code} 등록됨",
            ))
            if code in tt_codes:
                entry = next(t for t in tt["training_types"] if t["training_code"] == code)
                results.append(_check(
                    entry.get("related_document_id") == "ED-004",
                    f"training_types: {code} → related_document_id == 'ED-004'",
                    repr(entry.get("related_document_id")),
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "training_types.yml 검증 중 예외", tb))

    # ── 11. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in ED004_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR_ED004 / efname
        if not fpath.exists():
            continue
        try:
            import json
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                f"evidence 검증결과: {efname[:45]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:45]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# CL-001 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_CL001_MINIMAL: dict = {
    "check_date": "2026-04-25",
    "work_location": "외부비계 A구간 (3층~7층)",
    "checker_name": "홍길동",
}

SAMPLE_CL001_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "check_date": "2026-04-25",
    "work_date": "2026-04-25 ~ 2026-05-20",
    "checker_name": "홍길동",
    "supervisor_name": "이감독",
    "work_location": "외부비계 A구간 (3층~7층)",
    "scaffold_type": "강관비계",
    "scaffold_height": "21.0m",
    "scaffold_length": "45.0m",
    "scaffold_location": "건물 남측 외벽",
    "scaffold_work_type": "외부 마감 타일 공사",
    "pre_install_items": [
        {"item": "구조검토서 또는 조립도 작성 여부", "result": "○", "note": ""},
        {"item": "지반 침하·균열 및 지지 상태 확인 여부", "result": "○", "note": ""},
        {"item": "사용 자재 규격·수량 확인 여부", "result": "○", "note": ""},
        {"item": "작업 전 근로자 안전교육 실시 여부", "result": "○", "note": "TBM 실시"},
        {"item": "작업구역 출입통제 및 안전표지 설치 여부", "result": "○", "note": ""},
    ],
    "structure_items": [
        {"item": "비계 기둥 수직도 적합 여부", "result": "○", "note": ""},
        {"item": "벽이음 설치 간격 기준 충족 여부", "result": "○", "note": ""},
        {"item": "가새 설치 상태 (X형 또는 V형 적정 여부)", "result": "○", "note": ""},
        {"item": "받침철물(베이스플레이트) 및 깔판 설치 여부", "result": "○", "note": ""},
        {"item": "부재 손상·변형·부식 없음 여부", "result": "○", "note": ""},
    ],
    "nonconformance_items": [],
    "inspector_sign": "홍길동",
    "supervisor_sign": "이감독",
    "manager_sign": "최소장",
    "sign_date": "2026-04-25",
}

# CL-001 필수 섹션 제목
CL001_REQUIRED_HEADINGS = [
    "비계 설치 점검표",
    "① 현장 기본정보",
    "② 비계 기본정보",
    "③ 설치 전 확인사항",
    "④ 구조·재료 점검",
    "⑤ 작업발판·승강설비 점검",
    "⑥ 안전난간·낙하물 방지 점검",
    "⑦ 조립·해체·변경 작업 점검",
    "⑧ 사용 전·사용 중 점검 결과",
    "⑨ 부적합 및 시정조치",
    "⑩ 확인 서명",
]

# CL-001 필수 고정 문구
CL001_REQUIRED_PHRASES = [
    "거푸집동바리 점검은 CL-002 별도 서식으로 관리한다",
    "법령 조항 및 기준은 현행 법령 원문 확인 후 현장에 적용한다",
    "점검 결과 부적합 사항은 사용 전 시정 완료 후 확인 서명을 받아야 한다",
]

# CL-001 evidence ID
CL001_EXPECTED_EVIDENCE_IDS = ["CL-001-L1", "CL-001-L2"]

# CL-001 evidence 파일명
CL001_EXPECTED_EVIDENCE_FILES = [
    "CL-001-L1_safety_rule_article_57_scaffold.json",
    "CL-001-L2_safety_rule_scaffold_workboard_railing.json",
]

CL001_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_cl001_smoke_test() -> list[tuple[str, str, str]]:
    """CL-001 비계 설치 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "scaffold_installation_checklist" in supported,
        "registry: scaffold_installation_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("scaffold_installation_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "비계 설치 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("scaffold_installation_checklist", SAMPLE_CL001_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("scaffold_installation_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("scaffold_installation_checklist", SAMPLE_CL001_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 10개 포함 확인
        for heading in CL001_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 고정 문구 3개 포함 확인
        for phrase in CL001_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:35]}...'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8. CL-001 catalog 검증 ────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl001 = next((d for d in cat["documents"] if d["id"] == "CL-001"), None)
        cl002 = next((d for d in cat["documents"] if d["id"] == "CL-002"), None)

        results.append(_check(cl001 is not None, "catalog: CL-001 항목 존재"))

        if cl001:
            results.append(_check(
                cl001.get("implementation_status") == "DONE",
                "catalog: CL-001 implementation_status == DONE",
                repr(cl001.get("implementation_status")),
            ))
            results.append(_check(
                cl001.get("form_type") == "scaffold_installation_checklist",
                "catalog: CL-001 form_type == 'scaffold_installation_checklist'",
                repr(cl001.get("form_type")),
            ))
            ev_status = cl001.get("evidence_status", "")
            results.append(_check(
                ev_status in CL001_VALID_EV_STATUSES,
                f"catalog: CL-001 evidence_status in {CL001_VALID_EV_STATUSES}",
                repr(ev_status),
            ))
            ev_ids = cl001.get("evidence_id") or []
            for eid in CL001_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))
            # evidence_file 존재 확인
            for efname in CL001_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:50]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # CL-002 상태 확인 (DONE — CL-002 구현 완료)
        results.append(_check(cl002 is not None, "catalog: CL-002 항목 존재"))
        if cl002:
            results.append(_check(
                cl002.get("implementation_status") == "DONE",
                "catalog: CL-002 implementation_status == DONE",
                repr(cl002.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-001 catalog 검증 중 예외", tb))

    # ── 9. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL001_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:45]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:45]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# CL-006 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_CL006_MINIMAL: dict = {
    "check_date": "2026-04-25",
    "work_location": "A동 북측 타워크레인 (TC-01)",
    "checker_name": "홍길동",
}

SAMPLE_CL006_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "check_date": "2026-04-25",
    "work_date": "2026-04-25",
    "checker_name": "홍길동",
    "supervisor_name": "이감독",
    "work_location": "A동 북측 타워크레인 (TC-01)",
    "crane_model": "POTAIN MCT 385",
    "crane_reg_no": "TC-2024-0001",
    "crane_capacity": "16ton",
    "crane_height": "78.0m",
    "crane_work_radius": "60.0m",
    "installation_date": "2026-01-15",
    "next_inspection_date": "2026-07-15",
    "operator_name": "김조종",
    "operator_license_no": "타워-2020-0001",
    "doc_check_items": [
        {"item": "설치검사 또는 정기검사 유효 여부", "result": "○", "note": "유효기간 2026-07-15"},
        {"item": "검사증·등록증 현장 비치 여부",     "result": "○", "note": ""},
        {"item": "타워크레인 작업계획서 작성 여부",   "result": "○", "note": ""},
        {"item": "장비사용계획서 작성 여부",          "result": "○", "note": ""},
        {"item": "제조사 사용설명서 현장 비치 여부",  "result": "○", "note": ""},
    ],
    "rope_check_items": [
        {"item": "와이어로프 마모·단선·킹크·부식 없음 여부", "result": "○", "note": ""},
        {"item": "와이어로프 윤활 상태 적정 여부",            "result": "○", "note": ""},
        {"item": "훅 해지장치 정상 작동 여부",                "result": "○", "note": ""},
        {"item": "권과방지장치 정상 작동 여부",               "result": "○", "note": ""},
        {"item": "과부하방지장치 정상 작동 여부",             "result": "○", "note": ""},
    ],
    "signal_check_items": [
        {"item": "신호수 배치 여부 (타워크레인마다)",  "result": "○", "note": "2명 배치"},
        {"item": "신호수 신호방법 및 신호 일치 여부", "result": "○", "note": ""},
        {"item": "조종자 자격·면허 유효 여부",         "result": "○", "note": ""},
        {"item": "정격하중 초과 인양 없음 여부",       "result": "○", "note": ""},
        {"item": "인양물 결박 및 슬링 상태 확인 여부", "result": "○", "note": ""},
    ],
    "nonconformance_items": [],
    "daily_inspector_sign": "홍길동",
    "operator_sign": "김조종",
    "supervisor_sign": "이감독",
    "manager_sign": "최소장",
    "sign_date": "2026-04-25",
}

# CL-006 필수 섹션 제목
CL006_REQUIRED_HEADINGS = [
    "타워크레인 자체 점검표",
    "① 현장 기본정보",
    "② 타워크레인 기본정보",
    "③ 검사·인증·서류 확인",
    "④ 설치 상태 점검",
    "⑤ 구조부·마스트·지브 점검",
    "⑥ 와이어로프·훅·권과방지장치 점검",
    "⑦ 브레이크·제동·리미트 장치 점검",
    "⑧ 전기·제어·비상정지장치 점검",
    "⑨ 작업반경·충돌방지·출입통제 점검",
    "⑩ 신호수·조종자·작업방법 점검",
    "⑪ 부적합 및 시정조치",
    "⑫ 확인 서명",
]

# CL-006 필수 포함 키워드
CL006_REQUIRED_KEYWORDS = [
    "이동식 크레인 점검표를 대체하지 않는다",
    "신호수",
    "충돌방지",
    "권과방지장치",
]

# CL-006 필수 고정 문구
CL006_REQUIRED_PHRASES = [
    "이동식 크레인 점검표를 대체하지 않는다",
]

# CL-006 evidence ID
CL006_EXPECTED_EVIDENCE_IDS = ["CL-006-L1", "CL-006-L2", "CL-006-L3"]

# CL-006 evidence 파일명
CL006_EXPECTED_EVIDENCE_FILES = [
    "CL-006-L1_safety_rule_articles_142_145_tower_crane.json",
    "CL-006-L2_safety_rule_article_146_crane_work.json",
    "CL-006-L3_safety_rule_crane_inspection_common.json",
]

CL006_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_cl006_smoke_test() -> list[tuple[str, str, str]]:
    """CL-006 타워크레인 자체 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "tower_crane_self_inspection_checklist" in supported,
        "registry: tower_crane_self_inspection_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("tower_crane_self_inspection_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "타워크레인 자체 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("tower_crane_self_inspection_checklist", SAMPLE_CL006_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("tower_crane_self_inspection_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("tower_crane_self_inspection_checklist", SAMPLE_CL006_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 12개 + 제목 포함 확인
        for heading in CL006_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in CL006_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8. CL-006 catalog 검증 ────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl006 = next((d for d in cat["documents"] if d["id"] == "CL-006"), None)
        cl001 = next((d for d in cat["documents"] if d["id"] == "CL-001"), None)
        cl002 = next((d for d in cat["documents"] if d["id"] == "CL-002"), None)

        results.append(_check(cl006 is not None, "catalog: CL-006 항목 존재"))

        if cl006:
            results.append(_check(
                cl006.get("implementation_status") == "DONE",
                "catalog: CL-006 implementation_status == DONE",
                repr(cl006.get("implementation_status")),
            ))
            results.append(_check(
                cl006.get("form_type") == "tower_crane_self_inspection_checklist",
                "catalog: CL-006 form_type == 'tower_crane_self_inspection_checklist'",
                repr(cl006.get("form_type")),
            ))
            ev_status = cl006.get("evidence_status", "")
            results.append(_check(
                ev_status in CL006_VALID_EV_STATUSES,
                f"catalog: CL-006 evidence_status in {CL006_VALID_EV_STATUSES}",
                repr(ev_status),
            ))
            ev_ids = cl006.get("evidence_id") or []
            for eid in CL006_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))
            for efname in CL006_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:55]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # CL-001, CL-002 상태 변경 없음 확인
        if cl001:
            results.append(_check(
                cl001.get("implementation_status") == "DONE",
                "catalog: CL-001 상태 변경 없음 (== DONE)",
                repr(cl001.get("implementation_status")),
            ))
        if cl002:
            results.append(_check(
                cl002.get("implementation_status") == "DONE",
                "catalog: CL-002 상태 변경 없음 (== DONE)",
                repr(cl002.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-006 catalog 검증 중 예외", tb))

    # ── 9. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL006_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:50]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:50]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# CL-002 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_CL002_MINIMAL: dict = {
    "check_date":    "2026-04-25",
    "work_location": "지상 3층 슬래브",
    "checker_name":  "홍길동",
}

SAMPLE_CL002_FULL: dict = {
    "site_name":            "테스트건설(주) 테스트현장",
    "project_name":         "테스트현장 신축공사",
    "check_date":           "2026-04-25",
    "work_location":        "지상 3층 슬래브",
    "work_date":            "2026-04-25 ~ 2026-04-26",
    "supervisor_name":      "이감독",
    "structure_type":       "슬래브",
    "floor_level":          "지상 3층",
    "work_area":            "A동 북측",
    "formwork_type":        "합판 거푸집",
    "shoring_type":         "시스템동바리",
    "checker_name":         "홍길동",
    "inspector_sign":       "홍길동",
    "supervisor_sign":      "이감독",
    "work_commander_sign":  "박지휘",
    "manager_sign":         "장소장",
    "sign_date":            "2026-04-25",
    "nonconformance_items": [
        {
            "content":   "3번 동바리 베이스플레이트 미설치",
            "location":  "A동 북측 3열",
            "action":    "베이스플레이트 추가 설치",
            "deadline":  "2026-04-25",
            "completed": "완료",
        }
    ],
}

# CL-002 필수 섹션 제목 (13개)
CL002_REQUIRED_HEADINGS = [
    "거푸집 및 동바리 설치 점검표",
    "1. 현장 기본정보",
    "2. 구조물 및 작업구간 정보",
    "3. 조립도·작업계획서 확인",
    "4. 재료 및 부재 상태 점검",
    "5. 동바리 설치 상태 점검",
    "6. 거푸집 설치 상태 점검",
    "7. 침하·전도·변형 방지 조치",
    "8. 작업발판·통로·추락방지 조치",
    "9. 콘크리트 타설 전 점검",
    "10. 콘크리트 타설 중 점검",
    "11. 해체 전 안전조치",
    "12. 부적합 및 시정조치",
    "13. 확인 서명",
]

# CL-002 필수 고정 문구
CL002_REQUIRED_PHRASES = [
    "비계 점검표(CL-001)를 대체하지 않는다",
    "조립도와 작업계획서에 따라 설치·점검한다",
    "콘크리트 타설 전 부적합 사항은 사용 전 시정 완료 후 재확인한다",
    "법령 조항은 현행 원문 확인 후 현장에 적용한다",
]

# CL-002 필수 키워드 (조립도, 콘크리트, 해체)
CL002_REQUIRED_KEYWORDS = [
    "조립도",
    "콘크리트 타설 전 점검",
    "해체 전 안전조치",
]

# CL-002 evidence
CL002_EXPECTED_EVIDENCE_IDS = ["CL-002-L1", "CL-002-L2", "CL-002-L3"]
CL002_EXPECTED_EVIDENCE_FILES = [
    "CL-002-L1_safety_rule_articles_328_330_formwork_material.json",
    "CL-002-L2_safety_rule_articles_331_333_assembly_drawing.json",
    "CL-002-L3_safety_rule_articles_334_337_concrete_pour.json",
]
CL002_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_cl002_smoke_test() -> list[tuple[str, str, str]]:
    """CL-002 거푸집 및 동바리 설치 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "formwork_shoring_installation_checklist" in supported,
        "registry: formwork_shoring_installation_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("formwork_shoring_installation_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "거푸집 및 동바리 설치 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("formwork_shoring_installation_checklist", SAMPLE_CL002_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("formwork_shoring_installation_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~11. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("formwork_shoring_installation_checklist", SAMPLE_CL002_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트명에 "거푸집" 또는 "동바리" 포함
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("거푸집" in n or "동바리" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))

        # 섹션 제목 13개 포함
        for heading in CL002_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 고정 문구 4개 포함
        for phrase in CL002_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:50]}...'",
            ))

        # 필수 키워드 포함
        for kw in CL002_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 12. catalog + evidence 검증 ──────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl002 = next((d for d in cat["documents"] if d["id"] == "CL-002"), None)
        cl001 = next((d for d in cat["documents"] if d["id"] == "CL-001"), None)
        cl003 = next((d for d in cat["documents"] if d["id"] == "CL-003"), None)
        cl006 = next((d for d in cat["documents"] if d["id"] == "CL-006"), None)

        results.append(_check(cl002 is not None, "catalog: CL-002 항목 존재"))

        if cl002:
            results.append(_check(
                cl002.get("implementation_status") == "DONE",
                "catalog: CL-002 implementation_status == DONE",
                repr(cl002.get("implementation_status")),
            ))
            results.append(_check(
                cl002.get("form_type") == "formwork_shoring_installation_checklist",
                "catalog: CL-002 form_type == 'formwork_shoring_installation_checklist'",
                repr(cl002.get("form_type")),
            ))
            ev_status = cl002.get("evidence_status", "")
            results.append(_check(
                ev_status in CL002_VALID_EV_STATUSES,
                f"catalog: CL-002 evidence_status in {CL002_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 3개 포함
            ev_ids = cl002.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in CL002_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in CL002_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:65]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # CL-001, CL-003, CL-006 상태 변경 없음 확인
        if cl001:
            results.append(_check(
                cl001.get("implementation_status") == "DONE",
                "CL-001 상태 변경 없음 (DONE 유지)",
                repr(cl001.get("implementation_status")),
            ))
        if cl003:
            results.append(_check(
                cl003.get("implementation_status") == "DONE",
                "CL-003 상태 변경 없음 (DONE 유지)",
                repr(cl003.get("implementation_status")),
            ))
        if cl006:
            results.append(_check(
                cl006.get("implementation_status") == "DONE",
                "CL-006 상태 변경 없음 (DONE 유지)",
                repr(cl006.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-002 catalog/evidence 검증 중 예외", tb))

    # ── 13. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL002_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:60]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:60]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# CL-003 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_CL003_MINIMAL: dict = {
    "check_date":    "2026-04-25",
    "work_location": "테스트현장 1공구",
    "checker_name":  "홍길동",
}

SAMPLE_CL003_FULL: dict = {
    "site_name":            "테스트건설(주) 테스트현장",
    "project_name":         "테스트현장 신축공사",
    "check_date":           "2026-04-25",
    "work_location":        "1공구 굴착구간",
    "equipment_type":       "지게차",
    "equipment_model":      "현대 20BH",
    "equipment_reg_no":     "12가 3456",
    "equipment_capacity":   "2.0 ton",
    "operator_name":        "이운전",
    "operator_license_no":  "지게차운전기능사 제2020-12345호",
    "guide_worker_name":    "김유도",
    "work_commander_name":  "박지휘",
    "checker_name":         "홍길동",
    "operator_sign":        "이운전",
    "guide_worker_sign":    "김유도",
    "work_commander_sign":  "박지휘",
    "supervisor_sign":      "최감독",
    "manager_sign":         "장소장",
    "sign_date":            "2026-04-25",
    "nonconformance_items": [
        {
            "content":   "좌측 후미등 작동 불량",
            "location":  "후미등 좌측",
            "action":    "교체 조치",
            "deadline":  "2026-04-25",
            "completed": "완료",
        }
    ],
}

# CL-003 필수 섹션 제목 (13개)
CL003_REQUIRED_HEADINGS = [
    "건설장비 일일 사전점검표",
    "1. 현장 기본정보",
    "2. 장비 기본정보",
    "3. 운전자·유도자·작업지휘자 확인",
    "4. 서류·검사·보험 확인",
    "5. 외관·누유·타이어/궤도 점검",
    "6. 전조등·후미등·경광등·후진경보기 점검",
    "7. 제동장치·조향장치·비상정지 점검",
    "8. 전도·전락·지반 상태 점검",
    "9. 접촉·충돌·작업반경 통제 점검",
    "10. 적재·인양·하역 상태 점검",
    "11. 장비별 추가 점검",
    "12. 부적합 및 시정조치",
    "13. 확인 서명",
]

# CL-003 필수 고정 문구
CL003_REQUIRED_PHRASES = [
    "작업계획서 및 장비사용계획서를 대체하지 않는다",
    "타워크레인, 이동식 크레인, 비계, 거푸집동바리 전용 점검은 별도 서식으로 관리한다",
    "부적합 사항은 사용 전 시정 완료 후 재확인한다",
    "법령 조항은 현행 원문 확인 후 현장에 적용한다",
]

# CL-003 장비별 추가 점검 — 지게차 필수 포함 항목 키워드
CL003_FORKLIFT_KEYWORDS = [
    "후진경보기",
    "헤드가드",
    "백레스트",
    "좌석안전띠",
    "포크",
]

# CL-003 evidence
CL003_EXPECTED_EVIDENCE_IDS = ["CL-003-L1", "CL-003-L2", "CL-003-L3"]
CL003_EXPECTED_EVIDENCE_FILES = [
    "CL-003-L1_safety_rule_vehicle_construction_equipment.json",
    "CL-003-L2_safety_rule_material_handling_equipment_common.json",
    "CL-003-L3_safety_rule_forklift_specific.json",
]
CL003_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_cl003_smoke_test() -> list[tuple[str, str, str]]:
    """CL-003 건설장비 일일 사전점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "construction_equipment_daily_checklist" in supported,
        "registry: construction_equipment_daily_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("construction_equipment_daily_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "건설장비 일일 사전점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("construction_equipment_daily_checklist", SAMPLE_CL003_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("construction_equipment_daily_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~11. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("construction_equipment_daily_checklist", SAMPLE_CL003_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트 이름에 "건설장비" 또는 "점검" 포함
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("건설장비" in n or "점검" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))

        # 섹션 제목 13개 포함
        for heading in CL003_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 고정 문구 4개 포함
        for phrase in CL003_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:45]}...'",
            ))

        # 지게차 추가 점검 키워드 확인 (equipment_type = 지게차)
        for kw in CL003_FORKLIFT_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"지게차 점검 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 12. catalog + evidence 검증 ──────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl003 = next((d for d in cat["documents"] if d["id"] == "CL-003"), None)
        cl001 = next((d for d in cat["documents"] if d["id"] == "CL-001"), None)
        cl002 = next((d for d in cat["documents"] if d["id"] == "CL-002"), None)
        cl006 = next((d for d in cat["documents"] if d["id"] == "CL-006"), None)

        results.append(_check(cl003 is not None, "catalog: CL-003 항목 존재"))

        if cl003:
            results.append(_check(
                cl003.get("implementation_status") == "DONE",
                "catalog: CL-003 implementation_status == DONE",
                repr(cl003.get("implementation_status")),
            ))
            results.append(_check(
                cl003.get("form_type") == "construction_equipment_daily_checklist",
                "catalog: CL-003 form_type == 'construction_equipment_daily_checklist'",
                repr(cl003.get("form_type")),
            ))
            ev_status = cl003.get("evidence_status", "")
            results.append(_check(
                ev_status in CL003_VALID_EV_STATUSES,
                f"catalog: CL-003 evidence_status in {CL003_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 3개 포함
            ev_ids = cl003.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in CL003_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in CL003_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:60]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # CL-001, CL-002, CL-006 상태 변경 없음 확인
        if cl001:
            results.append(_check(
                cl001.get("implementation_status") == "DONE",
                "CL-001 상태 변경 없음 (DONE 유지)",
                repr(cl001.get("implementation_status")),
            ))
        if cl002:
            results.append(_check(
                cl002.get("implementation_status") == "DONE",
                "CL-002 상태 변경 없음 (DONE 유지)",
                repr(cl002.get("implementation_status")),
            ))
        if cl006:
            results.append(_check(
                cl006.get("implementation_status") == "DONE",
                "CL-006 상태 변경 없음 (DONE 유지)",
                repr(cl006.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-003 catalog/evidence 검증 중 예외", tb))

    # ── 13. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL003_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:55]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:55]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# PTW-002 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_PTW002_MINIMAL: dict = {
    "site_name":       "테스트건설(주) 테스트현장",
    "work_date":       "2026-04-25",
    "work_time":       "09:00 ~ 18:00",
    "work_location":   "지하 2층 기계실 배관 구간",
    "trade_name":      "소방 배관 공사",
    "work_content":    "소방 배관 용접 및 연결",
    "contractor":      "협력업체 A",
    "work_supervisor": "홍길동",
}

SAMPLE_PTW002_FULL: dict = {
    "site_name":          "테스트건설(주) 테스트현장",
    "project_name":       "테스트현장 신축공사",
    "permit_no":          "PTW-2026-0001",
    "work_date":          "2026-04-25",
    "work_time":          "09:00 ~ 18:00",
    "work_location":      "지하 2층 기계실 배관 구간",
    "trade_name":         "소방 배관 공사",
    "work_content":       "소방 배관 용접 및 연결 (배관 규격 65A)",
    "contractor":         "협력업체 A",
    "work_supervisor":    "홍길동",
    "equipment_list":     "전기 용접기, 그라인더",
    "combustibles_present": "주변 단열재 일부 존재",
    "combustibles_removed": "완료 — 작업반경 5m 내 제거",
    "fire_blanket_used":   "설치",
    "extinguisher_placed": "소화기 2대 비치 (3.3kg ABC)",
    "ventilation_status":  "환기 팬 1대 가동",
    "fire_watch_required": "필요",
    "fire_watch_name":     "이감시",
    "permit_issuer":       "김허가자",
    "supervisor_name":     "박감독",
    "validity_period":     "당일 1회 작업 한정",
    "work_end_time":       "18:00",
    "post_work_confirmer": "박감독",
    "work_types":          ["용접", "그라인더"],
    "pre_work_checks": [
        "작업내용·작업일시·안전점검 사항 게시 (제241조 제4항)",
        "작업장 주변 가연물 제거",
        "용접방화포 또는 방염포 설치",
        "소화기 또는 소화설비 비치",
    ],
    "workers": [
        {"name": "이순신", "job_type": "용접공"},
        {"name": "강감찬", "job_type": "보조"},
    ],
}

# PTW-002 필수 섹션 제목 (13개)
PTW002_REQUIRED_HEADINGS = [
    "화기작업 허가서",
    "1. 현장 기본정보",
    "2. 작업 기본정보",
    "3. 화기작업 유형",
    "4. 작업장소 및 가연물 확인",
    "5. 작업 전 안전조치",
    "6. 소화설비 및 방화조치",
    "7. 화재감시자 배치 확인",
    "8. 보호구 및 작업장비 확인",
    "9. 작업허가 승인",
    "10. 작업 중 점검",
    "11. 작업 종료 후 잔불 확인",
    "12. 사진 및 증빙자료",
    "13. 확인 서명",
]

# PTW-002 필수 키워드
PTW002_REQUIRED_KEYWORDS = [
    "화재감시자",
    "용접방화포",
    "소화기",
    "가연물 제거",
    "잔불 확인",
]

# PTW-002 필수 고정 문구
PTW002_REQUIRED_PHRASES = [
    "작업내용·작업일시·안전점검 사항 게시",
    "법정 안전보건교육 수료증을 대체하지 않는다",
    "화재감시자 배치 여부는",
    "현장 조건에 따라 최종 판단한다",
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다",
]

# PTW-002 evidence ID 목록
PTW002_EXPECTED_EVIDENCE_IDS = [
    "PTW-002-L1", "PTW-002-L2", "PTW-002-L3", "PTW-002-L4",
    "PTW-002-K1", "PTW-002-K2",
]

# PTW-002 evidence 파일명
PTW002_EXPECTED_EVIDENCE_FILES = [
    "PTW-002-L1_safety_rule_article_236_fire_risk_work.json",
    "PTW-002-L2_safety_rule_article_241_hot_work_measures.json",
    "PTW-002-L3_safety_rule_article_241_2_fire_watch.json",
    "PTW-002-L4_safety_rule_articles_243_244_fire_extinguishing.json",
    "PTW-002-K1_kosha_p_94_2021_hot_work_permit.json",
    "PTW-002-K2_kosha_f_1_2023_welding_cutting_fire_prevention.json",
]

PTW002_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}

# 기존 서식 상태 불변 확인 목록: {doc_id: (expected_form_type, expected_impl_status)}
PTW002_UNCHANGED_IDS = {
    "RA-001":  ("risk_assessment",       "DONE"),
    "RA-004":  ("tbm_log",               "DONE"),
    "PTW-001": ("confined_space_permit", "DONE"),
    "PTW-003": ("work_at_height_permit", "DONE"),  # PTW-003 구현 완료
}


def run_ptw002_smoke_test() -> list[tuple[str, str, str]]:
    """PTW-002 화기작업 허가서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "hot_work_permit" in supported,
        "registry: hot_work_permit 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("hot_work_permit")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "화기작업 허가서",
            "display_name == '화기작업 허가서'",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "workers",
            "repeat_field == 'workers'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("hot_work_permit", SAMPLE_PTW002_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("hot_work_permit", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~12. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("hot_work_permit", SAMPLE_PTW002_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트 제목에 "화기작업 허가서" 포함
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("화기" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))
        results.append(_check(
            "화기작업 허가서" in all_text,
            "시트 제목 '화기작업 허가서' 포함",
        ))

        # 섹션 제목 13개 포함
        for heading in PTW002_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in PTW002_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 필수 고정 문구 포함
        for phrase in PTW002_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:50]}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 13. catalog + evidence 검증 ──────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        docs = {d["id"]: d for d in cat["documents"]}
        ptw002 = docs.get("PTW-002")

        results.append(_check(ptw002 is not None, "catalog: PTW-002 항목 존재"))

        if ptw002:
            results.append(_check(
                ptw002.get("implementation_status") == "DONE",
                "catalog: PTW-002 implementation_status == DONE",
                repr(ptw002.get("implementation_status")),
            ))
            results.append(_check(
                ptw002.get("form_type") == "hot_work_permit",
                "catalog: PTW-002 form_type == 'hot_work_permit'",
                repr(ptw002.get("form_type")),
            ))
            ev_status = ptw002.get("evidence_status", "")
            results.append(_check(
                ev_status in PTW002_VALID_EV_STATUSES,
                f"catalog: PTW-002 evidence_status in {PTW002_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 6개 포함
            ev_ids = ptw002.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in PTW002_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in PTW002_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:65]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # RA-001, RA-004, PTW-001, PTW-003 상태 변경 없음 확인
        for doc_id, (expected_form_type, expected_status) in PTW002_UNCHANGED_IDS.items():
            d = docs.get(doc_id)
            if d is None:
                continue
            if expected_form_type is not None:
                results.append(_check(
                    d.get("form_type") == expected_form_type,
                    f"catalog: {doc_id} form_type 변경 없음 (== '{expected_form_type}')",
                    repr(d.get("form_type")),
                ))
            results.append(_check(
                d.get("implementation_status") == expected_status,
                f"catalog: {doc_id} implementation_status 변경 없음 (== {expected_status})",
                repr(d.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "PTW-002 catalog/evidence 검증 중 예외", tb))

    # ── 14. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in PTW002_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:60]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:60]}...", str(exc)))

    return results


# ---------------------------------------------------------------------------
# PTW-003 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_PTW003_MINIMAL: dict = {
    "site_name":       "테스트건설(주) 테스트현장",
    "work_date":       "2026-04-25",
    "work_time":       "09:00 ~ 18:00",
    "work_location":   "외부비계 A구간 5층",
    "trade_name":      "소방 배관 공사",
    "work_content":    "5층 외부 소방 배관 설치",
    "contractor":      "협력업체 A",
    "work_supervisor": "홍길동",
}

SAMPLE_PTW003_FULL: dict = {
    "site_name":          "테스트건설(주) 테스트현장",
    "project_name":       "테스트현장 신축공사",
    "permit_no":          "PTW-WAH-2026-0001",
    "work_date":          "2026-04-25",
    "work_time":          "09:00 ~ 18:00",
    "work_location":      "외부비계 A구간 5층 (높이 약 15m)",
    "trade_name":         "소방 배관 공사",
    "work_content":       "5층 외부 소방 배관 설치 및 지지대 고정",
    "contractor":         "협력업체 A",
    "work_supervisor":    "홍길동",
    "work_height":        "15.0",
    "equipment_list":     "비계 발판, 안전대",
    "equipment_type":     "강관 비계",
    "fall_risk_present":  "있음",
    "opening_present":    "없음",
    "workboard_installed": "설치",
    "railing_installed":  "설치",
    "lanyard_worn":       "착용",
    "anchor_confirmed":   "확인",
    "falling_zone_set":   "설정",
    "access_control":     "실시",
    "weather_confirmed":  "확인 — 풍속 3m/s, 맑음",
    "permit_issuer":      "김허가자",
    "supervisor_name":    "박감독",
    "safety_manager_sign": "최안전",
    "validity_period":    "당일 1회 작업 한정",
    "work_types":         ["비계 작업", "작업발판 작업"],
    "pre_work_checks": [
        "작업구역 추락위험 확인 (제42조)",
        "작업발판 설치 상태 확인 (제42조)",
        "안전난간 설치 상태 확인 (제43조)",
        "안전대 착용 확인 (제44조)",
        "기상조건 확인 (제37조 — 풍속 10m/s 이상 시 작업 중지 기준)",
    ],
    "workboard_checks": [
        "작업발판 폭·고정 상태 확인",
        "비계 사용 시 CL-001 비계 점검표 병행 확인 (제57조 이하)",
    ],
    "workers": [
        {"name": "이순신", "job_type": "배관공"},
        {"name": "강감찬", "job_type": "보조"},
    ],
}

# PTW-003 필수 섹션 제목 (13개)
PTW003_REQUIRED_HEADINGS = [
    "고소작업 허가서",
    "1. 현장 기본정보",
    "2. 작업 기본정보",
    "3. 고소작업 유형",
    "4. 작업장소 및 추락위험 확인",
    "5. 작업 전 안전조치",
    "6. 작업발판·비계·사다리 확인",
    "7. 고소작업대 확인",
    "8. 안전대·추락방지설비 확인",
    "9. 낙하물 방지 및 출입통제",
    "10. 작업허가 승인",
    "11. 작업 중 점검",
    "12. 사진 및 증빙자료",
    "13. 확인 서명",
]

# PTW-003 필수 키워드
PTW003_REQUIRED_KEYWORDS = [
    "추락위험",
    "작업발판",
    "안전난간",
    "개구부",
    "안전대",
    "부착설비",
    "고소작업대",
    "사다리",
    "낙하물",
]

# PTW-003 필수 고정 문구
PTW003_REQUIRED_PHRASES = [
    "비계 작업은 CL-001 비계 설치 점검표와 병행하여 확인한다",
    "장비 점검표 및 장비 상태를 별도로 확인한다",
    "법정 안전보건교육 수료증을 대체하지 않는다",
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다",
]

# PTW-003 evidence ID 목록
PTW003_EXPECTED_EVIDENCE_IDS = [
    "PTW-003-L1", "PTW-003-L2", "PTW-003-L3", "PTW-003-L4",
    "PTW-003-K1", "PTW-003-K2", "PTW-003-K3",
]

# PTW-003 evidence 파일명
PTW003_EXPECTED_EVIDENCE_FILES = [
    "PTW-003-L1_safety_rule_article_37_weather_stop.json",
    "PTW-003-L2_safety_rule_articles_42_45_fall_prevention.json",
    "PTW-003-L3_safety_rule_article_57_scaffold_supervision.json",
    "PTW-003-L4_safety_rule_articles_86_plus_aerial_work_platform.json",
    "PTW-003-K1_kosha_p_94_2021_work_permit.json",
    "PTW-003-K2_kosha_c_74_2015_aerial_work_platform.json",
    "PTW-003-K3_kosha_m_155_2023_mobile_aerial_work_platform.json",
]

PTW003_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"}

# PTW-003 구현 후 불변 확인 목록
PTW003_UNCHANGED_IDS = {
    "RA-001":  ("risk_assessment",       "DONE"),
    "RA-004":  ("tbm_log",               "DONE"),
    "PTW-002": ("hot_work_permit",       "DONE"),
    "CL-001":  ("scaffold_installation_checklist", "DONE"),
}


def run_ptw003_smoke_test() -> list[tuple[str, str, str]]:
    """PTW-003 고소작업 허가서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "work_at_height_permit" in supported,
        "registry: work_at_height_permit 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("work_at_height_permit")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "고소작업 허가서",
            "display_name == '고소작업 허가서'",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "workers",
            "repeat_field == 'workers'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("work_at_height_permit", SAMPLE_PTW003_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("work_at_height_permit", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~13. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("work_at_height_permit", SAMPLE_PTW003_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트 제목에 "고소작업 허가서" 포함
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("고소" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))
        results.append(_check(
            "고소작업 허가서" in all_text,
            "시트 제목 '고소작업 허가서' 포함",
        ))

        # 섹션 제목 13개 포함
        for heading in PTW003_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in PTW003_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 필수 고정 문구 포함
        for phrase in PTW003_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:50]}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 14. catalog + evidence 검증 ──────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        docs = {d["id"]: d for d in cat["documents"]}
        ptw003 = docs.get("PTW-003")

        results.append(_check(ptw003 is not None, "catalog: PTW-003 항목 존재"))

        if ptw003:
            results.append(_check(
                ptw003.get("implementation_status") == "DONE",
                "catalog: PTW-003 implementation_status == DONE",
                repr(ptw003.get("implementation_status")),
            ))
            results.append(_check(
                ptw003.get("form_type") == "work_at_height_permit",
                "catalog: PTW-003 form_type == 'work_at_height_permit'",
                repr(ptw003.get("form_type")),
            ))
            ev_status = ptw003.get("evidence_status", "")
            results.append(_check(
                ev_status in PTW003_VALID_EV_STATUSES,
                f"catalog: PTW-003 evidence_status in {PTW003_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 7개 포함
            ev_ids = ptw003.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in PTW003_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in PTW003_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:70]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # RA-001, RA-004, PTW-002, CL-001 상태 변경 없음 확인
        for doc_id, (expected_form_type, expected_status) in PTW003_UNCHANGED_IDS.items():
            d = docs.get(doc_id)
            if d is None:
                continue
            results.append(_check(
                d.get("form_type") == expected_form_type,
                f"catalog: {doc_id} form_type 변경 없음 (== '{expected_form_type}')",
                repr(d.get("form_type")),
            ))
            results.append(_check(
                d.get("implementation_status") == expected_status,
                f"catalog: {doc_id} implementation_status 변경 없음 (== {expected_status})",
                repr(d.get("implementation_status")),
            ))

        # related_documents 연계 확인 (RA-001, RA-004, CL-001, CL-007)
        if ptw003:
            rel_docs = ptw003.get("related_documents") or []
            for req_id in ("RA-001", "RA-004", "CL-001", "CL-007"):
                results.append(_check(
                    req_id in rel_docs,
                    f"catalog: PTW-003 related_documents에 {req_id} 포함",
                    repr(rel_docs) if req_id not in rel_docs else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "PTW-003 catalog/evidence 검증 중 예외", tb))

    # ── 15. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in PTW003_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:65]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:65]}...", str(exc)))

    return results


# ===========================================================================
# PTW-007 — 중량물 인양·중장비사용 작업 허가서
# ===========================================================================

SAMPLE_PTW007_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "work_date": "2026-04-25",
    "work_time": "08:00 ~ 17:00",
    "work_location": "3층 기계실",
    "trade_name": "기계설비",
    "work_content": "공조기 반입 인양 (3,200 kg)",
    "contractor": "협력업체 A",
    "work_supervisor": "홍길동",
}

SAMPLE_PTW007_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "work_date": "2026-04-25",
    "work_time": "08:00 ~ 17:00",
    "work_location": "3층 기계실 (EL+9.0m)",
    "trade_name": "기계설비",
    "work_content": "공조기 반입 인양 (3,200 kg, 2,400×1,600×1,800mm)",
    "contractor": "협력업체 A",
    "work_supervisor": "홍길동",
    "permit_no": "PTW-007-2026-001",
    "lifting_object_name": "공조기 AHU-1",
    "lifting_weight": "3.2 ton",
    "lifting_size": "2,400×1,600×1,800mm",
    "lifting_height": "12 m",
    "lifting_distance": "15 m",
    "lifting_route": "반입구 → 1층 → 리프트 → 3층 기계실",
    "equipment_name": "25ton 카고크레인",
    "equipment_rated_load": "25 ton",
    "work_radius": "12 m",
    "outrigger_installed": "설치 완료",
    "ground_condition": "아스팔트 포장, 지반 침하 없음",
    "rigging_method": "4점 슬링벨트",
    "rigging_gear": "슬링벨트 4ea (SWL 2ton), 샤클 4ea",
    "signal_worker_name": "김신호 (무전기 사용)",
    "permit_issuer": "이허가 (안전관리자)",
    "supervisor_name": "박감독",
    "safety_manager_sign": "이안전",
    "work_end_confirmer": "홍길동",
    "final_sign": "이허가",
    "during_work_issues": "없음",
    "work_end_time": "14:30",
    "lifting_types": ["카고크레인 인양"],
    "workplan_checks": [
        "WP-005 중량물 취급 작업계획서 확인 (제38조 ① 11호 + 별표4)",
        "CL-003 건설장비 일일점검표 확인",
        "인양물 중량 산정 근거 확인",
    ],
    "equipment_checks": [
        "장비명 및 장비번호 확인",
        "정격하중 표시 확인 (제133조)",
        "작업반경별 허용하중 확인 (제135조)",
        "아웃트리거 설치 확인 (KOSHA C-102-2023 참고)",
    ],
    "rigging_checks": [
        "와이어로프 손상 여부 (꼬임끊어짐·소선단선 7% 이상·마모·부식) (제163조·제166조)",
        "슬링벨트 손상 여부 (절단·손상·봉제부 풀림) (제163조·제169조)",
        "샤클 변형·균열 여부 (제168조)",
        "달기구 정격하중 확인 (제163조·제164조)",
    ],
    "signal_checks": [
        "신호수 배치 (제40조·제146조 ⑤)",
        "작업반경 내 출입통제 (제146조 ④)",
        "하부 근로자 출입금지 확인 (제146조 ④)",
    ],
    "pre_work_checks": [
        "작업 전 TBM 실시",
        "인양물 중량 확인 (제133조·제135조)",
        "하부 통제 (제146조 ④)",
        "장비 점검표(CL-003) 확인",
        "작업계획서(WP-005) 확인 (제38조 ① 11호)",
    ],
    "post_work_checks": [
        "인양물 설치 상태 확인",
        "달기구 해체 상태 확인",
        "장비 원위치 및 전원 차단",
        "작업구역 정리정돈",
        "종료 확인자 서명",
    ],
    "photo_items": [
        "줄걸이 체결 사진",
        "장비 및 아웃트리거 사진",
    ],
    "workers": [
        {"name": "홍길동", "job_type": "작업책임자"},
        {"name": "김신호", "job_type": "신호수"},
        {"name": "이인양", "job_type": "작업자"},
    ],
    "validity_period": "당일 1회 작업 한정",
}

# PTW-007 필수 섹션 제목 (14개)
PTW007_REQUIRED_HEADINGS = [
    "중량물 인양·중장비사용 작업 허가서",
    "1. 현장 기본정보",
    "2. 작업 기본정보 및 작업자 명단",
    "3. 작업계획서 확인",
    "4. 인양물 정보",
    "5. 양중기 선택 및 정격하중 확인",
    "6. 달기구·줄걸이 확인",
    "7. 인양 경로 및 위험구간 확인",
    "8. 작업장소·지반·아웃트리거 확인",
    "9. 신호수·통신·작업반경 통제",
    "10. 작업 전 안전조치",
    "11. 작업허가 승인",
    "12. 작업 중 점검",
    "13. 사진 및 증빙자료",
    "14. 작업 종료 및 확인 서명",
]

# PTW-007 필수 키워드
PTW007_REQUIRED_KEYWORDS = [
    "WP-005 중량물 취급 작업계획서",
    "장비점검표",
    "정격하중",
    "줄걸이",
    "샤클",
    "와이어로프",
    "슬링벨트",
    "신호수",
    "작업반경",
    "하부 근로자 출입금지",
    "아웃트리거",
]

# PTW-007 필수 고정 문구
PTW007_REQUIRED_PHRASES = [
    "법정 안전보건교육 수료증을 대체하지 않는다",
    "WP-005 중량물 취급 작업계획서를 대체하지 않는다",
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다",
    "KOSHA GUIDE P-94-2021 중장비사용작업허가 참고",
]

# PTW-007 evidence ID 목록
PTW007_EXPECTED_EVIDENCE_IDS = [
    "PTW-007-L1", "PTW-007-L2", "PTW-007-L3", "PTW-007-L4",
    "PTW-007-L5", "PTW-007-L6",
    "PTW-007-K1", "PTW-007-K2", "PTW-007-K3",
]

# PTW-007 evidence 파일명
PTW007_EXPECTED_EVIDENCE_FILES = [
    "PTW-007-L1_safety_rule_article_38_table4_heavy_lifting_plan.json",
    "PTW-007-L2_safety_rule_article_40_lifting_signal.json",
    "PTW-007-L3_safety_rule_articles_132_135_lifting_device_safety.json",
    "PTW-007-L4_safety_rule_articles_138_146_crane_work.json",
    "PTW-007-L5_safety_rule_articles_163_170_rigging_gears.json",
    "PTW-007-L6_safety_rule_articles_221_5_385_heavy_lifting.json",
    "PTW-007-K1_kosha_c_102_2023_mobile_crane_lifting_plan.json",
    "PTW-007-K2_kosha_m_186_2015_rigging_wire_rope.json",
    "PTW-007-K3_kosha_p_94_2021_work_permit.json",
]

PTW007_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"}

# PTW-007 구현 후 불변 확인 목록 (WP-005 구현 완료 반영)
PTW007_UNCHANGED_IDS = {
    "RA-001":  ("risk_assessment",              "DONE"),
    "RA-004":  ("tbm_log",                      "DONE"),
    "PTW-002": ("hot_work_permit",              "DONE"),
    "PTW-003": ("work_at_height_permit",        "DONE"),
    "CL-003":  ("construction_equipment_daily_checklist", "DONE"),
    "WP-005":  ("heavy_lifting_workplan",       "DONE"),
}


# ---------------------------------------------------------------------------
# WP-005 샘플 form_data
# ---------------------------------------------------------------------------

SAMPLE_WP005_MINIMAL: dict = {
    "object_name": "H빔 (H-200×200×8×12)",
    "object_weight": "0.85ton",
    "work_method": "카고크레인으로 직상방향 인양 후 설치 위치로 수평 이동",
    "emergency_measure": "인양 중 이상 발생 시 즉시 작업 중지, 안전한 위치에 하역 후 관리감독자 보고",
}

SAMPLE_WP005_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "work_location": "3층 기둥 설치 구간 (EL+9.0m)",
    "work_date": "2026-04-25",
    "supervisor": "홍길동 (반장)",
    "contractor": "협력업체 A",
    "prepared_by": "이안전",
    "object_name": "H빔 (H-200×200×8×12)",
    "object_shape": "H형강 직선 부재",
    "object_weight": "0.85ton",
    "object_size": "길이 6,000mm × 폭 200mm × 높이 200mm",
    "weight_basis": "설계도서 철골 부재표 확인 (구조도면 S-103)",
    "center_of_gravity": "부재 중앙부 (양단 균형)",
    "work_method": "카고크레인으로 직상방향 인양 후 설치 위치로 수평 이동, 신호수 유도 하에 볼트 체결",
    "transport_route": "야적장(동측) → 타워크레인 양중 → 3층 철골 설치 위치",
    "work_site_condition": "3층 슬래브 위, 주변 가설펜스 설치됨",
    "ground_condition": "콘크리트 슬래브 (설계하중 확인, 매설물 없음)",
    "access_control": "인양경로 하부 바리케이드 설치, 신호수 1명 배치",
    "equipment_name": "카고크레인 (25ton급)",
    "equipment_capacity": "25ton (작업반경 5m 기준 허용하중 8ton)",
    "auxiliary_equipment": "체인블록 (1ton), 마그네트 클램프",
    "rigging_method": "2줄걸이 (슬링벨트 양단 걸이)",
    "rigging_angle": "60°",
    "rigging_gear": "슬링벨트 (50mm × 4m, 사용하중 2ton) × 2조, 샤클 (WLL 3ton) × 2개",
    "rigging_safety_coeff": "슬링벨트 안전계수 7 이상 확인 (사용하중 2ton / 최대 0.85ton)",
    "work_commander": "홍길동 (제39조)",
    "signal_worker": "김신호 (무전기 사용, 제40조)",
    "worker_roles": "인양 작업자 2명 (상부 유도·체결), 하부 출입통제 1명",
    "fall_prevention": "작업자 안전대 착용 (D링 고리 구조체 체결), 개구부 덮개 설치",
    "drop_prevention": "인양경로 하부 낙하물 방지망 설치, 출입통제 바리케이드",
    "tipping_prevention": "장비 아웃트리거 완전 설치, 정격하중 이내 인양",
    "pinch_prevention": "인양물 이동경로 내 작업자 접근 금지, 신호수 통제",
    "collapse_prevention": "슬링 각도 60° 유지, 안전계수 7 이상 달기구 사용",
    "safety_steps": [
        {"task_step": "작업 전 준비", "hazard": "달기구 불량", "safety_measure": "슬링벨트·샤클 손상 여부 점검"},
        {"task_step": "줄걸이 체결", "hazard": "줄걸이 이탈", "safety_measure": "2줄걸이 체결 확인, 안전핀 고정"},
        {"task_step": "인양 시작", "hazard": "하중 쏠림", "safety_measure": "서행 인양, 수직 확인 후 권상"},
        {"task_step": "수평 이동", "hazard": "하부 협착", "safety_measure": "신호수 유도, 이동경로 통제"},
        {"task_step": "설치 위치 하강", "hazard": "충돌·협착", "safety_measure": "서행 하강, 안내 로프 사용"},
        {"task_step": "설치 및 체결", "hazard": "추락", "safety_measure": "안전대 착용 후 볼트 체결"},
        {"task_step": "달기구 해체", "hazard": "낙하", "safety_measure": "충분한 여유 확인 후 해체"},
    ],
    "pre_work_check_items": (
        "□ 중량물 중량·크기 확인 (설계도서 대조)\n"
        "□ 달기기구(슬링벨트·샤클) 이상 유무 점검\n"
        "□ 양중기 정격하중 확인 (0.85ton < 8ton @ 5m)\n"
        "□ 지반 상태 및 아웃트리거 설치 확인\n"
        "□ 신호수·작업지휘자 배치 확인\n"
        "□ 인양경로 내 하부 출입통제 확인\n"
        "□ TBM 실시 완료\n"
        "□ PTW-007 중량물 인양 작업허가서 발급 확인"
    ),
    "emergency_measure": "인양 중 이상 발생 시 즉시 작업 중지, 안전한 위치에 하역 후 관리감독자 보고. 부상자 발생 시 119 즉시 신고.",
    "photo_items": (
        "① 중량물 외관 및 표시 (중량·크기)\n"
        "② 달기기구 체결 상태 (줄걸이·샤클)\n"
        "③ 양중기 배치 및 아웃트리거 설치\n"
        "④ 인양 중 경로 하부 통제 확인\n"
        "⑤ 설치 완료 후 최종 상태"
    ),
    "sign_date": "2026-04-25",
}

# WP-005 필수 섹션 제목
WP005_REQUIRED_HEADINGS = [
    "중량물 취급 작업계획서",
    "기본 정보",
    "작업 개요 및 중량물 정보",
    "운반·인양 경로",
    "작업장소·지반 상태 및 사용 장비",
    "줄걸이·달기기구",
    "작업 인원 및 역할",
    "위험요인 및 방지대책",
    "작업 전 점검사항",
    "비상조치 계획",
    "사진·증빙 권장 항목",
    "확인 및 서명",
]

# WP-005 필수 포함 문구
WP005_REQUIRED_PHRASES = [
    "PTW-007 중량물 인양·중장비사용 작업 허가서의 선행 서류",
    "이 서식은 법정 안전보건교육 수료증을 대체하지 않습니다",
]

# WP-005 필수 키워드
WP005_REQUIRED_KEYWORDS = [
    "줄걸이",
    "달기기구",
    "작업지휘자",
    "신호수",
    "정격하중",
    "비상조치",
]

# WP-005 evidence ID 목록
WP005_EXPECTED_EVIDENCE_IDS = ["WP-005-L1", "WP-005-L2"]

# WP-005 evidence 파일명
WP005_EXPECTED_EVIDENCE_FILES = [
    "WP-005-L1_safety_rule_article_38_table4_heavy_lifting_workplan.json",
    "WP-005-L2_safety_rule_article_39_work_commander.json",
]

WP005_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"}


def run_wp005_smoke_test() -> list[tuple[str, str, str]]:
    """WP-005 중량물 취급 작업계획서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "heavy_lifting_workplan" in supported,
        "registry: heavy_lifting_workplan 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("heavy_lifting_workplan")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "중량물 취급 작업계획서",
            "display_name == '중량물 취급 작업계획서'",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "safety_steps",
            "repeat_field == 'safety_steps'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("heavy_lifting_workplan", SAMPLE_WP005_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("heavy_lifting_workplan", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("heavy_lifting_workplan", SAMPLE_WP005_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트명 확인
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("중량물" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))

        # 섹션 제목 포함
        for heading in WP005_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 문구 포함
        for phrase in WP005_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:60]}'",
            ))

        # 필수 키워드 포함
        for kw in WP005_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 개인정보 과다 필드 없음
        for keyword in FORBIDDEN_PII_KEYWORDS:
            found = keyword in all_text
            results.append(_check(
                not found,
                f"개인정보 과다 필드 없음: '{keyword}'",
                "FAIL — 기본 필드명에 해당 문구 발견" if found else "",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8~11. WP-005 catalog + evidence 검증 ─────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        wp005 = next(
            (d for d in cat["documents"] if d["id"] == "WP-005"), None
        )

        results.append(_check(wp005 is not None, "catalog: WP-005 항목 존재"))

        if wp005:
            results.append(_check(
                wp005.get("implementation_status") == "DONE",
                "catalog: WP-005 implementation_status == DONE",
                repr(wp005.get("implementation_status")),
            ))
            results.append(_check(
                wp005.get("form_type") == "heavy_lifting_workplan",
                "catalog: WP-005 form_type == 'heavy_lifting_workplan'",
                repr(wp005.get("form_type")),
            ))
            ev_status = wp005.get("evidence_status", "")
            results.append(_check(
                ev_status in WP005_VALID_EV_STATUSES,
                f"catalog: WP-005 evidence_status in {WP005_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 2개 포함
            ev_ids = wp005.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in WP005_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in WP005_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

            # PTW-007 선행관계 확인 (related_documents에 PTW-007)
            related = wp005.get("related_documents") or []
            results.append(_check(
                "PTW-007" in related,
                "catalog: WP-005 related_documents에 PTW-007 포함",
                repr(related),
            ))
            # CL-003 병행 확인 연계
            results.append(_check(
                "CL-003" in related,
                "catalog: WP-005 related_documents에 CL-003 포함",
                repr(related),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "WP-005 catalog/evidence 검증 중 예외", tb))

    # ── evidence 파일 내용 검증 ──────────────────────────────────────────
    for efname in WP005_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:65]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:65]}...", str(exc)))

    return results


def run_ptw007_smoke_test() -> list[tuple[str, str, str]]:
    """PTW-007 중량물 인양·중장비사용 작업 허가서 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "lifting_work_permit" in supported,
        "registry: lifting_work_permit 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("lifting_work_permit")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "중량물 인양·중장비사용 작업 허가서",
            "display_name == '중량물 인양·중장비사용 작업 허가서'",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "workers",
            "repeat_field == 'workers'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("lifting_work_permit", SAMPLE_PTW007_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("lifting_work_permit", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~. 전체 샘플 workbook 검증 ──────────────────────────────────────
    try:
        full_bytes = build_form_excel("lifting_work_permit", SAMPLE_PTW007_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 시트 제목에 "중량물" 포함
        sheet_names = wb_full.sheetnames
        results.append(_check(
            any("중량물" in n for n in sheet_names),
            f"시트명 확인 — {sheet_names}",
        ))
        results.append(_check(
            "중량물 인양·중장비사용 작업 허가서" in all_text,
            "시트 제목 '중량물 인양·중장비사용 작업 허가서' 포함",
        ))

        # 섹션 제목 14개 포함
        for heading in PTW007_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in PTW007_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 필수 고정 문구 포함
        for phrase in PTW007_REQUIRED_PHRASES:
            results.append(_check(
                phrase in all_text,
                f"고정 문구 포함: '{phrase[:60]}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── catalog + evidence 검증 ──────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        docs = {d["id"]: d for d in cat["documents"]}
        ptw007 = docs.get("PTW-007")

        results.append(_check(ptw007 is not None, "catalog: PTW-007 항목 존재"))

        if ptw007:
            results.append(_check(
                ptw007.get("implementation_status") == "DONE",
                "catalog: PTW-007 implementation_status == DONE",
                repr(ptw007.get("implementation_status")),
            ))
            results.append(_check(
                ptw007.get("form_type") == "lifting_work_permit",
                "catalog: PTW-007 form_type == 'lifting_work_permit'",
                repr(ptw007.get("form_type")),
            ))
            ev_status = ptw007.get("evidence_status", "")
            results.append(_check(
                ev_status in PTW007_VALID_EV_STATUSES,
                f"catalog: PTW-007 evidence_status in {PTW007_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # evidence_id 9개 포함
            ev_ids = ptw007.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in PTW007_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in PTW007_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:70]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # PTW-007 related_documents 확인 (WP-005, CL-003 포함)
        if ptw007:
            ptw007_related = ptw007.get("related_documents") or []
            results.append(_check(
                "WP-005" in ptw007_related,
                "catalog: PTW-007 related_documents에 WP-005 포함",
                repr(ptw007_related),
            ))
            results.append(_check(
                "CL-003" in ptw007_related,
                "catalog: PTW-007 related_documents에 CL-003 포함",
                repr(ptw007_related),
            ))

        # 불변 항목 확인 (RA-001, RA-004, PTW-002, PTW-003, CL-003, WP-005 상태 변경 없음)
        for doc_id, (expected_form_type, expected_status) in PTW007_UNCHANGED_IDS.items():
            d = docs.get(doc_id)
            if d is None:
                continue
            if expected_form_type is not None:
                results.append(_check(
                    d.get("form_type") == expected_form_type,
                    f"catalog: {doc_id} form_type 변경 없음 (== '{expected_form_type}')",
                    repr(d.get("form_type")),
                ))
            results.append(_check(
                d.get("implementation_status") == expected_status,
                f"catalog: {doc_id} implementation_status 변경 없음 (== {expected_status})",
                repr(d.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "PTW-007 catalog/evidence 검증 중 예외", tb))

    # evidence 파일 내용 검증
    for efname in PTW007_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:65]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:65]}...", str(exc)))

    return results


# ===========================================================================
# CL-007 — 추락 방호 설비 점검표
# ===========================================================================

SAMPLE_CL007_MINIMAL: dict = {
    "check_date": "2026-04-25",
    "work_location": "5층 외벽 단부",
    "checker_name": "홍길동",
}

SAMPLE_CL007_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "check_date": "2026-04-25",
    "work_location": "5층 외벽 단부 (EL+15.0m)",
    "checker_name": "홍길동",
    "supervisor_name": "이감독",
    "work_name": "외벽 마감 작업",
    "work_height": "15.0m",
    "work_period": "2026-04-25 ~ 2026-05-10",
    "ptw_no": "PTW-003-2026-001",
    "hazard_zone_items": [
        {"item": "개구부 위험구역 식별 및 방호조치 여부", "result": "○", "note": ""},
        {"item": "단부(끝단) 안전난간 또는 방호울 설치 여부", "result": "○", "note": ""},
    ],
    "platform_items": [
        {"item": "작업발판 폭 기준 충족 여부 (강관비계 40cm 이상)", "result": "○", "note": ""},
    ],
    "railing_items": [
        {"item": "상부 안전난간대 높이 기준 충족 여부", "result": "○", "note": ""},
    ],
    "harness_items": [
        {"item": "안전대 착용 여부 (2m 이상 전 작업자)", "result": "○", "note": "전원 착용"},
    ],
    "nonconformance_items": [],
    "work_verdict": "적합",
    "inspector_sign": "홍길동",
    "supervisor_sign": "이감독",
    "manager_sign": "최소장",
    "sign_date": "2026-04-25",
}

# CL-007 필수 섹션 제목
CL007_REQUIRED_HEADINGS = [
    "추락 방호 설비 점검표",
    "① 현장 기본정보",
    "② 점검 대상 작업 정보",
    "③ 추락 위험 구역 확인",
    "④ 작업발판 상태 점검",
    "⑤ 안전난간 점검",
    "⑥ 개구부 덮개 및 방호조치",
    "⑦ 안전대 및 부착설비 점검",
    "⑧ 추락방망 및 낙하 방지 설비",
    "⑨ 사다리·이동식 비계·고소작업대 추가 확인",
    "⑩ 부적합 사항 및 시정조치",
    "⑪ 작업 가능 여부 판정",
    "⑫ 확인 서명",
]

# CL-007 필수 키워드 (연계 검증)
CL007_REQUIRED_KEYWORDS = [
    "추락 방호 설비",
    "안전난간",
    "개구부",
    "안전대",
    "작업중지",
    "확인 서명",
]

# CL-007 evidence
CL007_EXPECTED_EVIDENCE_IDS = ["CL-007-L1", "CL-007-L2"]
CL007_EXPECTED_EVIDENCE_FILES = [
    "CL-007-L1_safety_rule_articles_42_45_fall_prevention.json",
    "CL-007-L2_safety_rule_article_13_safety_railing.json",
]
CL007_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_cl007_smoke_test() -> list[tuple[str, str, str]]:
    """CL-007 추락 방호 설비 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "fall_protection_checklist" in supported,
        "registry: fall_protection_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("fall_protection_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "추락 방호 설비 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("fall_protection_checklist", SAMPLE_CL007_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("fall_protection_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("fall_protection_checklist", SAMPLE_CL007_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 13개 포함 확인
        for heading in CL007_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함 확인
        for keyword in CL007_REQUIRED_KEYWORDS:
            results.append(_check(
                keyword in all_text,
                f"필수 키워드 포함: '{keyword}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8. CL-007 catalog 검증 ────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        docs = {d["id"]: d for d in cat["documents"]}
        cl007 = docs.get("CL-007")
        ptw003 = docs.get("PTW-003")
        cl001  = docs.get("CL-001")
        ra001  = docs.get("RA-001")
        ra004  = docs.get("RA-004")

        results.append(_check(cl007 is not None, "catalog: CL-007 항목 존재"))

        if cl007:
            results.append(_check(
                cl007.get("implementation_status") == "DONE",
                "catalog: CL-007 implementation_status == DONE",
                repr(cl007.get("implementation_status")),
            ))
            results.append(_check(
                cl007.get("form_type") == "fall_protection_checklist",
                "catalog: CL-007 form_type == 'fall_protection_checklist'",
                repr(cl007.get("form_type")),
            ))
            ev_status = cl007.get("evidence_status", "")
            results.append(_check(
                ev_status in CL007_VALID_EV_STATUSES,
                f"catalog: CL-007 evidence_status in {CL007_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # related_documents 확인 (PTW-003, RA-001, RA-004, CL-001)
            rel_docs = cl007.get("related_documents") or []
            for req_id in ("PTW-003", "RA-001", "RA-004", "CL-001"):
                results.append(_check(
                    req_id in rel_docs,
                    f"catalog: CL-007 related_documents에 {req_id} 포함",
                    repr(rel_docs) if req_id not in rel_docs else "",
                ))

            # evidence_id 포함 확인
            ev_ids = cl007.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in CL007_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in CL007_EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname[:70]}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # PTW-003 related_documents에 CL-007 포함 확인
        if ptw003:
            ptw003_rel = ptw003.get("related_documents") or []
            results.append(_check(
                "CL-007" in ptw003_rel,
                "catalog: PTW-003 related_documents에 CL-007 포함",
                repr(ptw003_rel) if "CL-007" not in ptw003_rel else "",
            ))

        # CL-001, PTW-003, RA-001, RA-004 상태 변경 없음 확인
        unchanged = {
            "CL-001":  ("scaffold_installation_checklist", "DONE"),
            "PTW-003": ("work_at_height_permit",            "DONE"),
        }
        for doc_id, (expected_ft, expected_st) in unchanged.items():
            d = docs.get(doc_id)
            if d is None:
                continue
            results.append(_check(
                d.get("form_type") == expected_ft,
                f"catalog: {doc_id} form_type 변경 없음",
                repr(d.get("form_type")),
            ))
            results.append(_check(
                d.get("implementation_status") == expected_st,
                f"catalog: {doc_id} implementation_status 변경 없음",
                repr(d.get("implementation_status")),
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-007 catalog/evidence 검증 중 예외", tb))

    # ── 9. evidence 파일 내용 검증 ────────────────────────────────────────
    for efname in CL007_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "NEEDS_VERIFICATION"),
                f"evidence 검증결과 유효값: {efname[:65]}... → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:65]}...", str(exc)))

    return results


def run_smoke_test() -> None:
    results: list[tuple[str, str, str]] = []
    overall = "PASS"

    supported = {f["form_type"] for f in list_supported_forms()}

    # ════════════════════════════════════════════════════════════
    # ED-003 검증
    # ════════════════════════════════════════════════════════════

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "special_education_log" in supported,
        "registry: special_education_log 등록됨",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    try:
        spec = get_form_spec("special_education_log")
        results.append(_check(isinstance(spec, dict),
                               "get_form_spec() → dict"))
        results.append(_check(spec.get("display_name") == "특별 안전보건교육 교육일지",
                               "display_name 확인",
                               repr(spec.get("display_name"))))
        results.append(_check(isinstance(spec.get("required_fields"), list)
                               and len(spec["required_fields"]) > 0,
                               "required_fields 비어있지 않음"))
        results.append(_check(spec.get("repeat_field") == "attendees",
                               "repeat_field == 'attendees'"))
        results.append(_check(spec.get("max_repeat_rows") == 30,
                               "max_repeat_rows == 30"))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("special_education_log", SAMPLE_ED003_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("special_education_log", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~8. 전체 샘플 workbook 검증 ────────────────────────────────────
    try:
        full_bytes = build_form_excel("special_education_log", SAMPLE_ED003_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = _all_cell_values(wb_full)
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션 제목 포함: '{heading}'",
            ))

        # 개인정보 과다 필드 없음
        for keyword in FORBIDDEN_PII_KEYWORDS:
            found = keyword in all_text
            results.append(_check(
                not found,
                f"개인정보 과다 필드 없음: '{keyword}'",
                "FAIL — 기본 필드명에 해당 문구 발견" if found else "",
            ))

        # 별표 4/5 기준 문구 포함 (VERIFIED 상태)
        for kw in REQUIRED_BYEOLPYO_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"별표 기준 문구 포함: '{kw}'",
            ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 9~15. ED-003 catalog + evidence 검증 ────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        ed003 = next(
            (d for d in cat["documents"] if d["id"] == "ED-003"), None
        )

        # 9. catalog 존재
        results.append(_check(ed003 is not None, "catalog: ED-003 항목 존재"))

        if ed003:
            # 10. implementation_status == DONE
            results.append(_check(
                ed003.get("implementation_status") == "DONE",
                "catalog: ED-003 implementation_status == DONE",
                repr(ed003.get("implementation_status")),
            ))
            # form_type 확인
            results.append(_check(
                ed003.get("form_type") == "special_education_log",
                "catalog: ED-003 form_type == 'special_education_log'",
                repr(ed003.get("form_type")),
            ))
            # 11. evidence_status VERIFIED 또는 PARTIAL_VERIFIED
            ev_status = ed003.get("evidence_status", "")
            results.append(_check(
                ev_status in VALID_EVIDENCE_STATUSES,
                f"catalog: ED-003 evidence_status in {VALID_EVIDENCE_STATUSES}",
                repr(ev_status),
            ))

            # 12. evidence_id 4개 포함
            ev_ids = ed003.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # 13. evidence_file 실제 존재 확인
            ev_files = ed003.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            for efname in EXPECTED_EVIDENCE_FILES:
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

        # 14~15. 각 evidence 파일 내용 검증
        for efname in EXPECTED_EVIDENCE_FILES:
            fpath = EVIDENCE_DIR / efname
            if not fpath.exists():
                continue
            try:
                ev_data = json.loads(fpath.read_text(encoding="utf-8"))
                vr = ev_data.get("verification_result", "")
                results.append(_check(
                    vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                    f"evidence 검증결과: {efname[:30]}... → {vr}",
                ))
                # 별표 4: special_education_hours 구조 확인
                if "table_4" in efname:
                    hours_data = ev_data.get("special_education_hours", {})
                    results.append(_check(
                        bool(hours_data.get("일반근로자")),
                        "별표 4: 일반근로자 특별교육 시간 구조화",
                        repr(hours_data.get("일반근로자", ""))[:60],
                    ))
                # 별표 5: range_phrase 확인
                if "table_5" in efname:
                    works_data = ev_data.get("special_education_target_works", {})
                    results.append(_check(
                        "제1호부터 제39호" in works_data.get("range_phrase", ""),
                        "별표 5: 대상 작업 범위 '제1호부터 제39호까지' 확인",
                        repr(works_data.get("range_phrase", ""))[:60],
                    ))
            except Exception as exc:
                results.append(_check(False, f"evidence 파일 로딩: {efname[:30]}...", str(exc)))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "catalog/evidence 검증 중 예외", tb))

    # ════════════════════════════════════════════════════════════
    # WP-011 검증
    # ════════════════════════════════════════════════════════════
    wp011_results = run_wp011_smoke_test()
    results.extend(wp011_results)

    # ════════════════════════════════════════════════════════════
    # WP-015 검증
    # ════════════════════════════════════════════════════════════
    wp015_results = run_wp015_smoke_test()
    results.extend(wp015_results)

    # ════════════════════════════════════════════════════════════
    # ED-004 검증
    # ════════════════════════════════════════════════════════════
    ed004_results = run_ed004_smoke_test()
    results.extend(ed004_results)

    # ════════════════════════════════════════════════════════════
    # CL-001 검증
    # ════════════════════════════════════════════════════════════
    cl001_results = run_cl001_smoke_test()
    results.extend(cl001_results)

    # ════════════════════════════════════════════════════════════
    # CL-006 검증
    # ════════════════════════════════════════════════════════════
    cl006_results = run_cl006_smoke_test()
    results.extend(cl006_results)

    # ════════════════════════════════════════════════════════════
    # CL-003 검증
    # ════════════════════════════════════════════════════════════
    cl003_results = run_cl003_smoke_test()
    results.extend(cl003_results)

    # ════════════════════════════════════════════════════════════
    # CL-002 검증
    # ════════════════════════════════════════════════════════════
    cl002_results = run_cl002_smoke_test()
    results.extend(cl002_results)

    # ════════════════════════════════════════════════════════════
    # PTW-002 검증
    # ════════════════════════════════════════════════════════════
    ptw002_results = run_ptw002_smoke_test()
    results.extend(ptw002_results)

    # ════════════════════════════════════════════════════════════
    # PTW-003 검증
    # ════════════════════════════════════════════════════════════
    ptw003_results = run_ptw003_smoke_test()
    results.extend(ptw003_results)

    # ════════════════════════════════════════════════════════════
    # PTW-007 검증
    # ════════════════════════════════════════════════════════════
    ptw007_results = run_ptw007_smoke_test()
    results.extend(ptw007_results)

    # ════════════════════════════════════════════════════════════
    # WP-005 검증
    # ════════════════════════════════════════════════════════════
    wp005_results = run_wp005_smoke_test()
    results.extend(wp005_results)

    # ════════════════════════════════════════════════════════════
    # CL-007 검증
    # ════════════════════════════════════════════════════════════
    cl007_results = run_cl007_smoke_test()
    results.extend(cl007_results)

    # ════════════════════════════════════════════════════════════
    # PTW-004 검증
    # ════════════════════════════════════════════════════════════
    ptw004_results = run_ptw004_smoke_test()
    results.extend(ptw004_results)

    # ════════════════════════════════════════════════════════════
    # CL-004 검증
    # ════════════════════════════════════════════════════════════
    cl004_results = run_cl004_smoke_test()
    results.extend(cl004_results)

    # ════════════════════════════════════════════════════════════
    # CL-005 검증
    # ════════════════════════════════════════════════════════════
    cl005_results = run_cl005_smoke_test()
    results.extend(cl005_results)

    # ════════════════════════════════════════════════════════════
    # HM-001 검증 (TEST_MISSING 해소 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    hm001_results = run_hm001_smoke_test()
    results.extend(hm001_results)

    # ════════════════════════════════════════════════════════════
    # HM-002 검증 (TEST_MISSING 해소 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    hm002_results = run_hm002_smoke_test()
    results.extend(hm002_results)

    # ════════════════════════════════════════════════════════════
    # 차량계 묶음 4건 검증 (WP-008/WP-009/EQ-001/EQ-002 — 중간 수준 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    vehicle_results = run_vehicle_workplan_smoke_test()
    results.extend(vehicle_results)

    # ════════════════════════════════════════════════════════════
    # 크레인 묶음 4건 검증 (WP-006/WP-007/EQ-003/EQ-004 — 중간 수준 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    crane_results = run_crane_workplan_smoke_test()
    results.extend(crane_results)

    # ════════════════════════════════════════════════════════════
    # RA 묶음 2건 검증 (RA-001/RA-004 — 중간 수준 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    ra_results = run_ra_forms_smoke_test()
    results.extend(ra_results)

    # ════════════════════════════════════════════════════════════
    # 밀폐공간 묶음 3건 검증 (WP-014/PTW-001/CL-010 — 중간 수준 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    confined_results = run_confined_space_forms_smoke_test()
    results.extend(confined_results)

    # ════════════════════════════════════════════════════════════
    # WP-001 굴착 작업계획서 검증 (중간 수준 — 2026-04-26)
    # ════════════════════════════════════════════════════════════
    wp001_results = run_wp001_smoke_test()
    results.extend(wp001_results)

    # ── 출력 ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 80)
    print("  KRAS P1 Smoke Test — ED-003 + WP-011 + WP-015 + ED-004 + CL-001 + CL-006 + CL-003 + CL-002 + PTW-002 + PTW-003 + PTW-007 + WP-005 + CL-007 + PTW-004 + CL-004 + CL-005 + HM-001 + HM-002 + WP-008/WP-009/EQ-001/EQ-002 + WP-006/WP-007/EQ-003/EQ-004 + RA-001/RA-004 + WP-014/PTW-001/CL-010 + WP-001")
    print("=" * 80)

    pass_cnt = warn_cnt = fail_cnt = 0
    for verdict, name, detail in results:
        icon = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌"}[verdict]
        line = f"  {icon} [{verdict}] {name}"
        if detail:
            line += f" — {detail}"
        print(line)
        if verdict == "PASS":   pass_cnt += 1
        elif verdict == "WARN": warn_cnt += 1
        else:                   fail_cnt += 1; overall = "FAIL"

    print("-" * 62)
    total = len(results)
    print(f"  합계: PASS {pass_cnt}/{total}  WARN {warn_cnt}  FAIL {fail_cnt}")
    print(f"\n  최종 판정: {overall}")
    print("=" * 62 + "\n")

    sys.exit(0 if overall != "FAIL" else 1)


# ===========================================================================
# PTW-004 — 전기작업 허가서 / LOTO
# ===========================================================================

SAMPLE_PTW004_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "work_date": "2026-04-25",
    "work_location": "3층 분전반실",
    "work_supervisor": "홍길동",
}

SAMPLE_PTW004_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "permit_no": "PTW-004-2026-001",
    "work_name": "분전반 차단기 교체",
    "work_date": "2026-04-25",
    "work_time": "2026-04-25 09:00 ~ 12:00",
    "work_location": "3층 분전반실 (EL+9.0m)",
    "voltage": "220V / 380V",
    "work_category": "정전작업",
    "contractor": "협력업체 전기팀",
    "work_supervisor": "홍길동",
    "permit_issuer": "김안전",
    "supervisor_name": "이감독",
    "safety_manager": "최안전",
    "loto_scope": "3층 분전반 ACB-01 이하 회로 전체",
    "loto_breaker_location": "3층 분전반실 주 차단기반",
    "loto_key_holder": "홍길동 (작업책임자)",
    "voltage_tester_used": "확인 완료",
    "voltage_zero_confirmed": "무전압 확인 완료",
    "residual_voltage_result": "방전 완료",
    "ground_confirmed": "접지 완료",
    "voltage_measurer_sign": "홍길동",
    "reenergize_approver": "김안전",
    "work_end_time": "12:00",
    "work_end_confirmer": "이감독",
    "during_work_issues": "이상 없음",
    "final_sign": "김안전",
    "work_types": ["정전작업", "분전반 작업"],
    "prereq_checks": [
        "RA-001 위험성평가표 — 전기감전·아크·화재 위험 포함 여부",
        "RA-004 TBM 일지 — 당일 전기작업 위험 공유 여부",
        "WP-011 전기 작업계획서 — 작업범위·전압·차단방법 기재 여부",
    ],
    "loto_checks": [
        "차단 대상 전로 확인 (제319조)",
        "차단기 차단 여부 확인",
        "잠금장치(Lockout) 설치 여부 확인",
        "표지(Tagout) 부착 여부 확인",
        "담당자 지정 여부 확인",
        "열쇠 보관자 지정 여부 확인",
        "재투입 방지 조치 여부 확인",
    ],
    "voltage_zero_checks": [
        "검전기로 무전압 확인 (제319조 제1항)",
        "잔류전압 방전 완료 확인",
        "접지 완료 확인 (제319조 제1항)",
        "측정자 서명 확인",
    ],
    "ppe_checks": [
        "절연장갑 지급·착용 확인",
        "절연화 지급·착용 확인",
        "보안면(아크 방호) 지급·착용 확인",
        "검전기 구비 및 교정 확인",
    ],
    "zone_control_checks": [
        "출입금지 표지 설치 확인",
        "방책·콘 설치 확인",
        "감시자 배치 확인",
        "비상연락체계 확인",
    ],
    "completion_checks": [
        "작업 완료 여부 확인",
        "인원 전원 철수 확인",
        "공구·장비 회수 확인",
        "잠금장치 해제 승인 완료",
        "재통전 전 감전위험 구역 최종 확인",
        "재통전 승인자 서명 완료",
    ],
    "workers": [
        {"name": "홍길동", "job_type": "전기공 (작업책임자)"},
        {"name": "김전기", "job_type": "전기공"},
    ],
    "nonconformance_items": [],
}

# PTW-004 필수 섹션 제목/문구
PTW004_REQUIRED_HEADINGS = [
    "전기작업 허가서",
    "LOTO",
    "1. 현장 기본정보",
    "2. 허가 대상 작업 정보",
    "3. 전기작업 유형",
    "4. 선행서류 확인",
    "5. 전로 차단 및 LOTO 확인",
    "6. 무전압·잔류전압 확인",
    "7. 활선·근접작업 허가 조건",
    "8. 보호구 및 장비 확인",
    "9. 작업구역 통제",
    "10. 작업 중 변경/작업중지 조건",
    "11. 작업 완료 및 재통전 승인",
    "12. 부적합 및 시정조치",
    "13. 허가 / 승인 / 확인 서명",
]

# PTW-004 필수 키워드
PTW004_REQUIRED_KEYWORDS = [
    "전기작업 허가서",
    "LOTO",
    "잠금장치",
    "전로 차단",
    "무전압",
    "재통전 승인",
    "작업중지",
    "확인 서명",
]

# PTW-004 evidence
PTW004_EXPECTED_EVIDENCE_IDS = ["PTW-004-L1", "PTW-004-L2", "PTW-004-P1"]
PTW004_EXPECTED_EVIDENCE_FILES = [
    "PTW-004-L1_elec_pack_loto_deenergized_work.json",
    "PTW-004-L2_elec_pack_live_work_protection.json",
    "PTW-004-P1_practical_electrical_permit.json",
]
PTW004_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"}

# related_documents 필수 포함 목록
PTW004_REQUIRED_REL_DOCS = ["RA-001", "RA-004", "WP-011", "CL-004"]


def run_ptw004_smoke_test() -> list[tuple[str, str, str]]:
    """PTW-004 전기작업 허가서 / LOTO smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "electrical_work_permit" in supported,
        "registry: electrical_work_permit 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("electrical_work_permit")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "전기작업 허가서 / LOTO",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "nonconformance_items",
            "repeat_field == 'nonconformance_items'",
            repr(spec.get("repeat_field")),
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("electrical_work_permit", SAMPLE_PTW004_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("electrical_work_permit", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        full_bytes = build_form_excel("electrical_work_permit", SAMPLE_PTW004_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in PTW004_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션/제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in PTW004_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8~12. PTW-004 catalog 검증 ───────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        ptw004 = next(
            (d for d in cat["documents"] if d["id"] == "PTW-004"), None
        )

        results.append(_check(ptw004 is not None, "catalog: PTW-004 항목 존재"))

        if ptw004:
            results.append(_check(
                ptw004.get("implementation_status") == "DONE",
                "catalog: PTW-004 implementation_status == DONE",
                repr(ptw004.get("implementation_status")),
            ))
            results.append(_check(
                ptw004.get("form_type") == "electrical_work_permit",
                "catalog: PTW-004 form_type == 'electrical_work_permit'",
                repr(ptw004.get("form_type")),
            ))
            ev_status = ptw004.get("evidence_status", "")
            results.append(_check(
                ev_status in PTW004_VALID_EV_STATUSES,
                f"catalog: PTW-004 evidence_status in {PTW004_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            # related_documents 필수 포함
            rel_docs = ptw004.get("related_documents") or []
            for req_doc in PTW004_REQUIRED_REL_DOCS:
                results.append(_check(
                    req_doc in rel_docs,
                    f"catalog: PTW-004 related_documents 포함 — {req_doc}",
                ))

            # evidence_id 포함
            ev_ids = ptw004.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in PTW004_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 실제 존재 확인
            for efname in PTW004_EXPECTED_EVIDENCE_FILES:
                fpath = Path("data/evidence/safety_law_refs") / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "PTW-004 catalog/evidence 검증 중 예외", tb))

    # ── 13. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in PTW004_EXPECTED_EVIDENCE_FILES:
        fpath = Path("data/evidence/safety_law_refs") / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"),
                f"evidence 검증결과: {efname[:45]} → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:45]}", str(exc)))

    return results


# ===========================================================================
# CL-004 — 전기설비 정기 점검표
# ===========================================================================

SAMPLE_CL004_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "inspection_date": "2026-04-25",
    "inspector": "홍길동",
}

SAMPLE_CL004_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "inspection_date": "2026-04-25",
    "inspection_no": "CL-004-2026-001",
    "equipment_name": "3층 분전반 (MCC-3F)",
    "voltage": "220V / 380V",
    "inspection_location": "3층 전기실",
    "responsible_person": "전기팀장 김길동",
    "inspector": "홍길동",
    "related_wp_no": "WP-011-2026-001",
    "related_ptw_no": "PTW-004-2026-001",
    "panel_checks": [
        "panel_lock",
        "panel_nameplate",
        "overcurrent_breaker",
        "elcb_installed",
        "elcb_test_button",
    ],
    "grounding_checks": [
        "ground_connection",
        "ground_terminal",
        "ground_resistance",
        "no_ground_missing",
    ],
    "wiring_checks": [
        "cable_sheath_ok",
        "no_wet_risk",
        "crossing_protect",
        "no_octopus_wiring",
    ],
    "equipment_checks": [
        "equip_housing_ok",
        "plug_ok",
        "grounded_plug",
        "equip_cable_ok",
        "equip_elcb",
    ],
    "temporary_checks": [
        "lamp_guard",
        "lamp_support_ok",
        "lamp_no_combustible",
    ],
    "hazard_checks": [
        "live_parts_guarded",
        "insulation_ok",
        "no_overheat_trace",
        "no_spark_trace",
        "no_dust_moisture",
    ],
    "ppe_checks": [
        "insulated_gloves",
        "insulated_shoes",
        "voltage_tester",
        "insulated_tools",
    ],
    "nonconformance_items": [
        {
            "content": "3번 차단기 명판 탈락",
            "action": "명판 재부착",
            "deadline": "2026-04-26",
            "completed": "",
        }
    ],
    "verdict": "조건부 적합",
    "verdict_condition": "3번 차단기 명판 재부착 완료 후 재확인",
    "inspector_sign": "홍길동",
    "supervisor_sign": "이관리",
    "safety_manager_sign": "박안전",
}

CL004_REQUIRED_HEADINGS = [
    "전기설비 정기 점검표",
    "1. 현장 기본정보",
    "2. 점검 대상 정보",
    "3. 분전반 및 차단기 점검",
    "4. 접지 상태 점검",
    "5. 배선 및 이동전선 점검",
    "6. 전기기계기구 및 전동공구 점검",
    "7. 임시전기 및 작업등 점검",
    "8. 감전·화재 위험 점검",
    "9. 보호구 및 측정장비 확인",
    "10. 부적합 사항 및 시정조치",
    "11. 점검 결과 판정",
    "12. 확인 서명",
]

CL004_REQUIRED_KEYWORDS = [
    "전기설비 정기 점검표",
    "누전차단기",
    "접지",
    "분전반",
    "이동전선",
    "사용중지",
    "확인 서명",
]

CL004_EXPECTED_EVIDENCE_IDS = ["CL-004-L1", "CL-004-L2", "CL-004-P1"]
CL004_EXPECTED_EVIDENCE_FILES = [
    "CL-004-L1_elec_pack_grounding_leakage_breaker.json",
    "CL-004-L2_elec_pack_electrical_tools_wiring.json",
    "CL-004-P1_practical_electrical_facility_checklist.json",
]
CL004_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"}

CL004_REQUIRED_REL_DOCS = ["RA-001", "RA-004", "WP-011", "PTW-004"]


def run_cl004_smoke_test() -> list[tuple[str, str, str]]:
    """CL-004 전기설비 정기 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "electrical_facility_checklist" in supported,
        "registry: electrical_facility_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("electrical_facility_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "전기설비 정기 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "nonconformance_items",
            "repeat_field == 'nonconformance_items'",
            repr(spec.get("repeat_field")),
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("electrical_facility_checklist", SAMPLE_CL004_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("electrical_facility_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        full_bytes = build_form_excel("electrical_facility_checklist", SAMPLE_CL004_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        for heading in CL004_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션/제목 포함: '{heading}'",
            ))

        for kw in CL004_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8~12. CL-004 catalog 검증 ────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl004 = next(
            (d for d in cat["documents"] if d["id"] == "CL-004"), None
        )

        results.append(_check(cl004 is not None, "catalog: CL-004 항목 존재"))

        if cl004:
            results.append(_check(
                cl004.get("implementation_status") == "DONE",
                "catalog: CL-004 implementation_status == DONE",
                repr(cl004.get("implementation_status")),
            ))
            results.append(_check(
                cl004.get("form_type") == "electrical_facility_checklist",
                "catalog: CL-004 form_type == 'electrical_facility_checklist'",
                repr(cl004.get("form_type")),
            ))
            ev_status = cl004.get("evidence_status", "")
            results.append(_check(
                ev_status in CL004_VALID_EV_STATUSES,
                f"catalog: CL-004 evidence_status in {CL004_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            rel_docs = cl004.get("related_documents") or []
            for req_doc in CL004_REQUIRED_REL_DOCS:
                results.append(_check(
                    req_doc in rel_docs,
                    f"catalog: CL-004 related_documents 포함 — {req_doc}",
                ))

            ev_ids = cl004.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in CL004_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            for efname in CL004_EXPECTED_EVIDENCE_FILES:
                fpath = Path("data/evidence/safety_law_refs") / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-004 catalog/evidence 검증 중 예외", tb))

    # ── 13. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL004_EXPECTED_EVIDENCE_FILES:
        fpath = Path("data/evidence/safety_law_refs") / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"),
                f"evidence 검증결과: {efname[:50]} → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:50]}", str(exc)))

    return results


# ===========================================================================
# CL-005 — 화재 예방 점검표
# ===========================================================================

SAMPLE_CL005_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "inspection_date": "2026-04-25",
    "inspector": "홍길동",
}

SAMPLE_CL005_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "inspection_no": "CL-005-2026-001",
    "work_name": "3층 철골 용접 작업",
    "work_location": "3층 철골 구조부",
    "work_datetime": "2026-04-25 09:00 ~ 18:00",
    "work_category": "용접",
    "related_ptw_no": "PTW-002-2026-001",
    "work_supervisor": "이순신",
    "inspector": "홍길동",
    "inspection_date": "2026-04-25",
    "fire_work_types": ["fw_welding", "fw_grinding"],
    "combustible_checks": [
        "comb_removed", "comb_protective", "comb_flammable",
        "comb_gas", "comb_ventilation",
    ],
    "spark_checks": [
        "spark_blanket", "spark_receiver", "spark_opening",
        "spark_lower", "spark_duct",
    ],
    "extinguisher_checks": [
        "ext_placed", "ext_hydrant", "ext_water",
        "ext_pressure", "ext_marked",
    ],
    "fire_watch_checks": [
        "fwatch_assigned", "fwatch_position", "fwatch_comm",
        "fwatch_during", "fwatch_after30",
    ],
    "gas_equip_checks": [
        "gas_fixed", "gas_flashback", "gas_hose",
        "gas_regulator", "gas_valve",
    ],
    "elec_fire_checks": [
        "ef_insulation", "ef_elcb", "ef_ground", "ef_overload",
    ],
    "post_work_checks": [
        "pw_ember", "pw_temp", "pw_30min", "pw_waste",
    ],
    "nonconformance_items": [],
    "verdict": "적합",
    "verdict_condition": "",
    "inspector_sign": "홍길동",
    "fire_watch_sign": "강감찬",
    "supervisor_sign": "김감독",
    "safety_manager_sign": "박안전",
}

CL005_REQUIRED_HEADINGS = [
    "화재 예방 점검표",
    "현장 기본정보",
    "점검 대상 작업 정보",
    "화기작업 유형",
    "가연물 제거 확인",
    "불티 비산 방지 조치",
    "소화설비 및 초기진화 준비",
    "화재감시자 배치",
    "가스·용접장비 점검",
    "전기화재 예방",
    "작업 후 화재 확인",
    "부적합 사항 및 시정조치",
    "작업 가능 여부 판정",
    "확인 서명",
]

CL005_REQUIRED_KEYWORDS = [
    "불티",
    "소화기",
    "화재감시자",
    "가연물",
    "작업중지",
]

CL005_EXPECTED_EVIDENCE_IDS = ["CL-005-P1"]
CL005_EXPECTED_EVIDENCE_FILES = [
    "CL-005-P1_fire_prevention_checklist_practical.json",
]
CL005_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"}
CL005_REQUIRED_REL_DOCS = ["RA-001", "RA-004", "PTW-002"]


def run_cl005_smoke_test() -> list[tuple[str, str, str]]:
    """CL-005 화재 예방 점검표 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "fire_prevention_checklist" in supported,
        "registry: fire_prevention_checklist 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("fire_prevention_checklist")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "화재 예방 점검표",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "nonconformance_items",
            "repeat_field == 'nonconformance_items'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 5,
            "max_repeat_rows == 5",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("fire_prevention_checklist", SAMPLE_CL005_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("fire_prevention_checklist", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("fire_prevention_checklist", SAMPLE_CL005_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        from io import BytesIO as _BytesIO
        from openpyxl import load_workbook as _load_wb
        wb_full = _load_wb(_BytesIO(full_bytes))
        all_vals = [
            str(cell.value or "")
            for ws_obj in wb_full.worksheets
            for r in ws_obj.iter_rows()
            for cell in r
        ]
        all_text = " ".join(all_vals)

        # 섹션 제목 포함
        for heading in CL005_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션/제목 포함: '{heading}'",
            ))

        # 필수 키워드 포함
        for kw in CL005_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8~12. CL-005 catalog 검증 ────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        cl005 = next(
            (d for d in cat["documents"] if d["id"] == "CL-005"), None
        )

        results.append(_check(cl005 is not None, "catalog: CL-005 항목 존재"))

        if cl005:
            results.append(_check(
                cl005.get("implementation_status") == "DONE",
                "catalog: CL-005 implementation_status == DONE",
                repr(cl005.get("implementation_status")),
            ))
            results.append(_check(
                cl005.get("form_type") == "fire_prevention_checklist",
                "catalog: CL-005 form_type == 'fire_prevention_checklist'",
                repr(cl005.get("form_type")),
            ))
            ev_status = cl005.get("evidence_status", "")
            results.append(_check(
                ev_status in CL005_VALID_EV_STATUSES,
                f"catalog: CL-005 evidence_status in {CL005_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            rel_docs = cl005.get("related_documents") or []
            for req_doc in CL005_REQUIRED_REL_DOCS:
                results.append(_check(
                    req_doc in rel_docs,
                    f"catalog: CL-005 related_documents 포함 — {req_doc}",
                ))

            ev_ids = cl005.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in CL005_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            for efname in CL005_EXPECTED_EVIDENCE_FILES:
                fpath = Path("data/evidence/safety_law_refs") / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 존재: {efname}",
                    str(fpath) if not fpath.exists() else "",
                ))

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "CL-005 catalog/evidence 검증 중 예외", tb))

    # ── 13. evidence 파일 내용 검증 ───────────────────────────────────────
    for efname in CL005_EXPECTED_EVIDENCE_FILES:
        fpath = Path("data/evidence/safety_law_refs") / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED", "PRACTICAL"),
                f"evidence 검증결과: {efname[:50]} → {vr}",
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:50]}", str(exc)))

    return results


# ===========================================================================
# HM-001 — 작업환경측정 실시 및 결과 관리대장
# 법적 근거: 산업안전보건법 제125조
# evidence : HM-001-L1 (구 G-02 명명에서 2026-04-26 정규화 배치로 표준명 전환)
# ===========================================================================

SAMPLE_HM001_MINIMAL: dict = {
    "target_process": "지하 1층 도장공정",
    "measurement_agency": "(주)한국작업환경측정",
    "measurement_date": "2026-04-25",
    "hazardous_agents": "톨루엔, 크실렌, 분진",
}

SAMPLE_HM001_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "work_location": "지하 1층 도장 부스",
    "measurement_period": "2026-04-25 ~ 2026-04-25",
    "supervisor": "홍길동",
    "contractor": "협력업체 도장",
    "prepared_by": "김안전",
    "target_process": "지하 1층 도장공정 (스프레이 도장)",
    "hazardous_agents": "톨루엔, 크실렌, 메틸이소부틸케톤, 분진",
    "measurement_agency": "(주)한국작업환경측정",
    "agency_contact": "02-0000-0000",
    "measurement_date": "2026-04-25",
    "result_received_date": "2026-05-02",
    "result_summary": "전 항목 노출기준 미만 (분진은 기준의 60% 수준)",
    "exceedance_status": "초과 없음",
    "exceedance_detail": "",
    "improvement_plan": "국소배기장치 풍속 점검 강화, 보호구 착용 재교육",
    "improvement_deadline": "2026-05-31",
    "improvement_done": "예정",
    "worker_notification": "조회 시 구두 설명 및 게시판 게시 (2026-05-03)",
    "original_attached": "첨부 완료 (외부 측정기관 발급 결과보고서 원본)",
    "measurement_rows": [
        {
            "target_location": "도장부스 A",
            "hazardous_agent": "톨루엔",
            "measured_value": "12.3 ppm",
            "exposure_limit": "20 ppm",
            "exceedance": "미초과",
        },
        {
            "target_location": "도장부스 A",
            "hazardous_agent": "크실렌",
            "measured_value": "18.0 ppm",
            "exposure_limit": "100 ppm",
            "exceedance": "미초과",
        },
        {
            "target_location": "혼합실",
            "hazardous_agent": "분진",
            "measured_value": "1.2 mg/㎥",
            "exposure_limit": "2 mg/㎥",
            "exceedance": "미초과",
        },
    ],
    "confirmer_name": "최보건",
    "confirmer_role": "보건관리자",
    "sign_date": "2026-05-03",
}

# HM-001 검증 항목 (사용자 요건):
#   - 문서 제목 / 주요 섹션명 / 법령·근거 표시 / 핵심 점검 항목 / 작성자·확인자·서명란
HM001_REQUIRED_HEADINGS = [
    "작업환경측정 실시 및 결과 관리대장",   # 문서 제목
    "측정 대상 및 측정기관 정보",            # 섹션
    "측정 결과 요약",                         # 섹션
    "측정 결과 상세",                         # 섹션 (full label: "(유해인자별)")
    "개선조치 및 근로자 고지",                # 섹션
    "원본 결과서 첨부 확인",                  # 섹션
    "확인 및 서명",                           # 서명 섹션
]

HM001_REQUIRED_KEYWORDS = [
    "산업안전보건법 제125조",   # 법령·근거 표시 영역
    "유해인자",                  # 핵심 점검 항목
    "측정값",                    # 핵심 점검 항목
    "노출기준",                  # 핵심 점검 항목
    "근로자 고지",               # 핵심 항목 (Rule 8 — 통지 의무)
    "원본",                      # 외부 결과보고서 원본 첨부 안내
]

HM001_REQUIRED_SIGNATURE_LABELS = [
    "작성자 서명",
    "확인자 서명",
    "작성일",
]

HM001_EXPECTED_EVIDENCE_IDS = ["HM-001-L1"]
# 2026-04-26 정규화 배치로 표준명 적용 (구 G-02 → HM-001-L1).
HM001_EXPECTED_EVIDENCE_FILES = [
    "HM-001-L1_industrial_safety_health_act_article_125.json",
]
HM001_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_hm001_smoke_test() -> list[tuple[str, str, str]]:
    """HM-001 작업환경측정 실시 및 결과 관리대장 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "work_environment_measurement" in supported,
        "registry: work_environment_measurement 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("work_environment_measurement")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "작업환경측정 실시 및 결과 관리대장",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "measurement_rows",
            "repeat_field == 'measurement_rows'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 10,
            "max_repeat_rows == 10",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("work_environment_measurement", SAMPLE_HM001_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("work_environment_measurement", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("work_environment_measurement", SAMPLE_HM001_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_text = " ".join(_all_cell_values(wb_full))

        # 문서 제목 + 주요 섹션명
        for heading in HM001_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션/제목 포함: '{heading}'",
            ))

        # 법령/근거 + 핵심 점검 항목
        for kw in HM001_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 작성자/확인자/서명란
        for label in HM001_REQUIRED_SIGNATURE_LABELS:
            results.append(_check(
                label in all_text,
                f"서명란 라벨 포함: '{label}'",
            ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8. catalog 검증 ───────────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        hm001 = next(
            (d for d in cat["documents"] if d["id"] == "HM-001"), None
        )

        results.append(_check(hm001 is not None, "catalog: HM-001 항목 존재"))

        if hm001:
            results.append(_check(
                hm001.get("implementation_status") == "DONE",
                "catalog: HM-001 implementation_status == DONE",
                repr(hm001.get("implementation_status")),
            ))
            results.append(_check(
                hm001.get("form_type") == "work_environment_measurement",
                "catalog: HM-001 form_type == 'work_environment_measurement'",
                repr(hm001.get("form_type")),
            ))
            ev_status = hm001.get("evidence_status", "")
            results.append(_check(
                ev_status in HM001_VALID_EV_STATUSES,
                f"catalog: HM-001 evidence_status in {HM001_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            ev_ids = hm001.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in HM001_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            # evidence_file 은 catalog 에서 full path 로 박혀 있을 수도 있고 basename 만 있을 수도 있음
            ev_files = hm001.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            for efname in HM001_EXPECTED_EVIDENCE_FILES:
                results.append(_check(
                    efname in ev_basenames,
                    f"catalog: evidence_file 등록 — {efname}",
                ))
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 실제 존재: {efname}",
                ))
    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "HM-001 catalog/evidence 검증 중 예외", tb))

    # ── 9. evidence 파일 내용 검증 ────────────────────────────────────────
    for efname in HM001_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                f"evidence 검증결과: {efname[:50]} → {vr}",
            ))
            # 법 제125조 명시 확인
            results.append(_check(
                ev_data.get("article_no") in ("125", "제125조"),
                f"evidence article_no == 제125조 — {efname[:30]}",
                repr(ev_data.get("article_no")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:50]}", str(exc)))

    return results


# ===========================================================================
# HM-002 — 특수건강진단 대상자 및 결과 관리대장
# 법적 근거: 산업안전보건법 제130조
# evidence : HM-002-L1 (구 G-03 명명에서 2026-04-26 정규화 배치로 표준명 전환)
# ===========================================================================

SAMPLE_HM002_MINIMAL: dict = {
    "exam_agency": "(주)한국특수건강진단의원",
    "exam_date": "2026-04-25",
    "exam_type": "정기",
    "hazardous_agents": "톨루엔, 분진",
}

SAMPLE_HM002_FULL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "project_name": "테스트현장 신축공사",
    "exam_target_work": "도장공정 / 분진 발생 작업",
    "exam_period": "2026-04-25 ~ 2026-04-25",
    "supervisor": "최보건",
    "contractor": "협력업체 도장",
    "prepared_by": "김안전",
    "exam_agency": "(주)한국특수건강진단의원",
    "agency_contact": "031-000-0000",
    "exam_date": "2026-04-25",
    "result_received_date": "2026-05-09",
    "exam_type": "정기",
    "hazardous_agents": "톨루엔, 크실렌, 분진",
    "judgment_summary": "A 8명, C1 1명, R 1명 (재검 권고 1명)",
    "followup_plan": "C1·R 대상자 보건관리자 면담 및 작업전환 검토",
    "non_exam_count": "1",
    "non_exam_reason": "병가",
    "non_exam_action": "복귀 후 1주 이내 검진 일정 재안내",
    "original_stored": "보관 완료 (보건관리자실 잠금 캐비닛)",
    "privacy_confirmed": "확인 완료 (열람 권한자 명단 게시)",
    "worker_rows": [
        {
            "employee_no": "20240001",
            "name": "이순신",
            "birth_year": "1985",
            "job_type": "도장공",
            "exam_done": "완료",
            "judgment": "A",
            "followup_needed": "불필요",
        },
        {
            "employee_no": "20240002",
            "name": "강감찬",
            "birth_year": "1990",
            "job_type": "도장공",
            "exam_done": "완료",
            "judgment": "C1",
            "followup_needed": "필요",
        },
    ],
    "confirmer_name": "최보건",
    "confirmer_role": "보건관리자",
    "sign_date": "2026-05-10",
}

HM002_REQUIRED_HEADINGS = [
    "특수건강진단 대상자 및 결과 관리대장",   # 문서 제목
    "검진기관 정보 및 검진 구분",              # 섹션
    "대상 근로자 목록 및 수검 결과",           # 섹션
    "판정 요약 및 사후관리",                   # 섹션
    "원본 보관 확인 및 개인정보 보호",         # 섹션
    "확인 및 서명",                            # 서명 섹션
]

HM002_REQUIRED_KEYWORDS = [
    "산업안전보건법 제130조",   # 법령·근거 표시 영역
    "검진기관",                  # 핵심 항목
    "검진 구분",                 # 핵심 항목 (배치전/정기/수시/임시)
    "유해인자",                  # 핵심 항목
    "판정",                      # 핵심 항목 (A/B/C1 등)
    "사후관리",                  # 핵심 항목
    "원본",                      # 외부 결과서 원본 별도 보관 안내
]

HM002_REQUIRED_SIGNATURE_LABELS = [
    "작성자 서명",
    "확인자 서명",
    "작성일",
]

HM002_EXPECTED_EVIDENCE_IDS = ["HM-002-L1"]
# 2026-04-26 정규화 배치로 표준명 적용 (구 G-03 → HM-002-L1).
HM002_EXPECTED_EVIDENCE_FILES = [
    "HM-002-L1_industrial_safety_health_act_article_130.json",
]
HM002_VALID_EV_STATUSES = {"VERIFIED", "PARTIAL_VERIFIED"}


def run_hm002_smoke_test() -> list[tuple[str, str, str]]:
    """HM-002 특수건강진단 대상자 및 결과 관리대장 smoke test. 결과 list 반환."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "special_health_examination" in supported,
        "registry: special_health_examination 등록됨",
    ))

    # ── 2. get_form_spec 검증 ─────────────────────────────────────────────
    try:
        spec = get_form_spec("special_health_examination")
        results.append(_check(isinstance(spec, dict), "get_form_spec() → dict"))
        results.append(_check(
            spec.get("display_name") == "특수건강진단 대상자 및 결과 관리대장",
            "display_name 확인",
            repr(spec.get("display_name")),
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), list)
            and len(spec["required_fields"]) > 0,
            "required_fields 비어있지 않음",
        ))
        results.append(_check(
            spec.get("repeat_field") == "worker_rows",
            "repeat_field == 'worker_rows'",
            repr(spec.get("repeat_field")),
        ))
        results.append(_check(
            spec.get("max_repeat_rows") == 15,
            "max_repeat_rows == 15",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec() 호출 성공", str(exc)))

    # ── 3. 최소 샘플 → bytes ─────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("special_health_examination", SAMPLE_HM002_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "최소 샘플 → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "최소 샘플 → bytes 생성", str(exc)))

    # ── 4. 공란 form_data → bytes ─────────────────────────────────────────
    try:
        empty_bytes = build_form_excel("special_health_examination", {})
        results.append(_check(
            isinstance(empty_bytes, bytes) and len(empty_bytes) > 0,
            "공란 form_data → bytes 생성 (오류 없음)",
            f"{len(empty_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "공란 form_data → bytes 생성", str(exc)))

    # ── 5~7. 전체 샘플 workbook 검증 ─────────────────────────────────────
    try:
        full_bytes = build_form_excel("special_health_examination", SAMPLE_HM002_FULL)
        results.append(_check(
            isinstance(full_bytes, bytes) and len(full_bytes) > 0,
            "전체 샘플 → bytes 생성",
            f"{len(full_bytes):,} bytes",
        ))
        wb_full = load_workbook(BytesIO(full_bytes))
        all_text = " ".join(_all_cell_values(wb_full))

        # 문서 제목 + 주요 섹션명
        for heading in HM002_REQUIRED_HEADINGS:
            results.append(_check(
                heading in all_text,
                f"섹션/제목 포함: '{heading}'",
            ))

        # 법령/근거 + 핵심 점검 항목
        for kw in HM002_REQUIRED_KEYWORDS:
            results.append(_check(
                kw in all_text,
                f"필수 키워드 포함: '{kw}'",
            ))

        # 작성자/확인자/서명란
        for label in HM002_REQUIRED_SIGNATURE_LABELS:
            results.append(_check(
                label in all_text,
                f"서명란 라벨 포함: '{label}'",
            ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "전체 샘플 처리 중 예외", tb))

    # ── 8. catalog 검증 ───────────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)
        hm002 = next(
            (d for d in cat["documents"] if d["id"] == "HM-002"), None
        )

        results.append(_check(hm002 is not None, "catalog: HM-002 항목 존재"))

        if hm002:
            results.append(_check(
                hm002.get("implementation_status") == "DONE",
                "catalog: HM-002 implementation_status == DONE",
                repr(hm002.get("implementation_status")),
            ))
            results.append(_check(
                hm002.get("form_type") == "special_health_examination",
                "catalog: HM-002 form_type == 'special_health_examination'",
                repr(hm002.get("form_type")),
            ))
            ev_status = hm002.get("evidence_status", "")
            results.append(_check(
                ev_status in HM002_VALID_EV_STATUSES,
                f"catalog: HM-002 evidence_status in {HM002_VALID_EV_STATUSES}",
                repr(ev_status),
            ))

            ev_ids = hm002.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in HM002_EXPECTED_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: evidence_id 포함 — {eid}",
                ))

            ev_files = hm002.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            for efname in HM002_EXPECTED_EVIDENCE_FILES:
                results.append(_check(
                    efname in ev_basenames,
                    f"catalog: evidence_file 등록 — {efname}",
                ))
                fpath = EVIDENCE_DIR / efname
                results.append(_check(
                    fpath.exists(),
                    f"evidence_file 실제 존재: {efname}",
                ))
    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "HM-002 catalog/evidence 검증 중 예외", tb))

    # ── 9. evidence 파일 내용 검증 ────────────────────────────────────────
    for efname in HM002_EXPECTED_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / efname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr in ("VERIFIED", "PARTIAL_VERIFIED"),
                f"evidence 검증결과: {efname[:50]} → {vr}",
            ))
            results.append(_check(
                ev_data.get("article_no") in ("130", "제130조"),
                f"evidence article_no == 제130조 — {efname[:30]}",
                repr(ev_data.get("article_no")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {efname[:50]}", str(exc)))

    return results


# ===========================================================================
# WP-008 / WP-009 / EQ-001 / EQ-002 — 차량계 묶음 (중간 수준 smoke test)
# ===========================================================================
#
# (c) 중간 수준 — full 검증 블록 4세트는 만들지 않는다.
#   1. registry: vehicle_construction_workplan, material_handling_workplan 등록 확인
#   2. get_form_spec 으로 두 form_type 조회 확인
#   3. 4개 document_id 가 smoke 파일 내 명시적으로 참조됨
#   4. vehicle_construction_workplan sample build 1건
#   5. material_handling_workplan sample build 1건
#   6. 4개 catalog 항목 evidence_* 필드 등록 확인
#   7. 신규 evidence 파일 4건 존재 + verification_result 확인
#
# 대상 문서: WP-008, WP-009, EQ-001, EQ-002 (audit doc_id 참조용 명시)
# ===========================================================================

VEHICLE_TARGET_DOC_IDS = ["WP-008", "WP-009", "EQ-001", "EQ-002"]

VEHICLE_EVIDENCE_BY_DOC = {
    "WP-008": (
        "WP-008-L1",
        "WP-008-L1_safety_rule_articles_196_199_vehicle_construction.json",
        "vehicle_construction_workplan",
        "VERIFIED",
    ),
    "WP-009": (
        "WP-009-L1",
        "WP-009-L1_safety_rule_articles_171_182_material_handling.json",
        "material_handling_workplan",
        "PARTIAL_VERIFIED",
    ),
    "EQ-001": (
        "EQ-001-L1",
        "EQ-001-L1_safety_rule_material_handling_equipment_specific.json",
        "material_handling_workplan",
        "PARTIAL_VERIFIED",
    ),
    "EQ-002": (
        "EQ-002-L1",
        "EQ-002-L1_safety_rule_vehicle_construction_equipment_specific.json",
        "vehicle_construction_workplan",
        "VERIFIED",
    ),
}

SAMPLE_VEHICLE_CONSTRUCTION_MINIMAL: dict = {
    "machine_type": "굴착기 (CAT 320D)",
    "machine_capacity": "1.0㎥, 작업반경 9m",
    "work_method": "북측 부지 굴착·상차 — 유도자 1명 배치, 작업반경 내 출입통제",
    "travel_route_text": "정문 → 북측 진입로 → 작업장 (편도 약 80m, 폭 4m)",
}

SAMPLE_MATERIAL_HANDLING_MINIMAL: dict = {
    "machine_type": "지게차 (도요타 8FB25, 2.5톤)",
    "machine_max_load": "2,500kg",
    "work_method": "자재창고 ↔ 야적장 자재 운반 — 운전자 1명, 보행자 동선 분리",
    "travel_route_text": "자재창고 → 본동 1층 → 야적장 (편도 60m, 폭 5m)",
    "emergency_measure": "사고 발생 시 즉시 정지·경광등 점등, 119/현장사무실 동시 신고",
}


def run_vehicle_workplan_smoke_test() -> list[tuple[str, str, str]]:
    """차량계 묶음 4건 (WP-008/WP-009/EQ-001/EQ-002) 중간 수준 smoke test."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "vehicle_construction_workplan" in supported,
        "registry: vehicle_construction_workplan 등록됨 (WP-008/EQ-002)",
    ))
    results.append(_check(
        "material_handling_workplan" in supported,
        "registry: material_handling_workplan 등록됨 (WP-009/EQ-001)",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    for form_type, expected_name in (
        ("vehicle_construction_workplan", "차량계 건설기계 작업계획서"),
        ("material_handling_workplan",   "차량계 하역운반기계 작업계획서"),
    ):
        try:
            spec = get_form_spec(form_type)
            results.append(_check(
                isinstance(spec, dict) and spec.get("display_name") == expected_name,
                f"get_form_spec('{form_type}') → display_name == '{expected_name}'",
                repr(spec.get("display_name")) if isinstance(spec, dict) else "",
            ))
            results.append(_check(
                isinstance(spec.get("required_fields"), (list, tuple))
                and len(spec["required_fields"]) > 0,
                f"{form_type}: required_fields 비어있지 않음",
            ))
        except Exception as exc:
            results.append(_check(False, f"get_form_spec('{form_type}') 호출 성공", str(exc)))

    # ── 3. sample build × 2 ──────────────────────────────────────────────
    for form_type, sample in (
        ("vehicle_construction_workplan", SAMPLE_VEHICLE_CONSTRUCTION_MINIMAL),
        ("material_handling_workplan",   SAMPLE_MATERIAL_HANDLING_MINIMAL),
    ):
        try:
            xlsx_bytes = build_form_excel(form_type, sample)
            results.append(_check(
                isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
                f"{form_type} sample build → bytes 생성",
                f"{len(xlsx_bytes):,} bytes",
            ))
        except Exception as exc:
            results.append(_check(False, f"{form_type} sample build", str(exc)))

    # ── 4. 4개 catalog 항목 검증 + evidence_file 존재 ────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)

        for doc_id in VEHICLE_TARGET_DOC_IDS:
            doc = next((d for d in cat["documents"] if d["id"] == doc_id), None)
            results.append(_check(doc is not None, f"catalog: {doc_id} 항목 존재"))
            if not doc:
                continue

            ev_id, ev_fname, expected_form_type, expected_status = VEHICLE_EVIDENCE_BY_DOC[doc_id]

            results.append(_check(
                doc.get("implementation_status") == "DONE",
                f"catalog: {doc_id} implementation_status == DONE",
                repr(doc.get("implementation_status")),
            ))
            results.append(_check(
                doc.get("form_type") == expected_form_type,
                f"catalog: {doc_id} form_type == '{expected_form_type}'",
                repr(doc.get("form_type")),
            ))

            ev_status = doc.get("evidence_status", "")
            results.append(_check(
                ev_status == expected_status,
                f"catalog: {doc_id} evidence_status == '{expected_status}'",
                repr(ev_status),
            ))

            ev_ids = doc.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            results.append(_check(
                ev_id in ev_ids,
                f"catalog: {doc_id} evidence_id 포함 — {ev_id}",
            ))

            ev_files = doc.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            results.append(_check(
                ev_fname in ev_basenames,
                f"catalog: {doc_id} evidence_file 등록 — {ev_fname[:50]}",
            ))
            fpath = EVIDENCE_DIR / ev_fname
            results.append(_check(
                fpath.exists(),
                f"evidence_file 실제 존재: {ev_fname[:50]}",
            ))
    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "차량계 catalog 검증 중 예외", tb))

    # ── 5. evidence 파일 내용 검증 (verification_result + document_id) ──
    for doc_id, (ev_id, ev_fname, _, expected_status) in VEHICLE_EVIDENCE_BY_DOC.items():
        fpath = EVIDENCE_DIR / ev_fname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr == expected_status,
                f"evidence {ev_id}: verification_result == '{expected_status}'",
                repr(vr),
            ))
            results.append(_check(
                ev_data.get("document_id") == doc_id,
                f"evidence {ev_id}: document_id == '{doc_id}'",
                repr(ev_data.get("document_id")),
            ))
            results.append(_check(
                ev_data.get("evidence_id") == ev_id,
                f"evidence {ev_id}: evidence_id 자기참조 일치",
                repr(ev_data.get("evidence_id")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {ev_fname[:50]}", str(exc)))

    return results


# ===========================================================================
# 크레인 묶음 smoke test — WP-006 / WP-007 / EQ-003 / EQ-004
#   1. registry: tower_crane_workplan, mobile_crane_workplan 등록 확인
#   2. get_form_spec 으로 두 form_type 조회 확인
#   3. tower_crane_workplan sample build 1건
#   4. mobile_crane_workplan sample build 1건
#   5. 4개 catalog 항목 evidence_* 필드 등록 확인
#   6. evidence 파일 존재 + verification_result 확인
#
# 대상 문서: WP-006, WP-007, EQ-003, EQ-004 (audit doc_id 참조용 명시)
# ===========================================================================

CRANE_TARGET_DOC_IDS = ["WP-006", "WP-007", "EQ-003", "EQ-004"]

CRANE_EVIDENCE_BY_DOC = {
    "WP-006": (
        "WP-006-L1",
        "WP-006-L1_safety_rule_article_38_142_148_tower_crane_workplan.json",
        "tower_crane_workplan",
        "PARTIAL_VERIFIED",
    ),
    "WP-007": (
        "WP-007-L1",
        "WP-007-L1_safety_rule_article_38_134_146_mobile_crane_workplan.json",
        "mobile_crane_workplan",
        "PARTIAL_VERIFIED",
    ),
    "EQ-003": (
        "EQ-003-L1",
        "EQ-003-L1_safety_rule_tower_crane_equipment_specific.json",
        "tower_crane_workplan",
        "PARTIAL_VERIFIED",
    ),
    "EQ-004": (
        "EQ-004-L1",
        "EQ-004-L1_safety_rule_mobile_crane_equipment_specific.json",
        "mobile_crane_workplan",
        "PARTIAL_VERIFIED",
    ),
}

SAMPLE_TOWER_CRANE_MINIMAL: dict = {
    "crane_type": "타워크레인 TC-400",
    "crane_capacity": "400톤m, 최대 작업반경 60m",
    "work_method": "PC 부재 양중 — 신호수 1명 배치, 인양반경 내 출입통제",
}

SAMPLE_MOBILE_CRANE_MINIMAL: dict = {
    "crane_type": "이동식 크레인 (트럭크레인 50톤)",
    "crane_capacity": "50톤, 아웃트리거 전개 후 작업",
    "work_method": "H빔 인양 — 신호수 1명, 인양물 하부 출입금지",
}


def run_crane_workplan_smoke_test() -> list[tuple[str, str, str]]:
    """크레인 묶음 4건 (WP-006/WP-007/EQ-003/EQ-004) 중간 수준 smoke test."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "tower_crane_workplan" in supported,
        "registry: tower_crane_workplan 등록됨 (WP-006/EQ-003)",
    ))
    results.append(_check(
        "mobile_crane_workplan" in supported,
        "registry: mobile_crane_workplan 등록됨 (WP-007/EQ-004)",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    for form_type, expected_name in (
        ("tower_crane_workplan", "타워크레인 작업계획서"),
        ("mobile_crane_workplan", "이동식 크레인 작업계획서"),
    ):
        try:
            spec = get_form_spec(form_type)
            results.append(_check(
                isinstance(spec, dict) and spec.get("display_name") == expected_name,
                f"get_form_spec('{form_type}') → display_name == '{expected_name}'",
                repr(spec.get("display_name")) if isinstance(spec, dict) else "",
            ))
            results.append(_check(
                isinstance(spec.get("required_fields"), (list, tuple))
                and len(spec["required_fields"]) > 0,
                f"{form_type}: required_fields 비어있지 않음",
            ))
        except Exception as exc:
            results.append(_check(False, f"get_form_spec('{form_type}') 호출 성공", str(exc)))

    # ── 3. sample build × 2 ──────────────────────────────────────────────
    for form_type, sample in (
        ("tower_crane_workplan", SAMPLE_TOWER_CRANE_MINIMAL),
        ("mobile_crane_workplan", SAMPLE_MOBILE_CRANE_MINIMAL),
    ):
        try:
            xlsx_bytes = build_form_excel(form_type, sample)
            results.append(_check(
                isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
                f"{form_type} sample build → bytes 생성",
                f"{len(xlsx_bytes):,} bytes",
            ))
        except Exception as exc:
            results.append(_check(False, f"{form_type} sample build", str(exc)))

    # ── 4. 4개 catalog 항목 검증 + evidence_file 존재 ────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)

        for doc_id in CRANE_TARGET_DOC_IDS:
            doc = next((d for d in cat["documents"] if d["id"] == doc_id), None)
            results.append(_check(doc is not None, f"catalog: {doc_id} 항목 존재"))
            if not doc:
                continue

            ev_id, ev_fname, expected_form_type, expected_status = CRANE_EVIDENCE_BY_DOC[doc_id]

            results.append(_check(
                doc.get("implementation_status") == "DONE",
                f"catalog: {doc_id} implementation_status == DONE",
                repr(doc.get("implementation_status")),
            ))
            results.append(_check(
                doc.get("form_type") == expected_form_type,
                f"catalog: {doc_id} form_type == '{expected_form_type}'",
                repr(doc.get("form_type")),
            ))

            ev_status = doc.get("evidence_status", "")
            results.append(_check(
                ev_status == expected_status,
                f"catalog: {doc_id} evidence_status == '{expected_status}'",
                repr(ev_status),
            ))

            ev_ids = doc.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            results.append(_check(
                ev_id in ev_ids,
                f"catalog: {doc_id} evidence_id 포함 — {ev_id}",
            ))

            ev_files = doc.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            results.append(_check(
                ev_fname in ev_basenames,
                f"catalog: {doc_id} evidence_file 등록 — {ev_fname[:50]}",
            ))
            fpath = EVIDENCE_DIR / ev_fname
            results.append(_check(
                fpath.exists(),
                f"evidence_file 실제 존재: {ev_fname[:50]}",
            ))
    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "크레인 catalog 검증 중 예외", tb))

    # ── 5. evidence 파일 내용 검증 (verification_result + document_id) ──
    for doc_id, (ev_id, ev_fname, _, expected_status) in CRANE_EVIDENCE_BY_DOC.items():
        fpath = EVIDENCE_DIR / ev_fname
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr == expected_status,
                f"evidence {ev_id}: verification_result == '{expected_status}'",
                repr(vr),
            ))
            results.append(_check(
                ev_data.get("document_id") == doc_id,
                f"evidence {ev_id}: document_id == '{doc_id}'",
                repr(ev_data.get("document_id")),
            ))
            results.append(_check(
                ev_data.get("evidence_id") == ev_id,
                f"evidence {ev_id}: evidence_id 자기참조 일치",
                repr(ev_data.get("evidence_id")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {ev_fname[:50]}", str(exc)))

    return results


# ===========================================================================
# RA 묶음 — RA-001(위험성평가표) / RA-004(TBM 안전점검 일지)
# ===========================================================================

RA_TARGET_DOC_IDS = ["RA-001", "RA-004"]

# doc_id → (primary_ev_id, primary_ev_fname, expected_form_type, expected_ev_status)
RA_EVIDENCE_BY_DOC: dict = {
    "RA-001": (
        "RA-001-L1",
        "RA-001-L1_industrial_safety_health_act_article_36.json",
        "risk_assessment",
        "VERIFIED",
    ),
    "RA-004": (
        "RA-004-L1",
        "RA-004-L1_notice_2023_19_tbm_daily_assessment.json",
        "tbm_log",
        "VERIFIED",
    ),
}

RA_ALL_EVIDENCE_FILES = [
    "RA-001-L1_industrial_safety_health_act_article_36.json",
    "RA-001-L2_industrial_safety_health_rule_article_37.json",
    "RA-001-L3_notice_2023_19_risk_assessment_method.json",
    "RA-004-L1_notice_2023_19_tbm_daily_assessment.json",
]


def run_ra_forms_smoke_test() -> list[tuple[str, str, str]]:
    """RA 묶음 2건 (RA-001/RA-004) 중간 수준 smoke test."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "risk_assessment" in supported,
        "registry: risk_assessment 등록됨 (RA-001)",
    ))
    results.append(_check(
        "tbm_log" in supported,
        "registry: tbm_log 등록됨 (RA-004)",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    for form_type, expected_name in (
        ("risk_assessment", "위험성평가표"),
        ("tbm_log",         "TBM 안전점검 일지"),
    ):
        try:
            spec = get_form_spec(form_type)
            results.append(_check(
                isinstance(spec, dict) and spec.get("display_name") == expected_name,
                f"get_form_spec('{form_type}') → display_name == '{expected_name}'",
                repr(spec.get("display_name")) if isinstance(spec, dict) else "",
            ))
            results.append(_check(
                isinstance(spec.get("required_fields"), (list, tuple)),
                f"{form_type}: required_fields 리스트 타입",
                repr(type(spec.get("required_fields")).__name__),
            ))
        except Exception as exc:
            results.append(_check(False, f"get_form_spec('{form_type}') 호출 성공", str(exc)))

    # ── 3. sample build × 2 ──────────────────────────────────────────────
    for form_type, sample in (
        ("risk_assessment", {"site_name": "테스트현장", "assessment_date": "2026-04-26"}),
        ("tbm_log",         {"site_name": "테스트현장", "tbm_date": "2026-04-26"}),
    ):
        try:
            xlsx_bytes = build_form_excel(form_type, sample)
            results.append(_check(
                isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
                f"{form_type} sample build → bytes 생성",
                f"{len(xlsx_bytes):,} bytes",
            ))
        except Exception as exc:
            results.append(_check(False, f"{form_type} sample build", str(exc)))

    # ── 4. catalog 항목 검증 + evidence_file 존재 ────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)

        for doc_id in RA_TARGET_DOC_IDS:
            doc = next((d for d in cat["documents"] if d["id"] == doc_id), None)
            results.append(_check(doc is not None, f"catalog: {doc_id} 항목 존재"))
            if not doc:
                continue

            ev_id, ev_fname, expected_form_type, expected_status = RA_EVIDENCE_BY_DOC[doc_id]

            results.append(_check(
                doc.get("implementation_status") == "DONE",
                f"catalog: {doc_id} implementation_status == DONE",
                repr(doc.get("implementation_status")),
            ))
            results.append(_check(
                doc.get("form_type") == expected_form_type,
                f"catalog: {doc_id} form_type == '{expected_form_type}'",
                repr(doc.get("form_type")),
            ))

            ev_status = doc.get("evidence_status", "")
            results.append(_check(
                ev_status == "READY",
                f"catalog: {doc_id} evidence_status == 'READY'",
                repr(ev_status),
            ))

            ev_ids = doc.get("evidence_ids") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            results.append(_check(
                ev_id in ev_ids,
                f"catalog: {doc_id} evidence_ids 포함 — {ev_id}",
            ))

            ev_files = doc.get("evidence_files") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            results.append(_check(
                ev_fname in ev_basenames,
                f"catalog: {doc_id} evidence_files 등록 — {ev_fname[:50]}",
            ))

            fpath = EVIDENCE_DIR / ev_fname
            results.append(_check(
                fpath.exists(),
                f"evidence_file 실제 존재: {ev_fname[:50]}",
            ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "RA catalog 검증 중 예외", tb))

    # ── 5. 모든 evidence 파일 내용 검증 ──────────────────────────────────
    for ev_fname in RA_ALL_EVIDENCE_FILES:
        fpath = EVIDENCE_DIR / ev_fname
        if not fpath.exists():
            results.append(_check(False, f"evidence 파일 존재: {ev_fname[:55]}"))
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr == "VERIFIED",
                f"evidence {ev_data.get('evidence_id', '?')}: verification_result == 'VERIFIED'",
                repr(vr),
            ))
            results.append(_check(
                bool(ev_data.get("document_id")),
                f"evidence {ev_data.get('evidence_id', '?')}: document_id 존재",
                repr(ev_data.get("document_id")),
            ))
            results.append(_check(
                bool(ev_data.get("evidence_id")),
                f"evidence {ev_fname[:40]}: evidence_id 존재",
                repr(ev_data.get("evidence_id")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {ev_fname[:55]}", str(exc)))

    return results



# ===========================================================================
# 밀폐공간 묶음 smoke test — WP-014 / PTW-001 / CL-010
#   1. registry: confined_space_workplan, confined_space_permit,
#                confined_space_checklist 등록 확인
#   2. get_form_spec × 3 조회 확인
#   3. sample build × 3
#   4. 3개 catalog 항목 evidence_* 필드 등록 확인
#   5. evidence 파일 존재 + verification_result 확인
# ===========================================================================

CONFINED_TARGET_DOC_IDS = ["WP-014", "PTW-001", "CL-010"]

CONFINED_EVIDENCE_BY_DOC = {
    "WP-014": (
        "WP-014-L1",
        "WP-014-L1_safety_rule_articles_619_confined_space_workplan.json",
        "confined_space_workplan",
        "VERIFIED",
    ),
    "PTW-001": (
        "PTW-001-L1",
        "PTW-001-L1_safety_rule_articles_619_622_confined_space_permit.json",
        "confined_space_permit",
        "VERIFIED",
    ),
    "CL-010": (
        "CL-010-L1",
        "CL-010-L1_safety_rule_articles_622_625_confined_space_checklist.json",
        "confined_space_checklist",
        "VERIFIED",
    ),
}

SAMPLE_CONFINED_WORKPLAN_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "work_date": "2026-04-26",
    "confined_space_location": "맨홀 MH-01",
}

SAMPLE_CONFINED_PERMIT_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "work_date": "2026-04-26",
    "confined_space_location": "오수받이 CS-02",
}

SAMPLE_CONFINED_CHECKLIST_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "check_date": "2026-04-26",
    "confined_space_location": "저류조 TS-01",
}


def run_confined_space_forms_smoke_test() -> list[tuple[str, str, str]]:
    """밀폐공간 묶음 3건 (WP-014/PTW-001/CL-010) 중간 수준 smoke test."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    for ft in ("confined_space_workplan", "confined_space_permit", "confined_space_checklist"):
        results.append(_check(ft in supported, f"registry: {ft} 등록됨"))

    # ── 2. get_form_spec × 3 ────────────────────────────────────────────
    expected_names = {
        "confined_space_workplan": "밀폐공간 작업계획서",
        "confined_space_permit":   "밀폐공간 작업허가서",
        "confined_space_checklist": "밀폐공간 사전 안전점검표",
    }
    for form_type, expected_name in expected_names.items():
        try:
            spec = get_form_spec(form_type)
            results.append(_check(
                isinstance(spec, dict) and spec.get("display_name") == expected_name,
                f"get_form_spec('{form_type}') → display_name == '{expected_name}'",
                repr(spec.get("display_name")) if isinstance(spec, dict) else "",
            ))
            results.append(_check(
                isinstance(spec.get("required_fields"), (list, tuple))
                and len(spec["required_fields"]) > 0,
                f"{form_type}: required_fields 비어있지 않음",
            ))
        except Exception as exc:
            results.append(_check(False, f"get_form_spec('{form_type}') 호출 성공", str(exc)))

    # ── 3. sample build × 3 ──────────────────────────────────────────────
    for form_type, sample in (
        ("confined_space_workplan",  SAMPLE_CONFINED_WORKPLAN_MINIMAL),
        ("confined_space_permit",    SAMPLE_CONFINED_PERMIT_MINIMAL),
        ("confined_space_checklist", SAMPLE_CONFINED_CHECKLIST_MINIMAL),
    ):
        try:
            xlsx_bytes = build_form_excel(form_type, sample)
            results.append(_check(
                isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
                f"{form_type} sample build → bytes 생성",
                f"{len(xlsx_bytes):,} bytes",
            ))
        except Exception as exc:
            results.append(_check(False, f"{form_type} sample build", str(exc)))

    # ── 4. 3개 catalog 항목 검증 + evidence_file 존재 ────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)

        for doc_id in CONFINED_TARGET_DOC_IDS:
            doc = next((d for d in cat["documents"] if d["id"] == doc_id), None)
            results.append(_check(doc is not None, f"catalog: {doc_id} 항목 존재"))
            if not doc:
                continue

            ev_id, ev_fname, expected_form_type, expected_status = CONFINED_EVIDENCE_BY_DOC[doc_id]

            results.append(_check(
                doc.get("implementation_status") == "DONE",
                f"catalog: {doc_id} implementation_status == DONE",
                repr(doc.get("implementation_status")),
            ))
            results.append(_check(
                doc.get("form_type") == expected_form_type,
                f"catalog: {doc_id} form_type == '{expected_form_type}'",
                repr(doc.get("form_type")),
            ))

            ev_status = doc.get("evidence_status", "")
            results.append(_check(
                ev_status == expected_status,
                f"catalog: {doc_id} evidence_status == '{expected_status}'",
                repr(ev_status),
            ))

            ev_ids = doc.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            results.append(_check(
                ev_id in ev_ids,
                f"catalog: {doc_id} evidence_id 포함 — {ev_id}",
            ))

            ev_files = doc.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            results.append(_check(
                ev_fname in ev_basenames,
                f"catalog: {doc_id} evidence_file 등록 — {ev_fname[:50]}",
            ))
            fpath = EVIDENCE_DIR / ev_fname
            results.append(_check(
                fpath.exists(),
                f"evidence_file 실제 존재: {ev_fname[:50]}",
            ))
    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "밀폐공간 catalog 검증 중 예외", tb))

    # ── 5. evidence 파일 내용 검증 ────────────────────────────────────────
    for doc_id, (ev_id, ev_fname, _, expected_status) in CONFINED_EVIDENCE_BY_DOC.items():
        fpath = EVIDENCE_DIR / ev_fname
        if not fpath.exists():
            results.append(_check(False, f"evidence 파일 존재: {ev_fname[:55]}"))
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr == expected_status,
                f"evidence {ev_id}: verification_result == '{expected_status}'",
                repr(vr),
            ))
            results.append(_check(
                ev_data.get("document_id") == doc_id,
                f"evidence {ev_id}: document_id == '{doc_id}'",
                repr(ev_data.get("document_id")),
            ))
            results.append(_check(
                ev_data.get("evidence_id") == ev_id,
                f"evidence {ev_id}: evidence_id 자기참조 일치",
                repr(ev_data.get("evidence_id")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {ev_fname[:50]}", str(exc)))

    return results


# ===========================================================================
# WP-001 — 굴착 작업계획서
# ===========================================================================

SAMPLE_WP001_MINIMAL: dict = {
    "site_name": "테스트건설(주) 테스트현장",
    "work_date": "2026-04-26",
    "work_location": "1공구 굴착구간",
    "work_supervisor": "홍길동",
    "excavation_depth": "3.5",
    "excavation_method": "오픈컷",
}

WP001_EVIDENCE_FILES = [
    "WP-001-L1_safety_rule_article_38_item6_excavation.json",
    "WP-001-L2_safety_rule_article_339_excavation_face.json",
]
WP001_EVIDENCE_IDS = ["WP-001-L1", "WP-001-L2"]


def run_wp001_smoke_test() -> list[tuple[str, str, str]]:
    """WP-001 굴착 작업계획서 중간 수준 smoke test."""
    results: list[tuple[str, str, str]] = []
    supported = {f["form_type"] for f in list_supported_forms()}

    # ── 1. registry 등록 확인 ────────────────────────────────────────────
    results.append(_check(
        "excavation_workplan" in supported,
        "registry: excavation_workplan 등록됨 (WP-001)",
    ))

    # ── 2. get_form_spec 검증 ────────────────────────────────────────────
    try:
        spec = get_form_spec("excavation_workplan")
        results.append(_check(
            isinstance(spec, dict),
            "get_form_spec('excavation_workplan') → dict",
        ))
        results.append(_check(
            isinstance(spec.get("required_fields"), (list, tuple))
            and len(spec["required_fields"]) > 0,
            "excavation_workplan: required_fields 비어있지 않음",
        ))
    except Exception as exc:
        results.append(_check(False, "get_form_spec('excavation_workplan') 호출 성공", str(exc)))

    # ── 3. sample build ──────────────────────────────────────────────────
    try:
        xlsx_bytes = build_form_excel("excavation_workplan", SAMPLE_WP001_MINIMAL)
        results.append(_check(
            isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
            "excavation_workplan sample build → bytes 생성",
            f"{len(xlsx_bytes):,} bytes",
        ))
    except Exception as exc:
        results.append(_check(False, "excavation_workplan sample build", str(exc)))

    # ── 4. catalog 항목 검증 ─────────────────────────────────────────────
    try:
        import yaml
        catalog_path = Path("data/masters/safety/documents/document_catalog.yml")
        with open(catalog_path, encoding="utf-8") as f:
            cat = yaml.safe_load(f)

        doc = next((d for d in cat["documents"] if d["id"] == "WP-001"), None)
        results.append(_check(doc is not None, "catalog: WP-001 항목 존재"))

        if doc:
            results.append(_check(
                doc.get("implementation_status") == "DONE",
                "catalog: WP-001 implementation_status == DONE",
                repr(doc.get("implementation_status")),
            ))
            results.append(_check(
                doc.get("form_type") == "excavation_workplan",
                "catalog: WP-001 form_type == 'excavation_workplan'",
                repr(doc.get("form_type")),
            ))
            ev_status = doc.get("evidence_status", "")
            results.append(_check(
                ev_status == "READY",
                "catalog: WP-001 evidence_status == 'READY'",
                repr(ev_status),
            ))

            ev_ids = doc.get("evidence_id") or []
            if isinstance(ev_ids, str):
                ev_ids = [ev_ids]
            for eid in WP001_EVIDENCE_IDS:
                results.append(_check(
                    eid in ev_ids,
                    f"catalog: WP-001 evidence_id 포함 — {eid}",
                ))

            ev_files = doc.get("evidence_file") or []
            if isinstance(ev_files, str):
                ev_files = [ev_files]
            ev_basenames = {Path(p).name for p in ev_files}
            for efname in WP001_EVIDENCE_FILES:
                results.append(_check(
                    efname in ev_basenames,
                    f"catalog: WP-001 evidence_file 등록 — {efname[:55]}",
                ))

    except Exception as exc:
        tb = traceback.format_exc().strip().splitlines()[-1]
        results.append(_check(False, "WP-001 catalog 검증 중 예외", tb))

    # ── 5. evidence 파일 존재 + 내용 검증 ────────────────────────────────
    for ev_fname, ev_id in zip(WP001_EVIDENCE_FILES, WP001_EVIDENCE_IDS):
        fpath = EVIDENCE_DIR / ev_fname
        results.append(_check(fpath.exists(), f"evidence_file 실제 존재: {ev_fname[:55]}"))
        if not fpath.exists():
            continue
        try:
            ev_data = json.loads(fpath.read_text(encoding="utf-8"))
            vr = ev_data.get("verification_result", "")
            results.append(_check(
                vr == "VERIFIED",
                f"evidence {ev_id}: verification_result == 'VERIFIED'",
                repr(vr),
            ))
            results.append(_check(
                ev_data.get("document_id") == "WP-001",
                f"evidence {ev_id}: document_id == 'WP-001'",
                repr(ev_data.get("document_id")),
            ))
            results.append(_check(
                ev_data.get("evidence_id") == ev_id,
                f"evidence {ev_id}: evidence_id 자기참조 일치",
                repr(ev_data.get("evidence_id")),
            ))
        except Exception as exc:
            results.append(_check(False, f"evidence 파일 로딩: {ev_fname[:50]}", str(exc)))

    return results


if __name__ == "__main__":
    run_smoke_test()

