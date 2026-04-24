"""
굴착 작업계획서 Excel builder 검증 스크립트.

검증 항목:
  1. xlsx bytes 정상 생성 여부
  2. openpyxl 재오픈 성공
  3. 시트명 == "작업계획서"
  4. 제목 셀(A1) == "굴착 작업계획서"
  5. 부제 셀(A2) == SHEET_SUBTITLE
  6. 법정항목 7개 라벨 존재
  7. 안전조치표 10행 존재 (순번 1~10)
  8. 인쇄영역 A1:H29
  9. 주요 병합셀 확인
  10. 빈 form_data → 공란 유지

실행:
  python scripts/validate_workplan_builder.py
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from engine.output.workplan_builder import (
    LEGAL_LABELS,
    MAX_STEPS,
    SHEET_HEADING,
    SHEET_NAME,
    SHEET_SUBTITLE,
    build_excavation_workplan_excel,
)

# ---------------------------------------------------------------------------
# 샘플 form_data
# ---------------------------------------------------------------------------
SAMPLE_FORM_DATA = {
    "site_name": "테스트 사업장",
    "project_name": "○○ 신축 공사 현장",
    "work_location": "부지 북측 경계 A구간 (STA. 0+000 ~ 0+200)",
    "work_date": "2026-04-25 ~ 2026-05-10",
    "supervisor": "홍길동 (010-1234-5678)",
    "contractor": "㈜한국건설",
    "excavation_method": "개착식 굴착, 굴착 심도 최대 5m, 1:1 경사면 유지",
    "earth_retaining": "H-Pile + 토류판 방식, 스트럿 2단 설치",
    "excavation_machine": "백호우 0.8m³ (CAT 320D), 덤프트럭 15t × 3대",
    "soil_disposal": "현장 외 토사장 반출 — 00환경 처리, 반출 거리 12km",
    "water_disposal": "웰포인트 공법 적용, 집수정 2개소, 수중펌프 상시 가동",
    "work_method": "굴착 → 흙막이 설치 → 스트럿 → 재굴착 순서 반복",
    "emergency_measure": "붕괴 징후 시 즉시 작업 중단 및 대피, 안전관리자 신고",
    "safety_steps": [
        {
            "task_step": "사전 지하매설물 확인",
            "hazard": "굴착 중 가스·전기·통신 매설물 파손",
            "safety_measure": "도면 확인 및 시험굴착 실시",
        },
        {
            "task_step": "굴착기 진입 및 굴착",
            "hazard": "굴착기 전도·후방 충돌",
            "safety_measure": "유도자 배치, 안전 이격거리 유지",
        },
        {
            "task_step": "흙막이 설치",
            "hazard": "H-Pile 항타 중 비래·붕괴",
            "safety_measure": "보호구 착용, 항타 반경 내 출입 통제",
        },
    ],
    "sign_date": "2026-04-25",
}

SAMPLE_EMPTY_DATA: dict = {}


# ---------------------------------------------------------------------------
# 검증 헬퍼
# ---------------------------------------------------------------------------

def _check(condition: bool, name: str, detail: str = "") -> bool:
    status = "PASS" if condition else "FAIL"
    msg = f"  [{status}] {name}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    return condition


def _all_cell_values(ws) -> set[str]:
    vals: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                vals.add(str(v))
    return vals


def _merged_ranges(ws) -> set[str]:
    return {str(m) for m in ws.merged_cells.ranges}


# ---------------------------------------------------------------------------
# 검증 함수
# ---------------------------------------------------------------------------

def validate_full(xlsx_bytes: bytes, label: str) -> list[bool]:
    results: list[bool] = []
    print(f"\n=== {label} ===")

    # 1. bytes 생성
    results.append(_check(
        isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0,
        "xlsx bytes 정상 생성",
        f"{len(xlsx_bytes):,} bytes",
    ))

    # 2. 재오픈
    try:
        wb = load_workbook(BytesIO(xlsx_bytes))
        results.append(_check(True, "openpyxl 재오픈 성공"))
    except Exception as e:
        results.append(_check(False, "openpyxl 재오픈", str(e)))
        return results

    # 3. 시트명
    results.append(_check(
        SHEET_NAME in wb.sheetnames,
        f"시트명 == '{SHEET_NAME}'",
        str(wb.sheetnames),
    ))
    ws = wb[SHEET_NAME]

    # 4. 제목 셀 A1
    title_val = ws.cell(row=1, column=1).value
    results.append(_check(
        title_val == SHEET_HEADING,
        f"제목(A1) == '{SHEET_HEADING}'",
        repr(title_val),
    ))

    # 5. 부제 셀 A2
    sub_val = ws.cell(row=2, column=1).value
    results.append(_check(
        sub_val == SHEET_SUBTITLE,
        f"부제(A2) == SHEET_SUBTITLE",
        repr(sub_val),
    ))

    # 6. 법정항목 7개 라벨 존재
    all_vals = _all_cell_values(ws)
    for label_text in LEGAL_LABELS:
        results.append(_check(
            label_text in all_vals,
            f"법정항목 라벨 존재: '{label_text}'",
        ))

    # 7. 안전조치표 10행 존재 (순번 1~10)
    step_numbers: list[int] = []
    for row in ws.iter_rows(values_only=True):
        v = row[0]  # A열
        if isinstance(v, int) and 1 <= v <= MAX_STEPS:
            step_numbers.append(v)
    results.append(_check(
        set(step_numbers) == set(range(1, MAX_STEPS + 1)),
        f"안전조치표 {MAX_STEPS}행(순번 1~{MAX_STEPS}) 존재",
        f"발견: {sorted(step_numbers)}",
    ))

    # 8. 인쇄영역 A1:H29
    # openpyxl은 시트명 포함 형식으로 반환: "'시트명'!$A$1:$H$29"
    print_area = ws.print_area or ""
    pa_norm = print_area.replace("$", "").split("!")[-1]
    results.append(_check(
        pa_norm == "A1:H29",
        "인쇄영역 == 'A1:H29'",
        repr(print_area),
    ))

    # 9. 주요 병합셀 확인
    merged = _merged_ranges(ws)
    key_merges = [
        "A1:H1",   # 제목
        "A2:H2",   # 부제
        "B3:D3",   # site_name
        "F3:H3",   # project_name
        "B4:H4",   # work_location
        "B5:D5",   # work_date
        "F5:H5",   # supervisor
        "A7:H7",   # 섹션헤더: 법정기재사항
        "B8:H8",   # excavation_method
        "A15:H15", # 섹션헤더: 안전조치
        "B16:C16", # 표헤더: 작업단계
        "D16:F16", # 표헤더: 위험요인
        "G16:H16", # 표헤더: 안전조치
        "A27:H27", # 섹션헤더: 확인및서명
        "A28:B28", # 작성자 라벨
        "E28:F28", # 검토자 라벨
    ]
    for mr in key_merges:
        results.append(_check(
            mr in merged,
            f"병합셀 존재: {mr}",
        ))

    return results


def validate_empty(xlsx_bytes: bytes) -> list[bool]:
    """빈 form_data로 생성한 xlsx: 값 필드가 모두 공란인지 확인."""
    results: list[bool] = []
    print("\n=== 공란 유지 검증 (빈 form_data) ===")

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb[SHEET_NAME]

    header_texts = {
        SHEET_HEADING,
        SHEET_SUBTITLE,
        "사업장명", "현장명", "작업 위치", "작업 일자", "작업 책임자",
        "도급업체", "작성일",
        "법정 기재 사항 (산업안전보건기준에 관한 규칙 제82조)",
        "굴착의 방법", "흙막이 지보공 및 방호망", "사용 기계 종류 및 능력",
        "토석 처리 방법", "용수 처리 방법", "작업 방법", "긴급조치 계획",
        "작업단계별 안전조치",
        "순번", "작업 단계", "위험 요인", "안전 조치",
        "확인 및 서명",
        "작성자", "검토자/확인자",
    }

    non_blank_non_header: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is None or v == "":
                continue
            sv = str(v)
            if sv in header_texts:
                continue
            if isinstance(v, int):  # 순번 정수 허용
                continue
            non_blank_non_header.append(sv)

    results.append(_check(
        len(non_blank_non_header) == 0,
        "모든 값 필드 공란 유지",
        f"비공란 비헤더 값: {non_blank_non_header[:5] if non_blank_non_header else '없음'}",
    ))
    return results


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def main() -> int:
    print("굴착 작업계획서 Excel Builder 검증 시작")
    print("=" * 50)

    all_results: list[bool] = []

    xlsx_full = build_excavation_workplan_excel(SAMPLE_FORM_DATA)
    all_results += validate_full(xlsx_full, "샘플 데이터 검증")

    xlsx_empty = build_excavation_workplan_excel(SAMPLE_EMPTY_DATA)
    all_results += validate_full(xlsx_empty, "빈 form_data 검증")
    all_results += validate_empty(xlsx_empty)

    total  = len(all_results)
    passed = sum(all_results)
    failed = total - passed

    print("\n" + "=" * 50)
    print(f"결과: {passed}/{total} PASS, {failed} FAIL")
    if failed == 0:
        print("최종 판정: PASS")
        return 0
    else:
        print("최종 판정: FAIL")
        return 1


if __name__ == "__main__":
    sys.exit(main())
