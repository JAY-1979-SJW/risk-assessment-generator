"""
P0 6종 + 기존 DONE 5종 Excel 생성 smoke test.
EQ-003/EQ-004 catalog form_type 연결 검증 포함.
HM-001/HM-002 보건관리 builder 검증 포함.

실행:
    cd <project_root>
    python scripts/smoke_test_p0_forms.py

판정:
    PASS — xlsx bytes 생성 + openpyxl로 재열기 성공
    WARN — bytes 생성됐으나 경고 발생
    FAIL — 예외 발생 또는 bytes 미생성
"""
from __future__ import annotations

import sys
import traceback
from io import BytesIO

from openpyxl import load_workbook

sys.path.insert(0, ".")

from engine.output.form_registry import build_form_excel, list_supported_forms
from engine.output.work_environment_measurement_builder import ORIGINAL_REPORT_NOTE
from engine.output.special_health_examination_builder import ORIGINAL_RESULT_NOTE

# ---------------------------------------------------------------------------
# 샘플 form_data (최소 required_fields 채움, optional 생략 가능)
# ---------------------------------------------------------------------------

SAMPLES: dict[str, dict] = {
    # ── 기존 DONE 5종 ──────────────────────────────────────────────────────
    "education_log": {
        "education_type": "정기교육",
        "education_date": "2026-04-24",
        "education_location": "현장 회의실",
        "education_duration_hours": "2",
        "education_target_job": "건설 근로자",
        "instructor_name": "홍길동",
        "instructor_qualification": "산업안전지도사",
        "confirmer_name": "김철수",
        "confirmer_role": "안전관리자",
    },
    "risk_assessment": {
        "company_name": "테스트건설(주)",
        "assessment_type": "최초평가",
        "assessment_date": "2026-04-24",
        "rows": [
            {
                "hazard_category": "추락",
                "hazard_detail": "비계 작업 중 추락",
                "current_measure": "안전난간 설치",
                "possibility": "2",
                "severity": "3",
                "risk_level": "중",
                "improvement": "안전대 착용 의무화",
            }
        ],
    },
    "excavation_workplan": {
        "excavation_method": "기계 굴착",
        "earth_retaining": "H-Pile + 토류판",
        "excavation_machine": "굴착기 (0.6㎥)",
        "soil_disposal": "덤프트럭 반출",
        "water_disposal": "집수정 + 수중펌프",
        "work_method": "굴착면 분할 굴착, 상부에서 하부 방향으로 시공",
        "emergency_measure": "119 즉시 신고, 현장 대피 경보 발령",
    },
    "vehicle_construction_workplan": {
        "machine_type": "굴착기",
        "machine_capacity": "0.8㎥",
        "work_method": "굴착 후 토사 적재",
        "travel_route_text": "현장 정문 → 굴착 구역 → 토사 적치장",
    },
    "material_handling_workplan": {
        "machine_type": "지게차",
        "machine_max_load": "3톤",
        "work_method": "자재 인양 후 지정 위치 이동",
        "travel_route_text": "자재창고 → 작업구역 (일방통행)",
        "emergency_measure": "119 신고, 유도자 즉시 대피 안내",
    },
    # ── P0 신규 6종 ─────────────────────────────────────────────────────────
    "tower_crane_workplan": {
        "crane_type": "타워크레인 (고정식)",
        "crane_capacity": "12톤",
        "work_method": "자재 인양 후 지정 위치 하강, 신호수 수신호에 따라 조작",
        "emergency_measure": "신호수 이상 신호 시 즉시 정지, 119 신고",
        "crane_reg_no": "서울-2024-001",
        "work_radius": "50m",
        "signal_worker": "신호수 1명 (지상 배치)",
        "anti_fall_measures": "작업반경 내 안전 펜스 설치, 낙하물 방지망 설치",
    },
    "mobile_crane_workplan": {
        "crane_type": "이동식 크레인 (카고크레인)",
        "crane_capacity": "25톤",
        "work_method": "아웃트리거 완전 전개 후 인양, 신호수 수신호 준수",
        "emergency_measure": "전도 위험 시 즉시 작업 중단, 119 신고",
        "vehicle_no": "경기12가3456",
        "outrigger_setup": "전후좌우 4점 완전 전개, 철판 받침 설치",
        "ground_condition": "콘크리트 포장면, 지지력 충분",
        "load_weight": "15톤",
        "work_radius": "15m",
        "rigging_method": "4점 걸이 (와이어로프 φ22mm)",
        "signal_worker": "신호수 2명 (전후 각 1명)",
        "anti_topple_measures": "수평 유지 확인, SWL 준수, 인양 경로 하부 출입 통제",
    },
    "confined_space_workplan": {
        "confined_space_location": "지하 2층 오수 집수정",
        "gas_measurement_plan": "작업 전·중 30분마다 가스 측정기로 산소·H₂S·CO 측정",
        "ventilation_plan": "이동식 송풍기 (1.5kW) 연속 가동, 배기구 반대편 설치",
        "emergency_measure": "이상 신호 시 즉시 대피, 119 신고, 공기호흡기 착용 후 구조",
        "confined_space_type": "오수 집수정 (맨홀형)",
        "worker_count": "2명",
        "monitor_placement": "외부 감시인 1명 상시 대기",
        "rescue_equipment": "구조용 로프, 안전대, 공기호흡기 2set",
        "access_control": "맨홀 덮개 제거 후 안전 펜스 설치, 출입자 명단 기록",
    },
    "tbm_log": {
        "tbm_date": "2026-04-24 08:00",
        "today_work": "지하 2층 밀폐공간 배관 보수 작업",
        "hazard_points": "산소결핍, 황화수소 발생 가능, 미끄러짐",
        "safety_instructions": "공기호흡기 착용 필수, 30분마다 가스 측정, 이상 시 즉시 대피",
        "tbm_location": "현장 조립식 사무실 앞",
        "trade_name": "기계설비 배관 보수",
        "pre_work_checks": "1. 가스 농도 측정 완료\n2. 환기 설비 가동 확인\n3. 안전장비 지급 완료",
        "permit_check": "밀폐공간 작업허가서 #CS-2026-001 확인",
        "ppe_check": "안전모, 안전화, 공기호흡기, 안전대 착용 확인",
        "attendees": [
            {"name": "홍길동", "job_type": "배관공"},
            {"name": "이순신", "job_type": "감시인"},
        ],
    },
    "confined_space_permit": {
        "permit_no": "CS-2026-001",
        "work_location": "지하 2층 오수 집수정 (맨홀 #3)",
        "work_content": "오수 배관 이음부 보수",
        "monitor_name": "이순신 (반장)",
        "oxygen_level": "20.9%",
        "permit_issuer": "김안전 (안전관리자)",
        "permit_datetime": "2026-04-24 08:30",
        "validity_period": "2026-04-24 08:30 ~ 17:00",
        "gas_h2s_level": "0.1ppm 이하 (허용: 10ppm)",
        "gas_co_level": "5ppm 이하 (허용: 25ppm)",
        "ventilation_status": "송풍기 정상 가동 확인",
        "rescue_equipment_check": "구조용 로프·안전대·공기호흡기 2set 현장 비치 확인",
        "workers": [
            {"name": "홍길동", "role": "배관공"},
        ],
    },
    "confined_space_checklist": {
        "check_date": "2026-04-24",
        "work_location": "지하 2층 오수 집수정",
        "checker_name": "김안전",
        "work_content": "오수 배관 이음부 보수",
        "check_items": [
            {"result": "○", "note": "산소 20.9%"},
            {"result": "○", "note": "H₂S 0.1ppm, CO 5ppm"},
            {"result": "○", "note": "송풍기 1.5kW 가동"},
            {"result": "○", "note": "이순신 반장 배치"},
            {"result": "○", "note": "로프·안전대·공기호흡기 2set"},
            {"result": "○", "note": "무선기 2대"},
            {"result": "○", "note": "공기호흡기 착용 완료"},
            {"result": "○", "note": "명단 기록 완료"},
            {"result": "○", "note": "119 및 안전관리자 연락처 확인"},
            {"result": "",  "note": "작업 종료 후 작성 예정"},
        ],
    },
    # ── HM-001/HM-002 보건관리 ───────────────────────────────────────────────
    "work_environment_measurement": {
        "site_name": "테스트사업장",
        "target_process": "용접·도장 작업 구역 (1공장 A라인)",
        "hazardous_agents": "소음, 분진, 톨루엔",
        "measurement_agency": "한국산업안전보건공단 인정 측정기관 (가나다측정원)",
        "agency_contact": "02-0000-0000",
        "measurement_date": "2026-04-10",
        "result_received_date": "2026-04-17",
        "result_summary": "소음 87dB(A) — 기준 90dB(A) 이하, 분진 0.8mg/m³ — 기준 이하, 톨루엔 15ppm — 기준 이하",
        "exceedance_status": "초과 없음",
        "improvement_plan": "소음 수준 지속 모니터링, 귀마개 착용 의무화",
        "worker_notification": "2026-04-18 전체 근로자 측정 결과 공지 (게시판 게시 및 반장 전달)",
        "original_attached": "첨부 완료 (안전보건 서류함 보관)",
        "measurement_rows": [
            {"target_location": "1공장 A라인", "hazardous_agent": "소음", "measured_value": "87dB(A)", "exposure_limit": "90dB(A)", "exceedance": "미초과"},
            {"target_location": "1공장 A라인", "hazardous_agent": "분진", "measured_value": "0.8mg/m³", "exposure_limit": "10mg/m³", "exceedance": "미초과"},
        ],
        "sign_date": "2026-04-25",
    },
    "special_health_examination": {
        "site_name": "테스트사업장",
        "exam_target_work": "소음 발생 작업, 분진 발생 작업",
        "exam_agency": "서울산업보건센터 (지정 검진기관)",
        "agency_contact": "02-1111-2222",
        "exam_date": "2026-04-05",
        "result_received_date": "2026-04-15",
        "exam_type": "정기",
        "hazardous_agents": "소음, 분진",
        "judgment_summary": "A(정상) 8명, C1(직업적 질병 요관찰) 2명, 미수검 1명",
        "followup_plan": "C1 판정자 2명 작업 환경 개선 및 3개월 후 재검진 실시",
        "non_exam_count": "1명",
        "non_exam_reason": "장기 출장",
        "non_exam_action": "귀임 후 1개월 내 수시건강진단 실시 예정",
        "original_stored": "보관 완료 (인사팀 잠금 서류함)",
        "privacy_confirmed": "확인 — 열람 권한자 제한 완료",
        "worker_rows": [
            {"employee_no": "E001", "name": "홍길동", "birth_year": "1985", "job_type": "용접공", "exam_done": "완료", "judgment": "A", "followup_needed": "없음"},
            {"employee_no": "E002", "name": "이순신", "birth_year": "1979", "job_type": "배관공", "exam_done": "완료", "judgment": "C1", "followup_needed": "필요"},
        ],
        "sign_date": "2026-04-25",
    },
}

# ---------------------------------------------------------------------------
# 필수 섹션 검사 — 각 form_type별 시트 제목 확인
# ---------------------------------------------------------------------------

REQUIRED_SHEET_HEADINGS: dict[str, str] = {
    "education_log":                   "안전보건교육",
    "risk_assessment":                 "위험성평가",
    "excavation_workplan":             "굴착",
    "vehicle_construction_workplan":   "차량계 건설기계",
    "material_handling_workplan":      "차량계 하역운반기계",
    "tower_crane_workplan":            "타워크레인",
    "mobile_crane_workplan":           "이동식 크레인",
    "confined_space_workplan":         "밀폐공간",
    "tbm_log":                         "TBM",
    "confined_space_permit":           "밀폐공간",
    "confined_space_checklist":        "밀폐공간",
    "work_environment_measurement":    "작업환경측정",
    "special_health_examination":      "특수건강진단",
}

# ---------------------------------------------------------------------------
# 검증 실행
# ---------------------------------------------------------------------------

def run_smoke_test() -> None:
    supported = {f["form_type"] for f in list_supported_forms()}
    results   = {}

    for form_type, form_data in SAMPLES.items():
        label = form_type
        try:
            # 1. registry에 등록됐는지 확인
            if form_type not in supported:
                results[label] = ("FAIL", f"registry에 미등록: {form_type}")
                continue

            # 2. bytes 생성
            xlsx_bytes = build_form_excel(form_type, form_data)
            if not isinstance(xlsx_bytes, bytes) or len(xlsx_bytes) == 0:
                results[label] = ("FAIL", "bytes 미생성 또는 0 bytes")
                continue

            # 3. openpyxl로 재열기
            wb = load_workbook(BytesIO(xlsx_bytes))
            ws = wb.active

            # 4. 시트 제목 포함 여부 (A1 셀 값으로 확인)
            a1_val = str(ws.cell(row=1, column=1).value or "")
            heading = REQUIRED_SHEET_HEADINGS.get(form_type, "")
            if heading and heading not in a1_val:
                results[label] = ("WARN", f"A1 제목에 '{heading}' 미포함: {a1_val!r}")
                continue

            results[label] = ("PASS", f"{len(xlsx_bytes):,} bytes")

        except Exception:
            tb = traceback.format_exc().strip().splitlines()[-1]
            results[label] = ("FAIL", tb)

    # ── 출력 ──────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  KRAS Form Builder Smoke Test")
    print("=" * 60)

    pass_cnt = warn_cnt = fail_cnt = 0
    for form_type, (verdict, msg) in results.items():
        icon = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌"}[verdict]
        print(f"  {icon} [{verdict}] {form_type:<36} {msg}")
        if verdict == "PASS": pass_cnt += 1
        elif verdict == "WARN": warn_cnt += 1
        else: fail_cnt += 1

    print("-" * 60)
    total = len(results)
    print(f"  합계: PASS {pass_cnt}/{total}  WARN {warn_cnt}  FAIL {fail_cnt}")

    if fail_cnt > 0:
        overall = "FAIL"
    elif warn_cnt > 0:
        overall = "WARN"
    else:
        overall = "PASS"

    # ── EQ-003/EQ-004 catalog form_type 연결 검증 ──────────────────
    print("  " + "-" * 56)
    print("  EQ-003/EQ-004 catalog 연결 검증")
    eq_links = {
        "EQ-003 (타워크레인 장비특화)":     "tower_crane_workplan",
        "EQ-004 (이동식크레인 장비특화)":   "mobile_crane_workplan",
    }
    eq_fail = 0
    for label, ft in eq_links.items():
        ok = ft in supported
        icon = "✅" if ok else "❌"
        print(f"  {icon} {label} → form_type={ft}")
        if not ok:
            eq_fail += 1
    if eq_fail:
        overall = "FAIL"

    # ── HM-001/HM-002 원본 첨부 문구 검증 ─────────────────────────
    print("  " + "-" * 56)
    print("  HM-001/HM-002 원본 첨부 문구 검증")
    hm_checks = [
        ("HM-001 (작업환경측정)", "work_environment_measurement",
         SAMPLES.get("work_environment_measurement", {}),
         "외부 전문 측정기관"),
        ("HM-002 (특수건강진단)", "special_health_examination",
         SAMPLES.get("special_health_examination", {}),
         "외부 지정 검진기관"),
    ]
    hm_fail = 0
    for label, ft, sample, note_keyword in hm_checks:
        try:
            xlsx_bytes = build_form_excel(ft, sample)
            wb2 = load_workbook(BytesIO(xlsx_bytes))
            found = any(
                note_keyword in str(cell.value or "")
                for ws2 in wb2.worksheets
                for row in ws2.iter_rows()
                for cell in row
            )
            icon = "✅" if found else "❌"
            msg  = "원본 첨부 문구 확인" if found else f"문구 미발견: '{note_keyword}'"
            print(f"  {icon} {label} → {msg}")
            if not found:
                hm_fail += 1
                overall = "FAIL"
        except Exception as exc:
            print(f"  ❌ {label} → 검증 중 예외: {exc}")
            hm_fail += 1
            overall = "FAIL"
    if hm_fail == 0 and "work_environment_measurement" in supported and "special_health_examination" in supported:
        print("  ✅ HM-001/HM-002 catalog form_type 연결 확인")
    else:
        for ft in ("work_environment_measurement", "special_health_examination"):
            if ft not in supported:
                print(f"  ❌ {ft} → registry 미등록")
                overall = "FAIL"

    print(f"\n  최종 판정: {overall}")
    print("=" * 60 + "\n")

    sys.exit(0 if overall != "FAIL" else 1)


if __name__ == "__main__":
    run_smoke_test()
