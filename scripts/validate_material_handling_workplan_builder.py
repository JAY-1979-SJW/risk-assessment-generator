"""
차량계 하역운반기계 작업계획서 builder 검증 스크립트.

검증 항목:
1. bytes 생성 성공 (빈 form_data)
2. openpyxl 재열기 성공
3. 시트명 일치 (차량계하역운반기계작업계획서)
4. 제목/부제 존재
5. 법정 필수 라벨 존재 (LEGAL_LABELS 전수)
6. 운행경로 텍스트 영역 존재 (Row 16)
7. 스케치 박스 헤더 존재 (Row 17)
8. 제179조 점검 6항목 존재
9. print_area 존재 및 A1:H51 일치
10. 빈 form_data에서 생성 가능
11. 샘플 form_data에서 주요 값 정상 렌더링
12. 총 행 수 검증 (51행)
13. 보행자 동선 분리 라벨 존재 (하역운반기계 전용)
14. 기계의 최대 하중 라벨 존재 (machine_max_load, ≠ machine_capacity)
"""
from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

# 프로젝트 루트 경로 추가
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from engine.output.material_handling_workplan_builder import (
    ARTICLE_179_CHECKS,
    LEGAL_LABELS,
    SHEET_NAME,
    SHEET_HEADING,
    SHEET_SUBTITLE,
    build_material_handling_workplan_excel,
)

# ---------------------------------------------------------------------------
# 검증 유틸
# ---------------------------------------------------------------------------

_PASS = 0
_FAIL = 0


def _check(label: str, condition: bool, detail: str = "") -> None:
    global _PASS, _FAIL
    status = "PASS" if condition else "FAIL"
    if condition:
        _PASS += 1
    else:
        _FAIL += 1
    suffix = f"  ({detail})" if detail else ""
    print(f"  [{status}] {label}{suffix}")


def _all_cell_values(ws) -> set[str]:
    """워크시트 전체 셀 값을 문자열 set으로 반환."""
    values: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                values.add(str(cell.value))
    return values


# ---------------------------------------------------------------------------
# 샘플 데이터
# ---------------------------------------------------------------------------

SAMPLE_FORM_DATA = {
    "site_name":     "해한건설 주식회사",
    "project_name":  "○○물류센터 신축공사",
    "work_location": "지하1층 하역장 구역",
    "work_date":     "2026-04-25 ~ 2026-04-30",
    "supervisor":    "김안전",
    "contractor":    "(주)안전운반",
    "prepared_by":   "이계획",
    # 법정 필수 — 제38조 제2항
    "machine_type":      "지게차 (전동식, 3.5톤)",
    "machine_max_load":  "3,500 kg",
    "machine_count":     "2대",
    "operator_name":     "박운전",
    "operator_license":  "지게차운전기능사",
    "work_method":       "팔레트 적재물을 하역장에서 창고 선반으로 순차 운반",
    "guide_worker_required": "필요 — 후방 사각지대 발생 시 유도자 1명 배치",
    "speed_limit":       "구내 최고속도 10 km/h 이하",
    "ground_survey":     "노면 상태 양호, 경사도 없음, 습기 없음 확인",
    "travel_route_text": "하역장(A구역) → 내부 통로 → 창고 선반(B구역), 보행자 동선과 분리된 전용 통로 이용",
    "access_control":    "운반 작업 중 작업 구역 출입 통제, 안전 표지 설치",
    "emergency_contact": "현장소장 010-1234-5678",
    "emergency_measure": "부상자 발생 시 119 신고, 작업 중단, 현장소장 보고",
    "pedestrian_separation": "작업 구역 내 보행자 전용 동선 황색 라인 도색, 지게차 운행 중 보행자 진입 금지",
    "hazard_items": [
        {"hazard": "지게차 후진 시 보행자 충돌", "safety_measure": "후방 경보음 작동, 유도자 배치"},
        {"hazard": "적재물 낙하",               "safety_measure": "적재물 고정 확인, 과적 금지"},
        {"hazard": "지게차 전복",               "safety_measure": "최대 하중 준수, 급선회 금지"},
    ],
    "sign_date": "2026-04-25",
}

# ---------------------------------------------------------------------------
# 검증 실행
# ---------------------------------------------------------------------------

def run_validation() -> None:
    print("=" * 60)
    print("차량계 하역운반기계 작업계획서 Builder 검증")
    print("=" * 60)

    # ── TEST 1: 빈 form_data bytes 생성 ──────────────────────────
    print("\n[1] 빈 form_data 생성 테스트")
    empty_bytes = None
    try:
        empty_bytes = build_material_handling_workplan_excel({})
        _check("bytes 생성 성공 (빈 form_data)", len(empty_bytes) > 0,
               f"{len(empty_bytes):,} bytes")
    except Exception as e:
        _check("bytes 생성 성공 (빈 form_data)", False, str(e))

    # ── TEST 2: openpyxl 재열기 ───────────────────────────────────
    print("\n[2] openpyxl 재열기 테스트")
    ws_empty = None
    if empty_bytes:
        try:
            wb = load_workbook(BytesIO(empty_bytes))
            _check("openpyxl 재열기 성공", True)
            _check("시트명 일치", SHEET_NAME in wb.sheetnames, SHEET_NAME)
            ws_empty = wb[SHEET_NAME]
        except Exception as e:
            _check("openpyxl 재열기 성공", False, str(e))

    # ── TEST 3: 제목/부제 ─────────────────────────────────────────
    print("\n[3] 제목 / 부제 확인")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        _check("제목 존재", SHEET_HEADING in vals, SHEET_HEADING)
        _check("부제 존재", SHEET_SUBTITLE in vals, SHEET_SUBTITLE)

    # ── TEST 4: 법정 필수 라벨 ────────────────────────────────────
    print("\n[4] 법정 필수 라벨 전수 확인")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        for label in LEGAL_LABELS:
            _check(f"라벨 존재: {label}", label in vals)

    # ── TEST 5: 운행경로 텍스트 영역 (Row 16) ─────────────────────
    print("\n[5] 운행경로 텍스트 영역 (Row 16)")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        _check("운행경로 기술 라벨 존재", "운행경로 기술" in vals)
        # Row 16에 실제로 라벨 셀이 있는지
        label_cell = ws_empty.cell(row=16, column=1).value
        _check("Row 16 A열 라벨 = 운행경로 기술", label_cell == "운행경로 기술",
               str(label_cell))

    # ── TEST 6: 스케치 박스 헤더 (Row 17) ────────────────────────
    print("\n[6] 스케치 박스 헤더 (Row 17)")
    if ws_empty:
        row17_val = ws_empty.cell(row=17, column=1).value or ""
        _check("Row 17 스케치 헤더 존재", "운행경로" in str(row17_val) or "스케치" in str(row17_val),
               str(row17_val)[:60])

    # ── TEST 7: 제179조 점검 6항목 ────────────────────────────────
    print("\n[7] 제179조 점검 6항목 확인")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        for item in ARTICLE_179_CHECKS:
            _check(f"점검항목: {item[:20]}…", item in vals)

    # ── TEST 8: print_area ────────────────────────────────────────
    print("\n[8] print_area 확인")
    if ws_empty:
        pa = ws_empty.print_area
        _check("print_area 존재", bool(pa), str(pa))
        # openpyxl은 시트명 접두사와 $절대주소 형식으로 반환 — 핵심 범위만 비교
        _check("print_area = A1:H51", "A1:H51" in str(pa).replace("$", ""), str(pa))

    # ── TEST 9: 보행자 동선 분리 (하역운반기계 전용) ──────────────
    print("\n[9] 보행자 동선 분리 라벨 (하역운반기계 전용)")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        _check("보행자 동선 분리 라벨 존재", "보행자 동선 분리" in vals)

    # ── TEST 10: 기계의 최대 하중 (≠ machine_capacity) ───────────
    print("\n[10] 기계의 최대 하중 (machine_max_load)")
    if ws_empty:
        vals = _all_cell_values(ws_empty)
        _check("기계의 최대 하중 라벨 존재", "기계의 최대 하중" in vals)
        _check("기계의 성능·최대작업능력 미포함 (건설기계 혼용 금지)",
               "기계의 성능·최대작업능력" not in vals)

    # ── TEST 11: 총 행 수 (51행) ──────────────────────────────────
    print("\n[11] 총 행 수 확인")
    if ws_empty:
        max_row = ws_empty.max_row
        _check("총 행 수 = 51", max_row == 51, f"max_row={max_row}")

    # ── TEST 12: 샘플 form_data 렌더링 ───────────────────────────
    print("\n[12] 샘플 form_data 렌더링 테스트")
    sample_bytes = None
    try:
        sample_bytes = build_material_handling_workplan_excel(SAMPLE_FORM_DATA)
        _check("bytes 생성 성공 (샘플 form_data)", len(sample_bytes) > 0,
               f"{len(sample_bytes):,} bytes")
    except Exception as e:
        _check("bytes 생성 성공 (샘플 form_data)", False, str(e))

    if sample_bytes:
        try:
            wb2 = load_workbook(BytesIO(sample_bytes))
            ws2 = wb2[SHEET_NAME]
            vals2 = _all_cell_values(ws2)
            _check("site_name 렌더링",    SAMPLE_FORM_DATA["site_name"]    in vals2)
            _check("machine_type 렌더링", SAMPLE_FORM_DATA["machine_type"] in vals2)
            _check("machine_max_load 렌더링", SAMPLE_FORM_DATA["machine_max_load"] in vals2)
            _check("travel_route_text 렌더링", SAMPLE_FORM_DATA["travel_route_text"] in vals2)
            _check("pedestrian_separation 렌더링", SAMPLE_FORM_DATA["pedestrian_separation"] in vals2)
            _check("hazard 1번 렌더링", "지게차 후진 시 보행자 충돌" in vals2)
        except Exception as e:
            _check("샘플 재열기", False, str(e))

    # ── 최종 결과 ─────────────────────────────────────────────────
    total = _PASS + _FAIL
    print("\n" + "=" * 60)
    print(f"결과: PASS {_PASS} / FAIL {_FAIL} / 합계 {total}")
    if _FAIL == 0:
        print("최종 판정: PASS ✓")
    else:
        print("최종 판정: FAIL ✗")
    print("=" * 60)

    return _FAIL == 0


if __name__ == "__main__":
    ok = run_validation()
    sys.exit(0 if ok else 1)
