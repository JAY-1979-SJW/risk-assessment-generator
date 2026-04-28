"""Excel form quality audit script.

Iterates all registered form_types, generates sample xlsx, and scores
each file against structural/compliance quality criteria.

Usage:
    python scripts/audit_excel_form_quality.py
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import engine.output.form_registry as regmod  # noqa: E402

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
CATALOG_PATH  = ROOT / "data/masters/safety/documents/document_catalog.yml"
JSON_OUT_DIR  = ROOT / "runs/form_quality_audit"
JSON_OUT_PATH = JSON_OUT_DIR / "latest_quality_audit.json"
MD_OUT_PATH   = ROOT / "docs/reports/excel_form_quality_audit.md"

# ---------------------------------------------------------------------------
# Dangerous phrases that imply official submission
# ---------------------------------------------------------------------------
OFFICIAL_PHRASES = [
    "공식 제출용", "별지 제출", "별지 서식으로 제출", "관할 노동관서 제출",
]
LEGAL_MANDATORY_PHRASE = "법정 기재사항"


# ---------------------------------------------------------------------------
# Catalog helpers
# ---------------------------------------------------------------------------
def _load_catalog() -> Dict[str, Dict[str, Any]]:
    with open(CATALOG_PATH, encoding="utf-8") as f:
        cat = yaml.safe_load(f)
    return {
        d["form_type"]: d
        for d in cat["documents"]
        if d.get("form_type") and d.get("implementation_status") == "DONE"
    }


# ---------------------------------------------------------------------------
# Sample data generator
# ---------------------------------------------------------------------------
def _make_sample_data(spec: regmod.FormSpec) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for field in spec.required_fields:
        if "date" in field or "year" in field:
            data[field] = "2025-01-01"
        elif "hours" in field or "qty" in field or "count" in field:
            data[field] = 1
        else:
            data[field] = f"[{field}]"
    return data


# ---------------------------------------------------------------------------
# Quality checks
# ---------------------------------------------------------------------------
def _all_cell_values(ws) -> List[str]:
    values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return values


def _check_xlsx(
    ws,
    form_type: str,
    legal_status: str,
) -> Tuple[List[str], List[str]]:
    """Return (fails, warns) lists."""
    fails: List[str] = []
    warns: List[str] = []

    all_vals = _all_cell_values(ws)
    joined   = "\n".join(all_vals)

    # ── FAIL checks ──────────────────────────────────────────────────────────
    for phrase in OFFICIAL_PHRASES:
        if phrase in joined:
            fails.append(f"공식 제출용 오해 문구 포함: '{phrase}'")

    if legal_status == "practical" and LEGAL_MANDATORY_PHRASE in joined:
        fails.append(f"practical 문서에 '{LEGAL_MANDATORY_PHRASE}' 문구 포함")

    # ── Sheet name ────────────────────────────────────────────────────────────
    if len(ws.title) > 31:
        fails.append(f"시트명 31자 초과: {len(ws.title)}자")

    # ── WARN checks ──────────────────────────────────────────────────────────
    # Title row
    first_val = str(ws.cell(1, 1).value or "")
    if not first_val.strip():
        warns.append("1행 제목 셀 비어 있음")

    # Subtitle / notice row
    second_val = str(ws.cell(2, 1).value or "")
    if not second_val.strip():
        warns.append("2행 부제목/notice 셀 비어 있음")

    # Basic info section keyword
    info_keywords = ["기본 정보", "기본정보", "현장명", "사업장명", "공사명"]
    if not any(k in joined for k in info_keywords):
        warns.append("기본 정보 섹션 없음")

    # Signature section
    sig_keywords = ["서명", "확인", "승인자", "작성자"]
    if not any(k in joined for k in sig_keywords):
        warns.append("서명/확인란 없음")

    # Print setup
    try:
        has_print = (
            ws.page_setup.orientation is not None
            or ws.page_margins.left != 0.75
        )
    except Exception:
        has_print = False
    if not has_print:
        warns.append("인쇄 설정(orientation/fitToPage/margins) 없음")

    # Column widths
    if not ws.column_dimensions:
        warns.append("열 너비 미설정")

    # Merge cells
    if not ws.merged_cells.ranges:
        warns.append("병합 셀 없음 (레이아웃 확인 권장)")

    # Empty cell ratio
    total_cells = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 1
    empty_count = sum(
        1 for row in ws.iter_rows()
        for cell in row if cell.value is None
    )
    empty_ratio = empty_count / total_cells
    if empty_ratio > 0.85:
        warns.append(f"빈 셀 비율 과다: {empty_ratio:.0%}")

    # Section count (▶ marker)
    section_count = sum(1 for v in all_vals if v.startswith("▶"))
    if section_count < 2:
        warns.append(f"섹션 수 부족: {section_count}개")

    # Repeating table presence (look for numbered rows: 1, 2, 3...)
    num_cells = [v for v in all_vals if v in ("1", "2", "3")]
    if len(num_cells) < 2:
        warns.append("반복 테이블(번호 행) 없음")

    return fails, warns


# ---------------------------------------------------------------------------
# Main audit loop
# ---------------------------------------------------------------------------
def run_audit() -> Dict[str, Any]:
    catalog = _load_catalog()
    registry = regmod._REGISTRY

    results: List[Dict[str, Any]] = []
    pass_count = warn_count = fail_count = 0

    for form_type, spec in sorted(registry.items()):
        cat_entry = catalog.get(form_type, {})
        legal_status = cat_entry.get("legal_status", "unknown")
        display_name = spec.display_name

        rec: Dict[str, Any] = {
            "form_type": form_type,
            "display_name": display_name,
            "legal_status": legal_status,
            "verdict": "PASS",
            "fails": [],
            "warns": [],
            "xlsx_bytes": 0,
        }

        # Step 1: build xlsx
        try:
            sample_data = _make_sample_data(spec)
            raw = spec.builder(sample_data)
            rec["xlsx_bytes"] = len(raw)
        except Exception as e:
            rec["fails"].append(f"xlsx 생성 실패: {e}")
            rec["verdict"] = "FAIL"
            results.append(rec)
            fail_count += 1
            continue

        # Step 2: load_workbook
        try:
            import io
            wb = openpyxl.load_workbook(io.BytesIO(raw))
            ws = wb.active
        except Exception as e:
            rec["fails"].append(f"load_workbook 실패: {e}")
            rec["verdict"] = "FAIL"
            results.append(rec)
            fail_count += 1
            continue

        # Step 3: quality checks
        fails, warns = _check_xlsx(ws, form_type, legal_status)
        rec["fails"].extend(fails)
        rec["warns"].extend(warns)

        if rec["fails"]:
            rec["verdict"] = "FAIL"
            fail_count += 1
        elif rec["warns"]:
            rec["verdict"] = "WARN"
            warn_count += 1
        else:
            rec["verdict"] = "PASS"
            pass_count += 1

        results.append(rec)

    summary = {
        "run_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total": len(results),
        "pass": pass_count,
        "warn": warn_count,
        "fail": fail_count,
        "results": results,
    }
    return summary


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------
def _write_json(summary: Dict[str, Any]) -> None:
    JSON_OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(JSON_OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)


def _write_markdown(summary: Dict[str, Any]) -> None:
    MD_OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    results = summary["results"]

    fails = [r for r in results if r["verdict"] == "FAIL"]
    warns = [r for r in results if r["verdict"] == "WARN"]

    # Collect common warn categories
    warn_counter: Dict[str, int] = {}
    for r in results:
        for w in r["warns"]:
            key = w.split(":")[0]
            warn_counter[key] = warn_counter.get(key, 0) + 1

    lines = [
        "# Excel 서식 품질 점검 보고서",
        "",
        f"생성일시: {summary['run_at']}",
        "",
        "## 전체 요약",
        "",
        f"| 항목 | 건수 |",
        f"|------|------|",
        f"| 전체 점검 | {summary['total']} |",
        f"| PASS | {summary['pass']} |",
        f"| WARN | {summary['warn']} |",
        f"| FAIL | {summary['fail']} |",
        "",
        "## FAIL 목록",
        "",
    ]

    if fails:
        for r in fails:
            lines.append(f"### {r['form_type']} — {r['display_name']} ({r['legal_status']})")
            for msg in r["fails"]:
                lines.append(f"- **FAIL**: {msg}")
            lines.append("")
    else:
        lines += ["FAIL 없음", ""]

    lines += ["## WARN 목록", ""]
    if warns:
        for r in warns:
            lines.append(f"### {r['form_type']} — {r['display_name']} ({r['legal_status']})")
            for msg in r["warns"]:
                lines.append(f"- WARN: {msg}")
            lines.append("")
    else:
        lines += ["WARN 없음", ""]

    lines += ["## 공통 문제 (WARN 빈도)", ""]
    for msg, cnt in sorted(warn_counter.items(), key=lambda x: -x[1]):
        lines.append(f"- {msg}: {cnt}건")
    lines.append("")

    lines += [
        "## 즉시 수정 필요",
        "",
        "- FAIL 항목 전체",
        "",
        "## 나중에 개선 가능",
        "",
        "- WARN 항목 (서식 품질 향상)",
        "",
        "## 다음 수정 우선순위",
        "",
        "1. FAIL 항목 → builder 수정 또는 catalog 재분류",
        "2. practical 문서 문구 오류 → SHEET_SUBTITLE 수정",
        "3. 공통 WARN(서명란/기본정보 누락) → 섹션 추가",
        "4. 인쇄 설정 누락 → _finalize_sheet 추가",
    ]

    with open(MD_OUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------
def _print_summary(summary: Dict[str, Any]) -> None:
    results = summary["results"]
    print("=" * 70)
    print(f"  Excel 서식 품질 점검  |  {summary['run_at']}")
    print("=" * 70)
    print(f"  전체: {summary['total']}종  |  PASS: {summary['pass']}  |  WARN: {summary['warn']}  |  FAIL: {summary['fail']}")
    print()

    fails = [r for r in results if r["verdict"] == "FAIL"]
    warns = [r for r in results if r["verdict"] == "WARN"]

    if fails:
        print("  [FAIL]")
        for r in fails:
            for msg in r["fails"]:
                print(f"    {r['form_type']}: {msg}")
        print()

    if warns:
        print("  [WARN 상위]")
        warn_counter: Dict[str, int] = {}
        for r in warns:
            for w in r["warns"]:
                key = w.split(":")[0]
                warn_counter[key] = warn_counter.get(key, 0) + 1
        for msg, cnt in sorted(warn_counter.items(), key=lambda x: -x[1])[:8]:
            print(f"    {msg}: {cnt}건")
        print()

    print(f"  JSON  → {JSON_OUT_PATH}")
    print(f"  보고서 → {MD_OUT_PATH}")
    print("=" * 70)
    verdict = "PASS" if summary["fail"] == 0 else "FAIL"
    print(f"  최종 판정: {verdict}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    summary = run_audit()
    _write_json(summary)
    _write_markdown(summary)
    _print_summary(summary)
    sys.exit(0 if summary["fail"] == 0 else 1)
