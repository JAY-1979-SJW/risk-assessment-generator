"""
trade_risk_recommendation payload → RA-001 위험성평가표 변환 어댑터 v1

주요 함수:
  build_ra001_input_from_trade_recommendation(recommendation)
      - payload → RA-001 builder 입력 dict 변환
  build_ra001_input_from_trade_id(trade_id, ...)
      - trade_id → recommender 호출 → RA-001 입력 반환
  build_ra001_excel_from_recommendation(recommendation)
      - payload → xlsx bytes (주의사항 시트 포함)
  validate_ra001_input(payload)
      - 필수 필드·참조 검사 → warning list

위험성 등급 v1 기본값:
  기본: 빈도 3 × 강도 3 = 위험도 9 (중)
  중대위험(FALL/CS/ELEC/HOT/LIFT): 빈도 3 × 강도 4 = 위험도 12 (높음)
  감소대책 후 공통: 빈도 2 × 강도 2 = 위험도 4 (낮음)
"""

from __future__ import annotations

import pathlib
import sys
from copy import deepcopy
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 프로젝트 루트 경로 보장
_REPO_ROOT = pathlib.Path(__file__).parent.parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from engine.output.form_excel_builder import render_form_sheet
from engine.recommendation.trade_risk_recommender import (
    build_trade_risk_recommendation,
    get_trade_preset,
    load_trade_risk_masters,
    merge_common_high_risk_presets,
)

# ──────────────────────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────────────────────

_DISCLAIMERS = [
    "본 위험성평가표는 공종별 프리셋 기반 초안이며, "
    "현장 조건에 따라 관리감독자 및 작업자가 검토·보완해야 한다.",
    "위험성 등급은 v1 기본 산정값이며, "
    "현장 실측·작업방법·인원·장비 조건에 따라 조정한다.",
    "관련 법령 및 서류는 현행 원문과 발주처/원청 기준을 확인 후 적용한다.",
]

# 중대위험 hazard_id → 강도 4 적용
_HIGH_SEVERITY_HAZARDS: frozenset[str] = frozenset((
    "FALL_FROM_HEIGHT",
    "CONFINED_SPACE_ASPHYXIATION",
    "ELECTRIC_SHOCK",
    "HOT_WORK_FIRE",
    "HEAVY_LIFTING",
))

_CATEGORY_KO: dict[str, str] = {
    "FALL":         "추락·낙하",
    "STRUCK_BY":    "비래·충돌",
    "CAUGHT_IN":    "협착·끼임",
    "FIRE_EXPLOSION": "화재·폭발",
    "ELECTRIC":     "전기",
    "HEALTH":       "건강장해",
    "ERGONOMIC":    "근골격계",
    "OTHER":        "기타",
}


# ──────────────────────────────────────────────────────────────
# 내부 헬퍼
# ──────────────────────────────────────────────────────────────

def _risk_scores(hazard_id: str) -> dict:
    if hazard_id in _HIGH_SEVERITY_HAZARDS:
        return {
            "probability": 3, "severity": 4, "risk_level": 12,
            "risk_grade": "높음",
            "residual_probability": 2, "residual_severity": 2,
            "residual_risk": 4, "residual_grade": "낮음",
        }
    return {
        "probability": 3, "severity": 3, "risk_level": 9,
        "risk_grade": "중",
        "residual_probability": 2, "residual_severity": 2,
        "residual_risk": 4, "residual_grade": "낮음",
    }


def _controls_text(controls: list[dict]) -> list[str]:
    lines = []
    for ctrl in controls:
        ctype = ctrl.get("control_type", "")
        pri = ctrl.get("priority", 5)
        desc = ctrl.get("description", "")
        cond = ctrl.get("applicable_conditions", "")
        line = f"[{ctype} P{pri}] {desc}"
        if cond:
            line += f"  ({cond})"
        lines.append(line)
    return lines


def _current_measures_text(controls: list[dict]) -> str:
    # priority 1~2의 ENGINEERING/ADMINISTRATIVE 조치를 현재 안전조치로 표시
    selected = [
        ctrl["description"]
        for ctrl in controls
        if ctrl.get("priority", 5) <= 2
        and ctrl.get("control_type") in ("ENGINEERING", "ADMINISTRATIVE", "ELIMINATION", "SUBSTITUTION")
    ]
    return "\n".join(selected) if selected else ""


def _category_ko(raw: str) -> str:
    return _CATEGORY_KO.get(raw.upper(), raw)


def _today_str() -> str:
    return date.today().isoformat()


# ──────────────────────────────────────────────────────────────
# 1. payload → RA-001 입력 변환
# ──────────────────────────────────────────────────────────────

def build_ra001_input_from_trade_recommendation(recommendation: dict) -> dict:
    """
    trade_risk_recommendation payload를 RA-001 builder 입력 구조로 변환.

    반환 dict 구조:
      - header keys (form_registry _RISK_ASSESSMENT_HEADER_KEYS):
          company_name, industry, site_name, representative,
          assessment_type, assessment_date, work_type
      - rows: list[dict] (form_excel_builder._COLUMNS 기준)
      - _meta: 추가 메타 (builder에 전달되지 않음, 검증/로깅용)
    """
    ctx = recommendation.get("site_context") or {}
    trade_id = recommendation.get("trade_id", "")
    trade_name = recommendation.get("trade_name", "")
    trade_group = recommendation.get("trade_group", "")
    work_location = ctx.get("work_location") or trade_name
    source_trace = recommendation.get("source_trace", [trade_id])

    # ── header ──────────────────────────────────────────────
    header = {
        "company_name": ctx.get("site_name") or "",
        "industry": "건설업",
        "site_name": ctx.get("site_name") or "",
        "representative": "",
        "assessment_type": "공종별 프리셋 초안 (v1)",
        "assessment_date": ctx.get("work_date") or _today_str(),
        "work_type": trade_name,
    }

    # ── rows ─────────────────────────────────────────────────
    rows: list[dict[str, Any]] = []
    risk_items = recommendation.get("risk_items", [])

    for idx, ri in enumerate(risk_items, start=1):
        hid = ri.get("hazard_id", "")
        scores = _risk_scores(hid)
        controls = ri.get("controls", [])
        related_docs = list(ri.get("related_documents", []))
        required_docs = list(recommendation.get("required_documents", []))
        all_docs_text = ", ".join(
            d for d in (related_docs + required_docs) if d
        ) or ""

        row: dict[str, Any] = {
            "no": idx,
            "process": trade_group,
            "sub_work": f"{trade_name}" + (f" / {work_location}" if work_location and work_location != trade_name else ""),
            "hazard_category_major": _category_ko(ri.get("category", "")),
            "hazard_category_minor": ri.get("hazard_name", hid),
            "hazard": (
                ri.get("hazard_name", "")
                + ("\n" + ri.get("risk_scenario", "") if ri.get("risk_scenario") else "")
            ),
            "legal_basis": all_docs_text,
            "current_measures": _current_measures_text(controls),
            "risk_scale": "빈도×강도",
            "probability": scores["probability"],
            "severity": scores["severity"],
            "risk_level": f"{scores['risk_level']} ({scores['risk_grade']})",
            "control_measures": _controls_text(controls),
            "residual_risk_level": f"{scores['residual_risk']} ({scores['residual_grade']})",
            "target_date": "",
            "completion_date": "",
            "responsible_person": "현장 관리감독자",
        }
        rows.append(row)

    # ── _meta ────────────────────────────────────────────────
    meta = {
        "trade_id": trade_id,
        "trade_name": trade_name,
        "trade_group": trade_group,
        "source_trace": source_trace,
        "source_status_summary": recommendation.get("source_status_summary", {}),
        "required_documents": recommendation.get("required_documents", []),
        "recommended_documents": recommendation.get("recommended_documents", []),
        "required_permits": recommendation.get("required_permits", []),
        "required_trainings": recommendation.get("required_trainings", []),
        "common_equipment": recommendation.get("common_equipment", []),
        "ppe": recommendation.get("ppe", []),
        "warnings": list(recommendation.get("warnings", [])),
        "disclaimers": _DISCLAIMERS,
        "risk_items_count": len(rows),
    }

    return {**header, "rows": rows, "_meta": meta}


# ──────────────────────────────────────────────────────────────
# 2. trade_id → RA-001 입력 변환
# ──────────────────────────────────────────────────────────────

def build_ra001_input_from_trade_id(
    trade_id: str,
    common_work_ids: list[str] | None = None,
    site_context: dict | None = None,
) -> dict:
    """
    trade_id → recommender → RA-001 입력 dict.
    common_work_ids가 있으면 공통 고위험작업을 merge한다.
    """
    if common_work_ids:
        base = get_trade_preset(trade_id)
        recommendation = merge_common_high_risk_presets(base, common_work_ids)
        if site_context:
            recommendation["site_context"] = {
                "site_name": site_context.get("site_name"),
                "work_location": site_context.get("work_location"),
                "work_date": site_context.get("work_date"),
                "workers_count": site_context.get("workers_count"),
                "equipment_used": site_context.get("equipment_used", []),
            }
    else:
        recommendation = build_trade_risk_recommendation(trade_id, site_context=site_context)

    return build_ra001_input_from_trade_recommendation(recommendation)


# ──────────────────────────────────────────────────────────────
# 3. xlsx 생성
# ──────────────────────────────────────────────────────────────

def build_ra001_excel_from_recommendation(recommendation: dict) -> bytes:
    """
    trade_risk_recommendation payload → xlsx bytes.
    시트 구성:
      Sheet 1: 위험성평가표 (RA-001 본표)
      Sheet 2: 주의사항 (disclaimer 3개 + 필수서류/교육/허가서)
    """
    ra001_input = build_ra001_input_from_trade_recommendation(recommendation)
    return _build_excel_bytes(ra001_input)


def build_ra001_excel_from_trade_id(
    trade_id: str,
    common_work_ids: list[str] | None = None,
    site_context: dict | None = None,
) -> bytes:
    """
    trade_id → xlsx bytes (편의 함수).
    """
    ra001_input = build_ra001_input_from_trade_id(
        trade_id, common_work_ids=common_work_ids, site_context=site_context
    )
    return _build_excel_bytes(ra001_input)


def _build_excel_bytes(ra001_input: dict) -> bytes:
    meta = ra001_input.get("_meta", {})

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "위험성평가표"

    # ── Sheet 1: 본표 ──────────────────────────────────────
    # render_form_sheet expects {"header": {...}, "rows": [...]}
    header_keys = {
        "company_name", "industry", "site_name", "representative",
        "assessment_type", "assessment_date", "work_type",
    }
    form_data_for_render = {
        "header": {k: ra001_input.get(k) for k in header_keys},
        "rows": ra001_input.get("rows", []),
    }
    render_form_sheet(ws_main, form_data_for_render)

    # ── Sheet 2: 주의사항 ──────────────────────────────────
    ws_notes = wb.create_sheet(title="주의사항")
    _write_notes_sheet(ws_notes, meta)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_notes_sheet(ws, meta: dict) -> None:
    FONT_TITLE = Font(name="맑은 고딕", size=12, bold=True)
    FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
    FONT_BODY = Font(name="맑은 고딕", size=10)
    FILL_WARN = PatternFill(fill_type="solid", fgColor="FFF2CC")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 80

    row = 1
    ws.cell(row=row, column=1, value="주의사항 및 필수 확인 사항").font = FONT_TITLE
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 24

    row += 1

    # 3개 필수 고정 문구
    for i, disc in enumerate(meta.get("disclaimers", _DISCLAIMERS), start=1):
        c_no = ws.cell(row=row, column=1, value=f"【주의 {i}】")
        c_no.font = FONT_BOLD
        c_no.fill = FILL_WARN
        c_no.alignment = ALIGN_CENTER
        c_body = ws.cell(row=row, column=2, value=disc)
        c_body.font = FONT_BODY
        c_body.fill = FILL_WARN
        c_body.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1

    # 필수 서류
    _write_notes_section(ws, row, "필수 서류", meta.get("required_documents", []), FONT_BOLD, FONT_BODY)
    row += len(meta.get("required_documents", [])) + 2

    # 필수 허가서
    _write_notes_section(ws, row, "필수 작업허가서", meta.get("required_permits", []), FONT_BOLD, FONT_BODY)
    row += len(meta.get("required_permits", [])) + 2

    # 필수 교육
    _write_notes_section(ws, row, "필수 교육", meta.get("required_trainings", []), FONT_BOLD, FONT_BODY)
    row += len(meta.get("required_trainings", [])) + 2

    # source_status_summary
    ss = meta.get("source_status_summary", {})
    if ss:
        ws.cell(row=row, column=1, value="검증상태").font = FONT_BOLD
        ws.cell(row=row, column=2, value=str(ss)).font = FONT_BODY
        row += 1

    # warnings
    warnings = meta.get("warnings", [])
    if warnings:
        row += 1
        ws.cell(row=row, column=1, value="경고").font = FONT_BOLD
        row += 1
        for w in warnings:
            ws.cell(row=row, column=2, value=w).font = FONT_BODY
            row += 1


def _write_notes_section(ws, start_row: int, label: str, items: list, font_bold, font_body) -> None:
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.cell(row=start_row, column=1, value=label).font = font_bold
    if not items:
        ws.cell(row=start_row, column=2, value="(없음)").font = font_body
        return
    ws.cell(row=start_row, column=2, value=items[0]).font = font_body
    for i, item in enumerate(items[1:], start=1):
        ws.cell(row=start_row + i, column=2, value=item).font = font_body


# ──────────────────────────────────────────────────────────────
# 4. RA-001 입력 검증
# ──────────────────────────────────────────────────────────────

def validate_ra001_input(payload: dict) -> list[str]:
    """
    RA-001 입력 dict 검사. warning 메시지 리스트 반환.
    """
    warnings: list[str] = []
    masters = load_trade_risk_masters()
    valid_doc_ids = masters["valid_doc_ids"]
    valid_training_codes = masters["valid_training_codes"]

    rows = payload.get("rows", [])

    # risk rows 1개 이상
    if not rows:
        warnings.append("[NO_ROWS] risk_rows가 0개입니다.")

    for i, row in enumerate(rows, start=1):
        if not row.get("hazard_category_minor"):
            warnings.append(f"[ROW_{i}] 위험요인명(hazard_category_minor) 누락")
        if not row.get("hazard"):
            warnings.append(f"[ROW_{i}] 위험상황(hazard) 누락")
        if not row.get("control_measures"):
            warnings.append(f"[ROW_{i}] 감소대책(control_measures) 누락")

    meta = payload.get("_meta", {})

    # document_id 참조 검증
    if valid_doc_ids:
        for did in meta.get("required_documents", []):
            if did not in valid_doc_ids:
                warnings.append(f"[UNKNOWN_DOC] required_documents '{did}' catalog에 없음")
        for did in meta.get("required_permits", []):
            if did not in valid_doc_ids:
                warnings.append(f"[UNKNOWN_PERMIT] required_permits '{did}' catalog에 없음")

    # training_code 참조 검증
    if valid_training_codes:
        for tcode in meta.get("required_trainings", []):
            if tcode not in valid_training_codes:
                warnings.append(f"[UNKNOWN_TRAINING] required_trainings '{tcode}' training_types에 없음")

    # source_status NEEDS_VERIFICATION 경고
    ss = meta.get("source_status_summary", {})
    if ss.get("NEEDS_VERIFICATION", 0) > 0:
        warnings.append(
            f"[NEEDS_VERIFY] source_status NEEDS_VERIFICATION {ss['NEEDS_VERIFICATION']}건 "
            f"— skeleton 공종 포함 가능, 사용 전 법령 검토 필요"
        )

    # 현장 정보 미입력 항목
    for field in ("company_name", "site_name", "assessment_date"):
        if not payload.get(field):
            warnings.append(f"[SITE_INFO] '{field}' 미입력 — 빈칸으로 출력됨")

    # disclaimer 포함 여부
    for disc in _DISCLAIMERS:
        if disc not in meta.get("disclaimers", []):
            warnings.append(f"[NO_DISCLAIMER] 필수 고정 문구 누락: '{disc[:30]}...'")

    return warnings
