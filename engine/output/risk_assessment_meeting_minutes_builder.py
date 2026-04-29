"""
위험성평가 참여 회의록 — Excel 출력 모듈 (v1.0).

법적 근거: 산업안전보건법 제36조 제3항 (근로자 참여 보장)
분류: RA-003 — 법정 서식 없음, 근로자 참여 증빙 서식

Optional form_data keys:
    site_name           str  사업장명
    project_name        str  현장명
    representative      str  사업주 / 현장 대표
    safety_manager      str  안전관리자

    meeting_date        str  회의 일시
    meeting_place       str  회의 장소
    work_type           str  대상 작업/공종
    meeting_purpose     str  회의 목적

    attendees  list[dict]  참석자 (max MAX_ATTENDEES)
        affiliation     str  소속
        position        str  직책
        name            str  성명

    discussion_items    str  위험성평가 주요 논의사항
    worker_opinions     str  근로자 의견 및 건의사항

    action_items  list[dict]  위험요인 및 개선대책 (max MAX_ACTIONS)
        seq             int/str  순번
        hazard          str      위험요인
        measure         str      개선대책
        assignee        str      조치 담당자
        due_date        str      조치 예정일
        status          str      조치 상태

    meeting_summary     str  회의 결과 요약

    prepared_by     str  작성자
    reviewed_by     str  검토자
    approved_by     str  승인자
    prepared_date   str  작성일
    reviewed_date   str  검토일
    approved_date   str  승인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "위험성평가참여회의록"
SHEET_HEADING  = "위험성평가 참여 회의록"
SHEET_SUBTITLE = "산업안전보건법 제36조 제3항 — 위험성평가 시 근로자 참여 보장"

MAX_ATTENDEES = 15
MAX_ACTIONS   = 10
TOTAL_COLS    = 10

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 6,   # 순번
    2: 12,  # 소속 / 위험요인
    3: 12,  # 직책 / 개선대책
    4: 12,  # 성명
    5: 12,  # 서명 / 조치 담당자
    6: 10,  # 조치 예정일
    7: 10,  # 조치 상태
    8: 10,
    9: 10,
    10: 11,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col1: int, label_col2: int,
              val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col1, label_col2, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_header_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"),    1, 1, 2, 5)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), 6, 6, 7, 10)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "사업주",  _v(data, "representative"), 1, 2, 3, 5)
    _write_lv(ws, r, "안전관리자", _v(data, "safety_manager"), 6, 7, 8, 10)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_meeting_overview(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "회의 개요",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "회의 일시", _v(data, "meeting_date"),    1, 2, 3, 5)
    _write_lv(ws, r, "회의 장소", _v(data, "meeting_place"),   6, 7, 8, 10)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "대상 작업/공종", _v(data, "work_type"),      1, 2, 3, 5)
    _write_lv(ws, r, "회의 목적",    _v(data, "meeting_purpose"),  6, 7, 8, 10)
    ws.row_dimensions[r].height = 20; r += 1
    return r


def _write_attendees_table(ws, start_row: int,
                            attendees: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "참석자 명단",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    # 헤더: 순번(1) / 소속(2-3) / 직책(4-5) / 성명(6-7) / 서명(8-10)
    for (c1, c2), h in (
        ((1, 1), "순번"),
        ((2, 3), "소속"),
        ((4, 5), "직책"),
        ((6, 7), "성명"),
        ((8, 10), "서명"),
    ):
        _write_cell(ws, r, c1, c2, h,
                    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=24)
    r += 1

    for i in range(MAX_ATTENDEES):
        entry = attendees[i] if i < len(attendees) else {}
        _write_cell(ws, r, 1,  1,  _v(entry, "seq") or (i + 1), font=_FONT_SMALL, align=_ALIGN_CENTER, height=22)
        _write_cell(ws, r, 2,  3,  _v(entry, "affiliation"),     font=_FONT_SMALL, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4,  5,  _v(entry, "position"),        font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 6,  7,  _v(entry, "name"),            font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8,  10, "",                           font=_FONT_SMALL, align=_ALIGN_CENTER)
        r += 1

    return r


def _write_text_section(ws, start_row: int,
                         section_title: str, value: Any) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, section_title,
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, TOTAL_COLS, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=54)
    r += 1
    return r


def _write_action_items_table(ws, start_row: int,
                               action_items: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "위험요인 및 개선대책 반영사항",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    headers = ["순번", "위험요인", "개선대책", "조치 담당자", "조치 예정일", "조치 상태"]
    col_spans = [(1, 1), (2, 3), (4, 6), (7, 8), (9, 9), (10, 10)]
    for (c1, c2), h in zip(col_spans, headers):
        _write_cell(ws, r, c1, c2, h,
                    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=24)
    r += 1

    for i in range(MAX_ACTIONS):
        entry = action_items[i] if i < len(action_items) else {}
        _write_cell(ws, r, 1,  1,  _v(entry, "seq") or (i + 1), font=_FONT_SMALL, align=_ALIGN_CENTER, height=22)
        _write_cell(ws, r, 2,  3,  _v(entry, "hazard"),    font=_FONT_SMALL, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4,  6,  _v(entry, "measure"),   font=_FONT_SMALL, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7,  8,  _v(entry, "assignee"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 9,  9,  _v(entry, "due_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 10, 10, _v(entry, "status"),    font=_FONT_SMALL, align=_ALIGN_CENTER)
        r += 1

    return r


def _write_signature_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작성 / 검토 / 승인",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 1,  "구분",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  "작성자", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  "검토자", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, "승인자", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20; r += 1

    _write_cell(ws, r, 1, 1,  "성명",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  _v(data, "prepared_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  _v(data, "reviewed_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, _v(data, "approved_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 28; r += 1

    _write_cell(ws, r, 1, 1,  "서명",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36; r += 1

    _write_cell(ws, r, 1, 1,  "일자",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  _v(data, "prepared_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  _v(data, "reviewed_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, _v(data, "approved_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20; r += 1

    return r


def build_risk_assessment_meeting_minutes_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 위험성평가 참여 회의록 xlsx 바이너리를 반환."""
    data         = form_data or {}
    attendees    = data.get("attendees")    or []
    action_items = data.get("action_items") or []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_header_block(ws, row, data)
    row = _write_meeting_overview(ws, row, data)
    row = _write_attendees_table(ws, row, attendees)
    row = _write_text_section(ws, row, "위험성평가 주요 논의사항",    _v(data, "discussion_items"))
    row = _write_text_section(ws, row, "근로자 의견 및 건의사항",      _v(data, "worker_opinions"))
    row = _write_action_items_table(ws, row, action_items)
    row = _write_text_section(ws, row, "회의 결과 요약",               _v(data, "meeting_summary"))
    _write_signature_block(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
