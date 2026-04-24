"""
굴착 작업계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제38조, 제82조 제1항
Layout:   docs/design/workplan_excel_layout_lock.md
Plan:     docs/design/workplan_builder_plan.md

Input — form_data dict:
    site_name            str|None   사업장명 [관행]
    project_name         str|None   현장명 [관행]
    work_location        str|None   작업 위치 [관행]
    work_date            str|None   작업 일자/기간 [관행]
    supervisor           str|None   작업 책임자 [관행]
    contractor           str|None   도급업체 [관행]
    excavation_method    str|None   굴착의 방법 [법정] 제82조 제1항 제1호
    earth_retaining      str|None   흙막이 지보공 및 방호망 [법정] 제82조 제1항 제2호
    excavation_machine   str|None   사용 기계 종류 및 능력 [법정] 제82조 제1항 제3호
    soil_disposal        str|None   토석 처리 방법 [법정] 제82조 제1항 제4호
    water_disposal       str|None   용수 처리 방법 [법정] 제82조 제1항 제5호
    work_method          str|None   작업 방법 [법정] 제38조 제2항
    emergency_measure    str|None   긴급조치 계획 [법정] 제38조 제2항
    guide_worker_required str|None  유도자 배치 여부 및 방법 [법정] 기준규칙 제172조, 제38조 제2항
    access_control       str|None   출입통제 방법 [법정] 기준규칙 제38조 제1항 제6호
    emergency_contact    str|None   비상연락망 [실무] 기준규칙 제38조 제2항 세부
    safety_steps         list[dict] 작업단계별 안전조치 (최대 MAX_STEPS=10)
        task_step        str|None   작업 단계
        hazard           str|None   위험 요인
        safety_measure   str|None   안전 조치
        responsible_person str|None (입력 허용, 레이아웃 제약으로 미출력)
        note             str|None   (입력 허용, 레이아웃 제약으로 미출력)
    sign_date            str|None   작성일 (서명란) [관행]

Output — xlsx bytes (in-memory). 파일 저장은 호출자 책임.

Principles:
- 법정 외 임의 필드 추가 금지.
- null/누락 → 빈 셀. 임의값 생성 금지.
- 서명란은 항상 공란 (수기 서명 전용).
- 안전조치 행은 MAX_STEPS=10 고정 출력.
- safety_steps > 10 시 초과분 무시.
- 굴착 작업계획서 1종만 구현 (v1).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME     = "작업계획서"
SHEET_HEADING  = "굴착 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제38조·제82조에 따른 작업계획서"
MAX_STEPS      = 10
TOTAL_COLS     = 8

# 법정 항목 라벨 (검증 스크립트 재사용)
LEGAL_LABELS = [
    "굴착의 방법",
    "흙막이 지보공 및 방호망",
    "사용 기계 종류 및 능력",
    "토석 처리 방법",
    "용수 처리 방법",
    "작업 방법",
    "긴급조치 계획",
    "유도자 배치",
    "출입통제 방법",
    "비상연락망",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 열 너비 (A=14, B~G=12, H=10)
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 14,  # A: 법정항목 라벨 / 순번
    2: 12,  # B
    3: 12,  # C
    4: 12,  # D
    5: 12,  # E
    6: 12,  # F
    7: 12,  # G
    8: 10,  # H
}

# 메타 블록 스팬 상수
_L1, _V1_START, _V1_END       = 1, 2, 4   # 좌: 라벨=A(1), 값=B:D(2-4)
_L2, _V2_START, _V2_END       = 5, 6, 8   # 우: 라벨=E(5), 값=F:H(6-8)
_FULL_VAL_START, _FULL_VAL_END = 2, 8     # 전폭 값: B:H(2-8)


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    """dict에서 값 반환. None이면 빈 문자열."""
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or PatternFill()
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    """[라벨 셀] [값 셀(병합)] 쌍 기록."""
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    """Row 1: 제목, Row 2: 부제 → 다음 row 반환."""
    # Row 1 — 제목
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    # Row 2 — 부제
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 사업장명(A,B:D) | 현장명(E,F:H)
    Row +1: 작업위치(A,B:H 전폭)
    Row +2: 작업일자(A,B:D) | 작업책임자(E,F:H)
    Row +3: 도급업체(A,B:D) | 작성일 공란(E,F:H)
    """
    H = 20
    r = start_row

    # Row 3: 사업장명 / 현장명
    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    # Row 4: 작업 위치 (전폭)
    r += 1
    _write_lv(ws, r, "작업 위치", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H

    # Row 5: 작업 일자 / 작업 책임자
    r += 1
    _write_lv(ws, r, "작업 일자", _v(data, "work_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 책임자", _v(data, "supervisor"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    # Row 6: 도급업체 / 작성일(공란)
    r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성일", "",        # 항상 공란 — 수기 기입
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    return r + 1


def _write_legal_items(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더 (법정기재사항)
    Row +1~+7: 법정 항목 7개 (라벨=A, 값=B:H)
    """
    r = start_row

    # 섹션 헤더
    _write_cell(ws, r, 1, TOTAL_COLS,
                "법정 기재 사항 (산업안전보건기준에 관한 규칙 제38조·제82조)",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    legal_rows = [
        ("굴착의 방법",          "excavation_method",      40),
        ("흙막이 지보공 및 방호망", "earth_retaining",        40),
        ("사용 기계 종류 및 능력", "excavation_machine",     30),
        ("토석 처리 방법",        "soil_disposal",          30),
        ("용수 처리 방법",        "water_disposal",         30),
        ("작업 방법",            "work_method",            40),
        ("긴급조치 계획",         "emergency_measure",      40),
        ("유도자 배치",           "guide_worker_required",  30),
        ("출입통제 방법",         "access_control",         30),
        ("비상연락망",            "emergency_contact",      30),
    ]

    for label, field, h in legal_rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL,
                    align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, field),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_step_table(ws, start_row: int,
                      steps: List[Dict[str, Any]]) -> int:
    """
    Row +0: 섹션 헤더 (작업단계별 안전조치)
    Row +1: 표 헤더 (순번|작업단계|위험요인|안전조치)
    Row +2~+11: 데이터행 MAX_STEPS=10 고정
    """
    r = start_row

    # 섹션 헤더
    _write_cell(ws, r, 1, TOTAL_COLS, "작업단계별 안전조치",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 표 헤더
    _write_cell(ws, r, 1, 1, "순번",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 3, "작업 단계", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 4, 6, "위험 요인", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 8, "안전 조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    # 데이터행: MAX_STEPS=10 고정, 초과분 무시
    for i in range(MAX_STEPS):
        step = steps[i] if i < len(steps) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 3, _v(step, "task_step"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4, 6, _v(step, "hazard"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7, 8, _v(step, "safety_measure"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더 (확인및서명)
    Row +1: 서명 라벨행
    Row +2: 서명 공란 / 작성일
    """
    r = start_row

    # 섹션 헤더
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # Row 28: 서명 라벨행
    _write_cell(ws, r, 1, 2, "작성자",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",  # 서명 공란
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "검토자/확인자",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 8, "",  # 서명 공란
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20
    r += 1

    # Row 29: 서명 공란 + 작성일
    _write_cell(ws, r, 1, 2, "",  # 서명 공간
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 3, 4, "",  # 서명 공간
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "",  # 서명 공간
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 40

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_excavation_workplan_sheet(ws, form_data: Dict[str, Any]) -> None:
    """
    주어진 worksheet에 굴착 작업계획서를 렌더링.

    외부 호출자가 직접 ws에 접근해야 할 때 사용.
    build_excavation_workplan_excel 은 단일 워크북/시트 래퍼.
    """
    data  = form_data or {}
    steps = data.get("safety_steps") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_legal_items(ws, row, data)
    row = _write_step_table(ws, row, steps)
    _write_confirmation(ws, row, data)

    # 인쇄 설정
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_area              = "A1:H32"
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_excavation_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 굴착 작업계획서 xlsx 바이너리를 반환.

    Args:
        form_data: workplan_builder_plan.md 입력 스키마 준수 dict.
            safety_steps > MAX_STEPS(10) 시 초과분 무시.
            responsible_person · note 는 입력 허용, 출력 없음.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_excavation_workplan_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
