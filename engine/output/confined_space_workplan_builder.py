"""
밀폐공간 작업계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제619조~제626조 (밀폐공간 작업 프로그램)
분류: GEN_INTERNAL — 조문 기반 설계

Required form_data keys:
    confined_space_location str  밀폐공간 위치   [법정] 제619조
    gas_measurement_plan    str  산소·유해가스 측정계획 [법정] 제619조
    ventilation_plan        str  환기계획         [법정] 제619조
    emergency_measure       str  비상연락·구조절차  [법정] 제619조

Optional form_data keys:
    site_name, project_name, work_date, supervisor, contractor, prepared_by
    work_content      str  작업 내용
    worker_count      str  작업 인원 수
    confined_space_type str 밀폐공간 유형 (맨홀/탱크/피트 등)
    monitor_placement str  감시인 배치 [법정] 제619조
    rescue_equipment  str  구조장비 비치 현황 [법정] 제619조
    emergency_contact str  비상연락망
    access_control    str  출입 통제 방법 [법정] 제619조
    work_before_check str  작업 전 확인사항 [실무 권장]
    work_during_check str  작업 중 확인사항 [실무 권장]
    work_after_check  str  작업 후 확인사항 [실무 권장]
    sign_date         str  작성일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "밀폐공간작업계획서"
SHEET_HEADING  = "밀폐공간 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제619조에 따른 밀폐공간 작업계획서"

TOTAL_COLS = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 기간", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 책임자", _v(data, "supervisor"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자",  _v(data, "prepared_by"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_space_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "밀폐공간 정보  (기준규칙 제619조)",
                font=_FONT_BOLD, fill=_FILL_WARN, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "밀폐공간 위치", _v(data, "confined_space_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1

    _write_lv(ws, r, "공간 유형", _v(data, "confined_space_type"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 인원", _v(data, "worker_count"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_cell(ws, r, _L1, _L1, "작업 내용",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "work_content"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_safety_plan(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "안전 계획  (기준규칙 제619조 법정 기재사항)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    for label, key, h in [
        ("산소·유해가스\n측정계획", "gas_measurement_plan", 50),
        ("환기계획",    "ventilation_plan",   50),
        ("감시인 배치", "monitor_placement",  40),
        ("구조장비",    "rescue_equipment",   40),
        ("출입 통제",   "access_control",     30),
        ("비상연락망",  "emergency_contact",  30),
        ("비상조치",    "emergency_measure",  50),
    ]:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, key),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_checklist(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업 단계별 확인사항  [실무 권장]",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    for label, key, h in [
        ("작업 전 확인사항", "work_before_check", 50),
        ("작업 중 확인사항", "work_during_check", 50),
        ("작업 후 확인사항", "work_after_check",  50),
    ]:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, key),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 2, "작성자 서명",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",        font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20; r += 1
    _write_cell(ws, r, 1, 4, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36
    return r + 1


def build_confined_space_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 밀폐공간 작업계획서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_space_info(ws, row, data)
    row = _write_safety_plan(ws, row, data)
    row = _write_checklist(ws, row, data)
    _write_confirmation(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
