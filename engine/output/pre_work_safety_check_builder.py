"""
작업 전 안전 확인서 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건법 제38조 (안전조치)
    산업안전보건법 제36조 (위험성평가 및 조치사항)
    산업안전보건기준에 관한 규칙 별표 3 (작업시작 전 점검사항)

분류: PRACTICAL — 공식 별지 제출서식 아님
      작업 전 안전확인 및 작업개시 판단 보조서식
      산업안전보건법상 안전조치 및 위험성평가 조치사항과 연계

역할 분리:
    DL-001: 하루 단위 현장 안전관리 종합 일지
    DL-002: 관리감독자 법정 업무 수행 기록
    DL-003: 순찰 지적·시정조치 이행관리
    DL-004: 기상 조건 및 작업중지 판단 기록
    DL-005(본 서식): 작업 시작 직전 작업개시 가능 여부 확인
    RA-004: TBM 회의 기록
    PTW: 작업허가서
    CL: 장비·설비별 점검표

Required form_data keys:
    site_name       str  현장명
    work_date       str  작업일
    confirmer       str  확인자 성명

Optional form_data keys:
    project_name            str  공사명
    work_day                str  요일
    work_start_time         str  작업 시작 예정시각
    work_zone               str  작업구역
    writer_name             str  작성자
    reviewer_name           str  검토자
    approver_name           str  승인자
    -- 섹션2: 작업 개요 --
    work_name               str  금일 작업명
    work_type               str  작업공종
    work_content            str  작업내용
    work_location           str  작업위치
    worker_count            str  투입 인원
    subcontractor           str  협력업체
    equipment_used          str  사용 장비
    materials_used          str  사용 자재
    high_risk_work          str  고위험 작업 여부
    special_notes           str  특이사항
    -- 섹션3: 작업계획서·허가서 확인 --
    workplan_required       str  작업계획서 필요 여부
    workplan_approved       str  작업계획서 승인 여부
    permit_required         str  작업허가서 필요 여부
    permit_issued           str  작업허가서 발급 여부
    hazardous_work_type     str  위험작업 종류
    permit_conditions_ok    str  허가 조건 확인
    related_form_number     str  관련 서식 번호
    stop_if_unmet           str  미충족 시 작업중지 여부
    -- 섹션4: 위험성평가·TBM 확인 --
    ra_conducted            str  위험성평가 실시 여부
    ra_shared               str  위험성평가 공유 여부
    residual_risk_checked   str  잔여 위험요인 확인
    tbm_conducted           str  TBM 실시 여부
    tbm_attendees           str  TBM 참석 인원
    tbm_absentee_action     str  불참자 조치
    new_worker_informed     str  신규 근로자 전달 여부
    foreign_worker_informed str  외국인 근로자 전달 여부
    -- 섹션5: 작업장 상태 확인 --
    passage_secured         str  통로 확보
    opening_protected       str  개구부 방호
    fall_prevented          str  추락방지 조치
    falling_object_prevented str 낙하물 방지 조치
    lighting_ok             str  조명 상태
    ventilation_ok          str  환기 상태
    housekeeping_ok         str  정리정돈
    temp_power_ok           str  가설전기 상태
    combustibles_removed    str  화기 주변 가연물 제거
    escape_route_secured    str  비상대피로 확보
    -- 섹션6: 장비·공구·보호구 확인 --
    equipment_inspected     str  장비 사전점검 완료 여부
    tools_ok                str  공구 상태
    elcb_checked            str  전동공구 누전차단 확인
    safety_device_ok        str  안전장치 정상 여부
    ppe_issued              str  보호구 지급 여부
    ppe_worn                str  보호구 착용 여부
    safety_harness_attached str  안전대 체결 여부
    defective_ppe_replaced  str  불량 보호구 교체 여부
    -- 섹션7: 기상·환경 조건 확인 --
    weather_checked         str  기상 조건 확인 여부
    weather_risk            str  강풍/강우/폭염/한파 위험 여부
    work_stop_criteria      str  작업중지 기준 해당 여부
    dl004_linked            str  DL-004 연계 여부
    gas_measurement_needed  str  유해가스/산소농도 측정 필요 여부
    hazardous_env_risk      str  소음/분진/유해물질 위험 여부
    -- 섹션8: 비상대응 확인 --
    emergency_contact_ok    str  비상연락망 확인
    first_aid_location      str  구급함 위치 확인
    extinguisher_location   str  소화기 위치 확인
    aed_location            str  AED 위치 확인
    fire_report_ok          str  119 신고 체계 확인
    evacuation_route_ok     str  대피경로 확인
    assembly_point_ok       str  집결지 확인
    first_aid_person        str  응급조치 담당자 확인
    -- 섹션9: 작업개시 판단 --
    work_start_approved     str  작업개시 가능 여부
    conditional_approval    str  조건부 승인 여부
    conditional_conditions  str  조건부 승인 조건
    work_stop_reason        str  작업중지 사유
    remedial_action         str  보완조치 내용
    remedial_person         str  보완 담당자
    remedial_eta            str  보완 완료 예정시각
    reconfirmer             str  재확인자
    final_approver          str  최종 승인자
    -- 섹션10: 확인 및 서명 --
    foreman_name            str  작업반장
    supervisor_name         str  관리감독자
    safety_manager_name     str  안전관리자
    site_manager_name       str  현장소장
    subcontractor_manager   str  협력업체 책임자
    confirm_date            str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "작업전안전확인서"
SHEET_HEADING = "작업 전 안전 확인서"
DOC_ID = "DL-005"

TOTAL_COLS = 10

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL = Font(name="맑은 고딕", size=9)

_FILL_LABEL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE = PatternFill()

_THIN = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LABEL = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}

# 섹션5 작업장 상태 항목 (label, key)
_WORKPLACE_ITEMS = [
    ("통로 확보",         "passage_secured"),
    ("개구부 방호",       "opening_protected"),
    ("추락방지 조치",     "fall_prevented"),
    ("낙하물 방지",       "falling_object_prevented"),
    ("조명 상태",         "lighting_ok"),
    ("환기 상태",         "ventilation_ok"),
    ("정리정돈",          "housekeeping_ok"),
    ("가설전기 상태",     "temp_power_ok"),
    ("화기 주변 가연물",  "combustibles_removed"),
    ("비상대피로 확보",   "escape_route_secured"),
]

# 섹션6 장비·공구·보호구 항목 (label, key)
_EQUIPMENT_ITEMS = [
    ("장비 사전점검",     "equipment_inspected"),
    ("공구 상태",         "tools_ok"),
    ("누전차단 확인",     "elcb_checked"),
    ("안전장치 정상",     "safety_device_ok"),
    ("보호구 지급",       "ppe_issued"),
    ("보호구 착용",       "ppe_worn"),
    ("안전대 체결",       "safety_harness_attached"),
    ("불량 보호구 교체",  "defective_ppe_replaced"),
]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value = "" if value is None else value
    cell.font = font or _FONT_DEFAULT
    cell.fill = fill or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vc1: int, vc2: int, height: int = 20) -> None:
    _cell(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vc1, vc2, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, f"▶ {title}",
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_LEFT, height=20)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:2"


# ---------------------------------------------------------------------------
# 섹션 작성 함수
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    notice = (
        "【 공식 제출 서식 아님 】  작업 전 안전확인 및 작업개시 판단 보조서식 | "
        f"문서번호: {DOC_ID}"
    )
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    basis = (
        "산업안전보건법상 안전조치 및 위험성평가 조치사항과 연계 | "
        "작업시작 전 점검사항은 산업안전보건기준 별표 3과 연계"
    )
    _cell(ws, row, 1, TOTAL_COLS, basis,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")
    _lv(ws, row, "현장명", _v(data, "site_name"), 1, 2, 5)
    _lv(ws, row, "공사명", _v(data, "project_name"), 6, 7, 10)
    row += 1
    _lv(ws, row, "작업일", _v(data, "work_date"), 1, 2, 3)
    _lv(ws, row, "요일", _v(data, "work_day"), 4, 5, 5)
    _lv(ws, row, "작업 시작 예정", _v(data, "work_start_time"), 6, 7, 8)
    _lv(ws, row, "작업구역", _v(data, "work_zone"), 9, 10, 10)
    row += 1
    _lv(ws, row, "확인자", _v(data, "confirmer"), 1, 2, 3)
    _lv(ws, row, "작성자", _v(data, "writer_name"), 4, 5, 5)
    _lv(ws, row, "검토자", _v(data, "reviewer_name"), 6, 7, 8)
    _lv(ws, row, "승인자", _v(data, "approver_name"), 9, 10, 10)
    row += 1
    return row


def _write_work_overview(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 작업 개요")
    _lv(ws, row, "금일 작업명", _v(data, "work_name"), 1, 2, 5)
    _lv(ws, row, "작업공종", _v(data, "work_type"), 6, 7, 10)
    row += 1
    _lv(ws, row, "작업내용", _v(data, "work_content"), 1, 2, 10, height=30)
    row += 1
    _lv(ws, row, "작업위치", _v(data, "work_location"), 1, 2, 4)
    _lv(ws, row, "투입 인원", _v(data, "worker_count"), 5, 6, 6)
    _lv(ws, row, "협력업체", _v(data, "subcontractor"), 7, 8, 10)
    row += 1
    _lv(ws, row, "사용 장비", _v(data, "equipment_used"), 1, 2, 5)
    _lv(ws, row, "사용 자재", _v(data, "materials_used"), 6, 7, 10)
    row += 1
    _lv(ws, row, "고위험 작업", _v(data, "high_risk_work"), 1, 2, 4)
    _lv(ws, row, "특이사항", _v(data, "special_notes"), 5, 6, 10)
    row += 1
    return row


def _write_permit_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 작업계획서·허가서 확인")

    notice = "작업계획서·작업허가서·TBM·장비점검·보호구 착용 확인 후 작업개시"
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "작업계획서 필요", _v(data, "workplan_required"), 1, 2, 3)
    _lv(ws, row, "작업계획서 승인", _v(data, "workplan_approved"), 4, 5, 5)
    _lv(ws, row, "작업허가서 필요", _v(data, "permit_required"), 6, 7, 8)
    _lv(ws, row, "작업허가서 발급", _v(data, "permit_issued"), 9, 10, 10)
    row += 1
    _lv(ws, row, "위험작업 종류", _v(data, "hazardous_work_type"), 1, 2, 4)
    _lv(ws, row, "허가 조건 확인", _v(data, "permit_conditions_ok"), 5, 6, 7)
    _lv(ws, row, "관련 서식 번호", _v(data, "related_form_number"), 8, 9, 10)
    row += 1
    _lv(ws, row, "미충족 시 작업중지", _v(data, "stop_if_unmet"), 1, 2, 10)
    row += 1
    return row


def _write_ra_tbm(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 위험성평가·TBM 확인")
    _lv(ws, row, "위험성평가 실시", _v(data, "ra_conducted"), 1, 2, 3)
    _lv(ws, row, "위험성평가 공유", _v(data, "ra_shared"), 4, 5, 5)
    _lv(ws, row, "잔여 위험요인 확인", _v(data, "residual_risk_checked"), 6, 7, 10)
    row += 1
    _lv(ws, row, "TBM 실시", _v(data, "tbm_conducted"), 1, 2, 3)
    _lv(ws, row, "TBM 참석 인원", _v(data, "tbm_attendees"), 4, 5, 5)
    _lv(ws, row, "불참자 조치", _v(data, "tbm_absentee_action"), 6, 7, 10)
    row += 1
    _lv(ws, row, "신규 근로자 전달", _v(data, "new_worker_informed"), 1, 2, 4)
    _lv(ws, row, "외국인 근로자 전달", _v(data, "foreign_worker_informed"), 5, 6, 10)
    row += 1
    return row


def _write_workplace_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 작업장 상태 확인")

    # 헤더 (2열 구성)
    for c, hdr in enumerate(["확인 항목", "확인 결과", "확인 항목", "확인 결과",
                               "확인 항목", "확인 결과", "확인 항목", "확인 결과",
                               "확인 항목", "확인 결과"], 1):
        pass  # 단순 2열 반복 레이아웃으로 처리

    items = _WORKPLACE_ITEMS
    half = (len(items) + 1) // 2
    for i in range(half):
        left_label, left_key = items[i]
        left_val = _v(data, left_key)
        _cell(ws, row, 1, 2, left_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5, left_val, align=_ALIGN_CENTER)

        if i + half < len(items):
            right_label, right_key = items[i + half]
            right_val = _v(data, right_key)
            _cell(ws, row, 6, 7, right_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
            _cell(ws, row, 8, 10, right_val, align=_ALIGN_CENTER)
        else:
            _cell(ws, row, 6, 10, "", align=_ALIGN_CENTER)

        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_equipment_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 장비·공구·보호구 확인")

    items = _EQUIPMENT_ITEMS
    half = (len(items) + 1) // 2
    for i in range(half):
        left_label, left_key = items[i]
        left_val = _v(data, left_key)
        _cell(ws, row, 1, 2, left_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5, left_val, align=_ALIGN_CENTER)

        if i + half < len(items):
            right_label, right_key = items[i + half]
            right_val = _v(data, right_key)
            _cell(ws, row, 6, 7, right_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
            _cell(ws, row, 8, 10, right_val, align=_ALIGN_CENTER)
        else:
            _cell(ws, row, 6, 10, "", align=_ALIGN_CENTER)

        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_weather_env(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 기상·환경 조건 확인")

    notice = "기상 위험은 DL-004 기상 조건 기록 일지와 연계"
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "기상 조건 확인", _v(data, "weather_checked"), 1, 2, 3)
    _lv(ws, row, "기상 위험 여부", _v(data, "weather_risk"), 4, 5, 5)
    _lv(ws, row, "작업중지 기준 해당", _v(data, "work_stop_criteria"), 6, 7, 10)
    row += 1
    _lv(ws, row, "DL-004 연계", _v(data, "dl004_linked"), 1, 2, 3)
    _lv(ws, row, "유해가스 측정 필요", _v(data, "gas_measurement_needed"), 4, 5, 7)
    _lv(ws, row, "유해물질 위험", _v(data, "hazardous_env_risk"), 8, 9, 10)
    row += 1
    return row


def _write_emergency(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 비상대응 확인")

    emergency_items = [
        ("비상연락망 확인",     "emergency_contact_ok"),
        ("구급함 위치",         "first_aid_location"),
        ("소화기 위치",         "extinguisher_location"),
        ("AED 위치",            "aed_location"),
        ("119 신고 체계",       "fire_report_ok"),
        ("대피경로 확인",       "evacuation_route_ok"),
        ("집결지 확인",         "assembly_point_ok"),
        ("응급조치 담당자",     "first_aid_person"),
    ]
    half = (len(emergency_items) + 1) // 2
    for i in range(half):
        left_label, left_key = emergency_items[i]
        left_val = _v(data, left_key)
        _cell(ws, row, 1, 2, left_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5, left_val, align=_ALIGN_CENTER)

        if i + half < len(emergency_items):
            right_label, right_key = emergency_items[i + half]
            right_val = _v(data, right_key)
            _cell(ws, row, 6, 7, right_label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
            _cell(ws, row, 8, 10, right_val, align=_ALIGN_CENTER)
        else:
            _cell(ws, row, 6, 10, "", align=_ALIGN_CENTER)

        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_start_decision(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. 작업개시 판단")

    notice = "미충족 사항이 있으면 작업중지 또는 조건부 승인으로 관리"
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "작업개시 가능", _v(data, "work_start_approved"), 1, 2, 3)
    _lv(ws, row, "조건부 승인", _v(data, "conditional_approval"), 4, 5, 5)
    _lv(ws, row, "작업중지 사유", _v(data, "work_stop_reason"), 6, 7, 10)
    row += 1
    _lv(ws, row, "조건부 승인 조건", _v(data, "conditional_conditions"), 1, 2, 10, height=30)
    row += 1
    _lv(ws, row, "보완조치 내용", _v(data, "remedial_action"), 1, 2, 6)
    _lv(ws, row, "보완 담당자", _v(data, "remedial_person"), 7, 8, 10)
    row += 1
    _lv(ws, row, "보완 완료 예정", _v(data, "remedial_eta"), 1, 2, 4)
    _lv(ws, row, "재확인자", _v(data, "reconfirmer"), 5, 6, 7)
    _lv(ws, row, "최종 승인자", _v(data, "final_approver"), 8, 9, 10)
    row += 1
    return row


def _write_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 확인 및 서명")

    notice = "사진·점검표 등 증빙자료는 별도 보관"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1

    sign_items = [
        ("작업반장",        "foreman_name"),
        ("관리감독자",      "supervisor_name"),
        ("안전관리자",      "safety_manager_name"),
        ("현장소장",        "site_manager_name"),
        ("협력업체 책임자", "subcontractor_manager"),
    ]
    for lbl, key in sign_items:
        _cell(ws, row, 1, 2, lbl, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5, _v(data, key), align=_ALIGN_CENTER)
        _cell(ws, row, 6, 7, "서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10, "", align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 28
        row += 1

    _lv(ws, row, "확인일", _v(data, "confirm_date"), 1, 2, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_pre_work_safety_check(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_work_overview(ws, row, data)
    row = _write_permit_check(ws, row, data)
    row = _write_ra_tbm(ws, row, data)
    row = _write_workplace_check(ws, row, data)
    row = _write_equipment_check(ws, row, data)
    row = _write_weather_env(ws, row, data)
    row = _write_emergency(ws, row, data)
    row = _write_start_decision(ws, row, data)
    row = _write_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
