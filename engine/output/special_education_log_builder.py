"""
특별 안전보건교육 교육일지 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건법 제29조 제3항, 시행규칙 제26조, 별표 4, 별표 5
Form: ED-003

입력 스키마 (form_data dict):
    site_name                str|None   사업장명
    site_address             str|None   사업장 소재지
    education_name           str|None   교육명
    education_date           str|None   교육 일자
    education_location       str|None   교육 장소
    education_target_work    str|None   교육 대상 작업명 (별표 5 기준)
    target_work_free_input   str|None   별표 5 미수집 시 자유입력 작업명
    related_education        str|None   신규/작업변경/특별교육 연계 여부
    duration_category        str|None   교육시간 구분 (별표 4 기준)
    actual_duration_hours    str|None   실제 교육시간 (h)
    remaining_hours          str|None   잔여 교육시간 (h)
    subjects                 list[dict] 교육 내용 반복 행
        subject_name         str|None   교육 과목명
        subject_content      str|None   교육 내용 요약
        subject_hours        str|None   시간
    instructor_name          str|None   강사 성명
    instructor_org           str|None   강사 소속
    instructor_role          str|None   강사 직책
    instructor_qualification str|None   강사 자격/경력
    attendees                list[dict] 교육대상자 명단 (최대 MAX_ATTENDEES=30)
        attendee_name        str|None   성명
        attendee_org         str|None   소속/협력업체
        attendee_job_type    str|None   직종
        attendee_birth_year  str|None   생년
        attendee_completed   str|None   이수 여부
    comprehension_verbal     str|None   이해도 확인 — 구두
    comprehension_checklist  str|None   이해도 확인 — 체크리스트
    comprehension_practice   str|None   이해도 확인 — 실습
    retraining_targets       str|None   미이수/재교육 대상
    attachments              str|None   첨부자료 목록
    confirmer_name           str|None   교육담당자 성명
    confirmer_role           str|None   교육담당자 직위
    supervisor_name          str|None   관리감독자 성명
    site_manager_name        str|None   현장소장 또는 안전관리자 성명
    confirm_date             str|None   확인 일자

출력: xlsx bytes (in-memory)

Principles:
- 주민등록번호 전체 필드 없음.
- null/누락 → 빈 셀. 임의값 생성 금지.
- attendee_signature 는 항상 공란 (수기 서명 전용).
- 수강자 행은 MAX_ATTENDEES=30 고정 출력.
- 별표 5 대상 작업명: 별표 5 제1호~제39호 39개 작업 중 해당 작업명 자유입력. 목록은 별표 5 원문 참조.
- 교육시간 구분은 별표 4 기준 안내 문구를 표시하고 시간값 자동 판정은 하지 않음 (입력값 기반).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME    = "특별교육일지"
SHEET_HEADING = "특별 안전보건교육 교육일지"
MAX_ATTENDEES = 30
_DEFAULT_SUBJECT_ROWS = 3

# 별표 4/5 기준 안내 문구 (VERIFIED — 2025.5.30 개정 별표 기준)
_BYEOLPYO4_NOTE = (
    "별표 4 기준 — 특별교육: 일반근로자 16시간 이상(최초 작업 전 4시간, 나머지 3개월내 분할) / "
    "단기·간헐적 작업 2시간 이상 / 일용근로자(별표 5 제39호 제외) 2시간 이상 / 일용근로자(제39호) 8시간 이상"
)
_BYEOLPYO5_NOTE = (
    "별표 5 기준 특별교육 대상 작업 — 제1호(고압실 내 작업)~제39호(타워크레인 신호업무) 39개 작업 "
    "(시행규칙 별표 5 제1호라목, 2025.5.30 개정). 해당 작업명을 아래에 기재하십시오."
)

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_NOTE     = Font(name="맑은 고딕", size=8, italic=True, color="808080")
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NOTE    = PatternFill(fill_type="solid", fgColor="FFFACD")

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 컬럼 정의 (8컬럼 A-H)
# A=6, B=14, C=14, D=12, E=10, F=12, G=12, H=12
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 6,
    2: 14,
    3: 14,
    4: 12,
    5: 10,
    6: 12,
    7: 12,
    8: 12,
}
TOTAL_COLS = 8

_L1, _V1_START, _V1_END = 1, 2, 4
_L2, _V2_START, _V2_END = 5, 6, 8


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or PatternFill()
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_label_value_pair(ws, row: int,
                             label: str, value: Any,
                             label_col: int,
                             val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _write_section_header(ws, row: int, title: str, height: int = 18) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER,
                height=height)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더링 함수
# ---------------------------------------------------------------------------

_SHEET_SUBTITLE = "「산업안전보건법」 제29조 제3항에 따른 특별 안전보건교육"


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=_SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_basic_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    교육 기본정보 섹션.
    """
    r = _write_section_header(ws, start_row, "교육 기본정보")
    H = 20

    # 사업장명 / 사업장 소재지
    _write_label_value_pair(ws, r, "사업장명", _v(data, "site_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "사업장 소재지", _v(data, "site_address"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    # 교육명 / 교육구분(고정: 특별교육)
    _write_label_value_pair(ws, r, "교육명", _v(data, "education_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "교육구분", "특별교육",
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    # 교육 일자 / 교육 장소
    _write_label_value_pair(ws, r, "교육 일자", _v(data, "education_date"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "교육 장소", _v(data, "education_location"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    # 교육 대상 작업명 (전체 폭)
    target_work = _v(data, "education_target_work") or _v(data, "target_work_free_input")
    _write_label_value_pair(ws, r, "대상 작업명", target_work,
                            _L1, _V1_START, TOTAL_COLS)
    ws.row_dimensions[r].height = H
    r += 1

    # 연계 교육 여부 (전체 폭)
    _write_label_value_pair(ws, r, "연계 교육", _v(data, "related_education"),
                            _L1, _V1_START, TOTAL_COLS)
    ws.row_dimensions[r].height = H
    r += 1

    return r


def _write_duration_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    교육시간 구분 섹션.
    별표 4 evidence 미수집 — 시간 기준 자동 판정 없이 입력값만 표시.
    """
    r = _write_section_header(ws, start_row, "교육시간 구분")

    # 별표 4 안내 행
    _write_cell(ws, r, 1, TOTAL_COLS, _BYEOLPYO4_NOTE,
                font=_FONT_NOTE, fill=_FILL_NOTE, align=_ALIGN_CENTER, height=16)
    r += 1

    # 교육시간 구분 / 실제 교육시간
    _write_label_value_pair(ws, r, "시간 구분", _v(data, "duration_category"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "실제 교육시간", _v(data, "actual_duration_hours"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    # 잔여 교육시간 (전체 폭)
    _write_label_value_pair(ws, r, "잔여 시간", _v(data, "remaining_hours"),
                            _L1, _V1_START, TOTAL_COLS)
    ws.row_dimensions[r].height = 20
    r += 1

    return r


def _write_target_work_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    특별교육 대상 작업 섹션.
    별표 5 evidence 미수집 — 자유입력 + 안내 문구 표시.
    """
    r = _write_section_header(ws, start_row, "특별교육 대상 작업")

    # 별표 5 안내 행
    _write_cell(ws, r, 1, TOTAL_COLS, _BYEOLPYO5_NOTE,
                font=_FONT_NOTE, fill=_FILL_NOTE, align=_ALIGN_CENTER, height=16)
    r += 1

    # 대상 작업명 입력 (전체 폭)
    target = _v(data, "education_target_work") or _v(data, "target_work_free_input")
    _write_cell(ws, r, 1, 2, "대상 작업명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, TOTAL_COLS, target,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 24
    r += 1

    return r


def _write_subject_table(ws, start_row: int, subjects: List[Dict[str, Any]]) -> int:
    """
    교육내용 섹션.
    헤더: 순번(A) | 교육 과목명(B:D) | 교육 내용 요약(E:G) | 시간(h)(H)
    """
    r = _write_section_header(ws, start_row, "교육내용")

    # 표 헤더
    _write_cell(ws, r, 1, 1, "순번",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4, "교육 과목명",   font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7, "교육 내용 요약", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "시간(h)",       font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    row_count = max(len(subjects), _DEFAULT_SUBJECT_ROWS)
    for i in range(row_count):
        subj = subjects[i] if i < len(subjects) else {}
        _write_cell(ws, r, 1, 1, i + 1,                       font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(subj, "subject_name"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 7, _v(subj, "subject_content"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 8, 8, _v(subj, "subject_hours"),   font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_instructor_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    교육강사 정보 섹션.
    """
    r = _write_section_header(ws, start_row, "교육강사 정보")
    H = 20

    _write_label_value_pair(ws, r, "강사 성명", _v(data, "instructor_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "강사 소속", _v(data, "instructor_org"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    _write_label_value_pair(ws, r, "강사 직책", _v(data, "instructor_role"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "자격/경력", _v(data, "instructor_qualification"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    return r


def _write_attendee_table(ws, start_row: int, attendees: List[Dict[str, Any]]) -> int:
    """
    교육대상자 명단 섹션.
    헤더: 번호(A) | 성명(B) | 소속/협력업체(C:D) | 직종(E) | 생년(F) | 이수여부(G) | 서명(H)
    서명란: 항상 공란 (수기 서명 전용).
    주민등록번호 전체 필드 없음.
    """
    r = _write_section_header(ws, start_row, "교육대상자 명단")

    _write_cell(ws, r, 1, 1, "번호",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 2, "성명",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 3, 4, "소속/협력업체", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 5, "직종",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 6, "생년",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "이수",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "서명",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_ATTENDEES):
        att = attendees[i] if i < len(attendees) else {}
        _write_cell(ws, r, 1, 1, i + 1,                          font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 2, _v(att, "attendee_name"),       font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 3, 4, _v(att, "attendee_org"),        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 5, _v(att, "attendee_job_type"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 6, _v(att, "attendee_birth_year"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7, 7, _v(att, "attendee_completed"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, "",                              font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_comprehension_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    이해도 확인 섹션.
    """
    r = _write_section_header(ws, start_row, "이해도 확인")
    H = 20

    _write_label_value_pair(ws, r, "구두 확인", _v(data, "comprehension_verbal"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "체크리스트", _v(data, "comprehension_checklist"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    _write_label_value_pair(ws, r, "실습 확인", _v(data, "comprehension_practice"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "재교육 대상", _v(data, "retraining_targets"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H
    r += 1

    return r


def _write_attachments_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    첨부자료 섹션.
    """
    r = _write_section_header(ws, start_row, "첨부자료")

    attach_val = _v(data, "attachments") or "교육자료 / 사진 / 참석자 서명부 / 시험·평가표"
    _write_cell(ws, r, 1, 2, "첨부 목록",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, TOTAL_COLS, attach_val,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 24
    r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    확인 서명 섹션 (교육담당자, 관리감독자, 현장소장/안전관리자 3인).
    """
    r = _write_section_header(ws, start_row, "확인 서명")

    # 헤더 행
    _write_cell(ws, r, 1, 1, "구분",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 3, "성명",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 4, 5, "직위",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 6, "확인 일자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 8, "서명",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    signers = [
        ("교육담당자", _v(data, "confirmer_name"),   _v(data, "confirmer_role"),   _v(data, "confirm_date")),
        ("관리감독자", _v(data, "supervisor_name"),   "",                           ""),
        ("현장소장/\n안전관리자", _v(data, "site_manager_name"), "",                ""),
    ]
    for role, name, title, date in signers:
        _write_cell(ws, r, 1, 1, role,  font=_FONT_BOLD,    fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 3, name,  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4, 5, title, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 6, date,  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7, 8, "",    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 30
        r += 1

    return r


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_special_education_log_sheet(ws, form_data: Dict[str, Any]) -> None:
    """
    주어진 worksheet에 특별 안전보건교육 교육일지를 렌더링.
    """
    data      = form_data or {}
    subjects  = data.get("subjects")  or []
    attendees = data.get("attendees") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_duration_section(ws, row, data)
    row = _write_target_work_section(ws, row, data)
    row = _write_subject_table(ws, row, subjects)
    row = _write_instructor_section(ws, row, data)
    row = _write_attendee_table(ws, row, attendees)
    row = _write_comprehension_section(ws, row, data)
    row = _write_attachments_section(ws, row, data)
    _write_confirmation(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_special_education_log_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 특별 안전보건교육 교육일지 xlsx 바이너리를 반환.

    Args:
        form_data: 입력 스키마 준수 dict. (상단 docstring 참조)

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_special_education_log_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
