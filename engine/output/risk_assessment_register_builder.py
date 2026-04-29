"""
위험성평가 관리 등록부 — Excel 출력 모듈 (v1.0).

법적 근거: 산업안전보건법 시행규칙 제37조 (위험성평가 결과 기록·보존 3년)
분류: RA-002 — 법정 서식 없음, 기록·보존 의무 서식

Required form_data keys: (모두 optional 처리 — 빈 칸 출력)

Optional form_data keys:
    site_name           str  사업장명
    project_name        str  현장명
    representative      str  사업주 / 현장 대표
    safety_manager      str  안전관리자

    entries  list[dict]  평가 등록 행 목록 (max MAX_ENTRIES)
        seq             int/str  순번
        work_type       str      작업/공종명
        assessment_date str      평가 실시일
        hazard_summary  str      주요 위험요인 요약
        risk_level      str      위험성 수준 (상/중/하)
        measure_status  str      개선대책 조치 상태 (완료/진행중/미착수)
        reviewer        str      검토자
        approver        str      승인자
        next_review     str      재검토 예정일
        remarks         str      비고

    prepared_by     str  작성자
    reviewed_by     str  검토자
    approved_by     str  승인자
    prepared_date   str  작성일
    reviewed_date   str  검토일
    approved_date   str  승인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "위험성평가관리등록부"
SHEET_HEADING  = "위험성평가 관리 등록부"
SHEET_SUBTITLE = "산업안전보건법 시행규칙 제37조 — 위험성평가 결과 기록·보존 (3년)"

MAX_ENTRIES = 30
TOTAL_COLS  = 10

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 6,   # 순번
    2: 16,  # 작업/공종명
    3: 13,  # 평가 실시일
    4: 24,  # 주요 위험요인 요약
    5: 8,   # 위험성 수준
    6: 12,  # 개선대책 조치 상태
    7: 10,  # 검토자
    8: 10,  # 승인자
    9: 13,  # 재검토 예정일
    10: 14, # 비고
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col1: int, label_col2: int,
              val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col1, label_col2, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_header_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"),    1, 1, 2, 4)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), 5, 5, 6, 8)
    _write_lv(ws, r, "사업주",  _v(data, "representative"), 9, 9, 10, 10)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "안전관리자", _v(data, "safety_manager"), 1, 2, 3, 10)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_entries_table(ws, start_row: int,
                          entries: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "평가 대상 작업/공종 목록",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    headers = ["순번", "작업/공종명", "평가 실시일", "주요 위험요인 요약",
               "위험성\n수준", "조치 상태", "검토자", "승인자", "재검토 예정일", "비고"]
    for col, h in enumerate(headers, start=1):
        _write_cell(ws, r, col, col, h,
                    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=32)
    r += 1

    for i in range(MAX_ENTRIES):
        entry = entries[i] if i < len(entries) else {}
        row_h = 22
        _write_cell(ws, r, 1,  1,  _v(entry, "seq") or (i + 1),        font=_FONT_SMALL, align=_ALIGN_CENTER, height=row_h)
        _write_cell(ws, r, 2,  2,  _v(entry, "work_type"),              font=_FONT_SMALL, align=_ALIGN_LEFT)
        _write_cell(ws, r, 3,  3,  _v(entry, "assessment_date"),        font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 4,  4,  _v(entry, "hazard_summary"),         font=_FONT_SMALL, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5,  5,  _v(entry, "risk_level"),             font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 6,  6,  _v(entry, "measure_status"),         font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7,  7,  _v(entry, "reviewer"),               font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8,  8,  _v(entry, "approver"),               font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 9,  9,  _v(entry, "next_review"),            font=_FONT_SMALL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 10, 10, _v(entry, "remarks"),                font=_FONT_SMALL, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_signature_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작성 / 검토 / 승인",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    # 역할 행
    _write_cell(ws, r, 1, 1,  "구분",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  "작성자",     font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  "검토자",     font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, "승인자",     font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20; r += 1

    # 성명 행
    _write_cell(ws, r, 1, 1,  "성명",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  _v(data, "prepared_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  _v(data, "reviewed_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, _v(data, "approved_by"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 28; r += 1

    # 서명 행
    _write_cell(ws, r, 1, 1,  "서명",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  "",           font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  "",           font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, "",           font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36; r += 1

    # 일자 행
    _write_cell(ws, r, 1, 1,  "일자",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4,  _v(data, "prepared_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7,  _v(data, "reviewed_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 10, _v(data, "approved_date"),  font=_FONT_SMALL, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 20; r += 1

    return r


def build_risk_assessment_register_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 위험성평가 관리 등록부 xlsx 바이너리를 반환."""
    data    = form_data or {}
    entries = data.get("entries") or []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_header_block(ws, row, data)
    row = _write_entries_table(ws, row, entries)
    _write_signature_block(ws, row, data)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:6"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
