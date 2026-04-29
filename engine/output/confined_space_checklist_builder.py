"""
밀폐공간 사전 안전점검표 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제622조~제625조
           (밀폐공간 작업 전 측정·환기·감시인 배치 의무)
분류: GEN_INTERNAL — 조문 기반 설계

기본 점검항목 10개는 제622조~제625조 의무사항에서 도출.
check_items 미제공 시 법정 기반 기본 10개 자동 적용.

Required form_data keys:
    check_date    str  점검 일자   [실무 권장]
    work_location str  작업 장소   [법정] 제619조
    checker_name  str  점검자 성명 [실무 권장]

Optional form_data keys:
    site_name, project_name
    work_content str  작업 내용

    check_items  list[dict]  점검 항목 (max 10, 미제공 시 기본 10개 적용)
        item    str  점검 항목명 (기본값 있음)
        result  str  결과 (○ / × / 해당없음)
        note    str  비고
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "밀폐공간사전점검표"
SHEET_HEADING  = "밀폐공간 작업 사전 안전점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제622조~제625조에 따른 사전 확인 점검"

MAX_CHECKS  = 10
TOTAL_COLS  = 8

# 기준규칙 제622조~제625조 의무사항 기반 기본 점검항목
_DEFAULT_CHECKS: List[Dict[str, str]] = [
    {"item": "출입 전 산소농도 측정 여부 (18% 이상 확인)",    "result": "", "note": ""},
    {"item": "유해가스(H₂S·CO 등) 측정 여부",               "result": "", "note": ""},
    {"item": "환기 실시 여부 (강제 환기 또는 자연 환기)",       "result": "", "note": ""},
    {"item": "감시인 배치 여부 (외부 상시 대기)",              "result": "", "note": ""},
    {"item": "구조장비 비치 여부 (구조용 로프·안전대·공기호흡기)", "result": "", "note": ""},
    {"item": "통신수단 확보 여부 (무선기·신호줄 등)",           "result": "", "note": ""},
    {"item": "보호구 착용 여부 (방독면 또는 공기호흡기)",        "result": "", "note": ""},
    {"item": "출입자 명단 확인 및 기록 여부",                  "result": "", "note": ""},
    {"item": "비상연락체계 확인 (119 신고·응급 연락망)",        "result": "", "note": ""},
    {"item": "작업 종료 후 인원 확인 완료",                    "result": "", "note": ""},
]

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 6, 2: 34, 3: 34, 4: 10, 5: 10, 6: 10, 7: 10, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_header_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "점검 일자", _v(data, "check_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "점검자",  _v(data, "checker_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 장소", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 내용", _v(data, "work_content"),  _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1
    return r


def _write_checklist_table(ws, start_row: int,
                            check_items: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "사전 안전점검 항목  (기준규칙 제622조~제625조 의무사항 기반)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    # 안내
    _write_cell(ws, r, 1, TOTAL_COLS,
                "※ 결과 기재 방법: 이상없음 → ○ / 이상있음 → × / 해당없음 → - (이상 발견 시 즉시 작업 중지)",
                font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_CENTER, height=16)
    r += 1

    # 표 헤더 (순번·점검항목·결과·비고)
    _write_cell(ws, r, 1, 1, "순번",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 5, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 6, "결과",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 8, "비고",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1

    for i in range(MAX_CHECKS):
        default = _DEFAULT_CHECKS[i] if i < len(_DEFAULT_CHECKS) else {}
        provided = check_items[i] if i < len(check_items) else {}
        item_name = provided.get("item") or default.get("item", "")
        result    = _v(provided, "result")
        note      = _v(provided, "note")

        _write_cell(ws, r, 1, 1, i + 1,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, item_name, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 6, result,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7, 8, note,      font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 28
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 2, "점검자 서명",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 8, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36
    return r + 1


def build_confined_space_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 밀폐공간 사전 안전점검표 xlsx 바이너리를 반환."""
    data        = form_data or {}
    check_items = data.get("check_items") or []
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_header_block(ws, row, data)
    row = _write_checklist_table(ws, row, check_items)
    _write_confirmation(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:9"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
