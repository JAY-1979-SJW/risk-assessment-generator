"""
건설 장비 보험·정기검사증 확인서 — Excel 출력 모듈 (v1).

법적 근거 / 실무 근거:
    건설기계관리법 제13조 (정기검사)
    자동차손해배상보장법 (의무보험 가입)
    산업안전보건법 제80조 (유해·위험기계 방호조치)
    산업안전보건법 제93조 (안전검사 의무)

분류: PRACTICAL — 공식 별지 제출서식 아님
      장비 보험·검사증 등 증빙서류 확인 보조서식
      보험증권·정기검사증·등록증·운전자 자격은 반입 및 사용 전 확인 필요

역할 분리:
    PPE-003(본 서식): 보험·정기검사증·등록증·자격증 등 증빙서류 상세 확인
    PPE-002: 장비 반입 신청·서류 확인·반입 승인
    CL-003: 반입 후 장비 일일 사전 점검
    EQ: 장비 사용계획서/작업계획서

Required form_data keys:
    site_name       str  현장명
    check_date      str  확인일
    equipment_name  str  장비명

Optional form_data keys:
    -- 섹션1: 문서 기본정보 --
    project_name            str  공사명
    entry_request_no        str  장비 반입 신청번호
    check_period            str  확인 대상 기간
    checker                 str  확인자
    writer                  str  작성자
    reviewer                str  검토자
    approver                str  승인자
    -- 섹션2: 장비 기본정보 --
    equipment_type          str  장비 종류
    manufacturer            str  제조사
    equipment_model         str  모델명
    serial_no               str  제조번호
    reg_no                  str  차량번호/등록번호
    equipment_capacity      str  규격/용량
    owner_name              str  소유자
    rental_company          str  임대업체
    planned_work            str  사용 예정 작업
    -- 섹션3: 장비 등록증 확인 --
    reg_cert_submitted      str  등록증 제출 여부
    reg_cert_no             str  등록번호
    owner_match             str  소유자 일치 여부
    equip_name_match        str  장비명 일치 여부
    serial_match            str  차대번호/제조번호 일치 여부
    reg_cert_valid          str  유효성 확인
    reg_cert_supplement     str  보완 필요사항
    -- 섹션4: 보험증권 확인 --
    insurance_type          str  보험 종류
    insurance_company       str  보험사
    policy_no               str  증권번호
    policy_holder           str  보험 가입자
    insured_equipment       str  피보험 장비
    insurance_start         str  보장 기간 시작일
    insurance_end           str  보장 기간 종료일
    liability_covered       str  대인/대물 보장 여부
    coverage_includes_period str 현장 사용 기간 포함 여부
    insurance_expiry_near   str  만료 임박 여부
    insurance_supplement    str  보완 필요사항
    -- 섹션5: 정기검사증 확인 --
    periodic_insp_required  str  정기검사 대상 여부
    insp_agency             str  검사기관
    insp_date               str  검사일
    insp_expiry             str  유효기간
    insp_result             str  검사 결과
    insp_nonconformance     str  부적합 사항
    reinspection_needed     str  재검사 필요 여부
    insp_expiry_near        str  만료 임박 여부
    insp_supplement         str  보완 필요사항
    -- 섹션6: 안전검사·성능검사 확인 --
    safety_insp_required    str  안전검사 대상 여부
    safety_cert_submitted   str  검사증 제출 여부
    safety_insp_agency      str  검사기관
    safety_insp_date        str  검사일
    safety_insp_expiry      str  유효기간
    safety_insp_passed      str  합격 여부
    safety_insp_conditions  str  조건부 합격 조건
    safety_insp_supplement  str  보완 필요사항
    -- 섹션7: 운전자 자격 및 교육 확인 --
    operator_name           str  운전자 성명
    operator_affiliation    str  소속
    license_type            str  자격/면허 종류
    license_no              str  자격/면허 번호
    license_valid           str  자격 유효 여부
    operable_scope          str  장비 운전 가능 범위
    safety_edu_done         str  특별교육/안전교육 이수 여부
    safety_edu_date         str  교육일
    operator_supplement     str  보완 필요사항
    -- 섹션8: 서류 원본/사본 및 보관 확인 --
    original_verified       str  원본 확인 여부
    copy_stored             str  사본 보관 여부
    electronic_stored       str  전자파일 보관 여부
    expiry_ledger_updated   str  유효기간 관리대장 반영 여부
    expiry_alert_needed     str  만료 알림 필요 여부
    pii_masked              str  개인정보 마스킹 여부
    storage_location        str  보관 위치
    -- 섹션9: 종합 판정 및 조치사항 --
    entry_allowed           str  반입 가능 여부
    use_allowed             str  사용 가능 여부
    conditional_approval    str  조건부 승인 여부
    approval_conditions     str  조건부 승인 조건
    missing_documents       str  미비 서류
    action_person           str  보완 담당자
    completion_due          str  완료 예정일
    recheck_date            str  재확인일
    final_confirmer         str  최종 확인자
    -- 섹션10: 연계 관리 및 승인 --
    ppe002_linked           str  PPE-002 반입 신청서 연계 여부
    cl003_linked            str  CL-003 일일 사전 점검 연계 여부
    eq_plan_linked          str  EQ 장비 사용계획서 연계 여부
    dl001_linked            str  DL-001 안전관리 일지 반영 여부
    applicant_company       str  신청업체 확인
    safety_manager_name     str  안전관리자
    supervisor_name         str  관리감독자
    site_manager_name       str  현장소장
    confirm_date            str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "보험검사증확인서"
SHEET_HEADING = "건설 장비 보험·정기검사증 확인서"
DOC_ID        = "PPE-003"
TOTAL_COLS    = 10

_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL   = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[float] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=value)
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    if height:
        ws.row_dimensions[row].height = height
    for c in range(col1, col2 + 1):
        ws.cell(row=row, column=c).border = _BORDER


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vs: int, ve: int, height: float = 20) -> None:
    _cell(ws, row, lc, lc,  label, font=_FONT_BOLD,   fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vs, ve,  value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, title,
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=22)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7


# ---------------------------------------------------------------------------
# 섹션 구현
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    _cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
          font=_FONT_TITLE, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=36)
    row += 1

    subtitle = (
        "공식 제출 서식 아님 — 장비 보험·검사증 등 증빙서류 확인 보조서식  |  "
        f"보험증권·정기검사증·등록증·운전자 자격은 반입 및 사용 전 확인 필요  ({DOC_ID})"
    )
    _cell(ws, row, 1, TOTAL_COLS, subtitle,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    notice = (
        "보험 및 검사 유효기간이 현장 사용 기간을 포함하는지 확인 필요  |  "
        "미비 서류 또는 유효기간 만료 시 반입·사용 보류  |  "
        "PPE-002 반입 신청서 및 CL-003 장비 일일 사전 점검표와 별도 관리  |  "
        "증빙자료 사본 또는 전자파일은 별도 보관"
    )
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")

    _lv(ws, row, "공사명",          _v(data, "project_name"),      1, 2, 5)
    _lv(ws, row, "현장명",          _v(data, "site_name"),         6, 7, 10)
    row += 1
    _lv(ws, row, "확인일",          _v(data, "check_date"),        1, 2, 4)
    _lv(ws, row, "반입 신청번호",   _v(data, "entry_request_no"),  5, 6, 7)
    _lv(ws, row, "확인 대상 기간",  _v(data, "check_period"),      8, 9, 10)
    row += 1
    _lv(ws, row, "확인자",          _v(data, "checker"),           1, 2, 4)
    _lv(ws, row, "작성자",          _v(data, "writer"),            5, 6, 7)
    _lv(ws, row, "검토자",          _v(data, "reviewer"),          8, 9, 10)
    row += 1
    _lv(ws, row, "승인자",          _v(data, "approver"),          1, 2, 10)
    row += 1

    notice = "개인정보·자격번호 등은 목적 범위 내 최소 기재"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_equipment_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 장비 기본정보")

    _lv(ws, row, "장비명",      _v(data, "equipment_name"),     1, 2, 5)
    _lv(ws, row, "장비 종류",   _v(data, "equipment_type"),     6, 7, 10)
    row += 1
    _lv(ws, row, "제조사",      _v(data, "manufacturer"),       1, 2, 4)
    _lv(ws, row, "모델명",      _v(data, "equipment_model"),    5, 6, 7)
    _lv(ws, row, "제조번호",    _v(data, "serial_no"),          8, 9, 10)
    row += 1
    _lv(ws, row, "차량번호/등록번호", _v(data, "reg_no"),       1, 2, 5)
    _lv(ws, row, "규격/용량",   _v(data, "equipment_capacity"), 6, 7, 10)
    row += 1
    _lv(ws, row, "소유자",      _v(data, "owner_name"),         1, 2, 5)
    _lv(ws, row, "임대업체",    _v(data, "rental_company"),     6, 7, 10)
    row += 1
    _lv(ws, row, "사용 예정 작업", _v(data, "planned_work"),    1, 2, 10, height=24)
    row += 1
    return row


def _write_reg_cert(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 장비 등록증 확인")

    items = [
        ("등록증 제출 여부",          "reg_cert_submitted", 1, 3, 5),
        ("등록번호",                  "reg_cert_no",        6, 8, 10),
        ("소유자 일치 여부",          "owner_match",        1, 3, 5),
        ("장비명 일치 여부",          "equip_name_match",   6, 8, 10),
        ("차대번호/제조번호 일치 여부", "serial_match",     1, 3, 5),
        ("유효성 확인",               "reg_cert_valid",     6, 8, 10),
    ]
    for label, key, lc, vs, ve in items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    _lv(ws, row, "보완 필요사항", _v(data, "reg_cert_supplement"), 1, 3, 10, height=24)
    row += 1
    return row


def _write_insurance(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 보험증권 확인")

    _lv(ws, row, "보험 종류",      _v(data, "insurance_type"),    1, 3, 5)
    _lv(ws, row, "보험사",         _v(data, "insurance_company"), 6, 8, 10)
    row += 1
    _lv(ws, row, "증권번호",       _v(data, "policy_no"),         1, 3, 5)
    _lv(ws, row, "보험 가입자",    _v(data, "policy_holder"),     6, 8, 10)
    row += 1
    _lv(ws, row, "피보험 장비",    _v(data, "insured_equipment"), 1, 3, 10)
    row += 1
    _lv(ws, row, "보장 시작일",    _v(data, "insurance_start"),   1, 3, 5)
    _lv(ws, row, "보장 종료일",    _v(data, "insurance_end"),     6, 8, 10)
    row += 1
    _lv(ws, row, "대인/대물 보장", _v(data, "liability_covered"), 1, 3, 5)
    _lv(ws, row, "현장 사용기간 포함", _v(data, "coverage_includes_period"), 6, 8, 10)
    row += 1
    _lv(ws, row, "만료 임박 여부", _v(data, "insurance_expiry_near"), 1, 3, 5)
    row += 1

    warn = "보험 및 검사 유효기간이 현장 사용 기간을 포함하는지 확인 필요"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {warn}",
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=16)
    row += 1
    _lv(ws, row, "보완 필요사항", _v(data, "insurance_supplement"), 1, 3, 10, height=24)
    row += 1
    return row


def _write_periodic_insp(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 정기검사증 확인")

    items = [
        ("정기검사 대상 여부", "periodic_insp_required", 1, 3, 5),
        ("검사기관",           "insp_agency",            6, 8, 10),
        ("검사일",             "insp_date",              1, 3, 5),
        ("유효기간",           "insp_expiry",            6, 8, 10),
        ("검사 결과",          "insp_result",            1, 3, 5),
        ("부적합 사항",        "insp_nonconformance",    6, 8, 10),
        ("재검사 필요 여부",   "reinspection_needed",    1, 3, 5),
        ("만료 임박 여부",     "insp_expiry_near",       6, 8, 10),
    ]
    for label, key, lc, vs, ve in items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    _lv(ws, row, "보완 필요사항", _v(data, "insp_supplement"), 1, 3, 10, height=24)
    row += 1
    return row


def _write_safety_insp(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 안전검사·성능검사 확인")

    items = [
        ("안전검사 대상 여부",    "safety_insp_required",  1, 3, 5),
        ("검사증 제출 여부",      "safety_cert_submitted", 6, 8, 10),
        ("검사기관",              "safety_insp_agency",    1, 3, 5),
        ("검사일",                "safety_insp_date",      6, 8, 10),
        ("유효기간",              "safety_insp_expiry",    1, 3, 5),
        ("합격 여부",             "safety_insp_passed",    6, 8, 10),
    ]
    for label, key, lc, vs, ve in items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    _lv(ws, row, "조건부 합격 조건", _v(data, "safety_insp_conditions"), 1, 3, 10, height=24)
    row += 1
    _lv(ws, row, "보완 필요사항",    _v(data, "safety_insp_supplement"), 1, 3, 10, height=24)
    row += 1
    return row


def _write_operator(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 운전자 자격 및 교육 확인")

    _lv(ws, row, "운전자 성명",        _v(data, "operator_name"),        1, 3, 5)
    _lv(ws, row, "소속",               _v(data, "operator_affiliation"), 6, 8, 10)
    row += 1
    _lv(ws, row, "자격/면허 종류",     _v(data, "license_type"),         1, 3, 5)
    _lv(ws, row, "자격/면허 번호",     _v(data, "license_no"),           6, 8, 10)
    row += 1
    _lv(ws, row, "자격 유효 여부",     _v(data, "license_valid"),        1, 3, 5)
    _lv(ws, row, "장비 운전 가능 범위", _v(data, "operable_scope"),      6, 8, 10)
    row += 1
    _lv(ws, row, "안전교육 이수 여부", _v(data, "safety_edu_done"),      1, 3, 5)
    _lv(ws, row, "교육일",             _v(data, "safety_edu_date"),      6, 8, 10)
    row += 1
    _lv(ws, row, "보완 필요사항",      _v(data, "operator_supplement"),  1, 3, 10, height=24)
    row += 1
    return row


def _write_document_storage(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 서류 원본/사본 및 보관 확인")

    items = [
        ("원본 확인 여부",           "original_verified",    1, 3, 5),
        ("사본 보관 여부",           "copy_stored",          6, 8, 10),
        ("전자파일 보관 여부",       "electronic_stored",    1, 3, 5),
        ("유효기간 관리대장 반영",   "expiry_ledger_updated", 6, 8, 10),
        ("만료 알림 필요 여부",      "expiry_alert_needed",  1, 3, 5),
        ("개인정보 마스킹 여부",     "pii_masked",           6, 8, 10),
    ]
    for label, key, lc, vs, ve in items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    _lv(ws, row, "보관 위치", _v(data, "storage_location"), 1, 3, 10)
    row += 1

    notice = "증빙자료 사본 또는 전자파일은 별도 보관"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_judgment(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. 종합 판정 및 조치사항")

    _lv(ws, row, "반입 가능 여부",    _v(data, "entry_allowed"),       1, 3, 5)
    _lv(ws, row, "사용 가능 여부",    _v(data, "use_allowed"),         6, 8, 10)
    row += 1
    _lv(ws, row, "조건부 승인 여부",  _v(data, "conditional_approval"), 1, 3, 5)
    _lv(ws, row, "조건부 승인 조건",  _v(data, "approval_conditions"), 6, 8, 10)
    row += 1
    _lv(ws, row, "미비 서류",         _v(data, "missing_documents"),   1, 3, 10, height=24)
    row += 1
    _lv(ws, row, "보완 담당자",       _v(data, "action_person"),       1, 3, 4)
    _lv(ws, row, "완료 예정일",       _v(data, "completion_due"),      5, 6, 7)
    _lv(ws, row, "재확인일",          _v(data, "recheck_date"),        8, 9, 10)
    row += 1
    _lv(ws, row, "최종 확인자",       _v(data, "final_confirmer"),     1, 3, 10)
    row += 1

    warn = "미비 서류 또는 유효기간 만료 시 반입·사용 보류"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {warn}",
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_linkage_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 연계 관리 및 승인")

    link_items = [
        ("PPE-002 반입 신청서 연계",   "ppe002_linked",     1, 3, 5),
        ("CL-003 일일 점검 연계",      "cl003_linked",      6, 8, 10),
        ("EQ 장비 사용계획서 연계",    "eq_plan_linked",    1, 3, 5),
        ("DL-001 안전관리 일지 반영",  "dl001_linked",      6, 8, 10),
    ]
    for label, key, lc, vs, ve in link_items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1

    _lv(ws, row, "신청업체 확인", _v(data, "applicant_company"), 1, 3, 10)
    row += 1

    sign_items = [
        ("안전관리자",  "safety_manager_name"),
        ("관리감독자",  "supervisor_name"),
        ("현장소장",    "site_manager_name"),
    ]
    for lbl, key in sign_items:
        _cell(ws, row, 1,  2,  lbl,             font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3,  5,  _v(data, key),  align=_ALIGN_CENTER)
        _cell(ws, row, 6,  7,  "서명",          font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10,  "",             align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 28
        row += 1

    _lv(ws, row, "확인일", _v(data, "confirm_date"), 1, 2, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_equipment_insurance_inspection_check(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_equipment_info(ws, row, data)
    row = _write_reg_cert(ws, row, data)
    row = _write_insurance(ws, row, data)
    row = _write_periodic_insp(ws, row, data)
    row = _write_safety_insp(ws, row, data)
    row = _write_operator(ws, row, data)
    row = _write_document_storage(ws, row, data)
    row = _write_judgment(ws, row, data)
    row = _write_linkage_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
