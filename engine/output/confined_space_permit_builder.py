"""
밀폐공간 작업허가서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제619조~제626조
           제622조 (사업주: 작업 전 산소·유해가스 측정 및 환기 의무)
           제625조 (사업주: 감시인 배치 의무)
분류: GEN_INTERNAL — 조문 기반 설계

Required form_data keys:
    permit_no       str  허가 번호     [실무 권장]
    work_location   str  작업 장소     [법정] 제619조
    work_content    str  작업 내용     [법정] 제619조
    monitor_name    str  감시인 성명   [법정] 제625조
    oxygen_level    str  산소농도 측정값 [법정] 제622조
    permit_issuer   str  허가자 (서명자) [실무 권장]

Optional form_data keys:
    site_name, project_name
    permit_datetime   str  허가 일시
    validity_period   str  유효기간
    gas_h2s_level     str  황화수소(H₂S) 농도
    gas_co_level      str  일산화탄소(CO) 농도
    gas_other         str  기타 유해가스 농도
    ventilation_status str  환기 상태
    rescue_equipment_check str 구조장비 비치 여부
    work_end_time     str  작업 종료 시각
    completion_confirm str 작업종료 확인자

    workers  list[dict]  작업자 명단 (max 10)
        name  str  성명
        role  str  역할/직종
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "밀폐공간작업허가서"
SHEET_HEADING  = "밀폐공간 작업 허가서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제619조~제626조에 따른 밀폐공간 작업 허가"

MAX_WORKERS = 10
TOTAL_COLS  = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_permit_header(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "허가 번호", _v(data, "permit_no"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "허가 일시", _v(data, "permit_datetime"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "유효기간", _v(data, "validity_period"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_work_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, _L1, _L1, "작업 장소",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "work_location"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1
    _write_cell(ws, r, _L1, _L1, "작업 내용",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "work_content"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1
    return r


def _write_measurement(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "측정값 확인  (기준규칙 제622조 — 허가 전 측정 의무)",
                font=_FONT_BOLD, fill=_FILL_WARN, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "산소농도(O₂)", _v(data, "oxygen_level"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "황화수소(H₂S)", _v(data, "gas_h2s_level"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "일산화탄소(CO)", _v(data, "gas_co_level"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "기타 가스",      _v(data, "gas_other"),    _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "환기 상태", _v(data, "ventilation_status"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 25; r += 1

    _write_lv(ws, r, "구조장비\n비치 여부", _v(data, "rescue_equipment_check"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1

    _write_lv(ws, r, "감시인 성명", _v(data, "monitor_name"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 20; r += 1

    return r


def _write_workers_table(ws, start_row: int, workers: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 4, "성명", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "역할/직종", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1

    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, r, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(item, "name"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 8, _v(item, "role"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_permit_sign(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "허가 및 종료 확인",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "허가자 (서명)", _v(data, "permit_issuer"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 종료 시각", _v(data, "work_end_time"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_cell(ws, r, _L1, _L1, "종료 확인자 서명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "completion_confirm"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 36
    r += 1

    return r


def build_confined_space_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 밀폐공간 작업허가서 xlsx 바이너리를 반환."""
    data    = form_data or {}
    workers = data.get("workers") or []
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_permit_header(ws, row, data)
    row = _write_work_info(ws, row, data)
    row = _write_measurement(ws, row, data)
    row = _write_workers_table(ws, row, workers)
    _write_permit_sign(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
