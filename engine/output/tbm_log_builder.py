"""
TBM (Tool Box Meeting) 안전점검 일지 — Excel 출력 모듈 (v1).

법적 근거: 고용노동부고시 제2023-19호 (상시평가 실시 권고),
           중대재해처벌법 시행령 제4조 (안전보건관리체계 구축 의무 증빙)
분류: GEN_INTERNAL — 법정 서식 없음, 실무 권장 서식

Required form_data keys:
    tbm_date            str  실시 일시  [실무 권장]
    today_work          str  오늘의 작업 내용 [실무 권장]
    hazard_points       str  오늘의 위험요인 [실무 권장]
    safety_instructions str  핵심 안전수칙  [실무 권장]

Optional form_data keys:
    site_name, project_name
    tbm_location  str  실시 장소
    ppe_check     str  보호구 확인사항
    worker_opinion str  근로자 의견
    action_items  str  조치사항

    attendees  list[dict]  참석자 명단 (max 20)
        name     str  성명
        job_type str  직종
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "TBM안전점검일지"
SHEET_HEADING  = "TBM (Tool Box Meeting) 안전점검 일지"
SHEET_SUBTITLE = "작업 전 안전점검 회의 기록 — 고용노동부고시 제2023-19호 상시평가 실시 권장"

MAX_ATTENDEES = 20
TOTAL_COLS    = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_header_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "실시 일시", _v(data, "tbm_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "실시 장소", _v(data, "tbm_location"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_content_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "오늘의 안전 브리핑",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    for label, key, h in [
        ("오늘의 작업 내용",   "today_work",          60),
        ("오늘의 위험요인",    "hazard_points",        70),
        ("핵심 안전수칙",     "safety_instructions",  70),
        ("보호구 확인사항",    "ppe_check",           40),
        ("근로자 의견",        "worker_opinion",      50),
        ("조치사항",           "action_items",        40),
    ]:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, key),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_attendees_table(ws, start_row: int,
                            attendees: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "참석자 명단 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    # Header: 순번|성명|직종|서명 (반복 2쌍으로 8열 채움)
    for col_offset in [(1, 2, 3, 4), (5, 6, 7, 8)]:
        no_c, name_c, job_c, sig_c = col_offset
        _write_cell(ws, r, no_c,   no_c,   "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
        _write_cell(ws, r, name_c, name_c, "성명", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
        _write_cell(ws, r, job_c,  job_c,  "직종", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
        _write_cell(ws, r, sig_c,  sig_c,  "서명", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    half = MAX_ATTENDEES // 2  # 10

    for i in range(half):
        # 좌측 쌍 (i번째)
        left  = attendees[i]    if i < len(attendees) else {}
        # 우측 쌍 (i + half번째)
        right = attendees[i + half] if (i + half) < len(attendees) else {}

        _write_cell(ws, r, 1, 1, i + 1,             font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 2, _v(left, "name"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 3, 3, _v(left, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4, 4, "",                 font=_FONT_DEFAULT, align=_ALIGN_CENTER)  # 서명 공란

        _write_cell(ws, r, 5, 5, i + half + 1,       font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 6, 6, _v(right, "name"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7, 7, _v(right, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 8, 8, "",                  font=_FONT_DEFAULT, align=_ALIGN_CENTER)  # 서명 공란
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 2, "진행자 서명",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 8, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36
    return r + 1


def build_tbm_log_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 TBM 안전점검 일지 xlsx 바이너리를 반환."""
    data      = form_data or {}
    attendees = data.get("attendees") or []
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_header_block(ws, row, data)
    row = _write_content_block(ws, row, data)
    row = _write_attendees_table(ws, row, attendees)
    _write_confirmation(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
