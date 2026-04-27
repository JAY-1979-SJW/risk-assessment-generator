"""
용접·용단·화기작업 계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제241조~제252조
분류: PRACTICAL — 법정 별지 서식 없음.
      화기작업 계획 및 위험요인 사전관리 목적.
      PTW-002 화기작업 허가서(허가/승인 절차)와 구분되는 사전 계획 문서.

DOC_ID   = "EQ-014"
FORM_TYPE = "hot_work_workplan"

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자/기간
    work_location    str  작업장소
    work_content     str  화기작업 내용 설명
    work_types       list[str]  화기작업 종류
    safety_measures  str  안전조치 계획 (종합)

Optional form_data keys:
    project_name     str   공사명
    trade_name       str   작업공종
    contractor       str   작업업체
    supervisor       str   작업책임자
    prepared_by      str   작성자
    sign_date        str   작성일
    equipment_list   str   사용 장비·도구
    work_period      str   작업 기간
    combustibles_removed   str  가연물 제거 계획
    spark_prevention       str  불티 비산 방지 계획
    fire_blanket_plan      str  용접방화포 계획
    extinguisher_plan      str  소화기 비치 계획
    ventilation_plan       str  환기 계획
    fire_watch_required    str  화재감시자 배치 필요 여부 (판단)
    fire_watch_plan        str  화재감시 계획
    post_work_plan         str  작업 후 잔불 확인 계획
    emergency_measure      str  긴급조치 계획
    emergency_contact      str  비상연락망
    overall_opinion        str  종합 의견

    hazard_items  list[dict]  위험요인 분석 (repeat_field)
        hazard_type          str  위험 유형
        hazard_description   str  위험 요인 내용
        legal_reference      str  관련 법령 조항
        preventive_measure   str  예방 대책
        responsible_person   str  담당자
        check_method         str  확인 방법
        status               str  이행 상태
        remarks              str  비고
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DOC_ID    = "EQ-014"
FORM_TYPE = "hot_work_workplan"
SHEET_NAME    = "화기작업계획서"
SHEET_HEADING = "용접·용단·화기작업 계획서"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제241조~제252조에 따른 "
    "화기작업 사전 계획 및 위험요인 관리 문서 (EQ-014)"
)

MAX_HAZARD_ROWS = 10
MIN_HAZARD_ROWS = 5
TOTAL_COLS = 8

# ---------------------------------------------------------------------------
# 화기작업 종류 고정 목록
# ---------------------------------------------------------------------------

HOT_WORK_TYPES = [
    "용접",
    "용단",
    "그라인더",
    "절단",
    "금속 가열",
    "건식 연마",
    "기타 불꽃 발생 작업",
]

# ---------------------------------------------------------------------------
# 공지 문구
# ---------------------------------------------------------------------------

NOTICE_PURPOSE = (
    "본 문서는 화기작업 실시 전 계획 수립 및 위험요인 사전관리를 위한 문서이며, "
    "PTW-002 화기작업 허가서(허가/승인 절차)를 대체하지 않는다."
)
NOTICE_FIRE_WATCH = (
    "화재감시자 배치 필요 여부는 제241조의2 기준에 따라 작업장소, 가연성물질 유무, "
    "작업반경, 소화설비 상태 등 현장 조건을 종합하여 판단한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SECTION3 = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=22)
    return row + 1


def _write_notice(ws, row: int, text: str) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=_FILL_NOTICE,
                align=_ALIGN_LEFT, height=24)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 32
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    row += 1
    row = _write_notice(ws, row, NOTICE_PURPOSE)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업공종", _v(data, "trade_name"),  _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업업체", _v(data, "contractor"),  _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업책임자", _v(data, "supervisor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작성자",    _v(data, "prepared_by"), _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s2_work_plan(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 화기작업 계획 내용")
    _write_lv(ws, row, "작업일자/기간", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업기간",      _v(data, "work_period"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업장소", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=30)
    row += 1
    _write_lv(ws, row, "작업내용", _v(data, "work_content"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=48)
    row += 1
    _write_lv(ws, row, "사용 장비·도구", _v(data, "equipment_list"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=30)
    row += 1
    return row


def _write_s3_work_types(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 화기작업 유형  (해당 항목 선택)")
    selected = set(data.get("work_types") or [])
    col_count = 4
    col_span  = TOTAL_COLS // col_count
    for i in range(0, len(HOT_WORK_TYPES), col_count):
        batch = HOT_WORK_TYPES[i:i + col_count]
        for j, item in enumerate(batch):
            mark = "■" if item in selected else "□"
            c1 = 1 + j * col_span
            c2 = c1 + col_span - 1
            _write_cell(ws, row, c1, c2, f"{mark} {item}",
                        font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        for j in range(len(batch), col_count):
            c1 = 1 + j * col_span
            c2 = c1 + col_span - 1
            _write_cell(ws, row, c1, c2, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_s4_equipment(ws, row: int, data: Dict[str, Any]) -> int:
    return row


def _write_s5_hazard_analysis(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 위험요인 분석  (제241조 제2항 기준)")
    # 헤더 행
    headers = ["위험 유형", "위험 요인 내용", "관련 법령", "예방 대책", "담당자", "확인 방법", "이행 상태", "비고"]
    col_widths_for_header = [1, 2, 2, 2, 1, 1, 1, 1]  # 상대 너비 (열 수 기준)
    # 실제 열 배정: 8열을 아이템 수에 맞게
    col_splits = [1, 2, 3, 4, 5, 6, 7, 8]
    for idx, (hdr, col) in enumerate(zip(headers, col_splits)):
        _write_cell(ws, row, col, col, hdr,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=20)
    row += 1

    hazard_items: List[Dict] = data.get("hazard_items") or []
    display_rows = max(MIN_HAZARD_ROWS, len(hazard_items))
    display_rows = min(display_rows, MAX_HAZARD_ROWS)

    for i in range(display_rows):
        item = hazard_items[i] if i < len(hazard_items) else {}
        _write_cell(ws, row, 1, 1, _v(item, "hazard_type"),        font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 2, _v(item, "hazard_description"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 3, 3, _v(item, "legal_reference"),     font=_FONT_SMALL,   align=_ALIGN_CENTER)
        _write_cell(ws, row, 4, 4, _v(item, "preventive_measure"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 5, 5, _v(item, "responsible_person"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 6, 6, _v(item, "check_method"),        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 7, 7, _v(item, "status"),              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, _v(item, "remarks"),             font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 28
        row += 1
    return row


def _write_s6_combustibles(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 가연물 제거 및 불티 비산 방지 계획  (제241조 제2항)")
    _write_lv(ws, row, "가연물 제거 계획", _v(data, "combustibles_removed"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "불티 비산 방지 계획", _v(data, "spark_prevention"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "용접방화포 계획", _v(data, "fire_blanket_plan"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "환기 계획", _v(data, "ventilation_plan"),
              _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s7_fire_extinguisher(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 소화설비 및 비상대응 계획  (제243조, 제244조)", fill=_FILL_SECTION2)
    _write_lv(ws, row, "소화기 비치 계획", _v(data, "extinguisher_plan"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    _write_lv(ws, row, "긴급조치 계획", _v(data, "emergency_measure"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "비상연락망", _v(data, "emergency_contact"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    return row


def _write_s8_fire_watch(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 화재감시자 배치 판단 및 계획  (제241조의2)")
    row = _write_notice(ws, row, NOTICE_FIRE_WATCH)
    _write_lv(ws, row, "배치 필요 여부",
              _v(data, "fire_watch_required") or "□ 필요   □ 불필요   □ 현장 판단",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    _write_lv(ws, row, "화재감시 계획", _v(data, "fire_watch_plan"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    return row


def _write_s9_post_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 작업 후 잔불 확인 계획", fill=_FILL_SECTION3)
    _write_lv(ws, row, "잔불 확인 계획", _v(data, "post_work_plan"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    return row


def _write_s10_opinion(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 종합 의견 및 작성자 확인")
    _write_lv(ws, row, "종합 의견", _v(data, "overall_opinion"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=48)
    row += 1
    _write_lv(ws, row, "작성자 서명",
              _v(data, "prepared_by") or "",
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "작성일",
              _v(data, "sign_date") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_hot_work_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 용접·용단·화기작업 계획서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_plan(ws, row, data)
    row = _write_s3_work_types(ws, row, data)
    row = _write_s5_hazard_analysis(ws, row, data)
    row = _write_s6_combustibles(ws, row, data)
    row = _write_s7_fire_extinguisher(ws, row, data)
    row = _write_s8_fire_watch(ws, row, data)
    row = _write_s9_post_work(ws, row, data)
    _write_s10_opinion(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
