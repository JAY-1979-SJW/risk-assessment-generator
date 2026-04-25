"""
전기 작업계획서 — Excel 출력 모듈 (v1.0)  [WP-011]

법적 근거: 산업안전보건기준에 관한 규칙 제38조 제1항 (별표4 제5호 — 전기작업),
           제301조 이하 (전기작업 안전기준 — NEEDS_VERIFICATION)
           (evidence: WP-011-L1, PARTIAL_VERIFIED)

주의:
- 법정 별지 서식 없음 — 현장 실무 표준서식 (practical/custom workplan).
- 본 서류는 PTW-004 전기작업 허가서를 대체하지 않는다.
- 활선작업·LOTO 계획은 본 서식으로 작성하되 실제 LOTO 실시 확인은
  별도 PTW-004 허가서와 연계하여 관리한다.
- 조항 번호는 임의 추정 없이 PARTIAL_VERIFIED 상태로 기재.

Input — form_data dict:
    site_name              str|None   사업장명
    project_name           str|None   현장명
    work_date              str|None   작성일
    work_period            str|None   작업기간
    contractor             str|None   협력업체
    prepared_by            str|None   작성자
    reviewer               str|None   검토자

    work_name              str|None   작업명
    work_location          str|None   작업 위치
    work_datetime          str|None   작업 일시
    voltage                str|None   작업 전압
    work_category          str|None   작업 구분
    work_supervisor        str|None   작업책임자

    # 전기작업 유형 체크 (□ / ○ / ✓ 등)
    type_outage            str|None   정전작업
    type_live              str|None   활선작업
    type_near              str|None   근접작업
    type_temp_elec         str|None   임시전기 설치
    type_panel             str|None   분전반 작업
    type_cable             str|None   케이블 포설
    type_power_tool        str|None   전동공구 사용
    type_test_measure      str|None   시험·측정 작업

    # 전기 위험요인 체크
    hazard_electric_shock  str|None   감전
    hazard_arc             str|None   아크
    hazard_short_circuit   str|None   단락
    hazard_leakage         str|None   누전
    hazard_fire            str|None   화재
    hazard_explosion       str|None   폭발
    hazard_fall            str|None   추락
    hazard_pinch           str|None   협착

    # 작업 전 선행서류 확인
    prereq_ra001           str|None   RA-001 위험성평가표
    prereq_ra004           str|None   RA-004 TBM 일지
    prereq_ptw004          str|None   PTW-004 전기작업 허가서/LOTO
    prereq_cl004           str|None   CL-004 전기설비 정기 점검표
    prereq_ppe001          str|None   PPE-001 보호구 지급 대장

    # 정전 및 LOTO 계획
    loto_scope             str|None   정전 범위
    loto_breaker_location  str|None   차단기 위치
    loto_lock_installed    str|None   잠금장치 설치 여부
    loto_sign_attached     str|None   표지 부착 여부
    loto_residual_voltage  str|None   잔류전압 확인
    loto_re_energize       str|None   재투입 방지 조치

    # 활선·근접작업 안전조치
    live_approach_limit    str|None   접근한계거리 준수
    live_insulation_ppe    str|None   절연보호구 착용
    live_insulation_tools  str|None   절연공구 사용
    live_monitor           str|None   감시자 배치
    live_energized_protect str|None   충전부 방호

    # 임시전기 및 분전반 안전조치
    temp_elcb              str|None   누전차단기 설치
    temp_grounding         str|None   접지 실시
    temp_wire_protect      str|None   배선 보호
    temp_waterproof        str|None   방수 조치
    temp_overload          str|None   과부하 방지
    temp_panel_lock        str|None   분전반 잠금

    # 전동공구 점검 (5항목)
    tool_body_damage       str|None   외함 손상 없음
    tool_wire_insulation   str|None   전선 피복 양호
    tool_plug              str|None   플러그 양호
    tool_ground_wire       str|None   접지선 연결 확인
    tool_elcb              str|None   누전차단기 작동 확인

    # 보호구 및 측정장비 확인 체크
    ppe_insulated_gloves   str|None   절연장갑
    ppe_insulated_shoes    str|None   절연화
    ppe_face_shield        str|None   보안면
    ppe_insulation_mat     str|None   절연매트
    equip_voltage_tester   str|None   검전기
    equip_insulation_meter str|None   절연저항계

    # 작업 중 관리계획
    mgmt_zone_control      str|None   작업구역 통제
    mgmt_monitor           str|None   감시자 배치
    mgmt_emergency_stop    str|None   비상정지 절차
    mgmt_fire_response     str|None   화재 대응
    mgmt_reenergize_proc   str|None   재통전 절차

    nonconformance_items   list[dict] 부적합 사항 (MAX_NC=5)
        content            str|None   부적합 내용
        action             str|None   시정조치
        deadline           str|None   완료 기한
        completed          str|None   완료 여부

    work_verdict           str|None   판정 (적합/조건부 적합/작업중지)
    verdict_condition      str|None   조건부 적합 조건

    sign_date              str|None   서명 날짜

Output — xlsx bytes (in-memory). 파일 저장은 호출자 책임.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME     = "전기작업계획서"
SHEET_HEADING  = "전기 작업계획서"
SHEET_SUBTITLE = (
    "산업안전보건기준에 관한 규칙 제38조 제1항(별표4 제5호)·제301조 이하에 따른 전기작업 안전관리 계획서"
)
ATTACH_NOTE = (
    "※ 본 서류는 법정 별지 서식이 아닌 실무 표준서식이며, "
    "PTW-004 전기작업 허가서·LOTO를 대체하지 않음"
)

MAX_NC      = 5
TOTAL_COLS  = 8

_DEFAULT_TOOL_CHECKS: List[Dict[str, Any]] = [
    {"check_item": "외함 손상 없음",         "result": "", "note": ""},
    {"check_item": "전선 피복 손상 없음",     "result": "", "note": ""},
    {"check_item": "플러그 이상 없음",        "result": "", "note": ""},
    {"check_item": "접지선 연결 확인",        "result": "", "note": ""},
    {"check_item": "누전차단기 작동 확인",    "result": "", "note": ""},
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8, italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 열 너비
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 14, 2: 12, 3: 12, 4: 12,
    5: 12, 6: 12, 7: 12, 8: 10,
}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int,
              height: int = 20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _section_header(ws, row: int, title: str, height: int = 18) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=height)
    return row + 1


def _write_4up_row(ws, row: int,
                   items: List[tuple],
                   height: int = 22) -> None:
    """4개 항목을 한 행에 배치 — (label, value) 4쌍, 각 2열 차지."""
    pairs = (items + [("", "")] * 4)[:4]
    for i, (lbl, val) in enumerate(pairs):
        c1 = i * 2 + 1
        c2 = c1 + 1
        _write_cell(ws, row, c1, c1, lbl,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
        _write_cell(ws, row, c2, c2, val or "",
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = height


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_basic_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)

    r += 1
    _write_lv(ws, r, "작성일", _v(data, "work_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업기간", _v(data, "work_period"),
              _L2, _V2_START, _V2_END)

    r += 1
    _write_lv(ws, r, "협력업체", _v(data, "contractor"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)

    r += 1
    _write_lv(ws, r, "작성자", _v(data, "prepared_by"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "검토자", _v(data, "reviewer"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_work_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 기본정보")

    _write_lv(ws, r, "작업명", _v(data, "work_name"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "작업 위치", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "작업 일시", _v(data, "work_datetime"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "작업 전압", _v(data, "voltage"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 구분", _v(data, "work_category"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "작업책임자", _v(data, "work_supervisor"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)

    return r + 1


def _write_work_types(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "전기작업 유형  (해당 항목에 ○ 표시)")

    _write_4up_row(ws, r, [
        ("정전작업",     _v(data, "type_outage")),
        ("활선작업",     _v(data, "type_live")),
        ("근접작업",     _v(data, "type_near")),
        ("임시전기 설치", _v(data, "type_temp_elec")),
    ])
    r += 1
    _write_4up_row(ws, r, [
        ("분전반 작업",    _v(data, "type_panel")),
        ("케이블 포설",    _v(data, "type_cable")),
        ("전동공구 사용",  _v(data, "type_power_tool")),
        ("시험·측정 작업", _v(data, "type_test_measure")),
    ])

    return r + 1


def _write_hazard_check(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "전기 위험요인 확인  (해당 항목에 ○ 표시)")

    _write_4up_row(ws, r, [
        ("감전",  _v(data, "hazard_electric_shock")),
        ("아크",  _v(data, "hazard_arc")),
        ("단락",  _v(data, "hazard_short_circuit")),
        ("누전",  _v(data, "hazard_leakage")),
    ])
    r += 1
    _write_4up_row(ws, r, [
        ("화재",  _v(data, "hazard_fire")),
        ("폭발",  _v(data, "hazard_explosion")),
        ("추락",  _v(data, "hazard_fall")),
        ("협착",  _v(data, "hazard_pinch")),
    ])

    return r + 1


def _write_prereq_docs(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 전 선행서류 확인")

    # 헤더
    _write_cell(ws, r, 1, 1, "서류 ID",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 5, "서류명",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 7, "확인 여부",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "비고",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    prereqs = [
        ("RA-001",  "위험성평가표",                  "prereq_ra001"),
        ("RA-004",  "TBM 일지",                      "prereq_ra004"),
        ("PTW-004", "전기작업 허가서 / LOTO",         "prereq_ptw004"),
        ("CL-004",  "전기설비 정기 점검표",            "prereq_cl004"),
        ("PPE-001", "보호구 지급 대장",               "prereq_ppe001"),
    ]
    for doc_id, doc_name, key in prereqs:
        _write_cell(ws, r, 1, 1, doc_id,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, doc_name,
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 7, _v(data, key),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, "",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_loto_plan(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "정전 및 LOTO 계획  (정전작업 시 필수 작성)")

    rows = [
        ("정전 범위",          "loto_scope",            25),
        ("차단기 위치",        "loto_breaker_location",  22),
        ("잔류전압 확인",      "loto_residual_voltage",  22),
        ("재투입 방지 조치",   "loto_re_energize",       22),
    ]
    for label, key, h in rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, key), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    _write_lv(ws, r, "잠금장치 설치", _v(data, "loto_lock_installed"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "표지 부착",     _v(data, "loto_sign_attached"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_live_work_safety(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "활선·근접작업 안전조치  (활선·근접작업 시 필수 작성)")

    rows = [
        ("접근한계거리 준수",   "live_approach_limit",    22),
        ("충전부 방호",         "live_energized_protect", 22),
        ("감시자 배치",         "live_monitor",           22),
    ]
    for label, key, h in rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, key), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    _write_lv(ws, r, "절연보호구 착용", _v(data, "live_insulation_ppe"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "절연공구 사용",   _v(data, "live_insulation_tools"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_temp_elec_safety(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "임시전기 및 분전반 안전조치  (임시전기·분전반 작업 시 필수 작성)")

    _write_lv(ws, r, "누전차단기 설치", _v(data, "temp_elcb"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "접지 실시",       _v(data, "temp_grounding"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "배선 보호",       _v(data, "temp_wire_protect"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "방수 조치",       _v(data, "temp_waterproof"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "과부하 방지",     _v(data, "temp_overload"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "분전반 잠금",     _v(data, "temp_panel_lock"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_tool_check_table(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "전동공구 및 이동식 전기기계기구 점검  (전동공구 사용 시 필수)")

    # 헤더
    _write_cell(ws, r, 1, 1, "순번",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 5, "점검 항목",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 7, "점검 결과",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "비고",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    tool_keys = [
        "tool_body_damage",
        "tool_wire_insulation",
        "tool_plug",
        "tool_ground_wire",
        "tool_elcb",
    ]
    for i, (default_item, key) in enumerate(
        zip([c["check_item"] for c in _DEFAULT_TOOL_CHECKS], tool_keys)
    ):
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, default_item,
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 7, _v(data, key),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, "",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_ppe_equipment_check(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "보호구 및 측정장비 확인  (해당 항목에 ○ 표시)")

    _write_4up_row(ws, r, [
        ("절연장갑",   _v(data, "ppe_insulated_gloves")),
        ("절연화",     _v(data, "ppe_insulated_shoes")),
        ("보안면",     _v(data, "ppe_face_shield")),
        ("절연매트",   _v(data, "ppe_insulation_mat")),
    ])
    r += 1
    _write_4up_row(ws, r, [
        ("검전기",      _v(data, "equip_voltage_tester")),
        ("절연저항계",  _v(data, "equip_insulation_meter")),
        ("",            ""),
        ("",            ""),
    ])

    return r + 1


def _write_work_mgmt(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 중 관리계획")

    rows = [
        ("작업구역 통제", "mgmt_zone_control",    22),
        ("감시자 배치",   "mgmt_monitor",          22),
        ("비상정지 절차", "mgmt_emergency_stop",   22),
        ("화재 대응",     "mgmt_fire_response",    22),
        ("재통전 절차",   "mgmt_reenergize_proc",  22),
    ]
    for label, key, h in rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, key), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_nonconformance(ws, start_row: int,
                          items: List[Dict[str, Any]]) -> int:
    r = _section_header(ws, start_row, "부적합 사항 및 시정조치")

    # 헤더
    _write_cell(ws, r, 1, 1, "순번",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4, "부적합 내용",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "시정조치",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "기한",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "완료",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_NC):
        item = items[i] if i < len(items) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(item, "content"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 6, _v(item, "action"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7, 7, _v(item, "deadline"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, _v(item, "completed"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 25
        r += 1

    return r


def _write_verdict(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 가능 여부 판정")

    verdict = _v(data, "work_verdict") or ""
    condition = _v(data, "verdict_condition") or ""

    verdict_text = "  적합  □    조건부 적합  □    작업중지  □"
    if verdict:
        verdict_text = f"  판정: {verdict}"
        if condition:
            verdict_text += f"  (조건: {condition})"

    _write_cell(ws, r, 1, TOTAL_COLS, verdict_text,
                font=_FONT_BOLD, fill=_FILL_WARN,
                align=_ALIGN_CENTER, height=26)
    r += 1

    _write_cell(ws, r, 1, TOTAL_COLS, ATTACH_NOTE,
                font=_FONT_SMALL, fill=_FILL_WARN,
                align=_ALIGN_CENTER, height=20)

    return r + 1


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "확인 서명")

    _write_cell(ws, r, 1, 2, "작성자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_cell(ws, r, 1, 2, "관리감독자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "안전관리자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 8, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_electrical_workplan_sheet(ws, form_data: Dict[str, Any]) -> None:
    """주어진 worksheet에 전기 작업계획서를 렌더링."""
    data  = form_data or {}
    nc_items = data.get("nonconformance_items") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_work_info(ws, row, data)
    row = _write_work_types(ws, row, data)
    row = _write_hazard_check(ws, row, data)
    row = _write_prereq_docs(ws, row, data)
    row = _write_loto_plan(ws, row, data)
    row = _write_live_work_safety(ws, row, data)
    row = _write_temp_elec_safety(ws, row, data)
    row = _write_tool_check_table(ws, row, data)
    row = _write_ppe_equipment_check(ws, row, data)
    row = _write_work_mgmt(ws, row, data)
    row = _write_nonconformance(ws, row, nc_items)
    row = _write_verdict(ws, row, data)
    _write_confirmation(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_electrical_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 전기 작업계획서 xlsx 바이너리를 반환.

    Args:
        form_data: 입력 스키마 준수 dict.
            nonconformance_items > MAX_NC(5) 시 초과분 무시.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_electrical_workplan_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
