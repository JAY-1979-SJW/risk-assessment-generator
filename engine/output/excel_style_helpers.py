"""
신규 안전서류 builder용 Excel 스타일 공통 helper.

신규 builder용 공통 helper이며 기존 builder는 아직 이 파일을 사용하지 않음.
기존 46종 builder는 각자 동일한 스타일 상수/함수를 인라인으로 유지한다.

Usage (신규 builder에서):
    from engine.output.excel_style_helpers import (
        FONT_TITLE, FONT_BOLD, FONT_DEFAULT,
        FILL_LABEL, FILL_SECTION, BORDER,
        ALIGN_CENTER, ALIGN_LEFT, ALIGN_LABEL,
        v, write_cell, border_rect, apply_col_widths, write_blank_row,
    )
"""
from __future__ import annotations

from collections.abc import Mapping
from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 폰트
# ---------------------------------------------------------------------------

FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
FONT_SMALL    = Font(name="맑은 고딕", size=9)
FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

# ---------------------------------------------------------------------------
# 배경색
# ---------------------------------------------------------------------------

FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FFFBE6")
FILL_NONE    = PatternFill()

# ---------------------------------------------------------------------------
# 테두리
# ---------------------------------------------------------------------------

THIN   = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# 정렬
# ---------------------------------------------------------------------------

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LABEL  = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# 함수
# ---------------------------------------------------------------------------

def v(data: Mapping[str, Any], key: str, default: str = "") -> Any:
    """dict에서 값을 반환한다. None이면 default(기본: 빈 문자열)를 반환한다."""
    val = data.get(key)
    return default if val is None else val


def write_cell(
    ws,
    row: int,
    col1: int,
    col2: int,
    value: Any,
    *,
    font=None,
    fill=None,
    align=None,
    border: bool = True,
    height: Optional[float] = None,
) -> None:
    """셀을 기록한다. col1 != col2이면 merge_cells를 수행한다.

    border=True(기본)이면 BORDER를 적용한다.
    align이 없으면 ALIGN_LEFT를 기본값으로 사용한다.
    """
    if col2 < col1:
        raise ValueError(f"col2 ({col2}) must be >= col1 ({col1})")
    if col2 > col1:
        ws.merge_cells(
            start_row=row, start_column=col1,
            end_row=row,   end_column=col2,
        )
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or FONT_DEFAULT
    cell.fill      = fill  or FILL_NONE
    cell.alignment = align or ALIGN_LEFT
    if border:
        border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """지정 범위의 모든 셀에 BORDER를 적용한다."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def apply_col_widths(ws, widths: Mapping[int, float]) -> None:
    """열 번호(int) → 너비(float) 매핑으로 열 너비를 설정한다."""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_blank_row(
    ws,
    row: int,
    total_cols: int,
    *,
    height: Optional[float] = None,
) -> None:
    """빈 행을 기록한다. 각 셀에 빈 문자열과 BORDER를 적용한다."""
    for c in range(1, total_cols + 1):
        write_cell(ws, row, c, c, "", height=height)


def apply_a4_page_setup(
    ws,
    *,
    landscape: bool = False,
    left: float = 0.5,
    right: float = 0.5,
    top: float = 0.5,
    bottom: float = 0.5,
) -> None:
    """A4 인쇄 설정을 worksheet에 적용한다.

    fitToWidth=1, fitToHeight=0으로 설정해 폭은 1페이지에 맞추고
    높이는 자연 분할한다. scale은 사용하지 않는다.
    """
    ps = ws.page_setup
    ps.paperSize  = 9          # A4
    ps.orientation = "landscape" if landscape else "portrait"
    ps.fitToWidth  = 1
    ps.fitToHeight = 0
    ps.scale       = None

    ws.page_margins = PageMargins(
        left=left, right=right, top=top, bottom=bottom,
        header=0.3, footer=0.3,
    )
    set_print_area_to_used_range(ws)


SIGNATURE_KEYWORDS = ("서명", "확인자", "승인자", "작성자", "검토자")


def apply_repeating_header_rows(ws, end_row: int) -> None:
    """인쇄 반복 행을 1:end_row 범위로 설정한다."""
    ws.print_title_rows = f"1:{end_row}"


def normalize_signature_row_heights(ws, min_height: float = 20.0) -> None:
    """서명 관련 키워드가 있는 행의 높이를 min_height 이상으로 보정한다."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and any(kw in str(cell.value) for kw in SIGNATURE_KEYWORDS):
                r = cell.row
                dim = ws.row_dimensions[r]
                current = dim.height if dim.height else 15.0
                if current < min_height:
                    ws.row_dimensions[r].height = min_height
                break  # 행당 1회만


def set_print_area_to_used_range(ws) -> None:
    """실제 사용 범위(max_row × max_column)를 print_area로 설정한다."""
    max_r = ws.max_row    or 1
    max_c = ws.max_column or 1
    end_col_letter = get_column_letter(max_c)
    ws.print_area = f"A1:{end_col_letter}{max_r}"
