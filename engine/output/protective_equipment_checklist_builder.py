"""
보호구 지급 및 관리 점검표 — Excel 출력 모듈 (v1.0)  [CL-008]

법적 근거:
    산업안전보건기준에 관한 규칙 제32조 — 보호구 지급 의무
    산업안전보건법 제39조 — 보호구 사용 및 관리

분류: PRACTICAL — 법정 별지 서식 없음, 실무 자체 표준서식

요약:
    현장에서 근로자에게 지급되는 보호구의 지급·관리·상태 점검을 기록하는 서식입니다.
    보호구 종류별 지급 현황, 상태 점검(파손·오염·유효기간), 착용 관리를 통합 관리합니다.
    부적합 항목 발견 시 개선조치를 기록하고 추적합니다.

Required form_data keys:
    site_name         str  사업장명
    check_date        str  점검일자
    inspector_name    str  점검자 성명

Optional form_data keys:
    project_name               str  현장명
    work_trade                 str  작업공종
    work_location              str  작업장소
    work_description           str  작업내용
    target_workers             str  대상 근로자 수
    work_commander_name        str  작업책임자
    department                 str  소속
    position                   str  직책
    helmet_supplied            str  안전모 지급 여부
    helmet_status              str  안전모 상태
    safety_shoes_supplied      str  안전화 지급 여부
    safety_shoes_status        str  안전화 상태
    safety_belt_supplied       str  안전대 지급 여부
    safety_belt_status         str  안전대 상태
    safety_glasses_supplied    str  보안경 지급 여부
    safety_glasses_status      str  보안경 상태
    dust_mask_supplied         str  방진마스크 지급 여부
    dust_mask_status           str  방진마스크 상태
    respirator_supplied        str  방독마스크 지급 여부
    respirator_status          str  방독마스크 상태
    ear_protection_supplied    str  귀마개/귀덮개 지급 여부
    ear_protection_status      str  귀마개/귀덮개 상태
    gloves_supplied            str  보호장갑 지급 여부
    gloves_status              str  보호장갑 상태
    face_shield_supplied       str  보안면 지급 여부
    face_shield_status         str  보안면 상태
    protective_clothing_supplied str  방열복/방염복 지급 여부
    protective_clothing_status   str  방열복/방염복 상태
    other_ppe_supplied         str  기타 보호구 지급 여부
    other_ppe_description      str  기타 보호구 설명
    other_ppe_status           str  기타 보호구 상태
    wearability_check          str  착용 가능 여부
    replacement_needed         str  교체 필요 여부
    pre_work_check             str  작업 전 착용 확인
    wearing_education          str  착용 방법 교육 여부
    non_wearer_action          str  미착용 근로자 조치
    storage_condition          str  보호구 보관 상태
    spare_ppe_status           str  예비 보호구 비치 여부
    nonconformance_items       list[dict]  부적합 항목 (no, content, action, responsible, deadline, completed)
    judgment                   str  종합 판정 (적합/조건부 적합/부적합)
    judgment_reason            str  판정 사유
    preparer_name              str  작성자
    supervisor_sign            str  관리감독자 서명
    safety_manager_sign        str  안전관리자 서명
    sign_date                  str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.output.excel_style_helpers import normalize_signature_row_heights

SHEET_NAME    = "보호구점검표"
SHEET_HEADING = "보호구 지급 및 관리 점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제32조에 따른 보호구 지급·관리 현황 점검"
DOC_ID = "CL-008"

TOTAL_COLS = 9
MAX_NONCONFORM = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 10, 7: 12, 8: 12, 9: 9}

_PPE_ITEMS = [
    ("안전모", "helmet"),
    ("안전화", "safety_shoes"),
    ("안전대", "safety_belt"),
    ("보안경", "safety_glasses"),
    ("방진마스크", "dust_mask"),
    ("방독마스크", "respirator"),
    ("귀마개/귀덮개", "ear_protection"),
    ("보호장갑", "gloves"),
    ("보안면", "face_shield"),
    ("방열복/방염복", "protective_clothing"),
    ("기타 보호구", "other_ppe"),
]

_JUDGMENT_OPTIONS = ["적합", "조건부 적합", "부적합"]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_protective_equipment_checklist(form_data: Dict[str, Any]) -> bytes:
    ws = Workbook().active
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    _apply_col_widths(ws)

    row = 1

    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
                font=_FONT_TITLE, align=_ALIGN_CENTER, height=25)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_SUBTITLE,
                font=_FONT_SUBTITLE, align=_ALIGN_CENTER)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, f"[{DOC_ID}]",
                font=_FONT_SMALL, align=_ALIGN_CENTER)
    row += 1

    row += 1

    _write_lv(ws, row, "사업장명", _v(form_data, "site_name"), 1, 2, 3)
    _write_lv(ws, row, "현장명", _v(form_data, "project_name"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "점검일자", _v(form_data, "check_date"), 1, 2, 3)
    _write_lv(ws, row, "점검자", _v(form_data, "inspector_name"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "소속", _v(form_data, "department"), 1, 2, 3)
    _write_lv(ws, row, "직책", _v(form_data, "position"), 5, 6, 7)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 보호구 지급 대상 정보",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "작업공종", _v(form_data, "work_trade"), 1, 2, 3)
    _write_lv(ws, row, "대상 근로자 수", _v(form_data, "target_workers"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "작업장소", _v(form_data, "work_location"), 1, 2, 3)
    _write_lv(ws, row, "작업책임자", _v(form_data, "work_commander_name"), 5, 6, 7)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, f"작업내용: {_v(form_data, 'work_description')}",
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 보호구 지급 및 상태 점검",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_cell(ws, row, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 2, 3, "보호구 종류", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 4, 4, "지급", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 5, 6, "상태 점검", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 7, 8, "개선조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 9, 9, "비고", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    for idx, (ppe_name, ppe_key) in enumerate(_PPE_ITEMS, 1):
        supplied = _v(form_data, f"{ppe_key}_supplied")
        status = _v(form_data, f"{ppe_key}_status")

        _write_cell(ws, row, 1, 1, str(idx), align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 3, ppe_name, align=_ALIGN_LEFT)
        _write_cell(ws, row, 4, 4, supplied, align=_ALIGN_CENTER)
        _write_cell(ws, row, 5, 6, status, align=_ALIGN_LEFT)
        _write_cell(ws, row, 7, 8, "", align=_ALIGN_LEFT)
        _write_cell(ws, row, 9, 9, "", align=_ALIGN_CENTER)
        row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 보호구 착용 및 관리",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "착용 가능 여부", _v(form_data, "wearability_check"), 1, 2, 4)
    _write_lv(ws, row, "교체 필요 여부", _v(form_data, "replacement_needed"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "작업 전 착용 확인", _v(form_data, "pre_work_check"), 1, 2, 4)
    _write_lv(ws, row, "착용 방법 교육", _v(form_data, "wearing_education"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "미착용자 조치", _v(form_data, "non_wearer_action"), 1, 2, 4)
    _write_lv(ws, row, "보관 상태", _v(form_data, "storage_condition"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "예비 보호구 비치", _v(form_data, "spare_ppe_status"), 1, 2, TOTAL_COLS)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 부적합 및 개선조치",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_cell(ws, row, 1, 1, "번호", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 2, 4, "부적합 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 5, 6, "개선조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 7, 7, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "예정일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 9, 9, "완료", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    nonconformance_items = form_data.get("nonconformance_items", [])
    for idx in range(1, MAX_NONCONFORM + 1):
        item = nonconformance_items[idx - 1] if idx <= len(nonconformance_items) else {}

        _write_cell(ws, row, 1, 1, str(idx), align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 4, item.get("content", ""), align=_ALIGN_LEFT)
        _write_cell(ws, row, 5, 6, item.get("action", ""), align=_ALIGN_LEFT)
        _write_cell(ws, row, 7, 7, item.get("responsible", ""), align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, item.get("deadline", ""), align=_ALIGN_CENTER)
        _write_cell(ws, row, 9, 9, item.get("completed", ""), align=_ALIGN_CENTER)
        row += 1

    row += 1

    _write_lv(ws, row, "종합 판정", _v(form_data, "judgment"), 1, 2, 4)
    _write_lv(ws, row, "판정 사유", _v(form_data, "judgment_reason"), 5, 6, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, 3, f"작성자: {_v(form_data, 'preparer_name')}", align=_ALIGN_LEFT)
    _write_cell(ws, row, 4, 6, f"관리감독자: {_v(form_data, 'supervisor_sign')}", align=_ALIGN_LEFT)
    _write_cell(ws, row, 7, TOTAL_COLS, f"서명일: {_v(form_data, 'sign_date')}", align=_ALIGN_LEFT)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, f"안전관리자: {_v(form_data, 'safety_manager_sign')}", align=_ALIGN_LEFT)

    normalize_signature_row_heights(ws, min_height=18.0)
    ws.print_title_rows = "1:4"
    output = BytesIO()
    ws.parent.save(output)
    output.seek(0)
    return output.read()
