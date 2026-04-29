"""
화기작업 허가서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제236조, 제240조, 제241조,
           제241조의2, 제243조, 제244조
           제241조 제4항 — 화재위험작업 작업내용·일시·안전점검 서면 게시 의무
분류: PRACTICAL — 법정 별지 서식 없음.
      KOSHA GUIDE P-94-2021 별지 양식1 참고, 법령 의무 항목 기반 자체 표준 서식.

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자
    work_time        str  작업시간
    work_location    str  작업장소
    trade_name       str  작업공종
    work_content     str  작업내용
    contractor       str  작업업체
    work_supervisor  str  작업책임자

Optional form_data keys:
    project_name        str   공사명
    permit_no           str   관리번호
    equipment_list      str   사용장비
    combustibles_present str  가연물 존재 여부
    combustibles_removed str  가연물 제거 여부
    spark_prevention    str   불티 비산 방지 조치
    fire_blanket_used   str   용접방화포 사용 여부
    extinguisher_placed str   소화기 비치 여부
    ventilation_status  str   환기 상태
    fire_watch_required str   화재감시자 배치 여부
    fire_watch_name     str   화재감시자 성명
    permit_issuer       str   작업허가자 (서명)
    supervisor_name     str   관리감독자 (서명)
    post_work_confirmer str   작업 종료 후 잔불 확인자
    final_sign          str   최종 확인 서명
    work_types          list[str]  화기작업 종류 선택 항목
    pre_work_checks     list[str]  작업 전 안전조치 이행 항목 (체크된 항목명)
    workers             list[dict] name, job_type — 작업자 명단 (max 10)
    photo_items         list[str]  사진 증빙 항목 (이행 표시)
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "화기작업허가서"
SHEET_HEADING = "화기작업 허가서"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제241조 제4항에 따른 "
    "화재위험작업 현장 게시·허가 기록서"
)

MAX_WORKERS = 10
TOTAL_COLS  = 8

# ---------------------------------------------------------------------------
# 고정 문구
# ---------------------------------------------------------------------------

NOTICE_NO_CERT = (
    "본 화기작업 허가서는 작업 전 화재위험 확인 및 안전조치 기록이며, "
    "법정 안전보건교육 수료증을 대체하지 않는다."
)
NOTICE_FIRE_WATCH = (
    "화재감시자 배치 여부는 작업장소, 가연물 위치, 작업반경, 소화설비 상태 등 "
    "현장 조건에 따라 최종 판단한다."
)
NOTICE_PERMIT_ISSUER = (
    "작업허가자는 작업 전 안전조치 이행 여부를 확인한 후 작업을 허가해야 한다."
)
NOTICE_POST_WORK = (
    "작업 종료 후 잔불, 불티, 가열부, 주변 가연물 상태를 확인하고 최종 서명한다."
)
NOTICE_PHOTO = (
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)

# ---------------------------------------------------------------------------
# 화기작업 종류 고정 목록
# ---------------------------------------------------------------------------

HOT_WORK_TYPES = [
    "용접",
    "용단",
    "그라인더",
    "절단",
    "금속 가열",
    "건식 연마",
    "기타 불꽃 발생 작업",
]

# ---------------------------------------------------------------------------
# 작업 전 안전조치 체크 항목
# ---------------------------------------------------------------------------

PRE_WORK_CHECK_ITEMS = [
    "작업내용·작업일시·안전점검 사항 게시 (제241조 제4항)",
    "작업장 주변 가연물 제거",
    "불티 비산 방지 조치",
    "용접방화포 또는 방염포 설치",
    "소화기 또는 소화설비 비치",
    "인화성 물질 보관 여부 확인",
    "가스용기 전도 방지",
    "산소·가스 호스 손상 여부 확인",
    "전기용접기 접지 및 누전 여부 확인",
    "환기 상태 확인",
    "작업반경 출입통제",
    "작업 후 잔불 확인 계획 수립",
]

# ---------------------------------------------------------------------------
# 소화설비 및 방화조치 체크 항목
# ---------------------------------------------------------------------------

FIRE_EXT_CHECK_ITEMS = [
    "소화기 비치 위치·수량 확인",
    "소화기 압력 정상 여부 확인",
    "소화설비(스프링클러 등) 작동 상태 확인",
    "방화 안전거리 확보 (제244조)",
    "인화성·가연성 물질 격리",
]

# ---------------------------------------------------------------------------
# 화재감시자 조건부 판단 항목
# ---------------------------------------------------------------------------

FIRE_WATCH_CONDITION_ITEMS = [
    "용접·용단 작업 해당 여부",
    "작업반경 11m 이내 가연성물질 존재 여부",
    "불티 비산 가능성",
    "소화기구 비치 여부",
    "경보설비 작동 상태",
    "상시·반복 작업 여부 (배치 면제 조건 해당 검토)",
]

# ---------------------------------------------------------------------------
# 보호구 및 작업장비 확인 항목
# ---------------------------------------------------------------------------

PPE_CHECK_ITEMS = [
    "용접 마스크·보안면",
    "용접 장갑 (가죽 또는 방염)",
    "방염 작업복",
    "안전화",
    "귀마개 또는 귀덮개 (소음 발생 시)",
    "방진마스크 또는 방독마스크 (흄 발생 시)",
    "안전대 (고소 병행 작업 시)",
]

# ---------------------------------------------------------------------------
# 작업 종료 후 확인 항목
# ---------------------------------------------------------------------------

POST_WORK_CHECK_ITEMS = [
    "잔불 확인",
    "불티 비산 잔여 여부 확인",
    "가열부 냉각 확인",
    "주변 가연물 재점검",
    "가스밸브 차단",
    "전원 차단",
    "작업장 정리정돈",
]

# ---------------------------------------------------------------------------
# 사진 증빙 항목
# ---------------------------------------------------------------------------

PHOTO_ITEMS = [
    "작업 전 작업장소 사진",
    "가연물 제거 사진",
    "소화기·방화포 비치 사진",
    "화기작업 진행 사진",
    "작업 종료 후 잔불 확인 사진",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=_FILL_NOTICE,
                align=_ALIGN_LEFT, height=24)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    row += 1
    row = _write_notice(ws, row, NOTICE_NO_CERT)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리번호", _v(data, "permit_no"),   _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 작업 기본정보")
    _write_lv(ws, row, "작업일자", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업시간", _v(data, "work_time"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업장소", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "작업공종", _v(data, "trade_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업업체", _v(data, "contractor"),    _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업내용", _v(data, "work_content"),  _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "작업책임자", _v(data, "work_supervisor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "사용장비",   _v(data, "equipment_list"),  _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s3_work_types(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 화기작업 유형  (해당 항목 선택)")
    selected = set(data.get("work_types") or [])
    col_count = 4
    items = HOT_WORK_TYPES
    for i in range(0, len(items), col_count):
        batch = items[i:i + col_count]
        col_span = TOTAL_COLS // col_count
        for j, item in enumerate(batch):
            mark = "■" if item in selected else "□"
            c1 = 1 + j * col_span
            c2 = c1 + col_span - 1
            _write_cell(ws, row, c1, c2, f"{mark} {item}",
                        font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        # fill remaining columns if batch is short
        for j in range(len(batch), col_count):
            c1 = 1 + j * col_span
            c2 = c1 + col_span - 1
            _write_cell(ws, row, c1, c2, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 20
        row += 1
    # 작업자 명단
    workers = data.get("workers") or []
    _write_cell(ws, row, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    _write_cell(ws, row, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 2, 5, "성명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 8, "직종/역할", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, row, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 5, _v(item, "name"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 8, _v(item, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_s4_combustibles(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 작업장소 및 가연물 확인")
    _write_lv(ws, row, "가연물 존재 여부", _v(data, "combustibles_present"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "가연물 제거 여부", _v(data, "combustibles_removed"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "위험물 보관 현황",
              _v(data, "hazmat_storage") or "확인 후 기재",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=30)
    row += 1
    _write_lv(ws, row, "인화성 증기·가스 발생 여부",
              _v(data, "flammable_vapor") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s5_pre_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 작업 전 안전조치  (이행 항목에 ■ 표시)")
    checked = set(data.get("pre_work_checks") or [])
    for item in PRE_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "비산방지 조치 내용",
              _v(data, "spark_prevention") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    return row


def _write_s6_fire_ext(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 소화설비 및 방화조치  (제243조, 제244조)")
    checked = set(data.get("fire_ext_checks") or [])
    for item in FIRE_EXT_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "소화기 비치 확인",
              _v(data, "extinguisher_placed") or "",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "환기 상태",
              _v(data, "ventilation_status") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "용접방화포 사용 여부",
              _v(data, "fire_blanket_used") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s7_fire_watch(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 화재감시자 배치 확인  (제241조의2)")
    row = _write_notice(ws, row, NOTICE_FIRE_WATCH)
    checked = set(data.get("fire_watch_conditions") or [])
    for item in FIRE_WATCH_CONDITION_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "화재감시자 배치 여부",
              _v(data, "fire_watch_required") or "□ 필요   □ 불필요   □ 현장 판단",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "화재감시자 성명",
              _v(data, "fire_watch_name") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "화재감시자 장비 지급 확인",
              _v(data, "fire_watch_equipment") or "□ 확성기  □ 휴대조명  □ 대피마스크",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    return row


def _write_s8_ppe(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 보호구 및 작업장비 확인  (제241조 제2항 제6호)")
    checked = set(data.get("ppe_checks") or [])
    for item in PPE_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s9_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 작업허가 승인", fill=_FILL_SECTION2)
    row = _write_notice(ws, row, NOTICE_PERMIT_ISSUER)
    _write_lv(ws, row, "작업신청자 (서명)", _v(data, "work_supervisor"),
              _L1, _V1_START, _V1_END, height=32)
    _write_lv(ws, row, "확인자 (서명)",     _v(data, "supervisor_name"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업허가자 (서명)", _v(data, "permit_issuer"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    _write_lv(ws, row, "허가 유효기간",
              _v(data, "validity_period") or "당일 1회 작업 한정 (KOSHA P-94-2021 참고)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    return row


def _write_s10_during_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 작업 중 점검")
    _write_cell(ws, row, 1, TOTAL_COLS,
                "이상 발생 (화재 징후, 가스 누출, 가연물 노출 등) 시 즉시 작업 중단 후 허가자에게 보고",
                font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=24)
    row += 1
    _write_lv(ws, row, "작업 중 이상 사항",
              _v(data, "during_work_issues") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    return row


def _write_s11_post_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 작업 종료 후 잔불 확인")
    row = _write_notice(ws, row, NOTICE_POST_WORK)
    checked = set(data.get("post_work_checks") or [])
    for item in POST_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "작업 종료 시각",
              _v(data, "work_end_time") or "",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "잔불 확인자 서명",
              _v(data, "post_work_confirmer") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s12_photos(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 사진 및 증빙자료  [OPTIONAL]")
    row = _write_notice(ws, row, NOTICE_PHOTO)
    taken_photos = set(data.get("photo_items") or [])
    for item in PHOTO_ITEMS:
        mark = "■" if item in taken_photos else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "사진 파일 목록",
              _v(data, "photo_file_list") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s13_final_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "13. 확인 서명")
    _write_lv(ws, row, "관리감독자 서명",
              _v(data, "supervisor_name") or "",
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "안전관리자 서명",
              _v(data, "safety_manager_sign") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "최종 확인 서명",
              _v(data, "final_sign") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_hot_work_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 화기작업 허가서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_info(ws, row, data)
    row = _write_s3_work_types(ws, row, data)
    row = _write_s4_combustibles(ws, row, data)
    row = _write_s5_pre_work(ws, row, data)
    row = _write_s6_fire_ext(ws, row, data)
    row = _write_s7_fire_watch(ws, row, data)
    row = _write_s8_ppe(ws, row, data)
    row = _write_s9_approval(ws, row, data)
    row = _write_s10_during_work(ws, row, data)
    row = _write_s11_post_work(ws, row, data)
    row = _write_s12_photos(ws, row, data)
    _write_s13_final_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
