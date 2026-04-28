"""
MSDS 비치 및 교육 확인서 — Excel 출력 모듈 (v1.0)  [PPE-004]

법적 근거 (중심):
    산업안전보건법 제114조 — 물질안전보건자료의 게시 및 교육
    산업안전보건법 제115조 — 물질안전보건자료대상물질 용기 등의 경고표시
    산업안전보건법 시행규칙 제167조 — 물질안전보건자료 게시·비치 방법
    산업안전보건법 시행규칙 제169조 — 물질안전보건자료 교육의 시기·내용·방법
    산업안전보건법 시행규칙 제170조 — 경고표시 방법 및 기재항목

법적 근거 (보조 — 비공개 승인·대체자료 항목 추가 시):
    산업안전보건법 제112조 — 물질안전보건자료의 비공개 승인 등

분류: PRACTICAL — 법정 별지 서식 없음
      MSDS 비치·경고표지 부착·취급 근로자 교육 의무(산안법 제114조·시행규칙 제167조·제169조) 이행 확인용
      실무 법정 연계 보조서식 (관계 기관 공식 법정서식 아님)

Required form_data keys:
    site_name       str  현장명(사업장명)
    check_date      str  확인일자
    checker         str  확인자

Optional form_data keys:
    project_name                str  공사명
    department                  str  소속
    position                    str  직책
    work_location               str  작업 장소
    work_type                   str  작업 내용
    -- 섹션2: 화학물질 기본정보 --
    chemical_name               str  화학물질명(제품명)
    cas_no                      str  CAS 번호
    manufacturer                str  제조사/공급사
    purpose                     str  사용 목적
    daily_amount                str  일일 취급량
    unit                        str  단위
    -- 섹션3: MSDS 비치 확인 --
    msds_available              str  MSDS 비치 여부
    msds_location               str  MSDS 비치 위치
    msds_accessible             str  근로자 접근 가능 여부
    msds_language               str  한국어 기재 여부
    msds_version_current        str  최신본 여부
    msds_remarks                str  비치 관련 비고
    -- 섹션4: 경고표지 부착 확인 --
    label_attached              str  경고표지 부착 여부
    label_location              str  부착 위치
    label_legible               str  표지 판독 가능 여부
    label_ghs_compliant         str  GHS 기준 적합 여부
    label_remarks               str  경고표지 관련 비고
    -- 섹션5: 취급 근로자 교육 확인 --
    edu_conducted               str  교육 실시 여부
    edu_date                    str  교육 일자
    edu_method                  str  교육 방법 (집합/OJT/VR 등)
    edu_duration                str  교육 시간
    edu_instructor              str  교육 강사
    edu_content                 str  교육 내용
    edu_participants            str  교육 대상 인원
    edu_actual                  str  실제 참석 인원
    edu_record_kept             str  교육 기록 보관 여부
    edu_remarks                 str  교육 관련 비고
    -- 섹션6: 보호구 및 응급조치 확인 --
    ppe_specified               str  MSDS 보호구 항목 확인 여부
    ppe_provided                str  지정 보호구 지급 여부
    ppe_types                   str  지급 보호구 종류
    first_aid_known             str  응급조치 요령 숙지 여부
    first_aid_kit_available     str  응급처치 키트 비치 여부
    -- 섹션7: 보관·환기·화기관리 확인 --
    storage_proper              str  전용 보관장소 여부
    storage_labeled             str  보관장소 표시 여부
    incompatible_separated      str  혼재 금지 물질 분리 여부
    ventilation_adequate        str  환기 설비 적절 여부
    open_flame_controlled       str  화기 관리 여부
    -- 섹션8: 누출·화재·응급상황 대응 확인 --
    spill_kit_available         str  누출 대응 키트 비치 여부
    fire_extinguisher_available str  소화기 비치 여부
    emergency_contact_posted    str  비상연락처 게시 여부
    emergency_procedure_known   str  비상대응 절차 숙지 여부
    -- 섹션9: 미흡사항 및 조치계획 --
    deficiency_items            list[dict]  미흡사항 (content, action, responsible, deadline, completed)
    overall_result              str  종합 결과 (적합/조건부 적합/미흡)
    overall_remarks             str  종합 비고
    -- 섹션10: 확인자/관리감독자 서명 --
    preparer_name               str  작성자
    supervisor_name             str  관리감독자
    safety_manager_name         str  안전관리자
    site_manager_name           str  현장소장
    confirm_date                str  서명일자
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "MSDS비치교육확인서"
SHEET_HEADING = "MSDS 비치 및 교육 확인서"
SHEET_SUBTITLE = (
    "「산업안전보건법」 제114조(MSDS 게시 및 교육), 제115조(경고표시), "
    "시행규칙 제167조·제169조·제170조 이행 확인"
)
DOC_ID = "PPE-004"

TOTAL_COLS   = 9
MAX_DEFECT   = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 16, 2: 14, 3: 12, 4: 12,
    5: 14, 6: 12, 7: 12, 8: 12, 9: 10,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_msds_posting_education_check(form_data: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left  = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top   = 0.75
    ws.page_margins.bottom = 0.75

    _apply_col_widths(ws)

    row = 1

    # 제목
    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
                font=_FONT_TITLE, align=_ALIGN_CENTER, height=25)
    row += 1
    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_SUBTITLE,
                font=_FONT_SUBTITLE, align=_ALIGN_CENTER)
    row += 1
    _write_cell(ws, row, 1, TOTAL_COLS, f"[{DOC_ID}]",
                font=_FONT_SMALL, align=_ALIGN_CENTER)
    row += 1
    row += 1  # 공백

    # ── 섹션1: 기본정보 ────────────────────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 1. 현장/공사 기본정보",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "현장명",   _v(form_data, "site_name"),    1, 2, 3)
    _write_lv(ws, row, "공사명",   _v(form_data, "project_name"), 5, 6, 8)
    row += 1
    _write_lv(ws, row, "확인일자", _v(form_data, "check_date"),   1, 2, 3)
    _write_lv(ws, row, "확인자",   _v(form_data, "checker"),      5, 6, 8)
    row += 1
    _write_lv(ws, row, "소속",     _v(form_data, "department"),   1, 2, 3)
    _write_lv(ws, row, "직책",     _v(form_data, "position"),     5, 6, 8)
    row += 1
    _write_lv(ws, row, "작업 장소", _v(form_data, "work_location"), 1, 2, 3)
    _write_lv(ws, row, "작업 내용", _v(form_data, "work_type"),    5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션2: 화학물질/제품 정보 ──────────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 2. 화학물질/제품 정보",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "화학물질명(제품명)", _v(form_data, "chemical_name"), 1, 2, 4)
    _write_lv(ws, row, "CAS 번호",          _v(form_data, "cas_no"),        5, 6, 8)
    row += 1
    _write_lv(ws, row, "제조사/공급사", _v(form_data, "manufacturer"), 1, 2, 4)
    _write_lv(ws, row, "사용 목적",    _v(form_data, "purpose"),      5, 6, 8)
    row += 1
    _write_lv(ws, row, "일일 취급량",  _v(form_data, "daily_amount"), 1, 2, 3)
    _write_lv(ws, row, "단위",         _v(form_data, "unit"),          5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션3: MSDS 비치 위치 확인 ────────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 3. MSDS 비치 위치 확인  (산안법 제114조, 시행규칙 제167조)",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "MSDS 비치 여부",        _v(form_data, "msds_available"),  1, 2, 3)
    _write_lv(ws, row, "비치 위치",              _v(form_data, "msds_location"),   5, 6, 8)
    row += 1
    _write_lv(ws, row, "근로자 접근 가능 여부",  _v(form_data, "msds_accessible"), 1, 2, 3)
    _write_lv(ws, row, "한국어 기재 여부",       _v(form_data, "msds_language"),   5, 6, 8)
    row += 1
    _write_lv(ws, row, "최신본 여부",            _v(form_data, "msds_version_current"), 1, 2, 3)
    _write_lv(ws, row, "비고",                   _v(form_data, "msds_remarks"),    5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션4: 경고표지 부착 확인 ──────────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 4. 경고표지 부착 확인  (산안법 제115조, 시행규칙 제170조)",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "경고표지 부착 여부",   _v(form_data, "label_attached"),    1, 2, 3)
    _write_lv(ws, row, "부착 위치",            _v(form_data, "label_location"),    5, 6, 8)
    row += 1
    _write_lv(ws, row, "표지 판독 가능 여부",  _v(form_data, "label_legible"),     1, 2, 3)
    _write_lv(ws, row, "GHS 기준 적합 여부",   _v(form_data, "label_ghs_compliant"), 5, 6, 8)
    row += 1
    _write_lv(ws, row, "비고", _v(form_data, "label_remarks"), 1, 2, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션5: 취급 근로자 교육 확인 ──────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 5. 취급 근로자 교육 확인  (산안법 제114조, 시행규칙 제169조)",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "교육 실시 여부",  _v(form_data, "edu_conducted"),   1, 2, 3)
    _write_lv(ws, row, "교육 일자",       _v(form_data, "edu_date"),         5, 6, 8)
    row += 1
    _write_lv(ws, row, "교육 방법",       _v(form_data, "edu_method"),       1, 2, 3)
    _write_lv(ws, row, "교육 시간",       _v(form_data, "edu_duration"),     5, 6, 8)
    row += 1
    _write_lv(ws, row, "교육 강사",       _v(form_data, "edu_instructor"),   1, 2, 3)
    _write_lv(ws, row, "교육 내용",       _v(form_data, "edu_content"),      5, 6, 8)
    row += 1
    _write_lv(ws, row, "교육 대상 인원",  _v(form_data, "edu_participants"), 1, 2, 3)
    _write_lv(ws, row, "실제 참석 인원",  _v(form_data, "edu_actual"),       5, 6, 8)
    row += 1
    _write_lv(ws, row, "교육 기록 보관",  _v(form_data, "edu_record_kept"),  1, 2, 3)
    _write_lv(ws, row, "비고",            _v(form_data, "edu_remarks"),      5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션6: 보호구 및 응급조치 확인 ────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 6. 보호구 및 응급조치 확인",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "MSDS 보호구 항목 확인",  _v(form_data, "ppe_specified"),          1, 2, 3)
    _write_lv(ws, row, "지정 보호구 지급 여부",   _v(form_data, "ppe_provided"),           5, 6, 8)
    row += 1
    _write_lv(ws, row, "지급 보호구 종류",        _v(form_data, "ppe_types"),              1, 2, 8)
    row += 1
    _write_lv(ws, row, "응급조치 요령 숙지",      _v(form_data, "first_aid_known"),        1, 2, 3)
    _write_lv(ws, row, "응급처치 키트 비치",      _v(form_data, "first_aid_kit_available"), 5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션7: 보관·환기·화기관리 확인 ───────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 7. 보관·환기·화기관리 확인",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "전용 보관장소 여부",    _v(form_data, "storage_proper"),     1, 2, 3)
    _write_lv(ws, row, "보관장소 표시 여부",    _v(form_data, "storage_labeled"),    5, 6, 8)
    row += 1
    _write_lv(ws, row, "혼재 금지 물질 분리",   _v(form_data, "incompatible_separated"), 1, 2, 3)
    _write_lv(ws, row, "환기 설비 적절 여부",   _v(form_data, "ventilation_adequate"),   5, 6, 8)
    row += 1
    _write_lv(ws, row, "화기 관리 여부",        _v(form_data, "open_flame_controlled"),  1, 2, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션8: 누출·화재·응급상황 대응 확인 ──────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 8. 누출·화재·응급상황 대응 확인",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "누출 대응 키트 비치",   _v(form_data, "spill_kit_available"),       1, 2, 3)
    _write_lv(ws, row, "소화기 비치 여부",      _v(form_data, "fire_extinguisher_available"), 5, 6, 8)
    row += 1
    _write_lv(ws, row, "비상연락처 게시",       _v(form_data, "emergency_contact_posted"),   1, 2, 3)
    _write_lv(ws, row, "비상대응 절차 숙지",    _v(form_data, "emergency_procedure_known"),  5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션9: 미흡사항 및 조치계획 ──────────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 9. 미흡사항 및 조치계획",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_cell(ws, row, 1, 1, "번호",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 2, 4, "미흡사항", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 5, 6, "조치계획", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 7, 7, "담당자",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "완료예정", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 9, 9, "완료",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    deficiency_items = form_data.get("deficiency_items", [])
    for idx in range(1, MAX_DEFECT + 1):
        item = deficiency_items[idx - 1] if idx <= len(deficiency_items) else {}
        _write_cell(ws, row, 1, 1, str(idx),                     align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 4, item.get("content", ""),      align=_ALIGN_LEFT)
        _write_cell(ws, row, 5, 6, item.get("action", ""),       align=_ALIGN_LEFT)
        _write_cell(ws, row, 7, 7, item.get("responsible", ""),  align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, item.get("deadline", ""),     align=_ALIGN_CENTER)
        _write_cell(ws, row, 9, 9, item.get("completed", ""),    align=_ALIGN_CENTER)
        row += 1

    row += 1  # 공백

    _write_lv(ws, row, "종합 결과", _v(form_data, "overall_result"),  1, 2, 4)
    _write_lv(ws, row, "종합 비고", _v(form_data, "overall_remarks"), 5, 6, 8)
    row += 1
    row += 1  # 공백

    # ── 섹션10: 확인자/관리감독자 서명 ───────────────────────────
    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 10. 확인자/관리감독자 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_cell(ws, row, 1, 1, "서명일자",     font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, 2, 3, _v(form_data, "confirm_date"), align=_ALIGN_LEFT)
    _write_cell(ws, row, 5, 5, "작성자",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, 6, 8, _v(form_data, "preparer_name"), align=_ALIGN_LEFT)
    row += 1

    _write_cell(ws, row, 1, 1, "관리감독자",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, 2, 3, _v(form_data, "supervisor_name"), align=_ALIGN_LEFT)
    _write_cell(ws, row, 5, 5, "안전관리자",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, 6, 8, _v(form_data, "safety_manager_name"), align=_ALIGN_LEFT)
    row += 1

    _write_cell(ws, row, 1, 1, "현장소장",     font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, 2, 8, _v(form_data, "site_manager_name"), align=_ALIGN_LEFT)
    row += 1

    # 법적 고지
    row += 1
    _write_cell(
        ws, row, 1, TOTAL_COLS,
        "※ 본 서식은 산업안전보건법 제114조(MSDS 게시 및 교육), 제115조(경고표시), "
        "시행규칙 제167조·제169조·제170조 이행 확인을 위한 실무 보조서식입니다. "
        "관계 기관 공식 법정서식이 아닙니다.",
        font=_FONT_SMALL, align=_ALIGN_LEFT,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
