"""
중량물 인양·중장비사용 작업 허가서 — Excel 출력 모듈 (v1).

법적 근거:
    산업안전보건기준에 관한 규칙 제38조 제1항 제11호 + 별표4 (중량물 취급 작업계획서 의무)
    제39조  (작업지휘자 지정)
    제40조  (신호방법 지정 — 양중기·중량물)
    제132조 (양중기 종류 및 정의)
    제133조 (정격하중 표시 의무)
    제135조 (정격하중 초과 사용 금지)
    제138조 (이동식 크레인 경사각 제한)
    제146조 ④⑤ (인양 중 근로자 탑승 금지, 신호수 배치 — 보이지 않는 하물)
    제163조~제170조 (달기기구 안전기준 — 와이어로프·훅·샤클·달기체인·슬링벨트)
    제221조의5 (굴착기를 이용한 인양작업 특칙)
    제385조  (중량물 취급 시 하역운반기계 사용 의무)
    KOSHA GUIDE C-102-2023 (이동식 크레인 중량물 작업계획서 작성지침 — 참고)
    KOSHA GUIDE M-186-2015 (달기기구·줄걸이 와이어로프 안전작업 지침 — 참고)
    KOSHA GUIDE P-94-2021 (안전작업허가지침 — 참고)

분류: PRACTICAL — 법정 별지 서식 없음.
    법령 의무 항목 기반 자체 표준서식 (KOSHA GUIDE 참고).
    evidence_status: NEEDS_VERIFICATION (조항 원문 API 미수집)

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자
    work_time        str  작업시간
    work_location    str  작업장소
    trade_name       str  작업공종
    work_content     str  작업내용
    contractor       str  작업업체
    work_supervisor  str  작업책임자

Optional form_data keys:
    project_name           str   공사명
    permit_no              str   관리번호
    lifting_object_name    str   인양물 명칭
    lifting_weight         str   인양물 중량 (ton/kg)
    lifting_size           str   인양물 크기 (mm/m)
    lifting_height         str   인양 높이 (m)
    lifting_distance       str   인양 거리 (m)
    lifting_route          str   인양 경로
    equipment_name         str   사용 양중기명
    equipment_rated_load   str   장비 정격하중 (ton)
    work_radius            str   작업반경 (m)
    outrigger_installed    str   아웃트리거 설치 여부
    ground_condition       str   지반 상태
    rigging_method         str   줄걸이 방법
    rigging_gear           str   사용 달기구
    signal_worker_name     str   신호수 성명
    permit_issuer          str   작업허가자
    supervisor_name        str   관리감독자
    safety_manager_sign    str   안전관리자
    work_end_confirmer     str   작업 종료 확인자
    final_sign             str   최종 확인 서명
    during_work_issues     str   작업 중 이상 사항
    work_end_time          str   작업 종료 시각
    photo_file_list        str   사진 파일 목록
    validity_period        str   허가 유효기간
    # list fields
    lifting_types          list[str]  인양작업 유형 선택
    workplan_checks        list[str]  작업계획서 확인 항목
    equipment_checks       list[str]  장비 및 정격하중 확인 항목
    rigging_checks         list[str]  달기구·줄걸이 확인 항목
    signal_checks          list[str]  신호수·통제 확인 항목
    pre_work_checks        list[str]  작업 전 안전조치 항목
    post_work_checks       list[str]  작업 종료 후 확인 항목
    photo_items            list[str]  사진 증빙 항목
    workers                list[dict] name, job_type — 작업자 명단 (max 10)
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "중량물인양중장비허가서"
SHEET_HEADING = "중량물 인양·중장비사용 작업 허가서"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제38조·제40조·제132조~제135조·제138조·"
    "제146조·제163조~제170조에 따른 인양위험 작업 전 안전조치 확인 기록서 "
    "/ KOSHA GUIDE P-94-2021 중장비사용작업허가 참고"
)

MAX_WORKERS = 10
TOTAL_COLS  = 8

# ---------------------------------------------------------------------------
# 고정 문구
# ---------------------------------------------------------------------------

NOTICE_NO_CERT = (
    "본 중량물 인양·중장비사용 작업 허가서는 작업 전 인양위험 확인 및 안전조치 기록이며, "
    "법정 안전보건교육 수료증을 대체하지 않는다."
)
NOTICE_SCOPE = (
    "본 허가서는 중량물 인양 작업 및 크레인·굴착기 등 중장비사용 작업을 포함하며, "
    "KOSHA GUIDE P-94-2021 '중장비사용작업허가' 구조를 참고하여 구현하였다."
)
NOTICE_WP005 = (
    "본 허가서는 WP-005 중량물 취급 작업계획서를 대체하지 않는다."
)
NOTICE_CRANE_PLAN = (
    "이동식크레인 등 양중기 사용 시 해당 작업계획서, 장비사용계획서, "
    "장비점검표를 별도로 확인해야 한다."
)
NOTICE_RATED_LOAD = (
    "정격하중, 작업반경, 줄걸이 상태, 신호수 배치, 하부 출입통제는 "
    "작업 전 현장에서 최종 확인한다."
)
NOTICE_PHOTO = (
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)
NOTICE_PERMIT_ISSUER = (
    "작업허가자는 작업 전 안전조치 이행 여부를 확인한 후 작업을 허가해야 한다."
)
NOTICE_POST_WORK = (
    "작업 종료 후 인양물 설치·고정 상태, 달기구 해체 상태, 장비 원위치를 확인하고 서명한다."
)

# ---------------------------------------------------------------------------
# 인양작업 유형 고정 목록
# ---------------------------------------------------------------------------

LIFTING_TYPES = [
    "이동식 크레인 인양",
    "카고크레인 인양",
    "체인블록 인양",
    "레버블록 인양",
    "지게차 보조 인양",
    "굴착기 인양작업 특칙 해당",
    "인력 운반 병행",
    "기타 중량물 취급",
]

# ---------------------------------------------------------------------------
# 작업계획서 확인 항목 (섹션 3)
# ---------------------------------------------------------------------------

WORKPLAN_CHECK_ITEMS = [
    "WP-005 중량물 취급 작업계획서 확인 (제38조 ① 11호 + 별표4)",
    "이동식크레인 사용 시 관련 작업계획서 확인",
    "장비사용계획서 확인",
    "CL-003 건설장비 일일점검표 확인",
    "인양물 중량 산정 근거 확인",
    "인양 경로 및 작업반경 확인",
    "작업자·신호수 역할 지정 확인 (제39조)",
]

# ---------------------------------------------------------------------------
# 장비 및 정격하중 확인 항목 (섹션 5)
# ---------------------------------------------------------------------------

EQUIPMENT_CHECK_ITEMS = [
    "장비명 및 장비번호 확인",
    "정격하중 표시 확인 (제133조)",
    "작업반경별 허용하중 확인 (제135조)",
    "정격하중 초과 여부 확인 (제135조 — 초과 사용 금지)",
    "과부하방지장치 확인",
    "권과방지장치 확인",
    "훅 해지장치 확인",
    "브레이크 작동 확인",
    "지반 침하 위험 확인 (KOSHA C-102-2023 참고)",
    "아웃트리거 설치 확인 (KOSHA C-102-2023 참고)",
]

# ---------------------------------------------------------------------------
# 달기구·줄걸이 확인 항목 (섹션 6)
# ---------------------------------------------------------------------------

RIGGING_CHECK_ITEMS = [
    "와이어로프 손상 여부 (꼬임끊어짐·소선단선 7% 이상·마모·부식) (제163조·제166조)",
    "슬링벨트 손상 여부 (절단·손상·봉제부 풀림) (제163조·제169조)",
    "샤클 변형·균열 여부 (제168조)",
    "훅 해지장치 상태 (제137조·제168조)",
    "달기체인 손상 여부 (늘어남 5% 초과 금지) (제167조·제168조)",
    "줄걸이 각도 적정 여부 (KOSHA M-186-2015 참고)",
    "인양물 중심 확인",
    "모서리 보호대 사용 여부",
    "달기구 정격하중 확인 (제163조·제164조)",
    "폐기기준 해당 여부 확인",
]

# ---------------------------------------------------------------------------
# 신호수·통제 확인 항목 (섹션 9)
# ---------------------------------------------------------------------------

SIGNAL_CHECK_ITEMS = [
    "신호수 배치 (제40조·제146조 ⑤)",
    "신호 방법 지정 (제40조 — 일원화된 신호방법)",
    "무전기 또는 수신호 확인",
    "작업반경 내 출입통제 (제146조 ④)",
    "하부 근로자 출입금지 확인 (제146조 ④)",
    "인양 경로 하부 통제",
    "유도자 배치",
    "작업 중 신호 일원화 확인",
]

# ---------------------------------------------------------------------------
# 작업 전 안전조치 체크 항목 (섹션 10)
# ---------------------------------------------------------------------------

PRE_WORK_CHECK_ITEMS = [
    "작업 전 TBM 실시",
    "인양물 중량 확인 (제133조·제135조)",
    "인양 경로 장애물 제거",
    "작업반경 출입통제",
    "하부 통제 (제146조 ④)",
    "기상조건 확인",
    "풍속 확인 (강풍 시 작업 중지 기준 확인)",
    "조명 상태 확인",
    "지반 상태 확인 (침하 위험)",
    "작업자 보호구 확인",
    "장비 점검표(CL-003) 확인",
    "작업계획서(WP-005) 확인 (제38조 ① 11호)",
]

# ---------------------------------------------------------------------------
# 작업 종료 후 확인 항목 (섹션 14)
# ---------------------------------------------------------------------------

POST_WORK_CHECK_ITEMS = [
    "인양물 설치 상태 확인",
    "임시 고정 상태 확인",
    "달기구 해체 상태 확인",
    "장비 원위치 및 전원 차단",
    "작업구역 정리정돈",
    "잔여 위험요인 확인",
    "종료 확인자 서명",
]

# ---------------------------------------------------------------------------
# 사진 증빙 항목 (섹션 13)
# ---------------------------------------------------------------------------

PHOTO_ITEMS = [
    "인양 전 작업장소 사진",
    "인양물 상태 사진",
    "줄걸이 체결 사진",
    "장비 및 아웃트리거 사진",
    "작업반경 통제 사진",
    "인양 완료 후 설치상태 사진",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SECTION3 = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=fill or _FILL_NOTICE,
                align=_ALIGN_LEFT, height=24)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    row += 1
    row = _write_notice(ws, row, NOTICE_NO_CERT)
    row = _write_notice(ws, row, NOTICE_SCOPE)
    row = _write_notice(ws, row, NOTICE_WP005)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리번호", _v(data, "permit_no"),   _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 작업 기본정보 및 작업자 명단")
    _write_lv(ws, row, "작업일자", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업시간", _v(data, "work_time"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업장소", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "작업공종", _v(data, "trade_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업업체", _v(data, "contractor"),    _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업내용", _v(data, "work_content"),  _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "작업책임자", _v(data, "work_supervisor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "신호수 성명", _v(data, "signal_worker_name"), _L2, _V2_START, _V2_END)
    row += 1
    # 작업자 명단
    _write_cell(ws, row, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    _write_cell(ws, row, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 2, 5, "성명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 8, "직종/역할", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    workers = data.get("workers") or []
    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, row, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 5, _v(item, "name"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 8, _v(item, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_s3_workplan_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 작업계획서 확인  (작업 전 선행 확인 필수)")
    row = _write_notice(ws, row, NOTICE_CRANE_PLAN)
    checked = set(data.get("workplan_checks") or [])
    for item in WORKPLAN_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s4_lifting_object(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 인양물 정보")
    _write_lv(ws, row, "인양물 명칭", _v(data, "lifting_object_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "인양물 중량", _v(data, "lifting_weight") or "__ ton / __ kg",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "인양물 크기", _v(data, "lifting_size"),   _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "인양 높이",   _v(data, "lifting_height") or "__ m", _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "인양 거리",   _v(data, "lifting_distance") or "__ m", _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "인양 경로",   _v(data, "lifting_route"),  _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s5_equipment(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 양중기 선택 및 정격하중 확인  (제132조~제135조·제138조)")
    row = _write_notice(ws, row, NOTICE_RATED_LOAD)
    _write_lv(ws, row, "사용 양중기",    _v(data, "equipment_name"),        _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "장비 정격하중",  _v(data, "equipment_rated_load") or "__ ton", _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업반경",       _v(data, "work_radius") or "__ m", _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "아웃트리거 설치",
              _v(data, "outrigger_installed") or "□ 설치   □ 미해당",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "지반 상태",      _v(data, "ground_condition"),      _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    # 인양작업 유형
    _write_cell(ws, row, 1, TOTAL_COLS, "인양작업 유형  (해당 항목 선택)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    selected = set(data.get("lifting_types") or [])
    col_span = TOTAL_COLS // 2
    for i in range(0, len(LIFTING_TYPES), 2):
        left  = LIFTING_TYPES[i]
        right = LIFTING_TYPES[i + 1] if i + 1 < len(LIFTING_TYPES) else None
        l_mark = "■" if left in selected else "□"
        _write_cell(ws, row, 1, col_span, f"{l_mark} {left}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        if right:
            r_mark = "■" if right in selected else "□"
            _write_cell(ws, row, col_span + 1, TOTAL_COLS, f"{r_mark} {right}",
                        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        else:
            _write_cell(ws, row, col_span + 1, TOTAL_COLS, "",
                        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        row += 1
    # 장비 점검 항목
    _write_cell(ws, row, 1, TOTAL_COLS, "장비 및 정격하중 확인  (이행 항목에 ■ 표시)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    checked = set(data.get("equipment_checks") or [])
    for item in EQUIPMENT_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s6_rigging(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 달기구·줄걸이 확인  (제163조~제170조)")
    _write_lv(ws, row, "줄걸이 방법", _v(data, "rigging_method"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "사용 달기구", _v(data, "rigging_gear"),   _L2, _V2_START, _V2_END)
    row += 1
    checked = set(data.get("rigging_checks") or [])
    for item in RIGGING_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s7_route(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 인양 경로 및 위험구간 확인")
    _write_lv(ws, row, "인양 경로",
              _v(data, "lifting_route") or "출발지 → 경유지 → 설치 위치",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    _write_lv(ws, row, "경로 장애물",
              "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    _write_lv(ws, row, "위험구간 특기사항",
              "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s8_ground(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 작업장소·지반·아웃트리거 확인  (KOSHA C-102-2023 참고)")
    _write_lv(ws, row, "작업장소 지반 상태",
              _v(data, "ground_condition") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "지반 침하 위험 여부",
              "□ 있음 — 지반 보강 조치 필요   □ 없음",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "아웃트리거 설치 여부",
              _v(data, "outrigger_installed") or "□ 설치 완료   □ 미해당",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "지반 조치 내용",
              "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s9_signal(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 신호수·통신·작업반경 통제  (제40조·제146조 ④⑤)")
    _write_lv(ws, row, "신호수 성명",
              _v(data, "signal_worker_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업반경",
              _v(data, "work_radius") or "__ m",
              _L2, _V2_START, _V2_END)
    row += 1
    checked = set(data.get("signal_checks") or [])
    for item in SIGNAL_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s10_pre_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 작업 전 안전조치  (이행 항목에 ■ 표시)")
    checked = set(data.get("pre_work_checks") or [])
    for item in PRE_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s11_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 작업허가 승인", fill=_FILL_SECTION2)
    row = _write_notice(ws, row, NOTICE_PERMIT_ISSUER)
    _write_lv(ws, row, "작업신청자 (서명)", _v(data, "work_supervisor"),
              _L1, _V1_START, _V1_END, height=32)
    _write_lv(ws, row, "관리감독자 (서명)", _v(data, "supervisor_name"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "안전관리자 (서명)", _v(data, "safety_manager_sign"),
              _L1, _V1_START, _V1_END, height=32)
    _write_lv(ws, row, "작업허가자 (서명)", _v(data, "permit_issuer"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "허가 유효기간",
              _v(data, "validity_period") or "당일 1회 작업 한정 (KOSHA P-94-2021 참고)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    return row


def _write_s12_during_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 작업 중 점검")
    _write_cell(ws, row, 1, TOTAL_COLS,
                "이상 발생(달기구 탈락·장비 이상·기상 악화·지반 침하 징후 등) 시 즉시 작업 중단 후 허가자에게 보고",
                font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=24)
    row += 1
    _write_lv(ws, row, "작업 중 이상 사항",
              _v(data, "during_work_issues") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    return row


def _write_s13_photos(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "13. 사진 및 증빙자료  [OPTIONAL]")
    row = _write_notice(ws, row, NOTICE_PHOTO)
    taken_photos = set(data.get("photo_items") or [])
    for item in PHOTO_ITEMS:
        mark = "■" if item in taken_photos else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "사진 파일 목록",
              _v(data, "photo_file_list") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s14_final_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "14. 작업 종료 및 확인 서명")
    row = _write_notice(ws, row, NOTICE_POST_WORK)
    checked = set(data.get("post_work_checks") or [])
    for item in POST_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "작업 종료 시각",
              _v(data, "work_end_time") or "",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "종료 확인자 서명",
              _v(data, "work_end_confirmer") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리감독자 서명",
              _v(data, "supervisor_name") or "",
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "안전관리자 서명",
              _v(data, "safety_manager_sign") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "최종 확인 서명",
              _v(data, "final_sign") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_lifting_work_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 중량물 인양·중장비사용 작업 허가서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_info(ws, row, data)
    row = _write_s3_workplan_check(ws, row, data)
    row = _write_s4_lifting_object(ws, row, data)
    row = _write_s5_equipment(ws, row, data)
    row = _write_s6_rigging(ws, row, data)
    row = _write_s7_route(ws, row, data)
    row = _write_s8_ground(ws, row, data)
    row = _write_s9_signal(ws, row, data)
    row = _write_s10_pre_work(ws, row, data)
    row = _write_s11_approval(ws, row, data)
    row = _write_s12_during_work(ws, row, data)
    row = _write_s13_photos(ws, row, data)
    _write_s14_final_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
