"""
안전관리 일지 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건법 제38조 (안전조치)
    산업안전보건법 제36조 (위험성평가 실시 및 조치사항 기록)
    산업안전보건기준에 관한 규칙 별표 3 (작업시작 전 점검사항)
분류: PRACTICAL — 공식 제출 서식 아님. 현장 일일 안전관리 및 점검 기록 보조서식.

요약:
    현장 관리자가 하루 단위로 안전관리 현황을 종합 기록하는 보조서식입니다.
    산업안전보건법상 안전조치 및 위험성평가 조치사항 관리와 연계합니다.
    작업시작 전 점검사항 및 TBM 결과를 일일 기록으로 관리합니다.
    사고·아차사고·응급조치 발생 시 관련 EM 서식과 별도 연계합니다.
    위험요인 발견 시 담당자·기한·이행상태까지 관리합니다.
    개인정보·민감정보 최소 기재. 사진·영상 등 증빙자료는 별도 보관.

Required form_data keys:
    site_name           str  현장명
    log_date            str  일지 작성 일자
    writer_name         str  작성자 성명
    weather             str  날씨
    work_summary        str  당일 주요 작업 요약

Optional form_data keys:
    project_name                str  공사명
    log_day_of_week             str  요일
    temperature                 str  기온
    work_start_time             str  작업 시작 시간
    work_end_time               str  작업 종료 시간
    reviewer_name               str  검토자
    approver_name               str  승인자
    work_type                   str  작업공종
    work_zone                   str  작업구역
    worker_count                str  투입 인원
    subcontractor_name          str  협력업체
    main_equipment              str  주요 장비
    high_risk_work              str  고위험 작업 여부
    night_work                  str  야간작업 여부
    work_remarks                str  작업 특이사항
    tbm_done                    str  TBM 실시 여부
    risk_assessment_shared      str  위험성평가 공유 여부
    work_plan_checked           str  작업계획서 확인 여부
    permit_required             str  작업허가서 필요 여부
    permit_issued               str  작업허가서 발급 여부
    ppe_check_done              str  보호구 착용 확인
    passage_opening_checked     str  통로/개구부 확인
    fall_protection_checked     str  추락방지 조치 확인
    falling_object_checked      str  낙하물 방지 조치 확인
    electrical_fire_checked     str  전기/화기 위험 확인
    equipment_check_done        str  장비 사전점검 확인
    emergency_contact_checked   str  비상연락망 확인
    hazard_items                list[dict]  위험요인 반복행 (max 10)
        found_time              str  발견 시간
        found_location          str  발견 장소
        hazard_description      str  위험요인
        risk_level              str  위험등급
        immediate_action        str  즉시 조치 내용
        responsible_person      str  담당자
        due_date                str  완료 예정일
        completed               str  완료 여부
        confirmed_by            str  확인자
        evidence                str  사진/증빙자료 여부
    patrol_items                list[dict]  안전순찰 반복행 (max 5)
        patrol_time             str  순찰 시간
        patrol_person           str  순찰자
        patrol_zone             str  순찰 구역
        findings                str  지적사항
        action_taken            str  조치사항
        reinspected             str  재점검 여부
        result                  str  확인 결과
    tbm_time                    str  TBM 실시 시간
    education_topic             str  교육 주제
    attendees_count             str  참석 인원
    absent_count                str  불참 인원
    education_note              str  전달사항
    foreign_worker_notified     str  외국인 근로자 전달 여부
    new_worker_educated         str  신규 근로자 교육 여부
    sign_sheet_attached         str  서명부 별첨 여부
    equipment_status            str  주요 장비 상태
    equipment_log_prepared      str  장비 점검표 작성 여부
    incoming_equipment_ok       str  반입 장비 이상 여부
    material_storage_ok         str  자재 적치 상태
    ppe_supply_ok               str  보호구 지급 상태
    ppe_wearing_ok              str  보호구 착용 상태
    defective_ppe_replaced      str  불량 보호구 교체 여부
    accident_occurred           str  사고 발생 여부
    near_miss_occurred          str  아차사고 발생 여부
    first_aid_occurred          str  응급조치 발생 여부
    em002_linked                str  EM-002 연계 여부
    em006_linked                str  EM-006 연계 여부
    serious_accident_report     str  중대재해 즉시보고 필요 여부
    incident_followup           str  후속 조치
    next_day_work               str  익일 예정 작업
    next_day_hazard             str  예상 위험요인
    preparation_needed          str  사전 준비사항
    subcontractor_note          str  협력업체 전달사항
    pending_actions             str  미완료 조치
    handover_person             str  인계자
    receiver_person             str  인수자
    safety_manager_name         str  안전관리자 성명
    supervisor_name             str  관리감독자 성명
    site_manager_name           str  현장소장 성명
    contractor_rep_name         str  협력업체 책임자 성명
    confirm_date                str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "안전관리일지"
SHEET_HEADING = "안전관리 일지"
SHEET_SUBTITLE = "공식 제출 서식 아님 | 현장 일일 안전관리 및 점검 기록 보조서식"
DOC_ID = "DL-001"

TOTAL_COLS = 10
MAX_HAZARD_ROWS = 5
MAX_PATROL_ROWS = 3

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL = Font(name="맑은 고딕", size=9, italic=True)
_FONT_NOTE = Font(name="맑은 고딕", size=8, italic=True, color="595959")

_FILL_LABEL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN = PatternFill(fill_type="solid", fgColor="FFF2CC")

_THIN = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_COL_WIDTHS: Dict[int, int] = {
    1: 12, 2: 10, 3: 14, 4: 12, 5: 10, 6: 14, 7: 10, 8: 12, 9: 12, 10: 10,
}

# layout: 5 label+value pairs per row (cols 1-2 / 3-4 / 5-6 / 7-8 / 9-10)
_PAIRS = [(1, 2, 3, 4), (5, 6, 7, 8)]  # (label_col, label_end, val_col, val_end)


def _v(data: Any, key: str) -> str:
    val = data.get(key, "") if isinstance(data, dict) else ""
    return str(val) if val is not None else ""


def _border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _BORDER


def _cell(ws, row: int, c1: int, c2: int, value: Any = "", *,
          font=None, fill=None, align=None, height: int = 20) -> None:
    if c1 != c2:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=value)
    cell.font = font or _FONT_DEFAULT
    cell.fill = fill or PatternFill()
    cell.alignment = align or _ALIGN_LEFT
    cell.border = _BORDER
    ws.row_dimensions[row].height = height


def _lv(ws, row: int, lc1: int, lc2: int, vc1: int, vc2: int,
        label: str, value: str, height: int = 20) -> None:
    ws.merge_cells(start_row=row, start_column=lc1, end_row=row, end_column=lc2)
    lbl = ws.cell(row=row, column=lc1, value=label)
    lbl.font = _FONT_BOLD
    lbl.fill = _FILL_LABEL
    lbl.alignment = _ALIGN_CENTER
    lbl.border = _BORDER

    ws.merge_cells(start_row=row, start_column=vc1, end_row=row, end_column=vc2)
    val = ws.cell(row=row, column=vc1, value=value)
    val.font = _FONT_DEFAULT
    val.alignment = _ALIGN_LEFT
    val.border = _BORDER
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, title,
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_title(ws, row: int) -> int:
    _cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
          font=_FONT_TITLE, align=_ALIGN_CENTER, height=28)
    row += 1
    _cell(ws, row, 1, TOTAL_COLS, SHEET_SUBTITLE,
          font=_FONT_SUBTITLE, align=_ALIGN_CENTER, height=15)
    row += 1
    _cell(ws, row, 1, TOTAL_COLS,
          "산업안전보건법상 안전조치 및 위험성평가 조치사항 관리와 연계 | "
          "작업시작 전 점검사항 및 TBM 결과를 일일 기록으로 관리",
          font=_FONT_NOTE, align=_ALIGN_CENTER, height=13)
    row += 1
    _cell(ws, row, 1, TOTAL_COLS,
          "사고·아차사고·응급조치 발생 시 관련 EM 서식과 별도 연계 | "
          "위험요인 발견 시 담당자·기한·이행상태까지 관리 | "
          "개인정보·민감정보 최소 기재 | 사진·영상 등 증빙자료는 별도 보관",
          font=_FONT_NOTE, align=_ALIGN_CENTER, height=13)
    row += 1
    _cell(ws, row, 1, TOTAL_COLS, f"문서 ID: {DOC_ID}",
          font=_FONT_SMALL, align=_ALIGN_CENTER, height=12)
    return row + 1


def _write_basic_info(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "1. 문서 기본정보")
    _lv(ws, row, 1, 1, 2, 4, "공사명", _v(data, "project_name"))
    _lv(ws, row, 5, 5, 6, 7, "현장명", _v(data, "site_name"))
    ws.cell(row=row, column=8, value="").border = _BORDER
    ws.cell(row=row, column=9, value="").border = _BORDER
    ws.cell(row=row, column=10, value="").border = _BORDER
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=10)
    row += 1
    _lv(ws, row, 1, 1, 2, 2, "작성일", _v(data, "log_date"))
    _lv(ws, row, 3, 3, 4, 4, "요일", _v(data, "log_day_of_week"))
    _lv(ws, row, 5, 5, 6, 6, "날씨", _v(data, "weather"))
    _lv(ws, row, 7, 7, 8, 8, "기온", _v(data, "temperature"))
    _lv(ws, row, 9, 9, 10, 10, "작업시간",
        f"{_v(data,'work_start_time')}~{_v(data,'work_end_time')}")
    row += 1
    _lv(ws, row, 1, 1, 2, 3, "작성자", _v(data, "writer_name"))
    _lv(ws, row, 4, 4, 5, 6, "검토자", _v(data, "reviewer_name"))
    _lv(ws, row, 7, 7, 8, 9, "승인자", _v(data, "approver_name"))
    ws.cell(row=row, column=10, value="").border = _BORDER
    row += 1
    return row


def _write_work_overview(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "2. 현장 작업 개요")
    _lv(ws, row, 1, 1, 2, 4, "금일 주요 작업", _v(data, "work_summary"))
    _lv(ws, row, 5, 5, 6, 7, "작업공종", _v(data, "work_type"))
    ws.cell(row=row, column=8, value="").border = _BORDER
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=10)
    row += 1
    _lv(ws, row, 1, 1, 2, 3, "작업구역", _v(data, "work_zone"))
    _lv(ws, row, 4, 4, 5, 5, "투입 인원", _v(data, "worker_count"))
    _lv(ws, row, 6, 6, 7, 8, "협력업체", _v(data, "subcontractor_name"))
    _lv(ws, row, 9, 9, 10, 10, "주요 장비", _v(data, "main_equipment"))
    row += 1
    _lv(ws, row, 1, 1, 2, 3, "고위험 작업", _v(data, "high_risk_work"))
    _lv(ws, row, 4, 4, 5, 5, "야간작업", _v(data, "night_work"))
    _lv(ws, row, 6, 6, 7, 10, "특이사항", _v(data, "work_remarks"))
    row += 1
    return row


def _write_precheck(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "3. 작업 전 안전점검")
    checks = [
        ("TBM 실시", "tbm_done"), ("위험성평가 공유", "risk_assessment_shared"),
        ("작업계획서 확인", "work_plan_checked"), ("작업허가서 필요", "permit_required"),
        ("작업허가서 발급", "permit_issued"), ("보호구 착용 확인", "ppe_check_done"),
        ("통로/개구부 확인", "passage_opening_checked"), ("추락방지 조치", "fall_protection_checked"),
        ("낙하물 방지 조치", "falling_object_checked"), ("전기/화기 위험 확인", "electrical_fire_checked"),
        ("장비 사전점검", "equipment_check_done"), ("비상연락망 확인", "emergency_contact_checked"),
    ]
    pairs = [(checks[i], checks[i + 1]) for i in range(0, len(checks) - 1, 2)]
    for left, right in pairs:
        _lv(ws, row, 1, 2, 3, 5, left[0], _v(data, left[1]))
        _lv(ws, row, 6, 7, 8, 10, right[0], _v(data, right[1]))
        row += 1
    return row


def _write_hazard_table(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "4. 위험요인 및 조치사항")
    headers = ["발견시간", "발견장소", "위험요인", "위험등급", "즉시조치", "담당자", "완료예정일", "완료여부", "확인자", "증빙"]
    for i, h in enumerate(headers, 1):
        ws.merge_cells(start_row=row, start_column=i, end_row=row, end_column=i)
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
    ws.row_dimensions[row].height = 18
    row += 1
    items = data.get("hazard_items", [])
    if not isinstance(items, list):
        items = []
    fields = ["found_time", "found_location", "hazard_description", "risk_level",
              "immediate_action", "responsible_person", "due_date",
              "completed", "confirmed_by", "evidence"]
    for i in range(MAX_HAZARD_ROWS):
        item = items[i] if i < len(items) else {}
        for c, fld in enumerate(fields, 1):
            cell = ws.cell(row=row, column=c, value=_v(item, fld))
            cell.font = _FONT_DEFAULT
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_patrol_table(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "5. 일일 안전순찰 기록")
    headers = ["순찰시간", "순찰자", "순찰구역", "지적사항", "조치사항", "재점검여부", "확인결과"]
    col_spans = [(1, 1), (2, 2), (3, 4), (5, 6), (7, 8), (9, 9), (10, 10)]
    for (c1, c2), h in zip(col_spans, headers):
        if c1 != c2:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=h)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
        if c1 != c2:
            for c in range(c1 + 1, c2 + 1):
                ws.cell(row=row, column=c).border = _BORDER
    ws.row_dimensions[row].height = 18
    row += 1
    items = data.get("patrol_items", [])
    if not isinstance(items, list):
        items = []
    fields = ["patrol_time", "patrol_person", "patrol_zone", "findings",
              "action_taken", "reinspected", "result"]
    for i in range(MAX_PATROL_ROWS):
        item = items[i] if i < len(items) else {}
        for (c1, c2), fld in zip(col_spans, fields):
            if c1 != c2:
                ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = ws.cell(row=row, column=c1, value=_v(item, fld))
            cell.font = _FONT_DEFAULT
            cell.alignment = _ALIGN_LEFT
            cell.border = _BORDER
            if c1 != c2:
                for c in range(c1 + 1, c2 + 1):
                    ws.cell(row=row, column=c).border = _BORDER
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_education(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "6. 교육 및 회의 기록")
    _lv(ws, row, 1, 1, 2, 3, "TBM 실시 시간", _v(data, "tbm_time"))
    _lv(ws, row, 4, 4, 5, 7, "교육 주제", _v(data, "education_topic"))
    _lv(ws, row, 8, 8, 9, 10, "참석 인원", _v(data, "attendees_count"))
    row += 1
    _lv(ws, row, 1, 1, 2, 3, "불참 인원", _v(data, "absent_count"))
    _lv(ws, row, 4, 4, 5, 10, "전달사항", _v(data, "education_note"))
    row += 1
    _lv(ws, row, 1, 2, 3, 4, "외국인 근로자 전달", _v(data, "foreign_worker_notified"))
    _lv(ws, row, 5, 6, 7, 8, "신규 근로자 교육", _v(data, "new_worker_educated"))
    _lv(ws, row, 9, 9, 10, 10, "서명부 별첨", _v(data, "sign_sheet_attached"))
    row += 1
    return row


def _write_equipment_material(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "7. 장비·자재·보호구 확인")
    _lv(ws, row, 1, 2, 3, 5, "주요 장비 상태", _v(data, "equipment_status"))
    _lv(ws, row, 6, 7, 8, 10, "장비 점검표 작성", _v(data, "equipment_log_prepared"))
    row += 1
    _lv(ws, row, 1, 2, 3, 5, "반입 장비 이상", _v(data, "incoming_equipment_ok"))
    _lv(ws, row, 6, 7, 8, 10, "자재 적치 상태", _v(data, "material_storage_ok"))
    row += 1
    _lv(ws, row, 1, 2, 3, 5, "보호구 지급 상태", _v(data, "ppe_supply_ok"))
    _lv(ws, row, 6, 7, 8, 10, "보호구 착용 상태", _v(data, "ppe_wearing_ok"))
    row += 1
    _lv(ws, row, 1, 2, 3, 5, "불량 보호구 교체", _v(data, "defective_ppe_replaced"))
    for c in range(6, TOTAL_COLS + 1):
        ws.cell(row=row, column=c).border = _BORDER
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=TOTAL_COLS)
    row += 1
    return row


def _write_incident(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "8. 사고·아차사고·응급조치 기록")
    _lv(ws, row, 1, 1, 2, 3, "사고 발생", _v(data, "accident_occurred"))
    _lv(ws, row, 4, 4, 5, 6, "아차사고 발생", _v(data, "near_miss_occurred"))
    _lv(ws, row, 7, 7, 8, 10, "응급조치 발생", _v(data, "first_aid_occurred"))
    row += 1
    _lv(ws, row, 1, 1, 2, 3, "EM-002 연계", _v(data, "em002_linked"))
    _lv(ws, row, 4, 4, 5, 6, "EM-006 연계", _v(data, "em006_linked"))
    _lv(ws, row, 7, 7, 8, 10, "중대재해 즉시보고", _v(data, "serious_accident_report"))
    row += 1
    _lv(ws, row, 1, 2, 3, 10, "후속 조치", _v(data, "incident_followup"), height=30)
    row += 1
    return row


def _write_handover(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "9. 내일 작업 및 인계사항")
    _lv(ws, row, 1, 2, 3, 10, "익일 예정 작업", _v(data, "next_day_work"))
    row += 1
    _lv(ws, row, 1, 2, 3, 10, "예상 위험요인", _v(data, "next_day_hazard"))
    row += 1
    _lv(ws, row, 1, 2, 3, 10, "사전 준비사항", _v(data, "preparation_needed"))
    row += 1
    _lv(ws, row, 1, 2, 3, 10, "협력업체 전달사항", _v(data, "subcontractor_note"))
    row += 1
    _lv(ws, row, 1, 2, 3, 10, "미완료 조치", _v(data, "pending_actions"))
    row += 1
    _lv(ws, row, 1, 1, 2, 4, "인계자", _v(data, "handover_person"))
    _lv(ws, row, 5, 5, 6, 10, "인수자", _v(data, "receiver_person"))
    row += 1
    return row


def _write_signatures(ws, row: int, data: Dict) -> int:
    row = _section(ws, row, "10. 확인 및 승인")
    sigs = [
        ("작성자", "writer_name"), ("안전관리자", "safety_manager_name"),
        ("관리감독자", "supervisor_name"), ("현장소장", "site_manager_name"),
        ("협력업체 책임자", "contractor_rep_name"),
    ]
    # 2 per row, last one spans full
    pairs = [(sigs[i], sigs[i + 1]) for i in range(0, len(sigs) - 1, 2)]
    for left, right in pairs:
        _lv(ws, row, 1, 2, 3, 5, left[0], _v(data, left[1]))
        _lv(ws, row, 6, 7, 8, 10, right[0], _v(data, right[1]))
        row += 1
    # last entry (협력업체 책임자) spans half width
    _lv(ws, row, 1, 2, 3, 5, sigs[-1][0], _v(data, sigs[-1][1]))
    _lv(ws, row, 6, 7, 8, 10, "확인일", _v(data, "confirm_date"))
    row += 1
    return row


def build_safety_management_log(form_data: Dict[str, Any]) -> bytes:
    """안전관리 일지 Excel 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    r = _write_title(ws, 1)
    r += 1
    r = _write_basic_info(ws, r, form_data)
    r += 1
    r = _write_work_overview(ws, r, form_data)
    r += 1
    r = _write_precheck(ws, r, form_data)
    r += 1
    r = _write_hazard_table(ws, r, form_data)
    r += 1
    r = _write_patrol_table(ws, r, form_data)
    r += 1
    r = _write_education(ws, r, form_data)
    r += 1
    r = _write_equipment_material(ws, r, form_data)
    r += 1
    r = _write_incident(ws, r, form_data)
    r += 1
    r = _write_handover(ws, r, form_data)
    r += 1
    r = _write_signatures(ws, r, form_data)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
