"""
고소작업 허가서 — Excel 출력 모듈 (v1).

법적 근거:
    산업안전보건기준에 관한 규칙 제37조 (악천후 시 작업 중지)
    제42조 (추락의 방지 — 작업발판·추락방호망·안전대 단계적 의무)
    제43조 (개구부 등의 방호조치)
    제44조 (안전대의 부착설비 등)
    제45조 (지붕 위에서의 위험 방지)
    제57조 이하 (비계 조립·해체·변경 — 관리감독자 배치 의무)
    제86조 이하 (고소작업대 설치 등의 조치)
    KOSHA GUIDE P-94-2021 안전작업허가지침 (참고)
    KOSHA GUIDE C-74-2015 건설공사의 고소작업대 안전보건작업지침 (참고)
    KOSHA GUIDE M-155-2023 이동식 고소작업대의 선정과 안전관리 (참고)

분류: PRACTICAL — 법정 별지 서식 없음.
    법령 제42조~제44조 안전조치 이행 확인 양식 + KOSHA P-94-2021 참고 자체 표준 서식.

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자
    work_time        str  작업시간
    work_location    str  작업장소
    trade_name       str  작업공종
    work_content     str  작업내용
    contractor       str  작업업체
    work_supervisor  str  작업책임자

Optional form_data keys:
    project_name         str   공사명
    permit_no            str   관리번호
    work_height          str   작업높이 (m)
    equipment_list       str   사용 설비
    equipment_type       str   사용 장비
    fall_risk_present    str   추락위험 존재 여부
    opening_present      str   개구부 존재 여부
    workboard_installed  str   작업발판 설치 여부
    railing_installed    str   안전난간 설치 여부
    lanyard_worn         str   안전대 착용 여부
    anchor_confirmed     str   안전대 부착설비 확인 여부
    falling_zone_set     str   낙하물 위험구역 설정 여부
    access_control       str   출입통제 여부
    weather_confirmed    str   기상조건 확인 여부
    permit_issuer        str   작업허가자
    supervisor_name      str   관리감독자
    safety_manager_sign  str   안전관리자
    work_end_confirmer   str   작업 종료 확인자
    final_sign           str   최종 확인 서명
    during_work_issues   str   작업 중 이상 사항
    work_end_time        str   작업 종료 시각
    photo_file_list      str   사진 파일 목록
    validity_period      str   허가 유효기간
    # list fields
    work_types           list[str]  고소작업 유형 선택
    pre_work_checks      list[str]  작업 전 안전조치 이행 항목
    workboard_checks     list[str]  작업발판·비계·사다리 확인 항목
    aerial_checks        list[str]  고소작업대 확인 항목
    harness_checks       list[str]  안전대·추락방지설비 확인 항목
    falling_checks       list[str]  낙하물 방지 확인 항목
    post_work_checks     list[str]  작업 종료 후 확인 항목
    photo_items          list[str]  사진 증빙 항목
    workers              list[dict] name, job_type — 작업자 명단 (max 10)
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "고소작업허가서"
SHEET_HEADING = "고소작업 허가서"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제42조~제44조에 따른 "
    "추락위험 작업 전 안전조치 확인 기록서"
)

MAX_WORKERS = 10
TOTAL_COLS  = 8

# ---------------------------------------------------------------------------
# 고정 문구
# ---------------------------------------------------------------------------

NOTICE_NO_CERT = (
    "본 고소작업 허가서는 작업 전 추락위험 확인 및 안전조치 기록이며, "
    "법정 안전보건교육 수료증을 대체하지 않는다."
)
NOTICE_SCAFFOLD = (
    "비계 작업은 CL-001 비계 설치 점검표와 병행하여 확인한다."
)
NOTICE_AERIAL = (
    "고소작업대 사용 시 장비 점검표 및 장비 상태를 별도로 확인한다."
)
NOTICE_HARNESS = (
    "안전대 착용 시 부착설비 상태를 함께 확인해야 한다."
)
NOTICE_PHOTO = (
    "사진은 법정 필수 고정항목이 아니라 점검 대응을 위한 권장 증빙으로 관리한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)
NOTICE_PERMIT_ISSUER = (
    "작업허가자는 작업 전 안전조치 이행 여부를 확인한 후 작업을 허가해야 한다."
)
NOTICE_POST_WORK = (
    "작업 종료 후 작업발판·사다리·장비를 정리하고 개구부 방호 원상복구 여부를 확인한다."
)

# ---------------------------------------------------------------------------
# 고소작업 유형 고정 목록
# ---------------------------------------------------------------------------

WORK_AT_HEIGHT_TYPES = [
    "비계 작업",
    "이동식 비계 작업",
    "작업발판 작업",
    "사다리 작업",
    "고소작업대 작업",
    "지붕 위 작업",
    "개구부 주변 작업",
    "샤프트/피트 주변 작업",
    "천장 내 작업",
    "기타 추락위험 작업",
]

# ---------------------------------------------------------------------------
# 작업 전 안전조치 체크 항목
# ---------------------------------------------------------------------------

PRE_WORK_CHECK_ITEMS = [
    "작업구역 추락위험 확인 (제42조)",
    "작업발판 설치 상태 확인 (제42조)",
    "안전난간 설치 상태 확인 (제43조)",
    "개구부 덮개 또는 방호조치 확인 (제43조)",
    "안전대 착용 확인 (제44조)",
    "안전대 부착설비 확인 (제44조)",
    "낙하물 방지조치 확인",
    "작업구역 출입통제 확인",
    "기상조건 확인 (제37조 — 풍속 10m/s 이상 시 작업 중지 기준)",
    "조명 상태 확인",
    "작업자 건강상태 확인",
    "TBM 실시 여부 확인",
]

# ---------------------------------------------------------------------------
# 작업발판·비계·사다리 확인 항목
# ---------------------------------------------------------------------------

WORKBOARD_CHECK_ITEMS = [
    "작업발판 폭·고정 상태 확인",
    "작업발판 미끄럼 방지 조치",
    "안전난간·중간난간·발끝막이판 설치 확인",
    "비계 사용 시 CL-001 비계 점검표 병행 확인 (제57조 이하)",
    "사다리 전도방지 조치 확인",
    "사다리 상부 고정 또는 보조자 배치 확인",
    "사다리 최상부 작업 금지 확인",
    "이동식 비계 바퀴 잠금 확인",
    "작업발판 적재하중 초과 여부 확인",
]

# ---------------------------------------------------------------------------
# 고소작업대 확인 항목
# ---------------------------------------------------------------------------

AERIAL_WORK_PLATFORM_CHECK_ITEMS = [
    "장비 점검표(CL-003) 확인 (제86조 이하)",
    "아웃트리거 설치 상태 확인",
    "작업대 난간 상태 확인",
    "비상정지장치 작동 확인",
    "비상하강장치 작동 확인",
    "과상승 방지장치 확인",
    "작업대 정격하중 확인",
    "작업대 내 안전대 착용 확인",
    "지반 침하 위험 확인",
    "장비 운전자 자격 확인",
]

# ---------------------------------------------------------------------------
# 안전대·추락방지설비 확인 항목
# ---------------------------------------------------------------------------

HARNESS_CHECK_ITEMS = [
    "안전대 착용 상태 확인 (제44조)",
    "안전대 훅 체결 상태 확인",
    "부착설비 강도 및 위치 확인 (제44조)",
    "수직/수평 구명줄 상태 확인",
    "추락방호망 설치 여부 확인 (제42조)",
    "안전블록 사용 여부 확인",
    "지붕 위 작업 시 미끄럼 방지 조치 (제45조)",
    "샤프트/피트 주변 방호조치 확인 (제43조)",
]

# ---------------------------------------------------------------------------
# 낙하물 방지 확인 항목
# ---------------------------------------------------------------------------

FALLING_CHECK_ITEMS = [
    "낙하물 방지망 설치 여부 확인",
    "낙하물 위험구역 설정 및 표지판 설치",
    "작업구역 하부 출입통제 확인",
    "낙하물 방지 덮개 설치 확인",
    "공구·자재 낙하 방지 조치 확인",
]

# ---------------------------------------------------------------------------
# 작업 종료 후 확인 항목
# ---------------------------------------------------------------------------

POST_WORK_CHECK_ITEMS = [
    "작업발판·사다리·장비 정리 확인",
    "개구부 방호 원상복구 확인",
    "낙하물 잔여 여부 확인",
    "안전난간 임시해체부 복구 확인",
    "작업구역 정리정돈",
    "이상 발견 시 보고 확인",
    "종료 확인자 서명",
]

# ---------------------------------------------------------------------------
# 사진 증빙 항목
# ---------------------------------------------------------------------------

PHOTO_ITEMS = [
    "작업 전 작업장소 사진",
    "작업발판/안전난간 사진",
    "안전대 착용 사진",
    "고소작업대 설치 사진",
    "개구부 방호 사진",
    "작업 종료 후 정리상태 사진",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=fill or _FILL_NOTICE,
                align=_ALIGN_LEFT, height=24)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    row += 1
    row = _write_notice(ws, row, NOTICE_NO_CERT)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리번호", _v(data, "permit_no"),   _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 작업 기본정보")
    _write_lv(ws, row, "작업일자", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업시간", _v(data, "work_time"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업장소", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    _write_lv(ws, row, "작업공종", _v(data, "trade_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업업체", _v(data, "contractor"),    _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업내용", _v(data, "work_content"),  _L1, _FULL_VAL_START, _FULL_VAL_END, height=40)
    row += 1
    _write_lv(ws, row, "작업책임자", _v(data, "work_supervisor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업높이",   _v(data, "work_height") or "__ m (최대 작업 높이)",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "사용 설비", _v(data, "equipment_list"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "사용 장비", _v(data, "equipment_type"), _L2, _V2_START, _V2_END)
    row += 1
    # 작업자 명단
    _write_cell(ws, row, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    _write_cell(ws, row, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 2, 5, "성명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 8, "직종/역할", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    workers = data.get("workers") or []
    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, row, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 5, _v(item, "name"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 8, _v(item, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_s3_work_types(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 고소작업 유형  (해당 항목 선택)")
    selected = set(data.get("work_types") or [])
    col_count = 2
    col_span  = TOTAL_COLS // col_count
    for item in WORK_AT_HEIGHT_TYPES:
        mark = "■" if item in selected else "□"
        _write_cell(ws, row, 1, col_span, f"{mark} {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        _write_cell(ws, row, col_span + 1, TOTAL_COLS, "",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        row += 1
    return row


def _write_s4_fall_risk(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 작업장소 및 추락위험 확인  (제42조~제43조)")
    _write_lv(ws, row, "추락위험 존재 여부",
              _v(data, "fall_risk_present") or "□ 있음   □ 없음",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "개구부 존재 여부",
              _v(data, "opening_present") or "□ 있음   □ 없음",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업발판 설치 여부",
              _v(data, "workboard_installed") or "□ 설치   □ 미설치",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "안전난간 설치 여부",
              _v(data, "railing_installed") or "□ 설치   □ 미설치",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "안전대 착용 여부",
              _v(data, "lanyard_worn") or "□ 착용   □ 미착용",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "안전대 부착설비 확인",
              _v(data, "anchor_confirmed") or "□ 확인   □ 미확인",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "낙하물 위험구역 설정",
              _v(data, "falling_zone_set") or "□ 설정   □ 미설정",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "출입통제 여부",
              _v(data, "access_control") or "□ 실시   □ 미실시",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "기상조건 확인",
              _v(data, "weather_confirmed") or "□ 확인 (풍속 10m/s 이상 시 작업 중지)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    return row


def _write_s5_pre_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 작업 전 안전조치  (이행 항목에 ■ 표시)")
    checked = set(data.get("pre_work_checks") or [])
    for item in PRE_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s6_workboard(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 작업발판·비계·사다리 확인  (제42조, 제57조 이하)")
    row = _write_notice(ws, row, NOTICE_SCAFFOLD)
    checked = set(data.get("workboard_checks") or [])
    for item in WORKBOARD_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s7_aerial(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 고소작업대 확인  (제86조 이하)")
    row = _write_notice(ws, row, NOTICE_AERIAL)
    checked = set(data.get("aerial_checks") or [])
    for item in AERIAL_WORK_PLATFORM_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s8_harness(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 안전대·추락방지설비 확인  (제44조)")
    row = _write_notice(ws, row, NOTICE_HARNESS)
    checked = set(data.get("harness_checks") or [])
    for item in HARNESS_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s9_falling(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 낙하물 방지 및 출입통제  (제42조~제43조)")
    checked = set(data.get("falling_checks") or [])
    for item in FALLING_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s10_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 작업허가 승인", fill=_FILL_SECTION2)
    row = _write_notice(ws, row, NOTICE_PERMIT_ISSUER)
    _write_lv(ws, row, "작업신청자 (서명)", _v(data, "work_supervisor"),
              _L1, _V1_START, _V1_END, height=32)
    _write_lv(ws, row, "관리감독자 (서명)", _v(data, "supervisor_name"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업허가자 (서명)", _v(data, "permit_issuer"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    _write_lv(ws, row, "허가 유효기간",
              _v(data, "validity_period") or "당일 1회 작업 한정 (KOSHA P-94-2021 참고)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=24)
    row += 1
    return row


def _write_s11_during_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 작업 중 점검")
    _write_cell(ws, row, 1, TOTAL_COLS,
                "이상 발생 (추락 징후, 낙하물 발생, 기상 악화 등) 시 즉시 작업 중단 후 허가자에게 보고",
                font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=24)
    row += 1
    _write_lv(ws, row, "작업 중 이상 사항",
              _v(data, "during_work_issues") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    return row


def _write_s12_photos(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 사진 및 증빙자료  [OPTIONAL]")
    row = _write_notice(ws, row, NOTICE_PHOTO)
    taken_photos = set(data.get("photo_items") or [])
    for item in PHOTO_ITEMS:
        mark = "■" if item in taken_photos else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "사진 파일 목록",
              _v(data, "photo_file_list") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s13_final_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "13. 확인 서명")
    # 작업 종료 후 확인
    row = _write_notice(ws, row, NOTICE_POST_WORK)
    checked = set(data.get("post_work_checks") or [])
    for item in POST_WORK_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "작업 종료 시각",
              _v(data, "work_end_time") or "",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "종료 확인자 서명",
              _v(data, "work_end_confirmer") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리감독자 서명",
              _v(data, "supervisor_name") or "",
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "안전관리자 서명",
              _v(data, "safety_manager_sign") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "최종 확인 서명",
              _v(data, "final_sign") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_work_at_height_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 고소작업 허가서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_info(ws, row, data)
    row = _write_s3_work_types(ws, row, data)
    row = _write_s4_fall_risk(ws, row, data)
    row = _write_s5_pre_work(ws, row, data)
    row = _write_s6_workboard(ws, row, data)
    row = _write_s7_aerial(ws, row, data)
    row = _write_s8_harness(ws, row, data)
    row = _write_s9_falling(ws, row, data)
    row = _write_s10_approval(ws, row, data)
    row = _write_s11_during_work(ws, row, data)
    row = _write_s12_photos(ws, row, data)
    _write_s13_final_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
