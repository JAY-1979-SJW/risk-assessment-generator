"""
안전보건교육일지 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건법 시행규칙 제32조, 별표 4~5
Form schema: docs/design/form_schema_from_law.md §2
Builder plan: docs/design/education_log_builder_plan.md
Generation lock: docs/design/form_generation_lock.md §3

Input  — form_data dict:
    site_name                str|None   사업장명 [관행]
    site_address             str|None   사업장 소재지 [관행]
    education_type           str|None   교육 종류 [법정]
    education_date           str|None   교육 일시 [법정]
    education_location       str|None   교육 장소 [법정]
    education_duration_hours str|None   교육 시간 합계 [법정/AUTO]
    education_target_job     str|None   교육 대상 [법정]
    instructor_name          str|None   강사명 [법정]
    instructor_qualification str|None   강사 자격 [법정]
    subjects                 list[dict] 교육 내용 반복 행 [법정]
        subject_name         str|None   교육 과목명
        subject_content      str|None   교육 내용 요약
        subject_hours        str|None   시간
    attendees                list[dict] 수강자 명단 [법정] (최대 30)
        attendee_name        str|None   성명
        attendee_job_type    str|None   직종(직위)
    confirmer_name           str|None   확인자 성명 [법정]
    confirmer_role           str|None   확인자 직위 [법정]
    confirm_date             str|None   확인 일자 [관행/AUTO]

Output — xlsx bytes (in-memory). 파일 저장은 호출자 책임.

Principles:
- 법정 외 필드 추가 금지.
- null/누락 → 빈 셀. 임의값 생성 금지.
- attendee_signature 는 항상 공란 (수기 서명 전용).
- 수강자 행은 MAX_ATTENDEES=30 고정 출력.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME = "교육일지"
SHEET_HEADING = "안전보건교육일지"
MAX_ATTENDEES = 30
_DEFAULT_SUBJECT_ROWS = 3  # subjects 없을 때 기본 빈 행 수

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 컬럼 정의 (8컬럼 A-H)
# A=6, B=16, C=14, D=12, E=10, F=14, G=14, H=14
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 6,   # A: 번호/레이블
    2: 16,  # B: 성명/값
    3: 14,  # C: 값
    4: 12,  # D: 직종/값
    5: 10,  # E: 레이블
    6: 14,  # F: 서명/값
    7: 14,  # G: 값
    8: 14,  # H: 값
}
TOTAL_COLS = 8

# 메타 블록 스팬: 라벨-값 분할
_L1, _V1_START, _V1_END = 1, 2, 4   # 왼쪽: 라벨=A, 값=B:D
_L2, _V2_START, _V2_END = 5, 6, 8   # 오른쪽: 라벨=E, 값=F:H


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    """dict에서 값 반환. None은 빈 문자열로."""
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    """범위 내 모든 셀에 _BORDER 적용."""
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    """셀(병합) 하나에 값+스타일 기록. 범위 전체에 테두리."""
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value = "" if value is None else value
    cell.font  = font  or _FONT_DEFAULT
    cell.fill  = fill  or PatternFill()
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_label_value_pair(ws, row: int,
                             label: str, value: Any,
                             label_col: int,
                             val_col1: int, val_col2: int) -> None:
    """[라벨 셀] [값 셀(병합)] 쌍을 한 행에 기록."""
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더링 함수
# ---------------------------------------------------------------------------

_SHEET_SUBTITLE = "「산업안전보건법」 제29조에 따른 안전보건교육"
_FONT_SUBTITLE  = Font(name="맑은 고딕", size=9, italic=True)


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    # P1: 법적 근거 부제 (시행규칙 제32조 / 법 제29조)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=_SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 사업장명 | site_name || 사업장 소재지 | site_address
    Row +1: 교육 종류 | education_type || 교육 장소 | education_location
    Row +2: 교육 일시 | education_date || 교육 시간 | education_duration_hours
    Row +3: 교육 대상 | education_target_job (전체 폭)
    Row +4: 강사명 | instructor_name || 강사 자격 | instructor_qualification
    """
    H = 20

    # row 0: 사업장명 / 사업장 소재지
    r = start_row
    _write_label_value_pair(ws, r, "사업장명", _v(data, "site_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "사업장 소재지", _v(data, "site_address"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    # row 1: 교육 종류 / 교육 장소
    r += 1
    _write_label_value_pair(ws, r, "교육 종류", _v(data, "education_type"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "교육 장소", _v(data, "education_location"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    # row 2: 교육 일시 / 교육 시간
    r += 1
    _write_label_value_pair(ws, r, "교육 일시", _v(data, "education_date"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "교육 시간", _v(data, "education_duration_hours"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    # row 3: 교육 대상 (전체 폭 값)
    r += 1
    _write_label_value_pair(ws, r, "교육 대상", _v(data, "education_target_job"),
                            _L1, _V1_START, TOTAL_COLS)
    # E 열에 빈 라벨 블록이 남지 않도록 처리 — 값이 B:H를 병합하므로 E 셀도 포함됨
    ws.row_dimensions[r].height = H

    # row 4: 강사명 / 강사 자격
    r += 1
    _write_label_value_pair(ws, r, "강사명", _v(data, "instructor_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "강사 자격", _v(data, "instructor_qualification"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    return r + 1


def _write_subject_table(ws, start_row: int,
                          subjects: List[Dict[str, Any]]) -> int:
    """
    교육 내용 섹션.
    헤더: 순번(A) | 교육 과목명(B:D) | 교육 내용 요약(E:G) | 시간(h)(H)
    데이터: 입력 subjects + 빈 행으로 _DEFAULT_SUBJECT_ROWS 보장.
    """
    # 섹션 헤더
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "교육 내용",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER,
                height=18)
    r += 1

    # 표 헤더
    _write_cell(ws, r, 1, 1, "순번",       font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4, "교육 과목명",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 7, "교육 내용 요약", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "시간(h)",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    # 데이터 행: 입력 subjects 기반, 최소 _DEFAULT_SUBJECT_ROWS 행 출력
    row_count = max(len(subjects), _DEFAULT_SUBJECT_ROWS)
    for i in range(row_count):
        subj = subjects[i] if i < len(subjects) else {}
        _write_cell(ws, r, 1, 1, i + 1,                      font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(subj, "subject_name"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 7, _v(subj, "subject_content"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 8, 8, _v(subj, "subject_hours"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 30  # P2: 20→30pt (내용 기록 공간)
        r += 1

    return r


def _write_attendee_table(ws, start_row: int,
                           attendees: List[Dict[str, Any]]) -> int:
    """
    수강자 명단 섹션.
    헤더: 번호(A) | 성명(B:C) | 직종(직위)(D:E) | 서명 또는 날인(F:H)
    데이터: MAX_ATTENDEES=30 행 고정. 입력 이후 행은 공란.
    서명란은 항상 공란 (수기 서명 전용).
    """
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "수강자 명단",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER,
                height=18)
    r += 1

    # 표 헤더
    _write_cell(ws, r, 1, 1, "번호",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 3, "성명",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 4, 5, "직종(직위)",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 8, "서명 또는 날인", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_ATTENDEES):
        att = attendees[i] if i < len(attendees) else {}
        _write_cell(ws, r, 1, 1, i + 1,                        font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 3, _v(att, "attendee_name"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4, 5, _v(att, "attendee_job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 8, "",                            font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22  # P3: 18→22pt (서명 공간)
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    확인 서명 섹션.
    Row 0: 확인자 성명 | confirmer_name || 확인자 직위 | confirmer_role
    Row 1: 서명란 라벨 | 공란(서명) || 확인 일자 | confirm_date
    """
    r = start_row
    _write_label_value_pair(ws, r, "확인자 성명", _v(data, "confirmer_name"),
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "확인자 직위", _v(data, "confirmer_role"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 22

    r += 1
    _write_label_value_pair(ws, r, "서명", "",
                            _L1, _V1_START, _V1_END)
    _write_label_value_pair(ws, r, "확인 일자", _v(data, "confirm_date"),
                            _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 40  # P4: 30→40pt (직인 공간)

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_education_log_sheet(ws, form_data: Dict[str, Any]) -> None:
    """
    주어진 worksheet에 교육일지를 렌더링.

    외부 호출자가 직접 ws에 접근해야 할 때 사용.
    build_education_log_excel은 단일 워크북/시트 래퍼.
    """
    data = form_data or {}
    subjects  = data.get("subjects")  or []
    attendees = data.get("attendees") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_subject_table(ws, row, subjects)
    row = _write_attendee_table(ws, row, attendees)
    _write_confirmation(ws, row, data)

    # 인쇄 설정
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_education_log_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 안전보건교육일지 xlsx 바이너리를 반환.

    Args:
        form_data: education_log_builder_plan.md 입력 스키마 준수 dict.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_education_log_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
