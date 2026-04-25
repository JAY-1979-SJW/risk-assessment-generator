"""
타워크레인 자체 점검표 — Excel 출력 모듈 (v1.0)  [CL-006]

법적 근거:
    산업안전보건기준에 관한 규칙 제142조~제146조
    (PARTIAL_VERIFIED: 제142조~제146조 타워크레인 설치·지지·작업 관련 조항.
     세부 점검항목 연결 조항은 원문 API 수집 후 확정 예정)

범위:
    본 서식은 타워크레인 자체 점검 전용입니다.
    이동식 크레인 점검표를 대체하지 않습니다.
    제147조·제148조(이동식 크레인 전용 조항)는 본 서식 근거에서 제외됩니다.

Required form_data keys:
    check_date    str  점검 일자
    work_location str  타워크레인 설치 위치
    checker_name  str  점검자 성명

Optional form_data keys:
    site_name             str  사업장명
    project_name          str  공사명
    work_date             str  작업 일자
    supervisor_name       str  관리감독자 성명
    crane_model           str  타워크레인 기종
    crane_reg_no          str  등록번호
    crane_capacity        str  정격하중 (ton)
    crane_height          str  설치 높이 (m)
    crane_work_radius     str  작업반경 (m)
    installation_date     str  설치 일자
    next_inspection_date  str  차기 검사 예정일
    operator_name         str  조종자 성명
    operator_license_no   str  조종자 면허번호
    doc_check_items       list[dict]  섹션3 항목 override (item, result, note)
    install_check_items   list[dict]  섹션4 항목 override
    structure_check_items list[dict]  섹션5 항목 override
    rope_check_items      list[dict]  섹션6 항목 override
    brake_check_items     list[dict]  섹션7 항목 override
    electric_check_items  list[dict]  섹션8 항목 override
    radius_check_items    list[dict]  섹션9 항목 override
    signal_check_items    list[dict]  섹션10 항목 override
    nonconformance_items  list[dict]  부적합 사항 (content, location, action, deadline, completed)
    daily_inspector_sign  str  일일 점검자 서명
    operator_sign         str  조종자 서명
    supervisor_sign       str  관리감독자 서명
    manager_sign          str  현장소장 서명
    sign_date             str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "타워크레인자체점검표"
SHEET_HEADING = "타워크레인 자체 점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제142조 이하 타워크레인 관련 조항에 따른 안전점검"

NOTICE_MOBILE  = "본 점검표는 타워크레인 자체 점검용이며, 이동식 크레인 점검표를 대체하지 않는다."
NOTICE_INSTALL = (
    "타워크레인 설치·해체·상승 작업은 관련 자격, 작업계획서, "
    "제조사 기준, 현장 안전조치 확인 후 수행한다."
)
NOTICE_NONCONFORM = "부적합 사항은 사용 전 시정 완료 후 재확인한다."
NOTICE_LAW        = "법령 조항은 현행 원문 확인 후 현장에 적용한다."

MAX_NONCONFORM = 5
TOTAL_COLS     = 9

# ──────────────────────────────────────────────
# 섹션별 기본 점검 항목
# ──────────────────────────────────────────────

_DOC_DEFAULTS: List[Dict[str, str]] = [
    {"item": "설치검사 또는 정기검사 유효 여부",               "result": "", "note": ""},
    {"item": "검사증·등록증 현장 비치 여부",                   "result": "", "note": ""},
    {"item": "타워크레인 작업계획서 작성 여부",                 "result": "", "note": ""},
    {"item": "장비사용계획서 작성 여부",                       "result": "", "note": ""},
    {"item": "제조사 사용설명서 현장 비치 여부",               "result": "", "note": ""},
]

_INSTALL_DEFAULTS: List[Dict[str, str]] = [
    {"item": "기초부(콘크리트 앵커) 침하·균열 없음 여부",       "result": "", "note": ""},
    {"item": "앵커프레임 볼트 체결 상태 이상 없음 여부",        "result": "", "note": ""},
    {"item": "마스트 수직도 적합 여부",                        "result": "", "note": ""},
    {"item": "벽체 연결(월 타이) 설치 상태 이상 없음 여부",     "result": "", "note": ""},
    {"item": "타워크레인 접지 설치 여부",                      "result": "", "note": ""},
]

_STRUCTURE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "마스트 볼트·핀 체결 및 균열·변형 없음 여부",      "result": "", "note": ""},
    {"item": "지브(붐) 균열·변형·부식 없음 여부",              "result": "", "note": ""},
    {"item": "카운터지브 균열·변형 없음 여부",                  "result": "", "note": ""},
    {"item": "턴테이블(선회환) 이상 없음 여부",                 "result": "", "note": ""},
    {"item": "캣헤드·타이바·텐션로드 이상 없음 여부",           "result": "", "note": ""},
]

_ROPE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "와이어로프 마모·단선·킹크·부식 없음 여부",         "result": "", "note": ""},
    {"item": "와이어로프 윤활 상태 적정 여부",                  "result": "", "note": ""},
    {"item": "훅 해지장치 정상 작동 여부",                     "result": "", "note": ""},
    {"item": "권과방지장치 정상 작동 여부",                     "result": "", "note": ""},
    {"item": "과부하방지장치 정상 작동 여부",                   "result": "", "note": ""},
]

_BRAKE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "권상 브레이크 정상 작동 여부",                   "result": "", "note": ""},
    {"item": "선회 브레이크 정상 작동 여부",                   "result": "", "note": ""},
    {"item": "트롤리 브레이크 정상 작동 여부",                  "result": "", "note": ""},
    {"item": "상·하 리미트 스위치 정상 작동 여부",             "result": "", "note": ""},
    {"item": "선회·트롤리 리미트 스위치 정상 작동 여부",        "result": "", "note": ""},
]

_ELECTRIC_DEFAULTS: List[Dict[str, str]] = [
    {"item": "전기판넬 내부 단선·누전·과열 없음 여부",          "result": "", "note": ""},
    {"item": "접지선 연결 상태 이상 없음 여부",                 "result": "", "note": ""},
    {"item": "비상정지장치 정상 작동 여부",                    "result": "", "note": ""},
    {"item": "운전실 내 조종장치 이상 없음 여부",              "result": "", "note": ""},
    {"item": "조명·경보장치 정상 작동 여부",                   "result": "", "note": ""},
]

_RADIUS_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업반경 내 장애물·충돌위험 확인 여부",           "result": "", "note": ""},
    {"item": "인접 타워크레인 간 충돌방지 조치 여부 (해당 시)", "result": "", "note": ""},
    {"item": "인양 중 하물 하부 출입통제 여부",                "result": "", "note": ""},
    {"item": "작업반경 내 근로자 출입 통제 여부",              "result": "", "note": ""},
    {"item": "강풍·악천후 시 작업중지 기준 적용 여부",         "result": "", "note": ""},
]

_SIGNAL_DEFAULTS: List[Dict[str, str]] = [
    {"item": "신호수 배치 여부 (타워크레인마다)",               "result": "", "note": ""},
    {"item": "신호수 신호방법 및 신호 일치 여부",              "result": "", "note": ""},
    {"item": "조종자 자격·면허 유효 여부",                    "result": "", "note": ""},
    {"item": "정격하중 초과 인양 없음 여부",                   "result": "", "note": ""},
    {"item": "인양물 결박 및 슬링 상태 확인 여부",            "result": "", "note": ""},
]

# ──────────────────────────────────────────────
# 스타일
# ──────────────────────────────────────────────

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_NOTICE   = Font(name="맑은 고딕", size=8,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FEF0E7")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 7, 3: 7, 4: 7, 5: 7, 6: 7,
    7: 8, 8: 10, 9: 10,
}

_L1, _V1S, _V1E = 1, 2, 5
_L2, _V2S, _V2E = 6, 7, 9
_LF, _VFS, _VFE = 1, 2, 9


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _wc(ws, row: int, c1: int, c2: int, value: Any, *,
        font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, c1, row, c2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _wlv(ws, row: int, label: str, value: Any,
         lc: int, vs: int, ve: int) -> None:
    _wc(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, vs, ve, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for idx, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    c = ws.cell(row=row, column=1, value=SHEET_HEADING)
    c.font = _FONT_TITLE; c.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    s = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    s.font = _FONT_SUBTITLE; s.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_notice_block(ws, row: int) -> int:
    for txt in (NOTICE_MOBILE, NOTICE_INSTALL, NOTICE_NONCONFORM, NOTICE_LAW):
        _wc(ws, row, 1, TOTAL_COLS, f"※ {txt}",
            font=_FONT_NOTICE, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
        row += 1
    return row


def _write_section1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "① 현장 기본정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "사업장명",  _v(data, "site_name"),    _L1, _V1S, _V1E)
    _wlv(ws, row, "공사명",    _v(data, "project_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검 일자", _v(data, "check_date"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "작업 일자", _v(data, "work_date"),    _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검자",    _v(data, "checker_name"),    _L1, _V1S, _V1E)
    _wlv(ws, row, "관리감독자", _v(data, "supervisor_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "설치 위치", _v(data, "work_location"), _LF, _VFS, _VFE)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_section2_crane_info(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "② 타워크레인 기본정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "기종",        _v(data, "crane_model"),        _L1, _V1S, _V1E)
    _wlv(ws, row, "등록번호",    _v(data, "crane_reg_no"),        _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "정격하중",    _v(data, "crane_capacity"),      _L1, _V1S, _V1E)
    _wlv(ws, row, "설치 높이",   _v(data, "crane_height"),        _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "작업반경",    _v(data, "crane_work_radius"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "설치 일자",   _v(data, "installation_date"),   _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "조종자",      _v(data, "operator_name"),       _L1, _V1S, _V1E)
    _wlv(ws, row, "면허번호",    _v(data, "operator_license_no"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "차기 검사일", _v(data, "next_inspection_date"), _LF, _VFS, _VFE)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_checklist_section(ws, row: int, section_no: str, title: str,
                              defaults: List[Dict[str, str]],
                              override: List[Dict[str, Any]]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, f"{section_no} {title}",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, TOTAL_COLS,
        "※ 결과: 이상없음 → ○ / 이상있음 → × / 해당없음 → -",
        font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_CENTER, height=14)
    row += 1
    _wc(ws, row, 1, 1, "No",       font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 6, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 7, 7, "결과",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 9, "비고",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    for i, default in enumerate(defaults):
        provided  = override[i] if i < len(override) else {}
        item_text = provided.get("item")   or default.get("item",   "")
        result    = provided.get("result") or default.get("result", "")
        note      = provided.get("note")   or default.get("note",   "")
        _wc(ws, row, 1, 1, i + 1,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 6, item_text, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 7, 7, result,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 8, 9, note,      font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 26
        row += 1

    return row


def _write_section11_nonconformance(ws, row: int,
                                     items: List[Dict[str, Any]]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "⑪ 부적합 및 시정조치",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, 1, "No",           font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 3, "부적합 내용",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 4, 5, "위치",         font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 6, 7, "시정조치 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 8, "완료기한",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 9, 9, "완료확인",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    for i in range(MAX_NONCONFORM):
        src = items[i] if i < len(items) else {}
        _wc(ws, row, 1, 1, i + 1,               font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 3, _v(src, "content"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 4, 5, _v(src, "location"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 6, 7, _v(src, "action"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 8, 8, _v(src, "deadline"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 9, 9, _v(src, "completed"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 26
        row += 1

    return row


def _write_section12_signatures(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "⑫ 확인 서명",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, 3, "서명일",
        font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, 4, 9, _v(data, "sign_date"),
        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = 20; row += 1

    labels    = ["일일 점검자 서명", "조종자 서명", "관리감독자 서명", "현장소장 서명"]
    sign_keys = ["daily_inspector_sign", "operator_sign", "supervisor_sign", "manager_sign"]
    # 4칸을 9열에 배분: (1-2), (3-4), (5-6), (7-9)
    col_ranges = [(1, 2), (3, 4), (5, 6), (7, 9)]

    for label, (cs, ce) in zip(labels, col_ranges):
        _wc(ws, row, cs, ce, label,
            font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    ws.row_dimensions[row].height = 18; row += 1

    for key, (cs, ce) in zip(sign_keys, col_ranges):
        _wc(ws, row, cs, ce, _v(data, key),
            font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 36; row += 1

    return row


# ──────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────

def build_tower_crane_self_inspection_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 타워크레인 자체 점검표 xlsx 바이너리를 반환.

    [CL-006]  타워크레인 자체 점검표
    법적 근거: 산업안전보건기준에 관한 규칙 제142조 이하 (PARTIAL_VERIFIED)
    이동식 크레인 점검과 분리 — 제147조·제148조 이동식 크레인 전용 조항 미적용.
    """
    data = form_data or {}

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_notice_block(ws, row)
    row = _write_section1_site_info(ws, row, data)
    row = _write_section2_crane_info(ws, row, data)

    row = _write_checklist_section(
        ws, row, "③", "검사·인증·서류 확인",
        _DOC_DEFAULTS, data.get("doc_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "④", "설치 상태 점검",
        _INSTALL_DEFAULTS, data.get("install_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑤", "구조부·마스트·지브 점검",
        _STRUCTURE_DEFAULTS, data.get("structure_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑥", "와이어로프·훅·권과방지장치 점검",
        _ROPE_DEFAULTS, data.get("rope_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑦", "브레이크·제동·리미트 장치 점검",
        _BRAKE_DEFAULTS, data.get("brake_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑧", "전기·제어·비상정지장치 점검",
        _ELECTRIC_DEFAULTS, data.get("electric_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑨", "작업반경·충돌방지·출입통제 점검",
        _RADIUS_DEFAULTS, data.get("radius_check_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑩", "신호수·조종자·작업방법 점검",
        _SIGNAL_DEFAULTS, data.get("signal_check_items") or [],
    )
    row = _write_section11_nonconformance(
        ws, row, data.get("nonconformance_items") or [],
    )
    _write_section12_signatures(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
