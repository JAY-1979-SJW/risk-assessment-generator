"""
보호구 지급 및 관리 점검표 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건기준에 관한 규칙 제32조 (보호구의 지급 등)
    산업안전보건기준에 관한 규칙 제33조 (보호구의 관리)
    산업안전보건기준에 관한 규칙 제34조 (전용 보호구 등)

분류: PRACTICAL — 공식 별지 제출서식 아님
      보호구 지급 및 관리 상태 점검 보조서식
      산업안전보건기준에 관한 규칙 제32조 보호구의 지급 등과 연계
      제33조 보호구의 관리 및 제34조 전용 보호구 관리와 연계

역할 분리:
    CL-008(본 서식): 보호구 지급·착용·손상·보관·재고·관리상태 점검
    PPE-001: 근로자별 보호구 지급·서명·교체·반납 이력
    DL-005: 작업 전 보호구 착용 등 작업개시 가능 여부 확인

Required form_data keys:
    site_name       str  현장명
    check_date      str  점검일
    inspector       str  점검자

Optional form_data keys:
    project_name            str  공사명
    check_zone              str  점검구역
    manager                 str  관리책임자
    writer                  str  작성자
    reviewer                str  검토자
    approver                str  승인자
    -- 섹션2: 보호구 지급 상태 점검 --
    helmet_issued           str  안전모 지급 여부
    safety_belt_issued      str  안전대 지급 여부
    safety_shoes_issued     str  안전화 지급 여부
    safety_glasses_issued   str  보안경 지급 여부
    face_shield_issued      str  보안면 지급 여부
    dust_mask_issued        str  방진마스크 지급 여부
    respirator_issued       str  방독마스크 지급 여부
    ear_protection_issued   str  청력보호구 지급 여부
    gloves_issued           str  보호장갑 지급 여부
    protective_clothing_issued str 보호복 지급 여부
    ppe_fit_for_work        str  작업조건 적합 여부
    -- 섹션3: 보호구 착용 상태 점검 --
    target_work             str  착용 대상 작업
    target_workers          str  착용 대상 인원
    actual_wearers          str  실제 착용 인원
    non_wearers             str  미착용 인원
    improper_wearing        str  착용 불량 사례
    non_wearing_reason      str  미착용 사유
    immediate_guidance      str  즉시 지도 여부
    work_stopped            str  작업중지 여부
    recheck_result          str  재확인 결과
    -- 섹션4: 보호구 손상·성능 상태 점검 --
    damaged_ppe_type        str  손상 보호구 종류
    damaged_qty             str  손상 수량
    damage_content          str  손상 내용
    expired_ppe             str  사용기한 경과 여부
    certification_ok        str  인증/검정 표시 확인
    disposal_needed         str  폐기 필요 여부
    replacement_needed      str  교체 필요 여부
    action_person           str  조치 담당자
    action_completed        str  완료 여부
    -- 섹션5: 개인 전용 보호구 관리 --
    individual_ppe_required str  개인 전용 지급 대상 여부
    infection_risk_ppe      str  감염 우려 보호구 여부
    shared_use              str  공동사용 여부
    id_marking              str  개인식별 표시 여부
    cleaning_status         str  세척/소독 상태
    storage_personal        str  보관 상태
    no_individual_reason    str  전용 보호구 미지급 사유
    individual_action_plan  str  조치 계획
    -- 섹션6: 보호구 보관 및 재고 관리 --
    storage_location        str  보관 장소
    storage_status          str  보관 상태
    contamination_exposure  str  오염/습기/직사광선 노출 여부
    stock_qty               str  재고 수량
    shortage_qty            str  부족 수량
    purchase_needed         str  추가 구매 필요 여부
    disposal_scheduled_qty  str  폐기 예정 수량
    next_inspection_date    str  다음 점검 예정일
    -- 섹션7: 교육·지도 및 서명 확인 --
    wearing_edu_done        str  착용 교육 실시 여부
    usage_guide_done        str  사용방법 안내 여부
    new_worker_edu          str  신규 근로자 교육 여부
    foreign_worker_guide    str  외국인 근로자 안내 여부
    edu_material_distributed str 교육자료 배포 여부
    signature_confirmed     str  서명부 확인 여부
    no_edu_action           str  미교육자 조치
    -- 섹션8: 미흡사항 및 개선조치 --
    deficiency_records      list[dict]  미흡사항 (content, risk_grade, action, person, due_date, completed_date, status, evidence, confirmer)
    -- 섹션9: PPE-001/DL-005 연계 확인 --
    ppe001_reflected        str  PPE-001 지급대장 반영 여부
    nonissue_supplemented   str  미지급자 지급대장 보완 여부
    dl005_linked            str  DL-005 작업 전 확인 연계 여부
    pre_work_recheck_needed str  작업 전 보호구 착용 재확인 필요 여부
    related_form_no         str  관련 서식 번호
    follow_up               str  후속 조치
    -- 섹션10: 확인 및 승인 --
    inspector_sign          str  점검자
    safety_manager_name     str  안전관리자
    supervisor_name         str  관리감독자
    site_manager_name       str  현장소장
    subcon_manager_name     str  협력업체 책임자
    confirm_date            str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "보호구관리점검표"
SHEET_HEADING = "보호구 지급 및 관리 점검표"
DOC_ID        = "CL-008"
TOTAL_COLS    = 10

_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL   = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}

MIN_TABLE_ROWS = 5

_PPE_ITEMS = [
    ("안전모",      "helmet"),
    ("안전대",      "safety_belt"),
    ("안전화",      "safety_shoes"),
    ("보안경",      "safety_glasses"),
    ("보안면",      "face_shield"),
    ("방진마스크",  "dust_mask"),
    ("방독마스크",  "respirator"),
    ("청력보호구",  "ear_protection"),
    ("보호장갑",    "gloves"),
    ("보호복",      "protective_clothing"),
]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[float] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=value)
    if font:  cell.font        = font
    if fill:  cell.fill        = fill
    if align: cell.alignment   = align
    if height:
        ws.row_dimensions[row].height = height
    for c in range(col1, col2 + 1):
        ws.cell(row=row, column=c).border = _BORDER


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vs: int, ve: int, height: float = 20) -> None:
    _cell(ws, row, lc, lc,  label, font=_FONT_BOLD,    fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vs, ve,  value, font=_FONT_DEFAULT,  align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, title,
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=22)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7


# ---------------------------------------------------------------------------
# 섹션 구현
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    _cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
          font=_FONT_TITLE, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=36)
    row += 1

    subtitle = (
        "공식 제출 서식 아님 — 보호구 지급 및 관리 상태 점검 보조서식  |  "
        f"산업안전보건기준에 관한 규칙 제32조 보호구의 지급 등과 연계  ({DOC_ID})"
    )
    _cell(ws, row, 1, TOTAL_COLS, subtitle,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    notice = (
        "제33조 보호구의 관리 및 제34조 전용 보호구 관리와 연계  |  "
        "PPE-001 보호구 지급 대장과 별도 관리  |  "
        "DL-005 작업 전 안전 확인서와 보호구 착용 확인 연계  |  "
        "손상·오염·사용기한 경과 보호구는 즉시 교체 또는 폐기 필요"
    )
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")

    _lv(ws, row, "공사명",    _v(data, "project_name"), 1, 2, 5)
    _lv(ws, row, "현장명",    _v(data, "site_name"),    6, 7, 10)
    row += 1
    _lv(ws, row, "점검일",    _v(data, "check_date"),   1, 2, 4)
    _lv(ws, row, "점검구역",  _v(data, "check_zone"),   5, 6, 7)
    _lv(ws, row, "점검자",    _v(data, "inspector"),    8, 9, 10)
    row += 1
    _lv(ws, row, "관리책임자", _v(data, "manager"),     1, 2, 4)
    _lv(ws, row, "작성자",     _v(data, "writer"),      5, 6, 7)
    _lv(ws, row, "검토자",     _v(data, "reviewer"),    8, 9, 10)
    row += 1
    _lv(ws, row, "승인자",     _v(data, "approver"),    1, 2, 4)
    row += 1

    notice = "개인정보·민감정보 최소 기재 — 성명·소속·직종 외 불필요한 개인정보 기재 금지"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_issue_status(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 보호구 지급 상태 점검")

    headers = ["No", "보호구 종류", "지급 여부", "비고"]
    col_spans = [(1, 1), (2, 5), (6, 8), (9, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    for i, (name, key) in enumerate(_PPE_ITEMS, 1):
        _cell(ws, row, 1, 1, i,                            align=_ALIGN_CENTER)
        _cell(ws, row, 2, 5, name,                         align=_ALIGN_LEFT)
        _cell(ws, row, 6, 8, _v(data, f"{key}_issued"),    align=_ALIGN_CENTER)
        _cell(ws, row, 9, 10, "",                          align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 18
        row += 1

    _lv(ws, row, "작업조건 적합 여부", _v(data, "ppe_fit_for_work"), 1, 3, 10)
    row += 1
    return row


def _write_wearing_status(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 보호구 착용 상태 점검")

    items = [
        ("착용 대상 작업",    "target_work",        1, 3, 10),
        ("착용 대상 인원",    "target_workers",     1, 3, 5),
        ("실제 착용 인원",    "actual_wearers",     6, 8, 10),
        ("미착용 인원",       "non_wearers",        1, 3, 5),
        ("착용 불량 사례",    "improper_wearing",   6, 8, 10),
        ("미착용 사유",       "non_wearing_reason", 1, 3, 10),
        ("즉시 지도 여부",    "immediate_guidance", 1, 3, 5),
        ("작업중지 여부",     "work_stopped",       6, 8, 10),
        ("재확인 결과",       "recheck_result",     1, 3, 10),
    ]
    for label, key, lc, vs, ve in items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    return row


def _write_damage_status(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 보호구 손상·성능 상태 점검")

    _lv(ws, row, "손상 보호구 종류",   _v(data, "damaged_ppe_type"),  1, 3, 7)
    _lv(ws, row, "손상 수량",          _v(data, "damaged_qty"),       8, 9, 10)
    row += 1
    _lv(ws, row, "손상 내용",          _v(data, "damage_content"),    1, 3, 10, height=24)
    row += 1
    _lv(ws, row, "사용기한 경과 여부", _v(data, "expired_ppe"),       1, 3, 5)
    _lv(ws, row, "인증/검정 표시 확인", _v(data, "certification_ok"), 6, 8, 10)
    row += 1
    _lv(ws, row, "폐기 필요 여부",     _v(data, "disposal_needed"),   1, 3, 5)
    _lv(ws, row, "교체 필요 여부",     _v(data, "replacement_needed"), 6, 8, 10)
    row += 1
    _lv(ws, row, "조치 담당자",        _v(data, "action_person"),     1, 3, 5)
    _lv(ws, row, "완료 여부",          _v(data, "action_completed"),  6, 8, 10)
    row += 1

    warn = "손상·오염·사용기한 경과 보호구는 즉시 교체 또는 폐기 필요"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {warn}",
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_individual_ppe(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 개인 전용 보호구 관리")

    _lv(ws, row, "개인 전용 지급 대상 여부",  _v(data, "individual_ppe_required"), 1, 3, 5)
    _lv(ws, row, "감염 우려 보호구 여부",      _v(data, "infection_risk_ppe"),      6, 8, 10)
    row += 1
    _lv(ws, row, "공동사용 여부",              _v(data, "shared_use"),              1, 3, 5)
    _lv(ws, row, "개인식별 표시 여부",         _v(data, "id_marking"),              6, 8, 10)
    row += 1
    _lv(ws, row, "세척/소독 상태",             _v(data, "cleaning_status"),         1, 3, 5)
    _lv(ws, row, "보관 상태",                  _v(data, "storage_personal"),        6, 8, 10)
    row += 1
    _lv(ws, row, "전용 보호구 미지급 사유",    _v(data, "no_individual_reason"),    1, 3, 10)
    row += 1
    _lv(ws, row, "조치 계획",                  _v(data, "individual_action_plan"),  1, 3, 10, height=24)
    row += 1
    return row


def _write_storage_stock(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 보호구 보관 및 재고 관리")

    _lv(ws, row, "보관 장소",            _v(data, "storage_location"),       1, 3, 5)
    _lv(ws, row, "보관 상태",            _v(data, "storage_status"),         6, 8, 10)
    row += 1
    _lv(ws, row, "오염/습기/직사광선 노출 여부", _v(data, "contamination_exposure"), 1, 3, 10)
    row += 1
    _lv(ws, row, "재고 수량",            _v(data, "stock_qty"),              1, 3, 4)
    _lv(ws, row, "부족 수량",            _v(data, "shortage_qty"),           5, 6, 7)
    _lv(ws, row, "추가 구매 필요",       _v(data, "purchase_needed"),        8, 9, 10)
    row += 1
    _lv(ws, row, "폐기 예정 수량",       _v(data, "disposal_scheduled_qty"), 1, 3, 5)
    _lv(ws, row, "다음 점검 예정일",     _v(data, "next_inspection_date"),   6, 8, 10)
    row += 1
    return row


def _write_education(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 교육·지도 및 서명 확인")

    edu_items = [
        ("착용 교육 실시 여부",     "wearing_edu_done",       1, 3, 5),
        ("사용방법 안내 여부",      "usage_guide_done",       6, 8, 10),
        ("신규 근로자 교육 여부",   "new_worker_edu",         1, 3, 5),
        ("외국인 근로자 안내 여부", "foreign_worker_guide",   6, 8, 10),
        ("교육자료 배포 여부",      "edu_material_distributed", 1, 3, 5),
        ("서명부 확인 여부",        "signature_confirmed",    6, 8, 10),
        ("미교육자 조치",           "no_edu_action",          1, 3, 10),
    ]
    for label, key, lc, vs, ve in edu_items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    return row


def _write_deficiency(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 미흡사항 및 개선조치")

    headers = ["No", "미흡사항", "위험등급", "개선조치", "담당자", "완료예정일", "완료일", "이행상태", "증빙", "확인자"]
    col_spans = [(1,1),(2,3),(4,4),(5,6),(7,7),(8,8),(9,9),(10,10),(10,10),(10,10)]
    # 컬럼 수 맞춰 10열 배분
    hdr_spans = [(1,1),(2,3),(4,4),(5,6),(7,7),(8,8),(9,9),(10,10)]
    hdr_labels = ["No", "미흡사항", "위험등급", "개선조치", "담당자", "완료예정일", "완료일·이행", "증빙/확인자"]
    for hdr, (cs, ce) in zip(hdr_labels, hdr_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    records: List[dict] = data.get("deficiency_records") or []
    n = max(len(records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1, 1, i + 1,                          align=_ALIGN_CENTER)
        _cell(ws, row, 2, 3, rec.get("content", ""),          align=_ALIGN_LEFT)
        _cell(ws, row, 4, 4, rec.get("risk_grade", ""),       align=_ALIGN_CENTER)
        _cell(ws, row, 5, 6, rec.get("action", ""),           align=_ALIGN_LEFT)
        _cell(ws, row, 7, 7, rec.get("person", ""),           align=_ALIGN_CENTER)
        _cell(ws, row, 8, 8, rec.get("due_date", ""),         align=_ALIGN_CENTER)
        _cell(ws, row, 9, 9, rec.get("completed_date", ""),   align=_ALIGN_CENTER)
        _cell(ws, row, 10, 10, rec.get("status", ""),         align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_linkage(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. PPE-001/DL-005 연계 확인")

    link_items = [
        ("PPE-001 지급대장 반영 여부",          "ppe001_reflected",       1, 3, 5),
        ("미지급자 지급대장 보완 여부",          "nonissue_supplemented",  6, 8, 10),
        ("DL-005 작업 전 확인 연계 여부",        "dl005_linked",           1, 3, 5),
        ("작업 전 보호구 착용 재확인 필요 여부", "pre_work_recheck_needed", 6, 8, 10),
        ("관련 서식 번호",                       "related_form_no",        1, 3, 5),
        ("후속 조치",                            "follow_up",              6, 8, 10),
    ]
    for label, key, lc, vs, ve in link_items:
        _lv(ws, row, label, _v(data, key), lc, vs, ve)
        if ve == 10:
            row += 1
    return row


def _write_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 확인 및 승인")

    sign_items = [
        ("점검자",        "inspector_sign"),
        ("안전관리자",    "safety_manager_name"),
        ("관리감독자",    "supervisor_name"),
        ("현장소장",      "site_manager_name"),
        ("협력업체 책임자", "subcon_manager_name"),
    ]
    for lbl, key in sign_items:
        _cell(ws, row, 1,  2,  lbl,               font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3,  5,  _v(data, key),     align=_ALIGN_CENTER)
        _cell(ws, row, 6,  7,  "서명",            font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10,  "",                align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 28
        row += 1

    _lv(ws, row, "확인일", _v(data, "confirm_date"), 1, 2, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_ppe_management_checklist(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_issue_status(ws, row, data)
    row = _write_wearing_status(ws, row, data)
    row = _write_damage_status(ws, row, data)
    row = _write_individual_ppe(ws, row, data)
    row = _write_storage_stock(ws, row, data)
    row = _write_education(ws, row, data)
    row = _write_deficiency(ws, row, data)
    row = _write_linkage(ws, row, data)
    row = _write_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
