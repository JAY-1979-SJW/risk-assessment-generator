"""
이동식 크레인 작업계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제38조 제1항 제14호, 제134조 이하
분류: GEN_INTERNAL — 원본 법정 서식 없음, 조문 기반 설계

Required form_data keys:
    crane_type        str  기계의 종류  [법정] 제38조
    crane_capacity    str  정격하중     [법정] 제134조
    work_method       str  작업방법     [법정] 제38조
    emergency_measure str  비상조치     [법정] 제38조

Optional form_data keys:
    site_name, project_name, work_location, work_date
    supervisor, contractor, prepared_by
    vehicle_no         str  차량번호
    outrigger_setup    str  아웃트리거 설치 상태 [실무 권장]
    ground_condition   str  지반 상태
    load_weight        str  인양물 중량
    work_radius        str  작업반경
    rigging_method     str  줄걸이 방법
    signal_worker      str  신호수 배치
    travel_route_text  str  이동 경로
    anti_topple_measures str 전도·협착·낙하 방지대책 [실무 권장]
    sign_date          str  작성일

    safety_steps  list[dict]  작업단계별 위험요소/안전조치 (max 10)
        task_step      str  작업단계
        hazard         str  위험요소
        safety_measure str  안전조치
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME     = "이동식크레인작업계획서"
SHEET_HEADING  = "이동식 크레인 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제38조 제1항 제14호에 따른 작업계획서"

MAX_STEPS  = 10
TOTAL_COLS = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SKETCH  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8

SKETCH_NOTE_DEFAULT = "※ 이동 경로 및 작업반경도를 아래 공간에 수기로 기재하시오 (아웃트리거 설치 위치·작업반경·위험구간 표시)"


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 위치", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 기간", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 책임자", _v(data, "supervisor"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자",  _v(data, "prepared_by"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_legal_items(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "법정 기재 사항 (기준규칙 제38조 제1항 제14호 · 제134조)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "장비명", _v(data, "crane_type"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "차량번호", _v(data, "vehicle_no"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "정격하중", _v(data, "crane_capacity"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업반경", _v(data, "work_radius"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "인양물 중량", _v(data, "load_weight"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "줄걸이 방법", _v(data, "rigging_method"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_cell(ws, r, _L1, _L1, "아웃트리거\n설치 상태",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "outrigger_setup"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "지반 상태",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "ground_condition"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "신호수 배치",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "signal_worker"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "작업방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "work_method"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_route_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "이동 경로  (기준규칙 제38조)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "이동 경로",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "travel_route_text"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    sketch_note = data.get("travel_route_sketch_note") or SKETCH_NOTE_DEFAULT
    _write_cell(ws, r, 1, TOTAL_COLS, sketch_note,
                font=_FONT_SMALL, fill=_FILL_SKETCH, align=_ALIGN_CENTER, height=16)
    r += 1

    for _ in range(4):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = ""
            cell.fill  = _FILL_NONE
            cell.alignment = _ALIGN_LEFT
            cell.border    = _BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_safety_section(ws, start_row: int, data: Dict[str, Any],
                           steps: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "위험요소 및 안전조치",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "전도·협착·낙하\n방지대책",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "anti_topple_measures"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, 1, 1, "순번",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 2, "작업단계", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 3, 5, "위험요소", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 8, "안전조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1

    for i in range(MAX_STEPS):
        item = steps[i] if i < len(steps) else {}
        _write_cell(ws, r, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 2, _v(item, "task_step"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 3, 5, _v(item, "hazard"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 8, _v(item, "safety_measure"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 28
        r += 1

    _write_cell(ws, r, _L1, _L1, "비상조치 방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "emergency_measure"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 2, "작성자 서명",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",        font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20; r += 1
    _write_cell(ws, r, 1, 4, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36
    return r + 1


def build_mobile_crane_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 이동식 크레인 작업계획서 xlsx 바이너리를 반환."""
    data  = form_data or {}
    steps = data.get("safety_steps") or []
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_legal_items(ws, row, data)
    row = _write_route_section(ws, row, data)
    row = _write_safety_section(ws, row, data, steps)
    _write_confirmation(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
