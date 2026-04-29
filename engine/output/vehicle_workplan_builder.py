"""
차량계 건설기계 작업계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제38조 제1항 제3호, 제170조
Layout:   docs/design/vehicle_workplan_builder_plan.md

Input — form_data dict:
    site_name                str|None   사업장명 [관행][AUTO]
    project_name             str|None   현장명 [관행][AUTO]
    work_location            str|None   작업 위치 [관행]
    work_date                str|None   작업 기간 [관행]
    supervisor               str|None   작업 책임자 [관행][EVID]
    contractor               str|None   도급업체 [관행]
    prepared_by              str|None   작성자 [관행]

    machine_type             str|None   기계의 종류 [법정] 제170조 제1호
    machine_capacity         str|None   기계의 성능·최대작업능력 [법정] 제170조 제1호
    operator_name            str|None   운전자 성명 [관행]
    operator_license         str|None   운전자 자격·면허 [관행][EVID]

    work_method              str|None   작업방법 [법정] 제170조 제3호
    guide_worker_required    str|None   유도자 배치 여부 및 방법 [관행] 제172조
    speed_limit              str|None   안전속도 제한 [관행] 제175조
    work_radius              str|None   작업반경 [관행]
    ground_survey            str|None   지형·지반 사전조사 [관행] 제171조

    travel_route_text        str|None   운행경로 텍스트 기술 [법정] 제170조 제2호
    travel_route_sketch_note str|None   스케치 박스 안내문구 (None이면 기본값 사용)

    access_control           str|None   출입통제 방법 [관행]
    emergency_contact        str|None   비상연락망 [관행]
    emergency_measure        str|None   비상조치 [관행]

    hazard_items     list[dict]  위험요소/안전조치 표 (MAX_HAZARD=10)
        hazard           str|None   위험 요소
        safety_measure   str|None   안전 조치
    pre_check_items  list[dict]  작업 전 점검 사항 (MAX_CHECKS=8)
        check_item       str|None   점검 항목
        result           str|None   이상유무
        note             str|None   비고

    sign_date                str|None   작성일 [관행][EVID]

Output — xlsx bytes (in-memory). 파일 저장은 호출자 책임.

Principles:
- null/누락 → 빈 셀. 임의값 생성 금지.
- 서명란은 항상 공란 (수기 서명 전용).
- hazard_items MAX_HAZARD=10 고정 출력, 초과분 무시.
- pre_check_items MAX_CHECKS=8 고정 출력, 초과분 무시.
- 운행경로 텍스트(Row 16)와 스케치 박스(Row 18~22) 반드시 분리.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME     = "차량계건설기계작업계획서"
SHEET_HEADING  = "차량계 건설기계 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제38조·제170조에 따른 작업계획서"

MAX_HAZARD = 10
MAX_CHECKS = 8
TOTAL_COLS = 8

SKETCH_NOTE_DEFAULT = "※ 운행경로 개략도를 아래 공간에 수기로 기재하시오 (진입로·구내동선·주요 위험구간 표시)"

# 법정 필수 라벨 목록 (검증 스크립트 재사용)
LEGAL_LABELS = [
    "기계의 종류",
    "기계의 성능·최대작업능력",
    "작업방법",
    "운행경로 기술",
    "유도자 배치 여부 및 방법",
    "출입통제 방법",
    "비상연락망",
    "비상조치",
    "지형·지반 사전조사",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8, italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SKETCH  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 열 너비 (A=14, B~G=12, H=10)
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 14,  # A
    2: 12,  # B
    3: 12,  # C
    4: 12,  # D
    5: 12,  # E
    6: 12,  # F
    7: 12,  # G
    8: 10,  # H
}

# 메타 블록 스팬 상수
_L1, _V1_START, _V1_END        = 1, 2, 4   # 좌: 라벨=A(1), 값=B:D(2-4)
_L2, _V2_START, _V2_END        = 5, 6, 8   # 우: 라벨=E(5), 값=F:H(6-8)
_FULL_VAL_START, _FULL_VAL_END = 2, 8      # 전폭 값: B:H(2-8)


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    """Row 1: 제목, Row 2: 부제."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 사업장명(A,B:D) | 현장명(E,F:H)
    Row +1: 작업위치(A,B:H 전폭)
    Row +2: 작업기간(A,B:D) | 작업책임자(E,F:H)
    Row +3: 도급업체(A,B:D) | 작성자(E,F:H)
    """
    H = 20
    r = start_row

    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "작업 위치", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "작업 기간", _v(data, "work_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업 책임자", _v(data, "supervisor"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자", _v(data, "prepared_by"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    return r + 1


def _write_legal_items(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더
    Row +1: 기계의 종류 [LAW]
    Row +2: 기계의 성능·최대작업능력 [LAW]
    Row +3: 작업방법 [LAW]
    Row +4: 운전자 성명(좌) / 운전자 자격·면허(우)
    Row +5: 유도자 배치 여부 및 방법
    Row +6: 안전속도 제한(좌) / 작업반경(우)
    Row +7: 지형·지반 사전조사
    """
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "법정 기재 사항 (산업안전보건기준에 관한 규칙 제38조·제170조)",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # Row: 기계의 종류 [LAW]
    _write_cell(ws, r, _L1, _L1, "기계의 종류",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "machine_type"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # Row: 기계의 성능·최대작업능력 [LAW]
    _write_cell(ws, r, _L1, _L1, "기계의 성능·최대작업능력",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "machine_capacity"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # Row: 작업방법 [LAW]
    _write_cell(ws, r, _L1, _L1, "작업방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "work_method"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # Row: 운전자 성명(좌) / 운전자 자격·면허(우) — 2분할
    _write_lv(ws, r, "운전자 성명", _v(data, "operator_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "운전자 자격·면허", _v(data, "operator_license"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    # Row: 유도자 배치 여부 및 방법
    _write_cell(ws, r, _L1, _L1, "유도자 배치 여부 및 방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "guide_worker_required"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # Row: 안전속도 제한(좌) / 작업반경(우) — 2분할
    _write_lv(ws, r, "안전속도 제한", _v(data, "speed_limit"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업반경", _v(data, "work_radius"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    # Row: 지형·지반 사전조사
    _write_cell(ws, r, _L1, _L1, "지형·지반 사전조사",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "ground_survey"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_travel_route(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더 (운행경로)
    Row +1: 운행경로 텍스트 기술 (travel_route_text) [LAW]
    Row +2: 스케치 박스 헤더 안내문구
    Row +3~+7: 스케치 박스 빈 영역 (5행)
    """
    r = start_row

    # 섹션 헤더
    _write_cell(ws, r, 1, TOTAL_COLS,
                "운행경로  (산업안전보건기준에 관한 규칙 제170조 제2호 법정 기재사항)",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 운행경로 텍스트 기술 [LAW]
    _write_cell(ws, r, _L1, _L1, "운행경로 기술",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=60)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "travel_route_text"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # 스케치 박스 헤더 — 안내문구
    sketch_note = data.get("travel_route_sketch_note") or SKETCH_NOTE_DEFAULT
    _write_cell(ws, r, 1, TOTAL_COLS, sketch_note,
                font=_FONT_SMALL, fill=_FILL_SKETCH,
                align=_ALIGN_CENTER, height=16)
    r += 1

    # 스케치 박스 — 5행, 병합 없음, 테두리만, 빈 영역
    for _ in range(5):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.value     = ""
            cell.fill      = _FILL_NONE
            cell.alignment = _ALIGN_LEFT
            cell.border    = _BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_access_emergency(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더
    Row +1: 출입통제 방법 (A, B:H)
    Row +2: 비상연락망(좌) / 비상조치(우)
    """
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "출입통제 · 유도자 · 비상연락망",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 출입통제 방법
    _write_cell(ws, r, _L1, _L1, "출입통제 방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "access_control"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # 비상연락망(좌) / 비상조치(우)
    _write_lv(ws, r, "비상연락망", _v(data, "emergency_contact"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "비상조치", _v(data, "emergency_measure"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    return r


def _write_hazard_table(ws, start_row: int,
                        hazard_items: List[Dict[str, Any]]) -> int:
    """
    Row +0: 섹션 헤더
    Row +1: 표 헤더 (순번|위험요소|안전조치)
    Row +2~+11: 데이터행 MAX_HAZARD=10 고정
    """
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "위험요소 및 안전조치",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 표 헤더
    _write_cell(ws, r, 1, 1, "순번",   font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4, "위험 요소", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "안전 조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_HAZARD):
        item = hazard_items[i] if i < len(hazard_items) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(item, "hazard"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 8, _v(item, "safety_measure"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_pre_check_table(ws, start_row: int,
                           pre_check_items: List[Dict[str, Any]]) -> int:
    """
    Row +0: 섹션 헤더
    Row +1: 표 헤더 (순번|점검항목|이상유무|비고)
    Row +2~+9: 데이터행 MAX_CHECKS=8 고정
    """
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "작업 전 점검 사항  (기준규칙 제175조)",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 표 헤더
    _write_cell(ws, r, 1, 1, "순번",   font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 5, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 7, "이상유무", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "비고",   font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_CHECKS):
        item = pre_check_items[i] if i < len(pre_check_items) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, _v(item, "check_item"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 7, _v(item, "result"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, _v(item, "note"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 25
        r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    """
    Row +0: 섹션 헤더
    Row +1: 서명란 (작성자 서명 | 공란 | 작업책임자 서명 | 공란 | 작성일 라벨 | 작성일값)
    """
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    # 서명란: A:B=작성자서명 라벨, C:D=공란, E:F=작업책임자서명 라벨, G=작성일 라벨, H=작성일값
    _write_cell(ws, r, 1, 2, "작성자 서명 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",  # 수기 서명 공란
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업책임자 서명 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 36

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_vehicle_workplan_sheet(ws, form_data: Dict[str, Any]) -> None:
    """
    주어진 worksheet에 차량계 건설기계 작업계획서를 렌더링.

    외부 호출자가 직접 ws에 접근해야 할 때 사용.
    build_vehicle_workplan_excel 은 단일 워크북/시트 래퍼.
    """
    data         = form_data or {}
    hazard_items = data.get("hazard_items") or []
    pre_checks   = data.get("pre_check_items") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_legal_items(ws, row, data)
    row = _write_travel_route(ws, row, data)
    row = _write_access_emergency(ws, row, data)
    row = _write_hazard_table(ws, row, hazard_items)
    row = _write_pre_check_table(ws, row, pre_checks)
    _write_confirmation(ws, row, data)

    # 인쇄 설정
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_area              = f"A1:H{row + 1}"
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:2"


def build_vehicle_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 차량계 건설기계 작업계획서 xlsx 바이너리를 반환.

    Args:
        form_data: vehicle_workplan_builder_plan.md 입력 스키마 준수 dict.
            hazard_items > MAX_HAZARD(10) 시 초과분 무시.
            pre_check_items > MAX_CHECKS(8) 시 초과분 무시.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_vehicle_workplan_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
