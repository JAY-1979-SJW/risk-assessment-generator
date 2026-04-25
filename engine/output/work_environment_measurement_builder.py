"""
작업환경측정 실시 및 결과 관리대장 — Excel 출력 모듈 (v1.0).

법적 근거: 산업안전보건법 제125조 (작업환경측정)
분류:     GEN_INTERNAL — 외부 측정기관 발급 결과보고서를 대체하지 않음
          이 문서는 측정 실시·결과·사후조치를 사업장에서 자체 관리하는 관리대장임.

주의:
- 공식 측정기관 발급 결과보고서 원본은 별도 첨부·보관 필수.
- 측정값 단위·노출기준은 입력값으로만 채움. 임의 기준치 자동 생성 금지.
- 노출기준 초과 여부는 입력값 없으면 빈칸("확인 필요" 입력 안내).

Input — form_data dict:
    site_name              str|None   사업장명
    project_name           str|None   현장/공장명
    work_location          str|None   측정 장소
    measurement_period     str|None   측정 기간
    supervisor             str|None   작업환경관리 담당자
    contractor             str|None   도급업체
    prepared_by            str|None   작성자

    target_process         str|None   측정 대상 공정/작업 장소
    hazardous_agents       str|None   유해인자 목록 (쉼표 구분)
    measurement_agency     str|None   측정기관명
    agency_contact         str|None   측정기관 연락처
    measurement_date       str|None   측정 실시일
    result_received_date   str|None   결과 수령일

    result_summary         str|None   측정 결과 요약
    exceedance_status      str|None   노출기준 초과 여부 ("초과 없음"/"초과 있음"/"확인 필요")
    exceedance_detail      str|None   초과 유해인자 및 수치

    improvement_plan       str|None   개선조치 계획
    improvement_deadline   str|None   개선 완료 목표일
    improvement_done       str|None   개선 완료 여부

    worker_notification    str|None   근로자 고지/설명 여부 및 방법
    original_attached      str|None   원본 결과서 첨부 여부 (예: "첨부 완료")

    measurement_rows list[dict]  측정 결과 상세 표 (MAX_ROWS=10)
        target_location    str|None   측정 장소
        hazardous_agent    str|None   유해인자
        measured_value     str|None   측정값 (단위 포함)
        exposure_limit     str|None   노출기준
        exceedance         str|None   초과 여부

    confirmer_name         str|None   확인자 성명
    confirmer_role         str|None   확인자 직위
    sign_date              str|None   작성일

Output — xlsx bytes (in-memory).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME    = "작업환경측정관리대장"
SHEET_HEADING = "작업환경측정 실시 및 결과 관리대장"
SHEET_SUBTITLE = "산업안전보건법 제125조에 따른 작업환경측정 실시·결과·사후조치 관리용 문서"

ORIGINAL_REPORT_NOTE = (
    "※ 외부 전문 측정기관이 발급한 작업환경측정 결과보고서 원본을 반드시 보관·첨부하여야 합니다."
    "  이 관리대장은 원본 결과서를 대체하지 않습니다."
)

MAX_ROWS   = 10
TOTAL_COLS = 8

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8, italic=True)
_FONT_NOTE     = Font(name="맑은 고딕", size=9, bold=True, color="C00000")

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NOTE    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10,
}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H = 20
    r = start_row

    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장/공장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "측정 장소", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "측정 기간", _v(data, "measurement_period"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "담당자", _v(data, "supervisor"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자", _v(data, "prepared_by"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    return r + 1


def _write_measurement_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "측정 대상 및 측정기관 정보",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "측정 대상\n공정/장소",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "target_process"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "유해인자\n목록",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "hazardous_agents"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_lv(ws, r, "측정기관명", _v(data, "measurement_agency"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "측정기관 연락처", _v(data, "agency_contact"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_lv(ws, r, "측정 실시일", _v(data, "measurement_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "결과 수령일", _v(data, "result_received_date"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20

    return r + 1


def _write_result_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "측정 결과 요약",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "결과 요약",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "result_summary"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    exceedance = _v(data, "exceedance_status") or "확인 필요"
    _write_cell(ws, r, _L1, _L1, "노출기준\n초과 여부",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                exceedance, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "초과 유해인자\n및 수치",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "exceedance_detail"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_measurement_table(ws, start_row: int,
                              measurement_rows: List[Dict[str, Any]]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "측정 결과 상세 (유해인자별)",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 1, "순번",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 3, "측정 장소", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 4, 5, "유해인자",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 6, "측정값",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "노출기준",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "초과여부",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_ROWS):
        item = measurement_rows[i] if i < len(measurement_rows) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 3, _v(item, "target_location"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 4, 5, _v(item, "hazardous_agent"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 6, _v(item, "measured_value"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7, 7, _v(item, "exposure_limit"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, _v(item, "exceedance"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def _write_improvement_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "개선조치 및 근로자 고지",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "개선조치 계획",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "improvement_plan"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_lv(ws, r, "개선 완료\n목표일", _v(data, "improvement_deadline"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "개선 완료\n여부", _v(data, "improvement_done"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30
    r += 1

    _write_cell(ws, r, _L1, _L1, "근로자 고지\n여부 및 방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "worker_notification"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_original_note(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "원본 결과서 첨부 확인",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "원본 첨부\n여부",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "original_attached"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, 1, TOTAL_COLS,
                ORIGINAL_REPORT_NOTE,
                font=_FONT_NOTE, fill=_FILL_NOTE,
                align=_ALIGN_CENTER, height=36)
    r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 2, "작성자 서명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "확인자 서명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_cell(ws, r, 1, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_work_environment_measurement_sheet(ws, form_data: Dict[str, Any]) -> None:
    data             = form_data or {}
    measurement_rows = data.get("measurement_rows") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_measurement_info(ws, row, data)
    row = _write_result_section(ws, row, data)
    row = _write_measurement_table(ws, row, measurement_rows)
    row = _write_improvement_section(ws, row, data)
    row = _write_original_note(ws, row, data)
    _write_confirmation(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_work_environment_measurement_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 작업환경측정 실시 및 결과 관리대장 xlsx 바이너리를 반환.

    이 문서는 외부 측정기관이 발급한 공식 결과보고서를 대체하지 않습니다.
    사업장 자체 실시·결과·사후조치 관리용 문서입니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_work_environment_measurement_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
