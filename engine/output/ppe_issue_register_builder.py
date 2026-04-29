"""
보호구 지급 대장 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건기준에 관한 규칙 제32조 (보호구의 지급 등)
    산업안전보건기준에 관한 규칙 제33조 (보호구의 관리)
    산업안전보건기준에 관한 규칙 제34조 (전용 보호구 등)

분류: PRACTICAL — 공식 별지 제출서식 아님
      보호구 지급·착용 및 교체 이력 관리 보조서식
      산업안전보건기준에 관한 규칙 제32조 보호구의 지급 등과 연계

역할 분리:
    PPE-001(본 서식): 보호구 지급, 착용 확인, 교체·반납, 개인 전용 보호구 관리, 지급 증빙 기록
    CL-008: 보호구 관리 점검표 (상태·수량 점검)
    DL-005: 작업 전 안전 확인서 (보호구 착용 여부 확인 항목 포함)

Required form_data keys:
    site_name       str  현장명
    manager         str  관리책임자

Optional form_data keys:
    project_name            str  공사명
    company_name            str  업체명
    period                  str  관리 기간
    prepared_date           str  작성일
    approver                str  승인자
    -- 섹션2: 지급 대상자 정보 --
    worker_name             str  근로자 성명
    worker_id               str  근로자 번호
    occupation              str  직종
    subcontractor           str  소속 협력업체
    work_location           str  작업 위치
    -- 섹션3: 보호구 지급 내역 --
    issue_records           list[dict]  지급 이력 (worker_name, ppe_type, spec, issue_date, qty, receipt_confirm)
    -- 섹션4: 보호구 종류별 관리 --
    ppe_type_records        list[dict]  종류별 관리 (ppe_type, standard, qty_required, qty_issued, qty_remain, remarks)
    -- 섹션5: 지급 전 적합성 확인 --
    ppe_standard_checked    str  규격·성능 기준 충족 여부
    qty_sufficient          str  작업 근로자 수 이상 지급 여부
    individual_ppe_needed   str  공동사용 감염 우려 개인 전용 지급 여부
    ppe_condition_ok        str  보호구 상태 이상 없음 여부
    -- 섹션6: 착용 교육 및 서명 --
    edu_conducted           str  사용·착용방법 교육 여부
    edu_date                str  교육 일시
    edu_instructor          str  교육 실시자
    edu_attendees           str  교육 참여 인원
    signature_records       list[dict]  서명 이력 (worker_name, date, signature_confirm)
    -- 섹션7: 교체·반납·폐기 이력 --
    replace_records         list[dict]  교체·반납 이력 (worker_name, ppe_type, action, date, reason, handler)
    -- 섹션8: 미지급·미착용 조치 --
    nonissue_records        list[dict]  미지급 조치 (worker_name, ppe_type, reason, action, action_date)
    nonwear_records         list[dict]  미착용 조치 (worker_name, ppe_type, date, action)
    -- 섹션9: 점검 및 재고 관리 --
    inspection_date         str  점검일
    inspector               str  점검자
    stock_records           list[dict]  재고 현황 (ppe_type, spec, stock_qty, issued_qty, remain_qty, status)
    inspection_result       str  점검 결과 요약
    -- 섹션10: 확인 및 승인 --
    supervisor_name         str  관리감독자
    safety_manager_name     str  안전관리자
    site_manager_name       str  현장소장
    confirm_date            str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "보호구지급대장"
SHEET_HEADING = "보호구 지급 대장"
DOC_ID        = "PPE-001"
TOTAL_COLS    = 10

_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL   = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}

MIN_TABLE_ROWS = 5


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[float] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=value)
    if font:  cell.font   = font
    if fill:  cell.fill   = fill
    if align: cell.alignment = align
    if height:
        ws.row_dimensions[row].height = height
    for c in range(col1, col2 + 1):
        ws.cell(row=row, column=c).border = _BORDER


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vs: int, ve: int, height: float = 20) -> None:
    _cell(ws, row, lc, lc,  label, font=_FONT_BOLD,    fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vs, ve,  value, font=_FONT_DEFAULT,  align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, title,
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=22)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7


# ---------------------------------------------------------------------------
# 섹션 구현
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    _cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
          font=_FONT_TITLE, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=36)
    row += 1

    subtitle = (
        "공식 제출 서식 아님 — 보호구 지급·착용 및 교체 이력 관리 보조서식  |  "
        f"산업안전보건기준에 관한 규칙 제32조 보호구의 지급 등과 연계  ({DOC_ID})"
    )
    _cell(ws, row, 1, TOTAL_COLS, subtitle,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    notice = (
        "작업조건에 맞는 보호구를 작업 근로자 수 이상으로 지급하고 착용 확인 필요  |  "
        "공동사용으로 감염 우려가 있는 경우 개인 전용 보호구 지급 필요  |  "
        "DL-005 작업 전 안전 확인서 및 CL-008 보호구 관리 점검표와 별도 관리"
    )
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")

    _lv(ws, row, "현장명",    _v(data, "site_name"),    1, 2, 5)
    _lv(ws, row, "공사명",    _v(data, "project_name"), 6, 7, 10)
    row += 1
    _lv(ws, row, "업체명",    _v(data, "company_name"), 1, 2, 5)
    _lv(ws, row, "관리책임자", _v(data, "manager"),     6, 7, 10)
    row += 1
    _lv(ws, row, "관리 기간", _v(data, "period"),       1, 2, 5)
    _lv(ws, row, "작성일",    _v(data, "prepared_date"), 6, 7, 8)
    _lv(ws, row, "승인자",    _v(data, "approver"),      9, 10, 10)
    row += 1

    notice = "개인정보·민감정보 최소 기재 — 성명·소속·직종 외 불필요한 개인정보 기재 금지"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_worker_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 지급 대상자 정보")

    _lv(ws, row, "근로자 성명", _v(data, "worker_name"),    1, 2, 4)
    _lv(ws, row, "근로자 번호", _v(data, "worker_id"),      5, 6, 7)
    _lv(ws, row, "직종",        _v(data, "occupation"),     8, 9, 10)
    row += 1
    _lv(ws, row, "협력업체",    _v(data, "subcontractor"),  1, 2, 5)
    _lv(ws, row, "작업 위치",   _v(data, "work_location"),  6, 7, 10)
    row += 1
    return row


def _write_issue_records(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 보호구 지급 내역")

    headers = ["No", "근로자 성명", "보호구 종류", "규격/성능", "지급일", "수량", "수령 확인"]
    col_spans = [(1, 1), (2, 3), (4, 5), (6, 7), (8, 8), (9, 9), (10, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    records: List[dict] = data.get("issue_records") or []
    n = max(len(records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1, 1, i + 1, align=_ALIGN_CENTER)
        _cell(ws, row, 2, 3, rec.get("worker_name", ""),    align=_ALIGN_LEFT)
        _cell(ws, row, 4, 5, rec.get("ppe_type", ""),       align=_ALIGN_LEFT)
        _cell(ws, row, 6, 7, rec.get("spec", ""),           align=_ALIGN_LEFT)
        _cell(ws, row, 8, 8, rec.get("issue_date", ""),     align=_ALIGN_CENTER)
        _cell(ws, row, 9, 9, rec.get("qty", ""),            align=_ALIGN_CENTER)
        _cell(ws, row, 10, 10, rec.get("receipt_confirm", ""), align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_ppe_type_mgmt(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 보호구 종류별 관리")

    headers = ["보호구 종류", "적용 규격", "필요 수량", "지급 수량", "잔여 수량", "비고"]
    col_spans = [(1, 2), (3, 4), (5, 6), (7, 7), (8, 8), (9, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    records: List[dict] = data.get("ppe_type_records") or []
    n = max(len(records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1, 2,  rec.get("ppe_type", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 3, 4,  rec.get("standard", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 5, 6,  rec.get("qty_required", ""),  align=_ALIGN_CENTER)
        _cell(ws, row, 7, 7,  rec.get("qty_issued", ""),    align=_ALIGN_CENTER)
        _cell(ws, row, 8, 8,  rec.get("qty_remain", ""),    align=_ALIGN_CENTER)
        _cell(ws, row, 9, 10, rec.get("remarks", ""),       align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_suitability_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 지급 전 적합성 확인")

    items = [
        ("규격·성능 기준 충족 여부",              "ppe_standard_checked"),
        ("작업 근로자 수 이상 지급 여부",          "qty_sufficient"),
        ("개인 전용 보호구 지급 여부 (감염 우려)", "individual_ppe_needed"),
        ("보호구 상태 이상 없음",                  "ppe_condition_ok"),
    ]
    for label, key in items:
        _lv(ws, row, label, _v(data, key), 1, 4, 10)
        row += 1

    notice = (
        "보호구 사용방법 및 착용방법 교육·서명 관리 필요  |  "
        "공동사용으로 감염 우려가 있는 경우 개인 전용 보호구 지급 필요"
    )
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_edu_signature(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 착용 교육 및 서명")

    _lv(ws, row, "교육 실시 여부",  _v(data, "edu_conducted"),  1, 2, 4)
    _lv(ws, row, "교육 일시",       _v(data, "edu_date"),        5, 6, 7)
    _lv(ws, row, "교육 실시자",     _v(data, "edu_instructor"),  8, 9, 10)
    row += 1
    _lv(ws, row, "교육 참여 인원",  _v(data, "edu_attendees"),  1, 2, 10)
    row += 1

    # 서명 테이블
    sign_headers = ["No", "근로자 성명", "서명 일시", "서명 확인"]
    sign_spans   = [(1, 1), (2, 4), (5, 7), (8, 10)]
    for hdr, (cs, ce) in zip(sign_headers, sign_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    sig_records: List[dict] = data.get("signature_records") or []
    n = max(len(sig_records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = sig_records[i] if i < len(sig_records) else {}
        _cell(ws, row, 1,  1,  i + 1,                       align=_ALIGN_CENTER)
        _cell(ws, row, 2,  4,  rec.get("worker_name", ""),  align=_ALIGN_LEFT)
        _cell(ws, row, 5,  7,  rec.get("date", ""),         align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10,  rec.get("signature_confirm", ""), align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_replace_records(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 교체·반납·폐기 이력")

    headers = ["No", "근로자 성명", "보호구 종류", "구분", "일시", "사유", "처리자"]
    col_spans = [(1, 1), (2, 3), (4, 5), (6, 6), (7, 7), (8, 9), (10, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    records: List[dict] = data.get("replace_records") or []
    n = max(len(records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1,  1,  i + 1,                     align=_ALIGN_CENTER)
        _cell(ws, row, 2,  3,  rec.get("worker_name", ""), align=_ALIGN_LEFT)
        _cell(ws, row, 4,  5,  rec.get("ppe_type", ""),    align=_ALIGN_LEFT)
        _cell(ws, row, 6,  6,  rec.get("action", ""),      align=_ALIGN_CENTER)
        _cell(ws, row, 7,  7,  rec.get("date", ""),        align=_ALIGN_CENTER)
        _cell(ws, row, 8,  9,  rec.get("reason", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 10, 10, rec.get("handler", ""),     align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_nonissue_action(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 미지급·미착용 조치")

    _cell(ws, row, 1, TOTAL_COLS, "▶ 미지급 조치",
          font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LEFT, height=18)
    row += 1
    headers = ["No", "근로자 성명", "보호구 종류", "미지급 사유", "조치 내용", "조치일"]
    col_spans = [(1, 1), (2, 3), (4, 5), (6, 7), (8, 9), (10, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    row += 1
    records: List[dict] = data.get("nonissue_records") or []
    n = max(len(records), 3)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1,  1,  i + 1,                     align=_ALIGN_CENTER)
        _cell(ws, row, 2,  3,  rec.get("worker_name", ""), align=_ALIGN_LEFT)
        _cell(ws, row, 4,  5,  rec.get("ppe_type", ""),    align=_ALIGN_LEFT)
        _cell(ws, row, 6,  7,  rec.get("reason", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 8,  9,  rec.get("action", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 10, 10, rec.get("action_date", ""), align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 18
        row += 1

    _cell(ws, row, 1, TOTAL_COLS, "▶ 미착용 조치",
          font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LEFT, height=18)
    row += 1
    headers2 = ["No", "근로자 성명", "보호구 종류", "미착용 일시", "조치 내용", "비고"]
    for hdr, (cs, ce) in zip(headers2, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    row += 1
    records2: List[dict] = data.get("nonwear_records") or []
    n2 = max(len(records2), 3)
    for i in range(n2):
        rec = records2[i] if i < len(records2) else {}
        _cell(ws, row, 1,  1,  i + 1,                     align=_ALIGN_CENTER)
        _cell(ws, row, 2,  3,  rec.get("worker_name", ""), align=_ALIGN_LEFT)
        _cell(ws, row, 4,  5,  rec.get("ppe_type", ""),    align=_ALIGN_LEFT)
        _cell(ws, row, 6,  7,  rec.get("date", ""),        align=_ALIGN_CENTER)
        _cell(ws, row, 8,  9,  rec.get("action", ""),      align=_ALIGN_LEFT)
        _cell(ws, row, 10, 10, "",                          align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 18
        row += 1
    return row


def _write_inspection(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. 점검 및 재고 관리")

    _lv(ws, row, "점검일",  _v(data, "inspection_date"), 1, 2, 5)
    _lv(ws, row, "점검자",  _v(data, "inspector"),       6, 7, 10)
    row += 1

    headers = ["보호구 종류", "규격", "재고 수량", "지급 수량", "잔여 수량", "상태"]
    col_spans = [(1, 2), (3, 4), (5, 6), (7, 7), (8, 8), (9, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    row += 1

    records: List[dict] = data.get("stock_records") or []
    n = max(len(records), MIN_TABLE_ROWS)
    for i in range(n):
        rec = records[i] if i < len(records) else {}
        _cell(ws, row, 1,  2,  rec.get("ppe_type", ""),   align=_ALIGN_LEFT)
        _cell(ws, row, 3,  4,  rec.get("spec", ""),       align=_ALIGN_LEFT)
        _cell(ws, row, 5,  6,  rec.get("stock_qty", ""),  align=_ALIGN_CENTER)
        _cell(ws, row, 7,  7,  rec.get("issued_qty", ""), align=_ALIGN_CENTER)
        _cell(ws, row, 8,  8,  rec.get("remain_qty", ""), align=_ALIGN_CENTER)
        _cell(ws, row, 9, 10,  rec.get("status", ""),     align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 18
        row += 1

    _lv(ws, row, "점검 결과", _v(data, "inspection_result"), 1, 2, 10, height=24)
    row += 1
    return row


def _write_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 확인 및 승인")

    sign_items = [
        ("관리감독자",  "supervisor_name"),
        ("안전관리자",  "safety_manager_name"),
        ("현장소장",    "site_manager_name"),
    ]
    for lbl, key in sign_items:
        _cell(ws, row, 1,  2,  lbl,                  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3,  5,  _v(data, key),        align=_ALIGN_CENTER)
        _cell(ws, row, 6,  7,  "서명",               font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10,  "",                   align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 28
        row += 1

    _lv(ws, row, "확인일", _v(data, "confirm_date"), 1, 2, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_ppe_issue_register(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_worker_info(ws, row, data)
    row = _write_issue_records(ws, row, data)
    row = _write_ppe_type_mgmt(ws, row, data)
    row = _write_suitability_check(ws, row, data)
    row = _write_edu_signature(ws, row, data)
    row = _write_replace_records(ws, row, data)
    row = _write_nonissue_action(ws, row, data)
    row = _write_inspection(ws, row, data)
    row = _write_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
