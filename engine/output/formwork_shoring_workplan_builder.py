"""
거푸집·동바리 작업계획서 — Excel 출력 모듈 (v1).

법적 근거: 산업안전보건기준에 관한 규칙 제38조·제39조·제331조, 제328조~제337조
           (evidence: WP-015-L1~L4, PARTIAL_VERIFIED)

주의:
- 구조계산을 builder가 자동 수행하지 않는다.
- 조립도 자체를 임의 생성하지 않는다.
- 구조검토서/조립도는 첨부·보관 확인 필드로만 구현한다.
- 이 서류는 구조계산서/구조검토서/조립도 원본을 대체하지 않는다.

Input — form_data dict:
    site_name              str|None   사업장명
    project_name           str|None   현장명
    work_date              str|None   작성일
    work_period            str|None   작업기간
    contractor             str|None   시공사/협력업체
    prepared_by            str|None   작성자
    reviewer               str|None   검토자
    approver               str|None   승인자

    work_location          str|None   작업 위치
    work_scope             str|None   작업 범위
    formwork_type          str|None   거푸집 종류
    shoring_type           str|None   동바리 종류
    floor_section          str|None   층/구간/타설 부위
    work_phases            str|None   작업 단계

    survey_ground          str|None   지반/바닥 상태
    survey_substructure    str|None   하부 구조물 상태
    survey_opening         str|None   개구부/단부 존재 여부
    survey_material_place  str|None   자재 적치 위치
    survey_equipment_route str|None   장비 진입 동선
    survey_weather         str|None   기상 조건
    survey_lighting        str|None   조명/통로 확보

    structural_review_done str|None   구조검토 실시 여부
    structural_reviewer    str|None   구조검토자
    assembly_drawing_done  str|None   조립도 작성 여부
    member_spec            str|None   동바리/멍에/장선 부재 규격
    install_interval       str|None   설치 간격
    joint_method           str|None   이음 방법
    brace_plan             str|None   수평연결재/가새/브레이싱 계획
    base_plate_plan        str|None   받침대/깔판/미끄럼 방지 조치
    structural_doc_attached str|None  구조검토서 첨부 여부
    assembly_drawing_attached str|None 조립도 첨부 여부

    work_sequence          str|None   작업 순서 및 방법 (전체 텍스트)

    hazard_items           list[dict] 주요 위험요인 (MAX_HAZARD=10)
        hazard             str|None   위험요인
        safety_measure     str|None   안전대책

    safety_measures_text   str|None   안전대책 전체 텍스트

    work_commander_name    str|None   작업지휘자 성명
    work_commander_org     str|None   소속/직책
    work_commander_contact str|None   연락처
    work_commander_duties  str|None   담당 작업
    work_commander_educated str|None  작업계획서 교육 여부

    pre_check_items        list[dict] 작업 전 점검표 (MAX_CHECKS=10)
        check_item         str|None   점검 항목
        result             str|None   점검 결과
        note               str|None   비고

    education_done         str|None   특별교육 실시 여부
    tbm_done               str|None   TBM 실시 여부

    sign_date              str|None   서명 날짜

Output — xlsx bytes (in-memory). 파일 저장은 호출자 책임.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME     = "거푸집동바리작업계획서"
SHEET_HEADING  = "거푸집·동바리 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제38조·제331조에 따른 작업계획서"

MAX_HAZARD  = 10
MAX_CHECKS  = 10
TOTAL_COLS  = 8

ATTACH_NOTE = (
    "※ 구조검토서 및 조립도 원본 첨부·보관 필요 "
    "(이 서류는 구조계산서·조립도 원본을 대체하지 않음)"
)

_DEFAULT_PRE_CHECKS: List[Dict[str, Any]] = [
    {"check_item": "구조검토서 확인",          "result": "", "note": ""},
    {"check_item": "조립도 확인",              "result": "", "note": ""},
    {"check_item": "자재 손상 여부",           "result": "", "note": ""},
    {"check_item": "동바리 설치상태",          "result": "", "note": ""},
    {"check_item": "수평연결재/가새 설치상태", "result": "", "note": ""},
    {"check_item": "작업발판/통로 상태",       "result": "", "note": ""},
    {"check_item": "개구부 방호 조치",         "result": "", "note": ""},
    {"check_item": "콘크리트 타설 전 승인",    "result": "", "note": ""},
    {"check_item": "해체 전 강도 확인",        "result": "", "note": ""},
    {"check_item": "보호구 착용 확인",         "result": "", "note": ""},
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8, italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 열 너비
# ---------------------------------------------------------------------------
_COL_WIDTHS: Dict[int, int] = {
    1: 14, 2: 12, 3: 12, 4: 12,
    5: 12, 6: 12, 7: 12, 8: 10,
}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int,
              height: int = 20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _section_header(ws, row: int, title: str, height: int = 18) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=height)
    return row + 1


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_basic_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)

    r += 1
    _write_lv(ws, r, "작성일", _v(data, "work_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업기간", _v(data, "work_period"),
              _L2, _V2_START, _V2_END)

    r += 1
    _write_lv(ws, r, "시공사/협력업체", _v(data, "contractor"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)

    r += 1
    _write_lv(ws, r, "작성자", _v(data, "prepared_by"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "검토자", _v(data, "reviewer"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_work_overview(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 개요")

    _write_lv(ws, r, "작업 위치", _v(data, "work_location"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "작업 범위", _v(data, "work_scope"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "거푸집 종류", _v(data, "formwork_type"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "동바리 종류", _v(data, "shoring_type"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "층/구간/타설 부위", _v(data, "floor_section"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1
    _write_lv(ws, r, "작업 단계", _v(data, "work_phases"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=30)

    return r + 1


def _write_preliminary_survey(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "사전조사  (산업안전보건기준에 관한 규칙 제38조)")

    survey_rows = [
        ("지반/바닥 상태",       "survey_ground",          30),
        ("하부 구조물 상태",     "survey_substructure",    30),
        ("개구부/단부 존재",     "survey_opening",         25),
        ("자재 적치 위치",       "survey_material_place",  25),
        ("장비 진입 동선",       "survey_equipment_route", 25),
        ("기상 조건",           "survey_weather",         25),
        ("조명/통로 확보",       "survey_lighting",        25),
    ]
    for label, field, h in survey_rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL,
                    align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, field), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    return r


def _write_structural_review(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "구조검토 및 조립도  (산업안전보건기준에 관한 규칙 제331조)")

    # 구조검토·조립도 작성 여부
    _write_lv(ws, r, "구조검토 실시", _v(data, "structural_review_done"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "구조검토자", _v(data, "structural_reviewer"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "조립도 작성", _v(data, "assembly_drawing_done"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    r += 1

    # 부재 규격
    struct_rows = [
        ("부재 규격",         "member_spec",          30),
        ("설치 간격",         "install_interval",     25),
        ("이음 방법",         "joint_method",         25),
        ("수평연결재/가새 계획", "brace_plan",           30),
        ("받침대/미끄럼 방지", "base_plate_plan",      25),
    ]
    for label, field, h in struct_rows:
        _write_cell(ws, r, _L1, _L1, label,
                    font=_FONT_BOLD, fill=_FILL_LABEL,
                    align=_ALIGN_LABEL, height=h)
        _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                    _v(data, field), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        r += 1

    # 첨부 여부
    _write_lv(ws, r, "구조검토서 첨부", _v(data, "structural_doc_attached"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "조립도 첨부", _v(data, "assembly_drawing_attached"),
              _L2, _V2_START, _V2_END)
    r += 1

    # 원본 첨부 주의 문구
    _write_cell(ws, r, 1, TOTAL_COLS, ATTACH_NOTE,
                font=_FONT_SMALL, fill=_FILL_WARN,
                align=_ALIGN_CENTER, height=20)
    r += 1

    return r


def _write_work_sequence(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "작업 순서 및 방법")
    _write_cell(ws, r, _L1, _L1, "작업 순서",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=60)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "work_sequence"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    return r + 1


def _write_hazard_table(ws, start_row: int,
                        hazard_items: List[Dict[str, Any]]) -> int:
    r = _section_header(ws, start_row, "주요 위험요인 및 안전대책")

    _write_cell(ws, r, 1, 1, "순번",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 4, "주요 위험요인",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "안전대책",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_HAZARD):
        item = hazard_items[i] if i < len(hazard_items) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 4, _v(item, "hazard"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 5, 8, _v(item, "safety_measure"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_safety_measures(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "안전대책 (종합)")
    _write_cell(ws, r, _L1, _L1, "안전대책",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "safety_measures_text"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    return r + 1


def _write_work_commander(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row,
                        "작업지휘자 및 역할  (산업안전보건기준에 관한 규칙 제39조)")

    _write_lv(ws, r, "작업지휘자 성명", _v(data, "work_commander_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "소속/직책", _v(data, "work_commander_org"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "연락처", _v(data, "work_commander_contact"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "교육 여부", _v(data, "work_commander_educated"),
              _L2, _V2_START, _V2_END)
    r += 1
    _write_lv(ws, r, "담당 작업", _v(data, "work_commander_duties"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=30)
    r += 1

    return r


def _write_pre_check_table(ws, start_row: int,
                           pre_checks: List[Dict[str, Any]]) -> int:
    r = _section_header(ws, start_row, "작업 전 점검표")

    _write_cell(ws, r, 1, 1, "순번",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 5, "점검 항목",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 7, "점검 결과",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "비고",
                font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    effective = pre_checks if pre_checks else _DEFAULT_PRE_CHECKS

    for i in range(MAX_CHECKS):
        item = effective[i] if i < len(effective) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, _v(item, "check_item"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 7, _v(item, "result"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, _v(item, "note"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 25
        r += 1

    return r


def _write_education_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "근로자 교육 및 TBM")

    _write_lv(ws, r, "특별교육 실시", _v(data, "education_done"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "TBM 실시", _v(data, "tbm_done"),
              _L2, _V2_START, _V2_END)

    return r + 1


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = _section_header(ws, start_row, "확인 서명")

    _write_cell(ws, r, 1, 2, "작성자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업지휘자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_cell(ws, r, 1, 2, "관리감독자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "안전관리자 (인)",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 8, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_formwork_shoring_workplan_sheet(ws, form_data: Dict[str, Any]) -> None:
    """주어진 worksheet에 거푸집·동바리 작업계획서를 렌더링."""
    data         = form_data or {}
    hazard_items = data.get("hazard_items") or []
    pre_checks   = data.get("pre_check_items") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_work_overview(ws, row, data)
    row = _write_preliminary_survey(ws, row, data)
    row = _write_structural_review(ws, row, data)
    row = _write_work_sequence(ws, row, data)
    row = _write_hazard_table(ws, row, hazard_items)
    row = _write_safety_measures(ws, row, data)
    row = _write_work_commander(ws, row, data)
    row = _write_pre_check_table(ws, row, pre_checks)
    row = _write_education_block(ws, row, data)
    _write_confirmation(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_formwork_shoring_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 거푸집·동바리 작업계획서 xlsx 바이너리를 반환.

    Args:
        form_data: 입력 스키마 준수 dict.
            hazard_items > MAX_HAZARD(10) 시 초과분 무시.
            pre_check_items 미제공 시 기본 10개 자동 적용.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_formwork_shoring_workplan_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
