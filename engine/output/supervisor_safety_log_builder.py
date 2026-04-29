"""
관리감독자 안전보건 업무 일지 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건법 제16조 (관리감독자의 안전보건 업무 의무)
    산업안전보건법 시행령 제15조 (관리감독자의 업무 등)
    산업안전보건기준에 관한 규칙 제35조 (관리감독자의 유해·위험 방지 업무 등)
    산업안전보건기준에 관한 규칙 별표 3 (작업시작 전 점검사항)

분류: PRACTICAL — 공식 별지 제출서식 아님 / 관리감독자 업무 수행·확인 기록 보조서식

Required form_data keys:
    site_name           str  현장명
    log_date            str  작성일
    supervisor_name     str  관리감독자 성명
    department          str  소속

Optional form_data keys:
    project_name                str  공사명
    day_of_week                 str  요일
    work_area                   str  작업구역
    position                    str  직책
    contact                     str  연락처
    writer_name                 str  작성자
    reviewer_name               str  검토자
    approver_name               str  승인자
    -- 섹션2: 금일 관리감독 대상 작업 --
    work_items                  list[dict]  작업 목록 (max 5)
        work_type               str  작업공종
        work_content            str  작업내용
        work_place              str  작업장소
        worker_count            int  작업인원
        subcontractor           str  협력업체
        main_equipment          str  주요 장비
        high_risk               str  고위험 작업 여부
        work_plan_needed        str  작업계획서 필요 여부
        work_permit_needed      str  작업허가서 필요 여부
        remarks                 str  특이사항
    -- 섹션3: 기계·기구·설비 안전보건 점검 --
    equipment_checks            list[dict]  점검 목록 (max 5)
        check_target            str  점검 대상
        check_time              str  점검 시간
        abnormal                str  이상 유무
        problem_found           str  발견 문제
        immediate_action        str  즉시 조치
        inspector               str  담당자
        completed               str  완료 여부
        recheck_result          str  재점검 결과
    -- 섹션4: 보호구·방호장치 점검 및 교육지도 --
    ppe_wear_checked            str  보호구 착용 확인
    ppe_condition_checked       str  보호구 상태 확인
    guard_installed_checked     str  방호장치 설치 확인
    guard_working_checked       str  방호장치 정상 작동 여부
    ppe_violation_guidance      str  미착용자 지도 내용
    edu_guidance_content        str  교육·지도 내용
    improvement_needed          str  개선 필요사항
    -- 섹션5: 작업장 정리정돈 및 통로 확보 --
    passage_status              str  통로 확보 상태
    material_storage_status     str  자재 적치 상태
    opening_area_status         str  개구부 주변 상태
    fall_risk_zone_status       str  추락·낙하 위험 구역 상태
    lighting_status             str  작업장 조명
    emergency_exit_secured      str  비상통로 확보
    housekeeping_order          str  정리정돈 지시사항
    housekeeping_done           str  조치 완료 여부
    -- 섹션6: 작업시작 전 점검사항(별표 3 연계) --
    pretask_check_items         list[dict]  별표 3 점검 항목 (max 5)
        target_work             str  점검 대상 작업
        check_item              str  점검 항목
        abnormal                str  이상 유무
        action_taken            str  이상 시 조치
    pretask_applicable          str  별표 3 해당 작업 여부
    pretask_conducted           str  작업시작 전 점검 실시 여부
    work_start_approved         str  작업 개시 승인 여부
    work_stopped                str  작업 중지 여부
    -- 섹션7: 근로자 교육·지도 및 TBM 연계 --
    tbm_conducted               str  TBM 실시 여부
    work_method_guidance        str  작업방법 지도
    risk_factor_communicated    str  위험요인 전달
    new_worker_guidance         str  신규 근로자 지도
    foreign_worker_communicated str  외국인 근로자 전달 여부
    unsafe_behavior_guidance    str  불안전 행동 지도
    edu_material_distributed    str  교육자료 배포 여부
    signature_sheet_attached    str  서명부 별첨 여부
    -- 섹션8: 산업재해·아차사고·응급조치 대응 --
    accident_occurred           str  산업재해 발생 여부
    near_miss_occurred          str  아차사고 발생 여부
    first_aid_needed            str  응급조치 필요 여부
    reported                    str  보고 여부
    first_aid_content           str  응급조치 내용
    em002_linked                str  EM-002 연계 여부
    em006_linked                str  EM-006 연계 여부
    followup_action             str  후속 조치
    -- 섹션9: 개선조치 및 이행관리 --
    improvement_actions         list[dict]  개선조치 목록 (max 8)
        seq                     int  번호
        issue                   str  지적사항
        action                  str  개선조치
        assignee                str  담당자
        due_date                str  완료 예정일
        completed_date          str  완료일
        status                  str  이행상태
        evidence_exists         str  증빙자료 여부
        confirmed_by            str  확인자
        incomplete_reason       str  미완료 사유
    -- 섹션10: 확인 및 승인 --
    confirm_supervisor          str  관리감독자
    confirm_safety_manager      str  안전관리자
    confirm_site_manager        str  현장소장
    confirm_contractor          str  협력업체 책임자
    confirm_date                str  확인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "관리감독자업무일지"
SHEET_HEADING = "관리감독자 안전보건 업무 일지"
DOC_ID = "DL-002"

TOTAL_COLS = 10
MAX_WORK_ROWS = 5
MAX_EQUIP_ROWS = 5
MAX_PRETASK_ROWS = 5
MAX_IMPROVE_ROWS = 8

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL = Font(name="맑은 고딕", size=9)

_FILL_LABEL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE = PatternFill()

_THIN = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LABEL = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 10, 3: 12, 4: 10, 5: 10,
    6: 12, 7: 12, 8: 12, 9: 12, 10: 10,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value = "" if value is None else value
    cell.font = font or _FONT_DEFAULT
    cell.fill = fill or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vc1: int, vc2: int, height: int = 20) -> None:
    _cell(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vc1, vc2, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, f"▶ {title}",
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_LEFT, height=20)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ---------------------------------------------------------------------------
# 섹션 작성 함수
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    notice_lines = [
        "공식 제출 서식 아님 / 관리감독자 업무 수행 및 확인 기록 보조서식",
        "산업안전보건법 제16조 및 시행령 제15조 관리감독자 업무와 연계",
        "기계·기구·설비 점검, 보호구·방호장치 점검, 산업재해 보고 및 응급조치, 작업장 정리정돈 확인을 기록",
        "작업시작 전 점검사항은 산업안전보건기준에 관한 규칙 별표 3과 연계 / DL-001 안전관리 일지와 별도 관리",
        "개인정보·민감정보 최소 기재 / 사진·영상 등 증빙자료는 별도 보관",
    ]
    for line in notice_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
        c = ws.cell(row=row, column=1, value=line)
        c.font = _FONT_SMALL
        c.fill = _FILL_NOTICE
        c.alignment = _ALIGN_CENTER
        _border_rect(ws, row, 1, row, TOTAL_COLS)
        ws.row_dimensions[row].height = 14
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    doc_id_cell = ws.cell(row=row, column=1, value=f"(문서ID: {DOC_ID})")
    doc_id_cell.font = _FONT_SMALL
    doc_id_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 12
    return row + 1


def _write_doc_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "문서 기본정보")
    _lv(ws, row, "공사명", _v(data, "project_name"), 2, 3, 5)
    _lv(ws, row, "현장명", _v(data, "site_name"), 6, 7, 9)
    _cell(ws, row, 10, 10, "", font=_FONT_DEFAULT)
    ws.row_dimensions[row].height = 20
    row += 1
    _lv(ws, row, "작성일", _v(data, "log_date"), 2, 3, 4)
    _lv(ws, row, "요일", _v(data, "day_of_week"), 5, 6, 6)
    _lv(ws, row, "작업구역", _v(data, "work_area"), 7, 8, 10)
    row += 1
    _lv(ws, row, "관리감독자 성명", _v(data, "supervisor_name"), 2, 3, 4)
    _lv(ws, row, "소속", _v(data, "department"), 5, 6, 6)
    _lv(ws, row, "직책", _v(data, "position"), 7, 8, 8)
    _lv(ws, row, "연락처", _v(data, "contact"), 9, 10, 10)
    row += 1
    _lv(ws, row, "작성자", _v(data, "writer_name"), 2, 3, 4)
    _lv(ws, row, "검토자", _v(data, "reviewer_name"), 5, 6, 7)
    _lv(ws, row, "승인자", _v(data, "approver_name"), 8, 9, 10)
    row += 1
    return row


def _write_work_items(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "금일 관리감독 대상 작업")
    headers = ["작업공종", "작업내용", "작업장소", "작업인원", "협력업체",
               "주요 장비", "고위험 작업", "작업계획서", "작업허가서", "특이사항"]
    cols = [(1, 1), (2, 3), (4, 4), (5, 5), (6, 6),
            (7, 7), (8, 8), (9, 9), (10, 10), (10, 10)]
    # 헤더 행
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "작업내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 4, "작업장소", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 5, 5, "작업인원", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 6, 6, "협력업체", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 7, 7, "주요 장비", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "고위험 작업", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "작업계획서", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "작업허가서", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("work_items", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_WORK_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "work_content"), height=20)
        _cell(ws, row, 4, 4, _v(item, "work_place"), height=20)
        _cell(ws, row, 5, 5, _v(item, "worker_count"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 6, 6, _v(item, "subcontractor"), height=20)
        _cell(ws, row, 7, 7, _v(item, "main_equipment"), height=20)
        _cell(ws, row, 8, 8, _v(item, "high_risk"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 9, 9, _v(item, "work_plan_needed"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "work_permit_needed"), align=_ALIGN_CENTER, height=20)
        row += 1
    return row


def _write_equipment_checks(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "기계·기구·설비 안전보건 점검")
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "점검 대상", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 4, "점검 시간", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 5, 5, "이상 유무", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 6, 7, "발견 문제/즉시 조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "완료 여부", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "재점검", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("equipment_checks", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_EQUIP_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "check_target"), height=20)
        _cell(ws, row, 4, 4, _v(item, "check_time"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 5, 5, _v(item, "abnormal"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 6, 7, _v(item, "problem_found"), height=20)
        _cell(ws, row, 8, 8, _v(item, "inspector"), height=20)
        _cell(ws, row, 9, 9, _v(item, "completed"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "recheck_result"), align=_ALIGN_CENTER, height=20)
        row += 1
    return row


def _write_ppe_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "보호구·방호장치 점검 및 교육지도")
    _lv(ws, row, "보호구 착용 확인", _v(data, "ppe_wear_checked"), 2, 3, 5)
    _lv(ws, row, "보호구 상태 확인", _v(data, "ppe_condition_checked"), 6, 7, 9)
    _cell(ws, row, 10, 10, "")
    row += 1
    _lv(ws, row, "방호장치 설치 확인", _v(data, "guard_installed_checked"), 2, 3, 5)
    _lv(ws, row, "방호장치 작동 여부", _v(data, "guard_working_checked"), 6, 7, 9)
    _cell(ws, row, 10, 10, "")
    row += 1
    _lv(ws, row, "미착용자 지도", _v(data, "ppe_violation_guidance"), 2, 3, 10)
    row += 1
    _lv(ws, row, "교육·지도 내용", _v(data, "edu_guidance_content"), 2, 3, 10)
    row += 1
    _lv(ws, row, "개선 필요사항", _v(data, "improvement_needed"), 2, 3, 10)
    row += 1
    return row


def _write_housekeeping(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "작업장 정리정돈 및 통로 확보")
    _lv(ws, row, "통로 확보 상태", _v(data, "passage_status"), 2, 3, 5)
    _lv(ws, row, "자재 적치 상태", _v(data, "material_storage_status"), 6, 7, 10)
    row += 1
    _lv(ws, row, "개구부 주변 상태", _v(data, "opening_area_status"), 2, 3, 5)
    _lv(ws, row, "추락·낙하 위험 구역", _v(data, "fall_risk_zone_status"), 6, 7, 10)
    row += 1
    _lv(ws, row, "작업장 조명", _v(data, "lighting_status"), 2, 3, 5)
    _lv(ws, row, "비상통로 확보", _v(data, "emergency_exit_secured"), 6, 7, 10)
    row += 1
    _lv(ws, row, "정리정돈 지시사항", _v(data, "housekeeping_order"), 2, 3, 8)
    _lv(ws, row, "조치 완료", _v(data, "housekeeping_done"), 9, 10, 10)
    row += 1
    return row


def _write_pretask_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "작업시작 전 점검사항 (산업안전보건기준에 관한 규칙 별표 3 연계)")
    _lv(ws, row, "별표3 해당 작업 여부", _v(data, "pretask_applicable"), 2, 3, 5)
    _lv(ws, row, "작업시작 전 점검 실시", _v(data, "pretask_conducted"), 6, 7, 10)
    row += 1

    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "점검 대상 작업", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 6, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 7, 7, "이상 유무", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 10, "이상 시 조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("pretask_check_items", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_PRETASK_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "target_work"), height=20)
        _cell(ws, row, 4, 6, _v(item, "check_item"), height=20)
        _cell(ws, row, 7, 7, _v(item, "abnormal"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 8, 10, _v(item, "action_taken"), height=20)
        row += 1

    _lv(ws, row, "작업 개시 승인", _v(data, "work_start_approved"), 2, 3, 5)
    _lv(ws, row, "작업 중지 여부", _v(data, "work_stopped"), 6, 7, 10)
    row += 1
    return row


def _write_edu_tbm(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "근로자 교육·지도 및 TBM 연계")
    _lv(ws, row, "TBM 실시 여부", _v(data, "tbm_conducted"), 2, 3, 5)
    _lv(ws, row, "작업방법 지도", _v(data, "work_method_guidance"), 6, 7, 10)
    row += 1
    _lv(ws, row, "위험요인 전달", _v(data, "risk_factor_communicated"), 2, 3, 5)
    _lv(ws, row, "신규 근로자 지도", _v(data, "new_worker_guidance"), 6, 7, 10)
    row += 1
    _lv(ws, row, "외국인 근로자 전달", _v(data, "foreign_worker_communicated"), 2, 3, 5)
    _lv(ws, row, "불안전 행동 지도", _v(data, "unsafe_behavior_guidance"), 6, 7, 10)
    row += 1
    _lv(ws, row, "교육자료 배포", _v(data, "edu_material_distributed"), 2, 3, 5)
    _lv(ws, row, "서명부 별첨", _v(data, "signature_sheet_attached"), 6, 7, 10)
    row += 1
    return row


def _write_accident_response(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "산업재해·아차사고·응급조치 대응")
    _lv(ws, row, "산업재해 발생 여부", _v(data, "accident_occurred"), 2, 3, 5)
    _lv(ws, row, "아차사고 발생 여부", _v(data, "near_miss_occurred"), 6, 7, 10)
    row += 1
    _lv(ws, row, "응급조치 필요 여부", _v(data, "first_aid_needed"), 2, 3, 5)
    _lv(ws, row, "보고 여부", _v(data, "reported"), 6, 7, 10)
    row += 1
    _lv(ws, row, "응급조치 내용", _v(data, "first_aid_content"), 2, 3, 10)
    row += 1
    _lv(ws, row, "EM-002 연계 여부", _v(data, "em002_linked"), 2, 3, 5)
    _lv(ws, row, "EM-006 연계 여부", _v(data, "em006_linked"), 6, 7, 10)
    row += 1
    _lv(ws, row, "후속 조치", _v(data, "followup_action"), 2, 3, 10)
    row += 1
    return row


def _write_improvement_actions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "개선조치 및 이행관리")
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "지적사항", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 5, "개선조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 6, 6, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 7, 7, "완료 예정일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "완료일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "이행상태", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "확인자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("improvement_actions", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_IMPROVE_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "issue"), height=20)
        _cell(ws, row, 4, 5, _v(item, "action"), height=20)
        _cell(ws, row, 6, 6, _v(item, "assignee"), height=20)
        _cell(ws, row, 7, 7, _v(item, "due_date"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 8, 8, _v(item, "completed_date"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 9, 9, _v(item, "status"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "confirmed_by"), height=20)
        row += 1
    return row


def _write_confirmation(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "확인 및 승인")
    _lv(ws, row, "관리감독자", _v(data, "confirm_supervisor"), 2, 3, 4)
    _lv(ws, row, "안전관리자", _v(data, "confirm_safety_manager"), 5, 6, 7)
    _lv(ws, row, "현장소장", _v(data, "confirm_site_manager"), 8, 9, 10)
    row += 1
    _lv(ws, row, "협력업체 책임자", _v(data, "confirm_contractor"), 2, 3, 5)
    _lv(ws, row, "작성일", _v(data, "log_date"), 6, 7, 7)
    _lv(ws, row, "확인일", _v(data, "confirm_date"), 8, 9, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_supervisor_safety_log(form_data: Dict[str, Any]) -> bytes:
    """관리감독자 안전보건 업무 일지 Excel 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    _apply_print_settings(ws)

    r = _write_title(ws, 1)
    r += 1
    r = _write_doc_info(ws, r, form_data)
    r += 1
    r = _write_work_items(ws, r, form_data)
    r += 1
    r = _write_equipment_checks(ws, r, form_data)
    r += 1
    r = _write_ppe_check(ws, r, form_data)
    r += 1
    r = _write_housekeeping(ws, r, form_data)
    r += 1
    r = _write_pretask_check(ws, r, form_data)
    r += 1
    r = _write_edu_tbm(ws, r, form_data)
    r += 1
    r = _write_accident_response(ws, r, form_data)
    r += 1
    r = _write_improvement_actions(ws, r, form_data)
    r += 1
    r = _write_confirmation(ws, r, form_data)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
