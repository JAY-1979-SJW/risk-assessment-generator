"""
비계 설치 점검표 — Excel 출력 모듈 (v1.0)  [CL-001]

법적 근거:
    산업안전보건기준에 관한 규칙 제57조 — 비계 등의 조립·해체 및 변경
    (PARTIAL_VERIFIED: 제57조 조문 확인. 이하 작업발판·안전난간·낙하물방지
     관련 조항은 NEEDS_VERIFICATION — 원문 API 수집 후 확정 예정)

범위:
    본 서식은 비계 설치·사용 상태 확인 전용입니다.
    거푸집동바리 점검은 CL-002 별도 서식으로 관리합니다.

Required form_data keys:
    check_date    str  점검 일자
    work_location str  작업 장소 / 비계 설치 위치
    checker_name  str  점검자 성명

Optional form_data keys:
    site_name           str  사업장명
    project_name        str  공사명
    work_date           str  작업 예정일 또는 기간
    supervisor_name     str  관리감독자 성명
    scaffold_type       str  비계 종류 (강관비계/강관틀비계/시스템비계/달비계/이동식비계 등)
    scaffold_height     str  설치 높이 (m)
    scaffold_length     str  설치 연장 (m)
    scaffold_location   str  세부 설치 위치
    scaffold_work_type  str  작업 내용 (외부 마감/내부 설치 등)
    pre_install_items   list[dict]  섹션3 항목 override (item, result, note)
    structure_items     list[dict]  섹션4 항목 override
    workboard_items     list[dict]  섹션5 항목 override
    railing_items       list[dict]  섹션6 항목 override
    assembly_items      list[dict]  섹션7 항목 override
    usage_items         list[dict]  섹션8 항목 override
    nonconformance_items list[dict]  부적합 사항 (no, content, location, action, deadline, completed)
    inspector_sign      str  점검자 서명
    supervisor_sign     str  관리감독자 서명
    manager_sign        str  현장소장 서명
    sign_date           str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "비계설치점검표"
SHEET_HEADING = "비계 설치 점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제57조 이하 비계 관련 조항에 따른 안전점검"

NOTICE_CL002 = (
    "본 점검표는 비계 설치·사용 상태 확인용이며, "
    "거푸집동바리 점검은 CL-002 별도 서식으로 관리한다."
)
NOTICE_LAW = "법령 조항 및 기준은 현행 법령 원문 확인 후 현장에 적용한다."
NOTICE_NONCONFORM = "점검 결과 부적합 사항은 사용 전 시정 완료 후 확인 서명을 받아야 한다."

MAX_NONCONFORM = 5
TOTAL_COLS     = 9

# ──────────────────────────────────────────────
# 기본 점검 항목 (섹션별, 미제공 시 자동 적용)
# ──────────────────────────────────────────────

_PRE_INSTALL_DEFAULTS: List[Dict[str, str]] = [
    {"item": "구조검토서 또는 조립도 작성 여부",              "result": "", "note": ""},
    {"item": "지반 침하·균열 및 지지 상태 확인 여부",         "result": "", "note": ""},
    {"item": "사용 자재 규격·수량 확인 여부",                 "result": "", "note": ""},
    {"item": "작업 전 근로자 안전교육 실시 여부",             "result": "", "note": ""},
    {"item": "작업구역 출입통제 및 안전표지 설치 여부",        "result": "", "note": ""},
]

_STRUCTURE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "비계 기둥 수직도 적합 여부",                    "result": "", "note": ""},
    {"item": "벽이음 설치 간격 기준 충족 여부",               "result": "", "note": ""},
    {"item": "가새 설치 상태 (X형 또는 V형 적정 여부)",        "result": "", "note": ""},
    {"item": "받침철물(베이스플레이트) 및 깔판 설치 여부",     "result": "", "note": ""},
    {"item": "부재 손상·변형·부식 없음 여부",                 "result": "", "note": ""},
]

_WORKBOARD_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업발판 폭 40cm 이상 여부",                    "result": "", "note": ""},
    {"item": "발판 틈새 3cm 이하 여부",                       "result": "", "note": ""},
    {"item": "발판 고정 상태 (탈락·움직임 없음)",              "result": "", "note": ""},
    {"item": "적재하중 표시 여부",                            "result": "", "note": ""},
    {"item": "승강통로 또는 사다리 설치 상태 적정 여부",       "result": "", "note": ""},
]

_RAILING_DEFAULTS: List[Dict[str, str]] = [
    {"item": "안전난간대 설치 여부 (높이 기준 충족)",          "result": "", "note": ""},
    {"item": "중간 난간대 설치 여부",                         "result": "", "note": ""},
    {"item": "발끝막이판 설치 여부",                          "result": "", "note": ""},
    {"item": "낙하물 방지망 설치 여부",                       "result": "", "note": ""},
    {"item": "방호선반 설치 여부 (해당 시)",                  "result": "", "note": ""},
]

_ASSEMBLY_DEFAULTS: List[Dict[str, str]] = [
    {"item": "조립·해체·변경 작업 전 관리감독자 지휘 여부",    "result": "", "note": ""},
    {"item": "강풍·강우·강설 등 악천후 시 작업 중지 여부",    "result": "", "note": ""},
    {"item": "안전대 부착설비 설치 및 착용 여부",             "result": "", "note": ""},
    {"item": "작업구역 하부 출입통제 여부",                   "result": "", "note": ""},
    {"item": "자재 낙하 방지 조치 여부",                     "result": "", "note": ""},
]

_USAGE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업 시작 전 사전 점검 실시 여부",              "result": "", "note": ""},
    {"item": "악천후 후 점검 실시 여부",                     "result": "", "note": ""},
    {"item": "이상 발견 시 보수 완료 후 사용 여부",           "result": "", "note": ""},
    {"item": "비계 허용 하중 준수 여부 (초과 사용 없음)",     "result": "", "note": ""},
    {"item": "무단 개조·변경 없음 여부",                     "result": "", "note": ""},
]

# ──────────────────────────────────────────────
# 스타일
# ──────────────────────────────────────────────

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_NOTICE   = Font(name="맑은 고딕", size=8,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER   = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FEF0E7")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# 9열 레이아웃: 1=순번(4), 2~6=항목(24), 7=결과(8), 8~9=비고(20)
_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 7, 3: 7, 4: 7, 5: 7, 6: 7,
    7: 8, 8: 10, 9: 10,
}

# 헤더 블록 열 분할
_L1, _V1S, _V1E = 1, 2, 5
_L2, _V2S, _V2E = 6, 7, 9
_LF, _VFS, _VFE = 1, 2, 9


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _wc(ws, row: int, c1: int, c2: int, value: Any, *,
        font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, c1, row, c2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _wlv(ws, row: int, label: str, value: Any,
         lc: int, vs: int, ve: int) -> None:
    _wc(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, vs, ve, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for idx, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ──────────────────────────────────────────────
# 섹션별 렌더링 함수
# ──────────────────────────────────────────────

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    c = ws.cell(row=row, column=1, value=SHEET_HEADING)
    c.font = _FONT_TITLE; c.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    s = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    s.font = _FONT_SUBTITLE; s.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_notice_block(ws, row: int) -> int:
    """고정 문구 3줄 출력."""
    for txt in (NOTICE_CL002, NOTICE_LAW, NOTICE_NONCONFORM):
        _wc(ws, row, 1, TOTAL_COLS, f"※ {txt}",
            font=_FONT_NOTICE, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
        row += 1
    return row


def _write_section1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    """섹션 1 — 현장 기본정보."""
    _wc(ws, row, 1, TOTAL_COLS, "① 현장 기본정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "사업장명",  _v(data, "site_name"),    _L1, _V1S, _V1E)
    _wlv(ws, row, "공사명",    _v(data, "project_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검 일자", _v(data, "check_date"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "작업 일자", _v(data, "work_date"),    _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검자",    _v(data, "checker_name"),    _L1, _V1S, _V1E)
    _wlv(ws, row, "관리감독자", _v(data, "supervisor_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "작업 장소", _v(data, "work_location"), _LF, _VFS, _VFE)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_section2_scaffold_info(ws, row: int, data: Dict[str, Any]) -> int:
    """섹션 2 — 비계 기본정보."""
    _wc(ws, row, 1, TOTAL_COLS, "② 비계 기본정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "비계 종류", _v(data, "scaffold_type"),     _L1, _V1S, _V1E)
    _wlv(ws, row, "설치 높이", _v(data, "scaffold_height"),   _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "설치 연장", _v(data, "scaffold_length"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "세부 위치", _v(data, "scaffold_location"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "작업 내용", _v(data, "scaffold_work_type"), _LF, _VFS, _VFE)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_checklist_section(ws, row: int, section_no: int, title: str,
                              defaults: List[Dict[str, str]],
                              override: List[Dict[str, Any]]) -> int:
    """범용 체크리스트 섹션 렌더링 (섹션 3~8)."""
    _wc(ws, row, 1, TOTAL_COLS, f"{section_no} {title}",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1

    # 안내
    _wc(ws, row, 1, TOTAL_COLS,
        "※ 결과: 이상없음 → ○ / 이상있음 → × / 해당없음 → -",
        font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_CENTER, height=14)
    row += 1

    # 표 헤더
    _wc(ws, row, 1, 1, "No",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 6, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 7, 7, "결과",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 9, "비고",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    for i, default in enumerate(defaults):
        provided = override[i] if i < len(override) else {}
        item_text = provided.get("item") or default.get("item", "")
        result    = provided.get("result") or default.get("result", "")
        note      = provided.get("note")   or default.get("note",   "")
        _wc(ws, row, 1, 1, i + 1,     font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 6, item_text,  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 7, 7, result,     font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 8, 9, note,       font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 26
        row += 1

    return row


def _write_section9_nonconformance(ws, row: int,
                                    items: List[Dict[str, Any]]) -> int:
    """섹션 9 — 부적합 및 시정조치."""
    _wc(ws, row, 1, TOTAL_COLS, "⑨ 부적합 및 시정조치",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1

    # 표 헤더
    _wc(ws, row, 1, 1, "No",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 3, "부적합 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 4, 5, "위치",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 6, 7, "시정조치 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 8, "완료기한", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 9, 9, "완료확인", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    for i in range(MAX_NONCONFORM):
        src = items[i] if i < len(items) else {}
        _wc(ws, row, 1, 1, i + 1,                 font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 3, _v(src, "content"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 4, 5, _v(src, "location"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 6, 7, _v(src, "action"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 8, 8, _v(src, "deadline"),   font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 9, 9, _v(src, "completed"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 26
        row += 1

    return row


def _write_section10_signatures(ws, row: int, data: Dict[str, Any]) -> int:
    """섹션 10 — 확인 서명."""
    _wc(ws, row, 1, TOTAL_COLS, "⑩ 확인 서명",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1

    _wc(ws, row, 1, 3, "서명일",
        font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, 4, 9, _v(data, "sign_date"),
        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = 20; row += 1

    # 서명란 3칸 (점검자 / 관리감독자 / 현장소장)
    labels  = ["점검자 서명",   "관리감독자 서명", "현장소장 서명"]
    sign_keys = ["inspector_sign", "supervisor_sign", "manager_sign"]
    col_ranges = [(1, 3), (4, 6), (7, 9)]

    for label, key, (cs, ce) in zip(labels, sign_keys, col_ranges):
        _wc(ws, row, cs, ce, label,
            font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    ws.row_dimensions[row].height = 18; row += 1

    for _, key, (cs, ce) in zip(labels, sign_keys, col_ranges):
        _wc(ws, row, cs, ce, _v(data, key),
            font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 36; row += 1

    return row


# ──────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────

def build_scaffold_installation_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 비계 설치 점검표 xlsx 바이너리를 반환.

    [CL-001]  비계 설치 점검표
    법적 근거: 산업안전보건기준에 관한 규칙 제57조 이하 (PARTIAL_VERIFIED)
    거푸집동바리는 CL-002에서 별도 관리.
    """
    data = form_data or {}

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_notice_block(ws, row)
    row = _write_section1_site_info(ws, row, data)
    row = _write_section2_scaffold_info(ws, row, data)

    row = _write_checklist_section(
        ws, row, "③", "설치 전 확인사항",
        _PRE_INSTALL_DEFAULTS, data.get("pre_install_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "④", "구조·재료 점검",
        _STRUCTURE_DEFAULTS, data.get("structure_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑤", "작업발판·승강설비 점검",
        _WORKBOARD_DEFAULTS, data.get("workboard_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑥", "안전난간·낙하물 방지 점검",
        _RAILING_DEFAULTS, data.get("railing_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑦", "조립·해체·변경 작업 점검",
        _ASSEMBLY_DEFAULTS, data.get("assembly_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑧", "사용 전·사용 중 점검 결과",
        _USAGE_DEFAULTS, data.get("usage_items") or [],
    )
    row = _write_section9_nonconformance(
        ws, row, data.get("nonconformance_items") or [],
    )
    _write_section10_signatures(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
