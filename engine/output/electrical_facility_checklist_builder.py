"""
전기설비 정기 점검표 — Excel 출력 모듈 (v1.0).

법적 근거:
    산업안전보건기준에 관한 규칙 제302조~제305조 (접지, 누전차단기)
    제301조·제309조·제313조 (전기기계기구, 이동전선, 전기작업 일반)
    제319조~제320조 (정전작업, 전로 차단)
    제323조 (절연용 보호구 착용)
    ELEC-001 공통 evidence pack(L2/L4/L5/L6) 재사용

분류: PRACTICAL — 법정 별지 서식 없음.
    전기설비 안전 점검 항목을 체계적으로 기록하는 실무 표준서식.

Required form_data keys:
    site_name        str  현장명
    inspection_date  str  점검일
    inspector        str  점검자

Optional form_data keys:
    project_name         str  공사명
    inspection_no        str  점검 번호
    equipment_name       str  설비명
    voltage              str  전압
    inspection_location  str  점검 위치
    responsible_person   str  담당자
    related_wp_no        str  관련 작업계획서 번호
    related_ptw_no       str  관련 PTW 번호
    # 체크리스트 항목 (list[str] — 해당 항목 키 전달)
    panel_checks          list[str]   분전반 및 차단기 점검 확인 항목 키
    grounding_checks      list[str]   접지 상태 점검 확인 항목 키
    wiring_checks         list[str]   배선 및 이동전선 점검 확인 항목 키
    equipment_checks      list[str]   전기기계기구 및 전동공구 점검 확인 항목 키
    temporary_checks      list[str]   임시전기 및 작업등 점검 확인 항목 키
    hazard_checks         list[str]   감전·화재 위험 점검 확인 항목 키
    ppe_checks            list[str]   보호구 및 측정장비 확인 항목 키
    # 부적합 사항 (list[dict])
    nonconformance_items  list[dict]  content, action, deadline, completed
    # 판정
    verdict               str         적합 / 조건부 적합 / 사용중지
    verdict_condition     str         조건부 적합 조건
    # 서명
    inspector_sign        str         점검자 서명
    supervisor_sign       str         관리감독자 서명
    safety_manager_sign   str         안전관리자 서명
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "전기설비점검표"
SHEET_HEADING = "전기설비 정기 점검표"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제302조~제305조·제319조·제323조에 따른 "
    "전기설비 정기 점검 기록서"
)

MAX_NC     = 5
TOTAL_COLS = 8

NOTICE_PRACTICAL = (
    "본 점검표는 법정 별지 서식이 아닌 실무 표준서식으로, "
    "전기설비 정기 점검 이행 여부를 확인·기록한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)

# ---------------------------------------------------------------------------
# 점검 항목 정의 (키: 표시 문구 쌍)
# ---------------------------------------------------------------------------

PANEL_ITEMS: List[tuple[str, str]] = [
    ("panel_lock",          "분전반 잠금 상태 양호 (제301조)"),
    ("panel_nameplate",     "차단기 명판 부착 및 판독 가능"),
    ("overcurrent_breaker", "과전류 차단장치 설치 및 정상 작동"),
    ("elcb_installed",      "누전차단기 설치 (제303조)"),
    ("elcb_test_button",    "누전차단기 시험 버튼 작동 확인 (제303조)"),
    ("no_exposed_live",     "노출 충전부 없음 (방호판·커버 설치)"),
]

GROUNDING_ITEMS: List[tuple[str, str]] = [
    ("ground_connection",   "접지선 연결 상태 양호 (제302조)"),
    ("ground_terminal",     "접지 단자 손상 없음"),
    ("ground_resistance",   "접지 저항 측정 여부 확인"),
    ("no_ground_missing",   "접지 누락 없음 (제302조)"),
]

WIRING_ITEMS: List[tuple[str, str]] = [
    ("cable_sheath_ok",     "이동전선 피복 손상 없음 (제313조)"),
    ("temp_wiring_protect", "임시 배선 보호관·트레이 사용"),
    ("no_wet_risk",         "물기·침수 위험 구간 배선 없음"),
    ("crossing_protect",    "통행로 횡단 배선 보호 (제313조)"),
    ("no_octopus_wiring",   "문어발식 배선 없음"),
]

EQUIPMENT_ITEMS: List[tuple[str, str]] = [
    ("equip_housing_ok",    "전기기계기구 외함 손상 없음 (제309조)"),
    ("plug_ok",             "플러그 손상·접촉 불량 없음"),
    ("grounded_plug",       "접지형 플러그 사용 (제302조)"),
    ("equip_cable_ok",      "기계기구 전선 피복 양호"),
    ("equip_elcb",          "이동형 기계기구 — 누전차단기 연결 (제304조)"),
]

TEMPORARY_ITEMS: List[tuple[str, str]] = [
    ("lamp_guard",          "작업등 보호망 설치"),
    ("lamp_waterproof",     "습윤 장소 방수형 등기구 사용"),
    ("lamp_support_ok",     "전등 지지 상태 양호 (흔들림 없음)"),
    ("lamp_no_combustible", "가연물 근접 없음"),
]

HAZARD_ITEMS: List[tuple[str, str]] = [
    ("live_parts_guarded",  "충전부 방호판·방호망 설치 (제301조)"),
    ("insulation_ok",       "절연 상태 양호 — 육안 확인"),
    ("no_overheat_trace",   "과열 흔적 없음"),
    ("no_spark_trace",      "스파크·탄화 흔적 없음"),
    ("no_dust_moisture",    "분진·수분 노출 위험 없음"),
]

PPE_ITEMS: List[tuple[str, str]] = [
    ("insulated_gloves",    "절연장갑 구비·상태 양호 (제323조)"),
    ("insulated_shoes",     "절연화 구비·상태 양호 (제323조)"),
    ("voltage_tester",      "검전기 구비·교정 유효"),
    ("insulation_tester",   "절연저항계(메거) 구비 (해당 시)"),
    ("insulated_tools",     "절연공구 사용 (제323조)"),
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SECTION3 = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=fill or _FILL_NOTICE,
                align=_ALIGN_LEFT, height=28)
    return row + 1


def _write_checklist_items(ws, row: int, items: List[tuple[str, str]],
                           checked_keys: set) -> int:
    """체크리스트 항목 행 작성. 키가 checked_keys에 있으면 ■, 없으면 □."""
    for key, label in items:
        mark = "■" if key in checked_keys else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {label}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 32
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 18
    row += 1
    row = _write_notice(ws, row, NOTICE_PRACTICAL)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "점검 번호", _v(data, "inspection_no") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_inspection_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 점검 대상 정보")
    _write_lv(ws, row, "점검일",
              _v(data, "inspection_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "점검 위치",
              _v(data, "inspection_location"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "설비명",
              _v(data, "equipment_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "전압",
              _v(data, "voltage") or "__ V / __ kV",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "담당자",
              _v(data, "responsible_person"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "점검자",
              _v(data, "inspector"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관련 작업계획서",
              _v(data, "related_wp_no") or "",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "관련 PTW 번호",
              _v(data, "related_ptw_no") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    return row


def _write_s3_panel(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 분전반 및 차단기 점검  (제301조·제303조)")
    checked = set(data.get("panel_checks") or [])
    return _write_checklist_items(ws, row, PANEL_ITEMS, checked)


def _write_s4_grounding(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 접지 상태 점검  (제302조)")
    checked = set(data.get("grounding_checks") or [])
    return _write_checklist_items(ws, row, GROUNDING_ITEMS, checked)


def _write_s5_wiring(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 배선 및 이동전선 점검  (제313조)")
    checked = set(data.get("wiring_checks") or [])
    return _write_checklist_items(ws, row, WIRING_ITEMS, checked)


def _write_s6_equipment(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 전기기계기구 및 전동공구 점검  (제309조·제304조)")
    checked = set(data.get("equipment_checks") or [])
    return _write_checklist_items(ws, row, EQUIPMENT_ITEMS, checked)


def _write_s7_temporary(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 임시전기 및 작업등 점검")
    checked = set(data.get("temporary_checks") or [])
    return _write_checklist_items(ws, row, TEMPORARY_ITEMS, checked)


def _write_s8_hazard(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 감전·화재 위험 점검", fill=_FILL_WARN)
    checked = set(data.get("hazard_checks") or [])
    return _write_checklist_items(ws, row, HAZARD_ITEMS, checked)


def _write_s9_ppe(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 보호구 및 측정장비 확인  (제323조)")
    checked = set(data.get("ppe_checks") or [])
    return _write_checklist_items(ws, row, PPE_ITEMS, checked)


def _write_s10_nonconformance(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 부적합 사항 및 시정조치")
    _write_cell(ws, row, 1, 3, "부적합 내용", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 4, 5, "시정조치",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 7, "완료 기한",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "완료",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    items = (data.get("nonconformance_items") or [])[:MAX_NC]
    for i in range(MAX_NC):
        item = items[i] if i < len(items) else {}
        _write_cell(ws, row, 1, 3, _v(item, "content"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=22)
        _write_cell(ws, row, 4, 5, _v(item, "action"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 7, _v(item, "deadline"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, _v(item, "completed"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        row += 1
    return row


def _write_s11_verdict(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 점검 결과 판정", fill=_FILL_SECTION2)
    verdict = _v(data, "verdict") or ""
    marks = {
        "적합":       "■" if verdict == "적합"       else "□",
        "조건부 적합": "■" if verdict == "조건부 적합" else "□",
        "사용중지":   "■" if verdict == "사용중지"   else "□",
    }
    verdict_text = (
        f"{marks['적합']} 적합    "
        f"{marks['조건부 적합']} 조건부 적합    "
        f"{marks['사용중지']} 사용중지"
    )
    _write_cell(ws, row, 1, TOTAL_COLS, verdict_text,
                font=_FONT_BOLD, align=_ALIGN_CENTER, height=24)
    row += 1
    _write_lv(ws, row, "조건부 조건",
              _v(data, "verdict_condition") or "(조건부 적합 시 기재)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    return row


def _write_s12_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 확인 서명", fill=_FILL_SECTION3)
    _write_lv(ws, row, "점검자 (서명)",
              _v(data, "inspector_sign") or _v(data, "inspector"),
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "관리감독자 (서명)",
              _v(data, "supervisor_sign"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "안전관리자 (서명)",
              _v(data, "safety_manager_sign"),
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_electrical_facility_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 전기설비 정기 점검표 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_inspection_info(ws, row, data)
    row = _write_s3_panel(ws, row, data)
    row = _write_s4_grounding(ws, row, data)
    row = _write_s5_wiring(ws, row, data)
    row = _write_s6_equipment(ws, row, data)
    row = _write_s7_temporary(ws, row, data)
    row = _write_s8_hazard(ws, row, data)
    row = _write_s9_ppe(ws, row, data)
    row = _write_s10_nonconformance(ws, row, data)
    row = _write_s11_verdict(ws, row, data)
    _write_s12_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
