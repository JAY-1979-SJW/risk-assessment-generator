"""
건설 장비 반입 신청서 — Excel 출력 모듈 (v2).

법적 근거 / 실무 근거:
    원청 안전 관리 규정 기반 — 법정 별지 서식 없음
    산업안전보건법 제80조 (유해·위험기계 사용 시 안전조치)
    건설기계관리법 제13조 (정기검사 의무)
    자동차손해배상보장법 (보험 가입 의무)

분류: PRACTICAL — 공식 별지 제출서식 아님
      건설 장비 반입 전 안전서류 확인 보조서식
      보험증권·정기검사증·운전자 자격·작업계획서 확인 후 반입 승인

역할 분리:
    PPE-002(본 서식): 장비 반입 신청, 서류 확인, 반입 승인
    PPE-003: 장비 보험·정기검사증 확인서 (서류 상세)
    CL-003: 장비 일일 사전 점검표 (매일 사용 전 점검)
    EQ 장비 사용계획서: 장비 사용 계획 상세 기술

Required form_data keys:
    site_name       str  현장명
    equipment_name  str  장비명
    apply_date      str  신청일

Optional form_data keys:
    -- 섹션1: 문서 기본정보 --
    project_name            str  공사명
    contractor              str  원청업체
    subcontractor           str  반입 협력업체
    applicant               str  신청자 성명
    applicant_position      str  신청자 직위
    manager                 str  관리책임자
    prepared_date           str  작성일
    -- 섹션2: 장비 기본정보 --
    equipment_type          str  장비 종류
    equipment_model         str  장비 모델명
    equipment_reg_no        str  등록번호
    manufacturer            str  제조사
    manufacture_year        str  제조연도
    equipment_capacity      str  규격/용량
    equipment_weight        str  중량
    owner_name              str  장비 소유자
    -- 섹션3: 반입 목적 및 작업내용 --
    work_purpose            str  반입 목적
    work_content            str  작업내용
    work_location           str  작업위치
    planned_entry_date      str  반입 예정일
    planned_exit_date       str  반출 예정일
    work_duration           str  작업 기간
    workplan_required       str  작업계획서 필요 여부
    workplan_no             str  작업계획서 번호
    -- 섹션4: 필수 서류 확인 --
    insurance_valid         str  보험증권 유효 여부
    insurance_expiry        str  보험 만료일
    inspection_valid        str  정기검사증 유효 여부
    inspection_expiry       str  검사 만료일
    operator_license_ok     str  운전자 면허 확인
    operator_license_type   str  면허 종류
    safety_inspection_ok    str  안전검사 합격 여부
    safety_cert_no          str  안전검사 합격번호
    -- 섹션5: 운전자 및 작업자 정보 --
    operator_name           str  운전자 성명
    operator_license_no     str  면허번호
    operator_experience     str  경력
    signal_worker_name      str  신호수 성명
    signal_worker_assigned  str  신호수 배치 여부
    banksman_name           str  유도원 성명
    worker_safety_edu       str  작업자 안전교육 이수 여부
    -- 섹션6: 반입 전 안전 확인 --
    access_route_ok         str  반입 동선 확인
    ground_bearing_ok       str  지반 지지력 확인
    overhead_hazard_ok      str  상부 장애물 확인
    underground_hazard_ok   str  지하 매설물 확인
    neighboring_safety_ok   str  인접 구조물 안전 확인
    exclusion_zone_ok       str  작업반경 통제 여부
    signal_system_ok        str  신호체계 확인
    ppe_check_ok            str  운전자 보호구 확인
    -- 섹션7: 현장 반입 조건 --
    entry_time_from         str  반입 가능 시간 (시작)
    entry_time_to           str  반입 가능 시간 (종료)
    entry_route             str  반입 경로
    parking_location        str  장비 대기 위치
    night_work_allowed      str  야간 작업 허용 여부
    weather_condition       str  기상 조건 제한
    load_limit              str  하중 제한
    speed_limit             str  현장 내 속도 제한
    -- 섹션8: 승인 조건 및 보완사항 --
    approval_status         str  승인 상태 (승인/조건부/반려)
    approval_conditions     str  조건부 승인 조건
    supplementary_items     str  보완 필요 사항
    supplement_deadline     str  보완 기한
    rejection_reason        str  반려 사유
    -- 섹션9: 반입 후 연계 관리 --
    cl003_linked            str  CL-003 일일 점검표 연계 여부
    eq_plan_linked          str  장비 사용계획서 연계 여부
    ppp003_linked           str  PPE-003 보험·검사증 연계 여부
    daily_inspection_due    str  일일 점검 예정일
    related_form_nos        str  관련 서식 번호
    follow_up_action        str  후속 조치
    -- 섹션10: 확인 및 승인 --
    applicant_sign          str  신청자
    safety_manager_name     str  안전관리자
    supervisor_name         str  관리감독자
    site_manager_name       str  현장소장
    confirm_date            str  승인일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "장비반입신청서"
SHEET_HEADING = "건설 장비 반입 신청서"
DOC_ID        = "PPE-002"
TOTAL_COLS    = 10

_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL   = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}

MIN_TABLE_ROWS = 3


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[float] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1, value=value)
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    if height:
        ws.row_dimensions[row].height = height
    for c in range(col1, col2 + 1):
        ws.cell(row=row, column=c).border = _BORDER


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vs: int, ve: int, height: float = 20) -> None:
    _cell(ws, row, lc, lc,  label, font=_FONT_BOLD,   fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vs, ve,  value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, title,
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=22)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7


# ---------------------------------------------------------------------------
# 섹션 구현
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    _cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
          font=_FONT_TITLE, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=36)
    row += 1

    subtitle = (
        "공식 제출 서식 아님 — 건설 장비 반입 전 안전서류 확인 보조서식  |  "
        f"보험증권·정기검사증·운전자 자격·작업계획서 확인 후 반입 승인  ({DOC_ID})"
    )
    _cell(ws, row, 1, TOTAL_COLS, subtitle,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    notice = (
        "반입 후 CL-003 장비 일일 사전 점검표와 연계  |  "
        "작업반경 통제, 신호수 배치, 이동동선 확인 필요  |  "
        "미비 서류 또는 안전조치 미충족 시 반입 보류  |  "
        "개인정보·민감정보 최소 기재"
    )
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")

    _lv(ws, row, "공사명",      _v(data, "project_name"),   1, 2, 5)
    _lv(ws, row, "현장명",      _v(data, "site_name"),      6, 7, 10)
    row += 1
    _lv(ws, row, "원청업체",    _v(data, "contractor"),     1, 2, 5)
    _lv(ws, row, "반입 협력업체", _v(data, "subcontractor"), 6, 7, 10)
    row += 1
    _lv(ws, row, "신청일",      _v(data, "apply_date"),     1, 2, 4)
    _lv(ws, row, "신청자",      _v(data, "applicant"),      5, 6, 7)
    _lv(ws, row, "직위",        _v(data, "applicant_position"), 8, 9, 10)
    row += 1
    _lv(ws, row, "관리책임자",  _v(data, "manager"),        1, 2, 5)
    _lv(ws, row, "작성일",      _v(data, "prepared_date"),  6, 7, 10)
    row += 1

    notice = "개인정보·민감정보 최소 기재 — 성명·소속·직위 외 불필요한 개인정보 기재 금지"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_equipment_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 장비 기본정보")

    _lv(ws, row, "장비명",      _v(data, "equipment_name"),  1, 2, 5)
    _lv(ws, row, "장비 종류",   _v(data, "equipment_type"),  6, 7, 10)
    row += 1
    _lv(ws, row, "모델명",      _v(data, "equipment_model"), 1, 2, 5)
    _lv(ws, row, "등록번호",    _v(data, "equipment_reg_no"), 6, 7, 10)
    row += 1
    _lv(ws, row, "제조사",      _v(data, "manufacturer"),    1, 2, 4)
    _lv(ws, row, "제조연도",    _v(data, "manufacture_year"), 5, 6, 7)
    _lv(ws, row, "규격/용량",   _v(data, "equipment_capacity"), 8, 9, 10)
    row += 1
    _lv(ws, row, "중량",        _v(data, "equipment_weight"), 1, 2, 5)
    _lv(ws, row, "장비 소유자", _v(data, "owner_name"),      6, 7, 10)
    row += 1
    return row


def _write_work_purpose(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 반입 목적 및 작업내용")

    _lv(ws, row, "반입 목적",   _v(data, "work_purpose"),       1, 2, 10, height=24)
    row += 1
    _lv(ws, row, "작업내용",    _v(data, "work_content"),       1, 2, 10, height=24)
    row += 1
    _lv(ws, row, "작업위치",    _v(data, "work_location"),      1, 2, 5)
    _lv(ws, row, "반입 예정일", _v(data, "planned_entry_date"), 6, 7, 10)
    row += 1
    _lv(ws, row, "반출 예정일", _v(data, "planned_exit_date"),  1, 2, 5)
    _lv(ws, row, "작업 기간",   _v(data, "work_duration"),      6, 7, 10)
    row += 1

    notice = "장비 사용계획서 또는 작업계획서 필요 여부 확인"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    _lv(ws, row, "작업계획서 필요",  _v(data, "workplan_required"), 1, 2, 5)
    _lv(ws, row, "작업계획서 번호",  _v(data, "workplan_no"),        6, 7, 10)
    row += 1
    return row


def _write_documents(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 필수 서류 확인")

    headers = ["서류 종류", "확인 여부", "유효기간/번호", "비고"]
    col_spans = [(1, 3), (4, 5), (6, 8), (9, 10)]
    for hdr, (cs, ce) in zip(headers, col_spans):
        _cell(ws, row, cs, ce, hdr,
              font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=20)
    row += 1

    doc_items = [
        ("보험증권",       "insurance_valid",      "insurance_expiry"),
        ("정기검사증",     "inspection_valid",     "inspection_expiry"),
        ("운전자 면허증",  "operator_license_ok",  "operator_license_type"),
        ("안전검사 합격증", "safety_inspection_ok", "safety_cert_no"),
    ]
    for label, valid_key, detail_key in doc_items:
        _cell(ws, row, 1,  3,  label,                        align=_ALIGN_LEFT)
        _cell(ws, row, 4,  5,  _v(data, valid_key),          align=_ALIGN_CENTER)
        _cell(ws, row, 6,  8,  _v(data, detail_key),         align=_ALIGN_LEFT)
        _cell(ws, row, 9, 10,  "",                           align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1

    warn = "미비 서류 또는 안전조치 미충족 시 반입 보류"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {warn}",
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_operator_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 운전자 및 작업자 정보")

    _lv(ws, row, "운전자 성명",   _v(data, "operator_name"),        1, 2, 4)
    _lv(ws, row, "면허번호",      _v(data, "operator_license_no"),  5, 6, 7)
    _lv(ws, row, "경력",          _v(data, "operator_experience"),  8, 9, 10)
    row += 1
    _lv(ws, row, "신호수 성명",   _v(data, "signal_worker_name"),   1, 2, 4)
    _lv(ws, row, "신호수 배치 여부", _v(data, "signal_worker_assigned"), 5, 6, 10)
    row += 1
    _lv(ws, row, "유도원 성명",   _v(data, "banksman_name"),        1, 2, 5)
    _lv(ws, row, "작업자 안전교육 이수", _v(data, "worker_safety_edu"), 6, 7, 10)
    row += 1
    return row


def _write_pre_entry_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 반입 전 안전 확인")

    check_items = [
        ("반입 동선 확인",        "access_route_ok"),
        ("지반 지지력 확인",      "ground_bearing_ok"),
        ("상부 장애물 확인",      "overhead_hazard_ok"),
        ("지하 매설물 확인",      "underground_hazard_ok"),
        ("인접 구조물 안전 확인", "neighboring_safety_ok"),
        ("작업반경 통제 여부",    "exclusion_zone_ok"),
        ("신호체계 확인",         "signal_system_ok"),
        ("운전자 보호구 확인",    "ppe_check_ok"),
    ]
    half = len(check_items) // 2
    for i in range(half):
        ll, lk = check_items[i]
        rl, rk = check_items[i + half]
        _cell(ws, row, 1, 2,  ll,                  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5,  _v(data, lk),        align=_ALIGN_CENTER)
        _cell(ws, row, 6, 7,  rl,                  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10, _v(data, rk),        align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 20
        row += 1

    notice = "작업반경 통제, 신호수 배치, 이동동선 확인 필요"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_entry_conditions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 현장 반입 조건")

    _lv(ws, row, "반입 가능 시간(시작)", _v(data, "entry_time_from"),  1, 3, 5)
    _lv(ws, row, "반입 가능 시간(종료)", _v(data, "entry_time_to"),    6, 8, 10)
    row += 1
    _lv(ws, row, "반입 경로",           _v(data, "entry_route"),      1, 3, 10, height=24)
    row += 1
    _lv(ws, row, "장비 대기 위치",      _v(data, "parking_location"), 1, 3, 5)
    _lv(ws, row, "야간 작업 허용",      _v(data, "night_work_allowed"), 6, 8, 10)
    row += 1
    _lv(ws, row, "기상 조건 제한",      _v(data, "weather_condition"), 1, 3, 5)
    _lv(ws, row, "하중 제한",           _v(data, "load_limit"),        6, 8, 10)
    row += 1
    _lv(ws, row, "현장 내 속도 제한",   _v(data, "speed_limit"),      1, 3, 10)
    row += 1
    return row


def _write_approval_conditions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 승인 조건 및 보완사항")

    _lv(ws, row, "승인 상태",        _v(data, "approval_status"),     1, 3, 10)
    row += 1
    _lv(ws, row, "조건부 승인 조건", _v(data, "approval_conditions"), 1, 3, 10, height=30)
    row += 1
    _lv(ws, row, "보완 필요 사항",   _v(data, "supplementary_items"), 1, 3, 10, height=30)
    row += 1
    _lv(ws, row, "보완 기한",        _v(data, "supplement_deadline"), 1, 3, 5)
    _lv(ws, row, "반려 사유",        _v(data, "rejection_reason"),    6, 8, 10)
    row += 1

    warn = "미비 서류 또는 안전조치 미충족 시 반입 보류"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {warn}",
          font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_linkage(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. 반입 후 연계 관리")

    link_items = [
        ("CL-003 일일 점검표 연계",    "cl003_linked"),
        ("장비 사용계획서 연계",       "eq_plan_linked"),
        ("PPE-003 보험·검사증 연계",   "ppp003_linked"),
        ("일일 점검 예정일",           "daily_inspection_due"),
        ("관련 서식 번호",             "related_form_nos"),
        ("후속 조치",                  "follow_up_action"),
    ]
    half = len(link_items) // 2
    for i in range(half):
        ll, lk = link_items[i]
        rl, rk = link_items[i + half]
        _cell(ws, row, 1, 2,  ll,               font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 5,  _v(data, lk),     align=_ALIGN_LEFT)
        _cell(ws, row, 6, 7,  rl,               font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10, _v(data, rk),     align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1

    notice = "반입 후 CL-003 장비 일일 사전 점검표와 연계"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 확인 및 승인")

    sign_items = [
        ("신청자",      "applicant_sign"),
        ("안전관리자",  "safety_manager_name"),
        ("관리감독자",  "supervisor_name"),
        ("현장소장",    "site_manager_name"),
    ]
    for lbl, key in sign_items:
        _cell(ws, row, 1,  2,  lbl,              font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3,  5,  _v(data, key),   align=_ALIGN_CENTER)
        _cell(ws, row, 6,  7,  "서명",           font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 10,  "",              align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 28
        row += 1

    _lv(ws, row, "승인일", _v(data, "confirm_date"), 1, 2, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_equipment_entry_application(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_equipment_info(ws, row, data)
    row = _write_work_purpose(ws, row, data)
    row = _write_documents(ws, row, data)
    row = _write_operator_info(ws, row, data)
    row = _write_pre_entry_check(ws, row, data)
    row = _write_entry_conditions(ws, row, data)
    row = _write_approval_conditions(ws, row, data)
    row = _write_linkage(ws, row, data)
    row = _write_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
