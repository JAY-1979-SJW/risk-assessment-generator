"""
굴착 작업 허가서 — Excel 출력 모듈 (v1).

법적 근거:
    산업안전보건기준에 관한 규칙 제38조 (사전조사 및 작업계획서)
    제82조~제88조 (굴착작업 안전조치 — 지하매설물 확인, 흙막이, 붕괴방지 등)
    KOSHA GUIDE P-94-2021 안전작업허가지침 (참고)

분류: PRACTICAL — 법정 별지 서식 없음.
    법령 제82조 이하 굴착 안전조치 이행 확인 양식 + KOSHA P-94-2021 참고 자체 표준 서식.
    WP-001 굴착 작업계획서 선행 작성 후 본 허가서 발급.

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자
    work_time        str  작업시간
    work_location    str  작업장소
    trade_name       str  작업공종
    work_content     str  작업내용
    contractor       str  작업업체
    work_supervisor  str  작업책임자

Optional form_data keys:
    project_name             str   공사명
    permit_no                str   관리번호
    excavation_depth         str   굴착 깊이 (m)
    excavation_area          str   굴착 범위
    validity_period          str   허가 유효기간
    permit_issuer            str   작업허가자
    supervisor_name          str   관리감독자
    safety_manager_sign      str   안전관리자
    work_end_confirmer       str   작업 종료 확인자
    final_sign               str   최종 확인 서명
    during_work_issues       str   작업 중 이상 사항
    work_end_time            str   작업 종료 시각
    photo_file_list          str   사진 파일 목록
    # list fields
    pre_check_items          list[str]  굴착 전 확인사항 이행 항목
    risk_factor_items        list[str]  주요 위험요인 확인 항목
    safety_measure_items     list[str]  안전조치 확인 항목
    ppe_items                list[str]  보호구 및 장비 확인 항목
    approval_conditions      list[str]  작업 승인 조건
    stop_conditions          list[str]  작업 중지 조건
    workers                  list[dict] name, job_type — 작업자 명단 (max 10)
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "굴착작업허가서"
SHEET_HEADING = "굴착 작업 허가서"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제38조, 제82조 이하에 따른 "
    "굴착작업 전 안전조치 확인 기록서"
)

MAX_WORKERS = 10
TOTAL_COLS  = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


_PRE_CHECK_DEFAULTS = [
    "지하매설물(가스·전기·통신·상하수도) 위치 도면 확인",
    "지하매설물 관계 기관 협의 및 위치 표시",
    "굴착 구역 주변 구조물·지반 현황 사전조사",
    "작업구역 통제선(바리케이드) 설치",
    "토질 조사 및 굴착 공법 결정",
    "WP-001 굴착 작업계획서 선행 작성 확인",
]

_RISK_FACTOR_DEFAULTS = [
    "붕괴·토사 유실 (절토면 기울기 미준수, 지반 약화)",
    "추락 (굴착면 인접 이동 중 추락)",
    "낙하 (굴착 주변 자재·장비 낙하)",
    "지하매설물 파손 (굴착기 접촉)",
    "장비 협착·충돌 (굴착기·덤프트럭 작업자 접근)",
    "침수·붕괴 (우기·용출수 지반 약화)",
]

_SAFETY_MEASURE_DEFAULTS = [
    "흙막이·버팀대 설치 상태 확인 (굴착 깊이 1.5m 초과 시 의무)",
    "굴착면 구배 준수 여부 확인 (산안규칙 별표 11)",
    "출입통제 및 안전표지 설치",
    "신호수 배치 (장비 후방·측면 사각지대)",
    "장비 유도 및 접근금지 구역 설정",
    "배수·침수 방지 조치 (집수정, 배수펌프)",
    "인접 구조물 계측·모니터링",
]

_PPE_DEFAULTS = [
    "안전모 (충격 흡수형)",
    "안전화 (절연·방수형)",
    "안전조끼 (형광)",
    "방진마스크",
    "굴착기 점검 완료 확인",
    "덤프트럭 후방경보장치 작동 확인",
]

_APPROVAL_CONDITIONS_DEFAULTS = [
    "지하매설물 위치 표시 및 관계 기관 협의 완료",
    "흙막이·버팀대 설치 완료 (해당 시)",
    "신호수 배치 완료",
    "작업구역 통제선 설치 완료",
    "WP-001 굴착 작업계획서 사전 검토 완료",
    "기상 조건 이상 없음",
]

_STOP_CONDITIONS_DEFAULTS = [
    "굴착면 균열·변형·용출수 발생 시",
    "지하매설물 손상 또는 손상 우려 발생 시",
    "강우·강풍 등 기상 악화로 붕괴 위험 증가 시",
    "계측값이 관리기준 초과 시",
    "작업자·장비 안전거리 미확보 시",
    "안전관리자 또는 현장소장 지시 시",
]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_site_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "현장 기본정보",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업업체", _v(data, "contractor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "공사명",  _v(data, "trade_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_permit_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업 허가 기본정보",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "관리번호", _v(data, "permit_no"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "허가 유효기간", _v(data, "validity_period"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업일시", _v(data, "work_date") + " " + _v(data, "work_time"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업책임자", _v(data, "work_supervisor"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업장소", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "굴착 깊이", _v(data, "excavation_depth"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "굴착 범위", _v(data, "excavation_area"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 내용", _v(data, "work_content"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 35; r += 1
    return r


def _write_checklist_section(ws, start_row: int, title: str,
                              items: List[str], fill=None) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 6, "확인 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "이행", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "비고", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1
    for i, item in enumerate(items, 1):
        _write_cell(ws, r, 1, 1, i, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 6, item, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7, 7, "□", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, "", font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1
    return r


def _write_workers_table(ws, start_row: int, workers: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 5, "성명", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 8, "직종", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1
    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, r, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 5, item.get("name", ""), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 8, item.get("job_type", ""), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 22
        r += 1
    return r


def _write_sign_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작성 / 검토 / 승인",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "작업허가자 (서명)", _v(data, "permit_issuer"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "관리감독자 (서명)", _v(data, "supervisor_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "안전관리자 (서명)", _v(data, "safety_manager_sign"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업종료 확인자", _v(data, "work_end_confirmer"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "작업 종료 시각", _v(data, "work_end_time"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "최종 확인 서명", _v(data, "final_sign"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "작업 중 이상사항", _v(data, "during_work_issues"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 40; r += 1
    return r


def build_excavation_work_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 굴착 작업 허가서 xlsx 바이너리를 반환."""
    data    = form_data or {}
    workers = data.get("workers") or []

    pre_checks     = data.get("pre_check_items")     or _PRE_CHECK_DEFAULTS
    risk_factors   = data.get("risk_factor_items")   or _RISK_FACTOR_DEFAULTS
    safety_measures= data.get("safety_measure_items")or _SAFETY_MEASURE_DEFAULTS
    ppe_items      = data.get("ppe_items")            or _PPE_DEFAULTS
    approval_conds = data.get("approval_conditions") or _APPROVAL_CONDITIONS_DEFAULTS
    stop_conds     = data.get("stop_conditions")     or _STOP_CONDITIONS_DEFAULTS

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_site_info(ws, row, data)
    row = _write_permit_info(ws, row, data)
    row = _write_checklist_section(ws, row, "굴착 전 확인사항", pre_checks)
    row = _write_checklist_section(ws, row, "주요 위험요인", risk_factors, fill=_FILL_WARN)
    row = _write_checklist_section(ws, row, "안전조치 확인", safety_measures)
    row = _write_checklist_section(ws, row, "보호구 및 장비 확인", ppe_items)
    row = _write_checklist_section(ws, row, "작업 승인 조건", approval_conds)
    row = _write_checklist_section(ws, row, "작업 중지 조건", stop_conds, fill=_FILL_WARN)
    row = _write_workers_table(ws, row, workers)
    _write_sign_section(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
