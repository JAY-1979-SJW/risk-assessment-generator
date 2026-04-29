"""
특수건강진단 대상자 및 결과 관리대장 — Excel 출력 모듈 (v1.0).

법적 근거: 산업안전보건법 제130조 (특수건강진단)
분류:     GEN_INTERNAL — 검진기관 발급 공식 결과서를 대체하지 않음
          이 문서는 특수건강진단 대상자·결과·사후관리를 사업장에서 자체 관리하는 관리대장임.

주의:
- 구체적 질병명·진단명·주민등록번호 전체 등 민감정보 필드는 포함하지 않음.
- 기본 식별 정보는 생년월일 또는 사번 중심으로 기재.
- 검진기관 발급 공식 결과서 원본은 별도 보관 필수.
- 개인건강정보는 열람 권한자를 제한하고 보안 관리.

Input — form_data dict:
    site_name             str|None   사업장명
    project_name          str|None   현장/공장명
    exam_target_work      str|None   대상 작업/유해인자
    exam_period           str|None   검진 실시 기간
    supervisor            str|None   건강관리 담당자
    contractor            str|None   도급업체
    prepared_by           str|None   작성자

    exam_agency           str|None   검진기관명
    agency_contact        str|None   검진기관 연락처
    exam_date             str|None   검진 실시일
    result_received_date  str|None   결과 수령일

    exam_type             str|None   검진 구분 (배치전/정기/수시/임시)
    hazardous_agents      str|None   해당 유해인자 목록 (쉼표 구분)

    worker_rows list[dict]  대상 근로자 목록 (MAX_WORKERS=15)
        employee_no        str|None   사번
        name               str|None   성명
        birth_year         str|None   생년(4자리)
        job_type           str|None   직종/작업
        exam_done          str|None   수검 여부 ("완료"/"미수검"/"예정")
        judgment           str|None   판정 구분 (A/B/C1/C2/D1/D2/R 등)
        followup_needed    str|None   사후관리 필요 여부

    judgment_summary      str|None   전체 판정 요약 (예: "A 10명, C1 2명, 미수검 1명")
    followup_plan         str|None   사후관리 계획
    non_exam_count        str|None   미수검자 수
    non_exam_reason       str|None   미수검 사유
    non_exam_action       str|None   미수검자 조치 계획

    original_stored       str|None   원본 결과서 보관 여부 (예: "보관 완료")
    privacy_confirmed     str|None   개인정보 보호 확인 여부

    confirmer_name        str|None   확인자 성명
    confirmer_role        str|None   확인자 직위
    sign_date             str|None   작성일

Output — xlsx bytes (in-memory).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME    = "특수건강진단관리대장"
SHEET_HEADING = "특수건강진단 대상자 및 결과 관리대장"
SHEET_SUBTITLE = "산업안전보건법 제130조에 따른 특수건강진단 대상자·결과·사후관리 관리용 문서"

ORIGINAL_RESULT_NOTE = (
    "※ 외부 지정 검진기관이 발급한 특수건강진단 결과서 원본을 별도 보관하여야 합니다."
    "  이 관리대장은 공식 결과서를 대체하지 않습니다."
)

PRIVACY_NOTE = (
    "※ 이 문서에 기재된 근로자 건강정보는 개인정보보호법 및 산업안전보건법에 따라"
    "  열람 권한자를 제한하고 안전하게 보관·관리하여야 합니다."
)

MAX_WORKERS = 15
TOTAL_COLS  = 8

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------
_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9, italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8, italic=True)
_FONT_NOTE     = Font(name="맑은 고딕", size=9, bold=True, color="C00000")

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_NOTE    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 10, 2: 10, 3: 10, 4: 12, 5: 14, 6: 10, 7: 10, 8: 10,
}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int,
              label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션 렌더링
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font      = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font      = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16

    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H = 20
    r = start_row

    _write_lv(ws, r, "사업장명", _v(data, "site_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장/공장명", _v(data, "project_name"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "대상 작업/\n유해인자", _v(data, "exam_target_work"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30

    r += 1
    _write_lv(ws, r, "검진 기간", _v(data, "exam_period"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "담당자", _v(data, "supervisor"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자", _v(data, "prepared_by"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H

    return r + 1


def _write_exam_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "검진기관 정보 및 검진 구분",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "검진기관명", _v(data, "exam_agency"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "검진기관 연락처", _v(data, "agency_contact"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_lv(ws, r, "검진 실시일", _v(data, "exam_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "결과 수령일", _v(data, "result_received_date"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_lv(ws, r, "검진 구분", _v(data, "exam_type"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "해당 유해인자", _v(data, "hazardous_agents"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20

    return r + 1


def _write_worker_table(ws, start_row: int,
                        worker_rows: List[Dict[str, Any]]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "대상 근로자 목록 및 수검 결과",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 1, "순번",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 2, 2, "사번",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 3, 3, "성명",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 4, 4, "생년",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 5, "직종/작업", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 6, "수검여부",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 7, "판정구분",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 8, 8, "사후관리",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 18
    r += 1

    for i in range(MAX_WORKERS):
        item = worker_rows[i] if i < len(worker_rows) else {}
        _write_cell(ws, r, 1, 1, i + 1,
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 2, _v(item, "employee_no"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 3, 3, _v(item, "name"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 4, 4, _v(item, "birth_year"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 5, 5, _v(item, "job_type"),
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 6, _v(item, "exam_done"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 7, 7, _v(item, "judgment"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 8, 8, _v(item, "followup_needed"),
                    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 20
        r += 1

    return r


def _write_result_followup(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "판정 요약 및 사후관리",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "전체 판정\n요약",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "judgment_summary"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "사후관리\n계획",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "followup_plan"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_lv(ws, r, "미수검자 수", _v(data, "non_exam_count"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "미수검 사유", _v(data, "non_exam_reason"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 25
    r += 1

    _write_cell(ws, r, _L1, _L1, "미수검자\n조치 계획",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END,
                _v(data, "non_exam_action"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_original_and_privacy(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS,
                "원본 보관 확인 및 개인정보 보호",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "원본 결과서\n보관 여부", _v(data, "original_stored"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "개인정보\n보호 확인", _v(data, "privacy_confirmed"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30
    r += 1

    _write_cell(ws, r, 1, TOTAL_COLS,
                ORIGINAL_RESULT_NOTE,
                font=_FONT_NOTE, fill=_FILL_NOTE,
                align=_ALIGN_CENTER, height=36)
    r += 1

    _write_cell(ws, r, 1, TOTAL_COLS,
                PRIVACY_NOTE,
                font=_FONT_NOTE, fill=_FILL_NOTE,
                align=_ALIGN_CENTER, height=36)
    r += 1

    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row

    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION,
                align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, 1, 2, "작성자 서명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "확인자 서명",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20
    r += 1

    _write_cell(ws, r, 1, 4, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "",
                font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36

    return r + 1


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def render_special_health_examination_sheet(ws, form_data: Dict[str, Any]) -> None:
    data        = form_data or {}
    worker_rows = data.get("worker_rows") or []

    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_exam_info(ws, row, data)
    row = _write_worker_table(ws, row, worker_rows)
    row = _write_result_followup(ws, row, data)
    row = _write_original_and_privacy(ws, row, data)
    _write_confirmation(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75


def build_special_health_examination_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_data dict를 받아 특수건강진단 대상자 및 결과 관리대장 xlsx 바이너리를 반환.

    이 문서는 검진기관이 발급한 공식 특수건강진단 결과서를 대체하지 않습니다.
    사업장 자체 대상자 관리·결과 기록·사후관리 관리용 문서입니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    render_special_health_examination_sheet(ws, form_data)
    ws.print_title_rows = "1:12"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
