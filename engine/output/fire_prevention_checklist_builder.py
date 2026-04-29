"""
화재 예방 점검표 — Excel 출력 모듈 (v1.0).

법적 근거:
    산업안전보건기준에 관한 규칙 제236조 (화재위험 작업 장소, 가연성물질 제거)
    제241조 (화기작업 안전조치, 불티비산방지)
    제241조의2 (화재감시자 배치)
    제243조·제244조 (소화설비, 잔불감시)
    PTW-002 evidence (L1~L4) 재사용

분류: PRACTICAL — 법정 별지 서식 없음.
    화기작업 사전·중·후 화재예방 이행 여부를 확인하는 실무 표준서식.
    PTW-002 화기작업 허가서와 연계하여 사용.

Required form_data keys:
    site_name        str  현장명
    inspection_date  str  점검일
    inspector        str  점검자

Optional form_data keys:
    project_name          str  공사명
    inspection_no         str  점검 번호
    work_name             str  작업명
    work_location         str  작업 위치
    work_datetime         str  작업 일시
    work_category         str  작업 종류
    related_ptw_no        str  관련 PTW 번호
    work_supervisor       str  작업책임자
    # 화기작업 유형 (list[str])
    fire_work_types       list[str]  화기작업 유형 확인 항목 키
    # 점검 항목 (list[str])
    combustible_checks    list[str]  가연물 제거 확인 항목 키
    spark_checks          list[str]  불티비산 방지 확인 항목 키
    extinguisher_checks   list[str]  소화설비 확인 항목 키
    fire_watch_checks     list[str]  화재감시자 확인 항목 키
    gas_equip_checks      list[str]  가스·용접장비 점검 확인 항목 키
    elec_fire_checks      list[str]  전기화재 예방 확인 항목 키
    post_work_checks      list[str]  작업 후 화재 확인 항목 키
    # 부적합 사항
    nonconformance_items  list[dict]  content, action, deadline, completed
    # 판정
    verdict               str  적합 / 조건부 적합 / 작업중지
    verdict_condition     str
    # 서명
    inspector_sign        str
    fire_watch_sign       str
    supervisor_sign       str
    safety_manager_sign   str
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "화재예방점검표"
SHEET_HEADING = "화재 예방 점검표"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제236조·제241조·제241조의2·제243조·제244조에 따른 "
    "화기작업 화재예방 이행 확인 기록서"
)

MAX_NC     = 5
TOTAL_COLS = 8

NOTICE_PRACTICAL = (
    "본 점검표는 법정 별지 서식이 아닌 실무 표준서식으로, "
    "화기작업 전·중·후 화재예방 조치 이행 여부를 확인·기록한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다. "
    "PTW-002 화기작업 허가서와 함께 사용한다."
)

# ---------------------------------------------------------------------------
# 점검 항목 정의 (키: 표시 문구 쌍)
# ---------------------------------------------------------------------------

FIRE_WORK_TYPE_ITEMS: List[tuple] = [
    ("fw_welding",     "용접 (아크·가스·CO₂ 등)"),
    ("fw_cutting",     "용단 (가스절단)"),
    ("fw_grinding",    "그라인더"),
    ("fw_sawing",      "전기톱·절단기"),
    ("fw_torch",       "토치 (가스버너)"),
    ("fw_heat_gun",    "열풍기 사용"),
    ("fw_other_spark", "기타 불꽃·고열 발생 작업"),
]

COMBUSTIBLE_ITEMS: List[tuple] = [
    ("comb_removed",    "작업 반경 11m 이내 가연성 물질 제거 (제236조)"),
    ("comb_protective", "제거 불가 가연물 — 방화보양재(방화포 등)로 덮음"),
    ("comb_flammable",  "인화성 물질(유류·도료·솔벤트) 격리 및 잠금"),
    ("comb_gas",        "가스·압축가스 용기 — 화기 이격·안전 보관"),
    ("comb_ventilation","밀폐·반밀폐 공간 환기 상태 확인"),
]

SPARK_ITEMS: List[tuple] = [
    ("spark_blanket",   "불꽃방지포(방화포) 설치 (제241조)"),
    ("spark_receiver",  "불티받이 설치 (낙하 불티 차단)"),
    ("spark_opening",   "개구부·구멍·슬리브 차단 조치"),
    ("spark_lower",     "하부층 낙하 불티 확인 및 감시 조치"),
    ("spark_duct",      "덕트·배관 관통부 차단 확인"),
]

EXTINGUISHER_ITEMS: List[tuple] = [
    ("ext_placed",      "소화기 작업장소 인근 비치 (제243조)"),
    ("ext_hydrant",     "옥내소화전 위치 확인 및 접근로 확보"),
    ("ext_water",       "방화수(물통) 준비"),
    ("ext_pressure",    "소화기 압력 정상 확인 (지시게이지 녹색)"),
    ("ext_marked",      "소화기 위치 표시 부착"),
]

FIRE_WATCH_ITEMS: List[tuple] = [
    ("fwatch_assigned", "화재감시자 지정 (제241조의2)"),
    ("fwatch_position", "감시 위치 설정 — 불티 낙하 방향 확인 가능"),
    ("fwatch_comm",     "비상연락 수단(무전기·휴대폰) 구비"),
    ("fwatch_during",   "작업 중 상시 감시 유지"),
    ("fwatch_after30",  "작업 완료 후 30분 이상 잔불 감시 (제244조)"),
]

GAS_EQUIP_ITEMS: List[tuple] = [
    ("gas_fixed",       "산소·연료가스 용기 직립 고정 및 전도 방지"),
    ("gas_flashback",   "역화방지기 설치 (산소·연료 양쪽)"),
    ("gas_hose",        "호스 손상·노화·접속부 누기 없음"),
    ("gas_regulator",   "조정기(레귤레이터) 상태 양호"),
    ("gas_valve",       "미사용 밸브 잠금 확인"),
]

ELEC_FIRE_ITEMS: List[tuple] = [
    ("ef_insulation",   "용접기·공구 전선 피복 손상 없음"),
    ("ef_elcb",         "누전차단기(ELCB) 사용 확인"),
    ("ef_ground",       "용접기 접지 상태 확인"),
    ("ef_overload",     "과부하·문어발 배선 없음"),
]

POST_WORK_ITEMS: List[tuple] = [
    ("pw_ember",        "잔불·불씨 소화 완료 확인 (제244조)"),
    ("pw_temp",         "열화상·온도 측정 이상 없음 (해당 시)"),
    ("pw_30min",        "작업 완료 후 30분 이상 감시 완료"),
    ("pw_waste",        "용접 슬래그·불꽃 찌꺼기 정리 및 처리 완료"),
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SECTION3 = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_FIRE     = PatternFill(fill_type="solid", fgColor="FCE4D6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=fill or _FILL_NOTICE,
                align=_ALIGN_LEFT, height=28)
    return row + 1


def _write_checklist_items(ws, row: int, items: List[tuple],
                           checked_keys: set) -> int:
    """체크리스트 항목 행 작성. 키가 checked_keys에 있으면 ■, 없으면 □."""
    for key, label in items:
        mark = "■" if key in checked_keys else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {label}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 32
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 18
    row += 1
    row = _write_notice(ws, row, NOTICE_PRACTICAL)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "점검 번호",
              _v(data, "inspection_no") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 점검 대상 작업 정보")
    _write_lv(ws, row, "작업명",
              _v(data, "work_name"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업 위치",
              _v(data, "work_location"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업 일시",
              _v(data, "work_datetime"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업 종류",
              _v(data, "work_category"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관련 PTW 번호",
              _v(data, "related_ptw_no") or "",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "점검자",
              _v(data, "inspector"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업책임자",
              _v(data, "work_supervisor"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    _write_lv(ws, row, "점검일",
              _v(data, "inspection_date"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s3_fire_work_types(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 화기작업 유형  (해당 항목 체크)", fill=_FILL_FIRE)
    checked = set(data.get("fire_work_types") or [])
    return _write_checklist_items(ws, row, FIRE_WORK_TYPE_ITEMS, checked)


def _write_s4_combustible(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 작업 전 가연물 제거 확인  (제236조)")
    checked = set(data.get("combustible_checks") or [])
    return _write_checklist_items(ws, row, COMBUSTIBLE_ITEMS, checked)


def _write_s5_spark(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 불티 비산 방지 조치  (제241조)")
    checked = set(data.get("spark_checks") or [])
    return _write_checklist_items(ws, row, SPARK_ITEMS, checked)


def _write_s6_extinguisher(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 소화설비 및 초기진화 준비  (제243조)")
    checked = set(data.get("extinguisher_checks") or [])
    return _write_checklist_items(ws, row, EXTINGUISHER_ITEMS, checked)


def _write_s7_fire_watch(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 화재감시자 배치  (제241조의2·제244조)", fill=_FILL_WARN)
    checked = set(data.get("fire_watch_checks") or [])
    return _write_checklist_items(ws, row, FIRE_WATCH_ITEMS, checked)


def _write_s8_gas_equip(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 가스·용접장비 점검")
    checked = set(data.get("gas_equip_checks") or [])
    return _write_checklist_items(ws, row, GAS_EQUIP_ITEMS, checked)


def _write_s9_elec_fire(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 전기화재 예방")
    checked = set(data.get("elec_fire_checks") or [])
    return _write_checklist_items(ws, row, ELEC_FIRE_ITEMS, checked)


def _write_s10_post_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 작업 후 화재 확인  (제244조)", fill=_FILL_FIRE)
    checked = set(data.get("post_work_checks") or [])
    return _write_checklist_items(ws, row, POST_WORK_ITEMS, checked)


def _write_s11_nonconformance(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 부적합 사항 및 시정조치")
    _write_cell(ws, row, 1, 3, "부적합 내용", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 4, 5, "시정조치",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 7, "완료 기한",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "완료",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    items = (data.get("nonconformance_items") or [])[:MAX_NC]
    for i in range(MAX_NC):
        item = items[i] if i < len(items) else {}
        _write_cell(ws, row, 1, 3, _v(item, "content"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=22)
        _write_cell(ws, row, 4, 5, _v(item, "action"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 7, _v(item, "deadline"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, _v(item, "completed"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        row += 1
    return row


def _write_s12_verdict(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 작업 가능 여부 판정", fill=_FILL_SECTION2)
    verdict = _v(data, "verdict") or ""
    marks = {
        "적합":       "■" if verdict == "적합"       else "□",
        "조건부 적합": "■" if verdict == "조건부 적합" else "□",
        "작업중지":   "■" if verdict == "작업중지"   else "□",
    }
    verdict_text = (
        f"{marks['적합']} 적합    "
        f"{marks['조건부 적합']} 조건부 적합    "
        f"{marks['작업중지']} 작업중지"
    )
    _write_cell(ws, row, 1, TOTAL_COLS, verdict_text,
                font=_FONT_BOLD, align=_ALIGN_CENTER, height=24)
    row += 1
    _write_lv(ws, row, "조건부 조건",
              _v(data, "verdict_condition") or "(조건부 적합 시 기재)",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    return row


def _write_s13_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "13. 확인 서명", fill=_FILL_SECTION3)
    _write_lv(ws, row, "점검자 (서명)",
              _v(data, "inspector_sign") or _v(data, "inspector"),
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "화재감시자 (서명)",
              _v(data, "fire_watch_sign"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "관리감독자 (서명)",
              _v(data, "supervisor_sign"),
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "안전관리자 (서명)",
              _v(data, "safety_manager_sign"),
              _L2, _V2_START, _V2_END)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_fire_prevention_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 화재 예방 점검표 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_info(ws, row, data)
    row = _write_s3_fire_work_types(ws, row, data)
    row = _write_s4_combustible(ws, row, data)
    row = _write_s5_spark(ws, row, data)
    row = _write_s6_extinguisher(ws, row, data)
    row = _write_s7_fire_watch(ws, row, data)
    row = _write_s8_gas_equip(ws, row, data)
    row = _write_s9_elec_fire(ws, row, data)
    row = _write_s10_post_work(ws, row, data)
    row = _write_s11_nonconformance(ws, row, data)
    row = _write_s12_verdict(ws, row, data)
    _write_s13_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
