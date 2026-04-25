"""
전기작업 허가서 / LOTO — Excel 출력 모듈 (v1.0).

법적 근거:
    산업안전보건기준에 관한 규칙 제319조 (정전전로에서의 전기작업)
    제320조 (정전전로 인근에서의 전기작업)
    제321조 (활선작업)
    제322조 (충전전로 인근에서의 전기작업 — 접근한계거리)
    제323조 (절연용 보호구 착용)
    제302조~제305조 (접지, 누전차단기)
    KOSHA GUIDE E-2-2023 전기작업 안전지침 (참고)

분류: PRACTICAL — 법정 별지 서식 없음.
    법령 제319조~제323조 안전조치 이행 확인 양식 + LOTO 절차 기록.
    전기작업 허가 및 재통전 승인 기록 포함.

Required form_data keys:
    site_name        str  현장명
    work_date        str  작업일자
    work_location    str  작업 위치
    work_supervisor  str  작업책임자

Optional form_data keys:
    project_name      str  공사명
    permit_no         str  허가번호
    work_time         str  작업 일시
    voltage           str  작업 전압 (V/kV)
    work_category     str  작업 구분 (정전/활선/근접 등)
    contractor        str  작업업체
    work_name         str  작업명
    permit_issuer     str  허가자
    supervisor_name   str  관리감독자
    safety_manager    str  안전관리자
    validity_period   str  허가 유효기간
    # 전기작업 유형 (list[str])
    work_types               list[str]  선택된 전기작업 유형 목록
    # 선행서류 확인 (list[str])
    prereq_checks            list[str]  확인된 선행서류 목록
    # LOTO 확인 항목 (list[str])
    loto_checks              list[str]  이행된 LOTO 항목 목록
    # 무전압·잔류전압 확인 (list[str])
    voltage_zero_checks      list[str]  이행된 무전압 확인 항목 목록
    # 활선·근접 허가 조건 (list[str])
    live_work_checks         list[str]  이행된 활선 안전조치 항목 목록
    # 보호구·장비 확인 (list[str])
    ppe_checks               list[str]  지급된/확인된 보호구 항목 목록
    # 작업구역 통제 (list[str])
    zone_control_checks      list[str]  이행된 구역통제 항목 목록
    # 작업중지 조건 (list[str])
    stop_conditions          list[str]  해당되는 작업중지 조건 목록
    # 작업완료 확인 (list[str])
    completion_checks        list[str]  이행된 작업완료 확인 항목 목록
    # 부적합 사항 (list[dict])
    nonconformance_items     list[dict] content, action, deadline, completed
    # 작업자 명단 (list[dict])
    workers                  list[dict] name, job_type
    # 단순 텍스트 필드
    loto_scope               str  차단 대상 전로
    loto_breaker_location    str  차단기 위치
    loto_key_holder          str  열쇠 보관자
    voltage_tester_used      str  검전기 사용 여부
    voltage_zero_confirmed   str  무전압 확인 결과
    residual_voltage_result  str  잔류전압 방전 결과
    ground_confirmed         str  접지 확인 결과
    voltage_measurer_sign    str  측정자 서명
    reenergize_approver      str  재통전 승인자
    work_end_time            str  작업 종료 시각
    work_end_confirmer       str  작업 종료 확인자
    during_work_issues       str  작업 중 이상 사항
    final_sign               str  최종 확인 서명
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "전기작업허가서"
SHEET_HEADING = "전기작업 허가서 / LOTO"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제319조~제323조에 따른 "
    "전기작업 전 안전조치 및 전로 차단(LOTO) 확인 기록서"
)

MAX_WORKERS = 8
MAX_NC      = 5
TOTAL_COLS  = 8

# ---------------------------------------------------------------------------
# 고정 문구
# ---------------------------------------------------------------------------

NOTICE_PRACTICAL = (
    "본 허가서는 법정 별지 서식이 아닌 실무 표준서식으로, "
    "법령 제319조~제323조 안전조치 이행 여부를 확인·기록한다."
)
NOTICE_LOTO = (
    "전로 차단 후 반드시 검전기로 무전압을 확인하고, 잠금장치 및 표지를 부착한 후 작업을 허가한다."
    " (제319조)"
)
NOTICE_LIVE = (
    "활선 또는 충전전로 근접작업 시 접근한계거리 이내 접근 금지. "
    "절연보호구 착용 및 절연공구 사용 의무. (제321조~제323조)"
)
NOTICE_REENERGIZE = (
    "재통전 전 작업구역 내 인원 전원 철수 확인 및 잠금장치 해제 승인 후에만 재통전 가능."
)
NOTICE_STOP = (
    "아래 조건 발생 시 즉시 작업 중지 후 허가자에게 보고한다."
)
NOTICE_LAW_REF = (
    "법령 조항 및 세부 기준은 현행 원문과 발주처·원청 기준을 확인 후 적용한다."
)

# ---------------------------------------------------------------------------
# 전기작업 유형 목록
# ---------------------------------------------------------------------------

WORK_TYPES = [
    "정전작업",
    "활선작업",
    "충전전로 근접작업",
    "임시전기 설치",
    "분전반 작업",
    "케이블 결선/해체",
    "시험·측정 작업",
    "기타 전기작업",
]

# ---------------------------------------------------------------------------
# 선행서류 확인 항목
# ---------------------------------------------------------------------------

PREREQ_ITEMS = [
    "RA-001 위험성평가표 — 전기감전·아크·화재 위험 포함 여부",
    "RA-004 TBM 일지 — 당일 전기작업 위험 공유 여부",
    "WP-011 전기 작업계획서 — 작업범위·전압·차단방법 기재 여부",
    "CL-004 전기설비 정기 점검표 — 해당 시 최근 점검 결과 확인",
]

# ---------------------------------------------------------------------------
# 전로 차단 및 LOTO 확인 항목
# ---------------------------------------------------------------------------

LOTO_CHECK_ITEMS = [
    "차단 대상 전로 확인 (제319조)",
    "차단기 차단 여부 확인",
    "잠금장치(Lockout) 설치 여부 확인",
    "표지(Tagout) 부착 여부 확인",
    "담당자 지정 여부 확인",
    "열쇠 보관자 지정 여부 확인",
    "재투입 방지 조치 여부 확인",
    "인근 충전전로 방호 여부 확인 (제320조)",
]

# ---------------------------------------------------------------------------
# 무전압·잔류전압 확인 항목
# ---------------------------------------------------------------------------

VOLTAGE_ZERO_ITEMS = [
    "검전기로 무전압 확인 (제319조 제1항)",
    "잔류전압 방전 완료 확인",
    "접지 완료 확인 (제319조 제1항)",
    "측정자 서명 확인",
]

# ---------------------------------------------------------------------------
# 활선·근접작업 허가 조건 항목
# ---------------------------------------------------------------------------

LIVE_WORK_ITEMS = [
    "접근한계거리 표시 및 확인 (제322조)",
    "충전부 방호판·방호망 설치 여부 확인",
    "절연장갑·절연화 착용 여부 확인 (제323조)",
    "절연공구 사용 여부 확인",
    "감시자 배치 여부 확인",
    "활선작업용 기구 사용 여부 확인 (제321조)",
]

# ---------------------------------------------------------------------------
# 보호구 및 장비 확인 항목
# ---------------------------------------------------------------------------

PPE_CHECK_ITEMS = [
    "절연장갑 지급·착용 확인",
    "절연화 지급·착용 확인",
    "보안면(아크 방호) 지급·착용 확인",
    "절연매트 설치 확인",
    "검전기 구비 및 교정 확인",
    "절연공구 상태 확인",
    "절연저항계(메거) 구비 확인 (해당 시)",
]

# ---------------------------------------------------------------------------
# 작업구역 통제 항목
# ---------------------------------------------------------------------------

ZONE_CONTROL_ITEMS = [
    "출입금지 표지 설치 확인",
    "방책·콘 설치 확인",
    "감시자 배치 확인",
    "비상연락체계 확인",
    "소화기 비치 확인",
]

# ---------------------------------------------------------------------------
# 작업중지 조건
# ---------------------------------------------------------------------------

STOP_CONDITION_ITEMS = [
    "작업범위 변경 발생",
    "재통전 필요 발생 (절차 미완료)",
    "보호구 미착용 또는 훼손 발견",
    "전압 확인 불가 또는 불명확",
    "우천·침수·누전 위험 발생",
    "감전 또는 아크 징후 발생",
    "허가자 지시 없는 범위 이탈",
]

# ---------------------------------------------------------------------------
# 작업완료 확인 항목
# ---------------------------------------------------------------------------

COMPLETION_CHECK_ITEMS = [
    "작업 완료 여부 확인",
    "인원 전원 철수 확인",
    "공구·장비 회수 확인",
    "잠금장치 해제 승인 완료",
    "재통전 전 감전위험 구역 최종 확인",
    "재통전 승인자 서명 완료",
]

# ---------------------------------------------------------------------------
# 스타일
# ---------------------------------------------------------------------------

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_NOTICE   = Font(name="맑은 고딕", size=9,  italic=True, color="666666")

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_SECTION2 = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SECTION3 = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FFFBE6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 16, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int, height=20) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _write_section_header(ws, row: int, title: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION,
                align=_ALIGN_CENTER, height=20)
    return row + 1


def _write_notice(ws, row: int, text: str, fill=None) -> int:
    _write_cell(ws, row, 1, TOTAL_COLS, text,
                font=_FONT_NOTICE, fill=fill or _FILL_NOTICE,
                align=_ALIGN_LEFT, height=28)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# 섹션별 렌더러
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 32
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 18
    row += 1
    row = _write_notice(ws, row, NOTICE_PRACTICAL)
    return row


def _write_s1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "1. 현장 기본정보")
    _write_lv(ws, row, "현장명",  _v(data, "site_name"),    _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "허가번호", _v(data, "permit_no") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    return row


def _write_s2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "2. 허가 대상 작업 정보")
    _write_lv(ws, row, "작업명", _v(data, "work_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업 위치", _v(data, "work_location"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업 일시", _v(data, "work_time") or _v(data, "work_date"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업 전압",
              _v(data, "voltage") or "__ V / __ kV",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업 구분",
              _v(data, "work_category") or "□ 정전   □ 활선   □ 근접   □ 기타",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "작업업체", _v(data, "contractor"), _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업책임자", _v(data, "work_supervisor"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "허가번호",  _v(data, "permit_no") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    # 작업자 명단
    _write_cell(ws, row, 1, TOTAL_COLS, "작업자 명단",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    row += 1
    _write_cell(ws, row, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 2, 5, "성명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 8, "직종/역할", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    workers = data.get("workers") or []
    for i in range(MAX_WORKERS):
        item = workers[i] if i < len(workers) else {}
        _write_cell(ws, row, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 5, _v(item, "name"),     font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 8, _v(item, "job_type"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1
    return row


def _write_s3_work_types(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "3. 전기작업 유형  (해당 항목 선택)")
    selected = set(data.get("work_types") or [])
    half = TOTAL_COLS // 2
    items = WORK_TYPES
    for i in range(0, len(items), 2):
        left  = items[i]
        right = items[i + 1] if i + 1 < len(items) else None
        lm = "■" if left in selected else "□"
        _write_cell(ws, row, 1, half, f"{lm} {left}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        if right:
            rm = "■" if right in selected else "□"
            _write_cell(ws, row, half + 1, TOTAL_COLS, f"{rm} {right}",
                        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        else:
            _write_cell(ws, row, half + 1, TOTAL_COLS, "",
                        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        row += 1
    return row


def _write_s4_prereq(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "4. 선행서류 확인  (확인된 서류에 ■ 표시)")
    checked = set(data.get("prereq_checks") or [])
    for item in PREREQ_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s5_loto(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "5. 전로 차단 및 LOTO 확인  (제319조~제320조)")
    row = _write_notice(ws, row, NOTICE_LOTO)
    _write_lv(ws, row, "차단 대상 전로",
              _v(data, "loto_scope") or "",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "차단기 위치",
              _v(data, "loto_breaker_location") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "열쇠 보관자",
              _v(data, "loto_key_holder") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    row += 1
    checked = set(data.get("loto_checks") or [])
    for item in LOTO_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s6_voltage_zero(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "6. 무전압·잔류전압 확인  (제319조 제1항)")
    _write_lv(ws, row, "검전기 확인",
              _v(data, "voltage_tester_used") or "□ 확인   □ 미확인",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "무전압 확인",
              _v(data, "voltage_zero_confirmed") or "□ 확인   □ 미확인",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "잔류전압 방전",
              _v(data, "residual_voltage_result") or "□ 완료   □ 미완료",
              _L1, _V1_START, _V1_END)
    _write_lv(ws, row, "접지 확인",
              _v(data, "ground_confirmed") or "□ 완료   □ 미완료",
              _L2, _V2_START, _V2_END)
    row += 1
    checked = set(data.get("voltage_zero_checks") or [])
    for item in VOLTAGE_ZERO_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "측정자 서명",
              _v(data, "voltage_measurer_sign") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=28)
    row += 1
    return row


def _write_s7_live_work(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "7. 활선·근접작업 허가 조건  (제321조~제323조)")
    row = _write_notice(ws, row, NOTICE_LIVE)
    checked = set(data.get("live_work_checks") or [])
    for item in LIVE_WORK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s8_ppe(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "8. 보호구 및 장비 확인  (제323조)")
    checked = set(data.get("ppe_checks") or [])
    for item in PPE_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s9_zone_control(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "9. 작업구역 통제")
    checked = set(data.get("zone_control_checks") or [])
    for item in ZONE_CONTROL_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    return row


def _write_s10_stop_conditions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "10. 작업 중 변경/작업중지 조건  (해당 시 ■ 표시)",
                                fill=_FILL_WARN)
    row = _write_notice(ws, row, NOTICE_STOP)
    checked = set(data.get("stop_conditions") or [])
    for item in STOP_CONDITION_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "작업 중 이상 사항",
              _v(data, "during_work_issues") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    return row


def _write_s11_completion(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "11. 작업 완료 및 재통전 승인", fill=_FILL_SECTION2)
    row = _write_notice(ws, row, NOTICE_REENERGIZE)
    checked = set(data.get("completion_checks") or [])
    for item in COMPLETION_CHECK_ITEMS:
        mark = "■" if item in checked else "□"
        _write_cell(ws, row, 1, TOTAL_COLS, f"{mark}  {item}",
                    font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=20)
        row += 1
    _write_lv(ws, row, "작업 종료 시각",
              _v(data, "work_end_time") or "",
              _L1, _V1_START, _V1_END, height=24)
    _write_lv(ws, row, "종료 확인자",
              _v(data, "work_end_confirmer") or "",
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "재통전 승인자 서명",
              _v(data, "reenergize_approver") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=32)
    row += 1
    return row


def _write_s12_nonconformance(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "12. 부적합 및 시정조치")
    # 헤더 행
    _write_cell(ws, row, 1, 3, "부적합 내용", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, row, 4, 5, "시정조치",   font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 6, 7, "완료 기한",  font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "완료",       font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
    row += 1
    items = (data.get("nonconformance_items") or [])[:MAX_NC]
    for i in range(MAX_NC):
        item = items[i] if i < len(items) else {}
        _write_cell(ws, row, 1, 3, _v(item, "content"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=22)
        _write_cell(ws, row, 4, 5, _v(item, "action"),    font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, row, 6, 7, _v(item, "deadline"),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, _v(item, "completed"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        row += 1
    return row


def _write_s13_sign(ws, row: int, data: Dict[str, Any]) -> int:
    row = _write_section_header(ws, row, "13. 허가 / 승인 / 확인 서명", fill=_FILL_SECTION3)
    _write_lv(ws, row, "작업신청자 (서명)", _v(data, "work_supervisor"),
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "관리감독자 (서명)", _v(data, "supervisor_name"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "작업허가자 (서명)", _v(data, "permit_issuer"),
              _L1, _V1_START, _V1_END, height=36)
    _write_lv(ws, row, "안전관리자 (서명)", _v(data, "safety_manager"),
              _L2, _V2_START, _V2_END)
    row += 1
    _write_lv(ws, row, "최종 확인 서명",
              _v(data, "final_sign") or "",
              _L1, _FULL_VAL_START, _FULL_VAL_END, height=36)
    row += 1
    row = _write_notice(ws, row, NOTICE_LAW_REF)
    return row


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_electrical_work_permit_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 전기작업 허가서 xlsx 바이너리를 반환."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_s1_site_info(ws, row, data)
    row = _write_s2_work_info(ws, row, data)
    row = _write_s3_work_types(ws, row, data)
    row = _write_s4_prereq(ws, row, data)
    row = _write_s5_loto(ws, row, data)
    row = _write_s6_voltage_zero(ws, row, data)
    row = _write_s7_live_work(ws, row, data)
    row = _write_s8_ppe(ws, row, data)
    row = _write_s9_zone_control(ws, row, data)
    row = _write_s10_stop_conditions(ws, row, data)
    row = _write_s11_completion(ws, row, data)
    row = _write_s12_nonconformance(ws, row, data)
    _write_s13_sign(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
