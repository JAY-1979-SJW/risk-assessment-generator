"""
작업 전 안전 확인서 — Excel 출력 모듈 (v1).

법적 근거: 중대재해처벌법 시행령 제4조 (안전보건관리체계 구축 의무)
분류: PRACTICAL — 법정 별지 서식 없음, 실무 자체 표준서식

요약:
    작업자 개인이 작업 개시 전 자신의 작업 안전성을 자가 점검하는 서식입니다.
    TBM 안전점검 일지(RA-004)와 연계하여, 개별 근로자의 작업 전 자가 점검을 기록합니다.
    점검 항목별 ○/△/✗ 판정 및 이상 발견 시 즉시 보고 및 조치를 포함합니다.

Required form_data keys:
    check_date        str  확인 일자
    worker_name       str  작업자 성명
    work_location     str  작업 장소

Optional form_data keys:
    site_name               str  사업장명
    project_name            str  현장명
    work_type               str  작업 공종
    work_time               str  작업 시간
    supervisor_name         str  작업책임자
    department              str  소속
    position                str  직책
    work_content            str  작업 내용
    hazard_found            str  이상 발견 여부
    hazard_description      str  발견 내용
    action_taken            str  조치 내용
    action_taken_by         str  조치 담당자
    action_completed        str  조치 완료 여부
    work_approval           str  작업 개시 가능 여부 (가능/조건부 가능/불가)
    work_approval_reason    str  불가 시 사유
    checker_name            str  확인자 (작업책임자 또는 관리감독자)
    manager_sign            str  안전관리자 서명
    check_datetime          str  확인 일시
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "작업전안전확인서"
SHEET_HEADING = "작업 전 안전 확인서"
SHEET_SUBTITLE = "「중대재해처벌법」 시행령 제4조에 따른 근로자 작업 개시 전 자가 점검"
DOC_ID = "DL-005"

TOTAL_COLS = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8

_SELF_CHECK_ITEMS = [
    "작업 내용과 위험요인을 이해했는가",
    "필요한 보호구를 올바르게 착용했는가",
    "작업장 주변 정리정돈이 되어 있는가",
    "통로와 대피로가 확보되어 있는가",
    "사용할 장비와 공구 상태를 확인했는가",
    "전기/화기/고소/굴착 등 특수위험 여부를 확인했는가",
    "작업 전 TBM 또는 안전 브리핑에 참석했는가",
    "이상 발견 시 작업중지 기준을 이해했는가",
]

_HAZARD_FACTORS = [
    "추락",
    "낙하",
    "협착",
    "감전",
    "화재",
    "전도",
    "붕괴",
    "질식",
    "기타",
]

_SAFETY_MEASURES = [
    "보호구 착용 (안전모, 안전화, 안전대 등)",
    "작업구역 통제 및 안전표지 설치",
    "신호수/감시자 배치 (필요 시)",
    "장비 사전점검 완료",
    "작업발판, 통로, 사다리 등 상태 확인",
    "환기, 채광, 조명 확인",
    "비상연락체계 확인 및 구조장비 비치",
]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    doc_id = ws.cell(row=row, column=1, value=f"(문서ID: {DOC_ID})")
    doc_id.font = _FONT_SMALL
    doc_id.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 12
    return row + 1


def _write_site_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "현장 기본정보",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "현장명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "확인 일자", _v(data, "check_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자",  _v(data, "worker_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "소속", _v(data, "department"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "직책",  _v(data, "position"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_work_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업 기본정보",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "작업 공종", _v(data, "work_type"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업장소",  _v(data, "work_location"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 시간", _v(data, "work_time"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업책임자",  _v(data, "supervisor_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 내용", _v(data, "work_content"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 35; r += 1
    return r


def _write_checklist_section(ws, start_row: int, title: str,
                              items: List[str], fill=None) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, title,
                font=_FONT_BOLD, fill=fill or _FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 1, "순번", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 6, "확인 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 7, 8, "결과", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1
    for i, item in enumerate(items, 1):
        _write_cell(ws, r, 1, 1, i, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 6, item, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 7, 8, "○/△/✗", font=_FONT_SMALL, align=_ALIGN_CENTER)
        ws.row_dimensions[r].height = 22
        r += 1
    return r


def _write_abnormality_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "이상 발견 및 조치사항",
                font=_FONT_BOLD, fill=_FILL_WARN, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "이상 발견 여부", _v(data, "hazard_found"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "발견 내용", _v(data, "hazard_description"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 25; r += 1
    _write_lv(ws, r, "조치 내용", _v(data, "action_taken"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "조치 담당자", _v(data, "action_taken_by"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "조치 완료 여부", _v(data, "action_completed"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_approval_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업 개시 가능 여부",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "판정", _v(data, "work_approval"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "불가 시 사유", _v(data, "work_approval_reason"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_sign_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작성 / 확인 / 승인",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "작성자 (작업자) 서명", _v(data, "worker_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "확인일시", _v(data, "check_datetime"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "확인자 (작업책임자/관리감독자) 서명", _v(data, "checker_name"), _L1, _V1_START, _V1_END)
    _write_cell(ws, r, _L2, _FULL_VAL_END, "", font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 30; r += 1
    _write_lv(ws, r, "안전관리자 서명", _v(data, "manager_sign"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1
    return r


def build_work_safety_checklist(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 작업 전 안전 확인서 xlsx 바이너리를 반환."""
    data = form_data or {}

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_site_info(ws, row, data)
    row = _write_work_info(ws, row, data)
    row = _write_checklist_section(ws, row, "작업 전 자가 점검 항목", _SELF_CHECK_ITEMS)
    row = _write_checklist_section(ws, row, "주요 위험요인 확인", _HAZARD_FACTORS, fill=_FILL_WARN)
    row = _write_checklist_section(ws, row, "안전조치 확인", _SAFETY_MEASURES)
    row = _write_abnormality_section(ws, row, data)
    row = _write_approval_section(ws, row, data)
    _write_sign_section(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
