"""
안전순찰 점검 일지 — Excel 출력 모듈 (v2).

법적 근거:
    산업안전보건법 제38조 (안전조치)
    산업안전보건법 제36조 (위험성평가)
    산업안전보건법 시행규칙 제37조 (위험성평가 기록보존)
    산업안전보건법 제17조 (안전관리자 직무)

분류: PRACTICAL — 공식 별지 제출서식 아님 / 현장 안전순찰 및 지적사항 이행관리 보조서식

Required form_data keys:
    site_name       str  현장명
    patrol_date     str  순찰 일자
    patrol_officer  str  순찰자 성명

Optional form_data keys:
    project_name            str  공사명
    patrol_time_start       str  순찰 시작 시간
    patrol_time_end         str  순찰 종료 시간
    patrol_route            str  순찰 구간/경로
    department              str  소속
    position                str  직책
    contact                 str  연락처
    writer_name             str  작성자
    reviewer_name           str  검토자
    approver_name           str  승인자
    weather                 str  날씨
    temperature             str  기온
    total_workers           str  당일 전체 작업 인원
    high_risk_work_today    str  당일 고위험 작업 여부
    -- 섹션3: 구역별 안전순찰 결과 --
    patrol_results          list[dict]  구역별 순찰 결과 (max 10)
        seq                 int  번호
        area                str  구역/위치
        check_time          str  점검 시간
        hazard_found        str  위험요인/결함 내용
        risk_level          str  위험 수준 (상/중/하)
        immediate_action    str  즉시 조치 여부
        responsible_person  str  담당자
        due_date            str  완료 기한
        status              str  이행 상태
        remarks             str  비고
    -- 섹션4: 위험유형별 점검 --
    fall_protection         str  추락·낙하 예방
    electrical_safety       str  감전·전기 안전
    fire_prevention         str  화재·폭발 예방
    equipment_safety        str  기계·장비 안전
    chemical_safety         str  유해화학물질 관리
    health_hazard           str  건강장해 요인
    traffic_safety          str  교통·운반 안전
    others_risk             str  기타 위험요인
    -- 섹션5: 즉시 시정조치 기록 --
    immediate_actions       list[dict]  즉시 조치 목록 (max 5)
        seq                 int  번호
        area                str  위치
        issue               str  지적 내용
        action_taken        str  즉시 조치 내용
        action_by           str  조치자
        action_time         str  조치 시간
        confirmed           str  확인 여부
    -- 섹션6: 개선조치 이행관리 --
    improvement_actions     list[dict]  개선조치 목록 (max 8)
        seq                 int  번호
        issue               str  지적사항
        improvement         str  개선조치 내용
        assignee            str  담당자
        due_date            str  완료 예정일
        completed_date      str  완료일
        status              str  이행상태
        evidence_exists     str  증빙 여부
        confirmed_by        str  확인자
    -- 섹션7: 반복 지적 및 재발 위험 --
    repeat_issues           str  반복 지적사항 내용
    repeat_risk_level       str  재발 위험 수준
    ra_reflected            str  위험성평가 반영 여부
    tbm_reflected           str  TBM 교육 반영 여부
    root_cause              str  근본 원인 분석
    prevention_measures     str  재발 방지 대책
    -- 섹션8: 사고·아차사고 연계 --
    accident_occurred       str  사고 발생 여부
    near_miss_occurred      str  아차사고 발생 여부
    em_form_linked          str  EM 서식 연계 여부
    em_form_type            str  연계 EM 서식 종류
    followup_needed         str  후속 조치 필요 여부
    followup_content        str  후속 조치 내용
    -- 섹션9: 종합 의견 및 인계사항 --
    overall_opinion         str  종합 의견
    handover_items          str  인계사항
    next_patrol_focus       str  차기 순찰 중점 사항
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "안전순찰점검일지"
SHEET_HEADING = "안전순찰 점검 일지"
DOC_ID = "DL-003"

TOTAL_COLS = 10
MAX_PATROL_ROWS = 10
MAX_IMMEDIATE_ROWS = 5
MAX_IMPROVE_ROWS = 8

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL = Font(name="맑은 고딕", size=9)

_FILL_LABEL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE = PatternFill()

_THIN = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LABEL = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value = "" if value is None else value
    cell.font = font or _FONT_DEFAULT
    cell.fill = fill or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vc1: int, vc2: int, height: int = 20) -> None:
    _cell(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vc1, vc2, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, f"▶ {title}",
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_LEFT, height=20)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ---------------------------------------------------------------------------
# 섹션 작성 함수
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    notice_lines = [
        "공식 제출 서식 아님 / 현장 안전순찰 및 지적사항 이행관리 보조서식",
        "산업안전보건법상 안전조치 및 위험성평가 조치사항 관리와 연계",
        "위험요인 발견 시 즉시조치·담당자·기한·이행상태까지 관리",
        "반복 지적사항은 위험성평가 및 TBM 교육에 반영 필요",
        "사고·아차사고·응급조치 발생 시 관련 EM 서식과 별도 연계",
        "개인정보·민감정보 최소 기재 / 사진·영상 등 증빙자료는 별도 보관",
    ]
    for line in notice_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
        c = ws.cell(row=row, column=1, value=line)
        c.font = _FONT_SMALL
        c.fill = _FILL_NOTICE
        c.alignment = _ALIGN_CENTER
        _border_rect(ws, row, 1, row, TOTAL_COLS)
        ws.row_dimensions[row].height = 14
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    doc_id_cell = ws.cell(row=row, column=1, value=f"(문서ID: {DOC_ID})")
    doc_id_cell.font = _FONT_SMALL
    doc_id_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 12
    return row + 1


def _write_doc_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "문서 기본정보")
    _lv(ws, row, "공사명", _v(data, "project_name"), 2, 3, 5)
    _lv(ws, row, "현장명", _v(data, "site_name"), 6, 7, 10)
    row += 1
    _lv(ws, row, "순찰 일자", _v(data, "patrol_date"), 2, 3, 4)
    _lv(ws, row, "시작", _v(data, "patrol_time_start"), 5, 6, 6)
    _lv(ws, row, "종료", _v(data, "patrol_time_end"), 7, 8, 8)
    _lv(ws, row, "날씨", _v(data, "weather"), 9, 10, 10)
    row += 1
    _lv(ws, row, "순찰자", _v(data, "patrol_officer"), 2, 3, 4)
    _lv(ws, row, "소속", _v(data, "department"), 5, 6, 6)
    _lv(ws, row, "직책", _v(data, "position"), 7, 8, 8)
    _lv(ws, row, "연락처", _v(data, "contact"), 9, 10, 10)
    row += 1
    _lv(ws, row, "작성자", _v(data, "writer_name"), 2, 3, 4)
    _lv(ws, row, "검토자", _v(data, "reviewer_name"), 5, 6, 7)
    _lv(ws, row, "승인자", _v(data, "approver_name"), 8, 9, 10)
    row += 1
    return row


def _write_patrol_overview(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "순찰 개요")
    _lv(ws, row, "순찰 구간/경로", _v(data, "patrol_route"), 2, 3, 6)
    _lv(ws, row, "기온", _v(data, "temperature"), 7, 8, 10)
    row += 1
    _lv(ws, row, "당일 작업 인원", _v(data, "total_workers"), 2, 3, 5)
    _lv(ws, row, "고위험 작업 여부", _v(data, "high_risk_work_today"), 6, 7, 10)
    row += 1
    return row


def _write_patrol_results(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "구역별 안전순찰 결과")
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "구역/위치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 4, "점검 시간", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 5, 6, "위험요인/결함", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 7, 7, "위험수준", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "완료기한", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "상태", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("patrol_results", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_PATROL_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "area"), height=20)
        _cell(ws, row, 4, 4, _v(item, "check_time"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 5, 6, _v(item, "hazard_found"), height=20)
        _cell(ws, row, 7, 7, _v(item, "risk_level"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 8, 8, _v(item, "responsible_person"), height=20)
        _cell(ws, row, 9, 9, _v(item, "due_date"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "status"), align=_ALIGN_CENTER, height=20)
        row += 1
    return row


def _write_risk_type_check(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "위험유형별 점검")
    _lv(ws, row, "추락·낙하 예방", _v(data, "fall_protection"), 2, 3, 5)
    _lv(ws, row, "감전·전기 안전", _v(data, "electrical_safety"), 6, 7, 10)
    row += 1
    _lv(ws, row, "화재·폭발 예방", _v(data, "fire_prevention"), 2, 3, 5)
    _lv(ws, row, "기계·장비 안전", _v(data, "equipment_safety"), 6, 7, 10)
    row += 1
    _lv(ws, row, "유해화학물질 관리", _v(data, "chemical_safety"), 2, 3, 5)
    _lv(ws, row, "건강장해 요인", _v(data, "health_hazard"), 6, 7, 10)
    row += 1
    _lv(ws, row, "교통·운반 안전", _v(data, "traffic_safety"), 2, 3, 5)
    _lv(ws, row, "기타 위험요인", _v(data, "others_risk"), 6, 7, 10)
    row += 1
    return row


def _write_immediate_actions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "즉시 시정조치 기록")
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "위치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 5, "지적 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 6, 7, "즉시 조치 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "조치자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "조치 시간", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "확인", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("immediate_actions", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_IMMEDIATE_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "area"), height=20)
        _cell(ws, row, 4, 5, _v(item, "issue"), height=20)
        _cell(ws, row, 6, 7, _v(item, "action_taken"), height=20)
        _cell(ws, row, 8, 8, _v(item, "action_by"), height=20)
        _cell(ws, row, 9, 9, _v(item, "action_time"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "confirmed"), align=_ALIGN_CENTER, height=20)
        row += 1
    return row


def _write_improvement_actions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "개선조치 이행관리")
    _cell(ws, row, 1, 1, "No", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 2, 3, "지적사항", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 4, 5, "개선조치 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 6, 6, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 7, 7, "완료 예정일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 8, 8, "완료일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 9, 9, "이행상태", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _cell(ws, row, 10, 10, "확인자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 18
    row += 1

    items = data.get("improvement_actions", [])
    if not isinstance(items, list):
        items = []
    for i in range(MAX_IMPROVE_ROWS):
        item = items[i] if i < len(items) else {}
        _cell(ws, row, 1, 1, str(i + 1), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 2, 3, _v(item, "issue"), height=20)
        _cell(ws, row, 4, 5, _v(item, "improvement"), height=20)
        _cell(ws, row, 6, 6, _v(item, "assignee"), height=20)
        _cell(ws, row, 7, 7, _v(item, "due_date"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 8, 8, _v(item, "completed_date"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 9, 9, _v(item, "status"), align=_ALIGN_CENTER, height=20)
        _cell(ws, row, 10, 10, _v(item, "confirmed_by"), height=20)
        row += 1
    return row


def _write_repeat_issues(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "반복 지적 및 재발 위험")
    _lv(ws, row, "반복 지적사항", _v(data, "repeat_issues"), 2, 3, 6)
    _lv(ws, row, "재발 위험 수준", _v(data, "repeat_risk_level"), 7, 8, 10)
    row += 1
    _lv(ws, row, "위험성평가 반영", _v(data, "ra_reflected"), 2, 3, 5)
    _lv(ws, row, "TBM 교육 반영", _v(data, "tbm_reflected"), 6, 7, 10)
    row += 1
    _lv(ws, row, "근본 원인 분석", _v(data, "root_cause"), 2, 3, 10)
    row += 1
    _lv(ws, row, "재발 방지 대책", _v(data, "prevention_measures"), 2, 3, 10)
    row += 1
    return row


def _write_accident_link(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "사고·아차사고 연계")
    _lv(ws, row, "사고 발생 여부", _v(data, "accident_occurred"), 2, 3, 5)
    _lv(ws, row, "아차사고 발생 여부", _v(data, "near_miss_occurred"), 6, 7, 10)
    row += 1
    _lv(ws, row, "EM 서식 연계 여부", _v(data, "em_form_linked"), 2, 3, 5)
    _lv(ws, row, "연계 EM 서식", _v(data, "em_form_type"), 6, 7, 10)
    row += 1
    _lv(ws, row, "후속 조치 필요", _v(data, "followup_needed"), 2, 3, 5)
    _lv(ws, row, "후속 조치 내용", _v(data, "followup_content"), 6, 7, 10)
    row += 1
    return row


def _write_overall(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "종합 의견 및 인계사항")
    _lv(ws, row, "종합 의견", _v(data, "overall_opinion"), 2, 3, 10)
    row += 1
    _lv(ws, row, "인계사항", _v(data, "handover_items"), 2, 3, 10)
    row += 1
    _lv(ws, row, "차기 중점 사항", _v(data, "next_patrol_focus"), 2, 3, 10)
    row += 1
    return row


def _write_confirmation(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "확인 및 승인")
    _lv(ws, row, "순찰자", _v(data, "patrol_officer"), 2, 3, 4)
    _lv(ws, row, "검토자", _v(data, "reviewer_name"), 5, 6, 7)
    _lv(ws, row, "승인자", _v(data, "approver_name"), 8, 9, 10)
    row += 1
    _lv(ws, row, "순찰 일자", _v(data, "patrol_date"), 2, 3, 5)
    _lv(ws, row, "확인 일자", _v(data, "patrol_date"), 6, 7, 10)
    row += 1
    return row


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_safety_patrol_inspection_log(form_data: Dict[str, Any]) -> bytes:
    """안전순찰 점검 일지 Excel 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    _apply_print_settings(ws)

    r = _write_title(ws, 1)
    r += 1
    r = _write_doc_info(ws, r, form_data)
    r += 1
    r = _write_patrol_overview(ws, r, form_data)
    r += 1
    r = _write_patrol_results(ws, r, form_data)
    r += 1
    r = _write_risk_type_check(ws, r, form_data)
    r += 1
    r = _write_immediate_actions(ws, r, form_data)
    r += 1
    r = _write_improvement_actions(ws, r, form_data)
    r += 1
    r = _write_repeat_issues(ws, r, form_data)
    r += 1
    r = _write_accident_link(ws, r, form_data)
    r += 1
    r = _write_overall(ws, r, form_data)
    r += 1
    r = _write_confirmation(ws, r, form_data)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
