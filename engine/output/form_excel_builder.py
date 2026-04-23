"""
KRAS 표준 위험성평가표 본표 — Excel 출력 모듈 (v1).

Input:
  - form_data: engine.kras_connector.form_builder.build_risk_assessment_form() 출력
    (kras_standard_form_v1 스키마)

Output:
  - xlsx bytes (in-memory 바이너리). 파일 저장은 호출자 책임.

Layout spec: docs/standards/excel_layout_spec.md
Form schema: docs/standards/kras_standard_form_v1.md

Principles:
- form_builder 로직 수정 금지, mapper/enrichment/API/라우터 무변경.
- 공종(work_category) 컬럼 추가 금지.
- null 은 빈 셀로 출력, 임의 값 생성 금지.
- references_detail 은 시트 비출력.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_TITLE = "위험성평가표"
SHEET_HEADING = "위험성평가표 (실시표)"

# (컬럼명, form_data 경로 키, 폭, 정렬)
# 경로 키가 "_BLANK" 이면 v1 에서는 항상 빈칸.
_COLUMNS: List[Tuple[str, str, int, str]] = [
    ("번호",                "no",                      6,  "center"),
    ("공정명",              "process",                 14, "left"),
    ("세부작업명",          "sub_work",                20, "left"),
    ("위험분류(대)",        "hazard_category_major",   12, "left"),
    ("위험분류(중)",        "hazard_category_minor",   12, "left"),
    ("유해위험요인",        "hazard",                  20, "left"),
    ("관련근거(법적기준)",  "legal_basis",             32, "left"),
    ("현재의 안전보건조치", "current_measures",        24, "left"),
    ("평가척도",            "risk_scale",              8,  "center"),
    ("가능성(빈도)",        "probability",             8,  "center"),
    ("중대성(강도)",        "severity",                8,  "center"),
    ("위험성",              "risk_level",              8,  "center"),
    ("위험성 감소대책",     "control_measures",        40, "left"),
    ("개선후 위험성",       "residual_risk_level",     10, "center"),
    ("개선 예정일",         "target_date",             12, "center"),
    ("완료일",              "completion_date",         12, "center"),
    ("담당자",              "responsible_person",      16, "left"),
]

_META_ROWS = [
    [("회사명",  "company_name"),    ("업종",     "industry")],
    [("현장명",  "site_name"),       ("대표자",   "representative")],
    [("평가종류", "assessment_type"), ("평가일자", "assessment_date")],
    [("작업유형", "work_type"),       ("",          None)],
]

_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_BOLD    = Font(name="맑은 고딕", size=10, bold=True)
_FONT_TITLE   = Font(name="맑은 고딕", size=14, bold=True)

_FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_LABEL  = PatternFill(fill_type="solid", fgColor="F2F2F2")

_THIN = Side(border_style="thin", color="808080")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _cell_value(row: Dict[str, Any], key: str) -> Any:
    """form_data row 에서 key 값을 꺼내 빈칸 친화적으로 변환."""
    v = row.get(key)
    if v is None:
        return ""
    if key == "control_measures" and isinstance(v, list):
        return "\n".join(str(c) for c in v if c)
    return v


def _write_title(ws, total_cols: int) -> None:
    ws.cell(row=1, column=1, value=SHEET_HEADING).font = _FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1).alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 26


def _write_meta(ws, header: Dict[str, Any], total_cols: int) -> int:
    """
    2~5행: 메타 영역.
    각 메타행은 [레이블1|값1 ... |레이블2|값2 ...] 2그룹 레이아웃.
    반환: 메타 다음(공백 포함) 시작 행 번호 (= 7).
    """
    if total_cols < 6:
        label_w1, value_w1 = 1, 2
        label_w2, value_w2 = 1, 2
    else:
        # 17컬럼 기준: label 2 + value 6 | label 2 + value 7
        label_w1, value_w1 = 2, 6
        label_w2, value_w2 = 2, 7

    for i, pair in enumerate(_META_ROWS):
        excel_row = 2 + i

        col = 1
        (label1, key1) = pair[0]
        _write_label_and_value(ws, excel_row, col, label_w1, value_w1, label1, header.get(key1) if key1 else "")
        col += label_w1 + value_w1

        (label2, key2) = pair[1]
        _write_label_and_value(ws, excel_row, col, label_w2, value_w2, label2, header.get(key2) if key2 else "")

        ws.row_dimensions[excel_row].height = 20

    # 6행 공백 구분자
    ws.row_dimensions[6].height = 6
    return 7


def _write_label_and_value(ws, row: int, start_col: int, label_w: int, value_w: int,
                           label: str, value: Any) -> None:
    """
    [label 셀(병합)] [value 셀(병합)] 1 라인 작성.
    label 이 공백이면 빈 영역으로만 두되 테두리는 유지.
    """
    label_end = start_col + label_w - 1
    value_start = label_end + 1
    value_end = value_start + value_w - 1

    # 레이블
    if label:
        cell = ws.cell(row=row, column=start_col, value=label)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_LABEL
        cell.alignment = _ALIGN_LABEL
    if label_w > 1:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=label_end)

    # 값
    v = "" if value is None else value
    vcell = ws.cell(row=row, column=value_start, value=v)
    vcell.font = _FONT_DEFAULT
    vcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    if value_w > 1:
        ws.merge_cells(start_row=row, start_column=value_start, end_row=row, end_column=value_end)

    # 테두리
    for c in range(start_col, value_end + 1):
        ws.cell(row=row, column=c).border = _BORDER_ALL


def _write_header_row(ws, excel_row: int) -> None:
    for idx, (name, _key, _width, _align) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=excel_row, column=idx, value=name)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_ALL
    ws.row_dimensions[excel_row].height = 32


def _write_data_rows(ws, rows: List[Dict[str, Any]], start_row: int) -> None:
    for r_i, row in enumerate(rows):
        excel_row = start_row + r_i
        for c_i, (_name, key, _width, align) in enumerate(_COLUMNS, start=1):
            value = _cell_value(row, key)
            cell = ws.cell(row=excel_row, column=c_i, value=value)
            cell.font = _FONT_DEFAULT
            cell.border = _BORDER_ALL
            cell.alignment = _ALIGN_CENTER if align == "center" else _ALIGN_LEFT
        ws.row_dimensions[excel_row].height = 60


def _apply_column_widths(ws) -> None:
    for idx, (_name, _key, width, _align) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def render_form_sheet(ws, form_data: Dict[str, Any]) -> None:
    """
    주어진 worksheet 에 공식 본표 헤더 구조로 form_data 를 렌더링.

    외부 호출자가 여러 시트(여러 케이스)를 1 워크북에 모으고 싶을 때 사용.
    build_form_excel 은 이 함수 위에 단일 시트 워크북을 래핑한 것.
    """
    header = (form_data or {}).get("header") or {}
    rows = (form_data or {}).get("rows") or []

    total_cols = len(_COLUMNS)

    _apply_column_widths(ws)
    _write_title(ws, total_cols)
    header_row_number = _write_meta(ws, header, total_cols)
    _write_header_row(ws, header_row_number)
    _write_data_rows(ws, rows, header_row_number + 1)

    ws.freeze_panes = ws.cell(row=header_row_number + 1, column=1)


def build_form_excel(form_data: Dict[str, Any]) -> bytes:
    """
    form_builder 출력 dict 를 받아 공식 본표 헤더 구조의 xlsx 바이너리를 반환.

    Args:
        form_data: kras_standard_form_v1 스키마의 dict.

    Returns:
        xlsx 파일 바이트.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE

    render_form_sheet(ws, form_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
