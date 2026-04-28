"""
기상 조건 기록 일지 — Excel 출력 모듈 (v1).

법적 근거:
    산업안전보건기준에 관한 규칙 제37조 (악천후 및 강풍 시 작업 중지)
    산업안전보건기준에 관한 규칙 제566조~제571조 (폭염작업 건강장해 예방조치)
    산업안전보건기준에 관한 규칙 제38조 (강풍 시 타워크레인 작업 중지)

분류: PRACTICAL — 공식 별지 제출서식 아님
      기상 위험 및 작업중지 판단 기록 보조서식
      산업안전보건기준에 관한 규칙 제37조 악천후 및 강풍 시 작업 중지와 연계

Required form_data keys:
    site_name       str  현장명
    record_date     str  기록 일자
    recorder        str  기록자 성명

Optional form_data keys:
    project_name            str  공사명
    department              str  소속
    position                str  직책
    total_workers           str  당일 근로자 수
    -- 섹션2: 기상 관측 정보 --
    obs_time_am             str  오전 관측 시간
    obs_time_pm             str  오후 관측 시간
    temperature_am          str  오전 기온(°C)
    temperature_pm          str  오후 기온(°C)
    humidity_am             str  오전 습도(%)
    humidity_pm             str  오후 습도(%)
    wind_speed_am           str  오전 풍속(m/s)
    wind_speed_pm           str  오후 풍속(m/s)
    wind_gust_am            str  오전 순간최대풍속(m/s)
    wind_gust_pm            str  오후 순간최대풍속(m/s)
    wind_direction          str  풍향
    precipitation           str  강수량(mm)
    precipitation_type      str  강수 형태 (비/눈/우박 등)
    snow_depth              str  적설 깊이(cm)
    visibility              str  가시거리(m)
    weather_forecast        str  기상 예보 요약
    weather_source          str  기상 정보 출처 (기상청/기상앱 등)
    -- 섹션3: 작업 영향 평가 --
    risk_level              str  기상 위험 수준 (정상/주의/경보/위험)
    affected_work_types     str  영향받는 작업 종류
    high_risk_equipment     str  고위험 장비 현황 (타워크레인 등)
    risk_assessment_summary str  위험 평가 요약
    -- 섹션4: 작업중지 판단 기록 --
    work_stop_decided       str  작업중지 결정 여부
    work_stop_time          str  작업중지 시간
    work_stop_scope         str  중지 대상 작업
    work_stop_reason        str  중지 사유
    work_stop_decision_by   str  결정자 성명/직책
    workers_evacuated       str  근로자 대피 여부
    evacuation_location     str  대피 장소
    -- 섹션5: 폭염·한파 건강장해 예방조치 --
    heat_index              str  열지수(°C)
    heat_alert_level        str  폭염 특보 수준 (없음/주의보/경보)
    cooling_measures        str  냉방·통풍 조치 내용
    work_hour_adjustment    str  작업시간 조정 내용
    rest_time_provided      str  휴식시간 부여 여부 및 주기
    water_supply            str  음료수 제공 여부
    heat_illness_symptoms   str  온열질환 증상 근로자 여부
    cold_alert_level        str  한파 특보 수준
    cold_prevention_measures str 한파 예방조치 내용
    -- 섹션6: 강풍·강우·강설 안전조치 --
    crane_work_suspended    str  타워크레인 등 장비 작업 중지 여부
    crane_wind_criterion    str  타워크레인 작업중지 순간풍속 기준(m/s)
    scaffold_checked        str  비계·거푸집 점검 여부
    drainage_measures       str  배수·누수 조치 내용
    slope_stability_checked str  사면·절토면 점검 여부
    snow_removal_done       str  제설 작업 실시 여부
    slippery_prevention     str  미끄럼 방지 조치
    -- 섹션7: 작업 재개 판단 --
    resume_decided          str  작업 재개 결정 여부
    resume_time             str  작업 재개 시간
    resume_check_weather    str  재개 전 기상 확인 여부
    resume_site_inspection  str  재개 전 현장점검 여부
    resume_decision_by      str  재개 결정자
    resume_conditions       str  재개 판단 기준·조건
    -- 섹션8: 조치사항 및 이행관리 --
    actions                 list[dict]  조치사항 목록 (max 8)
        seq                 int  번호
        category            str  구분 (작업중지/대피/장비조치/기타)
        content             str  조치 내용
        responsible_person  str  담당자
        action_time         str  조치 시간
        status              str  이행 상태
        evidence            str  증빙자료 여부
        remarks             str  비고
    -- 섹션9: 인계사항 --
    handover_items          str  인계사항 내용
    next_watch_focus        str  차기 기상 모니터링 중점 사항
    -- 섹션10: 확인 및 승인 --
    writer_name             str  작성자 서명
    reviewer_name           str  검토자 서명
    approver_name           str  승인자 서명
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "기상조건기록일지"
SHEET_HEADING = "기상 조건 기록 일지"
DOC_ID = "DL-004"

TOTAL_COLS = 10
MAX_ACTION_ROWS = 8

_FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT = Font(name="맑은 고딕", size=10)
_FONT_SMALL = Font(name="맑은 고딕", size=9)

_FILL_LABEL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="BDD7EE")
_FILL_NOTICE = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NONE = PatternFill()

_THIN = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_LABEL = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 12, 3: 13, 4: 11, 5: 11,
    6: 11, 7: 11, 8: 11, 9: 11, 10: 10,
}


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _cell(ws, row: int, col1: int, col2: int, value: Any, *,
          font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value = "" if value is None else value
    cell.font = font or _FONT_DEFAULT
    cell.fill = fill or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _lv(ws, row: int, label: str, value: Any,
        lc: int, vc1: int, vc2: int, height: int = 20) -> None:
    _cell(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _cell(ws, row, vc1, vc2, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = height


def _section(ws, row: int, title: str) -> int:
    _cell(ws, row, 1, TOTAL_COLS, f"▶ {title}",
          font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_LEFT, height=20)
    return row + 1


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_print_settings(ws) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ---------------------------------------------------------------------------
# 섹션 작성 함수
# ---------------------------------------------------------------------------

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1

    # 법적 근거 및 분류 안내
    notice_lines = (
        "【 공식 제출 서식 아님 】  기상 위험 및 작업중지 판단 기록 보조서식 | "
        f"문서번호: {DOC_ID}"
    )
    _cell(ws, row, 1, TOTAL_COLS, notice_lines,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1

    basis = (
        "산업안전보건기준에 관한 규칙 제37조 악천후 및 강풍 시 작업 중지와 연계 | "
        "비·눈·바람 등 기상상태로 근로자 위험 우려 시 작업중지 판단 필요"
    )
    _cell(ws, row, 1, TOTAL_COLS, basis,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_CENTER, height=18)
    row += 1
    return row


def _write_basic_info(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "1. 문서 기본정보")
    _lv(ws, row, "현장명", _v(data, "site_name"), 1, 2, 5)
    _lv(ws, row, "공사명", _v(data, "project_name"), 6, 7, 10)
    row += 1
    _lv(ws, row, "기록 일자", _v(data, "record_date"), 1, 2, 3)
    _lv(ws, row, "기록자", _v(data, "recorder"), 4, 5, 5)
    _lv(ws, row, "소속/직책", f"{_v(data,'department')} / {_v(data,'position')}", 6, 7, 8)
    _lv(ws, row, "당일 근로자 수", _v(data, "total_workers"), 9, 10, 10)
    row += 1
    return row


def _write_observation(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "2. 기상 관측 정보")

    # 헤더
    for c, hdr in enumerate(["구분", "오전", "오후", "구분", "오전", "오후",
                               "강수량(mm)", "강수형태", "적설(cm)", "가시거리(m)"], 1):
        _cell(ws, row, c, c, hdr, font=_FONT_BOLD, fill=_FILL_HEADER,
              align=_ALIGN_CENTER, height=18)
    row += 1

    obs_rows = [
        ("관측시간", "obs_time_am", "obs_time_pm",
         "풍속(m/s)", "wind_speed_am", "wind_speed_pm"),
        ("기온(°C)", "temperature_am", "temperature_pm",
         "순간최대풍속", "wind_gust_am", "wind_gust_pm"),
        ("습도(%)", "humidity_am", "humidity_pm",
         "풍향", "wind_direction", ""),
    ]
    extra = {
        "강수량(mm)": "precipitation",
        "강수형태": "precipitation_type",
        "적설(cm)": "snow_depth",
        "가시거리(m)": "visibility",
    }
    extra_keys = list(extra.values())
    extra_vals = [_v(data, k) for k in extra_keys]

    for i, (l1, k1a, k1p, l2, k2a, k2p) in enumerate(obs_rows):
        _cell(ws, row, 1, 1, l1, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 2, 2, _v(data, k1a), align=_ALIGN_CENTER)
        _cell(ws, row, 3, 3, _v(data, k1p), align=_ALIGN_CENTER)
        _cell(ws, row, 4, 4, l2, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        v2a = _v(data, k2a) if k2a else ""
        v2p = _v(data, k2p) if k2p else ""
        _cell(ws, row, 5, 5, v2a, align=_ALIGN_CENTER)
        _cell(ws, row, 6, 6, v2p, align=_ALIGN_CENTER)
        # 나머지 4열에 extra 값
        ev = extra_vals[i] if i < len(extra_vals) else ""
        el = list(extra.keys())[i] if i < len(extra) else ""
        _cell(ws, row, 7, 7, ev, align=_ALIGN_CENTER)
        _cell(ws, row, 8, 8, "", align=_ALIGN_CENTER)
        _cell(ws, row, 9, 9, "", align=_ALIGN_CENTER)
        _cell(ws, row, 10, 10, "", align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 18
        row += 1

    # 기상정보 출처/예보
    _lv(ws, row, "기상 예보", _v(data, "weather_forecast"), 1, 2, 7)
    _lv(ws, row, "정보 출처", _v(data, "weather_source"), 8, 9, 10)
    row += 1
    return row


def _write_impact_assessment(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "3. 작업 영향 평가")
    _lv(ws, row, "위험 수준", _v(data, "risk_level"), 1, 2, 3)
    _lv(ws, row, "영향 작업 종류", _v(data, "affected_work_types"), 4, 5, 10)
    row += 1
    _lv(ws, row, "고위험 장비", _v(data, "high_risk_equipment"), 1, 2, 5)
    _lv(ws, row, "위험 평가 요약", _v(data, "risk_assessment_summary"), 6, 7, 10)
    row += 1
    return row


def _write_work_stop(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "4. 작업중지 판단 기록")
    _lv(ws, row, "작업중지 결정", _v(data, "work_stop_decided"), 1, 2, 3)
    _lv(ws, row, "중지 시간", _v(data, "work_stop_time"), 4, 5, 5)
    _lv(ws, row, "결정자", _v(data, "work_stop_decision_by"), 6, 7, 10)
    row += 1
    _lv(ws, row, "중지 대상 작업", _v(data, "work_stop_scope"), 1, 2, 5)
    _lv(ws, row, "중지 사유", _v(data, "work_stop_reason"), 6, 7, 10)
    row += 1
    _lv(ws, row, "근로자 대피", _v(data, "workers_evacuated"), 1, 2, 3)
    _lv(ws, row, "대피 장소", _v(data, "evacuation_location"), 4, 5, 10)
    row += 1
    return row


def _write_heat_cold(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "5. 폭염·한파 건강장해 예방조치")

    notice = (
        "폭염작업 시 냉방·통풍, 작업시간 조정, 휴식시간 부여 등 건강장해 예방조치 필요 "
        "(산업안전보건기준에 관한 규칙 제566조~제571조)"
    )
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "열지수(°C)", _v(data, "heat_index"), 1, 2, 3)
    _lv(ws, row, "폭염 특보", _v(data, "heat_alert_level"), 4, 5, 5)
    _lv(ws, row, "한파 특보", _v(data, "cold_alert_level"), 6, 7, 10)
    row += 1
    _lv(ws, row, "냉방·통풍 조치", _v(data, "cooling_measures"), 1, 2, 10, height=22)
    row += 1
    _lv(ws, row, "작업시간 조정", _v(data, "work_hour_adjustment"), 1, 2, 5)
    _lv(ws, row, "휴식시간 부여", _v(data, "rest_time_provided"), 6, 7, 10)
    row += 1
    _lv(ws, row, "음료수 제공", _v(data, "water_supply"), 1, 2, 3)
    _lv(ws, row, "온열질환 증상자", _v(data, "heat_illness_symptoms"), 4, 5, 10)
    row += 1
    _lv(ws, row, "한파 예방조치", _v(data, "cold_prevention_measures"), 1, 2, 10, height=22)
    row += 1
    return row


def _write_wind_rain_snow(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "6. 강풍·강우·강설 안전조치")

    notice = (
        "타워크레인 등 장비 작업은 순간풍속 기준 확인 필요 "
        "(산업안전보건기준에 관한 규칙 제38조)"
    )
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "크레인 작업 중지", _v(data, "crane_work_suspended"), 1, 2, 3)
    _lv(ws, row, "순간풍속 기준", _v(data, "crane_wind_criterion"), 4, 5, 5)
    _lv(ws, row, "비계·거푸집 점검", _v(data, "scaffold_checked"), 6, 7, 10)
    row += 1
    _lv(ws, row, "배수·누수 조치", _v(data, "drainage_measures"), 1, 2, 5)
    _lv(ws, row, "사면 점검", _v(data, "slope_stability_checked"), 6, 7, 10)
    row += 1
    _lv(ws, row, "제설 작업", _v(data, "snow_removal_done"), 1, 2, 3)
    _lv(ws, row, "미끄럼 방지", _v(data, "slippery_prevention"), 4, 5, 10)
    row += 1
    return row


def _write_resume(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "7. 작업 재개 판단")

    notice = "작업 재개 전 기상 확인 및 현장점검 필요"
    _cell(ws, row, 1, TOTAL_COLS, notice,
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1

    _lv(ws, row, "재개 결정", _v(data, "resume_decided"), 1, 2, 3)
    _lv(ws, row, "재개 시간", _v(data, "resume_time"), 4, 5, 5)
    _lv(ws, row, "재개 결정자", _v(data, "resume_decision_by"), 6, 7, 10)
    row += 1
    _lv(ws, row, "기상 재확인", _v(data, "resume_check_weather"), 1, 2, 3)
    _lv(ws, row, "현장점검", _v(data, "resume_site_inspection"), 4, 5, 5)
    _lv(ws, row, "재개 판단 기준", _v(data, "resume_conditions"), 6, 7, 10)
    row += 1
    return row


def _write_actions(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "8. 조치사항 및 이행관리")

    headers = ["번호", "구분", "조치 내용", "담당자", "조치 시간", "이행 상태", "증빙", "비고"]
    col_spans = [(1, 1), (2, 3), (4, 6), (7, 7), (8, 8), (9, 9), (10, 10), (10, 10)]
    # 8개 컬럼을 10열에 배분
    col_map = [
        (1, 1), (2, 2), (3, 5), (6, 6), (7, 7), (8, 8), (9, 9), (10, 10)
    ]
    for (c1, c2), hdr in zip(col_map, headers):
        _cell(ws, row, c1, c2, hdr, font=_FONT_BOLD, fill=_FILL_HEADER,
              align=_ALIGN_CENTER, height=18)
    row += 1

    actions = data.get("actions") or []
    for i in range(MAX_ACTION_ROWS):
        act = actions[i] if i < len(actions) else {}
        vals = [
            act.get("seq", i + 1),
            act.get("category", ""),
            act.get("content", ""),
            act.get("responsible_person", ""),
            act.get("action_time", ""),
            act.get("status", ""),
            act.get("evidence", ""),
            act.get("remarks", ""),
        ]
        for (c1, c2), val in zip(col_map, vals):
            _cell(ws, row, c1, c2, val, align=_ALIGN_CENTER, height=18)
        row += 1

    notice = "사진·기상자료 등 증빙자료는 별도 보관"
    _cell(ws, row, 1, TOTAL_COLS, f"※ {notice}",
          font=_FONT_SMALL, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
    row += 1
    return row


def _write_handover(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "9. 인계사항")
    _lv(ws, row, "인계사항", _v(data, "handover_items"), 1, 2, 10, height=40)
    row += 1
    _lv(ws, row, "차기 모니터링\n중점사항", _v(data, "next_watch_focus"), 1, 2, 10, height=30)
    row += 1
    return row


def _write_approval(ws, row: int, data: Dict[str, Any]) -> int:
    row = _section(ws, row, "10. 확인 및 승인")

    for lbl, key in [("작성자", "writer_name"),
                     ("검토자", "reviewer_name"),
                     ("승인자", "approver_name")]:
        _cell(ws, row, 1, 2, lbl, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 3, 4, _v(data, key), align=_ALIGN_CENTER)
        _cell(ws, row, 5, 6, "서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_CENTER)
        _cell(ws, row, 7, 10, "", align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 30
        row += 1
    return row


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def build_weather_condition_log(form_data: Dict[str, Any]) -> bytes:
    """form_data → xlsx bytes."""
    data = form_data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _apply_col_widths(ws)
    _apply_print_settings(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_basic_info(ws, row, data)
    row = _write_observation(ws, row, data)
    row = _write_impact_assessment(ws, row, data)
    row = _write_work_stop(ws, row, data)
    row = _write_heat_cold(ws, row, data)
    row = _write_wind_rain_snow(ws, row, data)
    row = _write_resume(ws, row, data)
    row = _write_actions(ws, row, data)
    row = _write_handover(ws, row, data)
    row = _write_approval(ws, row, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
