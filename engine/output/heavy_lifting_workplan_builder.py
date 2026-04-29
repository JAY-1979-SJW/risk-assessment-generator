"""
중량물 취급 작업계획서 — Excel 출력 모듈 (v1).

법적 근거:
    산업안전보건기준에 관한 규칙 제38조 제1항 제11호 + 별표4 (중량물 취급 작업계획서 의무)
    제39조  (작업지휘자 지정)
    제40조  (신호방법 지정 — 중량물)
    제132조 (양중기 종류 및 정의)
    제133조 (정격하중 표시 의무)
    제163조~제170조 (달기기구 안전기준 — 와이어로프·훅·샤클·달기체인·슬링벨트)
    제385조  (중량물 취급 시 하역운반기계 사용 의무)

분류: PRACTICAL — 법정 별지 서식 없음.
    법령 의무 항목 기반 자체 표준서식.
    evidence_status: NEEDS_VERIFICATION

선행/후속 관계:
    이 서식(WP-005)은 PTW-007 중량물 인양·중장비사용 작업 허가서의 선행 작업계획서.
    PTW-007 허가 발급 전 WP-005 사전 작성·확인이 요구됨.

Required form_data keys:
    object_name      str  중량물 명칭                  [법정] 제38조+별표4
    object_weight    str  중량물 중량 (ton/kg)          [법정] 별표4
    work_method      str  작업방법                      [법정] 제38조+별표4
    emergency_measure str 비상조치 방법                 [법정] 제38조+별표4

Optional form_data keys:
    site_name         str  사업장명(현장명)
    project_name      str  공사명
    work_location     str  작업 위치
    work_date         str  작업 기간
    supervisor        str  작업책임자 (작업지휘자)
    contractor        str  도급업체
    prepared_by       str  작성자
    sign_date         str  작성일

    # 중량물 정보
    object_size        str  중량물 크기 (mm/m)
    object_shape       str  중량물 형상
    weight_basis       str  중량 산정 근거 (설계도서·계측)
    center_of_gravity  str  무게중심 위치

    # 운반/인양 경로
    transport_route    str  운반·인양 경로
    route_sketch_note  str  경로 스케치 안내 문구 (없으면 기본값 사용)

    # 작업장소 및 지반 상태
    work_site_condition  str  작업장소 상태
    ground_condition     str  지반 상태 (지지력·매설물)
    access_control       str  출입통제 방법

    # 사용 장비 및 보조기구
    equipment_name       str  주 양중기명 (크레인·호이스트 등)
    equipment_capacity   str  장비 정격하중 (ton)
    auxiliary_equipment  str  보조기구 (체인블록·전동호이스트 등)

    # 줄걸이/달기기구
    rigging_method       str  줄걸이 방법 (2줄걸이·4줄걸이 등)
    rigging_gear         str  사용 달기구 (와이어로프·슬링벨트·샤클 등)
    rigging_angle        str  슬링 각도 (°)
    rigging_safety_coeff str  안전계수 확인 여부

    # 작업 인원 및 역할
    work_commander       str  작업지휘자 성명 (제39조)
    signal_worker        str  신호수·유도자 성명 (제40조)
    worker_roles         str  작업자 역할 분담

    # 위험요인 및 방지대책
    fall_prevention      str  추락 방지대책
    drop_prevention      str  낙하 방지대책
    tipping_prevention   str  전도 방지대책
    pinch_prevention     str  협착 방지대책
    collapse_prevention  str  붕괴 방지대책

    # 작업 전 점검사항
    pre_work_check_items  str  작업 전 점검 사항 (자유기재)

    # 사진 권장
    photo_items           str  증빙 사진 권장 목록

    safety_steps  list[dict]  작업단계별 위험요소/안전조치 (max 10)
        task_step      str  작업단계
        hazard         str  위험요소
        safety_measure str  안전조치
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "중량물취급작업계획서"
SHEET_HEADING = "중량물 취급 작업계획서"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제38조 제1항 제11호 + 별표4에 따른 작업계획서"

MAX_STEPS  = 10
TOTAL_COLS = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_SKETCH  = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FCE4D6")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center", wrap_text=True)

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 10}

_L1, _V1_START, _V1_END        = 1, 2, 4
_L2, _V2_START, _V2_END        = 5, 6, 8
_FULL_VAL_START, _FULL_VAL_END = 2, 8

ROUTE_SKETCH_DEFAULT = (
    "※ 운반·인양 경로 및 위험구간을 아래 공간에 수기로 기재하시오 "
    "(장비 위치·인양 경로·출입통제 구간 표시)"
)


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    cell = ws.cell(row=row, column=1, value=SHEET_HEADING)
    cell.font = _FONT_TITLE
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    sub = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    sub.font = _FONT_SUBTITLE
    sub.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_meta_block(ws, start_row: int, data: Dict[str, Any]) -> int:
    H, r = 20, start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "기본 정보",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_lv(ws, r, "사업장명", _v(data, "site_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "공사명",  _v(data, "project_name"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 위치", _v(data, "work_location"), _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "작업 기간", _v(data, "work_date"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작업지휘자", _v(data, "supervisor"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    _write_lv(ws, r, "도급업체", _v(data, "contractor"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "작성자",   _v(data, "prepared_by"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = H; r += 1
    return r


def _write_object_info(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS,
                "작업 개요 및 중량물 정보  (기준규칙 제38조+별표4)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "중량물 명칭", _v(data, "object_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "중량물 형상", _v(data, "object_shape"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "중량 (ton/kg)", _v(data, "object_weight"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "크기 (mm/m)", _v(data, "object_size"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_cell(ws, r, _L1, _L1, "중량 산정\n근거",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=40)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "weight_basis"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "무게중심",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=30)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "center_of_gravity"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "작업방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "work_method"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    return r


def _write_route_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "운반·인양 경로  (기준규칙 제38조+별표4)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "운반·인양\n경로",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "transport_route"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    sketch_note = data.get("route_sketch_note") or ROUTE_SKETCH_DEFAULT
    _write_cell(ws, r, 1, TOTAL_COLS, sketch_note,
                font=_FONT_SMALL, fill=_FILL_SKETCH, align=_ALIGN_CENTER, height=16)
    r += 1

    for _ in range(4):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = ""
            cell.fill  = _FILL_NONE
            cell.alignment = _ALIGN_LEFT
            cell.border    = _BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    return r


def _write_site_equipment(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS,
                "작업장소·지반 상태 및 사용 장비  (기준규칙 제38조+별표4)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "작업장소 상태", _v(data, "work_site_condition"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1

    _write_lv(ws, r, "지반 상태\n(지지력·매설물)", _v(data, "ground_condition"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 40; r += 1

    _write_lv(ws, r, "출입통제 방법", _v(data, "access_control"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1

    _write_lv(ws, r, "주 양중기", _v(data, "equipment_name"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "정격하중 (ton)", _v(data, "equipment_capacity"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "보조기구", _v(data, "auxiliary_equipment"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 30; r += 1

    return r


def _write_rigging_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS,
                "줄걸이·달기기구  (기준규칙 제163조~제170조)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "줄걸이 방법", _v(data, "rigging_method"), _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "슬링 각도 (°)", _v(data, "rigging_angle"), _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 20; r += 1

    _write_lv(ws, r, "사용 달기구\n(와이어로프·슬링벨트·샤클)", _v(data, "rigging_gear"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 40; r += 1

    _write_lv(ws, r, "안전계수 확인", _v(data, "rigging_safety_coeff"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 25; r += 1

    return r


def _write_personnel_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS,
                "작업 인원 및 역할  (기준규칙 제39조·제40조)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_lv(ws, r, "작업지휘자\n(제39조)", _v(data, "work_commander"),
              _L1, _V1_START, _V1_END)
    _write_lv(ws, r, "신호수·유도자\n(제40조)", _v(data, "signal_worker"),
              _L2, _V2_START, _V2_END)
    ws.row_dimensions[r].height = 40; r += 1

    _write_lv(ws, r, "작업자 역할\n분담", _v(data, "worker_roles"),
              _L1, _FULL_VAL_START, _FULL_VAL_END)
    ws.row_dimensions[r].height = 40; r += 1

    return r


def _write_hazard_section(ws, start_row: int, data: Dict[str, Any],
                           steps: List[Dict[str, Any]]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS,
                "위험요인 및 방지대책",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    _write_cell(ws, r, _L1, _L1, "추락·낙하\n방지대책",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    fall_drop = (
        (_v(data, "fall_prevention") or "") +
        ("\n" if _v(data, "fall_prevention") and _v(data, "drop_prevention") else "") +
        (_v(data, "drop_prevention") or "")
    ) or ""
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, fall_drop,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    _write_cell(ws, r, _L1, _L1, "전도·협착·붕괴\n방지대책",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=50)
    tpc = (
        (_v(data, "tipping_prevention") or "") +
        ("\n" if (_v(data, "tipping_prevention") and _v(data, "pinch_prevention")) else "") +
        (_v(data, "pinch_prevention") or "") +
        ("\n" if (_v(data, "pinch_prevention") and _v(data, "collapse_prevention")) else "") +
        (_v(data, "collapse_prevention") or "")
    ) or ""
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, tpc,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1

    # 작업단계별 위험요소/안전조치 표
    _write_cell(ws, r, 1, 1, "순번",    font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _write_cell(ws, r, 2, 2, "작업단계", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 3, 5, "위험요소", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, r, 6, 8, "안전조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    r += 1

    for i in range(MAX_STEPS):
        item = steps[i] if i < len(steps) else {}
        _write_cell(ws, r, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _write_cell(ws, r, 2, 2, _v(item, "task_step"),      font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 3, 5, _v(item, "hazard"),         font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _write_cell(ws, r, 6, 8, _v(item, "safety_measure"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[r].height = 28
        r += 1

    return r


def _write_precheck_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "작업 전 점검사항",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1

    default_checks = (
        "□ 중량물 중량·크기 확인\n"
        "□ 달기기구(와이어로프·슬링벨트·샤클) 이상 유무 점검\n"
        "□ 양중기 정격하중 확인 (중량물 중량 < 정격하중)\n"
        "□ 지반 상태·아웃트리거 설치 확인\n"
        "□ 신호수·작업지휘자 배치 확인\n"
        "□ 인양경로 내 하부 출입통제 확인\n"
        "□ TBM 실시 확인\n"
        "□ PTW-007 작업허가서 확인"
    )
    check_text = _v(data, "pre_work_check_items") or default_checks
    _write_cell(ws, r, 1, TOTAL_COLS, check_text,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT, height=130)
    r += 1

    return r


def _write_emergency_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "비상조치 계획  (기준규칙 제38조+별표4)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, _L1, _L1, "비상조치\n방법",
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL, height=60)
    _write_cell(ws, r, _FULL_VAL_START, _FULL_VAL_END, _v(data, "emergency_measure"),
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    r += 1
    return r


def _write_photo_section(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "사진·증빙 권장 항목 (OPTIONAL)",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    default_photos = (
        "① 중량물 외관 및 표시 (중량·크기)\n"
        "② 달기기구 체결 상태 (줄걸이·샤클)\n"
        "③ 양중기 배치 및 아웃트리거 설치\n"
        "④ 인양 중 경로 하부 통제 확인\n"
        "⑤ 설치 완료 후 최종 상태"
    )
    photo_text = _v(data, "photo_items") or default_photos
    _write_cell(ws, r, 1, TOTAL_COLS, photo_text,
                font=_FONT_SMALL, align=_ALIGN_LEFT, height=80)
    r += 1
    return r


def _write_notice_section(ws, start_row: int) -> int:
    r = start_row
    notice = (
        "【주의】 ① 이 작업계획서는 PTW-007 중량물 인양·중장비사용 작업 허가서의 선행 서류로, "
        "허가 발급 전 작성·확인되어야 합니다.\n"
        "② 이 서식은 법정 안전보건교육 수료증을 대체하지 않습니다.\n"
        "③ 중량물 취급 작업계획서 (WP-005)는 산안규칙 제38조 제1항 제11호 + 별표4 의거 작성.\n"
        "④ evidence_status: NEEDS_VERIFICATION — 조항 원문 API 미수집"
    )
    _write_cell(ws, r, 1, TOTAL_COLS, notice,
                font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_LEFT, height=70)
    r += 1
    return r


def _write_confirmation(ws, start_row: int, data: Dict[str, Any]) -> int:
    r = start_row
    _write_cell(ws, r, 1, TOTAL_COLS, "확인 및 서명",
                font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    r += 1
    _write_cell(ws, r, 1, 2, "작성자 서명",    font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 3, 4, "",               font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 6, "작업지휘자 서명", font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 7, 7, "작성일",         font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, r, 8, 8, _v(data, "sign_date"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[r].height = 20; r += 1
    _write_cell(ws, r, 1, 4, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    _write_cell(ws, r, 5, 8, "", font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[r].height = 36
    return r + 1


def build_heavy_lifting_workplan_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 중량물 취급 작업계획서 xlsx 바이너리를 반환."""
    data  = form_data or {}
    steps = data.get("safety_steps") or []
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)
    row = 1
    row = _write_title(ws, row)
    row = _write_meta_block(ws, row, data)
    row = _write_object_info(ws, row, data)
    row = _write_route_section(ws, row, data)
    row = _write_site_equipment(ws, row, data)
    row = _write_rigging_section(ws, row, data)
    row = _write_personnel_section(ws, row, data)
    row = _write_hazard_section(ws, row, data, steps)
    row = _write_precheck_section(ws, row, data)
    row = _write_emergency_section(ws, row, data)
    row = _write_photo_section(ws, row, data)
    row = _write_notice_section(ws, row)
    _write_confirmation(ws, row, data)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
