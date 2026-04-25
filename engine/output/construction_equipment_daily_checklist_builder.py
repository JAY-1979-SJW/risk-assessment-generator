"""
건설장비 일일 사전점검표 — Excel 출력 모듈 (v1.0)  [CL-003]

법적 근거:
    [차량계 건설기계]
    산업안전보건기준에 관한 규칙 제196조~제199조 추정
    (NEEDS_VERIFICATION: 조항번호 원문 미수집. 전조등·낙하물 보호구조·전도방지 관련)

    [차량계 하역운반기계등 공통]
    산업안전보건기준에 관한 규칙 제171조~제178조 추정
    (PARTIAL_VERIFIED: WP-009 법령 연계. 공통 안전조치·전조등·전도방지 등)

    [지게차 세부 점검]
    산업안전보건기준에 관한 규칙 제179조~제183조 추정
    (PARTIAL_VERIFIED: WP-009 법령 연계. 전조등·후방확인·헤드가드·백레스트·안전띠 등)

범위:
    본 서식은 건설장비 일일 사용 전/사용 중 안전점검 기록용입니다.
    작업계획서(WP-008/WP-009) 및 장비사용계획서(EQ-001/EQ-002)를 대체하지 않습니다.
    타워크레인·이동식 크레인·비계·거푸집동바리 전용 점검은 별도 서식으로 관리합니다.
    대상 장비: 굴착기·로더·덤프트럭·롤러·콘크리트 펌프카·지게차·구내운반차·고소작업대·화물자동차

Required form_data keys:
    check_date    str  점검 일자
    work_location str  작업 장소
    checker_name  str  점검자 성명

Optional form_data keys:
    site_name               str  사업장명
    project_name            str  공사명
    equipment_type          str  장비 종류 (굴착기/지게차/고소작업대 등)
    equipment_model         str  장비 기종
    equipment_reg_no        str  장비 등록번호 또는 차량번호
    equipment_capacity      str  장비 최대작업용량
    operator_name           str  운전자 성명
    operator_license_no     str  운전자 면허번호
    guide_worker_name       str  유도자 성명
    work_commander_name     str  작업지휘자 성명
    doc_check_items         list[dict]  섹션4 항목 override (item, result, note)
    appearance_items        list[dict]  섹션5 항목 override
    light_items             list[dict]  섹션6 항목 override
    brake_items             list[dict]  섹션7 항목 override
    stability_items         list[dict]  섹션8 항목 override
    contact_items           list[dict]  섹션9 항목 override
    load_items              list[dict]  섹션10 항목 override
    additional_items        list[dict]  섹션11 장비별 추가 점검 override (item, result, note)
    nonconformance_items    list[dict]  부적합 사항 (content, location, action, deadline, completed)
    operator_sign           str  운전자 서명
    guide_worker_sign       str  유도자 서명
    work_commander_sign     str  작업지휘자 서명
    supervisor_sign         str  관리감독자 서명
    manager_sign            str  현장소장 서명
    sign_date               str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "건설장비일일사전점검표"
SHEET_HEADING = "건설장비 일일 사전점검표"
SHEET_SUBTITLE = (
    "차량계 건설기계 및 차량계 하역운반기계 일일 사용 전 안전점검"
    " | 「산업안전보건기준에 관한 규칙」 관련 조항에 따른 점검"
)

NOTICE_SCOPE      = "본 점검표는 건설장비 일일 사전점검용이며, 작업계획서 및 장비사용계획서를 대체하지 않는다."
NOTICE_EXCLUSIVE  = "타워크레인, 이동식 크레인, 비계, 거푸집동바리 전용 점검은 별도 서식으로 관리한다."
NOTICE_NONCONFORM = "부적합 사항은 사용 전 시정 완료 후 재확인한다."
NOTICE_LAW        = "법령 조항은 현행 원문 확인 후 현장에 적용한다."

MAX_NONCONFORM = 5
TOTAL_COLS     = 9


# ──────────────────────────────────────────────────────────────
# 섹션별 기본 점검 항목
# ──────────────────────────────────────────────────────────────

_DOC_DEFAULTS: List[Dict[str, str]] = [
    {"item": "정기검사증(안전검사증) 유효 여부",                    "result": "", "note": ""},
    {"item": "건설기계 등록증·건설기계 보험증권 비치 여부",          "result": "", "note": ""},
    {"item": "작업계획서(차량계 건설기계/하역운반기계) 작성 여부",   "result": "", "note": ""},
    {"item": "장비 제조사 사용설명서 현장 비치 여부",               "result": "", "note": ""},
    {"item": "작업지휘자(또는 유도자) 지정 여부",                   "result": "", "note": ""},
]

_APPEARANCE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "외관 손상·균열·변형 없음 여부",                       "result": "", "note": ""},
    {"item": "엔진오일·유압유·냉각수 누유·누수 없음 여부",           "result": "", "note": ""},
    {"item": "타이어 공기압·손상 또는 궤도(트랙) 상태 이상 없음",    "result": "", "note": ""},
    {"item": "볼트·너트 체결 상태 이상 없음 여부",                  "result": "", "note": ""},
    {"item": "연료 잔량 충분 여부",                                  "result": "", "note": ""},
]

_LIGHT_DEFAULTS: List[Dict[str, str]] = [
    {"item": "전조등 정상 작동 여부",                               "result": "", "note": ""},
    {"item": "후미등 정상 작동 여부",                               "result": "", "note": ""},
    {"item": "경광등(회전등) 설치 및 정상 작동 여부",               "result": "", "note": ""},
    {"item": "후진경보기 또는 후방감지기 정상 작동 여부",           "result": "", "note": ""},
    {"item": "후방 카메라(장착 시) 정상 작동 여부",                 "result": "", "note": ""},
]

_BRAKE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "주 제동장치(풋 브레이크) 작동 이상 없음 여부",        "result": "", "note": ""},
    {"item": "주차 브레이크 작동 이상 없음 여부",                   "result": "", "note": ""},
    {"item": "조향장치(핸들) 유격·작동 이상 없음 여부",             "result": "", "note": ""},
    {"item": "비상정지 스위치 정상 작동 여부",                      "result": "", "note": ""},
    {"item": "안전벨트(좌석안전띠) 착용 가능 상태 여부",            "result": "", "note": ""},
]

_STABILITY_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업 지반 지지력 확인 (연약지반·매립지 여부)",        "result": "", "note": ""},
    {"item": "작업 구역 내 경사·단차·굴착 인접 여부 확인",          "result": "", "note": ""},
    {"item": "전도 위험 구간 통행 금지 및 안전표지 설치 여부",       "result": "", "note": ""},
    {"item": "갓길 붕괴 위험 없음 여부 (하중 집중 구간 확인)",       "result": "", "note": ""},
    {"item": "강풍·우천 등 기상 조건 작업 중지 기준 초과 여부",      "result": "", "note": ""},
]

_CONTACT_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업반경 내 관계자 외 출입 통제 여부",                "result": "", "note": ""},
    {"item": "유도자 배치 위치 및 신호체계 확인 여부",               "result": "", "note": ""},
    {"item": "인접 구조물·가설물·지하매설물 안전거리 확보 여부",     "result": "", "note": ""},
    {"item": "보행자·근로자 동선 분리 여부",                        "result": "", "note": ""},
    {"item": "장비 간 충돌 위험 통제 여부 (협착·충돌 방지)",         "result": "", "note": ""},
]

_LOAD_DEFAULTS: List[Dict[str, str]] = [
    {"item": "적재 화물 최대 적재량 초과 없음 여부",                "result": "", "note": ""},
    {"item": "화물 고정 또는 적재 상태 이상 없음 여부",              "result": "", "note": ""},
    {"item": "인양 또는 하역 작업 시 신호수·작업지휘자 배치 여부",  "result": "", "note": ""},
    {"item": "근로자 탑승 금지 (탑승석 외 탑승 없음) 여부",          "result": "", "note": ""},
    {"item": "작업 시작 전 호이스트·포크·버킷 등 작동 상태 확인",   "result": "", "note": ""},
]

# 장비별 추가 점검 항목 기본값 — 장비 종류별 대표 항목
_ADDITIONAL_DEFAULTS_BY_TYPE: Dict[str, List[Dict[str, str]]] = {
    "굴착기": [
        {"item": "버킷·투스 마모 및 균열 없음 여부",               "result": "", "note": ""},
        {"item": "퀵커플러 안전핀 체결 상태 여부",                  "result": "", "note": ""},
        {"item": "붐·암 균열·변형 없음 여부",                      "result": "", "note": ""},
        {"item": "회전반경 통제 표시(안전로프·라바콘) 설치 여부",   "result": "", "note": ""},
    ],
    "덤프트럭": [
        {"item": "적재함 상태 및 덮개 고정 여부",                   "result": "", "note": ""},
        {"item": "적재함 잠금장치 이상 없음 여부",                  "result": "", "note": ""},
        {"item": "타이어 상태 및 공기압 이상 없음 여부",             "result": "", "note": ""},
        {"item": "적재량 초과 없음 여부 (과적 방지)",               "result": "", "note": ""},
    ],
    "롤러": [
        {"item": "진동장치 정상 작동 여부",                         "result": "", "note": ""},
        {"item": "후방 확인 장치(후방카메라·후진경보기) 작동 여부", "result": "", "note": ""},
        {"item": "전도 위험 구간(경사로) 작업 전 확인 여부",         "result": "", "note": ""},
        {"item": "작업 구역 통제 여부",                             "result": "", "note": ""},
    ],
    "콘크리트 펌프카": [
        {"item": "아웃트리거 전개 상태 및 지반 지지 확인 여부",     "result": "", "note": ""},
        {"item": "붐 균열·변형 없음 여부",                          "result": "", "note": ""},
        {"item": "호스 연결부 누출 없음 여부",                      "result": "", "note": ""},
        {"item": "압송라인 통제 및 작업반경 관계자 외 출입 통제",   "result": "", "note": ""},
    ],
    "지게차": [
        {"item": "포크 균열·변형 없음 여부",                        "result": "", "note": ""},
        {"item": "헤드가드 설치 및 손상 없음 여부",                  "result": "", "note": ""},
        {"item": "백레스트 설치 및 이상 없음 여부",                  "result": "", "note": ""},
        {"item": "좌석안전띠 정상 작동 여부",                       "result": "", "note": ""},
        {"item": "후진경보기 또는 경광등 후방감지기 작동 여부",     "result": "", "note": ""},
        {"item": "전조등·후미등 정상 작동 여부",                    "result": "", "note": ""},
    ],
    "구내운반차": [
        {"item": "적재 상태 이상 없음(과적·편적 없음) 여부",         "result": "", "note": ""},
        {"item": "전도 위험 없음(경사·급커브 등) 여부",              "result": "", "note": ""},
        {"item": "제동장치 정상 작동 여부",                         "result": "", "note": ""},
        {"item": "운행 통로 상태(바닥 파손·장애물 없음) 여부",       "result": "", "note": ""},
    ],
    "고소작업대": [
        {"item": "아웃트리거 전개 및 지반 지지 상태 확인 여부",     "result": "", "note": ""},
        {"item": "작업대(버킷·플랫폼) 난간 이상 없음 여부",         "result": "", "note": ""},
        {"item": "비상하강장치 정상 작동 여부",                     "result": "", "note": ""},
        {"item": "과상승 방지 장치 정상 작동 여부",                  "result": "", "note": ""},
        {"item": "탑승자 안전대·안전띠 착용 가능 상태 여부",        "result": "", "note": ""},
    ],
}

# 장비 종류 미지정 시 공통 기본값
_ADDITIONAL_DEFAULTS_COMMON: List[Dict[str, str]] = [
    {"item": "장비 특성에 따른 추가 점검 항목 1 (해당 항목 기재)",  "result": "", "note": ""},
    {"item": "장비 특성에 따른 추가 점검 항목 2 (해당 항목 기재)",  "result": "", "note": ""},
    {"item": "장비 특성에 따른 추가 점검 항목 3 (해당 항목 기재)",  "result": "", "note": ""},
]


# ──────────────────────────────────────────────────────────────
# 스타일 헬퍼
# ──────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_border() -> Border:
    m = Side(style="medium")
    return Border(left=m, right=m, top=m, bottom=m)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="1F4E79")


def _sub_fill() -> PatternFill:
    return PatternFill("solid", fgColor="2E75B6")


def _section_fill() -> PatternFill:
    return PatternFill("solid", fgColor="BDD7EE")


def _notice_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FFF2CC")


def _warn_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FCE4D6")


def _center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _white_bold(sz: int = 11) -> Font:
    return Font(bold=True, color="FFFFFF", size=sz)


def _black_bold(sz: int = 10) -> Font:
    return Font(bold=True, color="000000", size=sz)


def _black_normal(sz: int = 9) -> Font:
    return Font(bold=False, color="000000", size=sz)


def _apply(ws: Any, row: int, col: int, *,
           value: Any = None, font: Any = None,
           fill: Any = None, border: Any = None,
           align: Any = None) -> Any:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    return cell


def _merge_row(ws: Any, row: int, c1: int, c2: int, **kwargs) -> None:
    ws.merge_cells(
        start_row=row, start_column=c1,
        end_row=row,   end_column=c2,
    )
    _apply(ws, row, c1, **kwargs)


def _section_header(ws: Any, row: int, title: str) -> int:
    _merge_row(
        ws, row, 1, TOTAL_COLS,
        value=title,
        font=_black_bold(10),
        fill=_section_fill(),
        border=_thin_border(),
        align=_left(wrap=False),
    )
    return row + 1


def _col_header_row(ws: Any, row: int) -> int:
    headers = ["번호", "점검 항목", "", "", "", "", "결과 (○/✕/N/A)", "", "비고"]
    spans   = [(1, 1), (2, 6), None, None, None, None, (7, 8), None, (9, 9)]
    merged  = set()
    for i, (h, span) in enumerate(zip(headers, spans)):
        col = i + 1
        if col in merged:
            continue
        if span and span[0] != span[1]:
            ws.merge_cells(
                start_row=row, start_column=span[0],
                end_row=row,   end_column=span[1],
            )
            merged.update(range(span[0] + 1, span[1] + 1))
        _apply(ws, row, span[0] if span else col,
               value=h, font=_white_bold(9),
               fill=_sub_fill(), border=_thin_border(),
               align=_center(wrap=True))
    return row + 1


def _check_item_row(ws: Any, row: int, seq: int,
                    item: str, result: str = "", note: str = "") -> int:
    b = _thin_border()
    _apply(ws, row, 1, value=seq,   font=_black_normal(), border=b, align=_center())
    ws.merge_cells(start_row=row, start_column=2,
                   end_row=row,   end_column=6)
    _apply(ws, row, 2, value=item,  font=_black_normal(), border=b, align=_left())
    ws.merge_cells(start_row=row, start_column=7,
                   end_row=row,   end_column=8)
    _apply(ws, row, 7, value=result, font=_black_normal(), border=b, align=_center())
    _apply(ws, row, 9, value=note,  font=_black_normal(), border=b, align=_left())
    return row + 1


def _info_row_2col(ws: Any, row: int,
                   label1: str, val1: str,
                   label2: str, val2: str) -> int:
    b = _thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _apply(ws, row, 1, value=label1, font=_black_bold(9), fill=_section_fill(),
           border=b, align=_left(wrap=False))
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    _apply(ws, row, 3, value=val1, font=_black_normal(), border=b, align=_left())

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    _apply(ws, row, 5, value=label2, font=_black_bold(9), fill=_section_fill(),
           border=b, align=_left(wrap=False))
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    _apply(ws, row, 7, value=val2, font=_black_normal(), border=b, align=_left())
    return row + 1


def _info_row_wide(ws: Any, row: int, label: str, val: str) -> int:
    b = _thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _apply(ws, row, 1, value=label, font=_black_bold(9), fill=_section_fill(),
           border=b, align=_left(wrap=False))
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    _apply(ws, row, 3, value=val, font=_black_normal(), border=b, align=_left())
    return row + 1


def _notice_row(ws: Any, row: int, text: str,
                fill: Optional[Any] = None) -> int:
    _merge_row(
        ws, row, 1, TOTAL_COLS,
        value=text,
        font=_black_normal(8),
        fill=fill or _notice_fill(),
        border=_thin_border(),
        align=_left(wrap=True),
    )
    ws.row_dimensions[row].height = 22
    return row + 1


def _get_items(form_data: dict, key: str,
               defaults: List[Dict[str, str]]) -> List[Dict[str, str]]:
    raw = form_data.get(key)
    if raw and isinstance(raw, list):
        return raw
    return defaults


def _get_additional_defaults(form_data: dict) -> List[Dict[str, str]]:
    eq_type = (form_data.get("equipment_type") or "").strip()
    for key, items in _ADDITIONAL_DEFAULTS_BY_TYPE.items():
        if key in eq_type:
            return items
    return _ADDITIONAL_DEFAULTS_COMMON


# ──────────────────────────────────────────────────────────────
# 메인 빌더
# ──────────────────────────────────────────────────────────────

def build_construction_equipment_daily_checklist_excel(form_data: dict) -> bytes:
    """건설장비 일일 사전점검표 Excel 파일을 생성하고 bytes로 반환."""

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # 열 너비
    col_widths = [5, 8, 10, 10, 10, 10, 8, 8, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 제목 행 ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    _merge_row(
        ws, row, 1, TOTAL_COLS,
        value=SHEET_HEADING,
        font=Font(bold=True, color="FFFFFF", size=15),
        fill=_header_fill(),
        border=_medium_border(),
        align=_center(),
    )
    row += 1

    ws.row_dimensions[row].height = 28
    _merge_row(
        ws, row, 1, TOTAL_COLS,
        value=SHEET_SUBTITLE,
        font=Font(bold=False, color="FFFFFF", size=8),
        fill=_sub_fill(),
        border=_thin_border(),
        align=_center(wrap=True),
    )
    row += 1

    # ── 주의사항 행 ──────────────────────────────────────────────
    row = _notice_row(ws, row, f"[주의] {NOTICE_SCOPE}")
    row = _notice_row(ws, row, f"[주의] {NOTICE_EXCLUSIVE}")
    row = _notice_row(ws, row, f"[안내] {NOTICE_NONCONFORM}")
    row = _notice_row(ws, row, f"[안내] {NOTICE_LAW}")

    # ════════════════════════════════════════════════════════════
    # 섹션 1 — 현장 기본정보
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 1. 현장 기본정보 】")
    row = _info_row_2col(ws, row,
                         "사업장명", form_data.get("site_name") or "",
                         "공사명",   form_data.get("project_name") or "")
    row = _info_row_2col(ws, row,
                         "점검 일자", form_data.get("check_date") or "",
                         "작업 장소", form_data.get("work_location") or "")

    # ════════════════════════════════════════════════════════════
    # 섹션 2 — 장비 기본정보
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 2. 장비 기본정보 】")
    row = _info_row_2col(ws, row,
                         "장비 종류",     form_data.get("equipment_type") or "",
                         "장비 기종",     form_data.get("equipment_model") or "")
    row = _info_row_2col(ws, row,
                         "등록번호/차량번호", form_data.get("equipment_reg_no") or "",
                         "최대작업용량",  form_data.get("equipment_capacity") or "")

    # ════════════════════════════════════════════════════════════
    # 섹션 3 — 운전자·유도자·작업지휘자 확인
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 3. 운전자·유도자·작업지휘자 확인 】")
    row = _info_row_2col(ws, row,
                         "운전자 성명",     form_data.get("operator_name") or "",
                         "면허 번호",       form_data.get("operator_license_no") or "")
    row = _info_row_2col(ws, row,
                         "유도자 성명",     form_data.get("guide_worker_name") or "",
                         "작업지휘자 성명", form_data.get("work_commander_name") or "")
    row = _info_row_wide(ws, row,
                         "점검자 성명",     form_data.get("checker_name") or "")

    # ════════════════════════════════════════════════════════════
    # 섹션 4 — 서류·검사·보험 확인
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 4. 서류·검사·보험 확인 】")
    row = _col_header_row(ws, row)
    items4 = _get_items(form_data, "doc_check_items", _DOC_DEFAULTS)
    for seq, it in enumerate(items4, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 5 — 외관·누유·타이어/궤도 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 5. 외관·누유·타이어/궤도 점검 】")
    row = _col_header_row(ws, row)
    items5 = _get_items(form_data, "appearance_items", _APPEARANCE_DEFAULTS)
    for seq, it in enumerate(items5, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 6 — 전조등·후미등·경광등·후진경보기 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 6. 전조등·후미등·경광등·후진경보기 점검 】")
    row = _col_header_row(ws, row)
    items6 = _get_items(form_data, "light_items", _LIGHT_DEFAULTS)
    for seq, it in enumerate(items6, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 7 — 제동장치·조향장치·비상정지 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 7. 제동장치·조향장치·비상정지 점검 】")
    row = _col_header_row(ws, row)
    items7 = _get_items(form_data, "brake_items", _BRAKE_DEFAULTS)
    for seq, it in enumerate(items7, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 8 — 전도·전락·지반 상태 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 8. 전도·전락·지반 상태 점검 】")
    row = _col_header_row(ws, row)
    items8 = _get_items(form_data, "stability_items", _STABILITY_DEFAULTS)
    for seq, it in enumerate(items8, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 9 — 접촉·충돌·작업반경 통제 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 9. 접촉·충돌·작업반경 통제 점검 】")
    row = _col_header_row(ws, row)
    items9 = _get_items(form_data, "contact_items", _CONTACT_DEFAULTS)
    for seq, it in enumerate(items9, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 10 — 적재·인양·하역 상태 점검
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 10. 적재·인양·하역 상태 점검 】")
    row = _col_header_row(ws, row)
    items10 = _get_items(form_data, "load_items", _LOAD_DEFAULTS)
    for seq, it in enumerate(items10, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 11 — 장비별 추가 점검
    # ════════════════════════════════════════════════════════════
    eq_type_label = form_data.get("equipment_type") or "기타"
    row = _section_header(ws, row,
                          f"【 11. 장비별 추가 점검 】 (대상 장비: {eq_type_label})")
    row = _col_header_row(ws, row)
    add_defaults = _get_additional_defaults(form_data)
    items11 = _get_items(form_data, "additional_items", add_defaults)
    for seq, it in enumerate(items11, start=1):
        row = _check_item_row(ws, row, seq,
                              it.get("item", ""), it.get("result", ""), it.get("note", ""))

    # ════════════════════════════════════════════════════════════
    # 섹션 12 — 부적합 및 시정조치
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 12. 부적합 및 시정조치 】")
    # 열 헤더
    nf_headers = ["번호", "부적합 내용", "", "발생 위치", "", "시정 조치", "", "완료 여부", "비고"]
    nf_spans   = [(1,1),(2,3),(None),(4,5),(None),(6,7),(None),(8,8),(9,9)]
    merged_nf  = set()
    for i, (h, span) in enumerate(zip(nf_headers, nf_spans)):
        col = i + 1
        if col in merged_nf:
            continue
        if span and span[0] != span[1]:
            ws.merge_cells(start_row=row, start_column=span[0],
                           end_row=row,   end_column=span[1])
            merged_nf.update(range(span[0] + 1, span[1] + 1))
        _apply(ws, row, span[0] if span else col,
               value=h, font=_white_bold(9),
               fill=_sub_fill(), border=_thin_border(), align=_center(wrap=True))
    row += 1

    raw_nc = form_data.get("nonconformance_items")
    nc_items: List[Dict[str, str]] = raw_nc if (raw_nc and isinstance(raw_nc, list)) else []
    fill_count = max(MAX_NONCONFORM, len(nc_items))
    for i in range(fill_count):
        nc = nc_items[i] if i < len(nc_items) else {}
        b  = _thin_border()
        _apply(ws, row, 1, value=i + 1, font=_black_normal(), border=b, align=_center())
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        _apply(ws, row, 2, value=nc.get("content", ""), font=_black_normal(), border=b, align=_left())
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        _apply(ws, row, 4, value=nc.get("location", ""), font=_black_normal(), border=b, align=_left())
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        _apply(ws, row, 6, value=nc.get("action", ""), font=_black_normal(), border=b, align=_left())
        _apply(ws, row, 8, value=nc.get("completed", ""), font=_black_normal(), border=b, align=_center())
        _apply(ws, row, 9, value=nc.get("deadline", ""), font=_black_normal(), border=b, align=_center())
        row += 1

    # 재확인 안내 행
    row = _notice_row(ws, row,
                      "※ 부적합 사항은 사용 전 시정 완료 후 재확인한다. "
                      "시정 완료 전 장비를 사용해서는 안 된다.",
                      fill=_warn_fill())

    # ════════════════════════════════════════════════════════════
    # 섹션 13 — 확인 서명
    # ════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "【 13. 확인 서명 】")

    sign_date = form_data.get("sign_date") or form_data.get("check_date") or ""
    b = _thin_border()

    # 서명 행 헤더
    sign_labels = ["운전자", "유도자", "작업지휘자", "관리감독자", "현장소장"]
    sign_vals   = [
        form_data.get("operator_sign") or form_data.get("operator_name") or "",
        form_data.get("guide_worker_sign") or form_data.get("guide_worker_name") or "",
        form_data.get("work_commander_sign") or form_data.get("work_commander_name") or "",
        form_data.get("supervisor_sign") or "",
        form_data.get("manager_sign") or "",
    ]
    col_w = TOTAL_COLS // len(sign_labels)

    # 라벨 행
    for idx, label in enumerate(sign_labels):
        c1 = idx * col_w + 1
        c2 = c1 + col_w - 1 if idx < len(sign_labels) - 1 else TOTAL_COLS
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        _apply(ws, row, c1, value=label, font=_black_bold(9),
               fill=_section_fill(), border=b, align=_center())
    row += 1

    # 서명값 행
    for idx, val in enumerate(sign_vals):
        c1 = idx * col_w + 1
        c2 = c1 + col_w - 1 if idx < len(sign_vals) - 1 else TOTAL_COLS
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        _apply(ws, row, c1, value=val, font=_black_normal(),
               border=b, align=_center())
    ws.row_dimensions[row].height = 30
    row += 1

    # 서명 일자
    row = _info_row_wide(ws, row, "서명 일자", sign_date)

    # ── 하단 안내 ────────────────────────────────────────────────
    row = _notice_row(ws, row,
                      "※ 본 점검표는 건설장비 일일 사전점검용이며, "
                      "작업계획서 및 장비사용계획서를 대체하지 않는다.")
    row = _notice_row(ws, row,
                      "※ 타워크레인, 이동식 크레인, 비계, 거푸집동바리 전용 점검은 별도 서식으로 관리한다.")

    # 출력
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
