"""
유해화학물질 취급 점검표 — Excel 출력 모듈 (v1.0)  [CL-009]

법적 근거:
    산업안전보건기준에 관한 규칙 제441조 이하 — 유해화학물질 취급 기준

분류: PRACTICAL — 법정 별지 서식 없음, 실무 자체 표준서식

요약:
    현장에서 유해화학물질을 취급할 때 MSDS 비치, 라벨 표시, 보관 조건,
    환기·화재·누출 대응 등을 점검하는 통합 점검표입니다.
    부적합 항목 발견 시 개선조치를 기록하고 추적합니다.

Required form_data keys:
    site_name         str  사업장명
    check_date        str  점검일자
    inspector_name    str  점검자 성명

Optional form_data keys:
    project_name                str  현장명
    department                  str  소속
    position                    str  직책
    work_location               str  작업 장소
    work_description            str  작업 내용
    chemical_name               str  화학물질명
    cas_no                      str  CAS 번호
    chemical_purpose            str  취급 목적
    taking_amount               str  취급 수량
    taking_method               str  취급 방법
    storage_chemical_name       str  저장 화학물질명
    msds_available              str  MSDS 비치 여부
    msds_understanding          str  MSDS 내용 이해 여부
    container_label             str  용기 라벨 표시 상태
    label_legible               str  라벨 판독 가능 여부
    storage_condition           str  보관 조건 적절 여부
    storage_location            str  보관 위치
    incompatible_separated      str  부적합 물질 분리 보관 여부
    spill_kit_available         str  유출 대응 키트 비치 여부
    ventilation_status          str  환기 상태
    fire_extinguisher_available str  소화기 비치 여부
    emergency_response_plan     str  응급 대응 계획
    ppe_required                str  개인보호장비 필요 여부
    judgment                    str  종합 판정 (적합/조건부 적합/부적합)
    judgment_reason             str  판정 사유
    remarks                     str  비고
    nonconformance_items        list[dict]  부적합 항목 (content, action, responsible, deadline, completed)
    preparer_name               str  작성자
    reviewer_sign               str  검토자 서명
    approval_sign               str  승인자 서명
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.output.excel_style_helpers import normalize_signature_row_heights

SHEET_NAME    = "유해화학물질점검표"
SHEET_HEADING = "유해화학물질 취급 점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제441조 이하에 따른 화학물질 취급 현황 점검"
DOC_ID = "CL-009"

TOTAL_COLS = 9
MAX_NONCONFORM = 8

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=9)

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFE0E0")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {1: 14, 2: 12, 3: 12, 4: 12, 5: 12, 6: 10, 7: 12, 8: 12, 9: 9}

_JUDGMENT_OPTIONS = ["적합", "조건부 적합", "부적합"]


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _write_cell(ws, row: int, col1: int, col2: int, value: Any, *,
                font=None, fill=None, align=None,
                height: Optional[int] = None) -> None:
    if col2 > col1:
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
    cell = ws.cell(row=row, column=col1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, col1, row, col2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _write_lv(ws, row: int, label: str, value: Any,
              label_col: int, val_col1: int, val_col2: int) -> None:
    _write_cell(ws, row, label_col, label_col, label,
                font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _write_cell(ws, row, val_col1, val_col2, value,
                font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_hazardous_chemical_checklist(form_data: Dict[str, Any]) -> bytes:
    ws = Workbook().active
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    _apply_col_widths(ws)

    row = 1

    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_HEADING,
                font=_FONT_TITLE, align=_ALIGN_CENTER, height=25)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, SHEET_SUBTITLE,
                font=_FONT_SUBTITLE, align=_ALIGN_CENTER)
    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, f"[{DOC_ID}]",
                font=_FONT_SMALL, align=_ALIGN_CENTER)
    row += 1

    row += 1

    _write_lv(ws, row, "사업장명", _v(form_data, "site_name"), 1, 2, 3)
    _write_lv(ws, row, "현장명", _v(form_data, "project_name"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "점검일자", _v(form_data, "check_date"), 1, 2, 3)
    _write_lv(ws, row, "점검자", _v(form_data, "inspector_name"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "소속", _v(form_data, "department"), 1, 2, 3)
    _write_lv(ws, row, "직책", _v(form_data, "position"), 5, 6, 7)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 화학물질 취급 작업 정보",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "작업 장소", _v(form_data, "work_location"), 1, 2, 3)
    _write_lv(ws, row, "작업 내용", _v(form_data, "work_description"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "화학물질명", _v(form_data, "chemical_name"), 1, 2, 3)
    _write_lv(ws, row, "CAS 번호", _v(form_data, "cas_no"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "취급 목적", _v(form_data, "chemical_purpose"), 1, 2, 3)
    _write_lv(ws, row, "취급 수량", _v(form_data, "taking_amount"), 5, 6, 7)
    row += 1

    _write_lv(ws, row, "취급 방법", _v(form_data, "taking_method"), 1, 2, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ MSDS 및 표시·라벨 점검",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "MSDS 비치 여부", _v(form_data, "msds_available"), 1, 2, 3)
    _write_lv(ws, row, "MSDS 내용 이해 여부", _v(form_data, "msds_understanding"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "용기 라벨 표시 상태", _v(form_data, "container_label"), 1, 2, 3)
    _write_lv(ws, row, "라벨 판독 가능 여부", _v(form_data, "label_legible"), 5, 6, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 보관·취급 상태 점검",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "보관 조건 적절 여부", _v(form_data, "storage_condition"), 1, 2, 3)
    _write_lv(ws, row, "보관 위치", _v(form_data, "storage_location"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "부적합 물질 분리 보관", _v(form_data, "incompatible_separated"), 1, 2, 3)
    _write_lv(ws, row, "유출 대응 키트 비치", _v(form_data, "spill_kit_available"), 5, 6, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 환기·화재·누출 대응 점검",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "환기 상태", _v(form_data, "ventilation_status"), 1, 2, 3)
    _write_lv(ws, row, "소화기 비치 여부", _v(form_data, "fire_extinguisher_available"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "응급 대응 계획", _v(form_data, "emergency_response_plan"), 1, 2, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 보호구 점검",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_lv(ws, row, "개인보호장비 필요 여부", _v(form_data, "ppe_required"), 1, 2, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, TOTAL_COLS, "▣ 부적합 및 개선조치",
                font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    _write_cell(ws, row, 1, 1, "번호", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 2, 4, "부적합 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 5, 6, "개선조치", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 7, 7, "담당자", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 8, 8, "예정일", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _write_cell(ws, row, 9, 9, "완료", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1

    nonconformance_items = form_data.get("nonconformance_items", [])
    for idx in range(1, MAX_NONCONFORM + 1):
        item = nonconformance_items[idx - 1] if idx <= len(nonconformance_items) else {}

        _write_cell(ws, row, 1, 1, str(idx), align=_ALIGN_CENTER)
        _write_cell(ws, row, 2, 4, item.get("content", ""), align=_ALIGN_LEFT)
        _write_cell(ws, row, 5, 6, item.get("action", ""), align=_ALIGN_LEFT)
        _write_cell(ws, row, 7, 7, item.get("responsible", ""), align=_ALIGN_CENTER)
        _write_cell(ws, row, 8, 8, item.get("deadline", ""), align=_ALIGN_CENTER)
        _write_cell(ws, row, 9, 9, item.get("completed", ""), align=_ALIGN_CENTER)
        row += 1

    row += 1

    _write_lv(ws, row, "종합 판정", _v(form_data, "judgment"), 1, 2, 4)
    _write_lv(ws, row, "판정 사유", _v(form_data, "judgment_reason"), 5, 6, 8)
    row += 1

    _write_lv(ws, row, "비고", _v(form_data, "remarks"), 1, 2, 8)
    row += 1

    row += 1

    _write_cell(ws, row, 1, 3, f"작성자: {_v(form_data, 'preparer_name')}", align=_ALIGN_LEFT)
    _write_cell(ws, row, 4, 6, f"검토자: {_v(form_data, 'reviewer_sign')}", align=_ALIGN_LEFT)
    _write_cell(ws, row, 7, TOTAL_COLS, f"승인자: {_v(form_data, 'approval_sign')}", align=_ALIGN_LEFT)

    normalize_signature_row_heights(ws, min_height=18.0)
    ws.print_title_rows = "1:4"
    output = BytesIO()
    ws.parent.save(output)
    output.seek(0)
    return output.read()
