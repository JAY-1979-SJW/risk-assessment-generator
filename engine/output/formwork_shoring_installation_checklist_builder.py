"""
거푸집 및 동바리 설치 점검표 — Excel 출력 모듈 (v1.0)  [CL-002]

법적 근거:
    산업안전보건기준에 관한 규칙 제328조~제337조
    (PARTIAL_VERIFIED: WP-015 evidence 연계. 세부 조항번호는 원문 API 미수집으로
     NEEDS_VERIFICATION 포함 — 법령 원문 확인 후 현장 적용할 것)

    제328조: 거푸집동바리 재료 기준
    제329조: 거푸집 안전성 확보 (침하방지조치 등)
    제330조: 동바리 등의 설치 (수평연결재·가새·받침대)
    제331조: 조립도 (구조검토 후 조립도 작성 의무)
    제332조: 동바리 침하 방지
    제333조: 콘크리트 타설 작업 시 준수사항
    제334조: 콘크리트 타설 작업 중 점검
    제335조: 콘크리트 타설 시 조치
    제336조: 조립·해체·변경 시 준수사항
    제337조: 작업 중 위험 방지

    제338조 이후는 굴착작업 계열이므로 본 서식 근거로 확정하지 않음.

범위:
    본 서식은 거푸집 및 동바리 설치·사용 상태 확인 전용입니다.
    비계 점검표(CL-001)를 대체하지 않습니다.
    거푸집·동바리 작업계획서(WP-015) 및 구조검토서·조립도 원본을 대체하지 않습니다.

Required form_data keys:
    check_date    str  점검 일자
    work_location str  작업 장소 / 구조물 위치
    checker_name  str  점검자 성명

Optional form_data keys:
    site_name               str  사업장명
    project_name            str  공사명
    work_date               str  작업 예정일 또는 기간
    supervisor_name         str  관리감독자 성명
    structure_type          str  구조물 종류 (슬래브·보·기둥·벽체 등)
    floor_level             str  층·구간 (예: 지상 3층)
    work_area               str  작업 구역 (예: A동 북측)
    formwork_type           str  거푸집 종류 (합판·강재·시스템 등)
    shoring_type            str  동바리 종류 (파이프서포트·시스템동바리 등)
    drawing_items           list[dict]  섹션3 항목 override (item, result, note)
    material_items          list[dict]  섹션4 항목 override
    shoring_items           list[dict]  섹션5 항목 override
    formwork_items          list[dict]  섹션6 항목 override
    stability_items         list[dict]  섹션7 항목 override
    platform_items          list[dict]  섹션8 항목 override
    pre_pour_items          list[dict]  섹션9 항목 override
    during_pour_items       list[dict]  섹션10 항목 override
    removal_items           list[dict]  섹션11 항목 override
    nonconformance_items    list[dict]  부적합 사항 (content, location, action, deadline, completed)
    inspector_sign          str  점검자 서명
    supervisor_sign         str  관리감독자 서명
    work_commander_sign     str  작업지휘자 서명
    manager_sign            str  현장소장 서명
    sign_date               str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "거푸집동바리설치점검표"
SHEET_HEADING = "거푸집 및 동바리 설치 점검표"
SHEET_SUBTITLE = (
    "「산업안전보건기준에 관한 규칙」 제328조~제337조에 따른 거푸집·동바리 설치·사용 안전점검"
)

NOTICE_CL001     = "본 점검표는 거푸집 및 동바리 설치·사용 상태 확인용이며, 비계 점검표(CL-001)를 대체하지 않는다."
NOTICE_DRAWING   = "거푸집 및 동바리는 조립도와 작업계획서에 따라 설치·점검한다."
NOTICE_NONCONFORM = "콘크리트 타설 전 부적합 사항은 사용 전 시정 완료 후 재확인한다."
NOTICE_LAW       = "법령 조항은 현행 원문 확인 후 현장에 적용한다."

MAX_NONCONFORM = 5
TOTAL_COLS     = 9


# ──────────────────────────────────────────────────────────────────────
# 섹션별 기본 점검 항목
# ──────────────────────────────────────────────────────────────────────

_DRAWING_DEFAULTS: List[Dict[str, str]] = [
    {"item": "거푸집동바리 조립도 작성 및 현장 비치 여부 [제331조]",      "result": "", "note": ""},
    {"item": "조립도 구조검토 실시 여부 [제331조]",                        "result": "", "note": ""},
    {"item": "작업계획서(WP-015) 작성·비치 및 조립도와 일치 여부",         "result": "", "note": ""},
    {"item": "조립도에 부재 규격·설치간격·이음방법 기재 여부 [제331조]",   "result": "", "note": ""},
    {"item": "구조검토서·조립도 원본 또는 사본 보관 여부",                 "result": "", "note": ""},
]

_MATERIAL_DEFAULTS: List[Dict[str, str]] = [
    {"item": "거푸집동바리 재료 기준 적합 여부 [제328조]",                  "result": "", "note": ""},
    {"item": "변형·부식·손상 부재 사용 없음 여부 [제328조]",               "result": "", "note": ""},
    {"item": "부재 규격(직경·두께·길이) 조립도 기준 충족 여부",            "result": "", "note": ""},
    {"item": "합판 거푸집 두께·결함 없음 여부",                            "result": "", "note": ""},
    {"item": "사용 전 재료 육안 검사 실시 여부",                           "result": "", "note": ""},
]

_SHORING_DEFAULTS: List[Dict[str, str]] = [
    {"item": "동바리 침하 방지 받침목·베이스플레이트 설치 여부 [제330조]", "result": "", "note": ""},
    {"item": "동바리 상단·하단 고정 및 미끄럼 방지 조치 여부 [제330조]",  "result": "", "note": ""},
    {"item": "수평연결재(가새) 설치 여부 [제330조]",                        "result": "", "note": ""},
    {"item": "멍에·장선·동바리 설치 간격 조립도 기준 충족 여부",           "result": "", "note": ""},
    {"item": "연결부 볼트·클램프·핀 체결 상태 이상 없음 여부",             "result": "", "note": ""},
    {"item": "개구부 상부 동바리 받침 보강 여부",                           "result": "", "note": ""},
    {"item": "파이프서포트 최대 높이 준수 여부 (높이 조정 범위 이내)",      "result": "", "note": ""},
]

_FORMWORK_DEFAULTS: List[Dict[str, str]] = [
    {"item": "거푸집 설치 위치·수직도·수평도 기준 적합 여부",              "result": "", "note": ""},
    {"item": "거푸집 벌어짐 방지(폼타이·긴결재) 설치 여부 [제329조]",      "result": "", "note": ""},
    {"item": "거푸집 부상 방지 조치 여부",                                  "result": "", "note": ""},
    {"item": "이음부·접합부 틈새 없음 여부 (콘크리트 누출 방지)",           "result": "", "note": ""},
    {"item": "거푸집 박리제 도포 상태 (이형 시 손상 방지)",                 "result": "", "note": ""},
]

_STABILITY_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업 지반 지지력 확인 (연약지반·매립지 여부) [제332조]",      "result": "", "note": ""},
    {"item": "동바리 하부 지반 침하 위험 없음 여부 [제332조]",              "result": "", "note": ""},
    {"item": "동바리 전도·좌굴 방지 조치 여부",                             "result": "", "note": ""},
    {"item": "편심하중 집중 방지 여부",                                     "result": "", "note": ""},
    {"item": "악천후(강풍·우천) 시 거푸집·동바리 변형 점검 여부",           "result": "", "note": ""},
]

_PLATFORM_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업발판 설치 및 추락방지 조치 여부",                         "result": "", "note": ""},
    {"item": "이동통로 설치 및 안전난간 여부",                              "result": "", "note": ""},
    {"item": "작업구역 하부 출입통제 여부",                                 "result": "", "note": ""},
    {"item": "개구부 덮개 또는 안전망 설치 여부",                           "result": "", "note": ""},
    {"item": "작업 전 근로자 안전교육 실시 여부 [제336조]",                 "result": "", "note": ""},
]

_PRE_POUR_DEFAULTS: List[Dict[str, str]] = [
    {"item": "콘크리트 타설 전 조립도 준수 최종 확인 여부 [제333조]",       "result": "", "note": ""},
    {"item": "타설 순서·타설 속도·편심하중 관리 계획 확인 여부 [제334조]", "result": "", "note": ""},
    {"item": "동바리 수직도·간격·연결부 최종 점검 여부",                    "result": "", "note": ""},
    {"item": "거푸집 긴결재·이음부 이상 없음 최종 확인 여부",               "result": "", "note": ""},
    {"item": "작업지휘자 배치 여부 [제336조]",                              "result": "", "note": ""},
    {"item": "타설 중 이상 감시인(점검자) 배치 계획 확인 여부",             "result": "", "note": ""},
]

_DURING_POUR_DEFAULTS: List[Dict[str, str]] = [
    {"item": "타설 중 동바리 침하·변형·이상음 감시 여부 [제334조]",         "result": "", "note": ""},
    {"item": "타설 속도 과도 집중 없음 여부 (단위면적당 타설량 준수)",       "result": "", "note": ""},
    {"item": "편심타설 금지 준수 여부",                                     "result": "", "note": ""},
    {"item": "이상 발견 시 즉시 작업 중지 및 조치 여부 [제337조]",          "result": "", "note": ""},
    {"item": "타설 완료 후 거푸집·동바리 상태 확인 여부",                   "result": "", "note": ""},
]

_REMOVAL_DEFAULTS: List[Dict[str, str]] = [
    {"item": "콘크리트 강도 확인 후 해체 여부 (압축강도 시험 또는 동등 기준)", "result": "", "note": ""},
    {"item": "해체 전 조립도·작업계획서 확인 여부 [제336조]",                "result": "", "note": ""},
    {"item": "해체 작업 구역 출입통제 및 안전표지 설치 여부",                "result": "", "note": ""},
    {"item": "해체 순서 준수 여부 (설치 역순, 상부→하부)",                   "result": "", "note": ""},
    {"item": "해체 자재 즉시 지정 장소 집결 및 낙하방지 조치 여부",          "result": "", "note": ""},
    {"item": "악천후 시 해체 작업 중지 여부 [제337조]",                      "result": "", "note": ""},
]


# ──────────────────────────────────────────────────────────────────────
# 스타일 상수
# ──────────────────────────────────────────────────────────────────────

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_NOTICE   = Font(name="맑은 고딕", size=8,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)
_FONT_HEADER   = Font(name="맑은 고딕", size=9,  bold=True, color="FFFFFF")

_FILL_LABEL   = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION = PatternFill(fill_type="solid", fgColor="D6E4F7")
_FILL_COL_HDR = PatternFill(fill_type="solid", fgColor="375A7F")
_FILL_WARN    = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NOTICE  = PatternFill(fill_type="solid", fgColor="FEF0E7")
_FILL_DANGER  = PatternFill(fill_type="solid", fgColor="FCE4D6")
_FILL_NONE    = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 7, 3: 7, 4: 7, 5: 7, 6: 7,
    7: 8, 8: 10, 9: 10,
}

TOTAL_COLS = 9


# ──────────────────────────────────────────────────────────────────────
# 헬퍼
# ──────────────────────────────────────────────────────────────────────

def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _wc(ws, row: int, c1: int, c2: int, value: Any, *,
        font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, c1, row, c2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _wlv(ws, row: int, label: str, value: Any,
         lc: int, vs: int, ve: int) -> None:
    _wc(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, vs, ve, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _get_items(form_data: Dict[str, Any], key: str,
               defaults: List[Dict[str, str]]) -> List[Dict[str, str]]:
    raw = form_data.get(key)
    if raw and isinstance(raw, list):
        return raw
    return defaults


# ──────────────────────────────────────────────────────────────────────
# 렌더링 함수
# ──────────────────────────────────────────────────────────────────────

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    c = ws.cell(row=row, column=1, value=SHEET_HEADING)
    c.font = _FONT_TITLE; c.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 1
    _wc(ws, row, 1, TOTAL_COLS, SHEET_SUBTITLE,
        font=_FONT_SUBTITLE, align=_ALIGN_CENTER, height=16)
    row += 1
    return row


def _write_notices(ws, row: int) -> int:
    for text, fill in [
        (f"[주의] {NOTICE_CL001}",    _FILL_NOTICE),
        (f"[주의] {NOTICE_DRAWING}",   _FILL_NOTICE),
        (f"[안내] {NOTICE_NONCONFORM}", _FILL_WARN),
        (f"[안내] {NOTICE_LAW}",        _FILL_WARN),
    ]:
        _wc(ws, row, 1, TOTAL_COLS, text, font=_FONT_NOTICE, fill=fill,
            align=_ALIGN_LEFT, height=18)
        row += 1
    return row


def _write_section_header(ws, row: int, title: str) -> int:
    _wc(ws, row, 1, TOTAL_COLS, title,
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_LEFT, height=18)
    return row + 1


def _write_col_header(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    _wc(ws, row, 1, 1, "번호", font=_FONT_HEADER, fill=_FILL_COL_HDR, align=_ALIGN_CENTER)
    ws.cell(row=row, column=2).value = "점검 항목"
    ws.cell(row=row, column=2).font  = _FONT_HEADER
    ws.cell(row=row, column=2).fill  = _FILL_COL_HDR
    ws.cell(row=row, column=2).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.cell(row=row, column=7).value = "결과 (○/✕/N/A)"
    ws.cell(row=row, column=7).font  = _FONT_HEADER
    ws.cell(row=row, column=7).fill  = _FILL_COL_HDR
    ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
    _wc(ws, row, 9, 9, "비고", font=_FONT_HEADER, fill=_FILL_COL_HDR, align=_ALIGN_CENTER)
    _border_rect(ws, row, 1, row, TOTAL_COLS)
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_check_items(ws, row: int,
                       items: List[Dict[str, str]]) -> int:
    for seq, it in enumerate(items, start=1):
        _wc(ws, row, 1, 1, seq, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2).value     = it.get("item", "")
        ws.cell(row=row, column=2).font      = _FONT_DEFAULT
        ws.cell(row=row, column=2).alignment = _ALIGN_LEFT
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws.cell(row=row, column=7).value     = it.get("result", "")
        ws.cell(row=row, column=7).font      = _FONT_DEFAULT
        ws.cell(row=row, column=7).alignment = _ALIGN_CENTER
        _wc(ws, row, 9, 9, it.get("note", ""), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _border_rect(ws, row, 1, row, TOTAL_COLS)
        row += 1
    return row


def _write_section(ws, row: int, title: str,
                   items: List[Dict[str, str]]) -> int:
    row = _write_section_header(ws, row, title)
    row = _write_col_header(ws, row)
    row = _write_check_items(ws, row, items)
    return row


def _write_info_block(ws, row: int, form_data: Dict[str, Any]) -> int:
    """섹션 1~2: 현장 기본정보 + 구조물 정보."""
    row = _write_section_header(ws, row, "【 1. 현장 기본정보 】")
    _wlv(ws, row, "사업장명",  _v(form_data, "site_name"),     1, 2, 5)
    _wlv(ws, row, "공사명",    _v(form_data, "project_name"),  6, 7, 9)
    row += 1
    _wlv(ws, row, "점검 일자", _v(form_data, "check_date"),    1, 2, 5)
    _wlv(ws, row, "작업 장소", _v(form_data, "work_location"), 6, 7, 9)
    row += 1
    _wlv(ws, row, "작업 기간", _v(form_data, "work_date"),     1, 2, 5)
    _wlv(ws, row, "관리감독자", _v(form_data, "supervisor_name"), 6, 7, 9)
    row += 1

    row = _write_section_header(ws, row, "【 2. 구조물 및 작업구간 정보 】")
    _wlv(ws, row, "구조물 종류", _v(form_data, "structure_type"), 1, 2, 5)
    _wlv(ws, row, "층·구간",     _v(form_data, "floor_level"),    6, 7, 9)
    row += 1
    _wlv(ws, row, "작업 구역",  _v(form_data, "work_area"),       1, 2, 5)
    _wlv(ws, row, "점검자 성명", _v(form_data, "checker_name"),   6, 7, 9)
    row += 1
    _wlv(ws, row, "거푸집 종류", _v(form_data, "formwork_type"),  1, 2, 5)
    _wlv(ws, row, "동바리 종류", _v(form_data, "shoring_type"),   6, 7, 9)
    row += 1
    return row


def _write_nonconformance(ws, row: int, form_data: Dict[str, Any]) -> int:
    """섹션 12: 부적합 및 시정조치."""
    row = _write_section_header(ws, row, "【 12. 부적합 및 시정조치 】")

    # 열 헤더
    for (c1, c2, label) in [
        (1,1,"번호"), (2,3,"부적합 내용"), (4,5,"발생 위치"),
        (6,7,"시정 조치"), (8,8,"완료 여부"), (9,9,"완료 기한"),
    ]:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        ws.cell(row=row, column=c1).value     = label
        ws.cell(row=row, column=c1).font      = _FONT_HEADER
        ws.cell(row=row, column=c1).fill      = _FILL_COL_HDR
        ws.cell(row=row, column=c1).alignment = _ALIGN_CENTER
    _border_rect(ws, row, 1, row, TOTAL_COLS)
    ws.row_dimensions[row].height = 16
    row += 1

    raw_nc = form_data.get("nonconformance_items")
    nc_items: List[Dict[str, str]] = raw_nc if (raw_nc and isinstance(raw_nc, list)) else []
    fill_count = max(MAX_NONCONFORM, len(nc_items))
    for i in range(fill_count):
        nc = nc_items[i] if i < len(nc_items) else {}
        _wc(ws, row, 1, 1, i + 1, font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2).value     = nc.get("content", "")
        ws.cell(row=row, column=2).font      = _FONT_DEFAULT
        ws.cell(row=row, column=2).alignment = _ALIGN_LEFT
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.cell(row=row, column=4).value     = nc.get("location", "")
        ws.cell(row=row, column=4).font      = _FONT_DEFAULT
        ws.cell(row=row, column=4).alignment = _ALIGN_LEFT
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6).value     = nc.get("action", "")
        ws.cell(row=row, column=6).font      = _FONT_DEFAULT
        ws.cell(row=row, column=6).alignment = _ALIGN_LEFT
        _wc(ws, row, 8, 8, nc.get("completed", ""), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 9, 9, nc.get("deadline", ""),  font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _border_rect(ws, row, 1, row, TOTAL_COLS)
        row += 1

    _wc(ws, row, 1, TOTAL_COLS,
        "※ 콘크리트 타설 전 부적합 사항은 사용 전 시정 완료 후 재확인한다. "
        "시정 완료 전 타설을 진행해서는 안 된다.",
        font=_FONT_NOTICE, fill=_FILL_DANGER, align=_ALIGN_LEFT, height=22)
    row += 1
    return row


def _write_signature(ws, row: int, form_data: Dict[str, Any]) -> int:
    """섹션 13: 확인 서명."""
    row = _write_section_header(ws, row, "【 13. 확인 서명 】")

    sign_date = _v(form_data, "sign_date") or _v(form_data, "check_date")
    labels = ["점검자", "관리감독자", "작업지휘자", "현장소장"]
    vals   = [
        _v(form_data, "inspector_sign") or _v(form_data, "checker_name"),
        _v(form_data, "supervisor_sign"),
        _v(form_data, "work_commander_sign"),
        _v(form_data, "manager_sign"),
    ]
    # 4열 균등 분할 (9열 → 각 2열, 마지막 3열)
    spans = [(1,2),(3,4),(5,6),(7,9)]

    for (c1, c2), label in zip(spans, labels):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        ws.cell(row=row, column=c1).value     = label
        ws.cell(row=row, column=c1).font      = _FONT_BOLD
        ws.cell(row=row, column=c1).fill      = _FILL_LABEL
        ws.cell(row=row, column=c1).alignment = _ALIGN_CENTER
    _border_rect(ws, row, 1, row, TOTAL_COLS)
    row += 1

    for (c1, c2), val in zip(spans, vals):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        ws.cell(row=row, column=c1).value     = val
        ws.cell(row=row, column=c1).font      = _FONT_DEFAULT
        ws.cell(row=row, column=c1).alignment = _ALIGN_CENTER
    _border_rect(ws, row, 1, row, TOTAL_COLS)
    ws.row_dimensions[row].height = 28
    row += 1

    _wlv(ws, row, "서명 일자", sign_date, 1, 2, 9)
    ws.row_dimensions[row].height = 20
    row += 1

    _wc(ws, row, 1, TOTAL_COLS,
        "※ 본 점검표는 거푸집 및 동바리 설치·사용 상태 확인용이며, "
        "비계 점검표(CL-001)를 대체하지 않는다.",
        font=_FONT_NOTICE, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    _wc(ws, row, 1, TOTAL_COLS,
        "※ 거푸집·동바리 작업계획서(WP-015) 및 구조검토서·조립도 원본은 별도 보관하여야 한다.",
        font=_FONT_NOTICE, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=18)
    row += 1
    return row


# ──────────────────────────────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────────────────────────────

def build_formwork_shoring_installation_checklist_excel(form_data: dict) -> bytes:
    """거푸집 및 동바리 설치 점검표 Excel 파일을 생성하고 bytes로 반환."""

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for idx, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    row = 1
    row = _write_title(ws, row)
    row = _write_notices(ws, row)
    row = _write_info_block(ws, row, form_data)

    # 섹션 3 — 조립도·작업계획서 확인
    row = _write_section(ws, row, "【 3. 조립도·작업계획서 확인 】",
                         _get_items(form_data, "drawing_items", _DRAWING_DEFAULTS))

    # 섹션 4 — 재료 및 부재 상태 점검
    row = _write_section(ws, row, "【 4. 재료 및 부재 상태 점검 】",
                         _get_items(form_data, "material_items", _MATERIAL_DEFAULTS))

    # 섹션 5 — 동바리 설치 상태 점검
    row = _write_section(ws, row, "【 5. 동바리 설치 상태 점검 】",
                         _get_items(form_data, "shoring_items", _SHORING_DEFAULTS))

    # 섹션 6 — 거푸집 설치 상태 점검
    row = _write_section(ws, row, "【 6. 거푸집 설치 상태 점검 】",
                         _get_items(form_data, "formwork_items", _FORMWORK_DEFAULTS))

    # 섹션 7 — 침하·전도·변형 방지 조치
    row = _write_section(ws, row, "【 7. 침하·전도·변형 방지 조치 】",
                         _get_items(form_data, "stability_items", _STABILITY_DEFAULTS))

    # 섹션 8 — 작업발판·통로·추락방지 조치
    row = _write_section(ws, row, "【 8. 작업발판·통로·추락방지 조치 】",
                         _get_items(form_data, "platform_items", _PLATFORM_DEFAULTS))

    # 섹션 9 — 콘크리트 타설 전 점검
    row = _write_section(ws, row, "【 9. 콘크리트 타설 전 점검 】",
                         _get_items(form_data, "pre_pour_items", _PRE_POUR_DEFAULTS))

    # 섹션 10 — 콘크리트 타설 중 점검
    row = _write_section(ws, row, "【 10. 콘크리트 타설 중 점검 】",
                         _get_items(form_data, "during_pour_items", _DURING_POUR_DEFAULTS))

    # 섹션 11 — 해체 전 안전조치
    row = _write_section(ws, row, "【 11. 해체 전 안전조치 】",
                         _get_items(form_data, "removal_items", _REMOVAL_DEFAULTS))

    # 섹션 12 — 부적합 및 시정조치
    row = _write_nonconformance(ws, row, form_data)

    # 섹션 13 — 확인 서명
    row = _write_signature(ws, row, form_data)

    # ── A4 인쇄 설정 ──────────────────────────────────────────────────────────
    ws.page_setup.paperSize   = 9          # A4
    ws.page_setup.orientation  = "portrait"
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.print_title_rows = "1:2"  # 제목+부제 반복

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
