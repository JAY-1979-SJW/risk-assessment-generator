"""
안전보건관리자 직무교육 이수 확인서 — Excel 출력 모듈 (v1.0)

법적 근거: 산업안전보건법 제32조, 제15조, 제17조, 제18조
시행규칙 직무교육 조항 (NEEDS_VERIFICATION: 조항번호 재확인 필요)
Form: ED-004

입력 스키마 (form_data dict):
    [1. 사업장/현장 기본정보]
    site_name               str|None   사업장명
    field_name              str|None   현장명
    employer_name           str|None   사업주/시공사
    write_date              str|None   작성일
    department              str|None   담당부서

    [2. 직무교육 대상자 정보]
    person_name             str|None   성명
    person_org              str|None   소속
    person_title            str|None   직책
    role_type               str|None   역할 구분 (안전보건관리책임자 / 안전관리자 / 보건관리자)
    appointment_date        str|None   선임일 또는 채용일
    is_training_target      str|None   직무교육 대상 여부
    training_category       str|None   신규교육/보수교육 구분

    [3. 법정 교육 기준]
    legal_basis_text        str|None   관련 법령
    new_training_deadline   str|None   신규교육 기한
    refresher_cycle         str|None   보수교육 주기
    doctor_special_case     str|None   의사인 보건관리자 특례 여부
    training_exemption      str|None   교육 면제 또는 일부 면제 여부

    [4. 교육 이수 내역]
    training_org            str|None   교육기관명
    training_course         str|None   교육과정명
    training_start_date     str|None   교육 시작일
    training_end_date       str|None   교육 종료일
    training_hours          str|None   교육시간
    completion_no           str|None   이수번호 (수료증 번호 — 자유입력)
    certificate_date        str|None   수료증 발급일
    completion_status       str|None   수료 여부

    [5. 증빙자료 확인]
    cert_attached           str|None   수료증 첨부 여부
    agency_confirm_attached str|None   교육기관 확인서 첨부 여부
    appointment_doc_attached str|None  선임신고서 또는 선임 관련 자료 첨부 여부
    refresher_basis_attached str|None  보수교육 대상 산정자료 첨부 여부

    [6. 미이수자 관리]
    not_completed           str|None   미이수 여부
    not_completed_reason    str|None   미이수 사유
    action_plan             str|None   조치계획
    scheduled_training_date str|None   예정 교육일
    manager_name            str|None   담당자

    [7. 확인 및 서명]
    writer_name             str|None   작성자
    safety_manager_sign     str|None   안전관리자 또는 보건관리자
    supervisor_sign         str|None   관리감독자
    site_manager_sign       str|None   현장소장 또는 사업주 확인
    sign_date               str|None   확인 일자

출력: xlsx bytes (in-memory)

주의:
- 주민등록번호 전체 필드 없음
- 건강정보, 질병명 필드 없음
- 교육시간/주기 자동 판정 없음 (입력값 기반)
- 수료증 번호는 자유입력
- 본 확인서는 안전보건교육기관의 공식 수료증을 대체하지 않음
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
SHEET_NAME    = "직무교육이수확인서"
SHEET_HEADING = "안전보건관리자 직무교육 이수 확인서"

_HEADER_FILL  = PatternFill(patternType="solid", fgColor="1F4E79")
_SECTION_FILL = PatternFill(patternType="solid", fgColor="2E75B6")
_SUB_FILL     = PatternFill(patternType="solid", fgColor="BDD7EE")
_WARN_FILL    = PatternFill(patternType="solid", fgColor="FFF2CC")
_WHITE_FILL   = PatternFill(patternType="solid", fgColor="FFFFFF")

_HDR_FONT     = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=16)
_SEC_FONT     = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_SUB_FONT     = Font(name="맑은 고딕", bold=True, color="1F4E79", size=9)
_LABEL_FONT   = Font(name="맑은 고딕", bold=True, size=9)
_BODY_FONT    = Font(name="맑은 고딕", size=9)
_NOTE_FONT    = Font(name="맑은 고딕", italic=True, size=8, color="C00000")
_WARN_FONT    = Font(name="맑은 고딕", bold=True, size=8, color="7F0000")

_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_WRAP    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _thin_border() -> Border:
    s = Side(border_style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _medium_border() -> Border:
    s = Side(border_style="medium", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _v(form_data: dict, key: str) -> str:
    val = form_data.get(key)
    return str(val).strip() if val is not None else ""


def _apply(ws, row: int, col: int, value: Any,
           font=None, fill=None, alignment=None, border=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _merge_apply(ws, r1: int, c1: int, r2: int, c2: int, value: Any,
                 font=None, fill=None, alignment=None, border=None) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            if r == r1 and c == c1:
                continue
            ws.cell(row=r, column=c).border = border or _thin_border()


def _section_row(ws, row: int, label: str, col_span_end: int = 8) -> None:
    _merge_apply(ws, row, 1, row, col_span_end, label,
                 font=_SEC_FONT, fill=_SECTION_FILL, alignment=_LEFT)


def _label_value(ws, row: int, label_col: int, label: str,
                 value_col: int, value_col_end: int, value: str) -> None:
    _apply(ws, row, label_col, label,
           font=_LABEL_FONT, fill=_SUB_FILL, alignment=_CENTER,
           border=_thin_border())
    _merge_apply(ws, row, value_col, row, value_col_end, value,
                 font=_BODY_FONT, fill=_WHITE_FILL, alignment=_LEFT,
                 border=_thin_border())


# ---------------------------------------------------------------------------
# 메인 빌더
# ---------------------------------------------------------------------------

def build_manager_job_training_record_excel(form_data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False

    # 열 너비 설정 (A~H = 1~8)
    col_widths = [2, 14, 19, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 타이틀 ──────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 40
    _merge_apply(ws, row, 1, row, 8, SHEET_HEADING,
                 font=_HDR_FONT, fill=_HEADER_FILL, alignment=_CENTER,
                 border=_medium_border())
    row += 1

    # 법령 근거 안내
    ws.row_dimensions[row].height = 16
    _merge_apply(ws, row, 1, row, 8,
                 "법적 근거: 산업안전보건법 제32조 (직무교육), 제15조 (안전보건관리책임자), 제17조 (안전관리자), 제18조 (보건관리자)",
                 font=_BODY_FONT, fill=_WARN_FILL, alignment=_LEFT,
                 border=_thin_border())
    row += 1

    # ── 섹션 1: 사업장/현장 기본정보 ────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "1. 사업장/현장 기본정보")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "사업장명",    3, 4, _v(form_data, "site_name"))
    _label_value(ws, row, 5, "현장명",      6, 8, _v(form_data, "field_name"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "사업주/시공사", 3, 4, _v(form_data, "employer_name"))
    _label_value(ws, row, 5, "작성일",       6, 6, _v(form_data, "write_date"))
    _label_value(ws, row, 7, "담당부서",     8, 8, _v(form_data, "department"))
    row += 1

    # ── 섹션 2: 직무교육 대상자 정보 ────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "2. 직무교육 대상자 정보")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "성명",   3, 4, _v(form_data, "person_name"))
    _label_value(ws, row, 5, "소속",   6, 8, _v(form_data, "person_org"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "직책",   3, 4, _v(form_data, "person_title"))
    _label_value(ws, row, 5, "선임일/채용일", 6, 8, _v(form_data, "appointment_date"))
    row += 1

    ws.row_dimensions[row].height = 22
    _apply(ws, row, 2, "역할 구분",
           font=_LABEL_FONT, fill=_SUB_FILL, alignment=_CENTER, border=_thin_border())
    role_val = _v(form_data, "role_type")
    role_display = role_val if role_val else "안전보건관리책임자 / 안전관리자 / 보건관리자 중 선택"
    _merge_apply(ws, row, 3, row, 8, role_display,
                 font=_BODY_FONT, fill=_WHITE_FILL, alignment=_LEFT,
                 border=_thin_border())
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "직무교육 대상 여부", 3, 4, _v(form_data, "is_training_target"))
    _label_value(ws, row, 5, "신규교육/보수교육", 6, 8, _v(form_data, "training_category"))
    row += 1

    # ── 섹션 3: 법정 교육 기준 ──────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "3. 법정 교육 기준")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "관련 법령", 3, 8,
                 _v(form_data, "legal_basis_text") or
                 "산업안전보건법 제32조, 제15조/제17조/제18조, 시행규칙 직무교육 조항")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "신규교육 기한", 3, 4, _v(form_data, "new_training_deadline"))
    _label_value(ws, row, 5, "보수교육 주기", 6, 8, _v(form_data, "refresher_cycle"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "의사인 보건관리자 특례", 3, 4, _v(form_data, "doctor_special_case"))
    _label_value(ws, row, 5, "교육 면제 여부",         6, 8, _v(form_data, "training_exemption"))
    row += 1

    # 시행규칙 조항번호 NEEDS_VERIFICATION 안내
    ws.row_dimensions[row].height = 16
    _merge_apply(ws, row, 2, row, 8,
                 "※ 시행규칙 직무교육 조항 번호는 law.go.kr 원문 재확인 필요 (NEEDS_VERIFICATION)",
                 font=_NOTE_FONT, fill=_WARN_FILL, alignment=_LEFT, border=_thin_border())
    row += 1

    # ── 섹션 4: 교육 이수 내역 ──────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "4. 교육 이수 내역")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "교육기관명",   3, 4, _v(form_data, "training_org"))
    _label_value(ws, row, 5, "교육과정명",   6, 8, _v(form_data, "training_course"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "교육 시작일",  3, 4, _v(form_data, "training_start_date"))
    _label_value(ws, row, 5, "교육 종료일",  6, 8, _v(form_data, "training_end_date"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "교육시간",     3, 4, _v(form_data, "training_hours"))
    _label_value(ws, row, 5, "수료 여부",    6, 8, _v(form_data, "completion_status"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "이수번호",     3, 4, _v(form_data, "completion_no"))
    _label_value(ws, row, 5, "수료증 발급일", 6, 8, _v(form_data, "certificate_date"))
    row += 1

    # ── 섹션 5: 증빙자료 확인 ───────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "5. 증빙자료 확인")
    row += 1

    for label_key, label_text in [
        ("cert_attached",            "수료증 첨부 여부"),
        ("agency_confirm_attached",  "교육기관 확인서 첨부 여부"),
        ("appointment_doc_attached", "선임신고서/선임 자료 첨부 여부"),
        ("refresher_basis_attached", "보수교육 대상 산정자료 첨부 여부"),
    ]:
        ws.row_dimensions[row].height = 18
        _label_value(ws, row, 2, label_text, 3, 8, _v(form_data, label_key))
        row += 1

    # ── 섹션 6: 미이수자 관리 ───────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "6. 미이수자 관리")
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "미이수 여부",   3, 4, _v(form_data, "not_completed"))
    _label_value(ws, row, 5, "미이수 사유",   6, 8, _v(form_data, "not_completed_reason"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "조치계획",      3, 4, _v(form_data, "action_plan"))
    _label_value(ws, row, 5, "예정 교육일",   6, 8, _v(form_data, "scheduled_training_date"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "담당자",        3, 8, _v(form_data, "manager_name"))
    row += 1

    # ── 섹션 7: 확인 및 서명 ────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    _section_row(ws, row, "7. 확인 및 서명")
    row += 1

    ws.row_dimensions[row].height = 22
    _label_value(ws, row, 2, "작성자",          3, 3, _v(form_data, "writer_name"))
    _label_value(ws, row, 4, "안전관리자/보건관리자", 5, 5, _v(form_data, "safety_manager_sign"))
    _label_value(ws, row, 6, "관리감독자",       7, 7, _v(form_data, "supervisor_sign"))
    _label_value(ws, row, 8, "현장소장/사업주",  8, 8, _v(form_data, "site_manager_sign"))
    row += 1

    ws.row_dimensions[row].height = 18
    _label_value(ws, row, 2, "확인 일자", 3, 8, _v(form_data, "sign_date"))
    row += 1

    # ── 필수 고지 문구 ───────────────────────────────────────────────────
    for msg in [
        "본 확인서는 직무교육 이수 관리용 문서이며, 안전보건교육기관의 공식 수료증을 대체하지 않음",
        "공식 수료증 원본 또는 사본을 반드시 첨부·보관하여야 함",
        "법령 개정 시 교육 대상·주기·기한 재확인 필요",
    ]:
        ws.row_dimensions[row].height = 14
        _merge_apply(ws, row, 1, row, 8, f"※ {msg}",
                     font=_WARN_FONT, fill=_WARN_FILL, alignment=_LEFT,
                     border=_thin_border())
        row += 1

    # ── A4 인쇄 설정 ──────────────────────────────────────────────────────────
    ws.page_setup.paperSize   = 9          # A4
    ws.page_setup.orientation  = "portrait"
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.print_title_rows = "1:3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
