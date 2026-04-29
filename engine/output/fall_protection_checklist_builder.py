"""
추락 방호 설비 점검표 — Excel 출력 모듈 (v1.0)  [CL-007]

법적 근거:
    산업안전보건기준에 관한 규칙 제42조~제45조 — 추락방지/개구부방호/안전대부착설비/지붕위험
    (PARTIAL_VERIFIED: 제42조~제45조 원문 API 수집 완료. 제13조(안전난간) NEEDS_VERIFICATION)

범위:
    본 서식은 추락 방호 설비 현장 점검 전용입니다.
    비계 설치 점검은 CL-001, 거푸집동바리는 CL-002, 장비 점검은 CL-003을 사용합니다.
    고소작업 수행 허가는 PTW-003(고소작업 허가서)을 별도로 발행합니다.

Required form_data keys:
    check_date    str  점검 일자
    work_location str  작업 장소
    checker_name  str  점검자 성명

Optional form_data keys:
    site_name            str  사업장명
    project_name         str  공사명
    work_name            str  작업명
    work_height          str  작업 높이 (m)
    work_period          str  작업 기간
    ptw_no               str  관련 PTW 번호
    supervisor_name      str  관리감독자 성명
    hazard_zone_items    list[dict]  섹션3 항목 override (item, checked, note)
    platform_items       list[dict]  섹션4 항목 override (item, result, note)
    railing_items        list[dict]  섹션5 항목 override (item, result, note)
    opening_items        list[dict]  섹션6 항목 override (item, result, note)
    harness_items        list[dict]  섹션7 항목 override (item, result, note)
    net_items            list[dict]  섹션8 항목 override (item, result, note)
    special_equip_items  list[dict]  섹션9 항목 override (item, result, note)
    nonconformance_items list[dict]  부적합 사항 (no, content, location, action, deadline, completed)
    work_verdict         str  작업 가능 여부 판정 (적합/조건부 적합/작업중지)
    verdict_condition    str  조건부 적합 시 조건 내용
    inspector_sign       str  점검자 서명
    supervisor_sign      str  관리감독자 서명
    manager_sign         str  현장소장 서명
    sign_date            str  서명일
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME    = "추락방호설비점검표"
SHEET_HEADING = "추락 방호 설비 점검표"
SHEET_SUBTITLE = "「산업안전보건기준에 관한 규칙」 제42조~제45조에 따른 추락 위험 방호 설비 현장 점검"

NOTICE_PTW003   = (
    "고소작업(2m 이상) 수행 전 PTW-003(고소작업 허가서)을 별도로 발행하여야 한다."
)
NOTICE_CL001    = (
    "비계 설치 상태 점검은 CL-001(비계 설치 점검표)을 병행하여 확인한다."
)
NOTICE_LAW      = "법령 조항 및 기준은 현행 법령 원문 확인 후 현장에 적용한다."
NOTICE_NONCONFORM = "점검 결과 부적합 사항은 작업 전 시정 완료 후 확인 서명을 받아야 한다."

MAX_NONCONFORM = 5
TOTAL_COLS     = 9

# ──────────────────────────────────────────────
# 기본 점검 항목
# ──────────────────────────────────────────────

_HAZARD_ZONE_DEFAULTS: List[Dict[str, str]] = [
    {"item": "개구부 위험구역 식별 및 방호조치 여부",             "result": "", "note": ""},
    {"item": "단부(끝단) 안전난간 또는 방호울 설치 여부",         "result": "", "note": ""},
    {"item": "지붕 작업구역 가장자리 안전난간 설치 여부",          "result": "", "note": ""},
    {"item": "비계 작업발판 단부 방호조치 여부",                  "result": "", "note": ""},
    {"item": "이동식 비계 이동 잠금 및 단부 안전난간 여부",        "result": "", "note": ""},
    {"item": "고소작업대 사용 구역 안전거리 확보 여부",            "result": "", "note": ""},
    {"item": "사다리 상단 고정 및 사용 기준 준수 여부",           "result": "", "note": ""},
    {"item": "임시 작업발판 고정 및 적재하중 기준 충족 여부",      "result": "", "note": ""},
]

_PLATFORM_DEFAULTS: List[Dict[str, str]] = [
    {"item": "작업발판 폭 기준 충족 여부 (강관비계 40cm 이상)",    "result": "", "note": ""},
    {"item": "발판 고정 상태 — 탈락·흔들림 없음 여부",            "result": "", "note": ""},
    {"item": "발판 처짐·파손·균열 없음 여부",                    "result": "", "note": ""},
    {"item": "발판 표면 미끄럼 방지 조치 여부",                  "result": "", "note": ""},
    {"item": "발끝막이판(toe board) 설치 여부",                  "result": "", "note": ""},
]

_RAILING_DEFAULTS: List[Dict[str, str]] = [
    {"item": "상부 안전난간대 높이 기준 충족 여부",               "result": "", "note": ""},
    {"item": "중간 난간대 설치 여부",                            "result": "", "note": ""},
    {"item": "발끝막이판 설치 여부",                             "result": "", "note": ""},
    {"item": "난간 고정 상태 — 유동·변형 없음 여부",              "result": "", "note": ""},
    {"item": "개구부 방호울 또는 덮개 설치 여부",                 "result": "", "note": ""},
]

_OPENING_DEFAULTS: List[Dict[str, str]] = [
    {"item": "개구부 덮개 고정 상태 (임의 제거 불가 구조)",        "result": "", "note": ""},
    {"item": "개구부 식별 표시(경고 표지) 부착 여부",             "result": "", "note": ""},
    {"item": "덮개 임의 제거 방지 조치 여부",                    "result": "", "note": ""},
    {"item": "개구부 하부 출입통제 조치 여부",                   "result": "", "note": ""},
]

_HARNESS_DEFAULTS: List[Dict[str, str]] = [
    {"item": "안전대 착용 여부 (2m 이상 전 작업자)",              "result": "", "note": ""},
    {"item": "죔줄(lanyard) 상태 — 손상·마모 없음 여부",          "result": "", "note": ""},
    {"item": "안전블록(self-retracting) 설치 적정 여부 (해당 시)", "result": "", "note": ""},
    {"item": "구명줄 설치 여부 (해당 시)",                       "result": "", "note": ""},
    {"item": "앵커포인트 위치 적합 여부",                        "result": "", "note": ""},
    {"item": "부착설비 강도 확인 여부 (2kN 이상 견딜 구조)",      "result": "", "note": ""},
]

_NET_DEFAULTS: List[Dict[str, str]] = [
    {"item": "추락방망 설치 상태 — 처짐·파손 없음 여부",          "result": "", "note": ""},
    {"item": "낙하물 방지망 설치 여부 (해당 구역)",               "result": "", "note": ""},
    {"item": "공구·소재 낙하 방지 조치 여부 (연결줄 등)",          "result": "", "note": ""},
    {"item": "단부·난간 상부 자재 적치 금지 준수 여부",            "result": "", "note": ""},
]

_SPECIAL_EQUIP_DEFAULTS: List[Dict[str, str]] = [
    {"item": "이동식 비계 — 이동 잠금 후 작업 여부",             "result": "", "note": ""},
    {"item": "이동식 비계 — 최상부 안전난간 설치 여부",           "result": "", "note": ""},
    {"item": "고소작업대 — 작업 전 작동 상태 점검 여부",          "result": "", "note": ""},
    {"item": "고소작업대 — 전도 방지 아우트리거 전개 여부",        "result": "", "note": ""},
    {"item": "사다리 — 상단 60cm 이상 돌출 및 고정 여부",        "result": "", "note": ""},
    {"item": "사다리 — 높이 3.5m 이하(이동식), 최상부 발판 작업 금지", "result": "", "note": ""},
]

# ──────────────────────────────────────────────
# 스타일
# ──────────────────────────────────────────────

_FONT_TITLE    = Font(name="맑은 고딕", size=14, bold=True)
_FONT_SUBTITLE = Font(name="맑은 고딕", size=9,  italic=True)
_FONT_NOTICE   = Font(name="맑은 고딕", size=8,  italic=True)
_FONT_BOLD     = Font(name="맑은 고딕", size=10, bold=True)
_FONT_DEFAULT  = Font(name="맑은 고딕", size=10)
_FONT_SMALL    = Font(name="맑은 고딕", size=8,  italic=True)

_FILL_LABEL    = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SECTION  = PatternFill(fill_type="solid", fgColor="D9E1F2")
_FILL_HEADER   = PatternFill(fill_type="solid", fgColor="E2EFDA")
_FILL_WARN     = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_NOTICE   = PatternFill(fill_type="solid", fgColor="FEF0E7")
_FILL_VERDICT  = PatternFill(fill_type="solid", fgColor="FCE4D6")
_FILL_NONE     = PatternFill()

_THIN   = Side(border_style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LABEL  = Alignment(horizontal="center", vertical="center")

# 9열: 1=순번(4), 2~6=항목(26), 7=결과(8), 8~9=비고(18)
_COL_WIDTHS: Dict[int, int] = {
    1: 4, 2: 7, 3: 7, 4: 7, 5: 7, 6: 7,
    7: 8, 8: 9, 9: 9,
}

_L1, _V1S, _V1E = 1, 2, 5
_L2, _V2S, _V2E = 6, 7, 9
_LF, _VFS, _VFE = 1, 2, 9


def _v(data: Dict[str, Any], key: str) -> Any:
    val = data.get(key)
    return "" if val is None else val


def _border_rect(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _wc(ws, row: int, c1: int, c2: int, value: Any, *,
        font=None, fill=None, align=None, height: Optional[int] = None) -> None:
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value     = "" if value is None else value
    cell.font      = font  or _FONT_DEFAULT
    cell.fill      = fill  or _FILL_NONE
    cell.alignment = align or _ALIGN_LEFT
    _border_rect(ws, row, c1, row, c2)
    if height is not None:
        ws.row_dimensions[row].height = height


def _wlv(ws, row: int, label: str, value: Any,
         lc: int, vs: int, ve: int) -> None:
    _wc(ws, row, lc, lc, label, font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, vs, ve, value, font=_FONT_DEFAULT, align=_ALIGN_LEFT)


def _apply_col_widths(ws) -> None:
    for idx, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ──────────────────────────────────────────────
# 섹션별 렌더링
# ──────────────────────────────────────────────

def _write_title(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    c = ws.cell(row=row, column=1, value=SHEET_HEADING)
    c.font = _FONT_TITLE; c.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    s = ws.cell(row=row, column=1, value=SHEET_SUBTITLE)
    s.font = _FONT_SUBTITLE; s.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_notice_block(ws, row: int) -> int:
    for txt in (NOTICE_PTW003, NOTICE_CL001, NOTICE_LAW, NOTICE_NONCONFORM):
        _wc(ws, row, 1, TOTAL_COLS, f"※ {txt}",
            font=_FONT_NOTICE, fill=_FILL_NOTICE, align=_ALIGN_LEFT, height=16)
        row += 1
    return row


def _write_section1_site_info(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "① 현장 기본정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "사업장명",   _v(data, "site_name"),    _L1, _V1S, _V1E)
    _wlv(ws, row, "공사명",     _v(data, "project_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검 일자",  _v(data, "check_date"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "관리감독자", _v(data, "supervisor_name"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "점검자",     _v(data, "checker_name"),  _L1, _V1S, _V1E)
    _wlv(ws, row, "작업 장소",  _v(data, "work_location"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_section2_work_info(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "② 점검 대상 작업 정보",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    H = 20
    _wlv(ws, row, "작업명",     _v(data, "work_name"),   _L1, _V1S, _V1E)
    _wlv(ws, row, "작업 높이",  _v(data, "work_height"), _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    _wlv(ws, row, "작업 기간",  _v(data, "work_period"),  _L1, _V1S, _V1E)
    _wlv(ws, row, "PTW 번호",   _v(data, "ptw_no"),       _L2, _V2S, _V2E)
    ws.row_dimensions[row].height = H; row += 1
    return row


def _write_checklist_section(ws, row: int, section_label: str, title: str,
                              defaults: List[Dict[str, str]],
                              override: List[Dict[str, Any]]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, f"{section_label} {title}",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, TOTAL_COLS,
        "※ 결과: 이상없음 → ○ / 이상있음 → × / 해당없음 → -",
        font=_FONT_SMALL, fill=_FILL_WARN, align=_ALIGN_CENTER, height=14)
    row += 1
    _wc(ws, row, 1, 1, "No",      font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 6, "점검 항목", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 7, 7, "결과",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 9, "비고",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1
    for i, default in enumerate(defaults):
        provided  = override[i] if i < len(override) else {}
        item_text = provided.get("item")   or default.get("item",   "")
        result    = provided.get("result") or default.get("result", "")
        note      = provided.get("note")   or default.get("note",   "")
        _wc(ws, row, 1, 1, i + 1,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 6, item_text, font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 7, 7, result,    font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 8, 9, note,      font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        ws.row_dimensions[row].height = 26
        row += 1
    return row


def _write_section10_nonconformance(ws, row: int,
                                    items: List[Dict[str, Any]]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "⑩ 부적합 사항 및 시정조치",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, 1, "No",          font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER, height=18)
    _wc(ws, row, 2, 3, "부적합 내용",  font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 4, 5, "위치",         font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 6, 7, "시정조치 내용", font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 8, 8, "완료기한",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    _wc(ws, row, 9, 9, "완료확인",     font=_FONT_BOLD, fill=_FILL_HEADER, align=_ALIGN_CENTER)
    row += 1
    for i in range(MAX_NONCONFORM):
        src = items[i] if i < len(items) else {}
        _wc(ws, row, 1, 1, i + 1,              font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 2, 3, _v(src, "content"),  font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 4, 5, _v(src, "location"), font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 6, 7, _v(src, "action"),   font=_FONT_DEFAULT, align=_ALIGN_LEFT)
        _wc(ws, row, 8, 8, _v(src, "deadline"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        _wc(ws, row, 9, 9, _v(src, "completed"), font=_FONT_DEFAULT, align=_ALIGN_CENTER)
        ws.row_dimensions[row].height = 26
        row += 1
    return row


def _write_section11_verdict(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "⑪ 작업 가능 여부 판정",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1

    verdict = _v(data, "work_verdict") or ""
    def _mark(label: str) -> str:
        return f"☑ {label}" if verdict == label else f"□ {label}"

    _wc(ws, row, 1, 3, _mark("적합"),
        font=_FONT_BOLD, fill=_FILL_NONE, align=_ALIGN_CENTER, height=22)
    _wc(ws, row, 4, 6, _mark("조건부 적합"),
        font=_FONT_BOLD, fill=_FILL_NONE, align=_ALIGN_CENTER)
    _wc(ws, row, 7, 9, _mark("작업중지"),
        font=_FONT_BOLD, fill=_FILL_VERDICT, align=_ALIGN_CENTER)
    row += 1

    _wlv(ws, row, "조건 내용", _v(data, "verdict_condition"), _LF, _VFS, _VFE)
    ws.row_dimensions[row].height = 20; row += 1
    return row


def _write_section12_signatures(ws, row: int, data: Dict[str, Any]) -> int:
    _wc(ws, row, 1, TOTAL_COLS, "⑫ 확인 서명",
        font=_FONT_BOLD, fill=_FILL_SECTION, align=_ALIGN_CENTER, height=18)
    row += 1
    _wc(ws, row, 1, 3, "서명일",
        font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    _wc(ws, row, 4, 9, _v(data, "sign_date"),
        font=_FONT_DEFAULT, align=_ALIGN_LEFT)
    ws.row_dimensions[row].height = 20; row += 1

    labels    = ["점검자 서명",   "관리감독자 서명", "현장소장 서명"]
    sign_keys = ["inspector_sign", "supervisor_sign", "manager_sign"]
    col_ranges = [(1, 3), (4, 6), (7, 9)]

    for label, _, (cs, ce) in zip(labels, sign_keys, col_ranges):
        _wc(ws, row, cs, ce, label,
            font=_FONT_BOLD, fill=_FILL_LABEL, align=_ALIGN_LABEL)
    ws.row_dimensions[row].height = 18; row += 1

    for _, key, (cs, ce) in zip(labels, sign_keys, col_ranges):
        _wc(ws, row, cs, ce, _v(data, key),
            font=_FONT_DEFAULT, align=_ALIGN_CENTER)
    ws.row_dimensions[row].height = 36; row += 1
    return row


# ──────────────────────────────────────────────
# 공개 API
# ──────────────────────────────────────────────

def build_fall_protection_checklist_excel(form_data: Dict[str, Any]) -> bytes:
    """form_data dict를 받아 추락 방호 설비 점검표 xlsx 바이너리를 반환.

    [CL-007]  추락 방호 설비 점검표
    법적 근거: 산업안전보건기준에 관한 규칙 제42조~제45조 (PARTIAL_VERIFIED)
    비계 설치 점검은 CL-001, 고소작업 허가는 PTW-003 별도 사용.
    """
    data = form_data or {}

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _apply_col_widths(ws)

    row = 1
    row = _write_title(ws, row)
    row = _write_notice_block(ws, row)
    row = _write_section1_site_info(ws, row, data)
    row = _write_section2_work_info(ws, row, data)

    row = _write_checklist_section(
        ws, row, "③", "추락 위험 구역 확인",
        _HAZARD_ZONE_DEFAULTS, data.get("hazard_zone_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "④", "작업발판 상태 점검",
        _PLATFORM_DEFAULTS, data.get("platform_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑤", "안전난간 점검",
        _RAILING_DEFAULTS, data.get("railing_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑥", "개구부 덮개 및 방호조치",
        _OPENING_DEFAULTS, data.get("opening_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑦", "안전대 및 부착설비 점검",
        _HARNESS_DEFAULTS, data.get("harness_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑧", "추락방망 및 낙하 방지 설비",
        _NET_DEFAULTS, data.get("net_items") or [],
    )
    row = _write_checklist_section(
        ws, row, "⑨", "사다리·이동식 비계·고소작업대 추가 확인",
        _SPECIAL_EQUIP_DEFAULTS, data.get("special_equip_items") or [],
    )
    row = _write_section10_nonconformance(
        ws, row, data.get("nonconformance_items") or [],
    )
    row = _write_section11_verdict(ws, row, data)
    _write_section12_signatures(ws, row, data)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:2"  # 제목+부제 반복
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
